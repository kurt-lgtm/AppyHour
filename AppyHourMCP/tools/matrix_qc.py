"""
Production Matrix QC MCP tools — validate matrix Excel against Shopify orders,
generate per-order swap lists.

Consolidates compare_matrix.py and _gen_swap_csv.py into reusable MCP tools.
"""

import json
import time
import csv
import re
import os
from typing import Optional
from collections import Counter
from datetime import datetime
from pydantic import BaseModel, Field, ConfigDict

import requests

from utils import get_shopify_auth, format_error, to_json, APPYHOUR_ROOT
from tools.constants import NAME_TO_SKU, FOOD_PREFIXES


def _fetch_orders_by_tag(base, headers, tag, fields="id,name,line_items,email"):
    """Fetch all unfulfilled orders matching a specific tag."""
    all_orders = []
    url = f"{base}/orders.json"
    params = {
        "status": "open",
        "fulfillment_status": "unfulfilled",
        "limit": 250,
        "tag": tag,
        "fields": fields,
    }
    page = 0
    while url:
        page += 1
        resp = requests.get(url, headers=headers,
                            params=params if page == 1 else None, timeout=30)
        resp.raise_for_status()
        orders = resp.json().get("orders", [])
        all_orders.extend(orders)
        link = resp.headers.get("Link", "")
        url = None
        if 'rel="next"' in link:
            m = re.search(r'<([^>]+)>;\s*rel="next"', link)
            if m:
                url = m.group(1)
        time.sleep(0.1)
    return all_orders


def _parse_matrix(xlsx_path):
    """Parse production matrix Excel file. Returns (matrix, order_names, order_emails)."""
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Access_LIVE"]

    col_headers = {}
    for c in range(1, ws.max_column + 1):
        h = str(ws.cell(1, c).value or "")
        if h.startswith("AHB") and ": " in h:
            col_headers[c] = h.split(": ", 1)[1]

    matrix = {}
    order_names = {}
    order_emails = {}
    for r in range(2, ws.max_row + 1):
        oid = str(ws.cell(r, 1).value or "").strip()
        if not oid:
            continue
        order_names[oid] = str(ws.cell(r, 2).value or "")
        order_emails[oid] = str(ws.cell(r, 6).value or "")
        assignments = {}
        for c, name in col_headers.items():
            val = ws.cell(r, c).value
            if val and str(val).strip() not in ("", "0", "None"):
                sku = NAME_TO_SKU.get(name, f"??-{name[:15]}")
                assignments[sku] = int(float(str(val)))
        matrix[oid] = assignments

    return matrix, order_names, order_emails


def register(mcp):
    """Register matrix QC tools on the MCP server."""

    class ValidateMatrixInput(BaseModel):
        """Input for validating production matrix against Shopify."""
        model_config = ConfigDict(str_strip_whitespace=True)

        xlsx_path: str = Field(..., description="Path to the AHB_WeeklyProductionQuery xlsx file")
        rmfg_tag: str = Field(..., description="RMFG tag to filter Shopify orders (e.g. 'RMFG_20260323')")

    class GenerateSwapListInput(BaseModel):
        """Input for generating a per-order swap list."""
        model_config = ConfigDict(str_strip_whitespace=True)

        xlsx_path: str = Field(..., description="Path to the AHB_WeeklyProductionQuery xlsx file")
        rmfg_tag: str = Field(..., description="RMFG tag to filter Shopify orders (e.g. 'RMFG_20260323')")
        sku_prefix: str = Field("AC-", description="Only generate swaps for SKUs with this prefix (default: 'AC-')")
        output_path: str = Field("", description="Optional CSV output path. Default: ~/Downloads/swap-list-{date}.csv")

    @mcp.tool(
        name="appyhour_validate_production_matrix",
        annotations={
            "title": "Validate Production Matrix vs Shopify",
            "readOnlyHint": True,
            "destructiveHint": False,
            "idempotentHint": True,
            "openWorldHint": True,
        },
    )
    async def validate_production_matrix(params: ValidateMatrixInput) -> str:
        """Compare production matrix (Excel) against Shopify orders.

        Reads the weekly production query Excel file and fetches Shopify orders
        by RMFG tag. Reports SKUs present in the matrix but missing from Shopify
        and vice versa. Also analyzes replacement patterns.

        Args:
            params: Excel file path and RMFG tag.

        Returns:
            JSON with matrix_orders, shopify_orders, common_orders,
            missing_from_shopify, extra_on_shopify, and replacement analysis.
        """
        try:
            base, headers = get_shopify_auth()
            matrix, _, _ = _parse_matrix(params.xlsx_path)
            orders = _fetch_orders_by_tag(base, headers, params.rmfg_tag)

            # Build Shopify SKU map
            shopify = {}
            for o in orders:
                name = o["name"].replace("#", "")
                skus = {}
                for li in o.get("line_items", []):
                    sku = (li.get("sku") or "").strip()
                    fq = li.get("fulfillable_quantity", li.get("quantity", 0))
                    if sku and fq > 0:
                        skus[sku] = skus.get(sku, 0) + fq
                shopify[name] = skus

            common = set(matrix.keys()) & set(shopify.keys())

            missing_from_shopify = Counter()
            extra_on_shopify = Counter()
            for oid in common:
                m_food = {s for s in matrix[oid] if s.startswith(FOOD_PREFIXES)}
                s_food = {s for s in shopify[oid] if s.startswith(FOOD_PREFIXES)}
                for sku in m_food - s_food:
                    missing_from_shopify[sku] += 1
                for sku in s_food - m_food:
                    extra_on_shopify[sku] += 1

            return to_json({
                "matrix_orders": len(matrix),
                "shopify_orders": len(shopify),
                "common_orders": len(common),
                "matrix_only": len(set(matrix.keys()) - set(shopify.keys())),
                "shopify_only": len(set(shopify.keys()) - set(matrix.keys())),
                "missing_from_shopify": dict(missing_from_shopify.most_common()),
                "extra_on_shopify": dict(extra_on_shopify.most_common()),
            })
        except Exception as e:
            return format_error(e, "validate_production_matrix")

    @mcp.tool(
        name="appyhour_generate_swap_list",
        annotations={
            "title": "Generate Per-Order Swap List",
            "readOnlyHint": True,
            "destructiveHint": False,
            "idempotentHint": True,
            "openWorldHint": True,
        },
    )
    async def generate_swap_list(params: GenerateSwapListInput) -> str:
        """Generate a per-order swap list comparing matrix vs Shopify.

        For each order in both the matrix and Shopify, identifies SKUs that
        need to be removed (on Shopify but not in matrix) and added (in matrix
        but not on Shopify). Writes results to CSV.

        Args:
            params: Excel path, RMFG tag, SKU prefix filter, output path.

        Returns:
            JSON with swap rows, remove/add summary counts, and CSV path.
        """
        try:
            base, headers = get_shopify_auth()
            matrix, order_names, order_emails = _parse_matrix(params.xlsx_path)
            orders = _fetch_orders_by_tag(base, headers, params.rmfg_tag,
                                          fields="id,name,line_items,email")

            shopify = {}
            shopify_ids = {}
            for o in orders:
                name = o["name"].replace("#", "")
                skus = {}
                for li in o.get("line_items", []):
                    sku = (li.get("sku") or "").strip()
                    fq = li.get("fulfillable_quantity", li.get("quantity", 0))
                    if sku and fq > 0:
                        skus[sku] = skus.get(sku, 0) + fq
                shopify[name] = skus
                shopify_ids[name] = o["id"]

            prefix = params.sku_prefix
            rows = []
            for oid in sorted(set(matrix.keys()) & set(shopify.keys())):
                m_skus = set(matrix[oid].keys())
                s_skus = set(shopify[oid].keys())

                to_remove = sorted(s for s in s_skus - m_skus if s.startswith(prefix))
                to_add = sorted(s for s in m_skus - s_skus if s.startswith(prefix))

                to_add_iter = iter(to_add)
                for rem in to_remove:
                    add = next(to_add_iter, "")
                    rows.append({
                        "order_number": oid,
                        "order_id": shopify_ids.get(oid, ""),
                        "customer": order_names.get(oid, ""),
                        "email": order_emails.get(oid, ""),
                        "remove_sku": rem,
                        "add_sku": add,
                    })
                for add in to_add_iter:
                    rows.append({
                        "order_number": oid,
                        "order_id": shopify_ids.get(oid, ""),
                        "customer": order_names.get(oid, ""),
                        "email": order_emails.get(oid, ""),
                        "remove_sku": "",
                        "add_sku": add,
                    })

            today = datetime.now().strftime("%Y-%m-%d")
            outpath = params.output_path or str(
                APPYHOUR_ROOT.parent / "Downloads" / f"swap-list-{today}.csv"
            )
            with open(outpath, "w", newline="", encoding="utf-8") as f:
                writer = csv.DictWriter(f, fieldnames=[
                    "order_number", "order_id", "customer", "email", "remove_sku", "add_sku",
                ])
                writer.writeheader()
                writer.writerows(rows)

            remove_counts = Counter(r["remove_sku"] for r in rows if r["remove_sku"])
            add_counts = Counter(r["add_sku"] for r in rows if r["add_sku"])

            return to_json({
                "total_swaps": len(rows),
                "csv_path": outpath,
                "remove_summary": dict(remove_counts.most_common()),
                "add_summary": dict(add_counts.most_common()),
                "sample_rows": rows[:10],
            })
        except Exception as e:
            return format_error(e, "generate_swap_list")
