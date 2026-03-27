"""
Ops Summary Report & Cost of Issues builder.

Reads UPDATE_Operational Issues data, computes weekly pivots by FC,
writes formulas to the Ops Summary Report, and creates a Cost of Issues
tab with charts in Google Sheets.
"""

import json
import os
import re
import sys
from collections import Counter, defaultdict
from datetime import datetime, timedelta
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent / "GelPackCalculator"))

_APPDATA_SETTINGS = Path(os.environ.get("APPDATA", "")) / "AppyHour" / "gel_calc_shopify_settings.json"
SPREADSHEET_ID = "190AmXF8hy-M8lmt8q9uhOkyOMi7AmU0jJAd1KOpjWdA"
OPS_TAB = "Ops Summary Report "
DATA_TAB = "UPDATE_Operational Issues"
COST_TAB = "Cost of Issues"

# Resolution costs (col B of Ops Summary rows 35-66)
RESOLUTION_COSTS = {
    "Full Reship": 65.0,
    "Partial Reship": 35.0,
    "Credit Next Box::Amount $6": 6.0,
    "Credit Next Box::Amount $10": 10.0,
    "Credit Next Box::Amount $15": 15.0,
    "Credit Next Box::Amount $20": 20.0,
    "Credit Next Box::Amount $30": 30.0,
    "Credit Next Box::Amount $40+": 45.0,
    "Refund Order::Amount $6": 6.0,
    "Refund Order::Amount $10": 10.0,
    "Refund Order::Amount $15": 15.0,
    "Refund Order::Amount $20": 20.0,
    "Refund Order::Amount $30": 30.0,
    "Refund Order::Amount $40+": 45.0,
    "Refund Order::Full Amount": 110.0,
    "Refund Order::Specific Cheese(s)": 10.0,
    "Refund Order::Specific Meat(s)": 10.0,
    "Refund Order::Specific Accompaniment(s)": 6.0,
    "Refund Order::Specific Items (Multiple types)": 20.0,
    "Comp Item::Extra Cheese(s)": 5.5,
    "Comp Item::Extra Meat(s)": 4.0,
    "Comp Item::Extra Accompaniment(s)": 2.5,
    "Comp Item + Credit/Refund::$10 off + extra cheese": 15.5,
    "Comp Item + Credit/Refund::$15 off + extra cheese": 20.5,
    "Comp Item + Credit/Refund::$20 off + extra cheese": 25.5,
    "Comp Item + Credit/Refund::$10 off + extra meat": 14.0,
    "Comp Item + Credit/Refund::$15 off + extra meat": 19.0,
    "Comp Item + Credit/Refund::$20 off + extra meat": 24.0,
    "Comp Item + Credit/Refund::$10 off + extra accompaniment": 12.5,
    "Comp Item + Credit/Refund::$15 off + extra accompaniment": 17.5,
    "Comp Item + Credit/Refund::$20 off + extra accompaniment": 22.5,
    "Reship Box": 65.0,
    "FullReship": 65.0,
    "Full Refund": 110.0,
}

# Issue types in row order (rows 4-34 of Ops Summary)
ISSUE_TYPES = [
    "Shipping::cannot be delivered",
    "Shipping::Damaged in transit::Arrived Warm",
    "Shipping::Damaged in transit::Broken/Leaking Ice Pack",
    "Shipping::Delayed in transit::4+ Days in Transit",
    "Shipping::Delayed in transit::3 Days in Transit",
    "Shipping::Change Address::Customer Updated, Not applied",
    "Shipping::Change Address::Requested, Cannot be applied",
    "Shipping::Lost in Transit/Misdelivered",
    "Shipping::Damaged in transit::Full box damaged",
    "Shipping::Damaged in transit::Damaged cheese",
    "Shipping::Damaged in transit::Spilled accompaniment",
    "Shipping::Damaged in transit::Damaged meat",
    "Shipping::Damaged in transit::Spilled accompaniment or Broken Crackers or Broken/Leaking Jar",
    "Order::Missing item::All cheeses",
    "Order::Missing item::All Meats",
    "Order::Missing item::1+ cheeses",
    "Order::Missing item::1+ meats",
    "Order::Missing item::1+ accompaniment",
    "Order::Missing item::1+ meats",  # duplicate row in original
    "Order::Missing tasting guide",
    "Order::Wrong item::Cheese",
    "Order::Wrong item::Meat",
    "Order::Wrong item::Accompaniment",
    "Order::Wrong Order",
    "Order::Substitute complaint",
    "Order::Quality Complaint::Cheese",
    "Order::Quality Complaint::Meat",
    "Order::Quality Complaint::Accompaniment",
    "Order::Spoiled Item::Cheese",
    "Order::Spoiled Item::Meat",
    "Order::Spoiled Item::Accompaniment",
]

RESOLUTION_TYPES = list(RESOLUTION_COSTS.keys())

# Map variant resolutions to canonical names for cost lookup
RESOLUTION_ALIASES = {
    "FullReship": "Full Reship",
    "Reship Box": "Full Reship",
    "Full Refund": "Refund Order::Full Amount",
}

FCS = ["GRIPCA", "RMFG", "COG"]


def _load_settings() -> dict:
    with open(_APPDATA_SETTINGS, encoding="utf-8") as f:
        return json.load(f)


def _get_google_client():
    from google_integration import GoogleIntegration
    settings = _load_settings()
    creds = settings.get("google_credentials_path", "")
    if not creds or not os.path.exists(creds):
        creds = str(Path(__file__).resolve().parent.parent.parent
                     / "shipping-perfomance-review-accd39ac4b78.json")
    return GoogleIntegration(creds)


def _parse_date(date_str: str) -> datetime | None:
    """Parse date strings like 'May-15', 'Mar-18', 'Dec 6', 'December 5', '6/30/2025'."""
    date_str = date_str.strip()
    if not date_str:
        return None
    # Try various formats
    formats = [
        "%b-%d",       # May-15, Mar-18
        "%B-%d",       # March-15
        "%b %d",       # Dec 6, Mar 18
        "%B %d",       # December 5, March 18
        "%b-%d-%Y",    # May-15-2025
        "%m/%d/%Y",    # 6/30/2025
        "%Y-%m-%d",    # 2025-06-30
        "%m/%d/%y",    # 6/30/25
    ]
    for fmt in formats:
        try:
            dt = datetime.strptime(date_str, fmt)
            # If year is 1900 (no year in format), infer year
            if dt.year == 1900:
                now = datetime.now()
                dt = dt.replace(year=now.year)
                if dt > now + timedelta(days=60):
                    dt = dt.replace(year=now.year - 1)
            return dt
        except ValueError:
            continue
    return None


def _week_start(dt: datetime) -> datetime:
    """Get Monday of the week containing dt."""
    return dt - timedelta(days=dt.weekday())


def _load_shipment_volumes() -> dict:
    """Load weekly shipment volumes per FC from the Excel source.

    Returns {week_monday_datetime: {"GRIPCA": n, "RMFG": n, "COG": n, "SUM": n}}.
    """
    import openpyxl

    excel_path = Path(__file__).resolve().parent.parent.parent / "Issue & Resolution Guide (2).xlsx"
    if not excel_path.exists():
        return {}

    wb = openpyxl.load_workbook(str(excel_path), data_only=True)
    ws = wb["Ops Summary Report "]

    row1 = [c.value for c in ws[1][:150]]
    row2 = [c.value for c in ws[2][:150]]

    volumes = {}
    for i in range(3, len(row2), 4):
        date_val = row2[i]
        if not date_val:
            break
        if isinstance(date_val, datetime):
            week_dt = _week_start(date_val)
        elif isinstance(date_val, str):
            try:
                week_dt = _week_start(datetime.strptime(date_val.strip(), "%Y-%m-%d %H:%M:%S"))
            except ValueError:
                try:
                    week_dt = _week_start(datetime.strptime(date_val.strip(), "%m/%d/%Y"))
                except ValueError:
                    continue
        else:
            continue

        def _num(idx):
            v = row1[idx] if idx < len(row1) else None
            if v is None or v == "":
                return 0
            try:
                return float(v)
            except (ValueError, TypeError):
                return 0

        volumes[week_dt] = {
            "GRIPCA": _num(i),
            "RMFG": _num(i + 1),
            "COG": _num(i + 2),
            "SUM": _num(i + 3),
        }

    wb.close()
    return volumes


def _col_letter(n: int) -> str:
    """Convert 0-based column index to A, B, ..., Z, AA, AB, ..."""
    result = ""
    while True:
        result = chr(65 + n % 26) + result
        n = n // 26 - 1
        if n < 0:
            break
    return result


def build_ops_summary(dry_run: bool = False) -> dict:
    """Rebuild the Ops Summary Report from UPDATE_Operational Issues data.

    Reads all raw data, computes weekly pivots by FC, writes counts
    and cost formulas to the Ops Summary tab, then creates/updates
    the Cost of Issues tab with charts.
    """
    gclient = _get_google_client()
    sheets_svc = gclient._sheets

    # ── Read raw data ─────────────────────────────────────────────────
    raw = gclient.read_sheet(SPREADSHEET_ID, f"'{DATA_TAB}'!A:J")
    if not raw:
        return {"error": "No data in UPDATE_Operational Issues"}

    headers = raw[0]
    data_rows = raw[1:]

    # Parse into records
    records = []
    for row in data_rows:
        while len(row) < 10:
            row.append("")
        date_str, contact_reason, order_num, gorgias_link, carrier, \
            state, fc_tag, issue_type, resolution, comment = row[:10]

        dt = _parse_date(date_str)
        if not dt:
            continue

        records.append({
            "date": dt,
            "week": _week_start(dt),
            "fc": fc_tag.strip(),
            "issue_type": issue_type.strip(),
            "resolution": resolution.strip(),
            "state": state.strip(),
            "carrier": carrier.strip(),
        })

    if not records:
        return {"error": "No parseable records"}

    # ── Determine weeks ───────────────────────────────────────────────
    all_weeks = sorted(set(r["week"] for r in records))

    # ── Build pivot: (week, fc) -> Counter of issue types / resolutions
    issue_pivot = defaultdict(Counter)  # (week, fc) -> {issue_type: count}
    resolution_pivot = defaultdict(Counter)
    week_fc_total = Counter()  # (week, fc) -> total orders
    week_total = Counter()  # week -> total orders

    for r in records:
        fc = r["fc"] if r["fc"] in FCS else "RMFG"  # Default empty FC to RMFG (largest)
        r["fc"] = fc
        key = (r["week"], fc)
        if r["issue_type"]:
            issue_pivot[key][r["issue_type"]] += 1
        if r["resolution"]:
            resolution_pivot[key][r["resolution"]] += 1
        week_fc_total[key] += 1
        week_total[r["week"]] += 1

    # ── Load shipment volumes from Excel ─────────────────────────────
    shipment_volumes = _load_shipment_volumes()

    # ── Build the sheet data ──────────────────────────────────────────
    num_weeks = len(all_weeks)

    # Row 1: headers + total SHIPMENTS per FC per week (not issues)
    row1 = ["Issue vs. Resolution", "Category", "Issue"]
    row2 = ["", "", ""]
    row3 = ["", "", ""]

    for week in all_weeks:
        end = week + timedelta(days=6)
        vol = shipment_volumes.get(week, {})
        for fc in FCS:
            row1.append(str(int(vol.get(fc, 0))) if vol.get(fc) else "")
            row3.append(fc)
        row1.append(str(int(vol.get("SUM", 0))) if vol.get("SUM") else "")
        row3.append("SUM")
        row2.extend([week.strftime("%#m/%#d/%Y"), "to ", end.strftime("%#m/%#d/%Y"), ""])

    # Issue rows (4-34)
    issue_rows = []
    for itype in ISSUE_TYPES:
        category = "Shipping" if itype.startswith("Shipping") else "Order"
        row = ["Issue", category, itype]
        for week in all_weeks:
            for fc in FCS:
                count = issue_pivot.get((week, fc), {}).get(itype, 0)
                # Also check partial matches (Gorgias values may differ slightly)
                if count == 0:
                    for key_type, cnt in issue_pivot.get((week, fc), {}).items():
                        if key_type.startswith(itype.rstrip()):
                            count += cnt
                row.append(count if count else "")
            # SUM
            total = sum(
                issue_pivot.get((week, fc), {}).get(itype, 0) for fc in FCS
            )
            row.append(total if total else "")
        issue_rows.append(row)

    # Resolution rows (35-66)
    resolution_rows = []
    for rtype in RESOLUTION_TYPES:
        cost = RESOLUTION_COSTS[rtype]
        row = ["Resolution", cost if cost else "", rtype]
        for week in all_weeks:
            for fc in FCS:
                count = resolution_pivot.get((week, fc), {}).get(rtype, 0)
                if count == 0:
                    for key_type, cnt in resolution_pivot.get((week, fc), {}).items():
                        if key_type.startswith(rtype):
                            count += cnt
                row.append(count if count else "")
            total = sum(
                resolution_pivot.get((week, fc), {}).get(rtype, 0) for fc in FCS
            )
            row.append(total if total else "")
        resolution_rows.append(row)

    # Total Reships percent row (after Full Reship and Partial Reship)
    def _count_reships(week, fc):
        """Count all reship variants for a given week+FC."""
        data = resolution_pivot.get((week, fc), {})
        reship_keys = ("Full Reship", "Partial Reship", "Reship Box", "FullReship")
        return sum(data.get(k, 0) for k in reship_keys)

    reship_pct_row = ["Resolution", "", "Total Reships percent"]
    for week in all_weeks:
        vol = shipment_volumes.get(week, {})
        for fc in FCS:
            reships = _count_reships(week, fc)
            shipments = vol.get(fc, 0)
            pct = reships / shipments if shipments else 0
            reship_pct_row.append(f"{pct:.2%}" if pct else "0.00%")
        total_reships = sum(_count_reships(week, fc) for fc in FCS)
        total_shipments = vol.get("SUM", 0)
        pct = total_reships / total_shipments if total_shipments else 0
        reship_pct_row.append(f"{pct:.2%}" if pct else "0.00%")

    # Box setup + cost section (rows 67-71)
    blank_row = [""] * (3 + num_weeks * 4)
    box_setup_row = ["", "", "Box Set Up"] + [""] * (num_weeks * 4)

    # Cost rows
    cost_row_68 = ["", "", ""]
    cost_row_69 = ["", "", ""]
    cost_row_70 = ["", "", ""]  # Total cost per FC per week
    cost_row_71 = ["", "", ""]  # Cost per order

    for week in all_weeks:
        end = week + timedelta(days=6)
        cost_row_68.extend([week.strftime("%#m/%#d/%Y"), "to ", end.strftime("%#m/%#d/%Y"), ""])
        cost_row_69.extend(["GRIPCA", "RMFG", "COG", "SUM"])

        week_costs = {}
        for fc in FCS:
            fc_cost = 0.0
            fc_resolutions = resolution_pivot.get((week, fc), {})
            for res_name, count in fc_resolutions.items():
                # Direct match
                if res_name in RESOLUTION_COSTS:
                    fc_cost += count * RESOLUTION_COSTS[res_name]
                # Alias match
                elif res_name in RESOLUTION_ALIASES:
                    canonical = RESOLUTION_ALIASES[res_name]
                    if canonical in RESOLUTION_COSTS:
                        fc_cost += count * RESOLUTION_COSTS[canonical]
                # Prefix match
                else:
                    for rtype, unit_cost in RESOLUTION_COSTS.items():
                        if res_name.startswith(rtype):
                            fc_cost += count * unit_cost
                            break
            week_costs[fc] = fc_cost
            cost_row_70.append(round(fc_cost, 2))

        total_cost = sum(week_costs.values())
        cost_row_70.append(round(total_cost, 2))

        vol = shipment_volumes.get(week, {})
        total_shipments = vol.get("SUM", 0)
        cost_per_order = total_cost / total_shipments if total_shipments else 0
        for fc in FCS:
            fc_shipments = vol.get(fc, 0)
            fc_cpo = week_costs[fc] / fc_shipments if fc_shipments else 0
            cost_row_71.append(round(fc_cpo, 2))
        cost_row_71.append(round(cost_per_order, 2))

    # Assemble all rows
    # Insert reship pct after first 2 resolution rows
    all_res_rows = [resolution_rows[0], resolution_rows[1], reship_pct_row] + resolution_rows[2:]

    all_sheet_rows = (
        [row1, row2, row3]
        + issue_rows
        + all_res_rows
        + [box_setup_row, cost_row_68, cost_row_69, cost_row_70, cost_row_71]
    )

    if dry_run:
        return {
            "weeks": len(all_weeks),
            "records": len(records),
            "rows": len(all_sheet_rows),
            "cols": len(row1),
            "sample_cost_row": cost_row_70[:15],
        }

    # ── Write to Ops Summary ──────────────────────────────────────────
    import time as _time

    # Clear existing data first
    try:
        sheets_svc.spreadsheets().values().clear(
            spreadsheetId=SPREADSHEET_ID,
            range=f"'{OPS_TAB}'!A1:ZZ200",
        ).execute()
    except Exception:
        pass  # OK if nothing to clear
    _time.sleep(2)

    # Write in batches of 10 rows via batchUpdate to avoid 500 errors
    batch_size = 10
    for i in range(0, len(all_sheet_rows), batch_size):
        chunk = all_sheet_rows[i:i + batch_size]
        start_row = i + 1
        sheets_svc.spreadsheets().values().update(
            spreadsheetId=SPREADSHEET_ID,
            range=f"'{OPS_TAB}'!A{start_row}",
            valueInputOption="USER_ENTERED",
            body={"values": chunk},
        ).execute()
        _time.sleep(0.5)

    # ── Format the Ops Summary ──────────────────────────────────────
    _format_ops_summary(sheets_svc, num_weeks, len(all_sheet_rows))

    # ── Build Cost of Issues tab ──────────────────────────────────────
    _build_cost_of_issues_tab(gclient, all_weeks, cost_row_70, cost_row_71,
                               week_fc_total, week_total)

    return {
        "success": True,
        "weeks": len(all_weeks),
        "records": len(records),
        "rows_written": len(all_sheet_rows),
    }


def _format_ops_summary(sheets_svc, num_weeks, total_rows):
    """Apply formatting to the Ops Summary Report tab."""
    # Get sheet ID
    meta = sheets_svc.spreadsheets().get(
        spreadsheetId=SPREADSHEET_ID,
        fields="sheets.properties",
    ).execute()
    sheet_id = None
    for s in meta.get("sheets", []):
        if s["properties"]["title"] == OPS_TAB:
            sheet_id = s["properties"]["sheetId"]
            break
    if sheet_id is None:
        return

    total_cols = 3 + num_weeks * 4

    requests = [
        # Freeze first 3 rows and 3 columns
        {
            "updateSheetProperties": {
                "properties": {
                    "sheetId": sheet_id,
                    "gridProperties": {
                        "frozenRowCount": 3,
                        "frozenColumnCount": 3,
                    },
                },
                "fields": "gridProperties.frozenRowCount,gridProperties.frozenColumnCount",
            }
        },
        # Bold rows 1-3 (headers)
        {
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": 0,
                    "endRowIndex": 3,
                    "startColumnIndex": 0,
                    "endColumnIndex": total_cols,
                },
                "cell": {
                    "userEnteredFormat": {
                        "textFormat": {"bold": True},
                        "horizontalAlignment": "CENTER",
                        "backgroundColor": {"red": 0.9, "green": 0.9, "blue": 0.95},
                    }
                },
                "fields": "userEnteredFormat(textFormat,horizontalAlignment,backgroundColor)",
            }
        },
        # Bold column A-C (row labels)
        {
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": 3,
                    "endRowIndex": total_rows,
                    "startColumnIndex": 0,
                    "endColumnIndex": 3,
                },
                "cell": {
                    "userEnteredFormat": {
                        "textFormat": {"bold": True},
                    }
                },
                "fields": "userEnteredFormat(textFormat)",
            }
        },
        # Color issue rows (light blue)
        {
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": 3,
                    "endRowIndex": 3 + len(ISSUE_TYPES),
                    "startColumnIndex": 0,
                    "endColumnIndex": 3,
                },
                "cell": {
                    "userEnteredFormat": {
                        "backgroundColor": {"red": 0.85, "green": 0.92, "blue": 1.0},
                        "textFormat": {"bold": True},
                    }
                },
                "fields": "userEnteredFormat(backgroundColor,textFormat)",
            }
        },
        # Color resolution rows (light green)
        {
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": 3 + len(ISSUE_TYPES),
                    "endRowIndex": total_rows - 5,  # before cost rows
                    "startColumnIndex": 0,
                    "endColumnIndex": 3,
                },
                "cell": {
                    "userEnteredFormat": {
                        "backgroundColor": {"red": 0.85, "green": 1.0, "blue": 0.85},
                        "textFormat": {"bold": True},
                    }
                },
                "fields": "userEnteredFormat(backgroundColor,textFormat)",
            }
        },
        # Bold + color cost rows (last 2 data rows)
        {
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": total_rows - 2,
                    "endRowIndex": total_rows,
                    "startColumnIndex": 0,
                    "endColumnIndex": total_cols,
                },
                "cell": {
                    "userEnteredFormat": {
                        "textFormat": {"bold": True},
                        "backgroundColor": {"red": 1.0, "green": 0.95, "blue": 0.8},
                    }
                },
                "fields": "userEnteredFormat(textFormat,backgroundColor)",
            }
        },
        # Center-align all data cells
        {
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": 3,
                    "endRowIndex": total_rows,
                    "startColumnIndex": 3,
                    "endColumnIndex": total_cols,
                },
                "cell": {
                    "userEnteredFormat": {
                        "horizontalAlignment": "CENTER",
                    }
                },
                "fields": "userEnteredFormat(horizontalAlignment)",
            }
        },
        # Auto-resize column C (issue names)
        {
            "updateDimensionProperties": {
                "range": {
                    "sheetId": sheet_id,
                    "dimension": "COLUMNS",
                    "startIndex": 2,
                    "endIndex": 3,
                },
                "properties": {"pixelSize": 380},
                "fields": "pixelSize",
            }
        },
    ]

    # Add alternating week column colors (light gray for every other week)
    for w in range(0, num_weeks, 2):
        col_start = 3 + w * 4
        requests.append({
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": 3,
                    "endRowIndex": total_rows - 5,
                    "startColumnIndex": col_start,
                    "endColumnIndex": col_start + 4,
                },
                "cell": {
                    "userEnteredFormat": {
                        "backgroundColor": {"red": 0.96, "green": 0.96, "blue": 0.96},
                        "horizontalAlignment": "CENTER",
                    }
                },
                "fields": "userEnteredFormat(backgroundColor,horizontalAlignment)",
            }
        })

    sheets_svc.spreadsheets().batchUpdate(
        spreadsheetId=SPREADSHEET_ID,
        body={"requests": requests},
    ).execute()


def _build_cost_of_issues_tab(gclient, all_weeks, cost_row_70, cost_row_71,
                                week_fc_total, week_total):
    """Create/update the Cost of Issues tab with summary data and charts."""
    sheets_svc = gclient._sheets

    # Check if tab exists
    meta = sheets_svc.spreadsheets().get(
        spreadsheetId=SPREADSHEET_ID,
        fields="sheets.properties",
    ).execute()
    tab_exists = False
    cost_sheet_id = None
    for s in meta.get("sheets", []):
        if s["properties"]["title"] == COST_TAB:
            tab_exists = True
            cost_sheet_id = s["properties"]["sheetId"]
            break

    if not tab_exists:
        resp = sheets_svc.spreadsheets().batchUpdate(
            spreadsheetId=SPREADSHEET_ID,
            body={"requests": [{"addSheet": {"properties": {"title": COST_TAB}}}]},
        ).execute()
        cost_sheet_id = resp["replies"][0]["addSheet"]["properties"]["sheetId"]

    # Build data for Cost of Issues
    # Section 1: Weekly cost by FC (for bar chart)
    header = ["Week", "GRIPCA", "RMFG", "COG", "Total"]
    rows = [header]
    for i, week in enumerate(all_weeks):
        end = week + timedelta(days=6)
        week_label = f"{week.strftime('%#m/%#d')} - {end.strftime('%#m/%#d')}"
        col_base = 3 + i * 4  # 0-indexed into cost_row_70
        gripca = float(cost_row_70[col_base]) if col_base < len(cost_row_70) else 0
        rmfg = float(cost_row_70[col_base + 1]) if col_base + 1 < len(cost_row_70) else 0
        cog = float(cost_row_70[col_base + 2]) if col_base + 2 < len(cost_row_70) else 0
        total = float(cost_row_70[col_base + 3]) if col_base + 3 < len(cost_row_70) else 0
        rows.append([week_label, gripca, rmfg, cog, total])

    # Section 2: Cost per order (for line chart) - start 3 rows below
    gap = len(rows) + 2
    cpo_header = ["Week", "Cost per Order"]
    cpo_rows = [cpo_header]
    for i, week in enumerate(all_weeks):
        end = week + timedelta(days=6)
        week_label = f"{week.strftime('%#m/%#d')} - {end.strftime('%#m/%#d')}"
        col_base = 3 + i * 4
        total_cost_val = float(cost_row_70[col_base + 3]) if col_base + 3 < len(cost_row_70) else 0
        total_orders = week_total[week]
        cpo = total_cost_val / total_orders if total_orders else 0
        cpo_rows.append([week_label, round(cpo, 2)])

    # Section 3: Summary stats
    summary_start = gap + len(cpo_rows) + 2
    total_cost_all = sum(float(cost_row_70[3 + i * 4 + 3])
                         for i in range(len(all_weeks))
                         if 3 + i * 4 + 3 < len(cost_row_70))
    total_orders_all = sum(week_total[w] for w in all_weeks)
    avg_cpo = total_cost_all / total_orders_all if total_orders_all else 0

    summary_rows = [
        ["Summary"],
        ["Total Issues", total_orders_all],
        ["Total Resolution Cost", f"${total_cost_all:,.2f}"],
        ["Avg Cost per Issue", f"${avg_cpo:.2f}"],
        ["Weeks Tracked", len(all_weeks)],
        ["Date Range", f"{all_weeks[0].strftime('%m/%d/%Y')} - {all_weeks[-1].strftime('%m/%d/%Y')}"],
    ]

    # Clear and write
    sheets_svc.spreadsheets().values().clear(
        spreadsheetId=SPREADSHEET_ID,
        range=f"'{COST_TAB}'!A1:Z200",
    ).execute()

    # Write all sections
    all_data = rows + [[""], [""]] + cpo_rows + [[""], [""]] + summary_rows
    sheets_svc.spreadsheets().values().update(
        spreadsheetId=SPREADSHEET_ID,
        range=f"'{COST_TAB}'!A1",
        valueInputOption="USER_ENTERED",
        body={"values": all_data},
    ).execute()

    # ── Create charts ─────────────────────────────────────────────────
    num_weeks = len(all_weeks)

    # Delete existing charts on the tab
    existing_meta = sheets_svc.spreadsheets().get(
        spreadsheetId=SPREADSHEET_ID,
        fields="sheets.charts",
    ).execute()
    delete_requests = []
    for sheet in existing_meta.get("sheets", []):
        for chart in sheet.get("charts", []):
            delete_requests.append({"deleteEmbeddedObject": {"objectId": chart["chartId"]}})

    requests = delete_requests + [
        # Chart 1: Stacked bar — Cost of Issues per Week by FC
        {
            "addChart": {
                "chart": {
                    "position": {
                        "overlayPosition": {
                            "anchorCell": {"sheetId": cost_sheet_id, "rowIndex": 0, "columnIndex": 6},
                            "widthPixels": 900,
                            "heightPixels": 400,
                        }
                    },
                    "spec": {
                        "title": "Cost of Issues per Week",
                        "basicChart": {
                            "chartType": "COLUMN",
                            "stackedType": "STACKED",
                            "legendPosition": "BOTTOM_LEGEND",
                            "axis": [
                                {"position": "BOTTOM_AXIS", "title": "Week"},
                                {"position": "LEFT_AXIS", "title": "Cost ($)"},
                            ],
                            "domains": [{
                                "domain": {
                                    "sourceRange": {
                                        "sources": [{
                                            "sheetId": cost_sheet_id,
                                            "startRowIndex": 0,
                                            "endRowIndex": num_weeks + 1,
                                            "startColumnIndex": 0,
                                            "endColumnIndex": 1,
                                        }]
                                    }
                                }
                            }],
                            "series": [
                                {
                                    "series": {
                                        "sourceRange": {
                                            "sources": [{
                                                "sheetId": cost_sheet_id,
                                                "startRowIndex": 0,
                                                "endRowIndex": num_weeks + 1,
                                                "startColumnIndex": col,
                                                "endColumnIndex": col + 1,
                                            }]
                                        }
                                    },
                                    "targetAxis": "LEFT_AXIS",
                                }
                                for col in [1, 2, 3]  # GRIPCA, RMFG, COG
                            ],
                            "headerCount": 1,
                        },
                    },
                }
            }
        },
        # Chart 2: Line — Cost per Order trend
        {
            "addChart": {
                "chart": {
                    "position": {
                        "overlayPosition": {
                            "anchorCell": {"sheetId": cost_sheet_id, "rowIndex": num_weeks + 4, "columnIndex": 6},
                            "widthPixels": 900,
                            "heightPixels": 400,
                        }
                    },
                    "spec": {
                        "title": "Cost per Issue vs. Week",
                        "basicChart": {
                            "chartType": "LINE",
                            "legendPosition": "BOTTOM_LEGEND",
                            "axis": [
                                {"position": "BOTTOM_AXIS", "title": "Week"},
                                {"position": "LEFT_AXIS", "title": "Cost per Issue ($)"},
                            ],
                            "domains": [{
                                "domain": {
                                    "sourceRange": {
                                        "sources": [{
                                            "sheetId": cost_sheet_id,
                                            "startRowIndex": num_weeks + 3,
                                            "endRowIndex": num_weeks + 3 + num_weeks + 1,
                                            "startColumnIndex": 0,
                                            "endColumnIndex": 1,
                                        }]
                                    }
                                }
                            }],
                            "series": [{
                                "series": {
                                    "sourceRange": {
                                        "sources": [{
                                            "sheetId": cost_sheet_id,
                                            "startRowIndex": num_weeks + 3,
                                            "endRowIndex": num_weeks + 3 + num_weeks + 1,
                                            "startColumnIndex": 1,
                                            "endColumnIndex": 2,
                                        }]
                                    }
                                },
                                "targetAxis": "LEFT_AXIS",
                            }],
                            "headerCount": 1,
                        },
                    },
                }
            }
        },
    ]

    if requests:
        sheets_svc.spreadsheets().batchUpdate(
            spreadsheetId=SPREADSHEET_ID,
            body={"requests": requests},
        ).execute()


def _get_shopify_client():
    """Get a ShopifyClient instance for order lookups."""
    from gel_pack_shopify import ShopifyClient
    settings = _load_settings()
    store = settings.get("store_url", "")
    cid = settings.get("shopify_client_id", "")
    csecret = settings.get("shopify_secret", "")
    if not (store and cid and csecret):
        raise ValueError("Shopify credentials not configured in settings.")
    return ShopifyClient(store, cid, csecret)


def _classify_fc_from_tags(tags_str: str) -> str | None:
    """Classify an order's FC from its standalone Shopify tags.

    Looks for exact tags: RMFG, COG, GRIPCA (standalone, not prefixed).
    Falls back to RMFG_*, COG_*, GRIPCA_* prefixed tags.
    Returns the FC name or None if no match.
    """
    tags = [t.strip() for t in tags_str.split(",")]
    # Check standalone FC tags first (most reliable)
    for tag in tags:
        tag_upper = tag.upper()
        if tag_upper == "RMFG":
            return "RMFG"
        if tag_upper == "COG":
            return "COG"
        if tag_upper == "GRIPCA":
            return "GRIPCA"
    # Fallback: prefixed tags
    for tag in tags:
        tag_upper = tag.upper()
        if tag_upper.startswith("RMFG_") or tag_upper.startswith("RMFG-"):
            return "RMFG"
        if tag_upper.startswith("COG_") or tag_upper.startswith("COG-"):
            return "COG"
        if tag_upper.startswith("GRIPCA_") or tag_upper.startswith("GRIPCA-"):
            return "GRIPCA"
    return None


def _extract_ship_date(tags_str: str) -> datetime | None:
    """Extract ship date from _SHIP_YYYY-MM-DD tag.

    Returns the ship date as a datetime, or None if no tag found.
    """
    for tag in tags_str.split(","):
        tag = tag.strip()
        if tag.startswith("_SHIP_"):
            date_part = tag[6:]  # after "_SHIP_"
            try:
                return datetime.strptime(date_part, "%Y-%m-%d")
            except ValueError:
                continue
    return None


def build_shipments_from_shopify(
    weeks_back: int = 12,
    dry_run: bool = False,
) -> dict:
    """Query Shopify for fulfilled orders and count shipments by FC per week.

    Looks for orders with tags starting with RMFG_*, COG_*, or GRIPCA_*
    prefixes. Groups by week (Monday start) and FC. Writes results to the
    Shipments tab in Google Sheets.

    Args:
        weeks_back: How many weeks of history to pull (default 12).
        dry_run: If True, return counts without writing to sheet.

    Returns dict with summary and weekly counts.
    """
    import time as _time
    import requests as _requests

    client = _get_shopify_client()
    gclient = _get_google_client()
    sheets_svc = gclient._sheets

    # Calculate date range
    today = datetime.now()
    start_date = _week_start(today - timedelta(weeks=weeks_back))

    # Lightweight fetch — only tags + created_at (no addresses, items, fulfillments)
    all_orders = []
    from datetime import date as dt_date

    # Month-by-month chunking to avoid timeouts
    d = start_date.date() if hasattr(start_date, 'date') else dt_date(start_date.year, start_date.month, start_date.day)
    end_date = dt_date.today()
    chunks = []
    while d <= end_date:
        if d.month == 12:
            next_month = dt_date(d.year + 1, 1, 1)
        else:
            next_month = dt_date(d.year, d.month + 1, 1)
        chunk_end = min(next_month - timedelta(days=1), end_date)
        chunks.append((d, chunk_end))
        d = next_month

    for chunk_idx, (c_start, c_end) in enumerate(chunks, 1):
        params = {
            "status": "any",
            "fulfillment_status": "fulfilled",
            "limit": 250,
            "fields": "id,created_at,tags",  # lightweight — only what we need
            "created_at_min": f"{c_start}T00:00:00Z",
            "created_at_max": f"{c_end}T23:59:59Z",
        }
        url = client._url("orders.json")
        page = 0
        while url:
            page += 1
            resp = _requests.get(
                url, headers=client._headers(),
                params=params if page == 1 else None,
                timeout=30,
            )
            if resp.status_code != 200:
                raise Exception(f"Shopify HTTP {resp.status_code}: {resp.text[:200]}")
            orders = resp.json().get("orders", [])
            if not orders:
                break
            all_orders.extend(orders)
            # Next page via Link header
            link = resp.headers.get("Link", "")
            url = None
            if 'rel="next"' in link:
                for part in link.split(","):
                    if 'rel="next"' in part:
                        url = part.split("<")[1].split(">")[0]
                        params = None
                        break
            _time.sleep(0.3)

    # Count by week and FC
    week_fc_counts = defaultdict(lambda: {"GRIPCA": 0, "RMFG": 0, "COG": 0})

    classified = 0
    unclassified = 0

    for order in all_orders:
        tags_str = order.get("tags", "")

        # Only count orders with a _SHIP_ tag (indicates actual shipment)
        ship_date = _extract_ship_date(tags_str)
        if not ship_date:
            unclassified += 1
            continue

        fc = _classify_fc_from_tags(tags_str)
        if not fc:
            unclassified += 1
            continue

        classified += 1

        # Use ship date for week grouping (not order creation date)
        week = _week_start(ship_date)
        week_fc_counts[week][fc] += 1

    # Build Shipments tab data
    weeks = sorted(week_fc_counts.keys())
    ship_headers = ["Week Start", "Week End", "GRIPCA", "RMFG", "COG", "Total"]
    ship_rows = []

    for week in weeks:
        end = week + timedelta(days=6)
        counts = week_fc_counts[week]
        rn = len(ship_rows) + 2  # 1-indexed, after header
        ship_rows.append([
            week.strftime("%m/%d/%Y"),
            end.strftime("%m/%d/%Y"),
            counts["GRIPCA"] or "",
            counts["RMFG"] or "",
            counts["COG"] or "",
            f"=SUM(C{rn}:E{rn})",
        ])

    if not dry_run and ship_rows:
        # Ensure Shipments tab exists
        meta = sheets_svc.spreadsheets().get(
            spreadsheetId=SPREADSHEET_ID, fields="sheets.properties",
        ).execute()
        tab_exists = any(
            s["properties"]["title"] == "Shipments" for s in meta["sheets"]
        )
        if not tab_exists:
            sheets_svc.spreadsheets().batchUpdate(
                spreadsheetId=SPREADSHEET_ID,
                body={"requests": [{"addSheet": {"properties": {"title": "Shipments"}}}]},
            ).execute()
        else:
            sheets_svc.spreadsheets().values().clear(
                spreadsheetId=SPREADSHEET_ID, range="'Shipments'!A1:Z200",
            ).execute()

        gclient.write_sheet(SPREADSHEET_ID, "Shipments", ship_headers, ship_rows)

    return {
        "success": True,
        "total_orders": len(all_orders),
        "classified": classified,
        "unclassified": unclassified,
        "weeks": len(weeks),
        "dry_run": dry_run,
        "weekly_counts": [
            {
                "week": w.strftime("%m/%d/%Y"),
                "gripca": week_fc_counts[w]["GRIPCA"],
                "rmfg": week_fc_counts[w]["RMFG"],
                "cog": week_fc_counts[w]["COG"],
            }
            for w in weeks
        ],
    }


def register(mcp):
    """Register Ops Summary builder tools on the MCP server."""

    @mcp.tool()
    def rebuild_ops_summary(dry_run: bool = False) -> str:
        """Rebuild the Ops Summary Report and Cost of Issues from live data.

        Reads UPDATE_Operational Issues, computes weekly pivots by FC,
        writes to Ops Summary Report, and creates/updates the Cost of Issues
        tab with bar chart (cost by FC) and line chart (cost per order trend).

        Args:
            dry_run: If True, preview without writing (default False).

        Returns JSON summary.
        """
        try:
            result = build_ops_summary(dry_run=dry_run)
            return json.dumps(result, indent=2, default=str)
        except Exception as e:
            import traceback
            return json.dumps({"error": str(e), "trace": traceback.format_exc()})

    @mcp.tool()
    def build_shipments_tab(
        weeks_back: int = 12,
        dry_run: bool = False,
    ) -> str:
        """Build Shipments tab from Shopify fulfilled order data.

        Queries Shopify for fulfilled orders, classifies each by FC tag
        prefix (RMFG_*, COG_*, GRIPCA_*), counts per week, and writes
        to the Shipments tab. This feeds into the Ops Summary COUNTIFS
        and SUMPRODUCT formulas.

        Args:
            weeks_back: How many weeks of history to pull (default 12).
            dry_run: If True, preview without writing (default False).

        Returns JSON summary with weekly shipment counts.
        """
        try:
            result = build_shipments_from_shopify(
                weeks_back=weeks_back, dry_run=dry_run,
            )
            return json.dumps(result, indent=2, default=str)
        except Exception as e:
            import traceback
            return json.dumps({"error": str(e), "trace": traceback.format_exc()})
