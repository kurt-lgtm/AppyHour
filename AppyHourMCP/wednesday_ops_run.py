"""Wednesday afternoon ops run: sync Gorgias → sheet, then generate
the cohort carrier report for last week's Tuesday ship cohort.

Target cohort = Tuesday 8 days before the Wednesday this runs.
On Wed 2026-04-29 → target = Tue 2026-04-21.

Outputs:
- Live sheet update (append new tickets, enrich incomplete rows)
- XLSX report at C:/Users/Work/Downloads/cohort_report_YYYYMMDD.xlsx
- Log at AppyHourMCP/logs/wednesday_ops_YYYYMMDD.log (batch wrapper)
"""

from __future__ import annotations

import argparse
import json
import os
import sqlite3
import sys
import time
from collections import Counter, defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path

# Allow imports from AppyHourMCP
sys.path.insert(0, str(Path(__file__).resolve().parent))

from tools.gorgias_sheets_sync import (  # noqa: E402
    SPREADSHEET_ID,
    TAB_NAME,
    enrich_incomplete_rows,
    sync_gorgias_to_sheet,
)


SHIPPING_DB = os.path.expandvars(r"%APPDATA%\AppyHour\shipping.db")
DOWNLOADS = Path(os.path.expandvars(r"%USERPROFILE%\Downloads"))


def target_tuesday(today: date | None = None) -> date:
    """Return the Tuesday 8 days before this Wednesday (the T+8 cohort)."""
    today = today or date.today()
    # Find the Tuesday that is at least 7 days before today
    days_back = (today.weekday() - 1) % 7  # Mon=0 Tue=1
    most_recent_tue = today - timedelta(days=days_back)
    if (today - most_recent_tue).days < 7:
        most_recent_tue = most_recent_tue - timedelta(days=7)
    return most_recent_tue


def cohort_window(tue: date) -> tuple[str, str, str, str]:
    """Return (prev_friday_tag, tuesday_tag, pickup_mon_iso, pickup_wed_iso)."""
    fri = tue - timedelta(days=4)
    mon = tue - timedelta(days=1)
    wed = tue + timedelta(days=1)
    return (
        f"RMFG_{fri:%Y%m%d}",
        f"RMFG_{tue:%Y%m%d}",
        mon.isoformat(),
        wed.isoformat(),
    )


def _norm_carrier(c: str | None) -> str:
    c = (c or "").lower()
    if "laser" in c or "ontrac" in c:
        return "OnTrac/LaserShip"
    if "veho" in c:
        return "Veho"
    if "fedex" in c:
        return "FedEx"
    if "ups" in c:
        return "UPS"
    return c or "?"


def _classify_issue(issue_type: str) -> str:
    s = (issue_type or "").lower()
    if "arrived warm" in s:
        return "Warm"
    if "delayed" in s:
        return "Delay"
    if "lost" in s or "misdeliver" in s:
        return "Lost"
    if "missing" in s:
        return "Missing"
    if "wrong" in s or "substitute" in s:
        return "Wrong"
    if "damaged" in s:
        return "Damaged"
    return "Other"


def run_sync(days: int, dry_run: bool) -> dict:
    """Sync + 3-pass enrich, same pattern as run_gorgias_update.py."""
    delay = 30
    for attempt in range(1, 4):
        try:
            sync = sync_gorgias_to_sheet(days_back=days, dry_run=dry_run)
            break
        except Exception as e:  # noqa: BLE001
            if "429" in str(e) and attempt < 3:
                print(f"[sync] 429, sleeping {delay}s", flush=True)
                time.sleep(delay)
                delay *= 2
                continue
            raise

    enrich_passes = []
    for p in range(1, 4):
        enr = enrich_incomplete_rows(dry_run=dry_run)
        enrich_passes.append({"pass": p, "rows_enriched": enr["rows_enriched"]})
        if enr["rows_enriched"] == 0:
            break

    return {
        "sync": {
            "checked": sync["checked"],
            "new_rows": sync["new_rows"],
            "skipped_duplicate": sync["skipped_duplicate"],
            "skipped_invalid_tag": sync["skipped_invalid_tag"],
        },
        "enrich_passes": enrich_passes,
    }


def build_cohort_report(tue: date, out_path: Path) -> dict:
    """Join sheet rows against Kori DB by pickup_date to build the cohort report."""
    from google_integration import GoogleIntegration

    from tools._gorgias_internal import _load_settings

    fri_tag, tue_tag, pmin, pmax = cohort_window(tue)

    # 1. Read sheet
    settings = _load_settings()
    creds_path = settings.get("google_credentials_path", "")
    if not creds_path or not os.path.exists(creds_path):
        creds_path = str(
            Path(__file__).resolve().parent.parent
            / "shipping-perfomance-review-accd39ac4b78.json"
        )
    gclient = GoogleIntegration(creds_path)
    rows = gclient.read_sheet(SPREADSHEET_ID, f"'{TAB_NAME}'!A:J")
    if not rows:
        raise RuntimeError("No sheet data")
    rows = rows[1:]  # skip header

    # 2. Kori DB lookup for each row's order pickup
    con = sqlite3.connect(SHIPPING_DB)
    cur = con.cursor()

    cohort_rows = []
    for r in rows:
        while len(r) < 10:
            r.append("")
        date_str, reason, order, link, carrier, state, fc, issue, resolution, _ = r
        order_clean = order.lstrip("#").strip() if order else ""
        if not order_clean:
            continue
        cur.execute(
            "SELECT pickup_date, carrier FROM delivery_status "
            "WHERE order_number=? ORDER BY pickup_date DESC LIMIT 1",
            (order_clean,),
        )
        row = cur.fetchone()
        if not row or not row[0]:
            continue
        pickup, db_carrier = row[0], row[1]
        if not (pmin <= pickup <= pmax):
            continue
        cohort_rows.append({
            "order": order_clean,
            "pickup": pickup,
            "carrier": _norm_carrier(carrier or db_carrier),
            "state": state,
            "issue": issue,
            "issue_bucket": _classify_issue(issue),
            "resolution": resolution,
            "gorgias_link": link,
        })

    # 3. Shipment volume by carrier in pickup window
    cur.execute(
        "SELECT carrier, COUNT(*) FROM delivery_status "
        "WHERE pickup_date>=? AND pickup_date<=? GROUP BY carrier",
        (pmin, pmax),
    )
    totals = Counter()
    for c, n in cur.fetchall():
        totals[_norm_carrier(c)] += n
    grand = sum(totals.values())

    # 4. Tickets per carrier per issue
    issues = defaultdict(Counter)
    for row in cohort_rows:
        issues[row["carrier"]][row["issue_bucket"]] += 1

    # 5. Write xlsx
    _write_xlsx(
        out_path,
        tue=tue,
        fri_tag=fri_tag,
        tue_tag=tue_tag,
        pmin=pmin,
        pmax=pmax,
        totals=totals,
        grand=grand,
        issues=issues,
        cohort_rows=cohort_rows,
    )

    summary = {
        "target_tuesday": tue.isoformat(),
        "ship_tags": [fri_tag, tue_tag],
        "pickup_window": f"{pmin}..{pmax}",
        "shipments": grand,
        "tickets_in_cohort": len(cohort_rows),
        "out_path": str(out_path),
    }
    return summary


def _write_xlsx(
    out_path: Path,
    *,
    tue: date,
    fri_tag: str,
    tue_tag: str,
    pmin: str,
    pmax: str,
    totals: Counter,
    grand: int,
    issues: defaultdict,
    cohort_rows: list,
) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    HDR_FILL = PatternFill("solid", fgColor="1F3A5F")
    HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
    TITLE_FONT = Font(bold=True, size=14, color="1F3A5F")
    SUB_FONT = Font(italic=True, size=10, color="666666")
    RED_FILL = PatternFill("solid", fgColor="F8D7DA")
    YEL_FILL = PatternFill("solid", fgColor="FFF3CD")
    GRN_FILL = PatternFill("solid", fgColor="D4EDDA")
    TOT_FILL = PatternFill("solid", fgColor="E9ECEF")
    thin = Side(border_style="thin", color="CCCCCC")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

    def header(ws, row, cols):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = HDR_FILL
            cell.font = HDR_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = BORDER

    wb = Workbook()

    # Sheet 1 — Summary
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = f"Carrier Failure Rates — {tue:%b %d, %Y} cohort"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:J1")
    ws["A2"] = f"Ship tags: {fri_tag} + {tue_tag}   |   Pickup window: {pmin} .. {pmax}"
    ws["A2"].font = SUB_FONT
    ws.merge_cells("A2:J2")
    ws["A3"] = (
        f"Shipments: {grand}   |   Tickets in cohort: {len(cohort_rows)}   |   "
        f"Generated: {datetime.now():%Y-%m-%d %H:%M}"
    )
    ws["A3"].font = SUB_FONT
    ws.merge_cells("A3:J3")
    ws.append([])
    ws.append([])

    hdr = [
        "Carrier",
        "Shipped",
        "% of volume",
        "Warm",
        "Delay",
        "Lost",
        "Missing",
        "Wrong",
        "Total",
        "Rate %",
    ]
    ws.append(hdr)
    header(ws, ws.max_row, len(hdr))

    for c in sorted(totals, key=lambda x: -totals[x]):
        s = totals[c]
        warm = issues[c].get("Warm", 0)
        delay = issues[c].get("Delay", 0)
        lost = issues[c].get("Lost", 0)
        missing = issues[c].get("Missing", 0)
        wrong = issues[c].get("Wrong", 0)
        tot = warm + delay + lost + missing + wrong
        rate = tot / s if s else 0
        ws.append([
            c,
            s,
            s / grand if grand else 0,
            warm,
            delay,
            lost,
            missing,
            wrong,
            tot,
            rate,
        ])
        r = ws.max_row
        ws.cell(row=r, column=3).number_format = "0.0%"
        ws.cell(row=r, column=10).number_format = "0.00%"
        fill = (
            RED_FILL if rate >= 0.01 else (YEL_FILL if rate >= 0.005 else GRN_FILL)
        )
        ws.cell(row=r, column=10).fill = fill
        for col in range(1, 11):
            ws.cell(row=r, column=col).border = BORDER

    # Totals row
    tot_by_issue = {
        k: sum(issues[c].get(k, 0) for c in totals)
        for k in ("Warm", "Delay", "Lost", "Missing", "Wrong")
    }
    ta = sum(tot_by_issue.values())
    ws.append([
        "TOTAL",
        grand,
        1.0,
        tot_by_issue["Warm"],
        tot_by_issue["Delay"],
        tot_by_issue["Lost"],
        tot_by_issue["Missing"],
        tot_by_issue["Wrong"],
        ta,
        ta / grand if grand else 0,
    ])
    r = ws.max_row
    for col in range(1, 11):
        ws.cell(row=r, column=col).font = Font(bold=True)
        ws.cell(row=r, column=col).fill = TOT_FILL
        ws.cell(row=r, column=col).border = BORDER
    ws.cell(row=r, column=3).number_format = "0.0%"
    ws.cell(row=r, column=10).number_format = "0.00%"

    for i, w in enumerate([20, 10, 12, 8, 8, 8, 8, 8, 8, 10], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Sheet 2 — Tickets
    ws2 = wb.create_sheet("Tickets")
    ws2["A1"] = "Cohort Tickets"
    ws2["A1"].font = TITLE_FONT
    ws2.append([])
    ws2.append([])
    hdr = [
        "Order #",
        "Issue",
        "Bucket",
        "Carrier",
        "State",
        "Pickup",
        "Resolution",
        "Gorgias Link",
    ]
    ws2.append(hdr)
    header(ws2, ws2.max_row, len(hdr))
    for row in sorted(cohort_rows, key=lambda x: (x["carrier"], x["pickup"])):
        ws2.append([
            row["order"],
            row["issue"],
            row["issue_bucket"],
            row["carrier"],
            row["state"],
            row["pickup"],
            row["resolution"],
            row["gorgias_link"],
        ])
        r = ws2.max_row
        for col in range(1, 9):
            ws2.cell(row=r, column=col).border = BORDER
            ws2.cell(row=r, column=col).alignment = Alignment(vertical="top", wrap_text=True)
    for i, w in enumerate([12, 40, 10, 20, 15, 12, 20, 50], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    wb.save(out_path)


def main() -> int:
    ap = argparse.ArgumentParser(
        prog="wednesday-ops-run",
        description="Sync Gorgias + generate cohort report for T+8 Tuesday.",
    )
    ap.add_argument("--days", type=int, default=14, help="Sync lookback (default 14)")
    ap.add_argument("--dry-run", action="store_true", help="Preview only, no writes")
    ap.add_argument("--tuesday", help="Override target Tuesday (YYYY-MM-DD)")
    ap.add_argument("--skip-sync", action="store_true", help="Report only, no sync")
    args = ap.parse_args()

    if args.tuesday:
        tue = datetime.strptime(args.tuesday, "%Y-%m-%d").date()
    else:
        tue = target_tuesday()
    print(f"[ops-run] target cohort Tuesday: {tue}", flush=True)

    result: dict = {"target_tuesday": tue.isoformat()}

    if not args.skip_sync:
        print(f"[sync] pulling last {args.days} days...", flush=True)
        result["sync_result"] = run_sync(args.days, args.dry_run)
        print(json.dumps(result["sync_result"], indent=2), flush=True)

    out_path = DOWNLOADS / f"cohort_report_{tue:%Y%m%d}.xlsx"
    print(f"[report] building for {tue}, writing to {out_path}...", flush=True)
    result["report"] = build_cohort_report(tue, out_path)
    print(json.dumps(result["report"], indent=2), flush=True)

    print("\n[done]", flush=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
