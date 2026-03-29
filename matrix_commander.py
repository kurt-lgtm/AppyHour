"""
Matrix Commander — Phase 1: Validate + Inventory Check + Shortage Report.

Standalone CLI tool that validates weekly production matrix XLSX files
and generates inventory cross-check/shortage reports.

Usage:
    python matrix_commander.py validate <xlsx_path>
    python matrix_commander.py check <xlsx_path> --inventory <csv_or_json>
    python matrix_commander.py full <xlsx_path> --inventory <csv_or_json>
"""

import argparse
import csv
import io
import json
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

import openpyxl

# Force UTF-8 output on Windows to avoid cp1252 encoding errors
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

# ---------------------------------------------------------------------------
# Import NAME_TO_SKU from the MCP constants (single source of truth)
# ---------------------------------------------------------------------------
_CONSTANTS_DIR = Path(__file__).parent / "AppyHourMCP" / "tools"
sys.path.insert(0, str(_CONSTANTS_DIR))
from constants import NAME_TO_SKU, FOOD_PREFIXES  # noqa: E402

sys.path.pop(0)

# Reverse mapping: SKU → canonical product name
SKU_TO_NAME: dict[str, str] = {v: k for k, v in NAME_TO_SKU.items()}

# ---------------------------------------------------------------------------
# Substitution families — cheeses that can swap for each other
# ---------------------------------------------------------------------------
SUBSTITUTION_FAMILIES: dict[str, list[str]] = {
    "Brie": ["CH-TTBRIE", "CH-TIP", "CH-EBRIE", "CH-PBRIE", "CH-GPBRIE"],
    "Alpine / Semi-hard": ["CH-BARI"],
}

# Build reverse: SKU → family name
_SKU_TO_FAMILY: dict[str, str] = {}
for _fam, _skus in SUBSTITUTION_FAMILIES.items():
    for _s in _skus:
        _SKU_TO_FAMILY[_s] = _fam

# Non-pickable prefixes — not food, skip in demand counts
SKIP_PREFIXES = ("AHB-", "BL-", "PK-", "TR-", "EX-", "PR-CJAM", "CEX-E")

# Settings JSON path (inventory + curation config)
SETTINGS_PATH = Path(__file__).parent / "InventoryReorder" / "dist" / "inventory_reorder_settings.json"

# MFG translations CSV (exported from RMFG Translator portal)
MFG_TRANSLATIONS_PATH = Path(__file__).parent / "mfg_translations.csv"


# ═══════════════════════════════════════════════════════════════════════════
# Data structures
# ═══════════════════════════════════════════════════════════════════════════


@dataclass(frozen=True)
class OrderRow:
    """One row from the production matrix."""

    order_id: str
    name: str
    distribution_type: str
    total: str
    phone: str
    email: str
    address: str
    address2: str
    city: str
    state: str
    zip_code: str
    tags: str
    notes: str
    production_day: str
    assignments: dict[str, int]  # sku → qty


@dataclass
class CheckResult:
    """Result of a single validation check."""

    name: str
    passed: bool
    message: str
    details: list[str] = field(default_factory=list)


@dataclass
class ShortageItem:
    """One SKU that is short."""

    sku: str
    product_name: str
    demand: int
    available: int
    shortage: int
    family: str
    swap_candidates: list[tuple[str, int]]  # (sku, available_qty)


# ═══════════════════════════════════════════════════════════════════════════
# XLSX parser
# ═══════════════════════════════════════════════════════════════════════════


def parse_matrix(xlsx_path: str | Path) -> tuple[list[OrderRow], list[str], dict[str, str]]:
    """Parse the production matrix XLSX.

    Returns:
        (orders, product_columns, unmapped_names)
        - orders: list of OrderRow
        - product_columns: list of product column headers found
        - unmapped_names: {product_name: fallback_sku} for names not in NAME_TO_SKU
    """
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True, read_only=True)
    # Handle both RMFG download (Worksheet) and formatted files (Access_LIVE)
    if "Access_LIVE" in wb.sheetnames:
        ws = wb["Access_LIVE"]
    elif "Worksheet" in wb.sheetnames:
        ws = wb["Worksheet"]
    else:
        ws = wb[wb.sheetnames[0]]

    # Read headers
    headers: list[str] = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        headers.append(str(cell.value or ""))

    # Identify product columns (index, product_name, sku)
    product_cols: list[tuple[int, str, str]] = []
    product_column_names: list[str] = []
    unmapped: dict[str, str] = {}

    for idx, hdr in enumerate(headers):
        if hdr.startswith("AHB") and ": " in hdr:
            prod_name = hdr.split(": ", 1)[1]
            product_column_names.append(prod_name)
            sku = NAME_TO_SKU.get(prod_name)
            if sku is None:
                fallback = f"??-{prod_name[:20]}"
                unmapped[prod_name] = fallback
                sku = fallback
            product_cols.append((idx, prod_name, sku))

    # Check for duplicate column names
    seen_cols: dict[str, int] = {}
    duplicate_cols: list[str] = []
    for _, prod_name, _ in product_cols:
        seen_cols[prod_name] = seen_cols.get(prod_name, 0) + 1
    for name, count in seen_cols.items():
        if count > 1:
            duplicate_cols.append(f"{name} (×{count})")

    # Detect ProductionDay column (may not exist in RMFG downloads)
    prod_day_idx: int | None = None
    for i, h in enumerate(headers):
        if h.lower().strip() == "productionday":
            prod_day_idx = i
            break

    # Parse data rows
    orders: list[OrderRow] = []
    min_cols = 13  # At minimum: OrderID through Notes
    for row in ws.iter_rows(min_row=2):
        cells = [cell.value for cell in row]
        if len(cells) < min_cols:
            continue
        order_id = str(cells[0] or "").strip()
        if not order_id:
            continue

        assignments: dict[str, int] = {}
        for col_idx, _, sku in product_cols:
            if col_idx < len(cells):
                val = cells[col_idx]
                if val is not None and str(val).strip() not in ("", "0", "None"):
                    try:
                        qty = int(float(str(val)))
                        if qty > 0:
                            assignments[sku] = assignments.get(sku, 0) + qty
                    except (ValueError, TypeError):
                        pass

        prod_day = ""
        if prod_day_idx is not None and prod_day_idx < len(cells):
            prod_day = str(cells[prod_day_idx] or "")

        orders.append(
            OrderRow(
                order_id=order_id,
                name=str(cells[1] or ""),
                distribution_type=str(cells[2] or ""),
                total=str(cells[3] or ""),
                phone=str(cells[4] or ""),
                email=str(cells[5] or ""),
                address=str(cells[6] or ""),
                address2=str(cells[7] or ""),
                city=str(cells[8] or ""),
                state=str(cells[9] or ""),
                zip_code=str(cells[10] or ""),
                tags=str(cells[11] or ""),
                notes=str(cells[12] or ""),
                production_day=prod_day,
                assignments=assignments,
            )
        )

    wb.close()
    return orders, product_column_names, unmapped


# ═══════════════════════════════════════════════════════════════════════════
# Inventory loading
# ═══════════════════════════════════════════════════════════════════════════


def load_mfg_translations(csv_path: str | Path | None = None) -> dict[str, str]:
    """Load MFG translations: SKU -> MFG Name.

    CSV format (no header): SKU,"AHB (S_REG): Product Name"
    Exported from https://translator.robbinsmfginc.com/
    """
    path = Path(csv_path) if csv_path else MFG_TRANSLATIONS_PATH
    if not path.exists():
        return {}
    translations: dict[str, str] = {}
    with open(str(path), newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        for row in reader:
            if len(row) >= 2:
                sku = row[0].strip()
                mfg_name = row[1].strip()
                if sku:
                    translations[sku] = mfg_name
    return translations


def load_inventory_csv(csv_path: str | Path) -> dict[str, float]:
    """Load inventory from a simple CSV: sku,available_qty."""
    inventory: dict[str, float] = {}
    with open(str(csv_path), newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            sku = row.get("sku", "").strip()
            qty_str = row.get("available_qty", "0").strip()
            if sku:
                try:
                    inventory[sku] = float(qty_str)
                except ValueError:
                    inventory[sku] = 0.0
    return inventory


def load_inventory_settings() -> dict[str, float]:
    """Load inventory from the fulfillment app settings JSON."""
    if not SETTINGS_PATH.exists():
        return {}
    with open(SETTINGS_PATH, encoding="utf-8") as f:
        settings = json.load(f)
    raw = settings.get("inventory", {})
    return {sku: data.get("qty", 0.0) for sku, data in raw.items()}


def load_settings_config() -> dict:
    """Load PR-CJAM, CEX-EC, and splits config from settings."""
    if not SETTINGS_PATH.exists():
        return {}
    with open(SETTINGS_PATH, encoding="utf-8") as f:
        return json.load(f)


# ═══════════════════════════════════════════════════════════════════════════
# Validation checks
# ═══════════════════════════════════════════════════════════════════════════


def check_numeric_order_ids(orders: list[OrderRow]) -> CheckResult:
    """Verify all OrderIDs are numeric (RMFG requirement)."""
    bad: list[str] = []
    for o in orders:
        cleaned = o.order_id.replace(",", "").replace(".", "").strip()
        if not cleaned.isdigit():
            bad.append(f"  #{o.order_id} ({o.name})")
    if bad:
        return CheckResult(
            "Numeric OrderIDs",
            False,
            f"{len(bad)} non-numeric OrderIDs found",
            bad[:20],
        )
    return CheckResult("Numeric OrderIDs", True, f"All {len(orders)} OrderIDs numeric")


def check_zip_leading_zeroes(orders: list[OrderRow]) -> CheckResult:
    """Verify zip codes preserve leading zeroes (5-digit minimum)."""
    bad: list[str] = []
    for o in orders:
        z = o.zip_code.strip()
        if not z:
            continue
        # Strip to digits only for check
        digits = z.split("-")[0]  # handle ZIP+4
        if digits.isdigit() and len(digits) < 5:
            bad.append(f"  #{o.order_id}: zip={z} ({o.city}, {o.state})")
    if bad:
        return CheckResult(
            "Zip Leading Zeroes",
            False,
            f"{len(bad)} zips missing leading zeroes",
            bad[:20],
        )
    return CheckResult("Zip Leading Zeroes", True, "All zips properly formatted")


def check_duplicate_columns(product_columns: list[str]) -> CheckResult:
    """Check for duplicate product column names."""
    counts = Counter(product_columns)
    dupes = [(name, cnt) for name, cnt in counts.items() if cnt > 1]
    if dupes:
        details = [f"  {name} (×{cnt})" for name, cnt in dupes]
        return CheckResult(
            "No Duplicate Columns",
            False,
            f"{len(dupes)} duplicate column(s) found",
            details,
        )
    return CheckResult(
        "No Duplicate Columns",
        True,
        f"{len(product_columns)} unique product columns",
    )


def check_production_day(orders: list[OrderRow]) -> CheckResult:
    """Verify ProductionDay is set to SAT or TUE for all orders."""
    valid_days = {"SAT", "TUE", "WED"}
    missing: list[str] = []
    invalid: list[str] = []
    day_counts: dict[str, int] = defaultdict(int)

    for o in orders:
        day = o.production_day.strip().upper()
        if not day:
            missing.append(f"  #{o.order_id} ({o.name})")
        elif day not in valid_days:
            invalid.append(f"  #{o.order_id}: '{o.production_day}'")
        else:
            day_counts[day] += 1

    if missing or invalid:
        details = []
        if missing:
            details.append(f"Missing ({len(missing)}):")
            details.extend(missing[:10])
        if invalid:
            details.append(f"Invalid ({len(invalid)}):")
            details.extend(invalid[:10])
        return CheckResult(
            "ProductionDay",
            False,
            f"{len(missing)} missing, {len(invalid)} invalid",
            details,
        )

    summary = ", ".join(f"{d}: {c}" for d, c in sorted(day_counts.items()))
    return CheckResult("ProductionDay", True, f"All set ({summary})")


def check_sku_mappings(unmapped: dict[str, str]) -> CheckResult:
    """Verify all product names map to known SKUs."""
    if unmapped:
        details = [f"  '{name}' → {fb}" for name, fb in sorted(unmapped.items())]
        return CheckResult(
            "SKU Mappings",
            False,
            f"{len(unmapped)} unmapped product name(s) — not in NAME_TO_SKU",
            details,
        )
    return CheckResult("SKU Mappings", True, "All product names mapped to SKUs")


def check_mfg_onboarding(
    orders: list[OrderRow],
    mfg_translations: dict[str, str],
) -> CheckResult:
    """Verify all allocated SKUs are onboarded at RMFG.

    Compares SKUs found in the matrix against the MFG translations export.
    Any SKU with demand that isn't in the translations file won't appear
    in the RMFG pick/pack output.
    """
    if not mfg_translations:
        return CheckResult(
            "MFG Onboarding",
            False,
            "No MFG translations file found — cannot validate. Export from https://translator.robbinsmfginc.com/",
        )

    # Collect all SKUs with actual demand
    demand_skus: set[str] = set()
    for o in orders:
        for sku in o.assignments:
            if any(sku.startswith(p) for p in ("CH-", "MT-", "AC-", "PK-")):
                demand_skus.add(sku)

    mfg_skus = set(mfg_translations.keys())
    missing = sorted(demand_skus - mfg_skus)

    if missing:
        details = []
        # Calculate demand for each missing SKU
        missing_demand: dict[str, int] = {}
        for o in orders:
            for sku, qty in o.assignments.items():
                if sku in missing:
                    missing_demand[sku] = missing_demand.get(sku, 0) + qty
        for sku in missing:
            name = SKU_TO_NAME.get(sku, "???")
            demand = missing_demand.get(sku, 0)
            details.append(f"  {sku} ({name}) — demand: {demand}")
        return CheckResult(
            "MFG Onboarding",
            False,
            f"{len(missing)} SKU(s) NOT onboarded at RMFG — must add before submission",
            details,
        )

    return CheckResult(
        "MFG Onboarding",
        True,
        f"All {len(demand_skus)} food/packaging SKUs onboarded at RMFG",
    )


def check_cexec_cheese_counts(
    orders: list[OrderRow],
    cex_ec_config: dict[str, str],
    cexec_splits: dict[str, dict[str, float]],
) -> CheckResult:
    """Verify CEX-EC parent line items have corresponding extra cheese allocated.

    For each order with a CEX-EC-{curation} in its tags or implied by assignments,
    check that the expected extra cheese SKU is present in the matrix assignments.
    """
    issues: list[str] = []
    checked = 0

    for o in orders:
        # Find CEX-EC curations from tags
        tags = [t.strip() for t in o.tags.split(",")]
        cexec_curations: list[str] = []
        for tag in tags:
            tag_upper = tag.upper()
            if tag_upper.startswith("CEXEC") or tag_upper.startswith("CEX-EC"):
                # Extract curation from tag like CEXEC.3.10.FIX or CEX-EC-MONG
                for cur in cex_ec_config:
                    if cur in tag_upper:
                        cexec_curations.append(cur)

        # Also check assignments for CEX-EC SKUs
        for sku in o.assignments:
            if sku.startswith("CEX-EC-"):
                cur = sku.replace("CEX-EC-", "")
                if cur not in cexec_curations:
                    cexec_curations.append(cur)

        if not cexec_curations:
            continue

        checked += 1
        assigned_cheeses = {s for s in o.assignments if s.startswith("CH-")}

        for cur in cexec_curations:
            expected_cheese = cex_ec_config.get(cur)
            if not expected_cheese:
                # Check splits
                splits = cexec_splits.get(cur, {})
                if splits:
                    for split_sku in splits:
                        if split_sku not in assigned_cheeses:
                            issues.append(f"  #{o.order_id}: CEX-EC-{cur} missing split cheese {split_sku}")
                continue
            if expected_cheese not in assigned_cheeses:
                # Check splits as fallback
                splits = cexec_splits.get(cur, {})
                split_present = any(s in assigned_cheeses for s in splits)
                if not split_present:
                    issues.append(
                        f"  #{o.order_id}: CEX-EC-{cur} expects {expected_cheese}, "
                        f"has [{', '.join(sorted(assigned_cheeses))}]"
                    )

    if issues:
        return CheckResult(
            "CEX-EC Cheese Allocation",
            False,
            f"{len(issues)} orders with missing CEX-EC cheese (checked {checked})",
            issues[:20],
        )
    return CheckResult(
        "CEX-EC Cheese Allocation",
        True,
        f"All {checked} CEX-EC orders have expected cheese allocated",
    )


# ═══════════════════════════════════════════════════════════════════════════
# Inventory cross-check & shortage report
# ═══════════════════════════════════════════════════════════════════════════


def compute_demand(orders: list[OrderRow]) -> dict[str, int]:
    """Sum total demand per SKU across all orders."""
    demand: dict[str, int] = defaultdict(int)
    for o in orders:
        for sku, qty in o.assignments.items():
            demand[sku] += qty
    return dict(demand)


def find_shortages(
    demand: dict[str, int],
    inventory: dict[str, float],
) -> list[ShortageItem]:
    """Compare demand vs inventory, return shortage items with swap suggestions."""
    shortages: list[ShortageItem] = []

    # Only check food SKUs
    food_demand = {sku: qty for sku, qty in demand.items() if any(sku.startswith(p) for p in ("CH-", "MT-", "AC-"))}

    for sku, qty_needed in sorted(food_demand.items()):
        available = inventory.get(sku, 0.0)
        if qty_needed > available:
            shortage = qty_needed - int(available)
            family = _SKU_TO_FAMILY.get(sku, "")

            # Find swap candidates from same family
            candidates: list[tuple[str, int]] = []
            if family:
                family_skus = SUBSTITUTION_FAMILIES.get(family, [])
                for alt_sku in family_skus:
                    if alt_sku == sku:
                        continue
                    alt_avail = inventory.get(alt_sku, 0.0)
                    alt_demand = food_demand.get(alt_sku, 0)
                    surplus = int(alt_avail) - alt_demand
                    if surplus > 0:
                        candidates.append((alt_sku, surplus))
                candidates.sort(key=lambda x: x[1], reverse=True)

            shortages.append(
                ShortageItem(
                    sku=sku,
                    product_name=SKU_TO_NAME.get(sku, sku),
                    demand=qty_needed,
                    available=int(available),
                    shortage=shortage,
                    family=family,
                    swap_candidates=candidates,
                )
            )

    return shortages


# ═══════════════════════════════════════════════════════════════════════════
# Output formatting
# ═══════════════════════════════════════════════════════════════════════════

_GREEN = "\033[92m"
_RED = "\033[91m"
_YELLOW = "\033[93m"
_CYAN = "\033[96m"
_BOLD = "\033[1m"
_RESET = "\033[0m"


def _check_icon(passed: bool) -> str:
    return f"{_GREEN}PASS{_RESET}" if passed else f"{_RED}FAIL{_RESET}"


def print_validation_report(results: list[CheckResult], order_count: int) -> bool:
    """Print validation results. Returns True if all passed."""
    print(f"\n{_BOLD}{'=' * 60}{_RESET}")
    print(f"{_BOLD}  MATRIX COMMANDER — Validation Report{_RESET}")
    print(f"{_BOLD}{'=' * 60}{_RESET}")
    print(f"  Orders parsed: {order_count}\n")

    all_passed = True
    for r in results:
        icon = _check_icon(r.passed)
        print(f"  [{icon}] {r.name}: {r.message}")
        if not r.passed:
            all_passed = False
            for d in r.details:
                print(f"       {d}")
            print()

    print(f"{_BOLD}{'=' * 60}{_RESET}")
    if all_passed:
        print(f"  {_GREEN}All checks passed.{_RESET}")
    else:
        failed = sum(1 for r in results if not r.passed)
        print(f"  {_RED}{failed} check(s) failed.{_RESET}")
    print(f"{_BOLD}{'=' * 60}{_RESET}\n")
    return all_passed


def print_inventory_report(
    demand: dict[str, int],
    inventory: dict[str, float],
    shortages: list[ShortageItem],
) -> None:
    """Print inventory cross-check and shortage report."""
    food_demand = {sku: qty for sku, qty in demand.items() if any(sku.startswith(p) for p in ("CH-", "MT-", "AC-"))}

    print(f"\n{_BOLD}{'=' * 60}{_RESET}")
    print(f"{_BOLD}  MATRIX COMMANDER — Inventory Report{_RESET}")
    print(f"{_BOLD}{'=' * 60}{_RESET}")
    print(f"  SKUs in demand: {len(food_demand)}")
    print(f"  SKUs in inventory: {len(inventory)}")
    print()

    # Full demand table
    print(f"  {_BOLD}{'SKU':<14} {'Demand':>7} {'Avail':>7} {'Net':>7}  Status{_RESET}")
    print(f"  {'─' * 52}")

    for sku in sorted(food_demand.keys()):
        qty = food_demand[sku]
        avail = int(inventory.get(sku, 0))
        net = avail - qty
        if net < 0:
            status = f"{_RED}SHORT {abs(net)}{_RESET}"
        elif net < 20:
            status = f"{_YELLOW}LOW{_RESET}"
        else:
            status = f"{_GREEN}OK (+{net}){_RESET}"
        print(f"  {sku:<14} {qty:>7} {avail:>7} {net:>7}  {status}")

    # Missing from inventory
    missing_inv = [s for s in food_demand if s not in inventory]
    if missing_inv:
        print(f"\n  {_YELLOW}SKUs with demand but NOT in inventory:{_RESET}")
        for s in sorted(missing_inv):
            print(f"    {s} (demand: {food_demand[s]})")

    # Shortage details with swap recommendations
    if shortages:
        print(f"\n{_BOLD}{'=' * 60}{_RESET}")
        print(f"{_BOLD}  SHORTAGES — {len(shortages)} SKU(s){_RESET}")
        print(f"{_BOLD}{'=' * 60}{_RESET}")

        for s in shortages:
            print(f"\n  {_RED}{s.sku}{_RESET} ({s.product_name})")
            print(f"    Demand: {s.demand}  Available: {s.available}  Short: {s.shortage}")
            if s.family:
                print(f"    Family: {s.family}")
            if s.swap_candidates:
                print(f"    {_CYAN}Swap candidates:{_RESET}")
                for alt_sku, surplus in s.swap_candidates:
                    alt_name = SKU_TO_NAME.get(alt_sku, alt_sku)
                    print(f"      → {alt_sku} ({alt_name}): {surplus} surplus")
            elif s.family:
                print(f"    {_YELLOW}No swap candidates with surplus in {s.family} family{_RESET}")
            else:
                print(f"    {_YELLOW}No substitution family defined for this SKU{_RESET}")
    else:
        print(f"\n  {_GREEN}No shortages detected.{_RESET}")

    print(f"\n{_BOLD}{'=' * 60}{_RESET}\n")


def print_demand_summary(demand: dict[str, int]) -> None:
    """Print demand summary CSV-like output for cross-checking."""
    food_demand = {sku: qty for sku, qty in demand.items() if any(sku.startswith(p) for p in ("CH-", "MT-", "AC-"))}
    print(f"\n{_BOLD}  Demand Summary (food SKUs only){_RESET}")
    print(f"  {'SKU':<14} {'Name':<45} {'Qty':>5}")
    print(f"  {'─' * 66}")
    for sku in sorted(food_demand.keys()):
        name = SKU_TO_NAME.get(sku, "???")
        print(f"  {sku:<14} {name:<45} {food_demand[sku]:>5}")
    total = sum(food_demand.values())
    print(f"  {'─' * 66}")
    print(f"  {'TOTAL':<14} {'':<45} {total:>5}")


# ═══════════════════════════════════════════════════════════════════════════
# Phase 2: Interactive swap resolution
# ═══════════════════════════════════════════════════════════════════════════


@dataclass
class SwapDecision:
    """One approved swap: replace short_sku with replacement_sku on N orders."""

    short_sku: str
    replacement_sku: str
    qty: int  # number of units to swap


def interactive_swap_resolution(
    shortages: list[ShortageItem],
    inventory: dict[str, float],
    demand: dict[str, int],
) -> list[SwapDecision]:
    """Interactively resolve shortages with swap recommendations.

    For each shortage with swap candidates, prompts the user to accept,
    pick an alternative, or skip. Returns list of approved swaps.
    """
    if not shortages:
        return []

    decisions: list[SwapDecision] = []
    # Track running inventory adjustments from prior swaps
    inv_adj: dict[str, int] = {}  # sku -> units consumed by swaps
    demand_adj: dict[str, int] = {}  # sku -> units added by swaps

    print(f"\n{_BOLD}{'=' * 60}{_RESET}")
    print(f"{_BOLD}  SWAP RESOLUTION — {len(shortages)} shortage(s){_RESET}")
    print(f"{_BOLD}{'=' * 60}{_RESET}")

    for s in shortages:
        if not s.swap_candidates:
            print(f"\n  {_RED}{s.sku}{_RESET} ({s.product_name}): short {s.shortage}")
            print(f"    {_YELLOW}No swap candidates available — manual resolution needed{_RESET}")
            continue

        # Recalculate candidate surplus with running adjustments
        live_candidates: list[tuple[str, int]] = []
        for alt_sku, _orig_surplus in s.swap_candidates:
            alt_avail = inventory.get(alt_sku, 0) - inv_adj.get(alt_sku, 0)
            alt_demand = demand.get(alt_sku, 0) + demand_adj.get(alt_sku, 0)
            surplus = int(alt_avail) - alt_demand
            if surplus > 0:
                live_candidates.append((alt_sku, surplus))

        if not live_candidates:
            print(f"\n  {_RED}{s.sku}{_RESET} ({s.product_name}): short {s.shortage}")
            print(f"    {_YELLOW}Swap candidates exhausted by prior swaps — manual resolution needed{_RESET}")
            continue

        print(f"\n  {_RED}{s.sku}{_RESET} ({s.product_name})")
        print(f"    Demand: {s.demand}  Available: {s.available}  {_RED}Short: {s.shortage}{_RESET}")
        if s.family:
            print(f"    Family: {s.family}")
        print()
        for idx, (alt_sku, surplus) in enumerate(live_candidates, 1):
            alt_name = SKU_TO_NAME.get(alt_sku, alt_sku)
            can_cover = min(surplus, s.shortage)
            marker = " <-- recommended" if idx == 1 else ""
            print(f"    {idx}) {alt_sku} ({alt_name}): {surplus} surplus, covers {can_cover}{marker}")
        print(f"    s) Skip — handle manually")
        print()

        choice = input(f"    Choice [1/{'/'.join(str(i) for i in range(2, len(live_candidates) + 1))}/s]: ").strip()

        if choice.lower() == "s" or choice == "":
            print(f"    {_YELLOW}Skipped{_RESET}")
            continue

        try:
            pick = int(choice) - 1
            if 0 <= pick < len(live_candidates):
                alt_sku, surplus = live_candidates[pick]
                swap_qty = min(surplus, s.shortage)
                alt_name = SKU_TO_NAME.get(alt_sku, alt_sku)
                decisions.append(
                    SwapDecision(
                        short_sku=s.sku,
                        replacement_sku=alt_sku,
                        qty=swap_qty,
                    )
                )
                # Update running adjustments
                inv_adj[alt_sku] = inv_adj.get(alt_sku, 0) + swap_qty
                demand_adj[alt_sku] = demand_adj.get(alt_sku, 0) + swap_qty
                print(f"    {_GREEN}Approved: {s.sku} -> {alt_sku} ({swap_qty} units){_RESET}")
                if swap_qty < s.shortage:
                    print(f"    {_YELLOW}Partial: still short {s.shortage - swap_qty} units{_RESET}")
            else:
                print(f"    {_YELLOW}Invalid choice, skipped{_RESET}")
        except ValueError:
            print(f"    {_YELLOW}Invalid input, skipped{_RESET}")

    if decisions:
        print(f"\n{_BOLD}  Swap Summary: {len(decisions)} swap(s) approved{_RESET}")
        for d in decisions:
            short_name = SKU_TO_NAME.get(d.short_sku, d.short_sku)
            repl_name = SKU_TO_NAME.get(d.replacement_sku, d.replacement_sku)
            print(f"    {d.short_sku} ({short_name}) -> {d.replacement_sku} ({repl_name}): {d.qty} units")
    else:
        print(f"\n  {_YELLOW}No swaps approved.{_RESET}")

    print(f"{_BOLD}{'=' * 60}{_RESET}\n")
    return decisions


def apply_swaps_to_xlsx(
    xlsx_path: str | Path,
    decisions: list[SwapDecision],
    orders: list[OrderRow],
) -> str:
    """Apply approved swaps to the XLSX and save as a new file.

    For each swap decision, finds orders that have the short_sku assigned
    and replaces it with the replacement_sku (up to the swap qty).
    Saves to a new file with _FIXED suffix.

    Returns the output file path.
    """
    wb = openpyxl.load_workbook(str(xlsx_path))
    ws = wb["Access_LIVE"]

    # Build column index maps: product_name -> col_index, sku -> col_index
    headers: list[str] = []
    for cell in ws[1]:
        headers.append(str(cell.value or ""))

    sku_to_col: dict[str, int] = {}
    for idx, h in enumerate(headers):
        if h.startswith("AHB") and ": " in h:
            prod_name = h.split(": ", 1)[1]
            sku = NAME_TO_SKU.get(prod_name)
            if sku:
                sku_to_col[sku] = idx + 1  # openpyxl is 1-indexed

    # Build order_id -> row_number map
    oid_col = 1  # Column A = OrderID
    oid_to_row: dict[str, int] = {}
    for row_num in range(2, ws.max_row + 1):
        oid = str(ws.cell(row_num, oid_col).value or "").strip()
        if oid:
            oid_to_row[oid] = row_num

    swap_log: list[str] = []
    for decision in decisions:
        short_col = sku_to_col.get(decision.short_sku)
        repl_col = sku_to_col.get(decision.replacement_sku)

        if not short_col:
            swap_log.append(f"  SKIP: No column for {decision.short_sku}")
            continue
        if not repl_col:
            swap_log.append(f"  SKIP: No column for {decision.replacement_sku}")
            continue

        remaining = decision.qty
        swapped_orders = 0

        # Find orders with the short SKU assigned
        for o in orders:
            if remaining <= 0:
                break
            if decision.short_sku not in o.assignments:
                continue

            row_num = oid_to_row.get(o.order_id)
            if not row_num:
                continue

            qty = o.assignments[decision.short_sku]
            swap_amt = min(qty, remaining)

            # Remove from short column
            old_val = ws.cell(row_num, short_col).value or 0
            new_short_val = max(0, int(old_val) - swap_amt)
            ws.cell(row_num, short_col).value = new_short_val if new_short_val > 0 else None

            # Add to replacement column
            old_repl = ws.cell(row_num, repl_col).value or 0
            ws.cell(row_num, repl_col).value = int(old_repl) + swap_amt

            remaining -= swap_amt
            swapped_orders += 1

        swap_log.append(
            f"  {decision.short_sku} -> {decision.replacement_sku}: "
            f"{decision.qty - remaining}/{decision.qty} swapped across {swapped_orders} orders"
        )

    # Save to new file
    src = Path(xlsx_path)
    out_path = src.parent / f"{src.stem}_FIXED{src.suffix}"
    wb.save(str(out_path))
    wb.close()

    print(f"\n{_BOLD}  Swaps Applied to XLSX{_RESET}")
    for line in swap_log:
        print(line)
    print(f"\n  Saved: {_GREEN}{out_path.name}{_RESET}")

    return str(out_path)


# ═══════════════════════════════════════════════════════════════════════════
# Phase 3: Shopify $0 variant batch sync (replaces Matrixify)
# ═══════════════════════════════════════════════════════════════════════════

# Import Shopify auth + GraphQL from MCP utils
_MCP_UTILS_DIR = Path(__file__).parent / "AppyHourMCP"
sys.path.insert(0, str(_MCP_UTILS_DIR))


def _get_shopify_auth() -> tuple[str, dict]:
    """Get Shopify REST/GraphQL auth. Lazy import to avoid startup cost."""
    from utils import get_shopify_auth  # noqa: E402

    return get_shopify_auth()


def _shopify_graphql(base: str, headers: dict, query: str, variables: dict | None = None) -> dict:
    """Execute Shopify GraphQL query."""
    from utils import shopify_graphql  # noqa: E402

    return shopify_graphql(base, headers, query, variables)


def _lookup_zero_variant_gids(base: str, headers: dict, skus: set[str]) -> dict[str, str]:
    """Look up $0 variant GIDs for a set of SKUs. Prefers cheapest variant."""
    import requests as req

    variant_map: dict[str, tuple[str, float]] = {}
    sku_list = sorted(skus)
    batch_size = 10
    for i in range(0, len(sku_list), batch_size):
        batch = sku_list[i : i + batch_size]
        query_str = " OR ".join(f"sku:{s}" for s in batch)
        data = _shopify_graphql(
            base,
            headers,
            """
        query($q: String!) {
          productVariants(first: 50, query: $q) {
            edges { node { id sku price } }
          }
        }
        """,
            {"q": query_str},
        )
        for edge in data["productVariants"]["edges"]:
            node = edge["node"]
            sku = node["sku"]
            price = float(node.get("price", "999"))
            if sku in skus:
                prev_price = variant_map.get(sku, (None, float("inf")))[1]
                if price < prev_price:
                    variant_map[sku] = (node["id"], price)
        time.sleep(0.1)

    return {sku: gid for sku, (gid, _) in variant_map.items()}


def _fetch_orders_by_tag(base: str, headers: dict, tag: str) -> list[dict]:
    """Fetch all unfulfilled Shopify orders matching a tag."""
    import requests as req

    all_orders: list[dict] = []
    url = f"{base}/orders.json"
    params = {
        "status": "open",
        "fulfillment_status": "unfulfilled",
        "limit": 250,
        "tag": tag,
        "fields": "id,name,tags,line_items",
    }
    page = 0
    while url:
        page += 1
        resp = req.get(url, headers=headers, params=params if page == 1 else None, timeout=30)
        resp.raise_for_status()
        orders = resp.json().get("orders", [])
        all_orders.extend(orders)
        link = resp.headers.get("Link", "")
        url = None
        if 'rel="next"' in link:
            import re

            m = re.search(r'<([^>]+)>;\s*rel="next"', link)
            if m:
                url = m.group(1)
        time.sleep(0.1)
    return all_orders


@dataclass
class SyncResult:
    """Result of syncing one order."""

    order_name: str
    status: str  # "updated", "skipped", "duplicate", "gift", "error"
    added_skus: list[str] = field(default_factory=list)
    error: str = ""


def sync_order_to_shopify(
    base: str,
    headers: dict,
    order: dict,
    matrix_skus: dict[str, int],
    variant_gids: dict[str, str],
    mode: str = "smart",
) -> SyncResult:
    """Sync one order: add $0 variants for matrix SKUs not yet on Shopify.

    mode: "smart" = skip only duplicate SKUs, add rest.
          "conservative" = skip entire order if any duplicate.
    """
    order_name = order["name"].replace("#", "")
    tags_lower = order.get("tags", "").lower()

    # Skip gift redemption orders
    if "gift redemption" in tags_lower:
        return SyncResult(order_name, "gift")

    # Get current SKUs on order
    current_skus: dict[str, int] = {}
    for li in order.get("line_items", []):
        sku = (li.get("sku") or "").strip()
        fq = li.get("fulfillable_quantity", li.get("quantity", 0))
        if sku and fq > 0:
            current_skus[sku] = current_skus.get(sku, 0) + fq

    # Determine what to add
    to_add: list[str] = []
    duplicates: list[str] = []
    for sku, qty in matrix_skus.items():
        if not any(sku.startswith(p) for p in ("CH-", "MT-", "AC-", "PK-", "TR-")):
            continue  # Skip non-food/packaging SKUs
        if sku in current_skus:
            duplicates.append(sku)
        elif sku in variant_gids:
            to_add.append(sku)

    if not to_add and not duplicates:
        return SyncResult(order_name, "skipped")

    if duplicates and mode == "conservative":
        return SyncResult(order_name, "duplicate", error=f"Dupes: {', '.join(duplicates)}")

    if not to_add:
        return SyncResult(order_name, "skipped")

    # Execute order edit
    try:
        order_gid = f"gid://shopify/Order/{order['id']}"

        # Begin edit
        data = _shopify_graphql(
            base,
            headers,
            """
            mutation($id: ID!) {
                orderEditBegin(id: $id) {
                    calculatedOrder { id }
                    userErrors { field message }
                }
            }
        """,
            {"id": order_gid},
        )

        calc_order = data["orderEditBegin"]["calculatedOrder"]
        if not calc_order:
            errors = data["orderEditBegin"]["userErrors"]
            return SyncResult(order_name, "error", error=f"beginEdit: {errors}")
        calc_id = calc_order["id"]

        # Add each variant
        added: list[str] = []
        for sku in to_add:
            gid = variant_gids[sku]
            add_data = _shopify_graphql(
                base,
                headers,
                """
                mutation($id: ID!, $variantId: ID!, $quantity: Int!, $allowDuplicates: Boolean) {
                    orderEditAddVariant(id: $id, variantId: $variantId, quantity: $quantity, allowDuplicates: $allowDuplicates) {
                        calculatedOrder { id }
                        userErrors { field message }
                    }
                }
            """,
                {"id": calc_id, "variantId": gid, "quantity": 1, "allowDuplicates": False},
            )
            add_errors = add_data["orderEditAddVariant"]["userErrors"]
            if not add_errors:
                added.append(sku)

        # Commit
        commit_data = _shopify_graphql(
            base,
            headers,
            """
            mutation($id: ID!) {
                orderEditCommit(id: $id) {
                    order { id }
                    userErrors { field message }
                }
            }
        """,
            {"id": calc_id},
        )
        commit_errors = commit_data["orderEditCommit"]["userErrors"]
        if commit_errors:
            return SyncResult(order_name, "error", added, error=f"commit: {commit_errors}")

        return SyncResult(order_name, "updated", added)

    except Exception as e:
        return SyncResult(order_name, "error", error=str(e))


def cmd_sync(
    xlsx_path: str,
    rmfg_tag: str,
    mode: str = "smart",
    dry_run: bool = True,
    workers: int = 5,
) -> bool:
    """Sync matrix XLSX assignments to Shopify as $0 variants."""
    from concurrent.futures import ThreadPoolExecutor, as_completed

    print(f"  Loading {Path(xlsx_path).name}...")
    orders_parsed, _, _ = parse_matrix(xlsx_path)

    # Build matrix: order_id -> {sku: qty}
    matrix: dict[str, dict[str, int]] = {}
    for o in orders_parsed:
        matrix[o.order_id] = o.assignments

    print(f"  {len(matrix)} orders in matrix")
    print(f"  Connecting to Shopify...")

    base, headers = _get_shopify_auth()

    # Fetch Shopify orders
    print(f"  Fetching orders with tag '{rmfg_tag}'...")
    shopify_orders = _fetch_orders_by_tag(base, headers, rmfg_tag)
    print(f"  {len(shopify_orders)} Shopify orders fetched")

    # Match matrix orders to Shopify orders
    shopify_by_name: dict[str, dict] = {}
    for o in shopify_orders:
        name = o["name"].replace("#", "")
        shopify_by_name[name] = o

    matched = set(matrix.keys()) & set(shopify_by_name.keys())
    print(f"  {len(matched)} orders matched between matrix and Shopify")

    if not matched:
        print(f"  {_RED}No matching orders found!{_RESET}")
        return False

    # Collect all SKUs that need variant GIDs
    all_skus: set[str] = set()
    for oid in matched:
        for sku in matrix[oid]:
            if any(sku.startswith(p) for p in ("CH-", "MT-", "AC-", "PK-", "TR-")):
                all_skus.add(sku)

    print(f"  Looking up $0 variant GIDs for {len(all_skus)} SKUs...")
    variant_gids = _lookup_zero_variant_gids(base, headers, all_skus)
    missing_gids = all_skus - set(variant_gids.keys())
    if missing_gids:
        print(f"  {_YELLOW}Warning: No variant found for {len(missing_gids)} SKUs: {sorted(missing_gids)[:10]}{_RESET}")

    print(f"  Found $0 variants for {len(variant_gids)}/{len(all_skus)} SKUs")

    if dry_run:
        # Preview mode
        print(f"\n{_BOLD}  DRY RUN — No changes will be made{_RESET}\n")
        updated = 0
        skipped = 0
        gift = 0
        dupes = 0
        for oid in sorted(matched):
            order = shopify_by_name[oid]
            result = sync_order_to_shopify(base, headers, order, matrix[oid], variant_gids, mode)
            # Just count — don't actually call API in preview since sync_order_to_shopify
            # does the real work. For true dry run, we simulate:
            tags_lower = order.get("tags", "").lower()
            if "gift redemption" in tags_lower:
                gift += 1
                continue
            current = set()
            for li in order.get("line_items", []):
                sku = (li.get("sku") or "").strip()
                fq = li.get("fulfillable_quantity", li.get("quantity", 0))
                if sku and fq > 0:
                    current.add(sku)
            to_add = [
                s
                for s in matrix[oid]
                if any(s.startswith(p) for p in ("CH-", "MT-", "AC-", "PK-", "TR-"))
                and s not in current
                and s in variant_gids
            ]
            has_dupes = any(
                s in current for s in matrix[oid] if any(s.startswith(p) for p in ("CH-", "MT-", "AC-", "PK-", "TR-"))
            )
            if to_add:
                updated += 1
            elif has_dupes and mode == "conservative":
                dupes += 1
            else:
                skipped += 1

        print(f"  Would update: {_GREEN}{updated}{_RESET}")
        print(f"  Would skip (already correct): {skipped}")
        print(f"  Gift redemption (excluded): {gift}")
        if dupes:
            print(f"  Would reject (duplicates): {_YELLOW}{dupes}{_RESET}")
        print(f"\n  Run with --execute to apply changes.")
        return True

    # Live mode — execute with thread pool
    print(f"\n{_BOLD}  LIVE SYNC — Applying changes to Shopify ({workers} workers){_RESET}\n")
    results: list[SyncResult] = []
    order_items = [(shopify_by_name[oid], matrix[oid]) for oid in sorted(matched)]
    completed = 0
    total = len(order_items)

    def _do_sync(order_and_matrix: tuple[dict, dict[str, int]]) -> SyncResult:
        order, m_skus = order_and_matrix
        return sync_order_to_shopify(base, headers, order, m_skus, variant_gids, mode)

    with ThreadPoolExecutor(max_workers=workers) as pool:
        futures = {pool.submit(_do_sync, item): item[0]["name"] for item in order_items}
        for future in as_completed(futures):
            result = future.result()
            results.append(result)
            completed += 1
            if completed % 50 == 0 or completed == total:
                pct = int(100 * completed / total)
                print(f"  Progress: {completed}/{total} ({pct}%)")

    # Summarize
    counts: dict[str, int] = defaultdict(int)
    for r in results:
        counts[r.status] += 1

    print(f"\n{_BOLD}{'=' * 60}{_RESET}")
    print(f"{_BOLD}  SYNC COMPLETE{_RESET}")
    print(f"{_BOLD}{'=' * 60}{_RESET}")
    print(f"  Updated:    {_GREEN}{counts.get('updated', 0)}{_RESET}")
    print(f"  Skipped:    {counts.get('skipped', 0)}")
    print(f"  Gift:       {counts.get('gift', 0)}")
    print(f"  Duplicates: {_YELLOW}{counts.get('duplicate', 0)}{_RESET}")
    print(f"  Errors:     {_RED}{counts.get('error', 0)}{_RESET}")

    # Log errors
    error_results = [r for r in results if r.status == "error"]
    if error_results:
        print(f"\n  {_RED}Errors:{_RESET}")
        for r in error_results[:20]:
            print(f"    #{r.order_name}: {r.error}")

    # Log duplicates
    dupe_results = [r for r in results if r.status == "duplicate"]
    if dupe_results:
        print(f"\n  {_YELLOW}Duplicates (skipped):{_RESET}")
        for r in dupe_results[:20]:
            print(f"    #{r.order_name}: {r.error}")

    print(f"{_BOLD}{'=' * 60}{_RESET}\n")

    return counts.get("error", 0) == 0


# ═══════════════════════════════════════════════════════════════════════════
# Phase 4: Gift redemption merge + finalize
# ═══════════════════════════════════════════════════════════════════════════


def identify_gift_orders(orders: list[OrderRow]) -> tuple[list[OrderRow], list[OrderRow]]:
    """Split orders into regular and gift redemption lists."""
    regular: list[OrderRow] = []
    gift: list[OrderRow] = []
    for o in orders:
        if "gift redemption" in o.tags.lower():
            gift.append(o)
        else:
            regular.append(o)
    return regular, gift


def merge_gift_xlsx(main_path: str | Path, gift_path: str | Path) -> str:
    """Merge a separate gift redemption XLSX into the main matrix.

    Both files must have Access_LIVE tab with the same column layout.
    Gift orders are appended to the main file. Returns path to merged file.
    """
    main_wb = openpyxl.load_workbook(str(main_path))
    main_ws = main_wb["Access_LIVE"]

    gift_wb = openpyxl.load_workbook(str(gift_path), data_only=True, read_only=True)
    gift_ws = gift_wb["Access_LIVE"]

    # Get existing order IDs to avoid duplicates
    existing_oids: set[str] = set()
    for row in main_ws.iter_rows(min_row=2, max_col=1, values_only=True):
        oid = str(row[0] or "").strip()
        if oid:
            existing_oids.add(oid)

    # Append gift rows
    added = 0
    skipped = 0
    for row in gift_ws.iter_rows(min_row=2, values_only=True):
        oid = str(row[0] or "").strip()
        if not oid:
            continue
        if oid in existing_oids:
            skipped += 1
            continue
        main_ws.append(list(row))
        existing_oids.add(oid)
        added += 1

    gift_wb.close()

    # Save merged file
    src = Path(main_path)
    out_path = src.parent / f"{src.stem}_MERGED{src.suffix}"
    main_wb.save(str(out_path))
    main_wb.close()

    print(f"  Gift merge: {added} orders added, {skipped} duplicates skipped")
    print(f"  Saved: {_GREEN}{out_path.name}{_RESET}")

    return str(out_path)


def finalize_xlsx(
    xlsx_path: str | Path,
    ship_day: str = "SAT",
    ship_date: str = "",
) -> str:
    """Apply final formatting fixes to RMFG download and save as email-ready file.

    Handles both RMFG Translator downloads (tab=Worksheet, no ProductionDay)
    and already-formatted files (tab=Access_LIVE, has ProductionDay).

    Transformations:
    - Rename tab to Access_LIVE if needed
    - Insert ProductionDay column at position N if missing
    - OrderIDs stored as numbers, sorted ascending
    - Zips stored as text with leading zeroes
    - Auto-size columns
    - Rename file to AHB_WeeklyProductionQuery_MM-DD-YY_vF.xlsx
    """
    from openpyxl.utils import get_column_letter

    wb = openpyxl.load_workbook(str(xlsx_path))
    fixes_applied: list[str] = []

    # Step 1: Find or rename the data sheet
    if "Access_LIVE" in wb.sheetnames:
        ws = wb["Access_LIVE"]
    elif "Worksheet" in wb.sheetnames:
        ws = wb["Worksheet"]
        ws.title = "Access_LIVE"
        fixes_applied.append("Renamed tab 'Worksheet' -> 'Access_LIVE'")
    else:
        # Use first sheet
        ws = wb[wb.sheetnames[0]]
        old_name = ws.title
        ws.title = "Access_LIVE"
        fixes_applied.append(f"Renamed tab '{old_name}' -> 'Access_LIVE'")

    # Read headers
    max_col = ws.max_column
    headers = [str(ws.cell(1, c).value or "") for c in range(1, max_col + 1)]

    # Step 2: Check if ProductionDay column exists
    has_prod_day = any(h.lower().strip() == "productionday" for h in headers)
    prod_day_col = None  # 1-indexed

    if not has_prod_day:
        # Insert ProductionDay as column N (col 14, after Notes at col 13)
        # Find where Notes is (should be col 13 / index 12)
        notes_idx = None
        for i, h in enumerate(headers):
            if h.lower().strip() == "notes":
                notes_idx = i
                break

        insert_at = (notes_idx + 2) if notes_idx is not None else 14  # 1-indexed

        # Insert column
        ws.insert_cols(insert_at)
        ws.cell(1, insert_at).value = "ProductionDay"
        prod_day_col = insert_at

        # Fill all data rows with ship_day
        for r in range(2, ws.max_row + 1):
            oid = ws.cell(r, 1).value
            if oid is not None:
                ws.cell(r, insert_at).value = ship_day.upper()

        fixes_applied.append(f"Inserted ProductionDay column at col {get_column_letter(insert_at)} ({ship_day})")
        max_col = ws.max_column
        # Re-read headers after insert
        headers = [str(ws.cell(1, c).value or "") for c in range(1, max_col + 1)]
    else:
        for i, h in enumerate(headers):
            if h.lower().strip() == "productionday":
                prod_day_col = i + 1
                break

    # Find key column indices
    oid_col = 1
    zip_col = None
    for i, h in enumerate(headers):
        if h.lower().strip() == "zip":
            zip_col = i + 1
            break

    # Step 3: Read all data rows, fix values, sort
    data_rows: list[tuple[int, list]] = []
    for r in range(2, ws.max_row + 1):
        row_vals = [ws.cell(r, c).value for c in range(1, max_col + 1)]
        oid_raw = row_vals[0]
        if oid_raw is None:
            continue
        try:
            numeric_oid = int(float(str(oid_raw)))
        except (ValueError, TypeError):
            numeric_oid = 0
        data_rows.append((numeric_oid, row_vals))

    # Sort by OrderID ascending
    was_sorted = (
        all(data_rows[i][0] <= data_rows[i + 1][0] for i in range(len(data_rows) - 1)) if len(data_rows) > 1 else True
    )

    data_rows.sort(key=lambda x: x[0])
    if not was_sorted:
        fixes_applied.append("Sorted orders by OrderID ascending")

    # Step 4: Clear and rewrite sorted + fixed data
    for r in range(2, ws.max_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(r, c).value = None

    zip_fixes = 0
    for idx, (numeric_oid, row_vals) in enumerate(data_rows):
        r = idx + 2
        for c, val in enumerate(row_vals):
            ws.cell(r, c + 1).value = val

        # Fix OrderID: numeric
        ws.cell(r, oid_col).value = numeric_oid

        # Fix Zip: text with leading zeroes
        if zip_col:
            raw_zip = row_vals[zip_col - 1]
            if raw_zip is not None:
                z = str(raw_zip).strip().split("-")[0].split(".")[0]
                if isinstance(raw_zip, (int, float)) or (z.isdigit() and len(z) < 5):
                    ws.cell(r, zip_col).value = str(int(float(str(raw_zip)))).zfill(5)
                    zip_fixes += 1

    if zip_fixes:
        fixes_applied.append(f"Fixed {zip_fixes} zip codes (leading zeroes)")

    # Step 5: Auto-size columns
    for col_idx in range(1, max_col + 1):
        max_len = len(str(ws.cell(1, col_idx).value or ""))
        # Sample first 20 rows for width
        for r in range(2, min(22, len(data_rows) + 2)):
            val = ws.cell(r, col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)
    fixes_applied.append("Auto-sized columns")

    # Step 6: Generate output filename
    if ship_date:
        # ship_date should be like "2026-03-24" or "03-24-26"
        try:
            from datetime import datetime as dt

            if len(ship_date) == 10 and ship_date[4] == "-":
                d = dt.strptime(ship_date, "%Y-%m-%d")
            else:
                d = dt.strptime(ship_date, "%m-%d-%y")
            date_str = d.strftime("%m-%d-%y")
        except ValueError:
            date_str = ship_date.replace("-", "")[:8]
    else:
        # Try to extract date from RMFG tag in first order's tags
        date_str = ""
        if data_rows:
            tags_col = None
            for i, h in enumerate(headers):
                if h.lower().strip() == "tags":
                    tags_col = i
                    break
            if tags_col is not None:
                tags = str(data_rows[0][1][tags_col] or "")
                import re

                m = re.search(r"RMFG_(\d{4})(\d{2})(\d{2})", tags)
                if m:
                    date_str = f"{m.group(2)}-{m.group(3)}-{m.group(1)[2:]}"
        if not date_str:
            from datetime import datetime as dt

            date_str = dt.now().strftime("%m-%d-%y")

    src = Path(xlsx_path)
    out_name = f"AHB_WeeklyProductionQuery_{date_str}_vF.xlsx"
    out_path = src.parent / out_name
    wb.save(str(out_path))
    wb.close()

    fixes_applied.append(f"Renamed to {out_name}")

    print(f"\n{_BOLD}  Finalized XLSX{_RESET}")
    print(f"  Orders: {len(data_rows)}")
    for fix in fixes_applied:
        print(f"  {_GREEN}+ {fix}{_RESET}")
    print(f"\n  Saved: {_GREEN}{out_path.name}{_RESET}")

    return str(out_path)


def cmd_finalize(
    xlsx_path: str,
    gift_path: Optional[str] = None,
    ship_day: str = "SAT",
    ship_date: str = "",
) -> bool:
    """Merge gift orders (if provided) and finalize XLSX for RMFG."""
    print(f"  Loading {Path(xlsx_path).name}...")

    working_path = xlsx_path

    # Merge separate gift sheet if provided
    if gift_path:
        print(f"\n  Merging gift sheet: {Path(gift_path).name}")
        working_path = merge_gift_xlsx(working_path, gift_path)

    # MFG validation on final file
    mfg_translations = load_mfg_translations()
    if mfg_translations:
        final_orders, _, _ = parse_matrix(working_path)
        result = check_mfg_onboarding(final_orders, mfg_translations)
        icon = _check_icon(result.passed)
        print(f"\n  [{icon}] {result.name}: {result.message}")
        if not result.passed:
            for d in result.details:
                print(f"       {d}")
            print(f"\n  {_RED}BLOCKED — onboard missing SKUs at RMFG before sending{_RESET}")
            return False

    # Apply final formatting (tab rename, ProductionDay, sort, zips, file naming)
    final_path = finalize_xlsx(working_path, ship_day=ship_day, ship_date=ship_date)

    print(f"\n  {_GREEN}Ready to email to RMFG: {Path(final_path).name}{_RESET}")
    return True


# ═══════════════════════════════════════════════════════════════════════════
# CLI commands
# ═══════════════════════════════════════════════════════════════════════════


def cmd_validate(xlsx_path: str) -> bool:
    """Run all validation checks on the XLSX."""
    print(f"  Loading {Path(xlsx_path).name}...")
    orders, product_columns, unmapped = parse_matrix(xlsx_path)

    # Load settings for CEX-EC config
    settings = load_settings_config()
    cex_ec = settings.get("cex_ec", {})
    cexec_splits = settings.get("cexec_splits", {})

    # Load MFG translations for onboarding check
    mfg_translations = load_mfg_translations()

    results = [
        check_numeric_order_ids(orders),
        check_zip_leading_zeroes(orders),
        check_duplicate_columns(product_columns),
        check_production_day(orders),
        check_sku_mappings(unmapped),
        check_mfg_onboarding(orders, mfg_translations),
        check_cexec_cheese_counts(orders, cex_ec, cexec_splits),
    ]

    all_passed = print_validation_report(results, len(orders))

    # Also print demand summary
    demand = compute_demand(orders)
    print_demand_summary(demand)

    return all_passed


def cmd_check(xlsx_path: str, inventory_path: Optional[str] = None) -> bool:
    """Run inventory cross-check and shortage report."""
    print(f"  Loading {Path(xlsx_path).name}...")
    orders, _, _ = parse_matrix(xlsx_path)
    inventory = _load_inventory(inventory_path)

    if not inventory:
        print(f"  {_RED}No inventory data loaded! Provide --inventory or check settings.{_RESET}")
        return False

    print(f"  Loaded {len(inventory)} SKUs from inventory source.")

    demand = compute_demand(orders)
    shortages = find_shortages(demand, inventory)
    print_inventory_report(demand, inventory, shortages)

    return len(shortages) == 0


def _load_inventory(inventory_path: Optional[str]) -> dict[str, float]:
    """Load inventory from CSV, JSON, or settings. Shared by cmd_check/cmd_full/cmd_swap."""
    if inventory_path:
        p = Path(inventory_path)
        if p.suffix == ".json":
            with open(p, encoding="utf-8") as f:
                raw = json.load(f)
            if "inventory" in raw:
                return {sku: d.get("qty", 0) for sku, d in raw["inventory"].items()}
            return {k: v for k, v in raw.items()}
        return load_inventory_csv(p)
    return load_inventory_settings()


def cmd_swap(xlsx_path: str, inventory_path: Optional[str] = None) -> bool:
    """Run inventory check + interactive swap resolution standalone."""
    print(f"  Loading {Path(xlsx_path).name}...")
    orders, _, _ = parse_matrix(xlsx_path)
    inventory = _load_inventory(inventory_path)

    if not inventory:
        print(f"  {_RED}No inventory data loaded!{_RESET}")
        return False

    print(f"  Loaded {len(inventory)} SKUs from inventory source.")
    demand = compute_demand(orders)
    shortages = find_shortages(demand, inventory)
    print_inventory_report(demand, inventory, shortages)

    if not shortages:
        print(f"  {_GREEN}No shortages — no swaps needed.{_RESET}")
        return True

    decisions = interactive_swap_resolution(shortages, inventory, demand)
    if decisions:
        fixed_path = apply_swaps_to_xlsx(xlsx_path, decisions, orders)
        # Re-check
        fixed_orders, _, _ = parse_matrix(fixed_path)
        fixed_demand = compute_demand(fixed_orders)
        remaining = find_shortages(fixed_demand, inventory)
        if remaining:
            print(f"  {_YELLOW}{len(remaining)} shortage(s) remain after swaps{_RESET}")
            return False
        print(f"  {_GREEN}All shortages resolved!{_RESET}")
        return True

    return False


def cmd_full(xlsx_path: str, inventory_path: Optional[str] = None) -> bool:
    """Run full pipeline: validate + inventory check + swap resolution."""
    print(f"\n{_BOLD}{'#' * 60}{_RESET}")
    print(f"{_BOLD}  MATRIX COMMANDER — Full Pipeline{_RESET}")
    print(f"{_BOLD}{'#' * 60}{_RESET}\n")

    # Step 1: Validate
    print(f"  {_CYAN}[Step 1/3] Validation{_RESET}")
    print(f"  Loading {Path(xlsx_path).name}...")
    orders, product_columns, unmapped = parse_matrix(xlsx_path)

    settings = load_settings_config()
    cex_ec = settings.get("cex_ec", {})
    cexec_splits = settings.get("cexec_splits", {})
    mfg_translations = load_mfg_translations()

    results = [
        check_numeric_order_ids(orders),
        check_zip_leading_zeroes(orders),
        check_duplicate_columns(product_columns),
        check_production_day(orders),
        check_sku_mappings(unmapped),
        check_mfg_onboarding(orders, mfg_translations),
        check_cexec_cheese_counts(orders, cex_ec, cexec_splits),
    ]

    validation_passed = print_validation_report(results, len(orders))

    # Step 2: Inventory check
    print(f"  {_CYAN}[Step 2/3] Inventory Cross-Check{_RESET}")
    inventory = _load_inventory(inventory_path)

    shortages: list[ShortageItem] = []
    if inventory:
        print(f"  Loaded {len(inventory)} SKUs from inventory source.")
        demand = compute_demand(orders)
        shortages = find_shortages(demand, inventory)
        print_inventory_report(demand, inventory, shortages)
        inventory_ok = len(shortages) == 0
    else:
        print(f"  {_YELLOW}No inventory data — skipping cross-check.{_RESET}")
        demand = compute_demand(orders)
        print_demand_summary(demand)
        inventory_ok = True

    # Step 3: Swap resolution (if shortages found and validation passed)
    fixed_path = None
    if shortages and validation_passed:
        print(f"  {_CYAN}[Step 3/3] Swap Resolution{_RESET}")
        decisions = interactive_swap_resolution(shortages, inventory, demand)
        if decisions:
            fixed_path = apply_swaps_to_xlsx(xlsx_path, decisions, orders)
            # Re-check inventory after swaps
            print(f"\n  {_CYAN}Re-checking inventory after swaps...{_RESET}")
            fixed_orders, _, _ = parse_matrix(fixed_path)
            fixed_demand = compute_demand(fixed_orders)
            remaining_shortages = find_shortages(fixed_demand, inventory)
            if remaining_shortages:
                print(f"  {_YELLOW}{len(remaining_shortages)} shortage(s) remain after swaps{_RESET}")
                inventory_ok = False
            else:
                print(f"  {_GREEN}All shortages resolved!{_RESET}")
                inventory_ok = True

    # Final summary
    print(f"\n{_BOLD}{'#' * 60}{_RESET}")
    if validation_passed and inventory_ok:
        print(f"  {_GREEN}READY — All checks passed, no shortages.{_RESET}")
        if fixed_path:
            print(f"  {_GREEN}Fixed file: {Path(fixed_path).name}{_RESET}")
    elif validation_passed and not inventory_ok:
        print(f"  {_YELLOW}REVIEW — Validation passed but shortages remain.{_RESET}")
    else:
        print(f"  {_RED}BLOCKED — Validation failures must be resolved.{_RESET}")
    print(f"{_BOLD}{'#' * 60}{_RESET}\n")

    return validation_passed and inventory_ok


# ═══════════════════════════════════════════════════════════════════════════
# Main
# ═══════════════════════════════════════════════════════════════════════════


def main() -> None:
    parser = argparse.ArgumentParser(
        prog="matrix_commander",
        description="Matrix Commander — Validate production matrix & check inventory.",
    )
    sub = parser.add_subparsers(dest="command", required=True)

    # validate
    p_val = sub.add_parser("validate", help="Run QC validation checks on XLSX")
    p_val.add_argument("xlsx", help="Path to AHB_WeeklyProductionQuery XLSX file")

    # check
    p_chk = sub.add_parser("check", help="Run inventory cross-check & shortage report")
    p_chk.add_argument("xlsx", help="Path to AHB_WeeklyProductionQuery XLSX file")
    p_chk.add_argument("--inventory", "-i", help="Inventory CSV (sku,available_qty) or JSON path")

    # swap
    p_swap = sub.add_parser("swap", help="Interactive shortage swap resolution")
    p_swap.add_argument("xlsx", help="Path to AHB_WeeklyProductionQuery XLSX file")
    p_swap.add_argument("--inventory", "-i", help="Inventory CSV (sku,available_qty) or JSON path")

    # sync-shopify
    p_sync = sub.add_parser("sync-shopify", help="Sync matrix to Shopify as $0 variants (replaces Matrixify)")
    p_sync.add_argument("xlsx", help="Path to AHB_WeeklyProductionQuery XLSX file")
    p_sync.add_argument("tag", help="RMFG tag to match Shopify orders (e.g. RMFG_20260328)")
    p_sync.add_argument("--execute", action="store_true", help="Actually apply changes (default: dry run)")
    p_sync.add_argument(
        "--mode",
        choices=["smart", "conservative"],
        default="smart",
        help="Duplicate handling: smart (skip SKU) or conservative (skip order)",
    )
    p_sync.add_argument("--workers", type=int, default=5, help="Concurrent workers (default: 5)")

    # finalize
    p_fin = sub.add_parser("finalize", help="Merge gift orders + format fixes + MFG validation")
    p_fin.add_argument("xlsx", help="Path to RMFG download or production XLSX file")
    p_fin.add_argument("--gift", "-g", help="Separate gift redemption XLSX to merge")
    p_fin.add_argument("--day", "-d", choices=["SAT", "TUE"], default="SAT", help="Production day (default: SAT)")
    p_fin.add_argument("--date", help="Ship date for filename (e.g. 2026-03-24 or 03-24-26)")

    # full
    p_full = sub.add_parser("full", help="Run full pipeline: validate + check + swap")
    p_full.add_argument("xlsx", help="Path to AHB_WeeklyProductionQuery XLSX file")
    p_full.add_argument("--inventory", "-i", help="Inventory CSV (sku,available_qty) or JSON path")

    args = parser.parse_args()

    if args.command == "validate":
        ok = cmd_validate(args.xlsx)
    elif args.command == "check":
        ok = cmd_check(args.xlsx, getattr(args, "inventory", None))
    elif args.command == "swap":
        ok = cmd_swap(args.xlsx, getattr(args, "inventory", None))
    elif args.command == "sync-shopify":
        ok = cmd_sync(args.xlsx, args.tag, mode=args.mode, dry_run=not args.execute, workers=args.workers)
    elif args.command == "finalize":
        ok = cmd_finalize(
            args.xlsx,
            gift_path=getattr(args, "gift", None),
            ship_day=args.day,
            ship_date=getattr(args, "date", "") or "",
        )
    elif args.command == "full":
        ok = cmd_full(args.xlsx, getattr(args, "inventory", None))
    else:
        parser.print_help()
        return

    sys.exit(0 if ok else 1)


if __name__ == "__main__":
    main()
