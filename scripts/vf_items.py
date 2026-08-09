"""vf_items — validated, ledgered ITEM-MATRIX editing on a built vF production sheet.

🔴 CONSTRAINTS SSOT: `AppyHour/MATRIX_RULES.md` rule 24. Read it before changing this file.

The item sibling of `ShipRouting/scripts/vf_tags.py` (which owns col L / routing tags). The
submitted vF is THE AUTHORITY — RMFG picks from the sheet, not from Shopify — and the item side
was still hand-edited in Excel. That is how an invented "Farmstead Smoked Cumin Gouda" header
(2026-08-04, derived from a Shopify TITLE while the authority said "Farmstead Cumin Gouda") and a
comma'd `Walnut, Honey & …` header (wk0713, 545 units at risk) got written.

Everything authoritative is IMPORTED, never re-typed:
  MFG names   -> matrix_commander.load_mfg_translations(MFG_AUTHORITATIVE_PATH)  (rules 21/23)
  name gate   -> matrix_commander.validate_mfg_names / MfgOnboardingError
  gift merge  -> matrix_commander.merge_gift_xlsx                                (rule 20)

Hard invariants (each a real burn — MATRIX_RULES rule 24a-i):
  1 every header written is the VERBATIM authoritative name; a rename/replace re-validates the
    RESULT, so it cannot smuggle in a non-authoritative name. Unknown -> refuse + print MISSING.
  2 authority missing/empty = FAIL CLOSED (rule 23's fail-open, closed here)
  3 header shape checked before membership: no comma / tab / newline / doubled or edge whitespace /
    smart punctuation. Never normalized to pass — refused.
  4 a swap NEVER puts the same item twice in one order; offending rows are named and the whole
    write is refused (multi-target: least-used FREE target; no free target -> named, never skipped)
  5 never overwrite a sent file: --write emits <stem>_r2.xlsx, _r3.xlsx, refuses collisions
  6 col-D Total is RECOMPUTED on any row whose product cells changed; Tags/ProductionDay/meta
    are byte-preserved (an item edit touches items)
  7 semantic workbook preservation is VERIFIED after write, BY HEADER NAME (so add/drop/rename
    cannot shift a neighbour) — any discrepancy deletes the output and fails the run
  8 jsonl ledger of every change -> _outputs/cache/vf_item_edits_<stem>.jsonl (revert reads it)
  9 DRY-RUN BY DEFAULT; a REFUSED op writes NOTHING — no file, no ledger line

CLI (dry-run unless --write):
  validate       <sheet>                                       audit the item matrix (read-only)
  guides         <sheet> [--fix-mismatch] [--guide-policy p.json]   rule 25: exactly ONE guide/order
  swap           <sheet> --from CH-TETI --to CH-ETX,CH-WWHO [--orders ids|file | --all]
  qty            <sheet> --sku CH-ETX --orders ids --set 2 | --delta 1
  add-column     <sheet> --sku CH-ETX
  drop-column    <sheet> --sku CH-ETX | --header "AHB (S_REG): ..."
  rename-column  <sheet> --from "<old header>" --to "<authoritative header>"
  replace-name   <sheet> --find "Walnut, Honey" --replace "Walnut Honey"
  gift           <sheet> --vfgr <vFGR.xlsx> [--gift-drop OID,OID]
  revert         <sheet> [--orders ids] [--ledger path]
"""
from __future__ import annotations

import argparse
import csv
import io
import json
import shutil
import sys
import tempfile
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

APPYHOUR = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(APPYHOUR))

import openpyxl  # noqa: E402

import matrix_commander  # noqa: E402
from matrix_commander import (  # noqa: E402
    MFG_AUTHORITATIVE_PATH,
    GiftMergeError,
    MfgOnboardingError,
    load_mfg_translations,
    merge_gift_xlsx,
    validate_mfg_names,
)

SHEET = "Access_LIVE"
ORDER_COL = "OrderID"
TOTAL_COL = "Total"
TAGS_COL = "Tags"
HDR_PREFIX = "AHB (S_REG): "
RULE = "MATRIX_RULES.md rule 24"

# Meta (non-product) columns. Product columns are identified BY HEADER SHAPE ("AHB ("), the same
# test merge_gift_xlsx uses — never by position ("col O" is today's accident).
META_HEADERS = ("OrderID", "Name", "Distribution Type", "Total", "Phone Number", "Email",
                "Address", "Address 2", "City", "State", "Zip", "Tags", "Notes", "ProductionDay")

# Pasted-not-exported punctuation (MATRIX_RULES 23a): a curly apostrophe is invisible in a csv and
# reaches a sent vF as an un-pickable header. Never normalized away — refused.
SMART_PUNCT = {"‘": "left single quote", "’": "right single quote (curly apostrophe)",
               "“": "left double quote", "”": "right double quote",
               "–": "en-dash", "—": "em-dash", " ": "non-breaking space"}


# ── tasting-guide policy (MATRIX_RULES rule 25) ──────────────────────────────
# 🔴 A TABLE, not an if/else chain. Kurt's rule changes as DATA (edit the table or pass
# --guide-policy p.json); a branch would have to be rewritten and re-tested every time.
GUIDE_RULE = "MATRIX_RULES.md rule 25"

# status: "active"  = selectable
#         "retired" = never selected; RECOGNIZED so a sheet carrying it is IDENTIFIED, never
#                     silently accepted and never silently stripped (25b)
#         "parked"  = a backup Kurt may re-enable BY CONFIG; INFO when present, not an error (25b)
GUIDE_STATUS = {
    "PK-BITESGUIDE": "active",
    "PK-TCUST": "active",
    "PK-FCUST": "retired",   # Kurt 2026-08-09 "FCUST is retired." — kept visible, never chosen
    "PK-TMDT": "parked",     # Kurt 2026-08-09 "the TMDT is a weird backup we may use again."
}

# Ordered predicates, FIRST MATCH WINS. `when_prefix`: the order carries any SKU with that prefix.
# A mixed order (tray + regular) matches rule 1 — tray wins, bites guide ONLY (Kurt 2026-08-09).
GUIDE_POLICY = (
    {"when_prefix": "TR-", "guide": "PK-BITESGUIDE", "why": "order carries a TR- tray"},
    {"when": "default", "guide": "PK-TCUST", "why": "regular (non-tray) order"},
)


class GuidePolicy:
    """The ONLY place this tool answers 'which guide does this order get'."""

    def __init__(self, rules=GUIDE_POLICY, status=GUIDE_STATUS, source="built-in table"):
        self.rules = [dict(r) for r in rules]
        self.status = dict(status)
        self.source = source
        for r in self.rules:
            g = r.get("guide")
            if g not in self.status:
                raise ItemEditRefused(f"{GUIDE_RULE}: policy rule selects {g!r}, which is not in "
                                      f"the guide status table — every selectable guide must be "
                                      f"declared (never an ad-hoc SKU).")
            if self.status[g] != "active":
                raise ItemEditRefused(
                    f"{GUIDE_RULE}: policy rule selects {g!r} but its status is "
                    f"{self.status[g]!r}. A retired/parked guide is re-enabled by setting its "
                    f"status to 'active' in the SAME policy file — never by selecting it while it "
                    f"is still marked inert (25b).")

    @classmethod
    def load(cls, path=None):
        """--guide-policy json: {"rules":[{"when_prefix":"TR-","guide":"PK-X"},…],
        "status":{"PK-TMDT":"active"}} — re-enabling a parked guide is CONFIG, not a code edit."""
        if not path:
            return cls()
        p = Path(path)
        if not p.exists():
            raise ItemEditRefused(f"--guide-policy {p} does not exist. Never fall back to the "
                                  f"built-in table silently when an override was asked for.")
        data = json.loads(p.read_text(encoding="utf-8"))
        status = {**GUIDE_STATUS, **(data.get("status") or {})}
        rules = data.get("rules") or GUIDE_POLICY
        for r in rules:
            if "guide" not in r or not ({"when", "when_prefix", "when_sku"} & set(r)):
                raise ItemEditRefused(f"--guide-policy: malformed rule {r!r} — needs a 'guide' and "
                                      f"one of when/when_prefix/when_sku.")
        return cls(rules, status, source=str(p))

    @property
    def known(self) -> set:
        return set(self.status)

    def status_of(self, sku) -> str:
        return self.status.get(sku, "unknown")

    def select(self, skus_present) -> tuple[str, str]:
        """(guide_sku, why) for an order, from the SKUs it carries. Never returns None: rule 25 is
        'every order carries exactly one guide', so a table with no default is a refusal."""
        present = {str(s) for s in skus_present if s}
        for r in self.rules:
            if r.get("when") == "default":
                return r["guide"], r.get("why", "default rule (no earlier predicate matched)")
            pre = r.get("when_prefix")
            if pre and any(s.startswith(pre) for s in present):
                return r["guide"], r.get("why", f"carries a {pre} SKU")
            sku = r.get("when_sku")
            if sku and sku in present:
                return r["guide"], r.get("why", f"carries {sku}")
        raise ItemEditRefused(f"{GUIDE_RULE}: no policy rule matched and the table has no default "
                              f"— every order must get exactly one guide. Fix {self.source}.")


def _is_reship_tags(tags) -> bool:
    """🔴 25f — reships are the high-incidence class, and they are identified by TAG through the
    CANONICAL parser (`ShipRouting/lib/qc_gate.is_reship_tag`), which knows BOTH live formats
    (`Reship - <reason>` and nested `Shipping::<reason>::<x>`). A hand-rolled substring match misses
    the nested form. Unavailable import = FAIL LOUD, never a guess."""
    fn = _load_is_reship_tag()
    parts = [t.strip() for t in str(tags or "").split(",") if t.strip()]
    return any(fn(t) for t in parts)


_IS_RESHIP_TAG = None


def _load_is_reship_tag():
    global _IS_RESHIP_TAG
    if _IS_RESHIP_TAG is None:
        sr = r"C:\Users\Work\Claude Projects\ShipRouting"      # same bridge matrix_commander uses
        if sr not in sys.path:
            sys.path.insert(0, sr)
        try:
            from lib.qc_gate import is_reship_tag
        except Exception as e:                                  # noqa: BLE001 — surfaced, not eaten
            raise ItemEditRefused(
                f"{GUIDE_RULE} 25f: canonical reship parser unavailable "
                f"(lib.qc_gate: {type(e).__name__}: {e}). REFUSING rather than hand-rolling a "
                f"substring match that misses the nested `Shipping::<reason>::` form.") from e
        _IS_RESHIP_TAG = is_reship_tag
    return _IS_RESHIP_TAG


def _force_utf8():
    """Windows cp1252 stdout mangles the header/emoji unicode. Called from main() ONLY —
    rebinding sys.stdout at import time detonates pytest's capture."""
    for name in ("stdout", "stderr"):
        stream = getattr(sys, name, None)
        try:
            stream.reconfigure(encoding="utf-8", errors="replace")
        except (AttributeError, ValueError):
            if hasattr(stream, "buffer"):
                setattr(sys, name, io.TextIOWrapper(stream.buffer, encoding="utf-8",
                                                    errors="replace", line_buffering=True))


class ItemEditRefused(RuntimeError):
    """The plan is invalid, so NOTHING is written — no file, no ledger (invariant 9)."""


def order_key(v):
    """Normalize an OrderID from any source (sheet int, '#169610', ' 169610 ') to a join key."""
    s = str(v).strip().lstrip("#").strip()
    if s.endswith(".0") and s[:-2].isdigit():
        s = s[:-2]
    return s


def is_product_header(h) -> bool:
    return str(h or "").startswith("AHB (")


def qty_of(v) -> int:
    """A product cell is blank or an int. A non-numeric value is a defect, surfaced by validate."""
    if v is None or (isinstance(v, str) and not v.strip()):
        return 0
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return -1  # sentinel: unparseable


# ── authority ────────────────────────────────────────────────────────────────
class Authority:
    """The ONLY place this tool answers 'is that a real MFG name'.

    🔴 FAIL CLOSED (invariant 2). matrix_commander and validate_vf_sheet both warn-and-skip on a
    missing/empty authority (rule 23's OPEN item) — for a submission-bound EDIT that is the
    silent-degrade class: "0 names checked" reads exactly like "all names valid".
    """

    def __init__(self, path=None, allow_missing=False):
        self.path = Path(path) if path else MFG_AUTHORITATIVE_PATH
        # 🔴 An authority governs EVERY consumer ([[feedback-reference-file-governs-everything]]):
        # an --authority override rebinds matrix_commander's module-level path too, so the reused
        # rule-21 gate below and this class can never read two different files.
        matrix_commander.MFG_AUTHORITATIVE_PATH = self.path
        self.allow_missing = allow_missing
        self.sku_to_name = load_mfg_translations(self.path)
        if not self.sku_to_name:
            msg = (f"FATAL: MFG authority {self.path} is missing or empty. Every header this tool "
                   f"writes must come from it verbatim (MATRIX_RULES 24a). Refresh it from a "
                   f"meal-type export. --allow-missing-authority downgrades this to a warning and "
                   f"is recorded in the ledger.")
            if not allow_missing:
                raise ItemEditRefused(msg)
            print("⚠️  " + msg)
        self.name_to_sku: dict[str, str] = {}
        self.dupe_names: list[str] = []
        for sku, name in self.sku_to_name.items():
            if name in self.name_to_sku:
                self.dupe_names.append(name)      # rule 23c: last-row-wins is order-dependent
            self.name_to_sku[name] = sku
        self.names = set(self.name_to_sku)

    # -- naming -----------------------------------------------------------
    def header_for_sku(self, sku: str) -> str:
        name = self.sku_to_name.get(sku.strip())
        if not name:
            raise ItemEditRefused(
                f"MISSING: {sku!r} has no name in {self.path.name}. Never derive one from a "
                f"Shopify title (the 2026-08-04 'Farmstead Smoked Cumin Gouda' class) — onboard "
                f"it at {MfgOnboardingError.URL} and re-export.")
        return name

    def sku_for_header(self, header: str) -> str | None:
        return self.name_to_sku.get(str(header or "").strip())

    # -- validation -------------------------------------------------------
    def shape_problems(self, header: str) -> list[str]:
        """Shape BEFORE membership, so the message names the defect (wk0713 comma, 545 units)."""
        h = str(header)
        probs = []
        if not h.startswith(HDR_PREFIX):
            probs.append(f"header does not start with {HDR_PREFIX!r}")
        if "," in h:
            probs.append("contains a COMMA — the wk0713 walnut-crackers class; RMFG's translator "
                         "maps the no-comma form and 545 units failed its import mapping")
        if "\n" in h or "\r" in h or "\t" in h:
            probs.append("contains a newline/tab")
        if h != h.strip():
            probs.append("has leading/trailing whitespace")
        if "  " in h:
            probs.append("contains doubled whitespace")
        for ch, label in SMART_PUNCT.items():
            if ch in h:
                probs.append(f"contains {label} (pasted, not exported — MATRIX_RULES 23a)")
        if h.strip() == HDR_PREFIX.strip():
            probs.append("empty product name")
        return probs

    def header_problems(self, header: str) -> list[str]:
        probs = self.shape_problems(header)
        if not self.names:
            return probs                                   # --allow-missing-authority
        if str(header).strip() not in self.names:
            probs.append(f"NOT in {self.path.name} — invented/mistyped header. Never hand-edit a "
                         f"header to get past this; onboard + re-export ({MfgOnboardingError.URL})")
        return probs


# ── sheet ────────────────────────────────────────────────────────────────────
class Sheet:
    """The vF workbook + a header-name -> column-index map. Never index by position."""

    def __init__(self, path, sheet_name=SHEET):
        self.path = Path(path)
        self.sheet_name = sheet_name
        # data_only=False so FORMULAS round-trip as formulas (data_only writes cached values back
        # as literals — a silent de-formula-ing of the whole workbook).
        self.wb = openpyxl.load_workbook(self.path, data_only=False)
        if sheet_name not in self.wb.sheetnames:
            raise ItemEditRefused(f"FATAL: '{sheet_name}' tab not found in {self.path.name} "
                                  f"(has {self.wb.sheetnames})")
        self.ws = self.wb[sheet_name]
        self._index()

    def _index(self):
        self.headers = [str(c.value).strip() if c.value is not None else ""
                        for c in next(self.ws.iter_rows(min_row=1, max_row=1))]
        self.hdr: dict[str, int] = {}
        self.dupe_headers: list[str] = []
        for i, h in enumerate(self.headers, start=1):
            if not h:
                continue
            if h in self.hdr:
                self.dupe_headers.append(h)
            else:
                self.hdr[h] = i
        if ORDER_COL not in self.hdr:
            raise ItemEditRefused(f"FATAL: {ORDER_COL!r} header missing from {self.path.name}. "
                                  f"Do NOT fall back to a column index.")

    @property
    def product_headers(self) -> list[str]:
        return [h for h in self.headers if is_product_header(h)]

    def data_rows(self) -> list[tuple[int, str]]:
        out = []
        for r in range(2, self.ws.max_row + 1):
            v = self.ws.cell(r, self.hdr[ORDER_COL]).value
            if v is None or str(v).strip() == "":
                continue
            out.append((r, order_key(v)))
        return out

    def get(self, row: int, header: str):
        return self.ws.cell(row, self.hdr[header]).value

    def set(self, row: int, header: str, value):
        self.ws.cell(row, self.hdr[header]).value = value

    def row_total(self, row: int) -> int:
        tot = 0
        for h in self.product_headers:
            q = qty_of(self.get(row, h))
            if q > 0:
                tot += q
        return tot


# ── snapshot + semantic preservation verify (invariant 7) ────────────────────
def snapshot(path, sheet_name=SHEET) -> dict:
    """Semantic picture of the workbook: keyed by HEADER NAME, not column position, so an
    add/drop/rename cannot hide a shifted neighbour."""
    wb = openpyxl.load_workbook(path, data_only=False)
    try:
        snap = {"sheets": list(wb.sheetnames), "other": {}, "headers": [], "rows": []}
        for name in wb.sheetnames:
            ws = wb[name]
            if name != sheet_name:
                snap["other"][name] = {(r, c): ws.cell(r, c).value
                                       for r in range(1, ws.max_row + 1)
                                       for c in range(1, ws.max_column + 1)}
                continue
            hdrs = [str(c.value).strip() if c.value is not None else ""
                    for c in next(ws.iter_rows(min_row=1, max_row=1))]
            snap["headers"] = hdrs
            oi = hdrs.index(ORDER_COL) + 1
            for r in range(2, ws.max_row + 1):
                ov = ws.cell(r, oi).value
                if ov is None or str(ov).strip() == "":
                    continue
                cells = {}
                for i, h in enumerate(hdrs, start=1):
                    if h:
                        cells[h] = ws.cell(r, i).value
                snap["rows"].append((r, order_key(ov), cells))
        return snap
    finally:
        wb.close()


def verify_preserved(before: dict, after: dict, allowed_cells: dict,
                     added=(), dropped=(), renamed=None) -> list[str]:
    """🔴 INVARIANT 7 / MATRIX_RULES 24e — audit the artifact that LEFT.

    `allowed_cells`: {(rownum, header_after): expected_value}. Everything else must be identical.
    Returns a list of problems; a non-empty list means the output gets DELETED.
    """
    renamed = dict(renamed or {})
    probs: list[str] = []
    if before["sheets"] != after["sheets"]:
        return [f"sheet names changed: {before['sheets']} -> {after['sheets']}"]
    for name, cells in before["other"].items():
        if cells != after["other"].get(name):
            probs.append(f"[{name}] non-target sheet changed")
    exp_headers = []
    for h in before["headers"]:
        if h in dropped:
            continue
        exp_headers.append(renamed.get(h, h))
    exp_headers.extend(added)
    if [h for h in after["headers"] if h] != [h for h in exp_headers if h]:
        probs.append(f"header row not as planned:\n   expected {exp_headers}\n   got      "
                     f"{after['headers']}")
        return probs
    if len(before["rows"]) != len(after["rows"]):
        probs.append(f"row count changed: {len(before['rows'])} -> {len(after['rows'])}")
        return probs
    for (rb, ob, cb), (ra, oa, ca) in zip(before["rows"], after["rows"], strict=True):
        if (rb, ob) != (ra, oa):
            probs.append(f"row identity changed at row {rb}: {ob} -> {oa}")
            continue
        for h, vb in cb.items():
            if h in dropped:
                continue
            ha = renamed.get(h, h)
            va = ca.get(ha)
            if (ra, ha) in allowed_cells:
                if va != allowed_cells[(ra, ha)]:
                    probs.append(f"row {ra} ({oa}) {ha!r}: wrote {va!r}, plan said "
                                 f"{allowed_cells[(ra, ha)]!r}")
                continue
            if va != vb:
                probs.append(f"row {ra} ({oa}) {ha!r} changed unexpectedly: {vb!r} -> {va!r}")
            if len(probs) > 25:
                return probs
        for h in added:
            va = ca.get(h)
            if (ra, h) in allowed_cells:
                continue
            if va not in (None, ""):
                probs.append(f"row {ra} ({oa}) new column {h!r} not blank: {va!r}")
    return probs


# ── ledger (invariant 8) ─────────────────────────────────────────────────────
def ledger_path(sheet_path) -> Path:
    out = APPYHOUR.parent / "_outputs" / "cache"
    return out / f"vf_item_edits_{Path(sheet_path).stem}.jsonl"


def append_ledger(path, entries):
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "a", encoding="utf-8", newline="\n") as fh:
        for e in entries:
            fh.write(json.dumps(e, ensure_ascii=False) + "\n")


def read_ledger(path) -> list[dict]:
    p = Path(path)
    if not p.exists():
        return []
    out = []
    with open(p, encoding="utf-8") as fh:
        for line in fh:
            line = line.strip()
            if line:
                out.append(json.loads(line))
    return out


def next_revision(src) -> Path:
    """🔴 INVARIANT 5. NEVER overwrite a sent/dated file ([[never-delete-prior-output-files]])."""
    src = Path(src)
    n = 2
    while True:
        cand = src.with_name(f"{src.stem}_r{n}{src.suffix}")
        if not cand.exists():
            return cand
        n += 1


# ── the plan ─────────────────────────────────────────────────────────────────
class Plan:
    """A declared set of changes. Nothing touches disk until it validates clean."""

    def __init__(self, op: str, reason: str = ""):
        self.op = op
        self.reason = reason
        self.cells: dict[tuple[int, str], tuple] = {}   # (row, header) -> (before, after)
        self.orders: dict[int, str] = {}
        self.added: list[str] = []
        self.dropped: list[str] = []
        self.renamed: dict[str, str] = {}
        self.refusals: list[str] = []
        self.dropped_cells: dict[str, dict[str, object]] = {}   # header -> {order: qty} for revert

    def set_cell(self, row, order, header, before, after):
        self.cells[(row, header)] = (before, after)
        self.orders[row] = order

    def refuse(self, msg):
        self.refusals.append(msg)

    @property
    def touched_rows(self):
        return sorted({r for r, _ in self.cells})

    def __bool__(self):
        return bool(self.cells or self.added or self.dropped or self.renamed)


def recompute_totals(sheet: Sheet, plan: Plan):
    """Rule 0 / 24f: Total = sum of the row's product cells, on every row whose items changed."""
    if TOTAL_COL not in sheet.hdr:
        return
    prod = set(sheet.product_headers) | set(plan.added)
    rows = sorted({r for (r, h) in plan.cells if h in prod})
    for r in rows:
        tot = 0
        for h in prod:
            if h in plan.dropped:
                continue
            if (r, h) in plan.cells:
                q = qty_of(plan.cells[(r, h)][1])
            elif h in sheet.hdr:
                q = qty_of(sheet.get(r, h))
            else:
                q = 0
            if q > 0:
                tot += q
        before = sheet.get(r, TOTAL_COL)
        if before != tot:
            plan.set_cell(r, plan.orders.get(r, ""), TOTAL_COL, before, tot)


# ── apply ────────────────────────────────────────────────────────────────────
def apply_plan(sheet: Sheet, plan: Plan, auth: Authority, write=False,
               echo=True) -> tuple[int, Path | None]:
    """Print the plan, refuse on ANY problem (invariant 9), else write a new revision + ledger."""
    if echo:
        _print_plan(plan)
    if plan.refusals:
        print(f"\n🔴 REFUSING {plan.op}: {len(plan.refusals)} problem(s). NOTHING written — no "
              f"file, no ledger line ({RULE}).")
        for m in plan.refusals:
            print("   • " + m)
        return 1, None
    if not plan:
        print("nothing to do.")
        return 0, None
    if not write:
        print("\nDRY RUN — nothing written. Add --write to produce a new revision.")
        return 0, None

    before_snap = snapshot(sheet.path, sheet.sheet_name)
    out_path = next_revision(sheet.path)
    if out_path.exists():                                   # belt-and-braces (invariant 5)
        raise ItemEditRefused(f"refusing to overwrite {out_path}")

    # structural ops first, then cells (cells are addressed by header name after the rename)
    _apply_structure(sheet, plan)
    for (row, header), (_b, after) in plan.cells.items():
        sheet.set(row, header, after)
    sheet.wb.save(out_path)

    allowed = {(r, h): a for (r, h), (_b, a) in plan.cells.items()}
    probs = verify_preserved(before_snap, snapshot(out_path, sheet.sheet_name), allowed,
                             added=plan.added, dropped=plan.dropped, renamed=plan.renamed)
    if probs:
        out_path.unlink(missing_ok=True)
        print("\n🔴 WORKBOOK PRESERVATION FAILED — output deleted, nothing shipped, no ledger:")
        for p in probs[:25]:
            print("   " + p)
        return 1, None

    gate = final_name_gate(out_path, sheet.sheet_name, auth)
    if gate:
        out_path.unlink(missing_ok=True)
        print(f"\n🔴 MFG NAME GATE FAILED on the written revision — output deleted, no ledger:"
              f"\n   {gate}")
        return 1, None

    ts = datetime.now(timezone.utc).isoformat()
    entries = []
    base = {"ts": ts, "sheet": sheet.path.name, "output": out_path.name, "op": plan.op,
            "reason": plan.reason, "rule": RULE}
    if auth.allow_missing and not auth.names:
        base["authority_missing"] = True                    # recorded, never silent
    for (row, header), (b, a) in sorted(plan.cells.items(), key=lambda kv: (kv[0][0], kv[0][1])):
        entries.append({**base, "kind": "cell", "row": row, "order": plan.orders.get(row, ""),
                        "header": header, "sku": auth.sku_for_header(header),
                        "before": b, "after": a})
    for h in plan.added:
        entries.append({**base, "kind": "add-column", "header": h, "sku": auth.sku_for_header(h)})
    for h in plan.dropped:
        entries.append({**base, "kind": "drop-column", "header": h,
                        "sku": auth.sku_for_header(h), "cells": plan.dropped_cells.get(h, {})})
    for old, new in plan.renamed.items():
        entries.append({**base, "kind": "rename-column", "before": old, "after": new,
                        "sku": auth.sku_for_header(new)})
    append_ledger(ledger_path(sheet.path), entries)
    print(f"\n✅ wrote {out_path}  ({len(plan.touched_rows)} row(s) changed; workbook "
          f"preservation VERIFIED by header name)")
    print(f"   ledger: {ledger_path(sheet.path)}  (+{len(entries)} entries)")
    print("   🔴 re-open the NEW revision and run `vf_items validate` on it before sending "
          "(MATRIX_RULES 24e — audit the artifact that leaves).")
    return 0, out_path


def final_name_gate(path, sheet_name, auth: Authority) -> str:
    """🔴 24a/24e — the LAST gate, run on the artifact that would leave, not on the plan.

    REUSES rule 21's `matrix_commander.validate_mfg_names` (never reimplemented): it is handed
    {sku -> header} for every product column actually present in the written revision and raises
    `MfgOnboardingError` BY TYPE for any name outside the authoritative meal-type export.
    Returns "" when clean.
    """
    wb = openpyxl.load_workbook(path, data_only=False)
    try:
        hdrs = [str(c.value).strip() if c.value is not None else ""
                for c in next(wb[sheet_name].iter_rows(min_row=1, max_row=1))]
    finally:
        wb.close()
    translations = {}
    for h in hdrs:
        if h and is_product_header(h):
            translations[auth.sku_for_header(h) or f"<unmapped:{h}>"] = h
    try:
        validate_mfg_names(translations)
    except MfgOnboardingError as e:
        return str(e)
    return ""


def _apply_structure(sheet: Sheet, plan: Plan):
    for old, new in plan.renamed.items():
        sheet.ws.cell(1, sheet.hdr[old]).value = new
    for h in plan.dropped:
        sheet.ws.delete_cols(sheet.hdr[h])
        sheet._index()
    for h in plan.added:
        sheet.ws.cell(1, sheet.ws.max_column + 1).value = h
    if plan.renamed or plan.dropped or plan.added:
        sheet._index()


def _print_plan(plan: Plan):
    print(f"\n{plan.op}: {len(plan.cells)} cell change(s) on {len(plan.touched_rows)} row(s)"
          f"{'; +' + ','.join(plan.added) if plan.added else ''}"
          f"{'; -' + ','.join(plan.dropped) if plan.dropped else ''}"
          f"{'; rename ' + str(plan.renamed) if plan.renamed else ''}")
    if not plan.cells:
        return
    print("-" * 110)
    print(f"{'ORDER':<10} {'ROW':<6} {'COLUMN':<58} BEFORE -> AFTER")
    print("-" * 110)
    for shown, (row, header) in enumerate(sorted(plan.cells, key=lambda k: (k[0], k[1])), start=1):
        b, a = plan.cells[(row, header)]
        print(f"{plan.orders.get(row, ''):<10} {row:<6} {header[:56]:<58} {b!r} -> {a!r}")
        if shown >= 60:
            print(f"   … {len(plan.cells) - shown} more")
            break
    print("-" * 110)


# ── selection ────────────────────────────────────────────────────────────────
def load_orders(spec) -> list[str]:
    """--orders takes a comma list OR a path to a .csv/.txt of order numbers."""
    p = Path(spec)
    if p.exists():
        if p.suffix.lower() == ".csv":
            with open(p, encoding="utf-8-sig", newline="") as fh:
                rows = [r for r in csv.reader(fh) if r and any(c.strip() for c in r)]
            if not rows:
                return []
            head = [c.strip().lower() for c in rows[0]]
            idx = next((i for i, h in enumerate(head)
                        if h in ("orderid", "order", "order_number", "order number", "name")), None)
            body = rows[1:] if idx is not None else rows
            i = idx if idx is not None else 0
            return [order_key(r[i]) for r in body if len(r) > i and str(r[i]).strip()]
        return [order_key(x) for x in p.read_text(encoding="utf-8").split() if x.strip()]
    return [order_key(x) for x in str(spec).split(",") if x.strip()]


def select_rows(sheet: Sheet, args) -> list[tuple[int, str]]:
    """🔴 24h: selection is never INFERRED. Curation is not on the sheet (the box SKU never lands
    there), so there is no --curation — pass an order list."""
    rows = sheet.data_rows()
    if getattr(args, "orders", None):
        want = set(load_orders(args.orders))
        picked = [(r, o) for r, o in rows if o in want]
        missing = want - {o for _r, o in picked}
        if missing:
            raise ItemEditRefused(f"order(s) not on the sheet: {sorted(missing)} — never guess a "
                                  f"row; fix the list or the cohort.")
        return picked
    if getattr(args, "state", None):
        st = args.state.strip().upper()
        if "State" not in sheet.hdr:
            raise ItemEditRefused("no State column on this sheet")
        return [(r, o) for r, o in rows
                if str(sheet.get(r, "State") or "").strip().upper() == st]
    if getattr(args, "all", False):
        return rows
    raise ItemEditRefused("no selection: pass --orders <ids|file>, --state XX, or --all.")


# ── tasting-guide audit (rule 25) ────────────────────────────────────────────
class GuideAudit:
    """Read-only guide coverage of a built sheet. Every bucket is split reship vs regular (25f)."""

    def __init__(self, sheet: Sheet, auth: Authority, policy: GuidePolicy):
        self.sheet, self.auth, self.policy = sheet, auth, policy
        self.guide_cols: dict[str, str] = {}      # sku -> header, for guides ON the sheet
        for h in sheet.product_headers:
            sku = auth.sku_for_header(h)
            if sku in policy.known:
                self.guide_cols[sku] = h
        self.rows: list[dict] = []
        self.missing_cols: list[str] = []
        for r, o in sheet.data_rows():
            skus, carried = set(), []
            for h in sheet.product_headers:
                if qty_of(sheet.get(r, h)) > 0:
                    s = auth.sku_for_header(h) or h
                    skus.add(s)
                    if s in policy.known:
                        carried.append((s, qty_of(sheet.get(r, h))))
            want, why = policy.select(skus)
            tags = sheet.get(r, TAGS_COL) if TAGS_COL in sheet.hdr else ""
            self.rows.append({"row": r, "order": o, "want": want, "why": why,
                              "carried": carried, "reship": _is_reship_tags(tags)})
        for sku in {x["want"] for x in self.rows}:
            if sku not in self.guide_cols:
                self.missing_cols.append(sku)

    # -- buckets ----------------------------------------------------------
    def _pick(self, fn):
        return [x for x in self.rows if fn(x)]

    @property
    def none_(self):
        return self._pick(lambda x: not x["carried"])

    @property
    def exactly_one(self):
        return self._pick(lambda x: len(x["carried"]) == 1)

    @property
    def multi(self):
        return self._pick(lambda x: len(x["carried"]) > 1)

    @property
    def mismatched(self):
        """One guide, but not the one the table says — EXCLUDING retired/parked, which get their
        own buckets (25i: a retired guide is a policy statement, not a plain mismatch)."""
        return self._pick(lambda x: len(x["carried"]) == 1
                          and x["carried"][0][0] != x["want"]
                          and self.policy.status_of(x["carried"][0][0]) == "active")

    @property
    def retired_present(self):
        return self._pick(lambda x: any(self.policy.status_of(s) == "retired" for s, _q in
                                        x["carried"]))

    @property
    def parked_present(self):
        return self._pick(lambda x: any(self.policy.status_of(s) == "parked" for s, _q in
                                        x["carried"]))

    @property
    def bad_qty(self):
        return self._pick(lambda x: any(q != 1 for _s, q in x["carried"]))

    @property
    def tray_bearing(self):
        return self._pick(lambda x: x["want"] == "PK-BITESGUIDE")

    def report(self) -> int:
        pol = self.policy
        print(f"\n🎫 tasting-guide coverage ({GUIDE_RULE}) — policy: {pol.source}")
        for i, r in enumerate(pol.rules, start=1):
            cond = ("otherwise" if r.get("when") == "default"
                    else f"carries {r.get('when_prefix') or r.get('when_sku')}")
            print(f"   {i}. {cond:<22} -> {r['guide']}")
        inert = [f"{s} [{st}]" for s, st in pol.status.items() if st != "active"]
        if inert:
            print(f"   inert (never selected, still recognized): {', '.join(inert)}")
        print(f"   guide columns on the sheet: "
              f"{sorted(self.guide_cols) or 'NONE'}")

        def split(items):
            rs = sum(1 for x in items if x["reship"])
            return f"{len(items):>5}   (reship {rs} / regular {len(items) - rs})"

        n_rs = sum(1 for x in self.rows if x["reship"])
        print(f"\n   orders: {len(self.rows)}   (reship {n_rs} / regular {len(self.rows) - n_rs})")
        print(f"   tray-bearing (-> PK-BITESGUIDE):  {split(self.tray_bearing)}")
        for icon, label, items in (
            ("🔴", "NO guide", self.none_),
            ("✅", "exactly one guide", self.exactly_one),
            ("🔴", "MORE THAN ONE guide (never auto-fixed, 25h)", self.multi),
            ("🔴", "MISMATCHED vs the table", self.mismatched),
            ("🔴", "guide qty != 1 (25h)", self.bad_qty),
            ("🔴", "RETIRED guide present (25b — identify, never strip)", self.retired_present),
            ("🔵", "PARKED guide present (INFO — used deliberately)", self.parked_present),
        ):
            mark = "✅" if not items else icon
            print(f"   {mark} {label + ':':<48}{split(items)}")
            for x in items[:8]:
                got = ",".join(f"{s}x{q}" for s, q in x["carried"]) or "-"
                print(f"        • order {x['order']}{' [RESHIP]' if x['reship'] else ''}: "
                      f"want {x['want']} ({x['why']}), has {got}")
            if len(items) > 8:
                print(f"        … {len(items) - 8} more")
        if self.missing_cols:
            print(f"   ⚠️  guide column(s) the table needs but the sheet lacks: "
                  f"{sorted(self.missing_cols)} — `guides --write` adds them from the authority")
        fails = (len(self.none_) + len(self.multi) + len(self.mismatched) + len(self.bad_qty)
                 + len(self.retired_present))
        return fails


# ── commands ─────────────────────────────────────────────────────────────────
def cmd_guides(args) -> int:
    """🔴 rule 25 — bring every order to EXACTLY ONE correct tasting guide.

    Adds the guide where MISSING. `--fix-mismatch` also corrects an active-but-wrong guide.
    NEVER touches: duplicates (25h), qty!=1 (25h), retired/parked guides (25i) — each is named and
    left for a human. Every header written comes from the MFG authority (25d).
    """
    auth = Authority(args.authority, args.allow_missing_authority)
    policy = GuidePolicy.load(args.guide_policy)
    sheet = Sheet(args.sheet, args.sheet_name)
    audit = GuideAudit(sheet, auth, policy)
    audit.report()
    plan = Plan("guides", args.reason)

    # 🔴 25d: a needed guide column is created ONLY from the authority; absent -> MISSING, refuse.
    for sku in sorted(audit.missing_cols):
        try:
            header = auth.header_for_sku(sku)
        except ItemEditRefused as e:
            plan.refuse(str(e))
            continue
        probs = auth.header_problems(header)
        if probs:
            plan.refuse(f"{sku}: header {header!r} " + "; ".join(probs))
            continue
        plan.added.append(header)
        audit.guide_cols[sku] = header
    if plan.refusals:
        return apply_plan(sheet, plan, auth, write=False)[0]

    skipped = []
    for x in audit.rows:
        carried = x["carried"]
        if len(carried) > 1:
            skipped.append(f"order {x['order']}: {len(carried)} guides "
                           f"({','.join(s for s, _q in carried)}) — 25h, resolve by hand")
            continue
        if carried and carried[0][1] != 1:
            skipped.append(f"order {x['order']}: {carried[0][0]} qty={carried[0][1]} — 25h")
            continue
        if carried and policy.status_of(carried[0][0]) in ("retired", "parked"):
            skipped.append(f"order {x['order']}: carries {carried[0][0]} "
                           f"[{policy.status_of(carried[0][0])}] — 25i, never auto-replaced")
            continue
        target = audit.guide_cols[x["want"]]
        if not carried:
            plan.set_cell(x["row"], x["order"], target, sheet.get(x["row"], target)
                          if target in sheet.hdr else None, 1)
        elif carried[0][0] != x["want"]:
            if not args.fix_mismatch:
                skipped.append(f"order {x['order']}: has {carried[0][0]}, table says {x['want']} "
                               f"— pass --fix-mismatch to correct it")
                continue
            old = audit.guide_cols[carried[0][0]]
            plan.set_cell(x["row"], x["order"], old, sheet.get(x["row"], old), None)
            plan.set_cell(x["row"], x["order"], target, sheet.get(x["row"], target)
                          if target in sheet.hdr else None, 1)
    if skipped:
        print(f"\n⚠️  {len(skipped)} order(s) LEFT ALONE (named, never silently swallowed):")
        for m in skipped[:20]:
            print("   • " + m)
        if len(skipped) > 20:
            print(f"   … {len(skipped) - 20} more")
    recompute_totals(sheet, plan)
    return apply_plan(sheet, plan, auth, write=args.write)[0]


def cmd_validate(args) -> int:
    auth = Authority(args.authority, args.allow_missing_authority)
    sheet = Sheet(args.sheet, args.sheet_name)
    rows = sheet.data_rows()
    prod = sheet.product_headers
    print(f"\nvf_items validate — {sheet.path.name} [{sheet.sheet_name}]")
    print(f"  rows: {len(rows)}   columns: {len(sheet.headers)}   product columns: {len(prod)}")
    print(f"  authority: {auth.path.name} ({len(auth.names)} names)")

    bad_headers, unknown_meta = [], []
    for h in sheet.headers:
        if not h:
            continue
        if is_product_header(h):
            p = auth.header_problems(h)
            if p:
                bad_headers.append((h, p))
        elif h not in META_HEADERS:
            unknown_meta.append(h)

    dupe_cols = sorted(set(sheet.dupe_headers))
    sku_cols: dict[str, list[str]] = {}
    for h in prod:
        sku = auth.sku_for_header(h)
        if sku:
            sku_cols.setdefault(sku, []).append(h)
    multi_sku = {s: hs for s, hs in sku_cols.items() if len(hs) > 1}

    dupe_rows, total_drift, bad_cells = [], [], []
    for r, o in rows:
        seen: dict[str, list[str]] = {}
        for h in prod:
            q = qty_of(sheet.get(r, h))
            if q < 0:
                bad_cells.append((o, h, sheet.get(r, h)))
            elif q > 0:
                sku = auth.sku_for_header(h) or h
                seen.setdefault(sku, []).append(h)
        dup = {s: hs for s, hs in seen.items() if len(hs) > 1}
        if dup:
            dupe_rows.append((o, dup))
        if TOTAL_COL in sheet.hdr:
            want = sheet.row_total(r)
            got = qty_of(sheet.get(r, TOTAL_COL))
            if got != want:
                total_drift.append((o, got, want))

    def _section(title, items, fmt, limit=15):
        icon = "✅" if not items else "🔴"
        print(f"\n{icon} {title}: {len(items)}")
        for x in items[:limit]:
            print("   • " + fmt(x))
        if len(items) > limit:
            print(f"   … {len(items) - limit} more")

    _section("product headers NOT valid against the authority (24a/24b)", bad_headers,
             lambda x: f"{x[0]!r}: " + "; ".join(x[1]))
    _section("unrecognised non-product columns", unknown_meta, lambda x: repr(x))
    _section("duplicate header columns", dupe_cols, lambda x: repr(x))
    _section("one SKU spread over several columns", sorted(multi_sku.items()),
             lambda x: f"{x[0]}: {x[1]}")
    _section("rows carrying the SAME item twice (24c)", dupe_rows,
             lambda x: f"order {x[0]}: " + "; ".join(f"{s} in {hs}" for s, hs in x[1].items()))
    _section("non-numeric product cells", bad_cells, lambda x: f"order {x[0]} {x[1]!r} = {x[2]!r}")
    _section("Total != sum of product cells (rule 0)", total_drift,
             lambda x: f"order {x[0]}: sheet={x[1]} computed={x[2]}")

    guide_fails = GuideAudit(sheet, auth, GuidePolicy.load(args.guide_policy)).report()

    fails = (len(bad_headers) + len(dupe_cols) + len(multi_sku) + len(dupe_rows)
             + len(bad_cells) + len(total_drift) + guide_fails)
    print(f"\n{'✅ CLEAN' if not fails else '🔴 ' + str(fails) + ' DEFECT GROUP(S)'} — read-only, "
          f"nothing written.")
    return 0 if not fails else 1


def cmd_swap(args) -> int:
    """🔴 24c: a swap NEVER puts the same item twice in one order."""
    auth = Authority(args.authority, args.allow_missing_authority)
    sheet = Sheet(args.sheet, args.sheet_name)
    plan = Plan(f"swap {args.from_sku} -> {args.to}", args.reason)

    src_header = auth.header_for_sku(args.from_sku)
    targets = []
    for sku in [s.strip() for s in args.to.split(",") if s.strip()]:
        try:
            h = auth.header_for_sku(sku)                     # refuses a SKU absent from authority
        except ItemEditRefused as e:
            plan.refuse(str(e))
            continue
        probs = auth.header_problems(h)
        if probs:
            plan.refuse(f"target {sku}: header {h!r} " + "; ".join(probs))
            continue
        targets.append((sku, h))
    if src_header not in sheet.hdr:
        plan.refuse(f"source column {src_header!r} ({args.from_sku}) is not on the sheet")
    for sku, h in targets:
        if h not in sheet.hdr:
            plan.refuse(f"target column {h!r} ({sku}) is not on the sheet — add it first with "
                        f"`add-column --sku {sku}` (never hand-type a header)")
    if plan.refusals:
        return apply_plan(sheet, plan, auth, write=False)[0]

    rows = select_rows(sheet, args)
    used = Counter()
    blocked = []
    for r, o in rows:
        q = qty_of(sheet.get(r, src_header))
        if q <= 0:
            continue
        free = [(sku, h) for sku, h in targets if qty_of(sheet.get(r, h)) <= 0]
        if not free:
            blocked.append((o, [s for s, _h in targets]))
            continue
        sku, h = min(free, key=lambda t: (used[t[0]], [s for s, _ in targets].index(t[0])))
        used[sku] += 1
        plan.set_cell(r, o, src_header, sheet.get(r, src_header), None)
        plan.set_cell(r, o, h, sheet.get(r, h), q)
    for o, tg in blocked:
        plan.refuse(f"order {o}: DUPLICATE — already carries every target {tg}. A swap must never "
                    f"put the same item twice in one order (MATRIX_RULES 24c); resolve by hand or "
                    f"widen --to.")
    recompute_totals(sheet, plan)
    if used:
        print("  target distribution: " + ", ".join(f"{s}={n}" for s, n in used.items()))
    return apply_plan(sheet, plan, auth, write=args.write)[0]


def cmd_qty(args) -> int:
    auth = Authority(args.authority, args.allow_missing_authority)
    sheet = Sheet(args.sheet, args.sheet_name)
    header = auth.header_for_sku(args.sku)
    plan = Plan(f"qty {args.sku}", args.reason)
    if header not in sheet.hdr:
        plan.refuse(f"column {header!r} ({args.sku}) is not on the sheet")
        return apply_plan(sheet, plan, auth, write=False)[0]
    for r, o in select_rows(sheet, args):
        cur = qty_of(sheet.get(r, header))
        new = args.set if args.set is not None else cur + args.delta
        if new < 0:
            plan.refuse(f"order {o}: qty would go negative ({cur} -> {new})")
            continue
        val = None if new == 0 else new
        if val != sheet.get(r, header):
            plan.set_cell(r, o, header, sheet.get(r, header), val)
    recompute_totals(sheet, plan)
    return apply_plan(sheet, plan, auth, write=args.write)[0]


def cmd_add_column(args) -> int:
    auth = Authority(args.authority, args.allow_missing_authority)
    sheet = Sheet(args.sheet, args.sheet_name)
    plan = Plan(f"add-column {args.sku}", args.reason)
    header = auth.header_for_sku(args.sku)                   # authority or refuse
    probs = auth.header_problems(header)
    if probs:
        plan.refuse(f"{header!r}: " + "; ".join(probs))
    if header in sheet.hdr:
        plan.refuse(f"column {header!r} already exists — never create a second column for one SKU")
    if not plan.refusals:
        plan.added.append(header)
    return apply_plan(sheet, plan, auth, write=args.write)[0]


def cmd_drop_column(args) -> int:
    auth = Authority(args.authority, args.allow_missing_authority)
    sheet = Sheet(args.sheet, args.sheet_name)
    header = args.header or auth.header_for_sku(args.sku)
    plan = Plan(f"drop-column {header}", args.reason)
    if header not in sheet.hdr:
        plan.refuse(f"column {header!r} is not on the sheet")
        return apply_plan(sheet, plan, auth, write=False)[0]
    carried = {o: qty_of(sheet.get(r, header)) for r, o in sheet.data_rows()
               if qty_of(sheet.get(r, header)) > 0}
    if carried and not args.force:
        plan.refuse(f"{header!r} still carries {sum(carried.values())} unit(s) on "
                    f"{len(carried)} order(s) ({sorted(carried)[:8]}…). Dropping it deletes real "
                    f"demand — swap it out first, or --force --reason \"<why>\".")
    else:
        plan.dropped.append(header)
        plan.dropped_cells[header] = carried              # exact revert
        # rule 0: rows losing units get a recomputed Total (the dropped column's own cells vanish)
        if carried and TOTAL_COL in sheet.hdr:
            for r, o in sheet.data_rows():
                if o in carried:
                    tot = sheet.row_total(r) - carried[o]
                    if sheet.get(r, TOTAL_COL) != tot:
                        plan.set_cell(r, o, TOTAL_COL, sheet.get(r, TOTAL_COL), tot)
    return apply_plan(sheet, plan, auth, write=args.write)[0]


def cmd_rename_column(args) -> int:
    auth = Authority(args.authority, args.allow_missing_authority)
    sheet = Sheet(args.sheet, args.sheet_name)
    plan = Plan(f"rename-column {args.from_header} -> {args.to}", args.reason)
    if args.from_header not in sheet.hdr:
        plan.refuse(f"column {args.from_header!r} is not on the sheet")
    probs = auth.header_problems(args.to)                    # 🔴 the RESULT is validated
    if probs:
        plan.refuse(f"target {args.to!r}: " + "; ".join(probs))
    if args.to in sheet.hdr:
        plan.refuse(f"{args.to!r} already exists — a rename must not create a duplicate column")
    if not plan.refusals:
        plan.renamed[args.from_header] = args.to
    return apply_plan(sheet, plan, auth, write=args.write)[0]


def cmd_replace_name(args) -> int:
    """Formatting fix across headers. 🔴 every RESULT is re-validated — a find-and-replace cannot
    smuggle in a non-authoritative name (24a)."""
    auth = Authority(args.authority, args.allow_missing_authority)
    sheet = Sheet(args.sheet, args.sheet_name)
    plan = Plan(f"replace-name {args.find!r} -> {args.replace!r}", args.reason)
    hits = [h for h in sheet.product_headers if args.find in h]
    if not hits:
        print(f"no product header contains {args.find!r} — nothing to do.")
        return 0
    seen = set(sheet.headers)
    for h in hits:
        new = h.replace(args.find, args.replace)
        probs = auth.header_problems(new)
        if probs:
            plan.refuse(f"{h!r} -> {new!r}: " + "; ".join(probs))
            continue
        if new != h and new in seen:
            plan.refuse(f"{h!r} -> {new!r}: collides with an existing column")
            continue
        if new != h:
            plan.renamed[h] = new
            seen.add(new)
    return apply_plan(sheet, plan, auth, write=args.write)[0]


def cmd_gift(args) -> int:
    """Rule 20 via the ONE chokepoint (`merge_gift_xlsx`) — never a second gift-merge path.
    The source is copied first so the merge can never write beside a sent file."""
    auth = Authority(args.authority, args.allow_missing_authority)
    src = Path(args.sheet)
    drops = {order_key(x) for x in (args.gift_drop or "").split(",") if x.strip()}
    tmpdir = Path(tempfile.mkdtemp(prefix="vf_items_gift_"))
    try:
        work = tmpdir / src.name
        shutil.copy2(src, work)
        try:
            merged = Path(merge_gift_xlsx(work, args.vfgr, drop_oids=drops or None))
        except GiftMergeError as e:
            print(f"\n🔴 REFUSING gift merge — nothing written ({RULE}, rule 20c):\n   {e}")
            return 1
        before, after = snapshot(src), snapshot(merged)
        # 🔴 24a: the merge UNIONS gift-only PRODUCT columns in (rule 20d) — every resulting header
        # is authority-checked here, so a vFGR cannot introduce an invented name the generate-time
        # gate never saw.
        bad = {h: auth.header_problems(h) for h in after["headers"]
               if h and is_product_header(h) and auth.header_problems(h)}
        if bad:
            print(f"\n🔴 REFUSING gift merge — nothing written: {len(bad)} merged header(s) fail "
                  f"the MFG authority ({RULE} 24a):")
            for h, probs in list(bad.items())[:15]:
                print(f"   • {h!r}: " + "; ".join(probs))
            return 1
        changed = _diff_report(before, after)
        if not args.write:
            print("\nDRY RUN — nothing written. Add --write to produce a new revision.")
            return 0
        out = next_revision(src)
        shutil.copy2(merged, out)
        append_ledger(ledger_path(src), [{
            "ts": datetime.now(timezone.utc).isoformat(), "sheet": src.name, "output": out.name,
            "op": "gift", "kind": "gift-replace", "vfgr": str(args.vfgr),
            "gift_drop": sorted(drops), "rows_changed": changed, "reason": args.reason,
            "rule": f"{RULE} / MATRIX_RULES rule 20"}])
        print(f"\n✅ wrote {out}  ({changed} gift row(s) REPLACED from the vFGR)")
        print(f"   ledger: {ledger_path(src)}")
        print("   🔴 run `vf_items validate` on the new revision before sending.")
        return 0
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)


def _diff_report(before, after) -> int:
    bmap = {o: c for _r, o, c in before["rows"]}
    n = 0
    new_cols = [h for h in after["headers"] if h and h not in before["headers"]]
    if new_cols:
        print(f"  gift-only product column(s) unioned in: {new_cols}")
    for _r, o, c in after["rows"]:
        b = bmap.get(o)
        if b is None:
            print(f"  🔴 row {o} appeared — merge_gift_xlsx never appends unknown orders")
            n += 1
            continue
        diffs = {h: (b.get(h), v) for h, v in c.items() if b.get(h) != v}
        if diffs:
            n += 1
            if n <= 12:
                print(f"  order {o}: " + ", ".join(f"{h.replace(HDR_PREFIX, '')}: "
                                                   f"{bv!r}->{av!r}" for h, (bv, av) in
                                                   list(diffs.items())[:6]))
    print(f"  {n} row(s) would change")
    return n


def cmd_revert(args) -> int:
    auth = Authority(args.authority, args.allow_missing_authority)
    sheet = Sheet(args.sheet, args.sheet_name)
    lp = Path(args.ledger) if args.ledger else ledger_path(args.sheet)
    entries = read_ledger(lp)
    if not entries:
        raise ItemEditRefused(f"no ledger at {lp} — nothing to revert from. The ledger IS the "
                              f"authority for a prior value; never guess one.")
    want = set(load_orders(args.orders)) if args.orders else None
    plan = Plan("revert", args.reason or f"revert from {lp.name}")
    # last-write-wins per (row, header): walk newest→oldest, keep the OLDEST 'before' we see
    prior: dict[tuple[int, str], object] = {}
    for e in entries:
        if e.get("kind") != "cell":
            plan.refuse(f"ledger contains a structural op ({e.get('kind')} {e.get('header')!r}) — "
                        f"undo it with the inverse command, not revert.")
            continue
        if want is not None and str(e.get("order")) not in want:
            continue
        prior[(e["row"], e["header"])] = e["before"]
    for (row, header), before in prior.items():
        if header not in sheet.hdr:
            plan.refuse(f"column {header!r} no longer on the sheet")
            continue
        cur = sheet.get(row, header)
        if cur != before:
            plan.set_cell(row, order_key(sheet.get(row, ORDER_COL)), header, cur, before)
    return apply_plan(sheet, plan, auth, write=args.write)[0]


# ── CLI ──────────────────────────────────────────────────────────────────────
def build_parser():
    ap = argparse.ArgumentParser(prog="vf_items", description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    sub = ap.add_subparsers(dest="cmd", required=True)

    def common(p, write=True, select=False):
        p.add_argument("sheet", help="path to the built vF .xlsx")
        p.add_argument("--sheet-name", default=SHEET)
        p.add_argument("--authority", default=None, help="override mfg_names_authoritative.csv")
        p.add_argument("--allow-missing-authority", action="store_true",
                       help="downgrade a missing/empty authority to a warning (LOGGED)")
        p.add_argument("--guide-policy", default=None,
                       help="json overriding the rule-25 guide table (re-enable a PARKED guide by "
                            "CONFIG, never a code edit)")
        if write:
            p.add_argument("--write", action="store_true", help="emit <stem>_rN.xlsx (else dry-run)")
            p.add_argument("--reason", default="", help="why (recorded in the ledger)")
        if select:
            p.add_argument("--orders", help="comma list or path to .csv/.txt of order numbers")
            p.add_argument("--state", help="two-letter state")
            p.add_argument("--all", action="store_true", help="every row on the sheet")
        return p

    common(sub.add_parser("validate", help="audit the item matrix (read-only)"), write=False)

    p = common(sub.add_parser("swap", help="replace one SKU with another across a selection"),
               select=True)
    p.add_argument("--from", dest="from_sku", required=True)
    p.add_argument("--to", required=True, help="one SKU, or a comma list of alternatives")

    p = common(sub.add_parser("qty", help="set/adjust a SKU's quantity across a selection"),
               select=True)
    p.add_argument("--sku", required=True)
    g = p.add_mutually_exclusive_group(required=True)
    g.add_argument("--set", type=int)
    g.add_argument("--delta", type=int)

    p = common(sub.add_parser("add-column", help="add an item column (header from the authority)"))
    p.add_argument("--sku", required=True)

    p = common(sub.add_parser("drop-column", help="remove an item column"))
    p.add_argument("--sku")
    p.add_argument("--header")
    p.add_argument("--force", action="store_true", help="drop even if it carries units")

    p = common(sub.add_parser("rename-column", help="rename a header to an AUTHORITATIVE name"))
    p.add_argument("--from", dest="from_header", required=True)
    p.add_argument("--to", required=True)

    p = common(sub.add_parser("replace-name", help="find/replace across headers, results validated"))
    p.add_argument("--find", required=True)
    p.add_argument("--replace", required=True)

    p = common(sub.add_parser("gift", help="replace gift rows from the weekly vFGR (rule 20)"))
    p.add_argument("--vfgr", required=True)
    p.add_argument("--gift-drop", default="")

    p = common(sub.add_parser("guides", help="every order gets EXACTLY ONE tasting guide (rule 25)"))
    p.add_argument("--fix-mismatch", action="store_true",
                   help="also correct an active-but-wrong guide (retired/parked never auto-fixed)")

    p = common(sub.add_parser("revert", help="restore prior cell values from the ledger"),
               select=True)
    p.add_argument("--ledger")
    return ap


HANDLERS = {"validate": cmd_validate, "guides": cmd_guides, "swap": cmd_swap, "qty": cmd_qty,
            "add-column": cmd_add_column, "drop-column": cmd_drop_column,
            "rename-column": cmd_rename_column, "replace-name": cmd_replace_name,
            "gift": cmd_gift, "revert": cmd_revert}


def main(argv=None) -> int:
    _force_utf8()
    args = build_parser().parse_args(argv)
    for attr, default in (("write", False), ("reason", ""), ("force", False),
                          ("authority", None), ("allow_missing_authority", False),
                          ("guide_policy", None), ("fix_mismatch", False)):
        if not hasattr(args, attr):
            setattr(args, attr, default)
    try:
        return HANDLERS[args.cmd](args)
    except (ItemEditRefused, MfgOnboardingError) as e:
        print(f"\n🔴 REFUSED — nothing written ({RULE}):\n   {e}")
        return 1


if __name__ == "__main__":
    sys.exit(main())
