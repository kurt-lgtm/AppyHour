"""Export 11 CSV-overpacked Small Box orders to xlsx for review."""
from __future__ import annotations
import csv, json, sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

sys.path.insert(0, r"C:\Users\Work\Claude Projects\AppyHour")
from box_simulation import build_lookup, simulate, resolve_distvol, PHYSICAL_PREFIXES  # noqa: E402

CSV_PATH = Path(r"C:\Users\Work\Downloads\AHB_BoxAssignment_4-27-26.xlsx - SAT.csv")
CACHE = Path(r"C:\Users\Work\box_sim_cache__SHIP_2026-04-27.json")
OUT = Path(r"C:\Users\Work\Claude Projects\_outputs\reports\AHB_Overpacked_SmallBox_2026-04-27.xlsx")

def csv_box_kind(box: str) -> str:
    b = box.lower()
    if "small" in b: return "Small Box"
    if "medium" in b: return "Medium Box"
    if "large" in b or "tray" in b: return "Large Box"
    return box

def main():
    csv_rows = {}
    with open(CSV_PATH, newline="", encoding="utf-8-sig") as f:
        for r in csv.DictReader(f):
            oid = (r.get("OrderID") or "").strip()
            if oid:
                csv_rows[oid] = {
                    "name": r.get("Name", "").strip(),
                    "csv_items": int(r.get("ItemCnt") or 0),
                    "csv_vol": float(r.get("TotalVolume") or 0),
                    "csv_box": (r.get("Boxes") or "").strip(),
                }
    orders = json.loads(CACHE.read_text(encoding="utf-8"))
    lookup = build_lookup()
    sim = simulate(orders, lookup)
    by_name = {r["Order"].lstrip("#"): r for r in sim}
    raw_by_name = {o["name"].lstrip("#"): o for o in orders}

    rows = []
    for oid, c in csv_rows.items():
        s = by_name.get(oid)
        if not s: continue
        if csv_box_kind(c["csv_box"]) != s["Box Assignment"]: pass
        else: continue
        # only Small->Medium overpacked
        if s["Box Assignment"] not in ("Medium Box", "Large Box") or "Small" not in c["csv_box"]:
            continue
        raw = raw_by_name.get(oid)
        items_detail = []
        tray_count = 0
        for e in raw["lineItems"]["edges"]:
            n = e["node"]
            sku = (n.get("sku") or "").strip()
            cq = n.get("currentQuantity")
            q = int(cq if cq is not None else (n.get("quantity") or 0))
            if q <= 0 or not sku: continue
            if sku.startswith("TR-"):
                tray_count += q
            prefix = sku.split("-",1)[0] if "-" in sku else sku
            in_lookup = sku in lookup
            if not in_lookup and prefix not in PHYSICAL_PREFIXES: continue
            if sku.startswith("PK-"): continue
            dv, _ = resolve_distvol(sku, lookup)
            items_detail.append(f"{sku} x{q} ({round(dv*q, 2)})")
        rows.append({
            "OrderID": oid,
            "Customer": c["name"],
            "City": s["City"], "State": s["State"], "ZIP": s["ZIP"],
            "CSV_Box": c["csv_box"],
            "Should_Be": s["Box Assignment"],
            "DistVol": s["Total DistVol"],
            "Tray_Count": tray_count,
            "Item_Count": s["Item Count"],
            "Over_By_DV": round(s["Total DistVol"] - 2.99, 3),
            "Line_Items": " | ".join(items_detail),
        })
    rows.sort(key=lambda r: (-r["Tray_Count"], -r["DistVol"]))

    wb = Workbook()
    ws = wb.active
    ws.title = "Overpacked Small Box"
    cols = list(rows[0].keys())
    ws.append(cols)
    bold = Font(bold=True)
    fill = PatternFill("solid", fgColor="FFE4B5")
    redfill = PatternFill("solid", fgColor="FFB6B6")
    for c in ws[1]:
        c.font = bold
        c.fill = fill
    for r in rows:
        ws.append([r[c] for c in cols])
        # highlight tray-containing
        if r["Tray_Count"] > 0:
            for c in ws[ws.max_row]:
                c.fill = redfill
    # column widths
    widths = {"A":10,"B":22,"C":16,"D":6,"E":12,"F":22,"G":12,"H":9,"I":9,"J":9,"K":9,"L":120}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"

    # summary tab
    s2 = wb.create_sheet("Summary")
    s2.append(["Verification Date", "2026-04-29"])
    s2.append(["Cohort Tag", "_SHIP_2026-04-27 (SAT)"])
    s2.append(["Source CSV", str(CSV_PATH)])
    s2.append(["DistVol Truth", r"C:\Users\Work\Desktop\Onboarded Items with DistVol - Updated.xlsx (Sheet1)"])
    s2.append([])
    s2.append(["Issue", "Value"])
    s2.append(["Total CSV orders", 2206])
    s2.append(["Matched to Shopify", 2203])
    s2.append(["Box mismatches (this report)", len(rows)])
    s2.append(["With trays (must be Medium)", sum(1 for r in rows if r["Tray_Count"] > 0)])
    s2.append([])
    s2.append(["Rule", "Small Box max = 2.99 DV; Medium Box = 2.99-6.7 DV; trays force Medium"])

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Saved: {OUT}")
    print(f"Rows: {len(rows)}")
    for r in rows:
        flag = " <-- TRAYS" if r["Tray_Count"] > 0 else ""
        print(f"  {r['OrderID']}  trays={r['Tray_Count']}  DV={r['DistVol']}  CSV={r['CSV_Box']}{flag}")

if __name__ == "__main__":
    main()
