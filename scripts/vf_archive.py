"""vF archive DB — what we actually told RMFG to pack, per ship week, per product.

Constraints SSOT: AppyHour/VF_ARCHIVE_RULES.md (read it before changing anything here).
Source of record: C:\\AppyHourData\\vf_archive\\raw\\ — attachments off the messages we SENT.
A working copy in Downloads/_outputs is never ingested.

    python vf_archive.py ingest                 # scan raw/, load new files (idempotent, by hash)
    python vf_archive.py weeks                  # per ship-week roll-up: legs, orders, units
    python vf_archive.py product "KM39"         # per-week assigned for one product
    python vf_archive.py files --week 2026-08-31

Ship week = the UNION of that week's legs (Friday sheet + Tuesday sheet + drift-in re-runs).
"""

from __future__ import annotations

import argparse
import hashlib
import re
import sqlite3
import sys
from datetime import date, datetime, timedelta
from pathlib import Path

for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding="utf-8", errors="replace")  # type: ignore[union-attr]
    except AttributeError:
        pass

RAW_DIR = Path(r"C:/AppyHourData/vf_archive/raw")
DB_PATH = Path(r"C:/AppyHourData/vf_archive/vf_archive.db")
PRODUCT_PREFIX = "AHB (S_REG): "
FIXED_COLS = {
    "OrderID", "Name", "Distribution Type", "Total", "Phone Number", "Email",
    "Address", "Address 2", "City", "State", "Zip", "Tags", "Notes", "ProductionDay",
}

SCHEMA = """
CREATE TABLE IF NOT EXISTS vf_files (
    file_hash   TEXT PRIMARY KEY,
    filename    TEXT NOT NULL,
    sheet_date  TEXT,              -- date in the filename (the leg's own date)
    ship_week   TEXT,              -- Saturday-anchored week this leg belongs to
    leg         TEXT,              -- 'main' | 'tuesday' | 'driftin' | 'gift' | 'other'
    variant     TEXT,              -- vF / vFGR / _r2 / __dup2 ... verbatim suffix
    orders      INTEGER,
    product_cols INTEGER,
    units       INTEGER,
    bytes       INTEGER,
    source_path TEXT,
    ingested_at TEXT
);
CREATE TABLE IF NOT EXISTS vf_orders (
    file_hash TEXT, order_id TEXT, name TEXT, state TEXT, zip TEXT,
    tags TEXT, production_day TEXT, total REAL,
    PRIMARY KEY (file_hash, order_id)
);
CREATE TABLE IF NOT EXISTS vf_items (
    file_hash TEXT, order_id TEXT, mfg_name TEXT, qty REAL,
    PRIMARY KEY (file_hash, order_id, mfg_name)
);
CREATE INDEX IF NOT EXISTS ix_items_name ON vf_items (mfg_name);
CREATE INDEX IF NOT EXISTS ix_files_week ON vf_files (ship_week);
"""

DATE_RE = re.compile(r"_(\d{1,2})-(\d{1,2})-(\d{2})_")


def parse_sheet_date(name: str) -> date | None:
    m = DATE_RE.search(name)
    if not m:
        return None
    mm, dd, yy = (int(x) for x in m.groups())
    try:
        return date(2000 + yy, mm, dd)
    except ValueError:
        return None


def ship_week_of(d: date) -> str:
    """Anchor a leg to its week's Monday-shipping Saturday sheet date.

    The main sheet is dated the Saturday/Monday ship date; the Tuesday leg is +1..+3 days,
    and drift-in re-runs land later in the same week. Anchor everything to the most recent
    Monday-or-later Sunday boundary: weeks run Sunday->Saturday, labelled by their first day.
    """
    # weekday(): Mon=0 ... Sun=6. Anchor to the Sunday that opens the week.
    anchor = d - timedelta(days=(d.weekday() + 1) % 7)
    return anchor.isoformat()


def leg_of(d: date, name: str) -> str:
    low = name.lower()
    if "vfgr" in low:
        return "gift"
    wd = d.weekday()
    if wd == 0:  # Monday-dated main sheet
        return "main"
    if wd == 1:
        return "tuesday"
    return "driftin"


def variant_of(name: str) -> str:
    stem = Path(name).stem
    m = DATE_RE.search(stem + "_")
    return stem[m.end() - 1:].lstrip("_") if m else stem


def connect() -> sqlite3.Connection:
    DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    con = sqlite3.connect(DB_PATH)
    con.executescript(SCHEMA)
    return con


def ingest_file(con: sqlite3.Connection, path: Path) -> str:
    import openpyxl

    payload = path.read_bytes()
    fh = hashlib.sha256(payload).hexdigest()
    if con.execute("SELECT 1 FROM vf_files WHERE file_hash=?", (fh,)).fetchone():
        return "skip"

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        return "empty"

    # Columns by NAME, never index (VF_ARCHIVE_RULES rule 5).
    col = {str(h).strip(): i for i, h in enumerate(header) if h is not None}
    products = {k: i for k, i in col.items() if k.startswith(PRODUCT_PREFIX)}
    if not products:
        return "no-product-cols"

    def cell(r, key):
        i = col.get(key)
        return r[i] if i is not None and i < len(r) else None

    order_rows, item_rows, units, n = [], [], 0, 0
    for r in rows:
        oid = cell(r, "OrderID")
        if oid is None or str(oid).strip() == "":
            continue
        n += 1
        oid = str(oid).strip()
        order_rows.append((
            fh, oid, str(cell(r, "Name") or ""), str(cell(r, "State") or ""),
            str(cell(r, "Zip") or ""), str(cell(r, "Tags") or ""),
            str(cell(r, "ProductionDay") or ""),
            cell(r, "Total") if isinstance(cell(r, "Total"), (int, float)) else None,
        ))
        for hdr, i in products.items():
            v = r[i] if i < len(r) else None
            if isinstance(v, (int, float)) and v:  # blank/0/None = not in box (rule 6)
                item_rows.append((fh, oid, hdr[len(PRODUCT_PREFIX):], float(v)))
                units += v

    sd = parse_sheet_date(path.name)
    con.execute(
        "INSERT INTO vf_files VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
        (fh, path.name, sd.isoformat() if sd else None,
         ship_week_of(sd) if sd else None, leg_of(sd, path.name) if sd else "other",
         variant_of(path.name), n, len(products), int(units), len(payload),
         str(path), datetime.now().isoformat(timespec="seconds")),
    )
    con.executemany("INSERT OR REPLACE INTO vf_orders VALUES (?,?,?,?,?,?,?,?)", order_rows)
    con.executemany("INSERT OR REPLACE INTO vf_items VALUES (?,?,?,?)", item_rows)
    con.commit()
    return "new"


def cmd_ingest(args) -> int:
    con = connect()
    counts = {"new": 0, "skip": 0}
    for p in sorted(RAW_DIR.glob("*.xls*")):
        st = ingest_file(con, p)
        counts[st] = counts.get(st, 0) + 1
        if st not in ("skip",):
            print(f"  {st:14s} {p.name}")
    print(f"\n{counts} -> {DB_PATH}")
    return 0


# 🔴 Every file is KEPT (rule 2) but a week's count must use ONE file per leg, or a
# byte-different `__dup2` of the same sheet doubles the week: wk 2026-08-02 read 354 KM39
# against a true 178 because both copies of the main AND the Tuesday sheet were summed.
# Winner per (ship_week, leg, sheet_date) = most orders, then latest ingest. Losers are
# never deleted and `files` still lists them.
COUNTED = """
  SELECT file_hash FROM (
    SELECT file_hash,
           ROW_NUMBER() OVER (PARTITION BY ship_week, leg, sheet_date
                              ORDER BY orders DESC, ingested_at DESC) AS rn
    FROM vf_files WHERE ship_week IS NOT NULL AND leg != 'gift'
  ) WHERE rn = 1
"""


def cmd_weeks(args) -> int:
    con = connect()
    q = """
      SELECT ship_week,
             GROUP_CONCAT(leg || ':' || orders, '  ') AS legs,
             SUM(orders) AS orders, SUM(units) AS units, COUNT(*) AS files
      FROM vf_files WHERE file_hash IN (%s)
      GROUP BY ship_week ORDER BY ship_week DESC LIMIT ?
    """ % COUNTED
    print(f"{'ship_week':12s} {'orders':>7s} {'units':>8s} {'files':>5s}  legs")
    for w, legs, o, u, f in con.execute(q, (args.limit,)):
        print(f"{w:12s} {o:7d} {u:8d} {f:5d}  {legs}")
    return 0


def cmd_product(args) -> int:
    con = connect()
    q = """
      SELECT f.ship_week, SUM(i.qty) AS units, COUNT(DISTINCT i.order_id) AS orders,
             GROUP_CONCAT(DISTINCT f.leg)
      FROM vf_items i JOIN vf_files f USING (file_hash)
      WHERE i.mfg_name LIKE ? AND f.file_hash IN (%s)
      GROUP BY f.ship_week ORDER BY f.ship_week
    """ % COUNTED
    names = [r[0] for r in con.execute(
        "SELECT DISTINCT mfg_name FROM vf_items WHERE mfg_name LIKE ?", (f"%{args.name}%",))]
    if not names:
        print(f"no product header matching {args.name!r} — not guessing a mapping")
        return 1
    print("matched headers:", ", ".join(sorted(names)))
    print(f"\n{'ship_week':12s} {'units':>7s} {'orders':>7s}  legs")
    for w, u, o, legs in con.execute(q, (f"%{args.name}%",)):
        print(f"{w:12s} {u:7.0f} {o:7d}  {legs}")
    return 0


def cmd_files(args) -> int:
    con = connect()
    q = "SELECT ship_week, sheet_date, leg, variant, orders, units, filename FROM vf_files"
    p: tuple = ()
    if args.week:
        q += " WHERE ship_week=?"
        p = (args.week,)
    for row in con.execute(q + " ORDER BY sheet_date, filename", p):
        print("  ".join(str(x) for x in row))
    return 0


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    sub = ap.add_subparsers(dest="cmd", required=True)
    sub.add_parser("ingest").set_defaults(fn=cmd_ingest)
    w = sub.add_parser("weeks"); w.add_argument("--limit", type=int, default=20); w.set_defaults(fn=cmd_weeks)
    pr = sub.add_parser("product"); pr.add_argument("name"); pr.set_defaults(fn=cmd_product)
    fl = sub.add_parser("files"); fl.add_argument("--week"); fl.set_defaults(fn=cmd_files)
    args = ap.parse_args()
    return args.fn(args)


if __name__ == "__main__":
    raise SystemExit(main())
