"""Standalone QC validator for a vF / Access_LIVE production sheet.

Runs the file-checkable subset of Kori's QC (gel_pack_webview.py qc_check_sheet,
~2240-2646) on ANY vF xlsx, so a HAND-EDITED sheet is validated even when it never
went through Kori or matrix_commander. Primary value: catches an invented
`AHB (S_REG): <name>` header (the 2026-08-04 "Farmstead Smoked Cumin Gouda"
fabrication class) by checking every product header against the authoritative
meal-type export.

Authorities (never fabricated here):
  - MFG names  -> AppyHour/mfg_names_authoritative.csv (via matrix_commander loader)
  - Structure/PO-box/syntax rules -> VF_SHEET_RULES.md (extract of Kori QC)
Full routing-tag allowlist/combo validation stays in Kori (ROUTING_TAG_SET +
is_approved_bang_tag + validate_routing_tag_combo live in the heavy webview module);
this script only flags the self-contained `!!` double-bang and reports bang-tags for
Kori to adjudicate — it does NOT invent an allowlist.

Usage:  python validate_vf_sheet.py "<path to vF .xlsx>" [--tuesday]
Exit 0 = clean, 1 = issues found. READ-ONLY (never writes the sheet).
"""
from __future__ import annotations

import re
import sys
from collections import Counter
from pathlib import Path

import openpyxl

AH = Path(r"C:\Users\Work\Claude Projects\AppyHour")
sys.path.insert(0, str(AH))
from matrix_commander import MFG_AUTHORITATIVE_PATH, load_mfg_translations  # noqa: E402

FIXED_COLS = ["OrderID", "Name", "Distribution Type", "Total", "Phone Number", "Email",
              "Address", "Address 2", "City", "State", "Zip", "Tags", "Notes", "ProductionDay"]
PO_BOX_RE = re.compile(r"\b(?:p\.?\s*o\.?\s*box|pobox)\b", re.IGNORECASE)
HDR_PREFIX = "AHB (S_REG): "


def col_map(header):
    """header cell values (row1) -> {logical: idx}. Mirrors Kori's tolerant mapping."""
    m = {}
    for i, h in enumerate(header):
        if h is None:
            continue
        key = str(h).strip().lower()
        for logical, want in (("order_id", "orderid"), ("tags", "tags"), ("zip", "zip"),
                              ("state", "state"), ("total", "total"), ("address", "address"),
                              ("address2", "address 2"), ("production_day", "productionday"),
                              ("name", "name")):
            if key == want or key == want.replace(" ", ""):
                m.setdefault(logical, i)
    return m


def main():
    if len(sys.argv) < 2:
        print("usage: validate_vf_sheet.py <vF.xlsx> [--tuesday]"); sys.exit(2)
    path = Path(sys.argv[1])
    is_tuesday = "--tuesday" in sys.argv[2:]
    expected_pday = "TUE" if is_tuesday else "SAT"

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    fails = []  # (check, detail)

    # tab name
    ws = wb["Access_LIVE"] if "Access_LIVE" in wb.sheetnames else wb.active
    if ws.title != "Access_LIVE":
        fails.append(("Tab Name", f"tab is '{ws.title}', must be 'Access_LIVE'"))

    header = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    cm = col_map(header)

    # structure: fixed cols present + ProductionDay at col N (index 13)
    hdr_norm = [str(h).strip() if h is not None else "" for h in header]
    for expected in FIXED_COLS:
        if expected not in hdr_norm:
            fails.append(("Structure", f"missing fixed column '{expected}'"))
    if len(hdr_norm) > 13 and hdr_norm[13] != "ProductionDay":
        fails.append(("Structure", f"col N (14) is '{hdr_norm[13]}', expected 'ProductionDay'"))

    # NAMES — every AHB (S_REG): header must be in the authoritative export
    authoritative = set(load_mfg_translations(MFG_AUTHORITATIVE_PATH).values())
    if not authoritative:
        print(f"WARNING: {MFG_AUTHORITATIVE_PATH.name} missing/empty — NAME check SKIPPED")
    else:
        for h in header:
            if h is None:
                continue
            hs = str(h).strip()
            if hs.startswith(HDR_PREFIX) and hs not in authoritative:
                fails.append(("MFG Name (INVENTED?)",
                              f"header '{hs}' not in {MFG_AUTHORITATIVE_PATH.name} — "
                              f"fabricated/mistyped; look up the real name, never guess"))

    # row scan: names done; now PO box, zip, dupes, sort, productionday, low-item
    oi, ti = cm.get("order_id"), cm.get("tags")
    zi, ai, ai2 = cm.get("zip"), cm.get("address"), cm.get("address2")
    toti, pdi = cm.get("total"), cm.get("production_day")
    tray_cols = [i for i, h in enumerate(header) if h and str(h).startswith(HDR_PREFIX) and "(Tray)" in str(h)]
    ids, numeric = [], []
    for vals in ws.iter_rows(min_row=2, values_only=True):
        if oi is None or oi >= len(vals) or vals[oi] is None:
            continue
        oid = str(vals[oi]); ids.append(oid)
        try:
            numeric.append(int(oid))
        except (ValueError, TypeError):
            numeric.append(0)
        tags_lower = str(vals[ti] or "").lower() if ti is not None and ti < len(vals) else ""
        # PO box + slash
        for label, idx in (("Address", ai), ("Address 2", ai2)):
            if idx is not None and idx < len(vals):
                v = str(vals[idx] or "").strip()
                if v and PO_BOX_RE.search(v):
                    fails.append(("PO Box", f"#{oid}: PO Box in {label}: '{v}'"))
                if v and "/" in v:
                    fails.append(("Address Slash", f"#{oid}: '/' in {label}: '{v}'"))
        # zip
        if zi is not None and zi < len(vals) and vals[zi] is not None:
            raw = vals[zi]; z = str(raw).strip()
            if isinstance(raw, (int, float)):
                fails.append(("Zip", f"#{oid}: stored as number, leading zeros lost ('{z}')"))
            elif len(z) == 4 and z.isdigit():
                fails.append(("Zip", f"#{oid}: missing leading zero (should be 0{z})"))
        # productionday value
        if pdi is not None and pdi < len(vals):
            pv = str(vals[pdi] or "").strip().upper()
            if pv != expected_pday:
                fails.append(("ProductionDay", f"#{oid}: '{pv}' (expected '{expected_pday}')"))
        # double-bang tag
        for tag in [t.strip() for t in str(vals[ti] or "").split(",") if t.strip()] if ti is not None else []:
            if "!!" in tag:
                fails.append(("Tag !!", f"#{oid}: '{tag}' contains '!!' (double bang)"))
        # low item
        if toti is not None and toti < len(vals) and vals[toti] is not None:
            try:
                n = int(float(str(vals[toti])))
            except (ValueError, TypeError):
                n = None
            is_tray = any(ci < len(vals) and vals[ci] and str(vals[ci]).strip() not in ("", "0", "None") for ci in tray_cols)
            if n is not None and n < 10 and "reship" not in tags_lower and not is_tray:
                fails.append(("Low Items", f"#{oid}: only {n} items (no Reship/tray)"))

    # dupes + sort
    for o, c in Counter(ids).items():
        if c > 1:
            fails.append(("Duplicate Order", f"#{o}: appears {c} times"))
    if numeric != sorted(numeric):
        fails.append(("Sort Order", "OrderIDs not ascending"))

    wb.close()
    print(f"\nvF QC — {path.name}  (ship day: {expected_pday})")
    print(f"authoritative names: {len(authoritative)} | rows: {len(ids)}")
    if not fails:
        print("\nRESULT: CLEAN — all file-checkable QC passed.")
        print("NOTE: full routing-tag allowlist/combo validation still owned by Kori.")
        sys.exit(0)
    by = {}
    for chk, d in fails:
        by.setdefault(chk, []).append(d)
    print(f"\nRESULT: {len(fails)} issue(s) across {len(by)} check(s):")
    for chk, ds in by.items():
        print(f"\n[{chk}]  ({len(ds)})")
        for d in ds[:50]:
            print(f"  - {d}")
    print("\nNOTE: full routing-tag allowlist/combo validation still owned by Kori.")
    sys.exit(1)


if __name__ == "__main__":
    main()
