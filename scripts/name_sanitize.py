"""name_sanitize.py — strip RMFG-hostile characters from the vF `Name` column. Report-only DEFAULT.

🔴 WHY THIS EXISTS (total-loss class, Kurt 2026-08-09)
A special character in a name REJECTS THE WHOLE SHEET at RMFG — not the row, the sheet. One
instance holds the entire week's production until a corrected sheet is resubmitted. Nothing
anywhere validated characters before this: `qc_gate` has PO-box regex validation, `presend_check`
has ice/filename/ledger gates, `matrix_commander` has MFG-name validation — none of them look at
characters. Kurt believed it was covered. It was not.

Kurt's call on cosmetics, verbatim: "it doesn't matter if they looked fucked." Appearance is
overruled ON PURPOSE. That does NOT extend to meaning — see SCOPE.

🔴 SCOPE IS DELIBERATELY NARROW — DO NOT WIDEN WITHOUT KURT

  Name column ONLY.

  NEVER addresses. Survey of the submitted wk0810 sheet found `/` in real street numbers:
  `3226 1/2 LA AVENIDA DE SAN MARCOS` and `2714 8 1/2 ST APT 4`. Stripping there gives
  `3226 12 ...` — a DIFFERENT HOUSE. That is undeliverable mail, not an ugly label, and it is
  the one case the cosmetic overrule does not reach. Permanently out of scope.

  NEVER the `AHB (S_REG):` product headers. All 121 of them contain zero special characters in
  their cell values, and a stripped MFG header would itself be an INVENTED MFG NAME — which is
  its own whole-sheet reject ([[sku-mfg-name-validation-gate]], "we never make up MFG names").

🔴 THE MOTIVATING BURN — order #169525, `Shinsato/Johnson` (Kurt 2026-08-09, verbatim):

    "shinsato/johnson was found by them and I got yelled at for it."

That slash SHIPPED in the submitted sheet of record
(`_outputs/artifacts/vF-intent-2026-08-10/AHB_WeeklyProductionQuery_08-10-26_vF.xlsx`), RMFG caught
it, and Kurt took the complaint. This is not a theoretical reject class — it happened, to us, and
somebody was yelled at. Anyone tempted to soften this rule to "but it mangles surnames" should note
that argument was made and OVERRULED, twice: "it doesn't matter if they looked fucked" and
"just strip it."

🔴 RULES — A PIPELINE, AND THE ORDER IS LOAD-BEARING (Kurt 2026-08-09)

    1. Remove the C/O token   (case-insensitive)   "C/O Jane Smith"    -> "Jane Smith"
    2. Fraction -> word        1/2 -> half          "3226 1/2"          -> "3226 half"
    3. Remaining "/" REMOVED  (no space)            "Shinsato/Johnson"  -> "ShinsatoJohnson"
    4. Strip every remaining non-alphanumeric EXCEPT space and HYPHEN U+002D
    5. Collapse whitespace runs, trim

  Steps 1 and 2 MUST run before step 3 or they self-corrupt: reversed, "C/O Jane" becomes
  "COJane" and "1/2" becomes "12". The ordering is tested explicitly — it is the part most
  likely to regress under a later "simplification".

  KEEP letters, digits, space, hyphen ("you can keep the hyphens" — hyphenated surnames intact).
  `O'Brien` -> `OBrien`; appearance is explicitly NOT the constraint.
  Step 5 also cleans pre-existing double-space artifacts (`orene  r contini`, `Gary M  Bell`).

🔴 THE LOG IS THE DELIVERABLE. Kurt writes the permanent rule from real data, so the log must be
sufficient on its own — order number, column, before, after, and every removed character with its
unicode codepoint and name. One row per changed name, to a dated artifact under _outputs/reports/.

🔴 NEVER OVERWRITES ITS INPUT. Kurt's submitted artifacts are the record of what shipped
([[never-delete-prior-output-files]]). Default is report-only; --out writes a NEW file and refuses
to clobber an existing one.

Usage:
    python scripts/name_sanitize.py <vF.xlsx>                 # report + log only, no xlsx written
    python scripts/name_sanitize.py <vF.xlsx> --out <new.xlsx> # also write a sanitized COPY
    python scripts/name_sanitize.py <a.xlsx> <b.xlsx> --compare  # counts per file, no writes
"""
from __future__ import annotations

import argparse
import csv
import datetime as _dt
import re
import sys
import unicodedata
from pathlib import Path

for _s in (sys.stdout, sys.stderr):          # Windows cp1252 crashes on non-ASCII
    if hasattr(_s, "reconfigure"):
        _s.reconfigure(encoding="utf-8", errors="replace")

import openpyxl  # noqa: E402

NAME_COL = "Name"
SLASH = "/"
HYPHEN = "-"
REPORTS = Path(__file__).resolve().parents[2] / "_outputs" / "reports"

# 🔴 Step 1 — the C/O token. Case-insensitive, tolerates spaces around the slash and a trailing
# period. Requires the slash, so it can never fire on a name that merely contains "co".
CO_RE = re.compile(r"(?i)\bC\s*/\s*O\b\.?")

# 🔴 Step 2 — fraction -> word. Kurt 2026-08-09, verbatim: "1/2 turned into half. that's it."
# THIS EXACT TOKEN ONLY. Do NOT generalize to 1/4, 3/4 or anything else: those spellings would be
# INVENTED rather than observed, and inventing a token a person at RMFG reads is the same class of
# error as an invented MFG name, just smaller ([[never-fabricate]], [[sku-mfg-name-validation-gate]]).
# An unknown fraction is CORRECTLY handled by falling through to step 3 (slash removed) and landing
# in the change log — where Kurt sees it and decides. Unknown case gets LOGGED, never guessed at.
FRACTION_WORDS = {"1/2": "half"}


def is_kept(ch: str) -> bool:
    """KEEP letters, digits, space, hyphen. Everything else goes.

    `isalpha()` is unicode-aware, so accented letters (Renée, José, Zoë) are KEPT — the rule strips
    PUNCTUATION, and an accented character is a letter. Do not add a transliteration pass and do
    not chase RMFG's charset: that is a cosmetic question, and cosmetic questions lose here
    (see NAME_SANITIZE_RULES.md "Calibration").
    """
    return ch.isalpha() or ch.isdigit() or ch == " " or ch == HYPHEN


def sanitize(name: str) -> tuple[str, list[tuple[str, str, str]], list[str]]:
    """-> (clean, removed, transforms).

    removed    = [(char, 'U+XXXX', unicode-name), ...] for every character deleted in step 3/4
    transforms = human-readable token-level steps applied ('C/O REMOVED', '1/2 -> half')

    🔴 The five steps run IN ORDER. See module docstring — reordering silently corrupts.
    """
    transforms: list[str] = []
    s = name

    # --- 1. C/O token (Kurt: "c/o is not tolerated" -> remove the token, keep the name)
    s, n = CO_RE.subn(" ", s)
    if n:
        transforms.append(f"C/O token REMOVED x{n}")

    # --- 2. fraction -> word, BEFORE any general slash handling (else 1/2 -> 12)
    for frac, word in FRACTION_WORDS.items():
        if frac in s:
            s = s.replace(frac, word)
            transforms.append(f"{frac} -> {word}")

    # --- 3/4. remaining '/' is REMOVED (Kurt: "just strip it"), as is every other
    #          non-alphanumeric except space and hyphen.
    removed: list[tuple[str, str, str]] = []
    out: list[str] = []
    for ch in s:
        if is_kept(ch):
            out.append(ch)
            continue
        removed.append((ch, f"U+{ord(ch):04X}", _uname(ch)))

    # --- 5. collapse whitespace runs + trim
    clean = " ".join("".join(out).split())
    return clean, removed, transforms


def _uname(ch: str) -> str:
    try:
        return unicodedata.name(ch)
    except ValueError:
        return "<unnamed>"


def scan(path: Path) -> tuple[list[dict], int]:
    """-> (changes, rows_scanned). Read-only."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    hdr = [str(h) if h is not None else "" for h in next(it)]
    if NAME_COL not in hdr:
        wb.close()
        raise SystemExit(f"FAIL: no '{NAME_COL}' column in {path.name} (cols: {hdr[:8]})")
    iN = hdr.index(NAME_COL)
    iO = hdr.index("OrderID") if "OrderID" in hdr else 0

    changes, rows = [], 0
    for r in it:
        if r[iN] is None:
            continue
        rows += 1
        before = str(r[iN])
        after, removed, transforms = sanitize(before)
        oid = str(r[iO] or "").lstrip("#").strip()
        if after != before:
            acts = list(transforms) + [f"{c} {cp} REMOVED" for c, cp, _ in removed]
            changes.append({
                "order": oid, "column": NAME_COL, "before": before, "after": after,
                "removed": removed, "transforms": transforms,
                "actions": "; ".join(acts) or "WHITESPACE-COLLAPSE-ONLY",
            })
    wb.close()
    return changes, rows


def write_log(changes, rows, src: Path, stamp: str) -> tuple[Path, Path]:
    REPORTS.mkdir(parents=True, exist_ok=True)
    base = f"{stamp}-name-sanitize-{src.stem}"
    md, cs = REPORTS / f"{base}.md", REPORTS / f"{base}.csv"

    with open(cs, "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(["order", "column", "before", "after", "chars_removed", "codepoints", "actions"])
        for c in changes:
            w.writerow([c["order"], c["column"], c["before"], c["after"],
                        "".join(ch for ch, _, _ in c["removed"]),
                        " ".join(cp for _, cp, _ in c["removed"]), c["actions"]])

    lines = [
        f"# Name sanitize log — `{src.name}`", "",
        f"Generated {stamp} · source `{src}` · **read-only, source not modified**", "",
        f"- rows scanned: **{rows}**",
        f"- names changed: **{len(changes)}**", "",
        "Rules: KEEP letters/digits/space/hyphen · `/`->SPACE · else REMOVED · collapse+trim.",
        "Scope: `Name` column only — never addresses, never `AHB (S_REG):` headers.", "",
        "## Changes", "",
        "| order | before | after | removed | codepoints |",
        "|---|---|---|---|---|",
    ]
    for c in changes:
        rm = " ".join(f"`{ch}`" for ch, _, _ in c["removed"]) or "_(whitespace only)_"
        cp = " ".join(cp for _, cp, _ in c["removed"]) or "-"
        lines.append(f"| {c['order']} | `{c['before']}` | `{c['after']}` | {rm} | {cp} |")

    from collections import Counter
    tally = Counter((ch, cp, nm) for c in changes for ch, cp, nm in c["removed"])
    lines += ["", "## Character frequency — write the permanent rule from THIS", "",
              "| char | codepoint | unicode name | count | action |", "|---|---|---|---|---|"]
    for (ch, cp, nm), n in tally.most_common():
        lines.append(f"| `{ch}` | {cp} | {nm} | {n} | "
                     f"{'SLASH->SPACE' if ch == SLASH else 'REMOVED'} |")

    # 🔴 NO "needs review" SECTION. Every change above is an ordinary change. A decision queue
    # that isn't a decision costs Kurt's attention, which is the scarce thing here.
    md.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return md, cs


def write_sanitized(src: Path, out: Path) -> int:
    """Write a sanitized COPY. Refuses to overwrite. Never touches `src`."""
    if out.exists():
        raise SystemExit(f"REFUSED: {out} exists — never overwrite a prior output "
                         f"([[never-delete-prior-output-files]]). Pick a new name.")
    if out.resolve() == src.resolve():
        raise SystemExit("REFUSED: --out equals the input. This tool never modifies its source.")
    wb = openpyxl.load_workbook(src)                  # not read_only: we must save a copy
    ws = wb[wb.sheetnames[0]]
    hdr = [str(c.value) if c.value is not None else "" for c in ws[1]]
    iN = hdr.index(NAME_COL) + 1
    n = 0
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=iN)
        if cell.value is None:
            continue
        before = str(cell.value)
        after, _, _ = sanitize(before)
        if after != before:
            cell.value = after
            n += 1
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    wb.close()
    return n


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(description="Strip RMFG-hostile characters from the vF Name column.")
    ap.add_argument("xlsx", nargs="+")
    ap.add_argument("--out", default=None, help="write a sanitized COPY here (never overwrites)")
    ap.add_argument("--compare", action="store_true", help="counts for each input, no logs/writes")
    a = ap.parse_args(argv)
    stamp = _dt.date.today().isoformat()

    if a.compare:
        for p in a.xlsx:
            src = Path(p)
            changes, rows = scan(src)
            print(f"{src.name}: {rows} rows, {len(changes)} names changed")
        return 0

    if a.out and len(a.xlsx) > 1:
        raise SystemExit("--out takes exactly one input file.")

    rc = 0
    for p in a.xlsx:
        src = Path(p)
        if not src.exists():
            raise SystemExit(f"FAIL: no such file {src}")
        changes, rows = scan(src)
        md, cs = write_log(changes, rows, src, stamp)
        print(f"{src.name}: {rows} rows scanned, {len(changes)} names changed")
        for c in changes[:60]:
            print(f"   #{c['order']}  {c['before']!r} -> {c['after']!r}   [{c['actions']}]")
        print(f"log: {md}\n     {cs}")
        if a.out:
            n = write_sanitized(src, Path(a.out))
            print(f"sanitized copy written: {a.out}  ({n} names changed)")
    return rc


if __name__ == "__main__":
    sys.exit(main())
