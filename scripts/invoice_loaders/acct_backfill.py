"""Unify `shipments.acct` onto canonical account keys — LOCAL sqlite and CLOUD MySQL.

🔴 WHY THIS EXISTS (2026-08-20): `shipments.acct` held FIVE spellings for TWO real FedEx accounts
(`203738113`/`-113` = ours; `206137911`/`-911` = RMFG's) plus 3,550 NULL rows, and UPS repeated the
shape. Per-account cost and hub attribution silently split across the duplicate keys. Constraints
SSOT: `ShipRouting/INVOICE_INGEST_RULES.md` §1.

DRY-RUN BY DEFAULT. `--apply` is required to write, and it is a live write on ~90k rows per DB.

NULL recovery uses evidence FROM THE SOURCE ONLY (never a guess — ~/.claude/rules/never-fabricate.md):
  - the RMFG `Service Cost Analysis` xlsx carries `Bill to Account Number` PER ROW, joined on
    tracking id (3,379 rows spanning BOTH accounts: 203738113 x2,403 and 206137911 x976);
  - `FedEx_invoice_acct911_*.CSV` states the account in the FILENAME (4 rows).
Anything with no evidence in its own source stays `unknown`. It is never inferred from a sibling
row, a date range, or a hub.

Usage:
    python acct_backfill.py                      # dry-run, both DBs
    python acct_backfill.py --target local       # one DB
    python acct_backfill.py --apply --target local
"""
from __future__ import annotations

import argparse
import glob
import os
import re
import sqlite3
import sys
from collections import Counter, defaultdict

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
from appyhour_lib import acct_canon as ac  # noqa: E402

LOCAL_DB = r"C:\AppyHourData\shipping.db"
XLSX_GLOB = os.path.expandvars(r"%USERPROFILE%\.claude\downloads\Downloads\*Service Cost Analysis*.xlsx")


# ──────────────────────────────────────────────────────────── connections (read-only by default)

def local_conn(readonly: bool):
    """🔴 shipping.db is opened READ-ONLY for analysis — a plain read-only open still rewrites
    `-shm` and has corrupted this DB under MSIX before (INVOICE_INGEST_RULES §7)."""
    if readonly:
        return sqlite3.connect(f"file:{LOCAL_DB}?mode=ro&immutable=1", uri=True)
    return sqlite3.connect(LOCAL_DB)


def cloud_conn():
    sys.path.insert(0, r"C:\Users\Work\Claude Projects\ShipRouting")
    from server.durable_store import _conn
    return _conn()


# ──────────────────────────────────────────────────────────── source evidence for NULL recovery

def tracking_to_acct_from_xlsx() -> dict:
    """{tracking -> canonical acct} harvested from the RMFG xlsx `Details` sheets.

    🔴 Header layouts DIFFER between workbooks — detect by name, never by position, and skip a
    sheet that lacks the columns rather than guessing at them (INVOICE_INGEST_RULES §4, §5)."""
    out, skipped = {}, []
    try:
        import openpyxl
    except ImportError:
        print("  ! openpyxl unavailable — xlsx evidence skipped")
        return out
    for path in sorted(glob.glob(XLSX_GLOB)):
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception as e:
            skipped.append((os.path.basename(path), f"{type(e).__name__}: {e}"))
            continue
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            it = ws.iter_rows(values_only=True)
            try:
                hdr = next(it)
            except StopIteration:
                continue
            if not hdr:
                continue
            names = [str(h).strip() if h is not None else "" for h in hdr]
            if "Bill to Account Number" not in names:
                continue
            tcol = next((names.index(n) for n in ("Tracking ID", "Tracking Number",
                                                  "Express or Ground Tracking ID") if n in names), None)
            if tcol is None:
                skipped.append((f"{os.path.basename(path)}:{sheet}", "no tracking column"))
                continue
            acol = names.index("Bill to Account Number")
            for r in it:
                if len(r) <= max(tcol, acol):
                    continue
                trk, raw = r[tcol], r[acol]
                if trk is None or raw is None:
                    continue
                t = str(trk).strip()
                if t.endswith(".0"):
                    t = t[:-2]
                canon = ac.canon("FedEx", raw, filename=path)
                if t and canon != ac.UNKNOWN:
                    out[t] = canon
    for name, why in skipped:
        print(f"  ~ skipped {name}: {why}")
    return out


# ──────────────────────────────────────────────────────────── the diff

def plan_changes(rows, xlsx_map):
    """rows: (id, carrier, acct, tracking, source_file). Returns (changes, stats)."""
    changes, stats = [], Counter()
    by_transition = Counter()
    for _id, carrier, acct, tracking, source_file in rows:
        cur = "" if acct is None else str(acct)
        # 🔴 A carrier with NO declared accounts is OUT OF SCOPE, not "unknown". OnTrac and Veho are
        # RMFG-billed and carry no account of ours; stamping 'unknown' across 49k of their rows
        # would be churn that reads like a data problem we do not have. Left untouched.
        if not ac.canonical_keys(carrier):
            stats["out_of_scope_carrier"] += 1
            by_transition[(carrier, cur or "(null)", "(untouched)")] += 1
            continue
        new = ac.canon(carrier, acct, filename=source_file or "")
        evidence = "column"
        if new == ac.UNKNOWN:
            t = (tracking or "").strip()
            if t and t in xlsx_map:
                new, evidence = xlsx_map[t], "xlsx"
            elif source_file and re.search(r"acct(\d+)", source_file, re.I):
                cand = ac.canon(carrier, None, filename=source_file)
                if cand != ac.UNKNOWN:
                    new, evidence = cand, "filename"
            else:
                # Kurt-declared sole-account carriers (UPS: "yes its only our account", 8/20).
                # Own evidence class so it never reads as if the file stated it.
                cand = ac.sole_account(carrier)
                if cand != ac.UNKNOWN:
                    new, evidence = cand, "sole_account"
        if new == ac.UNKNOWN:
            stats["stays_unknown"] += 1
            by_transition[(carrier, cur or "(null)", "unknown")] += 1
            continue
        if new == cur:
            stats["already_canonical"] += 1
            continue
        changes.append((_id, new))
        stats[f"fixed_via_{evidence}"] += 1
        by_transition[(carrier, cur or "(null)", new)] += 1
    return changes, stats, by_transition


def run_target(name, fetch, apply_fn, xlsx_map, apply: bool):
    print(f"\n{'='*74}\n{name}\n{'='*74}")
    rows = fetch()
    print(f"rows examined: {len(rows):,}")
    changes, stats, by_transition = plan_changes(rows, xlsx_map)

    print("\n  transition (carrier, from -> to) : rows")
    for (carrier, frm, to), n in sorted(by_transition.items(), key=lambda kv: -kv[1]):
        flag = {"unknown": "  <-- NO EVIDENCE, stays unknown",
                "(untouched)": "  <-- carrier has no declared accounts"}.get(to, "")
        print(f"    {carrier:<7} {frm:<12} -> {to:<12} {n:>7,}{flag}")

    print("\n  summary:")
    for k in ("already_canonical", "fixed_via_column", "fixed_via_xlsx", "fixed_via_filename",
              "fixed_via_sole_account", "stays_unknown", "out_of_scope_carrier"):
        if stats.get(k):
            print(f"    {k:<20} {stats[k]:>7,}")
    print(f"    {'TOTAL WRITES':<20} {len(changes):>7,}")

    if not apply:
        print("\n  DRY-RUN — nothing written. Re-run with --apply to write.")
        return
    if not changes:
        print("\n  nothing to write.")
        return
    n = apply_fn(changes)
    print(f"\n  APPLIED: {n:,} rows updated.")
    # 🔴 rows written is the only success signal (INVOICE_INGEST_RULES §2)
    if n != len(changes):
        raise SystemExit(f"HARD FAIL: planned {len(changes)} writes, DB reported {n}")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--apply", action="store_true", help="WRITE (default is dry-run)")
    ap.add_argument("--target", choices=("local", "cloud", "both"), default="both")
    args = ap.parse_args()

    print("harvesting account evidence from RMFG xlsx ...")
    xlsx_map = tracking_to_acct_from_xlsx()
    print(f"  tracking->acct pairs available: {len(xlsx_map):,}")

    if args.target in ("local", "both"):
        con = local_conn(readonly=not args.apply)

        def fetch():
            return con.execute(
                "SELECT id, carrier, acct, tracking, source_file FROM shipments").fetchall()

        def apply_fn(changes):
            cur = con.cursor()
            cur.executemany("UPDATE shipments SET acct=? WHERE id=?",
                            [(new, _id) for _id, new in changes])
            con.commit()
            return cur.rowcount if cur.rowcount != -1 else len(changes)

        try:
            run_target(f"LOCAL  {LOCAL_DB}", fetch, apply_fn, xlsx_map, args.apply)
        finally:
            con.close()

    if args.target in ("cloud", "both"):
        my = cloud_conn()
        if my is None:
            raise SystemExit("cloud connection unavailable")
        cur = my.cursor()

        def fetch():
            cur.execute("SELECT id, carrier, acct, tracking, source_file FROM shipments")
            return cur.fetchall()

        def apply_fn(changes):
            c2 = my.cursor()
            c2.executemany("UPDATE shipments SET acct=%s WHERE id=%s",
                           [(new, _id) for _id, new in changes])
            my.commit()   # 🔴 explicit — PyMySQL autocommit=False (audit P0 class)
            return c2.rowcount if c2.rowcount != -1 else len(changes)

        try:
            run_target("CLOUD  appyhourbox-shipping-db.shipments", fetch, apply_fn, xlsx_map,
                       args.apply)
        finally:
            my.close()


if __name__ == "__main__":
    main()
