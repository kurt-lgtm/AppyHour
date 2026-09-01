"""Set-based `shipments.acct` canonicalization for the CLOUD. Same decisions, ~6 statements.

🔴 WHY THIS EXISTS instead of `acct_backfill.py --target cloud`:
that version issues one `UPDATE ... WHERE id=%s` per row through `executemany`, which PyMySQL does
NOT batch for UPDATEs. Measured 2026-08-20 against nyc3: **4.4 rows/sec** — 11,098 rows in 42
minutes, ~2.1 hours projected for 33,411. The work is set-based and the row loop was the mistake.
The run was killed at 25,723 rows by a process exit and rolled back cleanly, which is the one thing
the single-transaction design got right: the table was left exactly as it started, not half-canonical.

Same evidence classes as the row version, unchanged (INVOICE_INGEST_RULES §1):
  - alias collapse: the declared aliases of each real account, from `acct_canon.ACCOUNTS`
  - xlsx recovery: `Bill to Account Number` PER ROW, joined on tracking, for the 3,378 rows whose
    loader dropped the field
  - filename recovery: `FedEx_invoice_acct911_*.CSV` states the account in its NAME (4 rows)
  - sole-account: UPS, Kurt-declared 2026-08-20 ("yes its only our account") — its own class, never
    folded into `canon()`
  - everything else -> the literal `unknown`. NEVER a guess from a sibling row, date range, or hub.

Carriers with no declared accounts (OnTrac, Veho — 48k rows) are UNTOUCHED, not stamped `unknown`:
churn that would read like a data problem we do not have.

DRY-RUN by default. `--apply` writes.
"""
from __future__ import annotations

import argparse
import glob
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
sys.path.insert(0, r"C:\Users\Work\Claude Projects\ShipRouting")
from appyhour_lib import acct_canon as ac  # noqa: E402

XLSX_GLOB = os.path.expandvars(r"%USERPROFILE%\.claude\downloads\Downloads\*Service Cost Analysis*.xlsx")
MAP_TABLE = "_acct_recovery_map"


def _conn():
    from server.durable_store import _conn as c
    return c()


def xlsx_map() -> dict:
    """{tracking -> canonical acct}, harvested by COLUMN NAME (layouts differ between workbooks)."""
    out = {}
    try:
        import openpyxl
    except ImportError:
        print("  ! openpyxl unavailable — xlsx evidence skipped")
        return out
    for path in sorted(glob.glob(XLSX_GLOB)):
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception as e:                               # noqa: BLE001
            print(f"  ~ skip {os.path.basename(path)}: {type(e).__name__}")
            continue
        for sheet in wb.sheetnames:
            it = wb[sheet].iter_rows(values_only=True)
            try:
                hdr = next(it)
            except StopIteration:
                continue
            names = [str(h).strip() if h is not None else "" for h in (hdr or ())]
            if "Bill to Account Number" not in names:
                continue
            tcol = next((names.index(n) for n in ("Tracking ID", "Tracking Number",
                                                  "Express or Ground Tracking ID") if n in names), None)
            if tcol is None:
                continue
            acol = names.index("Bill to Account Number")
            for r in it:
                if len(r) <= max(tcol, acol) or r[tcol] is None or r[acol] is None:
                    continue
                t = str(r[tcol]).strip()
                if t.endswith(".0"):
                    t = t[:-2]
                canon = ac.canon("FedEx", r[acol], filename=path)
                if t and canon != ac.UNKNOWN:
                    out[t] = canon
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--apply", action="store_true")
    args = ap.parse_args()

    print("harvesting xlsx account evidence ...")
    tmap = xlsx_map()
    print(f"  tracking->acct pairs: {len(tmap):,}")

    my = _conn()
    if my is None:
        raise SystemExit("cloud connection unavailable")
    cur = my.cursor()

    # Aliases straight from the AUTHORITY — never a hand-typed list that can drift from it.
    plan = []
    for a in ac.ACCOUNTS:
        aliases = sorted({x for x in a.aliases if x != a.canonical} |
                         {a.canonical + ".0"})
        plan.append((a.carrier, a.canonical, aliases))

    print("\n  planned set-based statements:")
    total = 0
    for carrier, canonical, aliases in plan:
        marks = ",".join(["%s"] * len(aliases))
        cur.execute(f"SELECT COUNT(*) FROM shipments WHERE carrier=%s AND acct IN ({marks})",
                    [carrier] + aliases)
        n = cur.fetchone()[0]
        total += n
        print(f"    {carrier:<6} alias->{canonical:<12} {n:>7,} rows  (aliases {aliases})")

    # UPS sole-account: Kurt-declared, applied only to blank/NULL rows of that carrier.
    ups = ac.sole_account("UPS")
    cur.execute("SELECT COUNT(*) FROM shipments WHERE carrier='UPS' AND (acct IS NULL OR acct='')")
    ups_n = cur.fetchone()[0]
    total += ups_n
    print(f"    UPS    sole-account->{ups:<12} {ups_n:>7,} rows  (Kurt-declared 2026-08-20)")

    cur.execute("SELECT COUNT(*) FROM shipments WHERE carrier='FedEx' AND acct IS NULL "
                "AND source_file LIKE %s", ("%acct911%",))
    fn_n = cur.fetchone()[0]
    total += fn_n
    print(f"    FedEx  filename->206137911  {fn_n:>7,} rows")

    cur.execute("SELECT COUNT(*) FROM shipments WHERE carrier='FedEx' AND acct IS NULL")
    fedex_null = cur.fetchone()[0]
    print(f"    FedEx  xlsx-recovery + unknown for the remainder of {fedex_null:,} NULL rows")

    if not args.apply:
        print(f"\n  DRY-RUN — {total:,}+ rows would change. Re-run with --apply.")
        my.close()
        return

    print("\n  applying ...")
    for carrier, canonical, aliases in plan:
        marks = ",".join(["%s"] * len(aliases))
        cur.execute(f"UPDATE shipments SET acct=%s WHERE carrier=%s AND acct IN ({marks})",
                    [canonical, carrier] + aliases)
        print(f"    {carrier} alias->{canonical}: {cur.rowcount:,}")
    my.commit()

    cur.execute("UPDATE shipments SET acct=%s WHERE carrier='UPS' AND (acct IS NULL OR acct='')",
                (ups,))
    print(f"    UPS sole-account: {cur.rowcount:,}")
    my.commit()

    cur.execute("UPDATE shipments SET acct='206137911' WHERE carrier='FedEx' AND acct IS NULL "
                "AND source_file LIKE %s", ("%acct911%",))
    print(f"    FedEx filename: {cur.rowcount:,}")
    my.commit()

    # xlsx recovery via a temp map table — ONE multi-row insert then ONE UPDATE ... JOIN, instead
    # of 3,378 round trips.
    if tmap:
        cur.execute(f"DROP TEMPORARY TABLE IF EXISTS {MAP_TABLE}")
        cur.execute(f"CREATE TEMPORARY TABLE {MAP_TABLE} ("
                    f" tracking VARCHAR(128) NOT NULL PRIMARY KEY, acct VARCHAR(48) NOT NULL)"
                    f" ENGINE=InnoDB DEFAULT CHARSET=utf8mb4")
        items = list(tmap.items())
        for i in range(0, len(items), 1000):
            chunk = items[i:i + 1000]
            vals = ",".join(["(%s,%s)"] * len(chunk))
            cur.execute(f"INSERT IGNORE INTO {MAP_TABLE} (tracking, acct) VALUES {vals}",
                        [x for kv in chunk for x in kv])
        cur.execute(f"UPDATE shipments s JOIN {MAP_TABLE} m ON m.tracking = s.tracking "
                    f"SET s.acct = m.acct WHERE s.carrier='FedEx' AND s.acct IS NULL")
        print(f"    FedEx xlsx-recovery: {cur.rowcount:,}")
        my.commit()

    # 🔴 LAST, and only what is left: no evidence anywhere -> the literal `unknown`, so
    # "could not resolve" stays distinguishable from "nothing ever wrote here".
    cur.execute("UPDATE shipments SET acct=%s WHERE carrier IN ('FedEx','UPS') "
                "AND (acct IS NULL OR acct='')", (ac.UNKNOWN,))
    print(f"    residual -> unknown: {cur.rowcount:,}")
    my.commit()

    print("\n  final distribution:")
    cur.execute("SELECT carrier, acct, COUNT(*) FROM shipments WHERE carrier IN ('FedEx','UPS') "
                "GROUP BY carrier, acct ORDER BY 3 DESC")
    for r in cur.fetchall():
        print(f"    {r[0]:<6} {str(r[1]):<14} {r[2]:>7,}")
    my.close()


if __name__ == "__main__":
    main()
