"""Patch Onboarded Items DistVol xlsx Sheet1 with confirmed user values.

Updates per user confirmations 2026-04-29:
- All MT-* SKUs: CI = 7.37 (DV 0.07)
- AC-GLAW: CI = 12.636 (DV 0.12)
- Add missing: CH-BOSI=0.15, AC-TOK=0.12, AC-BLUCAR=0.12, MT-COPPA=0.07
"""
from __future__ import annotations
import shutil
from datetime import date
from pathlib import Path
from openpyxl import load_workbook

SRC = Path(r"C:\Users\Work\Desktop\Onboarded Items with DistVol - Updated.xlsx")
BAK = Path(rf"C:\Users\Work\Desktop\Onboarded Items with DistVol - Updated.{date.today().isoformat()}.bak.xlsx")
DV_RATE = 105.3

# CI = DV * rate
ci_for = lambda dv: round(dv * DV_RATE, 4)

UPDATES = {
    # SKU -> new CI
    "AC-GLAW": ci_for(0.12),
}
# Bulk-update all MT-* SKUs to CI=7.371 (DV 0.07)
MT_NEW_CI = round(0.07 * DV_RATE, 4)  # 7.371

ADDITIONS = [
    # (sku, ci, mfg_name_placeholder, category)
    ("CH-BOSI",   ci_for(0.15), "AHB: Robiola due Latte (Bosi)", "Cheese"),
    ("AC-TOK",    ci_for(0.12), "AHB: Tokyo (placeholder)",      "Accompaniments"),
    ("AC-BLUCAR", ci_for(0.12), "AHB: Blueberry Cardamom (placeholder)", "Accompaniments"),
    ("MT-COPPA",  MT_NEW_CI,    "AHB: Coppa Italiana",           "Meats"),
]

def main():
    if not BAK.exists():
        shutil.copy2(SRC, BAK)
        print(f"Backup -> {BAK.name}")

    wb = load_workbook(SRC)
    ws = wb["Sheet1"]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    sku_idx = header.index("SKU")
    ci_idx = header.index("Cubic Inches")
    dv_idx = header.index("DistVol")
    cat_idx = header.index("Category")

    existing_skus = set()
    mt_updated = 0
    explicit_updated = []

    for row in ws.iter_rows(min_row=2):
        sku_cell = row[sku_idx]
        sku = sku_cell.value
        if not sku:
            continue
        sku = str(sku).strip()
        existing_skus.add(sku)
        ci_cell = row[ci_idx]
        # MT bulk update
        if sku.startswith("MT-"):
            old = ci_cell.value
            if old != MT_NEW_CI:
                ci_cell.value = MT_NEW_CI
                mt_updated += 1
        # Explicit per-SKU
        if sku in UPDATES:
            old = ci_cell.value
            new = UPDATES[sku]
            if old != new:
                ci_cell.value = new
                explicit_updated.append((sku, old, new))
        # DistVol formula already references CI — leave intact

    # Append additions
    added = []
    next_row = ws.max_row + 1
    for sku, ci, mfg, cat in ADDITIONS:
        if sku in existing_skus:
            print(f"  SKIP add (exists): {sku}")
            continue
        # Build row matching header layout: Category, RMFG Inv No., SKU, MFG Name, Cubic Inches, DistVol, Exclude
        new_row = [None] * len(header)
        new_row[cat_idx] = cat
        new_row[sku_idx] = sku
        new_row[header.index("MFG Name")] = mfg
        new_row[ci_idx] = ci
        new_row[dv_idx] = f"=E{next_row}/{DV_RATE}"
        for col_offset, val in enumerate(new_row):
            ws.cell(row=next_row, column=col_offset + 1, value=val)
        added.append((sku, ci, round(ci/DV_RATE, 4)))
        next_row += 1

    wb.save(SRC)
    print(f"\nMT-* CI updated to {MT_NEW_CI}: {mt_updated} rows")
    print(f"Explicit updates ({len(explicit_updated)}):")
    for s, o, n in explicit_updated:
        print(f"  {s}: CI {o} -> {n}")
    print(f"Additions ({len(added)}):")
    for s, ci, dv in added:
        print(f"  {s}: CI={ci} DV={dv}")
    print(f"\nSaved: {SRC}")

if __name__ == "__main__":
    main()
