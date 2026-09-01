"""failed_tags_corpus.py — turn RMFG's rejection files into labelled regression ground truth.

🔴 WHY: RMFG hands us CLASSIFIED FAILURES FOR FREE and nothing has ever read them.
When a sheet is rejected, RMFG returns `AHB_Failed Tags_<M-D-YY>.xlsx` listing exactly the rows
they could not honor. That is a labelled test set produced by the authority itself — strictly
better than any fixture we could write, because it is THEIR verdict, not our model of it.

Validated 2026-08-09 against `AHB_Failed Tags_8-3-26.xlsx` (wk0803, 53 rows):

    REPRODUCED by the coverage authority : 53
    NOT reproduced                       :  0

53 of 53, zero misses, zero disagreements — every rejected row is an `!OnTrac Ground - <hub>!`
tag to a zip ABSENT from the OnTrac master. This is the empirical case for widening
`presend_check.py` to validate lane authorities (patch proposal P3).

🔴 SCALE, so nobody dismisses this as a 3-row curiosity: wk0803 shipped **53** bad lanes;
wk0810 shipped 3. A single bad token REJECTS THE WHOLE SHEET and holds the week's production
until a corrected sheet is resubmitted. The payoff is the week, not the row.

🔴 A MISSING FILE IS NOT EVIDENCE OF ACCEPTANCE. There is no `AHB_Failed Tags_8-10-26.xlsx` on
disk. That means only that no rejection file was found — NOT that wk0810 was accepted. It may
simply never have been downloaded. `--require` exists so a caller can assert presence explicitly
rather than inferring success from silence ([[feedback-join-zeroes-silently]]: absence of a row
is not a zero).

Usage:
    python scripts/failed_tags_corpus.py --scan                 # find rejection files
    python scripts/failed_tags_corpus.py --build                # -> corpus JSON fixture
    python scripts/failed_tags_corpus.py --validate             # replay authority vs RMFG verdict
    python scripts/failed_tags_corpus.py --require 8-10-26      # non-zero if that file is absent
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path

for _s in (sys.stdout, sys.stderr):
    if hasattr(_s, "reconfigure"):
        _s.reconfigure(encoding="utf-8", errors="replace")

import openpyxl  # noqa: E402

WS = Path(__file__).resolve().parents[2]              # Claude Projects
SHIPROUTING = WS / "ShipRouting"
CORPUS = WS / "_outputs" / "cache" / "failed_tags_corpus.json"
SEARCH_DIRS = [Path.home() / "Downloads", WS / "_outputs" / "artifacts", WS / "_outputs" / "cache"]

TAG_RE = re.compile(r"!(NO )?([A-Za-z]+)[^!]*? - ([A-Za-z ]+)_AHB!")
ZIP_RE = re.compile(r",\s*(\d{5})(?:-\d{4})?\s*$")
FNAME_RE = re.compile(r"AHB_Failed Tags_(\d{1,2}-\d{1,2}-\d{2})", re.I)


def find_files() -> list[Path]:
    out: list[Path] = []
    for d in SEARCH_DIRS:
        if d.exists():
            out.extend(sorted(d.glob("AHB_Failed Tags_*.xlsx")))
    # de-dupe Windows "(1)" copies by (cohort, size)
    seen, uniq = set(), []
    for p in out:
        m = FNAME_RE.search(p.name)
        key = (m.group(1) if m else p.name, p.stat().st_size)
        if key in seen:
            continue
        seen.add(key)
        uniq.append(p)
    return uniq


def parse(path: Path) -> list[dict]:
    """-> [{order, cohort, name, address, zip, tag, boxes, gel}] . Read-only."""
    m = FNAME_RE.search(path.name)
    cohort = m.group(1) if m else path.stem
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    hdr = [str(h) if h is not None else "" for h in next(it)]

    def idx(*names):
        for n in names:
            if n in hdr:
                return hdr.index(n)
        return None

    iO, iN = idx("OrderID"), idx("ReverseName", "Name")
    iA, iT = idx("DeliveryAddress", "Address"), idx("DeliveryTag", "Tags")
    iB, iG = idx("Boxes"), idx("GelPackTag")
    if iO is None or iT is None:
        wb.close()
        raise SystemExit(f"FAIL: {path.name} lacks OrderID/DeliveryTag (cols: {hdr})")

    rows = []
    for r in it:
        if r[iO] is None:
            continue
        addr = str(r[iA] or "") if iA is not None else ""
        zm = ZIP_RE.search(addr)
        rows.append({
            "order": str(r[iO]).lstrip("#").strip(),
            "cohort": cohort,
            "name": str(r[iN] or "") if iN is not None else "",
            "address": addr,
            "zip": zm.group(1) if zm else None,
            "tag": str(r[iT] or ""),
            "boxes": str(r[iB] or "") if iB is not None else "",
            "gel": str(r[iG] or "") if iG is not None else "",
            "source_file": path.name,
        })
    wb.close()
    return rows


def classify(rows: list[dict]) -> tuple[list[dict], list[dict]]:
    """Replay each rejected row through the coverage authority. -> (reproduced, missed)."""
    sys.path.insert(0, str(SHIPROUTING))
    from lib.features import CARRIER_HUBS
    from lib.hubs import HUB_CODE
    from lib.zip_loaders import load_ontrac

    ontrac = load_ontrac()
    reproduced, missed = [], []
    for row in rows:
        verdict = None
        for m in TAG_RE.finditer(row["tag"]):
            car, hub = m.group(2), m.group(3).strip()
            hubs = CARRIER_HUBS.get(car)
            if hubs is not None and hub not in hubs:
                verdict = f"ILLEGAL carrier x hub ({car} not at {hub})"
                break
            if car == "OnTrac" and row["zip"]:
                cell = ontrac.get(row["zip"])
                code = HUB_CODE.get(hub)
                hv = cell.get(code) if isinstance(cell, dict) and code else None
                if not (hv is not None and str(hv).strip() != ""):
                    verdict = ("zip absent from OnTrac master" if cell is None
                               else f"blank {code} cell (no {hub} lane)")
                    break
        rec = dict(row, verdict=verdict)
        (reproduced if verdict else missed).append(rec)
    return reproduced, missed


def main(argv=None) -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--scan", action="store_true")
    ap.add_argument("--build", action="store_true")
    ap.add_argument("--validate", action="store_true")
    ap.add_argument("--require", metavar="COHORT",
                    help="exit non-zero if no rejection file for this cohort (e.g. 8-10-26). "
                         "Absence is NOT evidence of acceptance — assert it deliberately.")
    a = ap.parse_args(argv)
    if not any([a.scan, a.build, a.validate, a.require]):
        a.validate = True

    files = find_files()
    if a.scan or not files:
        print(f"rejection files found: {len(files)}")
        for p in files:
            print(f"   {p}  ({p.stat().st_size} bytes)")
        if not files:
            print("\n🔴 none found. This is NOT evidence that any sheet was accepted — only that "
                  "no rejection file is on disk. Check email/Downloads before concluding anything.")
            return 0 if a.scan else 1

    if a.require:
        hit = [p for p in files if a.require.lower() in p.name.lower()]
        if not hit:
            print(f"🔴 NO rejection file for cohort {a.require}. This does NOT mean the sheet was "
                  f"accepted — it means no file was found. Verify before claiming success.")
            return 1
        print(f"rejection file present for {a.require}: {hit[0].name}")
        return 0

    rows: list[dict] = []
    for p in files:
        rows.extend(parse(p))
    print(f"parsed {len(rows)} rejected rows from {len(files)} file(s)")

    if a.build:
        CORPUS.parent.mkdir(parents=True, exist_ok=True)
        CORPUS.write_text(json.dumps(rows, indent=1), encoding="utf-8")
        print(f"corpus written: {CORPUS}")

    if a.validate:
        reproduced, missed = classify(rows)
        print(f"\nREPRODUCED by the coverage authority : {len(reproduced)}")
        print(f"NOT reproduced (authority says legal) : {len(missed)}")
        for r in missed[:20]:
            print(f"   🔴 MISS #{r['order']} {r['zip']} {r['tag']}")
        if missed:
            print("\n🔴 A MISS is a REAL GAP: RMFG rejected a row our authority calls legal. "
                  "Investigate before the next send — this is the check failing, not RMFG.")
            return 1
        print("\nCLEAN — the authority reproduces RMFG's verdict on every rejected row.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
