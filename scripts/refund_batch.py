# /// script
# requires-python = ">=3.10"
# dependencies = ["openpyxl", "requests", "urllib3"]
# ///

"""Canonical batch partial-refund CLI (formalizes _template_bulk_refund.py one-shots).

Constraints SSOT: scripts/REFUND_BATCH_RULES.md — READ FIRST.

Usage:
    refund_batch.py --orders file.xlsx --note "REASON — refund" [--amount 15] [--commit]
    refund_batch.py --ship-tag _SHIP_2026-08-10 --sku CH-FONT --note "..." [--commit]

Dry-run by default; --commit moves money. Idempotent: skips orders whose existing
refunds match the note keyword. After any --commit run:
    python InventoryReorder/Errors/detect_double_refunds_v2.py
"""

from __future__ import annotations

import argparse
import csv
import datetime as _dt
import json
import io
import sys
import time
from pathlib import Path

# 🔴 Windows cp1252 kills any non-ASCII print (arrows, emoji) MID-RUN — for a --apply tool that
# means a crash BETWEEN mutations, leaving the batch half-applied. Wrap stdout before anything
# prints, including argparse --help. (Live 2026-08-09: shorts_pass.py --help died on U+2192.)
if hasattr(sys.stdout, "buffer"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")


# Windows cp1252 stdout crashes on emoji/em-dash — force UTF-8 (workspace standard).
# Live 2026-08-09: an em-dash in a skip-summary line killed the run AFTER selection.
for _s in (sys.stdout, sys.stderr):
    if hasattr(_s, "reconfigure"):
        _s.reconfigure(encoding="utf-8", errors="replace")

SCRIPT_DIR = Path(__file__).resolve().parent
APPYHOUR = SCRIPT_DIR.parent
sys.path.insert(0, str(APPYHOUR / "AppyHourMCP"))

SETTINGS = APPYHOUR / "InventoryReorder" / "dist" / "inventory_reorder_settings.json"
OUT_ROOT = APPYHOUR.parent / "_outputs"

ORDER_FIELDS = "id,name,tags,line_items,refunds,cancelled_at"


# ── amount math (pure — unit-testable offline) ──────────────────────────────
def actual_paid_for_sku(order: dict, sku: str) -> float | None:
    """Actual-paid share for `sku` on an order: discounted line price + tax share.

    NEVER list price. Uses active_line_items() so removed/refunded lines don't count.
    Returns None if the SKU has no active line on the order (caller flags MISSING).
    """
    from utils import active_line_items  # AppyHourMCP/utils.py — canonical

    # active_line_items() nets removed/refunded qty INTO `quantity`. Financial
    # fields (discount_allocations, tax_lines) are for the ORIGINAL line, so
    # pull originals from order["line_items"] and pro-rate by net/orig qty.
    net_by_id = {li["id"]: int(li.get("quantity") or 0) for li in active_line_items(order)}
    total = 0.0
    found = False
    for li in order.get("line_items", []):
        if (li.get("sku") or "").strip() != sku or li["id"] not in net_by_id:
            continue
        found = True
        orig_qty = int(li.get("quantity") or 0)
        net_qty = net_by_id[li["id"]]
        gross = float(li.get("price") or 0.0) * orig_qty
        disc = sum(float(d.get("amount") or 0.0) for d in li.get("discount_allocations") or [])
        tax = sum(float(t.get("price") or 0.0) for t in li.get("tax_lines") or [])
        line_paid = gross - disc + tax
        if orig_qty and net_qty != orig_qty:
            line_paid *= net_qty / orig_qty
        total += line_paid
    return round(total, 2) if found else None


def near_miss_candidates(sku: str, present: set[str]) -> list[str]:
    """Prefix-overlap candidates for a SKU that matched nothing (PR-CJAM -> PR-CJAM-GEN)."""
    return sorted(p for p in present if p.startswith(sku) or sku.startswith(p))[:5]


def absent_sku_refusal(sku: str, present: set[str], cohort_n: int, tag: str) -> str | None:
    """🔴 Join-zero guard (2026-08-08 live smoke test: `--has PR-CJAM` matched 0 of 2321 orders
    because the real SKU is PR-CJAM-GEN). Matching is EXACT, so a near-miss SKU silently reads
    as "nobody is owed a refund" instead of "your SKU is wrong" — and a refund campaign that
    quietly refunds nobody is indistinguishable from one that ran. Returns a refusal message,
    or None when the SKU is present somewhere on the cohort."""
    if sku in present:
        return None
    cands = near_miss_candidates(sku, present)
    return (f"🔴 REFUSED: SKU {sku} appears on ZERO of {cohort_n} orders on {tag} — "
            f"that is a wrong/typo'd SKU, not an empty refund set.\n"
            f"  did you mean: {cands or '(no similar SKU in cohort)'}\n"
            f"  Re-run with --allow-absent-sku only if you truly expect it to be absent.")


def note_keyword(note: str) -> str:
    """Idempotency keyword — template's mechanism (first segment before em-dash)."""
    return note.split("—")[0].strip() if "—" in note else note[:20]


def has_existing_refund(refunds: list[dict], note: str) -> bool:
    kw = note_keyword(note)
    return any(kw in (r.get("note") or "") for r in refunds or [])


# ── I/O helpers ─────────────────────────────────────────────────────────────
def load_order_numbers(path: Path, skip: set[str]) -> list[str]:
    nums: list[str] = []
    if path.suffix.lower() == ".csv":
        rows = list(csv.reader(path.open(encoding="utf-8-sig")))
    else:
        import openpyxl

        ws = openpyxl.load_workbook(path).active
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
    for row in rows[1:]:
        raw = str(row[0] or "").strip().lstrip("#")
        if not raw or not raw.isdigit():
            continue
        if raw in skip:
            print(f"  SKIP #{raw} (exclusion list)")
            continue
        nums.append(raw)
    return nums


def _session():
    import requests
    from requests.adapters import HTTPAdapter
    from urllib3.util.retry import Retry

    s = requests.Session()
    s.mount("https://", HTTPAdapter(max_retries=Retry(total=3, backoff_factor=2, status_forcelist=[429, 500, 502, 503, 504])))
    return s


def _api():
    settings = json.loads(SETTINGS.read_text(encoding="utf-8"))
    base = f"https://{settings['shopify_store_url']}.myshopify.com/admin/api/2024-01"
    headers = {"X-Shopify-Access-Token": settings["shopify_access_token"], "Content-Type": "application/json"}
    return _session(), base, headers


def fetch_by_numbers(sess, base, headers, numbers: list[str]) -> tuple[list[dict], list[str]]:
    found, missing = [], []
    for i, num in enumerate(numbers):
        r = sess.get(f"{base}/orders.json", headers=headers,
                     params={"name": num, "status": "any", "limit": 5, "fields": ORDER_FIELDS}, timeout=30)
        r.raise_for_status()
        o = next((o for o in r.json().get("orders", []) if str(o.get("name", "")).lstrip("#") == num), None)
        (found.append(o) if o else missing.append(num))
        print(f"  [{i + 1}/{len(numbers)}] #{num} — {'found' if o else 'NOT FOUND'}")
        time.sleep(0.5)
    return found, missing


def select_cohort(all_orders: list[dict], tag: str, sku: str) -> tuple[list[dict], set[str], int]:
    """Pure selection over a fetched page-set — offline-testable, identical semantics to tag_where.

    Non-cancelled orders carrying `tag` whose active_line_items() include `sku`.
    Also returns (union of live SKUs on the cohort, cohort order count) for the join-zero guard.
    """
    from utils import active_line_items

    orders: list[dict] = []
    present: set[str] = set()
    cohort_n = 0
    for o in all_orders:
        tags = {t.strip() for t in (o.get("tags") or "").split(",")}
        if tag not in tags or o.get("cancelled_at"):
            continue
        cohort_n += 1
        live = [(li.get("sku") or "").strip() for li in active_line_items(o)]
        present |= {s for s in live if s}
        if sku in live:
            orders.append(o)
    return orders, present, cohort_n


def fetch_by_tag_sku(sess, base, headers, tag: str, sku: str) -> tuple[list[dict], set[str], int]:
    """Non-cancelled orders on tag that HAVE sku per active_line_items().

    🔴 Uses the SERVER-SIDE `tag` filter and the canonical `shopify_paginate` (utils.py) — the
    hand-rolled walk this replaced sent no `tag` param, so it paged the ENTIRE order history
    (~680 pages / 25+ min) and only agreed with tag_where.py if it was allowed to run to
    completion. Any interruption silently under-selected → under-refunded. It also split the
    Link header on "," instead of the canonical regex, which only survives because Shopify
    URL-encodes the `fields` commas as %2C.
    """
    from utils import shopify_paginate

    params = {"status": "any", "limit": 250, "tag": tag,
              # refunds REQUIRED or active_line_items silently no-ops (same gotcha as tag_where)
              "fields": ORDER_FIELDS}
    # resource="orders-live": money-adjacent read, never served from the 10m orders cache.
    all_orders = shopify_paginate(f"{base}/orders.json", headers, params=params,
                                  key="orders", resource="orders-live")
    return select_cohort(all_orders, tag, sku)


def get_transaction(sess, base, headers, order_id):
    r = sess.get(f"{base}/orders/{order_id}/transactions.json", headers=headers, timeout=30)
    r.raise_for_status()
    for txn in r.json().get("transactions", []):
        if txn.get("kind") in ("sale", "capture") and txn.get("status") == "success":
            return txn["id"], txn.get("gateway", "")
    return None, None


def issue_refund(sess, base, headers, order_id, txn_id, gateway, amount, note) -> int | None:
    payload = {"refund": {"notify": False, "note": note, "shipping": {"amount": 0},
               "transactions": [{"parent_id": txn_id, "amount": str(amount), "kind": "refund", "gateway": gateway}]}}
    r = sess.post(f"{base}/orders/{order_id}/refunds.json", headers=headers, json=payload, timeout=30)
    if r.status_code in (200, 201):
        return r.json().get("refund", {}).get("id")
    print(f"    FAILED HTTP {r.status_code} — {r.text[:300]}")
    return None


# ── main ────────────────────────────────────────────────────────────────────
def main() -> int:
    ap = argparse.ArgumentParser(description="Batch partial refund (dry-run default). SSOT: REFUND_BATCH_RULES.md")
    src = ap.add_mutually_exclusive_group(required=True)
    src.add_argument("--orders", type=Path, help="xlsx/csv, order numbers in col A")
    src.add_argument("--ship-tag", help="cohort tag (requires --sku)")
    ap.add_argument("--sku", help="SKU whose actual-paid share is refunded")
    ap.add_argument("--note", required=True, help="refund note — idempotency key, unique per campaign")
    ap.add_argument("--amount", type=float, help="fixed amount override (warns if > actual paid)")
    ap.add_argument("--amount-mode", default="actual-paid", choices=["actual-paid"])
    ap.add_argument("--skip", default="", help="comma-separated order numbers to exclude")
    ap.add_argument("--commit", action="store_true", help="MOVE MONEY (default: dry-run)")
    ap.add_argument("--allow-absent-sku", action="store_true",
                    help="permit a --sku that appears on ZERO cohort orders (default: refuse)")
    args = ap.parse_args()

    if args.ship_tag and not args.sku:
        ap.error("--ship-tag requires --sku")
    if not args.amount and not args.sku:
        ap.error("actual-paid mode needs --sku (or give a fixed --amount)")

    sess, base, headers = _api()
    skip = {s.strip().lstrip("#") for s in args.skip.split(",") if s.strip()}

    if args.orders:
        nums = load_order_numbers(args.orders, skip)
        print(f"{len(nums)} order numbers loaded; looking up...")
        orders, missing = fetch_by_numbers(sess, base, headers, nums)
        if missing:
            print(f"  NOT FOUND: {', '.join('#' + n for n in missing)}")
    else:
        print(f"Fetching cohort {args.ship_tag} with {args.sku} (active_line_items)...")
        hits, present, cohort_n = fetch_by_tag_sku(sess, base, headers, args.ship_tag, args.sku)
        # 🔴 Join-zero guard — a wrong SKU must never read as "no refunds owed" (rule 12).
        if not args.allow_absent_sku:
            refusal = absent_sku_refusal(args.sku, present, cohort_n, args.ship_tag)
            if refusal:
                print(refusal)
                return 2
        orders = [o for o in hits if str(o.get("name", "")).lstrip("#") not in skip]

    # plan amounts
    plan, warns = [], 0
    skipped_missing, skipped_zero = [], []
    for o in orders:
        name = str(o.get("name", "")).lstrip("#")
        paid = actual_paid_for_sku(o, args.sku) if args.sku else None
        if args.amount is not None:
            amt = args.amount
            if paid is not None and amt > paid:
                print(f"  WARN #{name}: fixed ${amt:.2f} > actual paid ${paid:.2f}")
                warns += 1
        elif paid is None:
            skipped_missing.append(name)
            continue
        elif paid <= 0:
            skipped_zero.append(name)
            continue
        else:
            amt = paid
        already = has_existing_refund(o.get("refunds"), args.note)
        plan.append((name, o["id"], amt, paid, already))

    total = sum(a for _, _, a, _, alr in plan if not alr)
    n_already = sum(1 for p in plan if p[4])
    mode = "COMMIT" if args.commit else "DRY-RUN"

    # 🔴 Selection reconciliation (2026-08-09): every order that entered MUST be accounted for.
    # Without this, 644 matched orders all skipped as $0 printed 644 look-alike lines and a
    # "0 orders" footer — reading as "the filter found nothing" (a matching bug) instead of
    # "this SKU is a $0 in-box component, nobody paid for it" (the truth). Rules gotcha 10.
    print(f"\nSELECTION: {len(orders)} orders carry {args.sku} -> planned {len(plan)} | "
          f"skipped {len(skipped_zero)} zero-paid, {len(skipped_missing)} SKU-line-missing")
    for label, names in (("zero-paid ($0 line — nobody paid, never refund)", skipped_zero),
                         ("SKU-line-missing (never inferred)", skipped_missing)):
        if names:
            head = ", ".join("#" + n for n in names[:10])
            print(f"  SKIPPED {len(names)} {label}: {head}{' …' if len(names) > 10 else ''}")
    if orders and not plan:
        print(f"  ⚠️  ALL {len(orders)} matching orders were skipped — the cohort+SKU filter WORKED; "
              f"the amount math found nothing owed. This is NOT an empty match.")

    print(f"\n[{mode}] {len(plan)} orders | {n_already} already refunded (note match) | "
          f"to move: ${total:.2f} | note: {args.note}")
    for name, _, amt, paid, alr in plan:
        print(f"  #{name}  ${amt:.2f}" + (f"  (paid ${paid:.2f})" if paid is not None else "") + ("  SKIP-EXISTING" if alr else ""))

    stamp = _dt.datetime.utcnow().strftime("%Y%m%dT%H%M%SZ")
    log_path = OUT_ROOT / "logs" / f"refund_batch_{stamp}.log"
    log_path.parent.mkdir(parents=True, exist_ok=True)
    lines = [f"refund_batch [{mode}] note={args.note} orders={len(plan)} total=${total:.2f} "
             f"matched={len(orders)} skipped_zero={len(skipped_zero)} skipped_missing={len(skipped_missing)}"]
    lines += [f"  #{n} ${a:.2f} paid={p} already={alr}" for n, _, a, p, alr in plan]

    moved = []
    if args.commit:
        for name, oid, amt, paid, alr in plan:
            if alr:
                print(f"    SKIP #{name}: already has matching refund")
                continue
            txn_id, gw = get_transaction(sess, base, headers, oid)
            if not txn_id:
                print(f"    FAILED #{name}: no payment transaction")
                lines.append(f"  FAILED #{name}: no transaction")
                continue
            time.sleep(0.3)
            rid = issue_refund(sess, base, headers, oid, txn_id, gw, amt, args.note)
            if rid:
                print(f"    OK #{name}: refunded ${amt:.2f} (refund {rid})")
                moved.append((name, amt, rid))
                lines.append(f"  OK #{name} ${amt:.2f} refund_id={rid}")
            else:
                lines.append(f"  FAILED #{name}")
            time.sleep(1.0)
        if moved:
            import openpyxl

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["order_number", "amount", "refund_id", "note"])
            for n, a, rid in moved:
                ws.append([n, a, rid, args.note])
            xlsx_path = OUT_ROOT / "reports" / f"refund_batch_{stamp}_moved.xlsx"
            xlsx_path.parent.mkdir(parents=True, exist_ok=True)
            wb.save(xlsx_path)
            print(f"\nMoved-money xlsx: {xlsx_path}")
        print(f"\nDone: {len(moved)} refunded, ${sum(a for _, a, _ in moved):.2f} total.")
        print("NEXT: python InventoryReorder/Errors/detect_double_refunds_v2.py  (REFUND_BATCH_RULES.md #6)")
    else:
        print("\nDRY-RUN complete. Re-run with --commit to move money (restate+go per live-writes rule).")

    log_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(f"Log: {log_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main() or 0)
