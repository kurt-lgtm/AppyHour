"""Regression tests for generate_matrix_xlsx product-column selection.

Guards MATRIX_RULES rule 19: every pickable prefix that can appear on an order
must become a product column. wk0720 (RMFG_20260717) dropped MR-JRNL
("Cheese Journal") from the generated sheet for 38 orders because the
`food_pkg_prefixes` filter omitted MR-.
"""

from __future__ import annotations

from pathlib import Path
from unittest.mock import patch

import openpyxl

import matrix_commander as mc


def _order(name: str, skus_qty: dict[str, int], tags: list[str] | None = None) -> dict:
    """Minimal Shopify order dict in the shape _fetch_orders_graphql returns."""
    edges = [
        {"node": {"sku": sku, "fulfillableQuantity": qty, "quantity": qty}}
        for sku, qty in skus_qty.items()
    ]
    return {
        "name": name,
        "tags": tags or [],
        "lineItems": {"edges": edges},
        "email": "cust@example.com",
        "phone": "5551234567",
        "note": "",
        "shippingAddress": {
            "firstName": "Test",
            "lastName": "Customer",
            "address1": "1 Main St",
            "address2": "",
            "city": "Indianapolis",
            "provinceCode": "IN",
            "zip": "46201",
            "phone": "5551234567",
        },
    }


_TRANSLATIONS = {
    "CH-BLR": "AHB (S_REG): Brie Locale Rouge",
    "MR-JRNL": "AHB (S_REG): Cheese Journal",
}


def _generate(tmp_path, orders):
    with (
        patch.object(mc, "_get_shopify_auth", return_value=("https://shop", {})),
        patch.object(mc, "_fetch_orders_graphql", return_value=orders),
        patch.object(mc, "load_mfg_translations", return_value=_TRANSLATIONS),
    ):
        out = mc.generate_matrix_xlsx(
            "RMFG_20260717", ship_date="2026-07-20", output_dir=str(tmp_path)
        )
    assert out, "generate_matrix_xlsx returned no path"
    ws = openpyxl.load_workbook(out, data_only=True).active
    headers = [str(c.value or "") for c in ws[1]]
    rows = {str(ws.cell(r, 2).value): r for r in range(2, ws.max_row + 1)}
    return ws, headers, rows


def test_unonboarded_sku_rejects_before_any_file_is_written(tmp_path):
    """Rule 19a: an un-onboarded SKU is an IMMEDIATE REJECT, never a fabricated column.

    wk0728 TUE: AC-QUIC ("Quicos") was live in Shopify but absent from mfg_translations.csv. The
    column builder emitted `AHB (S_REG): {SKU_TO_NAME.get(sku, sku)}` which, with no local mapping,
    resolved to the BARE SKU wrapped in the real header format — a phantom column that looks
    legitimate but names a product RMFG has never seen. It only printed a yellow warning.
    """
    import pytest as _pytest
    orders = [_order("2001", {"CH-BLR": 2, "AC-QUIC": 1})]
    with (
        patch.object(mc, "_get_shopify_auth", return_value=("https://shop", {})),
        patch.object(mc, "_fetch_orders_graphql", return_value=orders),
        patch.object(mc, "load_mfg_translations", return_value=_TRANSLATIONS),
        _pytest.raises(ValueError, match="MFG onboarding REJECT"),
    ):
        mc.generate_matrix_xlsx("RMFG_20260717", ship_date="2026-07-20", output_dir=str(tmp_path))

    # 🔴 NO file may exist — the old gate fired only after the xlsx was written, leaving a
    # legit-looking sheet in Downloads that could be (and was) sent.
    assert list(tmp_path.glob("*.xlsx")) == []


def test_no_header_may_ever_name_a_bare_sku(tmp_path):
    """The fabrication SHAPE itself is banned: `AHB (S_REG): <BARE-SKU>` must never be emitted.

    Guards the case where SKU_TO_NAME *does* have an entry — the header would then look even more
    plausible while still not being the RMFG translator's name. Only mfg_translations.csv may name
    a column.
    """
    import re
    orders = [_order("2001", {"CH-BLR": 2, "MR-JRNL": 1})]
    _, headers, _ = _generate(tmp_path, orders)
    bare = [h for h in headers if re.fullmatch(r"AHB \(S_REG\): (CH|MT|AC|PK|TR|MR)-[A-Z0-9-]+", h)]
    assert not bare, f"headers naming a bare SKU (fabrication shape): {bare}"
    for sku, name in _TRANSLATIONS.items():
        assert name in headers, f"{sku} must be named by mfg_translations, got {headers}"


def test_fill_reject_carries_typed_skus_not_just_a_message(tmp_path):
    """The fill-time reject is identified BY TYPE, and `.skus` carries the offenders.

    `test_unonboarded_sku_rejects_before_any_file_is_written` matches on the message string — the
    exact anti-pattern the MfgOnboardingError docstring and rule 19a forbid callers to use (a reword
    silently stops the §13.5 console from degrading only the SHEET stages and it goes back to failing
    whole jobs). This locks the contract routing depends on: catch the TYPE, read `.skus`.
    """
    import pytest as _pytest
    orders = [_order("2001", {"CH-BLR": 2, "AC-QUIC": 1, "MT-NOPE": 1})]
    with (
        patch.object(mc, "_get_shopify_auth", return_value=("https://shop", {})),
        patch.object(mc, "_fetch_orders_graphql", return_value=orders),
        patch.object(mc, "load_mfg_translations", return_value=_TRANSLATIONS),
        _pytest.raises(mc.MfgOnboardingError) as e,
    ):
        mc.generate_matrix_xlsx("RMFG_20260717", ship_date="2026-07-20", output_dir=str(tmp_path))
    # every un-onboarded pickable SKU is named, without the caller parsing the message
    assert set(e.value.skus) == {"AC-QUIC", "MT-NOPE"}, e.value.skus
    assert isinstance(e.value, ValueError)   # existing `except ValueError` handlers keep catching it
    assert list(tmp_path.glob("*.xlsx")) == []


def test_parent_line_never_false_rejects_nor_becomes_a_column(tmp_path):
    """A PR-CJAM / CEX-EC PARENT line item is not a pickable child — it must neither trip the
    onboarding reject nor render as a phantom column.

    Guards the wrong-source class (Kurt 2026-07-28 "you filled both in from CEX-EC not PR-CJAM"):
    parents live in SKIP_PREFIXES, so they are excluded from the pickable-column set and cannot be
    mistaken for a source that names a column. Only the mapped CH-/MT-/AC- children RMFG actually
    picks become columns. An order carrying both parents plus a mapped child generates clean.
    """
    assert "PR-CJAM" in mc.SKIP_PREFIXES and "CEX-E" in mc.SKIP_PREFIXES
    orders = [_order("2001", {"CH-BLR": 2, "PR-CJAM-MONG": 1, "CEX-EC-MONG": 1})]
    ws, headers, _ = _generate(tmp_path, orders)
    # the one mapped child is a column...
    assert "AHB (S_REG): Brie Locale Rouge" in headers
    # ...and neither parent leaked in under ANY header shape
    leaked = [h for h in headers if "PR-CJAM" in h or "CEX-EC" in h or "CEX-E" in h]
    assert not leaked, f"parent SKU leaked as a column: {leaked}"


def test_mfg_gate_and_column_builder_share_one_prefix_set():
    """Rule 19: the gate's prefix tuple and the column builder's MUST be the same object.

    They were separate literals and drifted: the gate omitted TR-/MR- while the column builder
    included them, so a tray or journal SKU with no MFG name became a column the gate never looked
    at — the AC-QUIC blind spot, one prefix over.
    """
    assert mc.PICKABLE_PREFIXES == ("CH-", "MT-", "AC-", "PK-", "TR-", "MR-")
    src = Path(mc.__file__).read_text(encoding="utf-8")
    # no stray copy of the old 4-prefix literal may remain
    assert '("CH-", "MT-", "AC-", "PK-")' not in src


def test_mr_jrnl_gets_column(tmp_path):
    """An order carrying MR-JRNL produces a 'Cheese Journal' column with the qty."""
    orders = [
        _order("2001", {"CH-BLR": 2, "MR-JRNL": 1}),
        _order("2002", {"CH-BLR": 1}),
    ]
    ws, headers, _ = _generate(tmp_path, orders)

    jrnl_hdr = "AHB (S_REG): Cheese Journal"
    assert jrnl_hdr in headers, f"MR-JRNL column missing from headers: {headers}"

    # both orders share a customer name, so locate rows by OrderID (col A)
    order_rows = {str(ws.cell(r, 1).value): r for r in range(2, ws.max_row + 1)}

    col = headers.index(jrnl_hdr) + 1  # openpyxl is 1-indexed
    assert ws.cell(order_rows["2001"], col).value == 1
    # order without a journal leaves the cell blank, not 0
    assert ws.cell(order_rows["2002"], col).value in (None, "")

    # col-D Total (rule 0) counts the journal qty too: 2 CH + 1 journal = 3
    total_col = headers.index("Total") + 1
    assert ws.cell(order_rows["2001"], total_col).value == 3


def test_notes_column_never_carries_the_shopify_note(tmp_path):
    """MATRIX_RULES rule 18 (Kurt 2026-07-10, restated 2026-08-18 "note shouldn't be pulled from
    shopify"): the Notes column ships EMPTY.

    The rule was documented and never implemented — `generate_matrix_xlsx` copied the Shopify order
    note verbatim into col 13, so internal workflow text reached RMFG's sheet. The 2026-08-18 cloud
    build carried "Order was held by Shopify Flow. Review customer orders: #162671" on a live row.
    The COLUMN must stay (RMFG's layout is fixed) — only the value is blank.
    """
    o = _order("2001", {"CH-BLR": 2})
    o["note"] = "Order was held by Shopify Flow. Review customer orders:\n\n  #162671\n"
    ws, headers, rows = _generate(tmp_path, [o])
    assert "Notes" in headers, "the Notes COLUMN must survive — RMFG's layout is fixed"
    col = headers.index("Notes") + 1
    for r in range(2, ws.max_row + 1):
        assert ws.cell(r, col).value in (None, ""), (
            f"row {r} Notes carries {ws.cell(r, col).value!r} — rule 18 says EMPTY")
