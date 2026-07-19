"""Regression tests for the wk0720 walnut-header bugs (MATRIX_RULES rule 15b).

1. merge_gift_xlsx: main matrix emits "Walnut, Honey & ..." (comma, Shopify title)
   while the gift sheet uses "Walnut Honey & ..." (no comma, RMFG translator form).
   A header-string column union treated them as two columns -> AC-FCWALN demand
   split and undercounted. Merge must normalize both sides to the no-comma form
   and keep ONE column carrying every order's walnut quantity.

2. constants.NAME_TO_SKU must map BOTH forms to AC-FCWALN so the rule-15b-correct
   submitted sheet no longer false-fails check_sku_mappings.
"""

from __future__ import annotations

import openpyxl

import matrix_commander as mc
from matrix_commander import NAME_TO_SKU

WALNUT_COMMA = "AHB (S_REG): Walnut, Honey & Extra Virgin Olive Oil Crackers"
WALNUT_15B = "AHB (S_REG): Walnut Honey & Extra Virgin Olive Oil Crackers"


def _write_xlsx(path, headers, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Access_LIVE"
    ws.append(headers)
    for r in rows:
        ws.append(r)
    wb.save(str(path))
    return str(path)


def _load_rows(path):
    ws = openpyxl.load_workbook(path, data_only=True)["Access_LIVE"]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    return rows[0], rows[1:]


def test_merge_keeps_single_walnut_column_with_full_demand(tmp_path):
    main = _write_xlsx(
        tmp_path / "main.xlsx",
        ["OrderID", "Name", WALNUT_COMMA],
        [[1001, "A", 2], [1002, "B", 1]],
    )
    gift = _write_xlsx(
        tmp_path / "gift.xlsx",
        ["OrderID", "Name", WALNUT_15B],
        [[2001, "G", 3]],
    )

    merged = mc.merge_gift_xlsx(main, gift)
    headers, rows = _load_rows(merged)

    walnut_cols = [h for h in headers if h and "Walnut" in str(h)]
    assert walnut_cols == [WALNUT_15B], f"expected one no-comma walnut column, got {walnut_cols}"

    wi = headers.index(WALNUT_15B)
    total = sum(int(r[wi] or 0) for r in rows)
    assert total == 6, f"walnut demand split/undercounted: {total} != 6"
    assert len(rows) == 3


def test_merge_aligns_gift_rows_by_header_not_position(tmp_path):
    # Gift sheet has a different column order + a gift-only product column
    main = _write_xlsx(
        tmp_path / "main.xlsx",
        ["OrderID", "Name", WALNUT_COMMA],
        [[1001, "A", 2]],
    )
    gift = _write_xlsx(
        tmp_path / "gift.xlsx",
        ["OrderID", WALNUT_15B, "Name", "AHB (S_REG): Glazed Walnuts"],
        [[2001, 5, "G", 1]],
    )

    merged = mc.merge_gift_xlsx(main, gift)
    headers, rows = _load_rows(merged)

    assert headers.index(WALNUT_15B) == 2  # normalized main column, unioned
    assert "AHB (S_REG): Glazed Walnuts" in headers  # gift-only column appended
    gift_row = next(r for r in rows if r[0] == 2001)
    assert gift_row[headers.index(WALNUT_15B)] == 5
    assert gift_row[headers.index("Name")] == "G"
    assert gift_row[headers.index("AHB (S_REG): Glazed Walnuts")] == 1


def test_merge_still_dedupes_order_ids(tmp_path):
    main = _write_xlsx(tmp_path / "main.xlsx", ["OrderID", WALNUT_COMMA], [[1001, 2]])
    gift = _write_xlsx(tmp_path / "gift.xlsx", ["OrderID", WALNUT_15B], [[1001, 9], [2001, 3]])

    merged = mc.merge_gift_xlsx(main, gift)
    _, rows = _load_rows(merged)
    assert [r[0] for r in rows] == [1001, 2001]


def test_name_to_sku_maps_both_walnut_forms():
    assert NAME_TO_SKU["Walnut Honey & Extra Virgin Olive Oil Crackers"] == "AC-FCWALN"
    assert NAME_TO_SKU["Walnut, Honey & Extra Virgin Olive Oil Crackers"] == "AC-FCWALN"


def test_resolver_resolves_rule15b_walnut_name():
    resolver = {mc._normalize_name(n): s for n, s in NAME_TO_SKU.items()}
    assert resolver[mc._normalize_name("Walnut Honey & Extra Virgin Olive Oil Crackers")] == "AC-FCWALN"


def test_check_sku_mappings_passes_on_rule15b_sheet():
    # With both forms in NAME_TO_SKU there is nothing unmapped -> PASS
    result = mc.check_sku_mappings({})
    assert result.passed
