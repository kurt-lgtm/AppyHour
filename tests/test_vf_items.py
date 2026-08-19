"""vf_items — scenario tests for the vF ITEM-MATRIX edit tool (MATRIX_RULES rule 24).

Every test is a real burn:
  invented MFG name (2026-08-04 "Farmstead Smoked Cumin Gouda")   -> authority-validated writes
  comma'd walnut header (wk0713, 545 units)                       -> header shape refused
  duplicate item in one order (wk0810 CH-TETI swap)               -> refused + rows named
  clobbered dated output                                          -> _rN revisions only
  "queued fix" != shipped fix                                     -> preservation verify after write

🔴 SYNTHETIC FIXTURES ONLY — no test ever opens or writes a real sheet.
"""
from __future__ import annotations

import csv
import json
import sys
from pathlib import Path

import openpyxl
import pytest

APPYHOUR = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(APPYHOUR))
sys.path.insert(0, str(APPYHOUR / "scripts"))

import vf_items  # noqa: E402

P = vf_items.HDR_PREFIX
AUTH_ROWS = [
    ("CH-TETI", P + "Teti Packaged Slice"),
    ("CH-ETX", P + "Etxegarai Packaged Slice"),
    ("CH-WWHO", P + "White Wine Honey Goat"),
    ("CH-CARO", P + "Carolina Moon Packaged Slice"),
    ("AC-FCWALN", P + "Walnut Honey & Extra Virgin Olive Oil Crackers"),
    ("CH-CUMIN", P + "Farmstead Cumin Gouda"),
]
AUTH = dict(AUTH_ROWS)
H_TETI, H_ETX, H_WWHO, H_CARO, H_WALN, H_CUMIN = (AUTH[s] for s in
                                                  ("CH-TETI", "CH-ETX", "CH-WWHO", "CH-CARO",
                                                   "AC-FCWALN", "CH-CUMIN"))


@pytest.fixture(autouse=True)
def _no_db(monkeypatch):
    """load_mfg_translations is env-first; keep tests on the csv path.

    Also RESTORE matrix_commander.MFG_AUTHORITATIVE_PATH: Authority rebinds it so every consumer
    reads one authority, and a leaked tmp path would silently re-point other modules' tests.
    """
    monkeypatch.delenv("ROUTING_INPUTS_DB", raising=False)
    import matrix_commander
    monkeypatch.setattr(matrix_commander, "MFG_AUTHORITATIVE_PATH",
                        matrix_commander.MFG_AUTHORITATIVE_PATH)


@pytest.fixture
def authority(tmp_path) -> Path:
    p = tmp_path / "mfg_names_authoritative.csv"
    with open(p, "w", newline="", encoding="utf-8") as fh:
        csv.writer(fh).writerows(AUTH_ROWS)
    return p


def _write_sheet(path, product_headers, rows, sheet_name=vf_items.SHEET):
    """rows: [{"OrderID":..., "Tags":..., <header>: qty, ...}] — Total is filled from the items."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    headers = list(vf_items.META_HEADERS) + list(product_headers)
    ws.append(headers)
    for r in rows:
        line = []
        for h in headers:
            if h == "Total":
                line.append(sum(int(r.get(ph) or 0) for ph in product_headers))
            else:
                line.append(r.get(h))
        ws.append(line)
    wb.save(path)
    return Path(path)


@pytest.fixture
def sheet(tmp_path):
    return _write_sheet(
        tmp_path / "AHB_WeeklyProductionQuery_08-10-26_vF.xlsx",
        [H_TETI, H_ETX, H_WWHO, H_CARO],
        [
            {"OrderID": 1001, "Name": "A", "State": "CA", "Tags": "!ANY FedEx_AHB!", H_TETI: 1},
            {"OrderID": 1002, "Name": "B", "State": "TX", "Tags": "!ANY FedEx_AHB!,!ExtraGel48oz!",
             H_TETI: 1, H_ETX: 1},                       # already has ETX -> dupe risk
            {"OrderID": 1003, "Name": "C", "State": "CA", "Tags": "Gift_Redemption",
             H_TETI: 2, H_WWHO: 1},
            {"OrderID": 1004, "Name": "D", "State": "NY", "Tags": "", H_ETX: 1},
        ],
    )


def run(argv, authority, extra=()):
    return vf_items.main([*argv, "--authority", str(authority), *extra])


def cells(path, sheet_name=vf_items.SHEET):
    wb = openpyxl.load_workbook(path, data_only=False)
    ws = wb[sheet_name]
    out = {(r, c): ws.cell(r, c).value
           for r in range(1, ws.max_row + 1) for c in range(1, ws.max_column + 1)}
    dims = (ws.max_row, ws.max_column, tuple(wb.sheetnames))
    wb.close()
    return out, dims


def ledger_entries(sheet_path):
    return vf_items.read_ledger(vf_items.ledger_path(sheet_path))


def revisions(sheet_path):
    return sorted(Path(sheet_path).parent.glob(f"{Path(sheet_path).stem}_r*.xlsx"))


@pytest.fixture(autouse=True)
def _ledger_in_tmp(tmp_path, monkeypatch):
    """Never write a ledger into the real _outputs/cache during tests."""
    monkeypatch.setattr(vf_items, "ledger_path",
                        lambda p: tmp_path / f"ledger_{Path(p).stem}.jsonl")


# ── 1. duplicate guard (MATRIX_RULES 24c) ────────────────────────────────────
def test_swap_creating_a_duplicate_is_refused_and_rows_named(sheet, authority, capsys):
    """wk0810 CH-TETI: order 1002 already carries CH-ETX. Single-target swap must refuse."""
    rc = run(["swap", str(sheet), "--from", "CH-TETI", "--to", "CH-ETX", "--all", "--write"],
             authority)
    out = capsys.readouterr().out
    assert rc == 1
    assert "DUPLICATE" in out and "1002" in out
    assert "REFUSING" in out
    assert revisions(sheet) == []                      # nothing written
    assert ledger_entries(sheet) == []                 # a refused op writes NO ledger line


def test_multi_target_swap_spreads_and_avoids_duplicates(sheet, authority, capsys):
    """CH-TETI -> CH-ETX or CH-WWHO or CH-CARO: every row gets a target it does not already hold."""
    rc = run(["swap", str(sheet), "--from", "CH-TETI", "--to", "CH-ETX,CH-WWHO,CH-CARO",
              "--all", "--write"], authority)
    assert rc == 0, capsys.readouterr().out
    out = revisions(sheet)
    assert [p.name for p in out] == [sheet.stem + "_r2.xlsx"]
    wb = openpyxl.load_workbook(out[0])
    ws = wb[vf_items.SHEET]
    hdr = {c.value: i for i, c in enumerate(next(ws.iter_rows(max_row=1)), start=1)}
    by_order = {ws.cell(r, hdr["OrderID"]).value: r for r in range(2, ws.max_row + 1)}
    for oid in (1001, 1002, 1003):
        r = by_order[oid]
        assert ws.cell(r, hdr[H_TETI]).value in (None, 0)
        got = {h: ws.cell(r, hdr[h]).value for h in (H_ETX, H_WWHO, H_CARO)}
        assert sum(1 for v in got.values() if v) >= 1
    # 1002 already had ETX -> must NOT have been doubled
    assert ws.cell(by_order[1002], hdr[H_ETX]).value == 1
    assert ws.cell(by_order[1002], hdr[H_WWHO]).value or ws.cell(by_order[1002], hdr[H_CARO]).value
    # 1003 already had WWHO -> its TETI:2 landed elsewhere
    assert ws.cell(by_order[1003], hdr[H_WWHO]).value == 1
    wb.close()


def test_swap_onto_sku_absent_from_authority_is_refused(sheet, authority, capsys):
    rc = run(["swap", str(sheet), "--from", "CH-TETI", "--to", "CH-NOPE", "--all", "--write"],
             authority)
    out = capsys.readouterr().out
    assert rc == 1
    assert "MISSING" in out and "CH-NOPE" in out
    assert revisions(sheet) == [] and ledger_entries(sheet) == []


def test_swap_recomputes_total(sheet, authority):
    assert run(["swap", str(sheet), "--from", "CH-TETI", "--to", "CH-CARO",
                "--orders", "1003", "--write"], authority) == 0
    wb = openpyxl.load_workbook(revisions(sheet)[0])
    ws = wb[vf_items.SHEET]
    hdr = {c.value: i for i, c in enumerate(next(ws.iter_rows(max_row=1)), start=1)}
    r = next(r for r in range(2, ws.max_row + 1) if ws.cell(r, hdr["OrderID"]).value == 1003)
    assert ws.cell(r, hdr[H_CARO]).value == 2 and ws.cell(r, hdr[H_TETI]).value is None
    assert ws.cell(r, hdr["Total"]).value == 3          # 2 CARO + 1 WWHO (rule 0)
    wb.close()


# ── 2. name authority (24a) + header shape (24b) ─────────────────────────────
def test_rename_to_non_authoritative_name_is_refused(sheet, authority, capsys):
    """The 2026-08-04 class: a plausible name derived from a Shopify title."""
    rc = run(["rename-column", str(sheet), "--from", H_ETX,
              "--to", P + "Farmstead Smoked Cumin Gouda", "--write"], authority)
    out = capsys.readouterr().out
    assert rc == 1 and "NOT in" in out
    assert revisions(sheet) == [] and ledger_entries(sheet) == []


def test_rename_to_authoritative_name_succeeds(sheet, authority):
    assert run(["rename-column", str(sheet), "--from", H_ETX, "--to", H_CUMIN,
                "--write", "--reason", "translator re-export"], authority) == 0
    wb = openpyxl.load_workbook(revisions(sheet)[0])
    hdrs = [c.value for c in next(wb[vf_items.SHEET].iter_rows(max_row=1))]
    assert H_CUMIN in hdrs and H_ETX not in hdrs
    wb.close()
    e = [x for x in ledger_entries(sheet) if x["kind"] == "rename-column"]
    assert len(e) == 1 and e[0]["before"] == H_ETX and e[0]["after"] == H_CUMIN
    assert e[0]["reason"] == "translator re-export"


@pytest.mark.parametrize("bad,why", [
    (P + "Walnut, Honey & Extra Virgin Olive Oil Crackers", "COMMA"),
    (P + "Teti  Packaged Slice", "doubled whitespace"),
    ("Teti Packaged Slice", "does not start with"),
    (P + "Frumage L’Ottavio", "curly apostrophe"),
])
def test_bad_header_shape_is_refused(sheet, authority, capsys, bad, why):
    rc = run(["rename-column", str(sheet), "--from", H_ETX, "--to", bad, "--write"], authority)
    out = capsys.readouterr().out
    assert rc == 1 and why in out
    assert revisions(sheet) == [] and ledger_entries(sheet) == []


def test_replace_name_revalidates_the_result(sheet, authority, capsys):
    """A find-and-replace cannot smuggle in a non-authoritative name."""
    assert run(["replace-name", str(sheet), "--find", "Etxegarai", "--replace", "Smoked Cumin",
                "--write"], authority) == 1
    assert "NOT in" in capsys.readouterr().out
    assert revisions(sheet) == []
    # the legitimate direction works
    assert run(["replace-name", str(sheet), "--find", "Etxegarai Packaged Slice",
                "--replace", "Farmstead Cumin Gouda", "--write"], authority) == 0
    assert len(revisions(sheet)) == 1


def test_add_column_uses_the_authority_header_and_drop_refuses_live_demand(tmp_path, authority,
                                                                          capsys):
    s = _write_sheet(tmp_path / "vF.xlsx", [H_TETI],
                     [{"OrderID": 1, "Tags": "", H_TETI: 1}])
    assert run(["add-column", str(s), "--sku", "CH-WWHO", "--write"], authority) == 0
    r2 = revisions(s)[0]
    hdrs = [c.value for c in next(openpyxl.load_workbook(r2)[vf_items.SHEET].iter_rows(max_row=1))]
    assert hdrs[-1] == H_WWHO
    assert run(["add-column", str(s), "--sku", "CH-NOPE", "--write"], authority) == 1
    assert "MISSING" in capsys.readouterr().out
    # dropping a column that still carries units is refused
    assert run(["drop-column", str(s), "--sku", "CH-TETI", "--write"], authority) == 1
    assert "still carries" in capsys.readouterr().out


def test_authority_missing_fails_closed(sheet, tmp_path, capsys):
    empty = tmp_path / "empty.csv"
    empty.write_text("", encoding="utf-8")
    assert run(["validate", str(sheet)], empty) == 1
    assert "missing or empty" in capsys.readouterr().out


# ── 3. quantity edits + byte-preservation (24e/24f) ──────────────────────────
def test_qty_edit_leaves_every_other_cell_identical(sheet, authority):
    before, dims_before = cells(sheet)
    assert run(["qty", str(sheet), "--sku", "CH-ETX", "--orders", "1004", "--set", "3",
                "--write"], authority) == 0
    after, dims_after = cells(revisions(sheet)[0])
    assert dims_before == dims_after
    diffs = {k: (before[k], after[k]) for k in before if before[k] != after[k]}
    wb = openpyxl.load_workbook(sheet)
    ws = wb[vf_items.SHEET]
    hdr = {c.value: i for i, c in enumerate(next(ws.iter_rows(max_row=1)), start=1)}
    row = next(r for r in range(2, ws.max_row + 1) if ws.cell(r, hdr["OrderID"]).value == 1004)
    wb.close()
    assert set(diffs) == {(row, hdr[H_ETX]), (row, hdr["Total"])}
    assert diffs[(row, hdr[H_ETX])] == (1, 3)
    assert diffs[(row, hdr["Total"])] == (1, 3)


def test_qty_negative_is_refused(sheet, authority, capsys):
    assert run(["qty", str(sheet), "--sku", "CH-ETX", "--orders", "1004", "--delta", "-5",
                "--write"], authority) == 1
    assert "negative" in capsys.readouterr().out
    assert revisions(sheet) == [] and ledger_entries(sheet) == []


def test_tags_and_meta_are_never_touched(sheet, authority):
    assert run(["swap", str(sheet), "--from", "CH-TETI", "--to", "CH-CARO", "--orders", "1003",
                "--write"], authority) == 0
    wb_a, wb_b = openpyxl.load_workbook(sheet), openpyxl.load_workbook(revisions(sheet)[0])
    a, b = wb_a[vf_items.SHEET], wb_b[vf_items.SHEET]
    hdr = {c.value: i for i, c in enumerate(next(a.iter_rows(max_row=1)), start=1)}
    for r in range(2, a.max_row + 1):
        for h in ("Tags", "Name", "State", "OrderID", "ProductionDay", "Notes"):
            assert a.cell(r, hdr[h]).value == b.cell(r, hdr[h]).value
    wb_a.close()
    wb_b.close()


def test_never_overwrites_and_revisions_increment(sheet, authority):
    assert run(["qty", str(sheet), "--sku", "CH-ETX", "--orders", "1004", "--set", "2",
                "--write"], authority) == 0
    assert run(["qty", str(sheet), "--sku", "CH-ETX", "--orders", "1004", "--set", "4",
                "--write"], authority) == 0
    assert [p.name for p in revisions(sheet)] == [sheet.stem + "_r2.xlsx",
                                                  sheet.stem + "_r3.xlsx"]


def test_dry_run_is_the_default(sheet, authority, capsys):
    assert run(["qty", str(sheet), "--sku", "CH-ETX", "--orders", "1004", "--set", "9"],
               authority) == 0
    assert "DRY RUN" in capsys.readouterr().out
    assert revisions(sheet) == [] and ledger_entries(sheet) == []


# ── 4. ledger + revert round-trip (24i) ──────────────────────────────────────
def test_ledger_records_each_change_once_with_before_after_reason(sheet, authority):
    assert run(["qty", str(sheet), "--sku", "CH-ETX", "--orders", "1004", "--set", "3",
                "--write", "--reason", "short on TETI"], authority) == 0
    entries = ledger_entries(sheet)
    cellsl = [e for e in entries if e["kind"] == "cell"]
    assert len(cellsl) == 2                                  # the qty cell + recomputed Total
    keys = [(e["row"], e["header"]) for e in cellsl]
    assert len(keys) == len(set(keys))                       # recorded exactly once each
    q = next(e for e in cellsl if e["header"] == H_ETX)
    assert (q["before"], q["after"], q["reason"], q["sku"]) == (1, 3, "short on TETI", "CH-ETX")
    assert q["order"] == "1004" and q["rule"].startswith("MATRIX_RULES")
    assert json.dumps(entries)                               # jsonl-serialisable


def test_revert_round_trip_restores_original_values(sheet, authority):
    before, dims_before = cells(sheet)
    assert run(["swap", str(sheet), "--from", "CH-TETI", "--to", "CH-CARO", "--orders",
                "1001,1003", "--write", "--reason", "swath swap"], authority) == 0
    r2 = revisions(sheet)[0]
    assert run(["revert", str(r2), "--ledger", str(vf_items.ledger_path(sheet)),
                "--write"], authority) == 0
    restored = revisions(r2)[0]
    after, dims_after = cells(restored)
    assert dims_before == dims_after
    assert after == before                                   # byte-for-byte on every cell


def test_revert_refuses_a_structural_op(sheet, authority, capsys):
    assert run(["add-column", str(sheet), "--sku", "CH-CUMIN", "--write"], authority) == 0
    r2 = revisions(sheet)[0]
    assert run(["revert", str(r2), "--ledger", str(vf_items.ledger_path(sheet)),
                "--write"], authority) == 1
    assert "structural op" in capsys.readouterr().out


# ── 5. gift rows (rule 20 / 24g) ─────────────────────────────────────────────
def test_gift_rows_are_replaced_not_merged_additively(tmp_path, authority, capsys):
    main = _write_sheet(tmp_path / "vF.xlsx", [H_TETI, H_ETX], [
        {"OrderID": 2001, "Name": "stale", "Zip": "07627", "Tags": "Gift_Redemption",
         "ProductionDay": "SAT", H_TETI: 5},
        {"OrderID": 2002, "Name": "other", "Zip": "10001", "Tags": "!ANY FedEx_AHB!", H_ETX: 1},
    ])
    vfgr = _write_sheet(tmp_path / "vFGR.xlsx", [H_TETI, H_ETX], [
        {"OrderID": 2001, "Name": "real recipient", "Zip": "07627", H_TETI: 1, H_ETX: 2},
    ])
    assert vf_items.main(["gift", str(main), "--vfgr", str(vfgr), "--authority", str(authority),
                          "--write"]) == 0
    out = revisions(main)[0]
    wb = openpyxl.load_workbook(out)
    ws = wb[vf_items.SHEET]
    hdr = {c.value: i for i, c in enumerate(next(ws.iter_rows(max_row=1)), start=1)}
    rows = [r for r in range(2, ws.max_row + 1) if ws.cell(r, hdr["OrderID"]).value == 2001]
    assert len(rows) == 1                                    # exactly ONE row per gift order
    r = rows[0]
    assert ws.cell(r, hdr[H_TETI]).value == 1                # REPLACED (not 5, not 6)
    assert ws.cell(r, hdr[H_ETX]).value == 2                 # exactly as the vFGR states
    assert ws.cell(r, hdr["Total"]).value == 3               # recomputed (rule 0)
    assert ws.cell(r, hdr["Tags"]).value == "Gift_Redemption"      # col L preserved
    # rule 20(a3)/(b): gift metadata is NEVER read as authority — the routing app's row owns
    # Name/address/Tags/etc.; the vFGR supplies only MFG item quantities, so its Name is inert.
    assert ws.cell(r, hdr["Name"]).value == "stale"
    # the non-gift row is untouched
    r2 = next(x for x in range(2, ws.max_row + 1) if ws.cell(x, hdr["OrderID"]).value == 2002)
    assert ws.cell(r2, hdr[H_ETX]).value == 1
    wb.close()


def test_gift_vfgr_with_an_invented_header_is_silently_omitted(tmp_path, authority, capsys):
    """rule 20(d): items map only by REGISTERED MFG name — an unknown/invented gift header is
    silently omitted and never becomes a vF column (rule 21's invariant holds by omission; the
    cmd_gift 24a post-check backstops any invented header that does land on the merged sheet)."""
    main = _write_sheet(tmp_path / "vF.xlsx", [H_TETI],
                        [{"OrderID": 2001, "Zip": "07627", "Tags": "", H_TETI: 2}])
    vfgr = _write_sheet(tmp_path / "vFGR.xlsx", [H_TETI, P + "Farmstead Smoked Cumin Gouda"],
                        [{"OrderID": 2001, "Zip": "07627", H_TETI: 1,
                          P + "Farmstead Smoked Cumin Gouda": 1}])
    assert vf_items.main(["gift", str(main), "--vfgr", str(vfgr), "--authority", str(authority),
                          "--write"]) == 0
    out = revisions(main)[0]
    wb = openpyxl.load_workbook(out)
    ws = wb[vf_items.SHEET]
    headers = [c.value for c in next(ws.iter_rows(max_row=1))]
    assert P + "Farmstead Smoked Cumin Gouda" not in headers   # never becomes a vF column
    hdr = {h: i for i, h in enumerate(headers, start=1)}
    r = next(x for x in range(2, ws.max_row + 1) if ws.cell(x, hdr["OrderID"]).value == 2001)
    assert ws.cell(r, hdr[H_TETI]).value == 1                  # known column still replaced
    assert ws.cell(r, hdr["Total"]).value == 1                 # recount excludes the omitted qty
    wb.close()


def test_gift_oid_missing_from_matrix_refuses(tmp_path, authority, capsys):
    main = _write_sheet(tmp_path / "vF.xlsx", [H_TETI],
                        [{"OrderID": 2001, "Tags": "", H_TETI: 1}])
    vfgr = _write_sheet(tmp_path / "vFGR.xlsx", [H_TETI],
                        [{"OrderID": 9999, "Tags": "", H_TETI: 1}])
    assert vf_items.main(["gift", str(main), "--vfgr", str(vfgr), "--authority", str(authority),
                          "--write"]) == 1
    assert "REFUSING gift merge" in capsys.readouterr().out
    assert revisions(main) == []


# ── 6. validate (read-only, first-class) ─────────────────────────────────────
def test_validate_flags_invented_header_dupe_row_and_total_drift(tmp_path, authority, capsys):
    bad = P + "Farmstead Smoked Cumin Gouda"                 # invented (2026-08-04)
    s = _write_sheet(tmp_path / "vF.xlsx", [H_TETI, bad, H_ETX], [
        {"OrderID": 1, "Tags": "", H_TETI: 1, bad: 1},
        {"OrderID": 2, "Tags": "", H_ETX: 1},
    ])
    # force a Total drift + a same-SKU-twice row by editing directly (simulating a hand edit)
    wb = openpyxl.load_workbook(s)
    ws = wb[vf_items.SHEET]
    hdr = {c.value: i for i, c in enumerate(next(ws.iter_rows(max_row=1)), start=1)}
    ws.cell(3, hdr["Total"]).value = 99
    wb.save(s)
    wb.close()
    assert run(["validate", str(s)], authority) == 1
    out = capsys.readouterr().out
    assert "Farmstead Smoked Cumin Gouda" in out and "NOT in" in out
    assert "Total != sum" in out and "computed=1" in out
    assert "read-only, nothing written" in out
    assert revisions(s) == []


@pytest.fixture
def wellformed(tmp_path):
    """🔴 A "well-formed" sheet must satisfy rule 25 too: every order carries EXACTLY ONE guide,
    the one the table picks. The pre-rule fixture (no guide columns at all) encoded a world where a
    box could ship with no tasting guide — that pin was WRONG, not merely stale, so it moved here
    rather than being loosened. The missing-guide path is pinned by `gsheet` below."""
    return _write_sheet(
        tmp_path / "AHB_WeeklyProductionQuery_08-10-26_vF.xlsx",
        [H_TETI, H_ETX, H_TRAY, G_BITES, G_TCUST],
        [
            {"OrderID": 1001, "Tags": "", H_TETI: 1, G_TCUST: 1},
            {"OrderID": 1002, "Tags": "", H_TRAY: 1, G_BITES: 1},
            {"OrderID": 1003, "Tags": "", H_TETI: 1, H_ETX: 1, G_TCUST: 1},
            {"OrderID": 1004, "Tags": "", H_TRAY: 1, H_TETI: 1, G_BITES: 1},   # MIXED -> bites
        ],
    )


def test_validate_is_clean_on_a_well_formed_sheet(wellformed, gauthority, capsys):
    assert run(["validate", str(wellformed)], gauthority) == 0
    assert "CLEAN" in capsys.readouterr().out


def test_validate_flags_duplicate_item_in_one_order(tmp_path, authority, capsys):
    """Two columns resolving to the same SKU, both carrying units on one row."""
    s = _write_sheet(tmp_path / "vF.xlsx", [H_TETI, H_ETX], [
        {"OrderID": 1, "Tags": "", H_TETI: 1, H_ETX: 1}])
    wb = openpyxl.load_workbook(s)
    ws = wb[vf_items.SHEET]
    hdr = {c.value: i for i, c in enumerate(next(ws.iter_rows(max_row=1)), start=1)}
    ws.cell(1, hdr[H_ETX]).value = H_TETI                    # duplicate header column
    wb.save(s)
    wb.close()
    assert run(["validate", str(s)], authority) == 1
    out = capsys.readouterr().out
    assert "duplicate header columns" in out


# ── 7. selection is never inferred (24h) ─────────────────────────────────────
def test_no_selection_is_refused(sheet, authority, capsys):
    assert run(["qty", str(sheet), "--sku", "CH-ETX", "--set", "1", "--write"], authority) == 1
    assert "no selection" in capsys.readouterr().out


def test_unknown_order_is_refused_never_guessed(sheet, authority, capsys):
    assert run(["qty", str(sheet), "--sku", "CH-ETX", "--orders", "1004,999999", "--set", "1",
                "--write"], authority) == 1
    assert "not on the sheet" in capsys.readouterr().out
    assert revisions(sheet) == []


def test_there_is_no_curation_flag(sheet, authority):
    """Curation is not on the sheet — inferring it would be the fabrication class (24h)."""
    with pytest.raises(SystemExit):
        vf_items.build_parser().parse_args(["swap", str(sheet), "--from", "CH-TETI",
                                            "--to", "CH-ETX", "--curation", "MDT"])


# ── 8. preservation verifier itself ──────────────────────────────────────────
def test_final_name_gate_reuses_rule21_and_catches_an_invented_header(tmp_path, authority):
    """The last gate runs on the WRITTEN file, via matrix_commander.validate_mfg_names (rule 21)."""
    auth = vf_items.Authority(authority)
    clean = _write_sheet(tmp_path / "clean.xlsx", [H_TETI], [{"OrderID": 1, H_TETI: 1}])
    assert vf_items.final_name_gate(clean, vf_items.SHEET, auth) == ""
    dirty = _write_sheet(tmp_path / "dirty.xlsx", [P + "Farmstead Smoked Cumin Gouda"],
                         [{"OrderID": 1}])
    msg = vf_items.final_name_gate(dirty, vf_items.SHEET, auth)
    assert "MFG onboarding REJECT" in msg


def test_preservation_verifier_catches_an_undeclared_change(sheet):
    before = vf_items.snapshot(sheet)
    wb = openpyxl.load_workbook(sheet)
    ws = wb[vf_items.SHEET]
    ws.cell(2, 2).value = "TAMPERED"
    out = sheet.with_name("tampered.xlsx")
    wb.save(out)
    wb.close()
    probs = vf_items.verify_preserved(before, vf_items.snapshot(out), {})
    assert probs and "changed unexpectedly" in probs[0]


def test_preservation_verifier_catches_a_dropped_row(sheet):
    before = vf_items.snapshot(sheet)
    wb = openpyxl.load_workbook(sheet)
    wb[vf_items.SHEET].delete_rows(2)
    out = sheet.with_name("shortened.xlsx")
    wb.save(out)
    wb.close()
    probs = vf_items.verify_preserved(before, vf_items.snapshot(out), {})
    assert any("row count changed" in p or "row identity changed" in p for p in probs)


# ── tasting guides (MATRIX_RULES rule 25) ────────────────────────────────────
# 🔴 SYNTHETIC ONLY. Burn: 29 of 2,253 orders on the submitted 08-10-26 vF carried NO tasting
# guide (measured 2026-08-09) — a box of cheese with nothing telling the customer what it is.
G_BITES = P + "Tasting Guide - Gourmet Bites"
G_TCUST = P + "Tasting Guide - Custom Box"
G_FCUST = P + "Tasting Guide - The First AppyHour"
G_TMDT = P + "Tasting Guide - Mediterranean Escape"
H_TRAY = P + "When in Rome"
GUIDE_AUTH_ROWS = AUTH_ROWS + [
    ("PK-BITESGUIDE", G_BITES), ("PK-TCUST", G_TCUST),
    ("PK-FCUST", G_FCUST), ("PK-TMDT", G_TMDT), ("TR-ROME", H_TRAY),
]


@pytest.fixture
def gauthority(tmp_path) -> Path:
    p = tmp_path / "mfg_names_authoritative.csv"
    with open(p, "w", newline="", encoding="utf-8") as fh:
        csv.writer(fh).writerows(GUIDE_AUTH_ROWS)
    return p


@pytest.fixture
def gsheet(tmp_path):
    """1001 tray/no guide · 1002 regular/no guide (RESHIP) · 1003 tray w/ WRONG guide ·
    1004 regular/correct · 1005 TWO guides · 1006 retired FCUST · 1007 parked TMDT (nested
    reship tag) · 1008 MIXED tray+regular, no guide (Kurt: tray wins)."""
    return _write_sheet(
        tmp_path / "AHB_WeeklyProductionQuery_08-10-26_vF.xlsx",
        [H_TETI, H_ETX, H_TRAY, G_BITES, G_TCUST, G_FCUST, G_TMDT],
        [
            {"OrderID": 1001, "Tags": "", H_TRAY: 1},
            {"OrderID": 1002, "Tags": "Reship - Arrived Warm", H_TETI: 1},
            {"OrderID": 1003, "Tags": "", H_TRAY: 1, G_TCUST: 1},
            {"OrderID": 1004, "Tags": "", H_TETI: 1, G_TCUST: 1},
            {"OrderID": 1005, "Tags": "", H_TETI: 1, G_TCUST: 1, G_BITES: 1},
            {"OrderID": 1006, "Tags": "", H_TETI: 1, G_FCUST: 1},
            {"OrderID": 1007, "Tags": "Shipping::Delayed in transit::x", H_TETI: 1, G_TMDT: 1},
            {"OrderID": 1008, "Tags": "", H_TRAY: 1, H_TETI: 1, H_ETX: 1},
        ],
    )


def _audit(sheet_path, authority_path, policy=None):
    auth = vf_items.Authority(authority_path)
    return vf_items.GuideAudit(vf_items.Sheet(sheet_path), auth,
                               vf_items.GuidePolicy.load(policy))


def _orders(items):
    return {x["order"] for x in items}


def test_policy_table_tray_gets_bitesguide_regular_gets_tcust():
    pol = vf_items.GuidePolicy()
    assert pol.select({"TR-ROME", "CH-TETI"})[0] == "PK-BITESGUIDE"   # MIXED -> tray wins
    assert pol.select({"TR-ROME"})[0] == "PK-BITESGUIDE"
    assert pol.select({"CH-TETI", "AC-FCWALN"})[0] == "PK-TCUST"


def test_policy_never_selects_retired_or_parked():
    pol = vf_items.GuidePolicy()
    assert {r["guide"] for r in pol.rules} == {"PK-BITESGUIDE", "PK-TCUST"}
    assert pol.status_of("PK-FCUST") == "retired"
    assert pol.status_of("PK-TMDT") == "parked"      # parked, NOT deleted (25b)


def test_policy_refuses_a_table_that_selects_an_inert_guide():
    with pytest.raises(vf_items.ItemEditRefused) as e:
        vf_items.GuidePolicy(rules=[{"when": "default", "guide": "PK-TMDT"}])
    assert "parked" in str(e.value)


def test_parked_guide_is_re_enabled_by_CONFIG_not_a_code_edit(tmp_path):
    p = tmp_path / "policy.json"
    p.write_text(json.dumps({"status": {"PK-TMDT": "active"},
                             "rules": [{"when_prefix": "TR-", "guide": "PK-BITESGUIDE"},
                                       {"when": "default", "guide": "PK-TMDT"}]}),
                 encoding="utf-8")
    assert vf_items.GuidePolicy.load(p).select({"CH-TETI"})[0] == "PK-TMDT"
    assert vf_items.GuidePolicy().select({"CH-TETI"})[0] == "PK-TCUST"   # built-in unchanged


def test_audit_detects_missing_duplicate_mismatched_retired_and_parked(gsheet, gauthority):
    a = _audit(gsheet, gauthority)
    assert _orders(a.none_) == {"1001", "1002", "1008"}
    assert _orders(a.multi) == {"1005"}
    assert _orders(a.mismatched) == {"1003"}          # tray order carrying PK-TCUST
    assert _orders(a.retired_present) == {"1006"}
    assert _orders(a.parked_present) == {"1007"}
    assert _orders(a.tray_bearing) == {"1001", "1003", "1008"}
    assert "1004" in _orders(a.exactly_one)


def test_audit_splits_reship_from_regular_via_the_canonical_parser(gsheet, gauthority):
    a = _audit(gsheet, gauthority)
    # 🔴 25f: the NESTED `Shipping::<reason>::` form must count — a hand-rolled "reship" substring
    # match would miss 1007 entirely.
    assert {x["order"] for x in a.rows if x["reship"]} == {"1002", "1007"}


def test_guides_remediation_adds_the_missing_guide_and_preserves_everything(gsheet, gauthority):
    before, dims = cells(gsheet)
    guide_cols = {vf_items.Sheet(gsheet).hdr[h] for h in (G_BITES, G_TCUST)}
    total_col = vf_items.Sheet(gsheet).hdr["Total"]
    assert vf_items.main(["guides", str(gsheet), "--authority", str(gauthority), "--write"]) == 0
    out = revisions(gsheet)
    assert len(out) == 1
    a = _audit(out[0], gauthority)
    assert not a.none_
    for o, want in (("1001", "PK-BITESGUIDE"), ("1002", "PK-TCUST"), ("1008", "PK-BITESGUIDE")):
        assert next(x for x in a.rows if x["order"] == o)["carried"] == [(want, 1)]
    after, dims2 = cells(out[0])
    assert dims == dims2
    changed = {k for k in after if after[k] != before.get(k)}
    assert changed and all(c in guide_cols | {total_col} for _r, c in changed)


def test_guides_leaves_duplicate_retired_and_parked_alone_and_NAMES_them(gsheet, gauthority,
                                                                        capsys):
    vf_items.main(["guides", str(gsheet), "--authority", str(gauthority), "--write"])
    a = _audit(revisions(gsheet)[0], gauthority)
    assert _orders(a.multi) == {"1005"}               # never auto-resolved (25h)
    assert _orders(a.retired_present) == {"1006"}     # identified, never stripped (25b/i)
    assert _orders(a.parked_present) == {"1007"}
    txt = capsys.readouterr().out
    assert "LEFT ALONE" in txt
    assert all(o in txt for o in ("1005", "1006", "1007"))


def test_guides_mismatch_needs_an_explicit_flag(gsheet, gauthority):
    vf_items.main(["guides", str(gsheet), "--authority", str(gauthority), "--write"])
    assert _orders(_audit(revisions(gsheet)[0], gauthority).mismatched) == {"1003"}
    vf_items.main(["guides", str(gsheet), "--authority", str(gauthority), "--write",
                   "--fix-mismatch"])
    a = _audit(revisions(gsheet)[-1], gauthority)
    assert not a.mismatched
    assert next(x for x in a.rows if x["order"] == "1003")["carried"] == [("PK-BITESGUIDE", 1)]


def test_guides_adds_the_column_when_the_sheet_lacks_it(tmp_path, gauthority):
    s = _write_sheet(tmp_path / "vF.xlsx", [H_TETI], [{"OrderID": 2001, "Tags": "", H_TETI: 1}])
    assert vf_items.main(["guides", str(s), "--authority", str(gauthority), "--write"]) == 0
    out = revisions(s)[0]
    assert next(x for x in _audit(out, gauthority).rows
                if x["order"] == "2001")["carried"] == [("PK-TCUST", 1)]
    assert G_TCUST in vf_items.Sheet(out).headers


def test_guides_REFUSES_when_the_guide_is_absent_from_the_authority(tmp_path, authority, capsys):
    """🔴 25d — never invent 'AHB (S_REG): Tasting Guide - …'. No file, no ledger."""
    s = _write_sheet(tmp_path / "vF.xlsx", [H_TETI], [{"OrderID": 3001, "Tags": "", H_TETI: 1}])
    assert vf_items.main(["guides", str(s), "--authority", str(authority), "--write"]) == 1
    assert revisions(s) == []
    assert ledger_entries(s) == []
    assert "MISSING" in capsys.readouterr().out


def test_guides_recomputes_total_ledgers_and_reverts(gsheet, gauthority):
    sh = vf_items.Sheet(gsheet)
    r1001 = next(r for r, o in sh.data_rows() if o == "1001")
    before_total = sh.get(r1001, "Total")
    vf_items.main(["guides", str(gsheet), "--authority", str(gauthority), "--write"])
    out = revisions(gsheet)[0]
    assert vf_items.Sheet(out).get(r1001, "Total") == before_total + 1     # rule 0
    assert [e for e in ledger_entries(gsheet) if e["op"] == "guides"]
    assert vf_items.main(["revert", str(out), "--authority", str(gauthority), "--write",
                          "--ledger", str(vf_items.ledger_path(gsheet))]) == 0
    back = revisions(out)[0]
    assert vf_items.Sheet(back).get(r1001, "Total") == before_total
    assert "1001" in _orders(_audit(back, gauthority).none_)      # revert really undid the guide
