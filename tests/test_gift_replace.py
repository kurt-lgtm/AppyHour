"""MATRIX_RULES rule 20 — vFGR gift-order REPLACE semantics (Kurt 2026-07-24, wk0727 by hand).

Gift redemption orders are UNEDITABLE in Shopify, so the matrix rows built from Shopify carry
stale/too-few items. The weekly *_vFGR.xlsx is the ITEM-TRUTH:
  (a) A-suffix twin rows (Simple Bundles "Associated Order") fold onto the parent FIRST;
      pre-combined vFGRs (remove=1, no A row) fold nothing — never sum twice.
  (b) REPLACE matrix rows by OrderID: meta + items from vFGR; PRESERVE matrix Tags (engine
      col L) + ProductionDay; Notes blank (rule 18); Total recomputed (rule 0).
  (c) vFGR OrderID missing from the matrix (e.g. _HOLD) = loud GiftMergeError; explicit
      drop only via drop_oids.
"""

from __future__ import annotations

import openpyxl
import pytest

import matrix_commander as mc
from matrix_commander import GiftMergeError

META = ["OrderID", "Name", "Total", "Zip", "Tags", "Notes", "ProductionDay"]
P1 = "AHB (S_REG): Montasio"
P2 = "AHB (S_REG): Barista"
P3 = "AHB (S_REG): Chorizo Seco"
ENGINE_L = ("!ExtraGel48oz!, !ExtraGel48oz!, !FedEx Home Delivery - Indianapolis_AHB!, "
            "Gift_Redemption, MaxIce, MaxIce")


def _write_xlsx(path, headers, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Access_LIVE"
    ws.append(headers)
    for r in rows:
        ws.append(r)
    wb.save(str(path))
    return str(path)


def _load(path):
    ws = openpyxl.load_workbook(path, data_only=True)["Access_LIVE"]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    return rows[0], {str(r[0]): r for r in rows[1:] if r[0] is not None}


def _main(tmp_path):
    # stale matrix row for gift 164878 (too few items) + a regular order
    return _write_xlsx(
        tmp_path / "main.xlsx",
        META + [P1, P2, P3],
        [
            [164878, "Old Stale Name", 1, "92026", ENGINE_L, "junk note", "SAT", 1, None, None],
            [150000, "Regular", 2, "10001", "!UPS Ground - Dallas_AHB!", None, "SAT", 1, 1, None],
        ],
    )


def test_replace_changes_only_items_and_preserves_all_app_fields(tmp_path):
    gift = _write_xlsx(
        tmp_path / "gift.xlsx",
        META + [P1, P2],
        [[164878, "True Recipient", 99, "92026", "Gift Redemption, RMFG", "gift note", "SAT", None, 2]],
    )
    headers, rows = _load(mc.merge_gift_xlsx(_main(tmp_path), gift))
    row = rows["164878"]

    assert row[headers.index("Name")] == "Old Stale Name"          # app metadata wins
    assert row[headers.index(P1)] is None                            # stale item WIPED
    assert row[headers.index(P2)] == 2                               # item-truth refilled
    assert row[headers.index("Tags")] == ENGINE_L                    # engine col L PRESERVED
    assert row[headers.index("ProductionDay")] == "SAT"
    assert row[headers.index("Notes")] == "junk note"
    assert row[headers.index("Total")] == 1                          # fixed app cell unchanged
    # untouched regular order
    assert rows["150000"][headers.index(P1)] == 1


def test_twin_fold_sums_onto_parent_and_drops_a_row(tmp_path):
    gift = _write_xlsx(
        tmp_path / "gift.xlsx",
        META + [P1, P2, "remove"],
        [
            [164878, "Parent", 0, "92026", "", None, "SAT", 1, 1, None],
            ["164878A", "Twin", 0, "92026", "", None, "SAT", 2, None, None],
        ],
    )
    headers, rows = _load(mc.merge_gift_xlsx(_main(tmp_path), gift))

    assert "164878A" not in rows                                     # twin never ships
    row = rows["164878"]
    assert row[headers.index(P1)] == 3                               # 1 + 2 summed
    assert row[headers.index(P2)] == 1
    assert row[headers.index("Total")] == 1
    assert "remove" not in headers                                   # bookkeeping col not unioned


def test_no_a_row_means_nothing_is_folded(tmp_path):
    """wk0727 shape: a single parent row, no A row — nothing to fold, so nothing doubles.

    This used to be called `test_precombined_remove_flag_never_sums_twice` and the `remove=1` cell
    was believed to be a "pre-combined" flag. It is not: `remove` is a placeholder MFG name for a
    BL- (bulk, non-fulfillable) SKU, so that 1 is a QUANTITY (Kurt 2026-07-28). The no-double
    guarantee comes from there being no A row, never from the flag.
    """
    gift = _write_xlsx(
        tmp_path / "gift.xlsx",
        META + [P1, "remove"],
        [[164878, "Parent", 16, "92026", "", None, "SAT", 3, 1]],
    )
    headers, rows = _load(mc.merge_gift_xlsx(_main(tmp_path), gift))
    assert rows["164878"][headers.index(P1)] == 3                    # untouched, not doubled
    assert rows["164878"][headers.index("Total")] == 1


def test_placeholder_column_is_dropped_and_never_counted(tmp_path):
    """A column named `remove` is a placeholder for a BL- SKU RMFG cannot pick — drop it.

    Its quantity must not reach the sheet OR the col-D Total (rule 0): a bulk SKU counted as a
    pick line inflates the box and misstates demand.
    """
    gift = _write_xlsx(
        tmp_path / "gift.xlsx",
        META + [P1, "remove"],
        [[164878, "Parent", 99, "92026", "", None, "SAT", 3, 1]],
    )
    headers, rows = _load(mc.merge_gift_xlsx(_main(tmp_path), gift))
    assert "remove" not in headers
    assert not any(str(h).strip().lower() in {"remove", "delete", "ignore"} for h in headers)
    assert rows["164878"][headers.index("Total")] == 1               # app Total is untouched


def test_placeholder_column_dropped_even_if_shaped_like_a_product(tmp_path):
    """Onboarding that BL- SKU properly must not sneak an unfulfillable column onto the sheet.

    Today `remove` lacks the "AHB (" prefix so it falls out anyway — an accident of the broken
    generator, not a guarantee. The drop is by NAME and runs before the product-shape check.
    """
    assert mc._is_placeholder_header("remove")
    assert mc._is_placeholder_header("  REMOVE ")
    # exact match only — a fuzzy "contains" would eat a real product name someday
    assert not mc._is_placeholder_header("AHB (S_REG): Removable Rind Brie")


def test_trailing_all_blank_row_is_ignored(tmp_path):
    """The generator appends an all-None row in 5 of 8 weekly vFGRs (QC pass 2026-07-28).

    Both paths drop it on the empty OrderID. Pinned because it was previously untested — a refactor
    that iterates rows without the OID check would silently process a blank order.
    """
    gift = _write_xlsx(
        tmp_path / "gift.xlsx",
        META + [P1],
        [
            [164878, "Parent", 0, "92026", "", None, "SAT", 3],
            [None] * 8,
        ],
    )
    headers, rows = _load(mc.merge_gift_xlsx(_main(tmp_path), gift))
    assert rows["164878"][headers.index(P1)] == 3
    assert "" not in rows and "None" not in rows


def test_gift_zip_is_ignored_even_when_numeric(tmp_path):
    """A zip stored as a NUMBER has already lost its leading zero — never overwrite Shopify's with it.

    07627 -> 7627 mis-routes a cold-chain box and is only found after it ships.
    """
    gift = _write_xlsx(
        tmp_path / "gift.xlsx",
        META + [P1],
        [[164878, "Parent", 0, 7627, "", None, "SAT", 3]],
    )
    headers, rows = _load(mc.merge_gift_xlsx(_main(tmp_path), gift))
    assert rows["164878"][headers.index("Zip")] == "92026"


def test_renamed_gift_meta_header_is_inert(tmp_path):
    """Alignment is by header name: a renamed meta column would silently stop being overwritten."""
    meta_renamed = [h if h != "Zip" else "ZIP " for h in META]
    gift = _write_xlsx(
        tmp_path / "gift.xlsx",
        meta_renamed + [P1],
        [[164878, "Parent", 0, "92026", "", None, "SAT", 3]],
    )
    headers, rows = _load(mc.merge_gift_xlsx(_main(tmp_path), gift))
    assert rows["164878"][headers.index("Zip")] == "92026"


def test_unknown_mfg_name_is_silently_omitted(tmp_path):
    unknown = "AHB (S_REG): Definitely Not A Registered MFG Name"
    gift = _write_xlsx(
        tmp_path / "gift.xlsx", META + [P1, unknown, "remove"],
        [[164878, "Gift Name", 99, "00000", "gift tags", "gift note", "TUE", 2, 7, 1]],
    )
    headers, rows = _load(mc.merge_gift_xlsx(_main(tmp_path), gift))
    assert unknown not in headers and "remove" not in headers
    assert rows["164878"][headers.index(P1)] == 2


def test_duplicate_normalized_product_headers_collapse_and_sum(tmp_path):
    main = _write_xlsx(
        tmp_path / "main.xlsx", META + [P1, P1],
        [[164878, "App Name", 9, "92026", ENGINE_L, "app note", "SAT", 1, 2]],
    )
    gift = _write_xlsx(
        tmp_path / "gift.xlsx", META + [P1, P1],
        [[164878, "Gift Name", 99, "00000", "gift tags", "gift note", "TUE", 3, 4]],
    )
    headers, rows = _load(mc.merge_gift_xlsx(main, gift))
    assert headers.count(P1) == 1
    assert rows["164878"][headers.index(P1)] == 7
    assert rows["164878"][headers.index("Name")] == "App Name"


def test_orphan_twin_row_is_loud_error(tmp_path):
    gift = _write_xlsx(
        tmp_path / "gift.xlsx", META + [P1], [["164878A", "Twin", 0, "92026", "", None, "SAT", 2]]
    )
    with pytest.raises(GiftMergeError, match="164878A"):
        mc.merge_gift_xlsx(_main(tmp_path), gift)


def test_missing_cohort_oid_is_loud_error(tmp_path):
    # wk0727 #165505: _HOLD order in the vFGR but not the cohort — never silent
    gift = _write_xlsx(
        tmp_path / "gift.xlsx",
        META + [P1],
        [[164878, "P", 0, "92026", "", None, "SAT", 1], [165505, "Held", 0, "91915", "", None, "SAT", 1]],
    )
    with pytest.raises(GiftMergeError, match="165505"):
        mc.merge_gift_xlsx(_main(tmp_path), gift)


def test_explicit_gift_drop_excludes_and_reports(tmp_path, capsys):
    gift = _write_xlsx(
        tmp_path / "gift.xlsx",
        META + [P1],
        [[164878, "P", 0, "92026", "", None, "SAT", 1], [165505, "Held", 0, "91915", "", None, "SAT", 1]],
    )
    headers, rows = _load(mc.merge_gift_xlsx(_main(tmp_path), gift, drop_oids={"165505"}))
    assert "165505" not in rows
    assert rows["164878"][headers.index(P1)] == 1
    assert "165505" in capsys.readouterr().out                       # drop surfaced, never silent
