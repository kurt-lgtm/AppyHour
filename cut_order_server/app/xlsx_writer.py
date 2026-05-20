"""xlsx writer — Calibri throughout, hidden source cols, NET color coding.

Two tabs:
  1. Cut Order — raw-ingredient rows with cut qty + hidden contributing SKUs
  2. Demand — finished-SKU demand with hidden RC/SH/first-order/override breakdown
"""
from __future__ import annotations

import io
import json
from datetime import datetime
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

CALIBRI = "Calibri"
HEADER_FILL = PatternFill("solid", fgColor="1f2328")
HEADER_FONT = Font(name=CALIBRI, bold=True, color="FFFFFF", size=11)
ROW_FONT = Font(name=CALIBRI, size=11)
THIN = Side(border_style="thin", color="BBBBBB")
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# NET color thresholds (cut_lbs)
NET_RED = PatternFill("solid", fgColor="fde0de")
NET_YELLOW = PatternFill("solid", fgColor="fff4c2")
NET_GREEN = PatternFill("solid", fgColor="d8f5d8")


def _set_widths(ws, widths: dict[int, float]) -> None:
    for col_idx, w in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = w


def _write_header(ws, row: int, headers: list[str]) -> None:
    for ci, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = CELL_BORDER


def _write_row(ws, row: int, values: list[Any], fills: dict[int, PatternFill] | None = None) -> None:
    for ci, v in enumerate(values, start=1):
        c = ws.cell(row=row, column=ci, value=v)
        c.font = ROW_FONT
        c.border = CELL_BORDER
        if fills and ci in fills:
            c.fill = fills[ci]


def _net_fill(cut_lbs: float) -> PatternFill:
    if cut_lbs > 0:
        return NET_RED
    if cut_lbs < 0:
        return NET_GREEN
    return NET_YELLOW  # zero/borderline


def build_xlsx(
    *,
    demand_result: dict,        # serialized DemandResult dict
    calc_result: dict,          # serialized CalcResult dict
    multiplier_knob: float,
    multiplier_ratios: dict,
) -> bytes:
    wb = openpyxl.Workbook()

    # ─── Tab 1: Cut Order ────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Cut Order"

    title = ws1.cell(row=1, column=1, value=f"Cut Order — WK1 {demand_result['wk1_start']} → {demand_result['wk1_end']}")
    title.font = Font(name=CALIBRI, bold=True, size=14)
    ws1.merge_cells(start_row=1, end_row=1, start_column=1, end_column=8)

    sub = ws1.cell(row=2, column=1, value=f"Snapshot: {calc_result.get('snapshot_date') or 'unknown'} · Ship tags: {', '.join(demand_result.get('ship_tags', []))} · Multiplier knob: {multiplier_knob}")
    sub.font = Font(name=CALIBRI, italic=True, color="6b6e74", size=10)
    ws1.merge_cells(start_row=2, end_row=2, start_column=1, end_column=8)

    headers = ["Raw Ingredient", "Pack Size", "UoM", "Box Demand (lbs)", "Tray Demand (lbs)", "Total Demand (lbs)", "Available (lbs)", "Cut (lbs)", "Cut (packs)", "Contributing SKUs (hidden)"]
    _write_header(ws1, 4, headers)
    _set_widths(ws1, {1: 42, 2: 10, 3: 8, 4: 16, 5: 16, 6: 18, 7: 14, 8: 12, 9: 12, 10: 50})

    for i, r in enumerate(calc_result["rows"], start=5):
        contrib = ", ".join(f"{k}={v}" for k, v in sorted(r.get("contributing_skus", {}).items()))
        _write_row(ws1, i, [
            r["raw_name"], r["pack_size"], r["uom"],
            r["box_demand_lbs"], r["tray_demand_lbs"], r["total_demand_lbs"],
            r["available_lbs"], r["cut_lbs"], r["cut_packs"], contrib,
        ], fills={8: _net_fill(r["cut_lbs"])})

    # Hide contributing-SKUs col by default
    ws1.column_dimensions["J"].hidden = True
    ws1.freeze_panes = "A5"

    # ─── Tab 2: Demand ───────────────────────────────────────────────────
    ws2 = wb.create_sheet("Demand")
    title2 = ws2.cell(row=1, column=1, value=f"Demand — finished SKUs (WK1 {demand_result['wk1_start']})")
    title2.font = Font(name=CALIBRI, bold=True, size=14)
    ws2.merge_cells(start_row=1, end_row=1, start_column=1, end_column=7)

    headers2 = ["SKU", "Total Demand", "RC (queued)", "SH (orders)", "First-order subset", "Override?", "Empirical ratio"]
    _write_header(ws2, 3, headers2)
    _set_widths(ws2, {1: 20, 2: 14, 3: 12, 4: 12, 5: 18, 6: 12, 7: 16})

    rc = demand_result.get("rc_by_sku", {}) or {}
    sh = demand_result.get("sh_by_sku", {}) or {}
    first = demand_result.get("first_order_by_sku", {}) or {}
    overrides = demand_result.get("overrides_applied", {}) or {}

    all_skus = sorted(set(demand_result.get("per_sku", {}).keys()))
    global_ratio = (multiplier_ratios or {}).get("__global__", 1.0)
    row_n = 4
    for sku in all_skus:
        is_override = sku in overrides
        ratio = (multiplier_ratios or {}).get(sku, global_ratio if first.get(sku) else "")
        _write_row(ws2, row_n, [
            sku,
            demand_result["per_sku"].get(sku, 0),
            rc.get(sku, 0),
            sh.get(sku, 0),
            first.get(sku, 0),
            "YES" if is_override else "",
            ratio,
        ])
        row_n += 1

    # Hide source breakdown cols (per feedback_cut_order_hidden_source_cols)
    for col in ("C", "D", "E", "G"):
        ws2.column_dimensions[col].hidden = True
    ws2.freeze_panes = "A4"

    # ─── Save to bytes ───────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
