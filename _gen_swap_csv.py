
import openpyxl, json, requests, time, csv

with open('C:/Users/Work/Claude Projects/AppyHour/InventoryReorder/dist/inventory_reorder_settings.json', encoding='utf-8') as f:
    settings = json.load(f)

STORE = settings['shopify_store_url']
TOKEN = settings['shopify_access_token']
BASE = f"https://{STORE}.myshopify.com/admin/api/2024-01"
HEADERS = {"X-Shopify-Access-Token": TOKEN, "Content-Type": "application/json"}

NAME_TO_SKU = {
    "Praline Pecans": "AC-PRPE", "Gooseberry Elderflower Mini Jam": "AC-GBEF",
    "Caramelized Fig & Pear with Honey Mini Jam": "AC-CFPH",
    "Piri Piri Cocktail Mix": "AC-PPCM", "Stone Ground Mustard for Cheese": "AC-MUSTCH",
    "Artisan Crisps - Tart Cherry Cranberry & Cacao Nib": "AC-TCRISP",
    "Artisan Crisp Apricot Pistachio Brandy": "AC-ACRISP",
    "Baked Lemon Ricotta": "CH-BLR", "Blackberry Balsamic Jam": "AC-BLBALS",
    "Dried Tart Cherries": "AC-DTCH", "Lonza": "MT-LONZ",
    "Marinated Australian Farmstead Cheese": "CH-MAFT", "Prairie Breeze": "CH-BRZ",
    "Toscano Salame": "MT-TUSC", "Leonora": "CH-LEON",
    "McCalls Grassfed Irish Porter Cheddar": "CH-MCPC",
    "Wensleydale with Mango & Ginger": "CH-WMANG", "Jambon Honey & Herb": "MT-JAHH",
    "Los Cameros de Romero": "CH-LOSC", "Tipperary Brie": "CH-TIP",
    "Strawberry Rhubarb Mini Jam": "AC-SRHUB", "Wooly Wooly Diablo": "CH-WWDI",
    "Petit Truffle Triple Cream Brie": "CH-TTBRIE", "Barista": "CH-BARI",
    "Finocchiona": "MT-SFEN", "Apple Maple Butter Mini": "AC-APMB",
    "Lemon Feta & Olive Mix": "AC-LFOLIVE",
    "Rustic Bakery Organic Olive Oil & Sel Gris Flatbread Bites": "AC-RBOL",
    "Sottocenere with Truffles": "CH-SOT", "Sopressata": "MT-SOP",
    "Raw Honey with Comb": "AC-HON", "Prosciutto": "MT-PRO",
    "Jamon Serrano": "MT-JAMS", "Sun-Dried Turkish Figs": "AC-SDF",
    "Toma Provence": "CH-TOPR", "Dark Chocolate Covered Almonds": "AC-DALM",
    "Strawberry Lemon Lavendar Jam": "AC-SLL", "Everything Flatbreads": "AC-EFLAT",
    "Bacon Marmalade": "AC-BACO", "Manchego Aurora": "CH-MAU3",
    "Smoked Paprika Chorizo": "MT-SPAP", "Applewood Smoked Speck": "MT-ASPK",
    "Farmstead Smoked Gouda": "CH-MSMG", "Red Wine & Garlic Sliced Salami": "MT-SLRWG",
    "Sweet & Spicy Prosciutto": "MT-PSS", "Spanish Salted Marcona Almonds": "AC-MARC",
    "Capocollo": "MT-CAPO", "Dark Chocolate Covered Cranberries": "AC-DCRAN",
    "Spiced Peach Bourbon Jam": "AC-SBPBJ", "Everything Bagel Cheese Curds": "CH-EBCC",
    "Fontal": "CH-FONTAL", "Fiddlehead IPA Cheddar": "CH-IPAC",
    "Piave Vecchio": "CH-PVEC", "California Clothbound Cheddar": "CH-FOWC",
    "Fig Lemon Honey & Honey Preserves": "AC-FLH", "Sweet & Smoky Almonds": "AC-SMAL",
    "KM39": "CH-KM39", "Barricato Al Pepe": "CH-BAP",
    "Ubriaco Pinot Rose": "CH-UROSE", "Ubriacone": "CH-UCONE",
    "Petit Garlic & Pepper Triple Cream Brie": "CH-GPBRIE",
    "Peach Bellini Preserves": "AC-PBLINI",
    "Tasting Guide - The First AppyHour": "PK-TSTBX0",
    "Tasting Guide - Custom Box": "PK-TSTG", "Tasting Guide - Curated Box": "PK-TSTBL",
    "Hazelnuts": "AC-HAZEL", "Raspberry Hibiscus Jam": "AC-RHB",
    "Red Pepper Jelly Mini": "AC-RPJ",
}

wb = openpyxl.load_workbook('C:/Users/Work/Claude Projects/AppyHour/AHB_WeeklyProductionQuery_03-17-26_vF.xlsx', data_only=True)
ws = wb['Access_LIVE']
headers = {}
for c in range(1, ws.max_column+1):
    h = str(ws.cell(1, c).value or '')
    if h.startswith('AHB') and ': ' in h:
        headers[c] = h.split(': ', 1)[1]

matrix = {}
order_names = {}
order_emails = {}
for r in range(2, ws.max_row+1):
    oid = str(ws.cell(r, 1).value or '').strip()
    if not oid: continue
    order_names[oid] = str(ws.cell(r, 2).value or '')
    order_emails[oid] = str(ws.cell(r, 6).value or '')
    skus = {}
    for c, name in headers.items():
        val = ws.cell(r, c).value
        if val and str(val).strip() not in ('', '0', 'None'):
            sku = NAME_TO_SKU.get(name, f'??-{name[:15]}')
            skus[sku] = int(float(str(val)))
    matrix[oid] = skus

# Fetch Shopify
orders = []
url = f'{BASE}/orders.json'
params = {'status': 'open', 'fulfillment_status': 'unfulfilled', 'limit': 250, 'tag': 'RMFG_20260317', 'fields': 'id,name,line_items,email'}
while url:
    resp = requests.get(url, headers=HEADERS, params=params if not orders else None, timeout=30)
    resp.raise_for_status()
    orders.extend(resp.json().get('orders', []))
    link = resp.headers.get('Link', '')
    url = None
    if 'rel="next"' in link:
        for part in link.split(','):
            if 'rel="next"' in part:
                url = part.split('<')[1].split('>')[0]
    time.sleep(0.5)

shopify = {}
shopify_ids = {}
for o in orders:
    name = o['name'].replace('#', '')
    skus = {}
    for li in o.get('line_items', []):
        sku = (li.get('sku') or '').strip()
        fq = li.get('fulfillable_quantity', li.get('quantity', 0))
        if sku and fq > 0:
            skus[sku] = skus.get(sku, 0) + fq
    shopify[name] = skus
    shopify_ids[name] = o['id']

# Generate per-order swap list
rows = []
for oid in sorted(set(matrix.keys()) & set(shopify.keys())):
    m_skus = set(matrix[oid].keys())
    s_skus = set(shopify[oid].keys())
    
    # Only AC- swaps
    to_remove = {s for s in s_skus - m_skus if s.startswith('AC-')}
    to_add = {s for s in m_skus - s_skus if s.startswith('AC-')}
    
    for rem in to_remove:
        add = to_add.pop() if to_add else ''
        rows.append({
            'order_number': oid,
            'order_id': shopify_ids.get(oid, ''),
            'customer': order_names.get(oid, ''),
            'email': order_emails.get(oid, ''),
            'remove_sku': rem,
            'add_sku': add,
        })
    # Any remaining adds without a paired remove
    for add in to_add:
        rows.append({
            'order_number': oid,
            'order_id': shopify_ids.get(oid, ''),
            'customer': order_names.get(oid, ''),
            'email': order_emails.get(oid, ''),
            'remove_sku': '',
            'add_sku': add,
        })

outpath = 'C:/Users/Work/Downloads/ac-swap-list-2026-03-17.csv'
with open(outpath, 'w', newline='', encoding='utf-8') as f:
    w = csv.DictWriter(f, fieldnames=['order_number','order_id','customer','email','remove_sku','add_sku'])
    w.writeheader()
    w.writerows(rows)

print(f'Wrote {len(rows)} swaps to {outpath}')

from collections import Counter
remove_counts = Counter(r['remove_sku'] for r in rows if r['remove_sku'])
add_counts = Counter(r['add_sku'] for r in rows if r['add_sku'])
print(f'\nRemove summary:')
for sku, cnt in remove_counts.most_common():
    print(f'  {sku}: {cnt}')
print(f'\nAdd summary:')
for sku, cnt in add_counts.most_common():
    print(f'  {sku}: {cnt}')
