"""Compare production matrix (Excel) vs Shopify orders for RMFG_20260317."""
import openpyxl, json, requests, time

with open('InventoryReorder/dist/inventory_reorder_settings.json', encoding='utf-8') as f:
    settings = json.load(f)

STORE = settings['shopify_store_url']
TOKEN = settings['shopify_access_token']
BASE = f"https://{STORE}.myshopify.com/admin/api/2024-01"
HEADERS = {"X-Shopify-Access-Token": TOKEN, "Content-Type": "application/json"}

NAME_TO_SKU = {
    "Praline Pecans": "AC-PRPE",
    "Artisan Crisps - Tart Cherry Cranberry & Cacao Nib": "AC-TCRISP",
    "Artisan Crisp Apricot Pistachio Brandy": "AC-ACRISP",
    "Baked Lemon Ricotta": "CH-BLR",
    "Blackberry Balsamic Jam": "AC-BLBALS",
    "Dried Tart Cherries": "AC-DTCH",
    "Gooseberry Elderflower Mini Jam": "AC-GBEF",
    "Lonza": "MT-LONZ",
    "Marinated Australian Farmstead Cheese": "CH-MAFT",
    "Prairie Breeze": "CH-BRZ",
    "Toscano Salame": "MT-TUSC",
    "Leonora": "CH-LEON",
    "McCalls Grassfed Irish Porter Cheddar": "CH-MCPC",
    "Wensleydale with Mango & Ginger": "CH-WMANG",
    "Jambon Honey & Herb": "MT-JAHH",
    "Stone Ground Mustard for Cheese": "AC-MUSTCH",
    "Los Cameros de Romero": "CH-LOSC",
    "Tipperary Brie": "CH-TIP",
    "Strawberry Rhubarb Mini Jam": "AC-SRHUB",
    "Wooly Wooly Diablo": "CH-WWDI",
    "Petit Truffle Triple Cream Brie": "CH-TTBRIE",
    "Barista": "CH-BARI",
    "Finocchiona": "MT-SFEN",
    "Apple Maple Butter Mini": "AC-APMB",
    "Piri Piri Cocktail Mix": "AC-PPCM",
    "Lemon Feta & Olive Mix": "AC-LFOLIVE",
    "Rustic Bakery Organic Olive Oil & Sel Gris Flatbread Bites": "AC-RBOL",
    "Sottocenere with Truffles": "CH-SOT",
    "Sopressata": "MT-SOP",
    "Raw Honey with Comb": "AC-HON",
    "Prosciutto": "MT-PRO",
    "Jamon Serrano": "MT-JAMS",
    "Sun-Dried Turkish Figs": "AC-SDF",
    "Toma Provence": "CH-TOPR",
    "Dark Chocolate Covered Almonds": "AC-DALM",
    "Strawberry Lemon Lavendar Jam": "AC-SLL",
    "Everything Flatbreads": "AC-EFLAT",
    "Bacon Marmalade": "AC-BACO",
    "Manchego Aurora": "CH-MAU3",
    "Smoked Paprika Chorizo": "MT-SPAP",
    "Applewood Smoked Speck": "MT-ASPK",
    "Farmstead Smoked Gouda": "CH-MSMG",
    "Red Wine & Garlic Sliced Salami": "MT-SLRWG",
    "Caramelized Fig & Pear with Honey Mini Jam": "AC-CFPH",
    "Sweet & Spicy Prosciutto": "MT-PSS",
    "Spanish Salted Marcona Almonds": "AC-MARC",
    "Capocollo": "MT-CAPO",
    "Dark Chocolate Covered Cranberries": "AC-DCRAN",
    "Spiced Peach Bourbon Jam": "AC-SBPBJ",
    "Everything Bagel Cheese Curds": "CH-EBCC",
    "Fontal": "CH-FONTAL",
    "Fiddlehead IPA Cheddar": "CH-IPAC",
    "Piave Vecchio": "CH-PVEC",
    "California Clothbound Cheddar": "CH-FOWC",
    "Fig Lemon Honey & Honey Preserves": "AC-FLH",
    "Sweet & Smoky Almonds": "AC-SMAL",
    "KM39": "CH-KM39",
    "Barricato Al Pepe": "CH-BAP",
    "Ubriaco Pinot Rose": "CH-UROSE",
    "Tasting Guide - The First AppyHour": "PK-TSTBX0",
    "Tasting Guide - Custom Box": "PK-TSTG",
    "Tasting Guide - Curated Box": "PK-TSTBL",
    "Ubriacone": "CH-UCONE",
    "Hazelnuts": "AC-HAZEL",
    "Raspberry Hibiscus Jam": "AC-RHB",
    "Red Pepper Jelly Mini": "AC-RPJ",
    "Petit Garlic & Pepper Triple Cream Brie": "CH-GPBRIE",
}

# Build matrix
wb = openpyxl.load_workbook('AHB_WeeklyProductionQuery_03-17-26_vF.xlsx', data_only=True)
ws = wb['Access_LIVE']
headers = {}
for c in range(1, ws.max_column+1):
    h = str(ws.cell(1, c).value or '')
    if h.startswith('AHB') and ': ' in h:
        name = h.split(': ', 1)[1]
        headers[c] = name

matrix = {}
for r in range(2, ws.max_row+1):
    order_id = str(ws.cell(r, 1).value or '').strip()
    if not order_id:
        continue
    assignments = {}
    for c, name in headers.items():
        val = ws.cell(r, c).value
        if val and str(val).strip() not in ('', '0', 'None'):
            sku = NAME_TO_SKU.get(name, f'??-{name[:15]}')
            assignments[sku] = int(float(str(val)))
    if assignments:
        matrix[order_id] = assignments

print(f"Matrix orders with assignments: {len(matrix)}")

# Fetch Shopify orders
print("Fetching Shopify orders...")
orders = []
url = f"{BASE}/orders.json"
params = {"status": "open", "fulfillment_status": "unfulfilled", "limit": 250, "tag": "RMFG_20260317", "fields": "id,name,line_items"}
page = 0
while url:
    page += 1
    resp = requests.get(url, headers=HEADERS, params=params if page == 1 else None, timeout=30)
    resp.raise_for_status()
    orders.extend(resp.json().get("orders", []))
    link = resp.headers.get("Link", "")
    url = None
    if 'rel="next"' in link:
        for part in link.split(","):
            if 'rel="next"' in part:
                url = part.split("<")[1].split(">")[0]
                params = None
    time.sleep(0.5)

print(f"Shopify orders: {len(orders)}")

# Build Shopify map
shopify = {}
for o in orders:
    name = o['name'].replace('#', '')
    skus = {}
    for li in o.get('line_items', []):
        sku = (li.get('sku') or '').strip()
        fq = li.get('fulfillable_quantity', li.get('quantity', 0))
        if sku and fq > 0:
            skus[sku] = skus.get(sku, 0) + fq
    shopify[name] = skus

common_orders = set(matrix.keys()) & set(shopify.keys())
print(f"Orders in both: {len(common_orders)}")

food_prefixes = ('CH-', 'MT-', 'AC-')
missing_from_shopify = {}
extra_on_shopify = {}

for oid in common_orders:
    m_food = {s for s in matrix[oid] if s.startswith(food_prefixes)}
    s_food = {s for s in shopify[oid] if s.startswith(food_prefixes)}

    for sku in m_food - s_food:
        missing_from_shopify[sku] = missing_from_shopify.get(sku, 0) + 1
    for sku in s_food - m_food:
        extra_on_shopify[sku] = extra_on_shopify.get(sku, 0) + 1

print(f"\n=== IN MATRIX BUT MISSING FROM SHOPIFY ===")
for sku, count in sorted(missing_from_shopify.items(), key=lambda x: -x[1]):
    print(f"  {sku}: {count} orders")

print(f"\n=== ON SHOPIFY BUT NOT IN MATRIX ===")
for sku, count in sorted(extra_on_shopify.items(), key=lambda x: -x[1]):
    print(f"  {sku}: {count} orders")

# --- AC-PRPE replacement analysis ---
# For orders that have AC-PRPE on Shopify but NOT in matrix,
# find what the matrix has that Shopify doesn't (the replacement)
from collections import Counter
praline_col = None
for c_idx, name in headers.items():
    if 'praline' in name.lower():
        praline_col = c_idx
        break

# Build a per-order matrix using product NAMES (avoid SKU mapping issues)
# Then for orders with AC-PRPE on Shopify but no Praline in matrix,
# count how many have each matrix-only item
replacement_counts = Counter()
prpe_orders_no_matrix = 0

# Build title->SKU map from Shopify data
title_to_sku = {}
for o in orders:
    for li in o.get('line_items', []):
        sku = (li.get('sku') or '').strip()
        title = (li.get('title') or '').strip()
        if sku and title:
            title_to_sku[title] = sku

# For each order with AC-PRPE on Shopify
for o in orders:
    oid = o['name'].replace('#', '')
    has_prpe = any((li.get('sku') or '').strip() == 'AC-PRPE' and li.get('fulfillable_quantity', 0) > 0
                   for li in o.get('line_items', []))
    m_has_prpe = matrix.get(oid, {}).get('AC-PRPE', 0) > 0 if oid in matrix else False

    if has_prpe and not m_has_prpe and oid in matrix:
        prpe_orders_no_matrix += 1
        # Get all Shopify SKUs for this order
        s_skus = set()
        for li in o.get('line_items', []):
            sku = (li.get('sku') or '').strip()
            if sku and li.get('fulfillable_quantity', 0) > 0:
                s_skus.add(sku)

        # Get all matrix SKUs for this order (using our mapping)
        m_skus = set(matrix[oid].keys())

        # Items in matrix but not on Shopify = the replacement
        # But we need correct SKU mapping... use NAME_TO_SKU
        for sku in m_skus - s_skus:
            if sku.startswith(('CH-', 'MT-', 'AC-')) and not sku.startswith('??'):
                replacement_counts[sku] += 1

print(f"\n=== AC-PRPE REPLACEMENT ANALYSIS ===")
print(f"Orders with AC-PRPE on Shopify but NOT in matrix: {prpe_orders_no_matrix}")
print(f"\nWhat the matrix has instead (replacement candidates):")
for sku, count in replacement_counts.most_common(15):
    print(f"  {sku}: {count} orders")
