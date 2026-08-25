"""Sheet (pick list of record) <-> Shopify comparison.

🔴 The sheet must be built from currentQuantity. Two real defects found this way, both
sheet-side: #175526 carried the live 10 trays AND the 10 removed originals; #174939
omitted AC-KETT x2, a paid BL-4USA board component.
"""
from __future__ import annotations
import collections
from .checks import live, sku, tags
from .rules import CHILD

MFG_PREFIX = "AHB (S_REG):"


def load_sheet(path: str, tab: str | None = None):
    """-> {order_id: dict(name, items, guides, tags, zip, per_column)}"""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[tab] if tab else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    h_i = next(i for i, r in enumerate(rows)
               if any(str(c or "").strip() == "OrderID" for c in r))
    hdr = rows[h_i]
    I = {h: i for i, h in enumerate(hdr) if h}
    mfg = [i for i, h in enumerate(hdr) if h and str(h).startswith(MFG_PREFIX)]
    guide = [i for i in mfg if "Tasting Guide" in str(hdr[i])]
    item = [i for i in mfg if i not in guide]
    out = {}
    for r in rows[h_i + 1:]:
        oid = r[I["OrderID"]]
        if oid is None:
            continue
        out[str(oid).strip().lstrip("#")] = {
            "name": r[I.get("Name", 0)],
            "items": sum(int(r[i]) for i in item if isinstance(r[i], (int, float))),
            "guides": sum(int(r[i]) for i in guide if isinstance(r[i], (int, float))),
            "tags": r[I["Tags"]] or "" if "Tags" in I else "",
            "zip": str(r[I["Zip"]] or "") if "Zip" in I else "",
            "columns": {str(hdr[i]).replace(MFG_PREFIX + " ", ""): int(r[i])
                        for i in item if isinstance(r[i], (int, float)) and r[i]},
        }
    return out


def compare(sheet: dict, shop: dict):
    """-> list of dicts for orders whose sheet total != live Shopify child count."""
    out = []
    for oid, s in sheet.items():
        o = shop.get(oid)
        if not o:
            continue
        n = sum(x["current_quantity"] for x in live(o) if sku(x).startswith(CHILD))
        if n == s["items"]:
            continue
        removed = collections.Counter()
        for x in o["line_items"]:
            if sku(x).startswith(CHILD) and x["current_quantity"] == 0:
                removed[sku(x)] += x["quantity"]
        out.append({"order": "#" + oid, "customer": s["name"], "sheet": s["items"],
                    "shopify": n, "delta": n - s["items"],
                    "removed_units": sum(removed.values()),
                    "removed": " ".join(f"{k}x{v}" for k, v in removed.most_common())})
    return out


def missing_guide(sheet: dict, shop: dict):
    out = []
    for oid, s in sheet.items():
        o = shop.get(oid)
        if o and any(t.strip().lower() == "gift redemption" for t in tags(o)):
            continue
        if s["guides"] != 1:
            out.append({"order": "#" + oid, "customer": s["name"], "guides": s["guides"]})
    return out


def tag_mismatch(sheet: dict, shop: dict, rmfg: str, ship: str):
    """6a - every sheet order must carry BOTH this run's tags.

    -> (mismatches, not_pulled). 🔴 Keep them SEPARATE: an order older than --since is a
    WINDOW artifact, not a tag defect. Reporting them together turned 0 real problems into
    "32 tag mismatches" on the 8_24 run.
    """
    out, unpulled = [], []
    for oid in sheet:
        o = shop.get(oid)
        if not o:
            unpulled.append({"order": "#" + oid, "customer": sheet[oid]["name"]})
            continue
        t = tags(o)
        if rmfg not in t or ship not in t:
            out.append({"order": "#" + oid, "issue": "missing run tag",
                        "run_tags": ",".join(x for x in t
                                             if x.startswith(("RMFG_", "_SHIP_")) or x == rmfg)})
    return out, unpulled


def not_on_sheet(sheet: dict, shop: dict, rmfg: str):
    """6c - Shopify order tagged for the run but absent from the pick list (the serious one)."""
    return [{"order": o["name"], "tags": o.get("tags", "")[:80]}
            for oid, o in shop.items()
            if rmfg in tags(o) and oid not in sheet and not o.get("cancelled_at")]
