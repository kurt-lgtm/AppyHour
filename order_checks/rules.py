"""RULE SET loader. 🔴 Read ORDER_CHECKS_RULES.md before changing anything here."""
from __future__ import annotations
import re, os, functools

# EX-PS "Party Size Upgrade" — absent from the RULE SET xlsx. Kurt 2026-08-25.
PARTY = {"CH-": 2, "MT-": 2, "AC-": 2}
# Priced null-SKU packs whose item count is not in the variant title.
PACK_ITEMS = {"curator's choice - extra meat, cheese & accompaniment": 3}
# Paid boards -> components. A REMOVED board forfeits its components' allowance (#176565).
BOARD = {"BL-USA":  ("AC-KETT", "CH-FAG", "AC-BLUCAR", "MT-PARM"),
         "BL-4USA": ("AC-KETT", "CH-FAG", "AC-BLUCAR", "MT-PARM")}
CHILD = ("AC-", "MT-", "CH-", "TR-")

# Blank-SKU StayAI / Simple Bundles wrapper lines, keyed by the bare PRODUCT TITLE
# (GraphQL `title`; REST `name` appends the variant and the lookup silently never fires).
# "Free Brie for a Year" contributes +1 CH- -- this is why AHB-MED measures CH3 against a
# LIKELY row of CH2 on 46 of 82 orders.
WRAPPER_ADD = {
    "AppyHour Box + Free Brie for a Year": {"CH-": 1},
    # the pairing ships as its own PR-CJAM- line, so the wrapper itself adds nothing
    "AppyHour Box + FREE Artisan Cheese & Jam Pairings for Life": {},
    "AppyHour Cheese Box + FREE Artisan Cheese & Jam Pairings for Life": {},
    # --- priced blank-SKU add-on packs. Contribution is the DESCRIPTION, per Kurt
    # 2026-08-28: "those are prosciutto packs, just follow the description." Without
    # these the pack lands in c1_norule and the order gets no count check at all
    # (38 orders on RMFG_20260828).
    "Prosciutto Variety Pack (3-Pack)": {"MT-": 3},
    "Prosciutto (5-Pack)": {"MT-": 5},
    "Curator's Choice - Extra Meat, Cheese & Accompaniment": {"MT-": 1, "CH-": 1, "AC-": 1},
    # mix read off the product DESCRIPTION: Prairie Breeze (cheese), Prosciutto (meat),
    # Sweet & Smoky Almonds + Bacon Marmalade (accompaniments). Kurt 2026-08-28.
    "Ultimate Add-on Package: Summer Cookout": {"CH-": 1, "MT-": 1, "AC-": 2},
}

# AHB-CUR-* "Monthly Curation" rides with AHB-MED/AHB-LGE and adds NO child of its own.
# 🔴 PREPAY-*: a PAYMENT wrapper ("Your Prepaid AppyHour Boxes"), not a box parent. The
# real AHB- box is on the order alongside it, so giving it a rule DOUBLE-COUNTS -- 24 of
# 27 prepay orders failed the moment it was treated as AHB-MED/AHB-LGE. Contributes
# nothing. (Its q is the number of prepaid months, not children in this box.)
ZERO_CONTRIB = ("AHB-CUR-", "PREPAY-")

# Parents that are NOT in Dan's RULE SET xlsx. Merged over it at load time so the order
# gets a count check instead of landing in c1_norule, where nothing is checked at all.
# 🔴 Every entry here is Kurt's stated contribution, never inferred from the SKU name.
EXTRA_RULES = {
    # "like medium, AHB-MED, but 2 cheese instead of 2 meat" (Kurt 2026-08-28).
    # AHB-MED is MT2 CH2 AC3; the two meat slots become cheese. Still 7 total.
    "AHB-CMED": {"MT-": 0, "CH-": 4, "AC-": 3},
}

# Sheet column title -> SKU, where the sheet's MFG name does not fuzzy-match any live
# product title. Dan's list, RUN_2026-08-25; each was validated by matching the column's
# week total against that SKU's Shopify total.
ALIAS_OVERRIDE = {
    "figlemonhoneyhoneypreserves": "AC-FLH",
    "redwhitebluekettlecorn": "AC-KETT",
    "allnaturalwasabipeas": "AC-WASP2",
    "alpblossomfloralcheese": "CH-ALP",
    "fontal": "CH-FONTAL",
    "km39": "CH-KM39",
    "bresaolaitaliana": "MT-IBRES",
    "jambonhoneyherb": "MT-JAHH",
    "tastingguidecustombox": "PK-TCUST",
    "roastededamame": "AC-MAME",            # "Edamame, Roasted & Salted"
    "miticatokettionion": "AC-TOK",         # "Toketti"
    "tastingguidegourmetbites": "PK-BITESGUIDE",
    "loscamerosderomero": "CH-LOSC",
    "loubergier": "CH-LOU",
    "honeyclovergoudapackagedslice": "CH-HCGU",
}


def clean_title(t):
    """Sheet column header / product title -> comparable key."""
    t = re.sub(r"^AHB \(S_REG\):\s*", "", str(t))
    return re.sub(r"[^a-z0-9]", "", t.lower().replace("*", ""))

# A route pin looks like "!UPS Ground - Dallas_AHB!" / "!ANY FedEx - Chicago_AHB!".
# 🔴 Match a WHOLE tag, never a regex across the joined tag string: "!.*?_AHB!" spans
# commas and swallows "!ExtraGel24oz!, !ExtraGel48oz!, !UPS Ground - Dallas_AHB!" as one
# token, so a correctly pinned order reports as a mismatch. Split first, then match.
ROUTE_TAG = re.compile(r"^!.*_AHB!$")


def route_tags(tags) -> list:
    """Route pins from a tag string or list. Whole-tag match only."""
    items = tags if isinstance(tags, list) else str(tags or "").split(",")
    return [t.strip() for t in items if ROUTE_TAG.match(t.strip())]
FIXED_ROUTE_TAG = "fixed_route"
MILITARY_TAG = "military"

# A CEX-CR slot must deliver an actual CRACKER, not merely any AC-. Dan's set,
# RUN_2026-08-25. Catches what a count check cannot see: #176361 (9/9) and #176392
# (11/11) are full but their cracker slot holds something else.
# AC-TOK (Mitica Toketti Onion) added by Kurt 2026-08-25: it is what the CEX-CR slot is
# actually being filled with. #176361 and #176392 already shipped it before anyone
# touched them, and it was the approved fill for the five orders whose AC-FCROSE re-add
# Matrixify bounced. Absent from Dan's set, so his run reports all seven as non-cracker
# fills - tell him when this changes.
CRACKERS = {"AC-FCROSE", "AC-FCEVOO", "AC-ACRISP", "AC-TCRISP",
            "AC-EFLAT", "AC-FCWALN", "AC-PFLAT", "AC-TOK"}
# CEX-/EX- parents -> the child type each contributes.
SLOT_TYPE = {"EX-EM": "MT-", "EX-EC": "CH-", "EX-EA": "AC-",
             "CEX-EM": "MT-", "CEX-EC": "CH-", "CEX-EA": "AC-", "CEX-CR": "AC-"}


@functools.lru_cache(maxsize=4)
def load_rule_set(path: str) -> dict:
    """Parse the RULE SET + SKU ADDS BY DISC CODE tabs. LIKELY is deliberately ignored."""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    out = {"any": {}, "typed": {}, "disc": {}}
    ws = wb["RULE SET"]
    rows = list(ws.iter_rows(values_only=True))
    # 🔴 Header row is NOT always row 0 - Dan ships the file both ways (the 08-18 backup
    # has a blank leading row). Find it, never assume.
    h_i = next(i for i, r in enumerate(rows)
               if str((r[0] or "")).strip().lower() == "prefix")
    hdr = [str(h or "").strip() for h in rows[h_i]]
    for r in rows[h_i + 1:]:
        pre = str(r[0] or "").strip()
        if not pre or pre.startswith("*"):
            continue
        typed = {}
        for i, h in enumerate(hdr[1:], start=1):
            if h in CHILD and isinstance(r[i], (int, float)):
                typed[h] = int(r[i])
            if h == "ANY" and isinstance(r[i], (int, float)):
                out["any"][pre] = int(r[i])
        if typed:
            out["typed"][pre] = typed
    if "SKU ADDS BY DISC CODE" in wb.sheetnames:
        for r in wb["SKU ADDS BY DISC CODE"].iter_rows(values_only=True):
            code = str(r[0] or "").strip().lower()
            if not code or code in ("discount code", "none", "sku addition"):
                continue
            adds = {k: int(v) for k, v in zip(("MT-", "CH-", "AC-"), r[1:4])
                    if isinstance(v, (int, float))}
            if adds:
                out["disc"][code] = adds
    return out


def box_expect(prefix: str, rs: dict):
    """(any_total, typed_dict, deferred). Longest matching prefix wins."""
    if prefix.startswith("AHB-X") or prefix.startswith("BL-"):
        return None, {}, True                       # docx: added separately
    for key in sorted(rs["any"], key=len, reverse=True):
        if prefix.startswith(key):
            return rs["any"][key], {}, False
    for key in sorted(rs["typed"], key=len, reverse=True):
        if prefix.startswith(key):
            return None, dict(rs["typed"][key]), False
    return None, {}, True


def resolve_box(line):
    """Rule-set prefix for a box line, or None.

    🔴 The box parent may have a NULL SKU (226 orders in one cohort). Resolve from
    variant_title, or it reads as NO_BOX.
    """
    s = (line.get("sku") or "").strip().upper()
    if s.startswith(("AHB-", "BL-")):
        return s
    if not s and (line.get("name") or "").lower().startswith(("appyhour box", "appyhour cheese box")):
        v = (line.get("variant_title") or "").lower()
        if v.startswith("medium"):
            return "AHB-MED"
        if v.startswith("large"):
            return "AHB-LGE"
    return None


def paid_pack_items(line) -> int:
    """Items a priced null-SKU add-on pack contributes (0 if not one)."""
    if (line.get("sku") or "").strip() or net(line) <= 0:
        return 0
    name = (line.get("name") or "").lower()
    if name.startswith(("appyhour box", "appyhour cheese box")):
        return 0
    q = line["current_quantity"]
    m = (re.search(r"(\d+)\s*items?", line.get("variant_title") or "", re.I)
         or re.search(r"(\d+)\s*items?", line.get("name") or "", re.I))
    if m:
        return int(m.group(1)) * q
    for k, c in PACK_ITEMS.items():
        if k in name:
            return c * q
    return 0


def net(line) -> float:
    """What the customer PAID for this line: price - LINE-LEVEL discount.

    🔴 NOT pre_tax_price — 0.00 on orders with no discount at all (#176576).
    🔴 NOT discount_allocations — those are ORDER-level codes; #176576's "AppyHour Credit"
       $103 spreads over the box + a paid CH-MAFT the customer did pay for.
    total_discount is the box-builder signal: #174407 CEX-EA price 5.50, discount 5.50 -> 0.
    """
    q = line["current_quantity"]
    return round(float(line["price"]) * q - float(line.get("total_discount") or 0), 2)
