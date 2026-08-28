"""Checks 1/2/3/5/6/8 - faithful port of Dan's run_checks2.py onto the GraphQL shape.

RUN_2026-08-25_SHIP_2026-08-31.

Ported WHOLESALE, not constant-by-constant. His expectations are computed against his
data shape and his bucket model; transplanting individual constants into different
plumbing produced 19 flags with ZERO overlap against his 3 (2026-08-25).

The bucket model is the part that matters and that a flat SHORT/OVER verdict destroys:

  c1_unresolved  shortfall equals EXACTLY the unfilled CEX- slots -> the box is not
                 short, its slots simply have not been filled yet. Not a failure.
  must_have      a CEX- rule row states what the slot must DELIVER: a floor. It is NOT
                 added to the expected total -- that is the whole slot-vs-additive
                 distinction (superseded 2026-08-18 after re-measuring on FREE children
                 only: free CEX-EA sits at +0 in 8 of 9 orders).
  exp_any        an ANY allotment absorbs any type, so typed expectations become FLOORS,
                 never equalities.
  c1_norule      an unknown parent gets its own bucket instead of corrupting a count.
  c1_ahbx        AHB-X{N} may state N inclusive of the PK guide -- report BOTH readings
                 rather than silently picking one.
"""
from __future__ import annotations
import collections
import re

from .rules import CHILD, WRAPPER_ADD, ZERO_CONTRIB

AHBX_RE = re.compile(r"^AHB-X(\d+)")
BCPC_TAG = "BOX_CUSTOMIZED_POST_CHECKOUT"
SLOT_COL = "Contributes"
# Fallback only -- the RULE SET "Contributes" column is authoritative when present.
NON_ADDITIVE_DEFAULT = ("CEX-CR", "CEX-EC", "CEX-EM", "CEX-EA")


def match_prefix(sku, table):
    """Longest matching prefix wins -- AHB-MCUST-TRAY must beat AHB-MCUST-."""
    hits = [p for p in table if sku.startswith(p)]
    return max(hits, key=len) if hits else None


def load_rules(path):
    """-> (RULE, LIKELY, DISC, NON_ADDITIVE)."""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)

    def rows_of(name):
        return [r for r in wb[name].iter_rows(values_only=True)
                if r[0] and not str(r[0]).startswith("*")]

    def tab(name):
        rs = rows_of(name)
        hdr = rs[0]
        return {str(r[0]).strip(): {hdr[i]: r[i] for i in range(1, len(hdr))
                                    if hdr[i] and hdr[i] != SLOT_COL and r[i] is not None}
                for r in rs[1:]}

    def slots(name):
        rs = rows_of(name)
        hdr = list(rs[0])
        if SLOT_COL not in hdr:
            return None                       # column absent -> keep the default
        j = hdr.index(SLOT_COL)
        return tuple(str(r[0]).strip() for r in rs[1:]
                     if len(r) > j and str(r[j] or "").strip().upper() == "SLOT")

    disc = {}
    rs = [r for r in wb["SKU ADDS BY DISC CODE"].iter_rows(values_only=True) if r[0]]
    hdr = rs[0]
    for r in rs[1:]:
        code = str(r[0]).strip().lower()
        if code in ("discount code", "sku addition"):
            continue
        disc[code] = {hdr[i]: r[i] for i in range(1, len(hdr)) if hdr[i] and r[i] is not None}
    return tab("RULE SET"), tab("LIKELY"), disc, (slots("RULE SET") or NON_ADDITIVE_DEFAULT)


def run(orders, sheet, rules, rmfg_tag, ship_tag):
    """orders: {oid: graphql node}. sheet: {oid: {'columns': {sku: qty}, 'tags': str}}."""
    RULE, LIKELY, DISC, NON_ADDITIVE = rules
    R = collections.defaultdict(list)
    norule_obs = collections.defaultdict(list)
    scope = []

    for oid, o in sorted(orders.items()):
        tags = o.get("tags") or []
        if any(re.search(r"reship", t, re.I) for t in tags):
            R["reship_excluded"].append({"order": oid, "tags": sorted(set(tags))})
            continue

        items_all = [e["node"] for e in o["lineItems"]["edges"]]
        # Scope test runs on ALL lines, not just live ones: Simple Bundles zeroes a paid
        # board's parent and explodes it into components, so the board sits at
        # currentQuantity 0 while its children inflate the count (#176565, +4 on exactly
        # BL-4USA's four components).
        xbl = sorted({(i["sku"] or "").strip() for i in items_all
                      if (i["sku"] or "").strip().startswith(("AHB-X", "BL-"))})
        if xbl:
            R["xbl_excluded"].append({"order": oid, "skus": xbl})
        scope.append(oid)

        items = [i for i in items_all if (i.get("currentQuantity") or 0) > 0]
        child, child_sku = collections.Counter(), collections.Counter()
        parents, pk, no_rule = [], [], []
        wrap_add, ahbx = collections.Counter(), []

        for it in items:
            sku = (it["sku"] or "").strip()
            q = it["currentQuantity"]
            hit = [c for c in CHILD if sku.startswith(c)]
            if hit:
                child[hit[0]] += q
                child_sku[sku] += q
                continue
            if sku.startswith("PK-"):
                pk.append((sku, q))
                continue
            if not sku:
                w = WRAPPER_ADD.get(it["title"])
                if w is not None:
                    for k, v in w.items():
                        wrap_add[k] += v * q
                    continue
                no_rule.append((f"(blank SKU) {it['title']}", q))
                continue
            if sku.startswith(ZERO_CONTRIB):
                continue                      # curation marker, contributes nothing
            mx = AHBX_RE.match(sku)
            if mx:
                ahbx.append((sku, int(mx.group(1)), q))
                parents.append(("AHB-X#", sku, q, "AHBX"))
                continue
            m = match_prefix(sku, RULE)
            if m:
                parents.append((m, sku, q, "RULE"))
            else:
                lm = match_prefix(sku, LIKELY)
                if lm:
                    parents.append((lm, sku, q, "LIKELY"))
                else:
                    no_rule.append((sku, q))

        tot = sum(child.values())
        n_cex = sum(q for _, s, q, _ in parents if s.startswith(NON_ADDITIVE))
        exp, must_have = collections.Counter(), collections.Counter()
        exp_any = 0
        for k, v in wrap_add.items():
            exp[k] += v
        for _, nx, q in ahbx:
            exp_any += nx * q
        for m, s, q, src in parents:
            if src == "AHBX":
                continue
            tbl = RULE[m] if src == "RULE" else LIKELY[m]
            for k, v in tbl.items():
                if s.startswith(NON_ADDITIVE):
                    if k != "ANY":
                        must_have[k] = max(must_have[k], v)     # floor, NOT additive
                elif k == "ANY":
                    exp_any += v * q
                else:
                    exp[k] += v * q

        dcodes = [str(c).strip().lower() for c in (o.get("discountCodes") or [])]
        disc_add = collections.Counter()
        for c in dcodes:
            for k, v in DISC.get(c, {}).items():
                disc_add[k] += v
        for k, v in disc_add.items():
            exp[k] += v
        exp_total = sum(exp.values()) + exp_any

        has_ex_addon = any(s.startswith("EX-") for _, s, _, _ in parents)
        unresolved = (exp_total and tot < exp_total and tot + n_cex == exp_total
                      and not has_ex_addon)

        probs = []
        for k, v in exp.items():
            if exp_any:                       # ANY absorbs any type -> floor
                if child.get(k, 0) < v:
                    probs.append(f"{k} expected at least {v}, got {child.get(k, 0)}")
            elif child.get(k, 0) != v:
                probs.append(f"{k} expected {v}, got {child.get(k, 0)}")
        if exp_total and tot != exp_total and not unresolved:
            probs.append(f"total children expected {exp_total}, got {tot}")
        if not unresolved:
            for k, v in must_have.items():
                if child.get(k, 0) < v:
                    probs.append(f"{k} required (swap line present), got {child.get(k, 0)}")

        base = {"order": oid, "parents": [p[1] for p in parents], "child": dict(child),
                "total": tot, "disc": dcodes or None}

        if xbl:
            pass                                          # out of Check 1 ONLY
        elif ahbx and not no_rule:
            n_off = sum(q for _, _, q in ahbx)
            R["c1_ahbx"].append({**base,
                                 "ahbx": [f"{s} (N={n})" for s, n, _ in ahbx],
                                 "required_N": exp_total,
                                 "required_N_minus_1": exp_total - n_off,
                                 "ok_as_written": tot == exp_total,
                                 "ok_if_pk_inclusive": tot == exp_total - n_off,
                                 "slots": n_cex})
        elif unresolved and not no_rule:
            R["c1_unresolved"].append({**base, "resolved": tot, "slots": n_cex,
                                       "required": exp_total})
        elif not parents and not no_rule:
            R["c1_noparent"].append({**base, "has_pk": bool(pk)})
        elif no_rule:
            for s, _ in no_rule:
                norule_obs[s].append(oid)
            R["c1_norule"].append({**base, "no_rule": [s for s, _ in no_rule],
                                   "notes": probs})
        elif probs:
            R["c1_fail"].append({**base, "problems": probs, "expected_total": exp_total})
        else:
            lexp, complete = collections.Counter(), True
            for m, s, q, src in parents:
                if s.startswith(NON_ADDITIVE):
                    continue
                lm = match_prefix(s, LIKELY)
                if not lm:
                    complete = False
                    continue
                for k, v in LIKELY[lm].items():
                    if k != "ANY":
                        lexp[k] += v * q
            for k, v in disc_add.items():
                lexp[k] += v
            devs = [f"{k} likely {v}, got {child.get(k, 0)}"
                    for k, v in sorted(lexp.items()) if child.get(k, 0) != v]
            if complete and lexp and devs:
                R["c1_warn_bcpc" if BCPC_TAG in tags else "c1_warn"].append(
                    {**base, "dev": devs})

        # ---------- Check 3 : exactly one PK- brochure ----------
        npk = sum(q for _, q in pk)
        if npk != 1:
            R["c3"].append({"order": oid, "pk": [f"{s} x{q}" for s, q in pk], "n": npk})

        srow = sheet.get(oid) or {}
        sheet_tags = [t.strip() for t in str(srow.get("tags") or "").split(",") if t.strip()]

        # ---------- Check 5 : gel pack tag ----------
        gel = [t for t in (tags + sheet_tags) if "gel" in t.lower()]
        tray = any(s.startswith("AHB-MCUST-TRAY") for _, s, _, _ in parents)
        if not gel and not tray:
            R["c5"].append({"order": oid, "tags": tags, "sheet_tags": sheet_tags})

        # ---------- Check 2 : sheet vs Shopify, PER SKU ----------
        gift = any("gift redemption" in t.lower() for t in tags)
        sheet_child = collections.Counter(
            {k: v for k, v in (srow.get("columns_sku") or {}).items() if k.startswith(CHILD)})
        s_child = sum(sheet_child.values())
        diff = [(k, sheet_child.get(k, 0), child_sku.get(k, 0))
                for k in set(sheet_child) | set(child_sku)
                if sheet_child.get(k, 0) != child_sku.get(k, 0)]
        if srow and (s_child != tot or diff) and not gift:
            R["c2"].append({"order": oid, "sheet_child_sum": s_child,
                            "shopify_children": tot, "item_diffs": sorted(diff)})

        # ---------- Check 6 : both of this run's tags ----------
        c6 = []
        if rmfg_tag not in tags:
            c6.append(f"missing {rmfg_tag}")
        st = [t for t in tags if t.startswith("_SHIP_")]
        if ship_tag not in st:
            c6.append(f"ship-week tag is {st or ['(none)']}, expected {ship_tag}")
        if c6:
            R["c6"].append({"order": oid, "issues": c6,
                            "fulfillment": o.get("displayFulfillmentStatus")})

        # ---------- Check 8 : duplicate child SKU alongside a CEX-/EX- line ----------
        dups = sorted([(s, n) for s, n in child_sku.items() if n > 1])
        swaps = sorted({s for _, s, _, _ in parents if s.startswith(("CEX-", "EX-"))})
        if dups:
            R["c8"].append({"order": oid, "dupes": [f"{s} x{n}" for s, n in dups],
                            "swap_lines": swaps, "has_swap": bool(swaps)})

    R["scope"] = scope
    R["norule_obs"] = dict(norule_obs)
    return R
