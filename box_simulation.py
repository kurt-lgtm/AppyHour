"""Box Simulation — compute DistVol per Shopify order and assign to box size.

Usage: python box_simulation.py [SHIP_TAG]
Default SHIP_TAG: _SHIP_2026-04-27
"""
from __future__ import annotations

import json
import os
import re
import sys
from datetime import date
from pathlib import Path
from collections import Counter

sys.path.insert(0, r"C:\Users\Work\Claude Projects\AppyHour\AppyHourMCP")
from utils import get_shopify_auth, shopify_graphql  # noqa: E402

from openpyxl import Workbook, load_workbook  # noqa: E402

# DISTVOL_XLSX env override (P7 portability 2026-07-27) — unset resolves to the same Desktop literal,
# so Windows behavior is byte-identical. 🔴 This file is the DistVol source of truth AND a single copy
# on the Desktop; the env hook is what lets a non-Windows host (the DO droplet) point at its own copy
# until the `distvol` table replaces it.
XLSX_LOOKUP = Path(os.environ.get("DISTVOL_XLSX")
                   or r"C:\Users\Work\Desktop\Onboarded Items with DistVol - Updated.xlsx")
OUT_DIR = Path(r"C:\Users\Work\Desktop")

SMALL_MAX = 2.99
LARGE_MAX = 6.7

PREFIX_DEFAULTS = {
    "AC": 0.12,
    "CH": 0.20,
    "MT": 0.07,
    "PK": 0.00,
    "TR": 1.00,
}
PHYSICAL_PREFIXES = set(PREFIX_DEFAULTS.keys())

MANUAL_OVERRIDES: dict[str, float] = {
    # Historical overrides retained until verified in Sheet1
    "CH-GFLEECE": 0.18,
    # wk0803 onboards (Kurt's DistVol sheet 2026-07-31) — not yet in the Desktop xlsx;
    # CH prefix fallback (.20) over-sizes CH-OTTA (.15). Move to Sheet1 when it lands there.
    "AC-QUIC": 0.12,
    "CH-OTTA": 0.15,
    # CH-BOSI / AC-GLAW / AC-TOK / AC-BLUCAR / MT-COPPA / all MT-*
    # now live in Sheet1 directly (patched 2026-04-29)
}


DISTVOL_RATE = 105.3  # cu in per 1.00 DistVol


def _eval_num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().lstrip("=")
    if all(ch in "0123456789.+-*/() " for ch in s) and s:
        try:
            return float(eval(s, {"__builtins__": {}}, {}))
        except Exception:
            return None
    return None


_REF_RE = re.compile(r"\$?([A-Z]+)\$?(\d+)")


def _resolve_ci_column(ws, ci_col_letter: str) -> dict[int, float]:
    """Two-pass resolve Cubic Inches column, substituting cell refs."""
    raw: dict[int, object] = {}
    ci_col_idx = ord(ci_col_letter) - ord("A")
    for r in ws.iter_rows(min_row=2):
        raw[r[0].row] = r[ci_col_idx].value

    ci: dict[int, float] = {}
    for row, v in raw.items():
        n = _eval_num(v)
        if n is not None:
            ci[row] = n

    for row, v in raw.items():
        if row in ci or not isinstance(v, str) or not v.startswith("="):
            continue
        expr = v[1:]
        bad = False

        def repl(m):
            nonlocal bad
            col, ref_row = m.group(1), int(m.group(2))
            if col == ci_col_letter and ref_row in ci:
                return str(ci[ref_row])
            bad = True
            return "0"

        sub = _REF_RE.sub(repl, expr)
        if not bad:
            n = _eval_num(sub)
            if n is not None:
                ci[row] = n
    return ci


# ── MySQL `distvol` replica read path (retirement step 2, ShipRouting/server/DATA_CANON_RULES.md).
# Env-first: ROUTING_INPUTS_DB=1 AND DATABASE_URL → read the table; default = the Desktop xlsx,
# byte-identical. The xlsx STAYS the authority; the table is a replica (etl_history --load-inputs).
# Empty table → loud fallback to xlsx; half-loaded (< floor) → RAISE, never serve a shrunken map.
DB_DISTVOL_MIN_ROWS = 200      # healthy replica ≈239 skus (2026-08-05 load)


def _distvol_db() -> dict[str, float] | None:
    if os.environ.get("ROUTING_INPUTS_DB") != "1":
        return None
    url = os.environ.get("DATABASE_URL", "")
    m = re.match(r"mysql(?:\+\w+)?://([^:]+):([^@]+)@([^:/]+):(\d+)/([^?]+)", url)
    if not m:
        print("WARNING: ROUTING_INPUTS_DB=1 but DATABASE_URL missing/unparseable - using xlsx")
        return None
    import pymysql
    usr, pw, host, port, db = m.groups()
    try:
        con = pymysql.connect(host=host, port=int(port), user=usr, password=pw,
                              database=db, ssl={"ssl": {}})
        try:
            cur = con.cursor()
            cur.execute("SELECT sku, distvol FROM distvol")
            rows = cur.fetchall()
        finally:
            con.close()
    except Exception as e:
        # Transient db failure must not kill box sizing — the xlsx authority is right there.
        print(f"WARNING: distvol db read failed ({type(e).__name__}: {e}) - using xlsx")
        return None
    if not rows:
        print("distvol db replica EMPTY - falling back to xlsx")
        return None
    if len(rows) < DB_DISTVOL_MIN_ROWS:
        raise RuntimeError(
            f"distvol table has only {len(rows)} rows (< {DB_DISTVOL_MIN_ROWS}) - refusing a "
            f"half-loaded replica. Re-run etl_history.py --load-inputs --tables distvol.")
    print(f"distvol: db replica, {len(rows)} skus")
    return {sku: float(dv) for sku, dv in rows}


def build_lookup(xlsx_path=None) -> dict[str, float]:
    # Explicit xlsx_path (manual-upload ingest, phase 2b) bypasses the db-replica branch on
    # purpose — an upload is being validated/loaded, so it must read the given file.
    if xlsx_path is None:
        db_lookup = _distvol_db()
        if db_lookup is not None:
            # MANUAL_OVERRIDES still win on top — same semantics as the xlsx path below, and the code
            # dict is the fresher authority between loads (wk0803 pattern: override lands in code first).
            db_lookup.update(MANUAL_OVERRIDES)
            return db_lookup
    wb = load_workbook(xlsx_path or XLSX_LOOKUP)
    ws = wb["Sheet1"]
    lookup: dict[str, float] = {}
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    sku_idx = header.index("SKU")
    ci_idx = header.index("Cubic Inches")
    dv_idx = header.index("DistVol")
    ci_letter = chr(ord("A") + ci_idx)  # E for our sheet
    ci_map = _resolve_ci_column(ws, ci_letter)

    for row in ws.iter_rows(min_row=2):
        r = row[0].row
        sku = row[sku_idx].value
        if not sku:
            continue
        dv = _eval_num(row[dv_idx].value)
        if dv is None:
            ci = ci_map.get(r)
            if ci is not None:
                dv = ci / DISTVOL_RATE
        if dv is not None:
            lookup[str(sku).strip()] = dv
    lookup.update(MANUAL_OVERRIDES)
    return lookup


def resolve_distvol(sku: str, lookup: dict[str, float]) -> tuple[float, bool]:
    """Return (distvol, flagged). Flag = not present in xlsx reference.
    DV always rounded to 2 decimals (per user pref 2026-04-29)."""
    s = sku.strip()
    if s in lookup:
        return (round(lookup[s], 2), False)
    prefix = s.split("-", 1)[0] if "-" in s else s
    return (round(PREFIX_DEFAULTS.get(prefix, 0.10), 2), True)


QUERY = """
query($cursor: String, $q: String!) {
  orders(first: 100, after: $cursor, query: $q) {
    pageInfo { hasNextPage endCursor }
    edges { node {
      id name tags
      currentSubtotalPriceSet { shopMoney { amount } }
      customer { firstName lastName }
      shippingAddress { city provinceCode zip }
      lineItems(first: 50) {
        # first:50 not 100 (Kurt 2026-06-19): orders(first:100) x lineItems(first:100) blew Shopify's
        # query-cost budget, so Shopify SILENTLY DROPPED the heaviest order nodes (multi-line bundle
        # orders) from each page -> they never entered the cohort / never got routed or ice-sized.
        # first:50 keeps every order's line items in one page (0 deep-paginations on the current cohort)
        # AND stays under cost so no order is dropped. >50-line-item orders still deep-paginate below.
        pageInfo { hasNextPage }
        edges { node { sku name quantity currentQuantity } }
      }
    } }
  }
}
"""

DEEP_QUERY = """
query($id: ID!, $cursor: String) {
  order(id: $id) {
    lineItems(first: 250, after: $cursor) {
      pageInfo { hasNextPage endCursor }
      edges { node { sku name quantity currentQuantity } }
    }
  }
}
"""


CANCELLED_NAMES_Q = """
query($cursor: String, $q: String!) {
  orders(first: 250, after: $cursor, query: $q) {
    pageInfo { hasNextPage endCursor }
    edges { node { name cancelledAt cancelReason } }
  }
}
"""


def cohort_query(tag: str) -> str:
    """THE cohort search string. 🔴 A cohort is UNFULFILLED + OPEN (non-cancelled) orders only
    (Kurt 2026-08-11: "the app should always take unfulfilled open orders only").

    *What went wrong (2026-08-11, ROUTE_TEST_811):* Shopify STRIPS an order's `_SHIP_` tag when it
    is cancelled but LEAVES the cohort tag on. Two cancelled orders (#171480, #171841) therefore
    entered the cohort carrying no ship week, and the verify-tag gate reported "MIXED ship weeks"
    — a true failure with a false diagnosis that sent the investigation hunting ship-week drift.
    The fix belongs in the READ, never in the tags.

    - `-status:cancelled`, NOT `status:open`: `status:open` also excludes ARCHIVED (closed) orders,
      which is a different fact and would silently shrink a cohort. Cancelled is the one we mean.
    - `status` / `financial_status` / `displayFulfillmentStatus` are three different fields —
      never substitute one for another.
    - The RMFG_ carve-out is DELIBERATE and unchanged: RMFG_ tags mark packs that are already
      fulfilled, so filtering them to unfulfilled would empty the cut/matrix reads. Cancellation
      exclusion applies to BOTH branches.
    """
    status_filter = "" if tag.startswith("RMFG_") else " fulfillment_status:unfulfilled"
    return f"tag:{tag}{status_filter} AND -status:cancelled"


def fetch_cancelled_tagged(base: str, hdr: dict, tag: str) -> list[dict]:
    """Orders that STILL carry `tag` after being cancelled — a real (minor) hygiene signal, so it
    gets surfaced as its own line rather than silently dropped by `cohort_query` (silence must fail
    loudly). Read-only. Returns [{name, cancelledAt, cancelReason}]."""
    cursor, out = None, []
    q = f"tag:{tag} AND status:cancelled"
    while True:
        data = shopify_graphql(base, hdr, CANCELLED_NAMES_Q, {"cursor": cursor, "q": q},
                               resource="orders-live")
        out.extend(e["node"] for e in data["orders"]["edges"])
        if not data["orders"]["pageInfo"]["hasNextPage"]:
            return out
        cursor = data["orders"]["pageInfo"]["endCursor"]


def fetch_all_orders(base: str, hdr: dict, tag: str, live: bool = False) -> list[dict]:
    """Fetch a cohort's orders. `live=True` BYPASSES the response cache.

    🔴 The default path is cached (`resource="default"` = 30 min, AppyHourMCP/tools/cache.py
    CACHE_TTL). That is right for a build — the same cohort is fetched repeatedly inside one
    run — and WRONG for anything whose purpose is to observe current Shopify state. Kurt hit
    this on the console's Verify-tag step 2026-08-11: pressing it repeatedly returned the same
    answer for half an hour while orders were being cancelled underneath it, so the gate kept
    failing on data that no longer existed. The same trap produced a "0 of 108 tags landed"
    false alarm the same day when a post-apply verification read went through this function.

    Rule: a caller ASKING WHAT IS TRUE RIGHT NOW passes live=True. Only "orders-live" (any
    resource absent from CACHE_TTL) bypasses; there is no shorter way to say it.
    """
    q = cohort_query(tag)
    res = "orders-live" if live else "default"
    cursor = None
    orders: list[dict] = []
    while True:
        data = shopify_graphql(base, hdr, QUERY, {"cursor": cursor, "q": q}, resource=res)
        edges = data["orders"]["edges"]
        for e in edges:
            n = e["node"]
            if n["lineItems"]["pageInfo"]["hasNextPage"]:
                # Deep paginate line items
                extra = []
                sub_cursor = None
                while True:
                    d2 = shopify_graphql(base, hdr, DEEP_QUERY, {"id": n["id"], "cursor": sub_cursor},
                                        resource=res)
                    li = d2["order"]["lineItems"]
                    extra.extend([x["node"] for x in li["edges"]])
                    if not li["pageInfo"]["hasNextPage"]:
                        break
                    sub_cursor = li["pageInfo"]["endCursor"]
                n["lineItems"]["edges"] = [{"node": x} for x in extra]
            orders.append(n)
        if not data["orders"]["pageInfo"]["hasNextPage"]:
            break
        cursor = data["orders"]["pageInfo"]["endCursor"]
    return orders


def simulate(orders: list[dict], lookup: dict[str, float]) -> list[dict]:
    results = []
    for o in orders:
        name = o["name"]
        cust = o.get("customer") or {}
        customer = f"{cust.get('firstName') or ''} {cust.get('lastName') or ''}".strip()
        addr = o.get("shippingAddress") or {}
        # V = order subtotal (post-discount, post-refund product value, excl. ship/tax) — drives the
        # per-order warm_cost in the ShipRouting air/ground decision. None if absent (engine falls back).
        _sub = ((o.get("currentSubtotalPriceSet") or {}).get("shopMoney") or {}).get("amount")
        subtotal = float(_sub) if _sub is not None else None
        lis = [e["node"] for e in o["lineItems"]["edges"]]
        total = 0.0
        tray_count = 0
        item_count = 0
        flagged: list[str] = []
        has_tray = False
        for li in lis:
            sku = (li.get("sku") or "").strip()
            cq = li.get("currentQuantity")
            qty = int(cq if cq is not None else (li.get("quantity") or 0))
            if qty <= 0 or not sku:
                continue
            prefix = sku.split("-", 1)[0] if "-" in sku else sku
            in_lookup = sku in lookup
            # Admin SKUs not in lookup = skip. Anything in lookup = count (even non-physical prefix).
            if not in_lookup and prefix not in PHYSICAL_PREFIXES:
                continue
            if sku.startswith("PK-"):
                continue
            if sku.startswith("TR-"):
                has_tray = True
                tray_count += qty
            dv, flag = resolve_distvol(sku, lookup)
            if flag:
                flagged.append(sku)
            total += dv * qty
            item_count += qty

        if total > LARGE_MAX:
            box = "Large Box"
            slack = LARGE_MAX - total
            over = True
        elif has_tray or total > SMALL_MAX:
            box = "Medium Box"
            slack = LARGE_MAX - total
            over = False
        else:
            box = "Small Box"
            slack = SMALL_MAX - total
            over = False

        results.append({
            "Order": name,
            "Customer": customer,
            "City": addr.get("city") or "",
            "State": addr.get("provinceCode") or "",
            "ZIP": addr.get("zip") or "",
            "Item Count": item_count,
            "Tray Count": tray_count,
            "Total DistVol": round(total, 3),
            "Box Assignment": box,
            "Subtotal": subtotal,
            "Slack": round(slack, 3),
            "Over Capacity?": over,
            "Flagged SKUs": ", ".join(sorted(set(flagged))),
        })
    return results


def write_xlsx(results: list[dict], out_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Simulation"
    cols = list(results[0].keys()) if results else [
        "Order", "Customer", "City", "State", "ZIP", "Item Count", "Tray Count",
        "Total DistVol", "Box Assignment", "Slack", "Over Capacity?", "Flagged SKUs",
    ]
    ws.append(cols)
    sort_key = lambda r: (r["Box Assignment"], -r["Total DistVol"])
    for r in sorted(results, key=sort_key):
        ws.append([r[c] for c in cols])

    # Summary
    s = wb.create_sheet("Summary")
    n = len(results)
    small = [r for r in results if r["Box Assignment"] == "Small Box"]
    large = [r for r in results if r["Box Assignment"] == "Large Tray Box"]
    over = [r for r in results if r["Over Capacity?"]]
    flagged_orders = [r for r in results if r["Flagged SKUs"]]
    all_flagged = Counter()
    for r in results:
        if r["Flagged SKUs"]:
            for sku in r["Flagged SKUs"].split(", "):
                all_flagged[sku] += 1

    avg_small = round(sum(r["Total DistVol"] for r in small) / len(small), 3) if small else 0
    avg_large = round(sum(r["Total DistVol"] for r in large) / len(large), 3) if large else 0
    pct_small = round(100 * len(small) / n, 1) if n else 0
    pct_large = round(100 * len(large) / n, 1) if n else 0

    buckets = [("0–1", 0, 1), ("1–2", 1, 2), ("2–2.75", 2, 2.75),
               ("2.75–5", 2.75, 5), ("5–6.5", 5, 6.5), (">6.5", 6.5, float("inf"))]
    hist = {name: 0 for name, _, _ in buckets}
    for r in results:
        v = r["Total DistVol"]
        for name, lo, hi in buckets:
            if lo <= v < hi or (hi == float("inf") and v >= lo):
                hist[name] += 1
                break

    s.append(["Total orders", n])
    s.append(["Small Box count", len(small), f"{pct_small}%"])
    s.append(["Large Tray Box count", len(large), f"{pct_large}%"])
    s.append(["Over-capacity count", len(over)])
    s.append(["Orders with flagged SKUs", len(flagged_orders)])
    s.append(["Avg DistVol — Small", avg_small])
    s.append(["Avg DistVol — Large", avg_large])
    s.append([])
    s.append(["Histogram", "Count"])
    for name, _, _ in buckets:
        s.append([name, hist[name]])
    s.append([])
    s.append(["Over-capacity orders", "Total DistVol"])
    for r in over:
        s.append([r["Order"], r["Total DistVol"]])
    s.append([])
    s.append(["Flagged SKU", "Frequency"])
    for sku, cnt in all_flagged.most_common():
        s.append([sku, cnt])

    wb.save(out_path)


def main() -> None:
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    tag = args[0] if args else "_SHIP_2026-04-27"
    force_refresh = "--refresh" in sys.argv
    cache_path = Path(rf"C:\Users\Work\box_sim_cache_{tag}.json")
    lookup = build_lookup()
    print(f"Loaded {len(lookup)} SKUs from lookup.")
    if cache_path.exists() and not force_refresh:
        orders = json.loads(cache_path.read_text(encoding="utf-8"))
        print(f"Loaded {len(orders)} orders from cache ({cache_path.name}).")
    else:
        base, hdr = get_shopify_auth()
        orders = fetch_all_orders(base, hdr, tag)
        cache_path.write_text(json.dumps(orders), encoding="utf-8")
        print(f"Fetched {len(orders)} orders with tag {tag} -> cached.")
    results = simulate(orders, lookup)
    today = date.today().isoformat()
    safe_tag = tag.strip("_")
    out = OUT_DIR / f"Box_Simulation_{safe_tag}_{today}.xlsx"
    i = 2
    while True:
        try:
            out.touch(exist_ok=True)
            with open(out, "a"):
                pass
            break
        except PermissionError:
            out = OUT_DIR / f"Box_Simulation_{safe_tag}_{today}_v{i}.xlsx"
            i += 1
    write_xlsx(results, out)
    n = len(results)
    small = sum(1 for r in results if r["Box Assignment"] == "Small Box")
    large = sum(1 for r in results if r["Box Assignment"] == "Large Tray Box")
    flagged = len({s for r in results for s in r["Flagged SKUs"].split(", ") if s})
    over = sum(1 for r in results if r["Over Capacity?"])
    a = round(100 * small / n, 1) if n else 0
    b = round(100 * large / n, 1) if n else 0
    print(f"{n} orders — {a}% small, {b}% large, {flagged} flagged SKUs, {over} over-capacity")
    print(f"Output: {out}")


if __name__ == "__main__":
    main()
