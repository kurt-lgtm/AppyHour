"""Parser for Veho shipping breakdown XLSX files.

File pattern: AHB_NNNNN_Veho Shipping Breakdown_AHB_M-D-YY.xlsx
Sheet: 'Query result'
29 columns. Key fields:
  Tracking ID, Total Rate, Origin Zip, Injection Market,
  Delivery Market, Delivery Zip, Zone, Tendered/Created Timestamp,
  Charge Name 1 (service), External ID (order ref)

Notes:
- Veho does NOT include delivery dates → transit_days is None
- Hub derived from 'Injection Market' (Indianapolis / Nashville / Dallas / Inland Empire -> Anaheim)
- State derived from delivery zip prefix (bootstrapped from shipping.db)
"""

import os
import sqlite3
from datetime import datetime
from typing import Dict, List, Optional

from .common import Shipment, parse_date_flexible

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# Veho's Injection Market -> AHB hub naming
_INJECTION_TO_HUB = {
    'indianapolis': 'Indianapolis',
    'nashville': 'Nashville',
    'dallas': 'Dallas',
    'inland empire': 'Anaheim',   # Veho's SoCal DC ~= our Anaheim hub
    'los angeles': 'Anaheim',
}


_ZIP3_STATE_CACHE: Optional[Dict[str, str]] = None


def _load_zip3_state_map(db_path: str) -> Dict[str, str]:
    """Bootstrap zip3 -> state map from existing shipments table."""
    global _ZIP3_STATE_CACHE
    if _ZIP3_STATE_CACHE is not None:
        return _ZIP3_STATE_CACHE
    mapping: Dict[str, str] = {}
    try:
        conn = sqlite3.connect(db_path)
        cur = conn.cursor()
        cur.execute(
            "SELECT substr(zip_code,1,3) z3, state FROM shipments "
            "WHERE zip_code != '' AND state != '' "
            "GROUP BY z3, state"
        )
        # Prefer most-common state per zip3 (in case of straddlers)
        counts: Dict[str, Dict[str, int]] = {}
        cur.execute(
            "SELECT substr(zip_code,1,3), state, COUNT(*) FROM shipments "
            "WHERE zip_code != '' AND state != '' GROUP BY 1,2"
        )
        for z3, st, ct in cur.fetchall():
            if not z3:
                continue
            counts.setdefault(z3, {})[st] = ct
        for z3, by_state in counts.items():
            mapping[z3] = max(by_state.items(), key=lambda kv: kv[1])[0]
        conn.close()
    except sqlite3.Error:
        pass
    _ZIP3_STATE_CACHE = mapping
    return mapping


def _default_db_path() -> str:
    """Canonical path to shipping.db for zip3 bootstrap.

    Sources from appyhour_lib.paths so we don't drift from the canonical
    location. Falls back to legacy ShippingReports/output/shipments.db
    only if appyhour_lib isn't importable (e.g. parser used standalone).
    """
    try:
        # Add AppyHour repo root to sys.path so appyhour_lib resolves
        import sys
        here = os.path.dirname(os.path.abspath(__file__))
        root = os.path.normpath(os.path.join(here, '..', '..'))
        if root not in sys.path:
            sys.path.insert(0, root)
        from appyhour_lib.paths import db_path  # noqa: E402
        return str(db_path())
    except ImportError:
        here = os.path.dirname(os.path.abspath(__file__))
        return os.path.normpath(
            os.path.join(here, '..', 'output', 'shipments.db')
        )


def _derive_hub(origin_zip: str, injection_market: str) -> str:
    inj = (injection_market or '').strip().lower()
    if inj in _INJECTION_TO_HUB:
        return _INJECTION_TO_HUB[inj]
    # Fallback by origin zip (46xxx -> Indianapolis, 37xxx -> Nashville, 75xxx -> Dallas)
    oz = (origin_zip or '').strip()
    if oz.startswith('46'):
        return 'Indianapolis'
    if oz.startswith('37'):
        return 'Nashville'
    if oz.startswith('75'):
        return 'Dallas'
    if oz.startswith('9'):
        return 'Anaheim'
    return 'Unknown'


def parse_veho_xlsx(filepath: str, zip3_map: Optional[Dict[str, str]] = None) -> List[Shipment]:
    """Parse a Veho invoice XLSX into Shipment records."""
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl required for Veho XLSX parsing. Install: pip install openpyxl")

    if zip3_map is None:
        zip3_map = _load_zip3_state_map(_default_db_path())

    shipments: List[Shipment] = []
    invoice_id = os.path.basename(filepath).split('_')[1] if '_' in os.path.basename(filepath) else ''

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active  # Veho invoices have single sheet "Query result"
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col_idx = {h: i for i, h in enumerate(headers) if h}

    for row in ws.iter_rows(min_row=2, values_only=True):
        def g(name):
            idx = col_idx.get(name)
            return row[idx] if idx is not None and idx < len(row) else None

        tracking = g('Tracking ID')
        if not tracking:
            continue

        origin_zip = str(g('Origin Zip') or '').strip()
        injection = str(g('Injection Market') or '').strip()
        hub = _derive_hub(origin_zip, injection)

        # Skip HQ/unknown origins we don't want counted
        if hub == 'HQ_IGNORE':
            continue

        deliv_zip = str(g('Delivery Zip') or '').strip()[:5]
        state = zip3_map.get(deliv_zip[:3], '') if deliv_zip else ''

        tendered = g('Tendered Timestamp')
        created = g('Created Timestamp')
        ship_dt = None
        for val in (tendered, created):
            if isinstance(val, datetime):
                ship_dt = val.date()
                break
            parsed = parse_date_flexible(val)
            if parsed:
                ship_dt = parsed
                break

        try:
            cost = float(g('Total Rate') or 0)
        except (TypeError, ValueError):
            cost = 0.0

        zone_raw = g('Zone')
        zone = str(zone_raw) if zone_raw is not None else ''

        service = str(g('Charge Name 1') or g('Charge Code 1') or '').strip()

        shipments.append(Shipment(
            tracking=str(tracking),
            carrier='Veho',
            service=service,
            hub=hub,
            state=state,
            zip_code=deliv_zip,
            city=str(g('Delivery Market') or '').strip(),
            zone=zone,
            cost=cost,
            ship_date=ship_dt,
            delivery_date=None,       # Veho invoices lack POD date
            invoice_id=invoice_id,
            source_file=filepath,
        ))

    wb.close()
    return shipments
