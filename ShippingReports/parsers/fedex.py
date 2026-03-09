"""Parser for FedEx shipping breakdown XLSX files."""

import os
from typing import List
from .common import Shipment, parse_date_flexible, identify_hub

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def parse_fedex_xlsx(filepath: str) -> List[Shipment]:
    """Parse a FedEx shipping breakdown XLSX into Shipment records.

    Expected file pattern: AHB_NNNNN_FedEx Shipping Breakdown_AHB_M-D-YY.XLSX
    Key columns: Service Type, Shipment Date, POD Delivery Date, Net Charge Amount,
                 Recipient State/City/Zip Code, Shipper City/State, Zone Code
    """
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl required for FedEx XLSX parsing. Install: pip install openpyxl")

    shipments = []
    invoice_id = os.path.basename(filepath).split('_')[1] if '_' in os.path.basename(filepath) else ''

    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col_idx = {h: i for i, h in enumerate(headers) if h}

    for row in ws.iter_rows(min_row=2, values_only=True):
        def g(name):
            idx = col_idx.get(name)
            return row[idx] if idx is not None and idx < len(row) else None

        tracking = g('Express or Ground Tracking ID')
        if not tracking:
            continue

        hub = identify_hub(
            shipper_city=str(g('Shipper City') or ''),
            shipper_state=str(g('Shipper State') or ''),
        )

        # Skip HQ shipments (Woburn, MA)
        if hub == 'HQ_IGNORE':
            continue

        ship_date = parse_date_flexible(g('Shipment Date'))
        pod_date = parse_date_flexible(g('POD Delivery Date'))

        net = g('Net Charge Amount')
        try:
            cost = float(net)
        except (TypeError, ValueError):
            cost = 0.0

        shipments.append(Shipment(
            tracking=str(tracking),
            carrier='FedEx',
            service=str(g('Service Type') or '').strip(),
            hub=hub,
            state=str(g('Recipient State') or '').strip().upper(),
            zip_code=str(g('Recipient Zip Code') or '').strip()[:5],
            city=str(g('Recipient City') or '').strip(),
            zone=str(g('Zone Code') or '').strip(),
            cost=cost,
            ship_date=ship_date,
            delivery_date=pod_date,
            invoice_id=invoice_id,
            source_file=filepath,
        ))

    wb.close()
    return shipments
