#!/usr/bin/env python

# /// script
# requires-python = ">=3.10"
# dependencies = ["google-api-python-client", "google-auth", "openpyxl", "requests"]
# ///

"""
Cut Order Excel v2 — three-tab polished workbook.

Tab 1: "Cut Order" — urgency-grouped SKU demand with SUMIF-linked assignments
Tab 2: "Assignments" — PR-CJAM, CEX-EC, and MONTHLY slot tables
Tab 3: "Raw Materials" — cheese wheels and bulk accompaniment potential

Pulls demand from Recharge (subscriptions) and Shopify (orders).
"""

from __future__ import annotations

import argparse
import csv as _csv
import io
import json
import os
import re as _re
import sys
import time
from collections import defaultdict
from datetime import date, datetime, timedelta

import openpyxl
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

BASE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, BASE)
from inventory_demand_report import (
    INV_CSV,
    PICKABLE_PREFIXES,
    SAT_DEPLETION,
    TUE_DEPLETION,
    WK1_END,
    WK1_SHIP_TAGS,
    WK1_START,
    WK2_END,
    WK2_START,
    fetch_recharge_api,
    fetch_shopify_orders,
    load_inventory_csv,
    load_settings,
    parse_depletion_xlsx,
    SETTINGS_PATH,
)
from fulfillment_web.invoice_processor import extract_bulk_weights

# Ship weeks ALWAYS start Monday, so the cut order is labeled by its ship-week
# Monday — the forward ship-tag date (next week's _SHIP_ Monday), NOT WK1_END
# (the upcoming Saturday, which is only the Recharge billing boundary). (Kurt 2026-07-06)
SHIP_WEEK_MONDAY = date.fromisoformat(WK1_SHIP_TAGS[-1].replace("_SHIP_", ""))

# ── Design Tokens ────────────────────────────────────────────────────

HEADER_BG = "1E293B"
HEADER_FG = "FFFFFF"
SHORTAGE_BG = "FEE2E2"
SHORTAGE_FG = "991B1B"
TIGHT_BG = "FEF3C7"
TIGHT_FG = "92400E"
OK_BG = "F0FDF4"
OK_FG = "166534"
INPUT_BG = "EEF2FF"
INPUT_FG = "3730A3"
SURFACE = "F8FAFC"
MUTED = "94A3B8"
WHEEL_POT = "7C3AED"
SECTION_ACCENT = "0F172A"

# ── Fonts ────────────────────────────────────────────────────────────

F_SKU = Font(name="Calibri", size=10)
F_NAME = Font(name="Calibri", size=10)
F_NUM = Font(name="Calibri", size=12)
F_NUM_BOLD = Font(name="Calibri", size=12, bold=True)
F_NUM_MUTED = Font(name="Calibri", size=12, color=MUTED)
F_NUM_WHEEL = Font(name="Calibri", size=12, color=WHEEL_POT)
F_HDR = Font(name="Calibri", size=11, bold=True, color=HEADER_FG)
F_SECTION = Font(name="Calibri", size=10, bold=True)
F_INPUT = Font(name="Calibri", size=12, bold=True, color=INPUT_FG)
F_GOOD = Font(name="Calibri", size=9, bold=True)
F_EDIT = Font(name="Calibri", size=10, bold=True, color=INPUT_FG)
F_TITLE = Font(name="Calibri", size=14, bold=True, color=HEADER_FG)
F_SUBTITLE = Font(name="Calibri", size=10, color=MUTED)

# ── Fills ────────────────────────────────────────────────────────────

FILL_HEADER = PatternFill("solid", fgColor=HEADER_BG)
FILL_SHORTAGE = PatternFill("solid", fgColor=SHORTAGE_BG)
FILL_TIGHT = PatternFill("solid", fgColor=TIGHT_BG)
FILL_OK = PatternFill("solid", fgColor=OK_BG)
FILL_INPUT = PatternFill("solid", fgColor=INPUT_BG)
FILL_SURFACE = PatternFill("solid", fgColor=SURFACE)

# ── Alignment ────────────────────────────────────────────────────────

A_RIGHT = Alignment(horizontal="right", vertical="center")
A_CENTER = Alignment(horizontal="center", vertical="center")
A_LEFT = Alignment(horizontal="left", vertical="center")


# ── Helpers ──────────────────────────────────────────────────────────


def _sku_category(sku: str) -> str:
    if sku.startswith("CH-"):
        return "CHEESE"
    if sku.startswith("MT-"):
        return "MEAT"
    if sku.startswith("AC-"):
        return "ACCOMPANIMENTS"
    return "OTHER"


def _set_col_widths(ws: Worksheet, widths: dict[int, float]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def _dark_header_row(ws: Worksheet, row: int, headers: list[str], col_start: int = 1) -> None:
    for ci, h in enumerate(headers):
        c = ws.cell(row=row, column=col_start + ci, value=h)
        c.font = F_HDR
        c.fill = FILL_HEADER
        c.alignment = A_CENTER


def _merge_title_bar(ws: Worksheet, row: int, text: str, last_col: int) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)
    c = ws.cell(row=row, column=1, value=text)
    c.font = F_TITLE
    c.fill = FILL_HEADER
    c.alignment = A_LEFT
    # Fill merged range
    for ci in range(2, last_col + 1):
        ws.cell(row=row, column=ci).fill = FILL_HEADER


def _merge_subtitle(ws: Worksheet, row: int, text: str, last_col: int) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)
    c = ws.cell(row=row, column=1, value=text)
    c.font = F_SUBTITLE
    c.fill = FILL_HEADER
    c.alignment = A_LEFT
    for ci in range(2, last_col + 1):
        ws.cell(row=row, column=ci).fill = FILL_HEADER


def _section_header(ws: Worksheet, row: int, text: str, bg: str, fg: str, last_col: int) -> None:
    fill = PatternFill("solid", fgColor=bg)
    font = Font(name="Calibri", size=10, bold=True, color=fg)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)
    c = ws.cell(row=row, column=1, value=text)
    c.font = font
    c.fill = fill
    for ci in range(2, last_col + 1):
        ws.cell(row=row, column=ci).fill = fill


# ── SKU Aliases ──────────────────────────────────────────────────────
# Maps retired/duplicate SKUs to their canonical equivalents.
# Demand from aliased SKUs is folded into the canonical SKU.
SKU_ALIASES: dict[str, str] = {
    "CH-RAGPT": "CH-RP",
}


def _normalize_sku(sku: str) -> str:
    """Return canonical SKU, resolving aliases."""
    return SKU_ALIASES.get(sku, sku)


# ── Data Fetching (copied from v1 with cleanup) ─────────────────────


def _fetch_all_data(settings: dict) -> dict:
    """Fetch inventory, Recharge, and Shopify data. Returns a dict of all data needed."""
    sku_translations = settings.get("sku_translations", {})
    recharge_token = settings.get("recharge_api_token", "")
    inv_settings = settings.get("inventory", {})

    # -- Load inventory --
    print("Loading inventory...")
    try:
        inventory = load_inventory_csv(INV_CSV)
    except (KeyError, ValueError):
        inventory = {}

    if not inventory:
        with open(INV_CSV, encoding="utf-8-sig") as _f:
            reader = _csv.reader(_f)
            hdr = next(reader)
            avail_col = next((i for i, h in enumerate(hdr) if "available" in (h or "").lower()), 4)
            for row in reader:
                sku = (row[0] if row else "").strip()
                if sku and sku.startswith(PICKABLE_PREFIXES):
                    try:
                        inventory[sku] = int(float(row[avail_col] or 0))
                    except (ValueError, IndexError):
                        pass
        print(f"  Loaded {len(inventory)} SKUs from template check (col {avail_col})")

    # -- Load CSV as DictReader for bulk weights and SKU names --
    csv_rows: list[dict] = []
    csv_sku_names: dict[str, str] = {}
    try:
        with open(INV_CSV, encoding="utf-8-sig") as _f:
            csv_rows = list(_csv.DictReader(_f))
        for row in csv_rows:
            sku = (row.get("Product SKU") or "").strip()
            name = (row.get("Ingredient") or "").strip()
            if sku and name:
                csv_sku_names[sku] = name
        print(f"  Loaded {len(csv_sku_names)} SKU names from inventory CSV")
    except Exception as e:
        print(f"  Warning: Could not load CSV as DictReader: {e}")

    # -- Load Shopify SKU database as final fallback --
    shopify_sku_names: dict[str, str] = {}
    _sku_db_path = os.path.join(os.path.dirname(BASE), "AppyHourMCP", "data", "sku_database.json")
    if os.path.exists(_sku_db_path):
        try:
            with open(_sku_db_path, encoding="utf-8") as _sdf:
                shopify_sku_names = json.load(_sdf)
        except Exception:
            pass

    def sku_name(sku: str) -> str:
        # 1. Dropbox inventory CSV (canonical "Cheese Slice, ..." format)
        if sku in csv_sku_names:
            return csv_sku_names[sku]
        # 2. Settings inventory
        data = inv_settings.get(sku, {})
        settings_name = data.get("name", "") if isinstance(data, dict) else ""
        if settings_name:
            return settings_name
        # 3. Shopify SKU database (raw Shopify titles)
        return shopify_sku_names.get(sku, "")

    bulk_weights = extract_bulk_weights(csv_rows) if csv_rows else {}

    # -- Depletions --
    print("Parsing depletions...")
    sat_dep: dict[str, int] = {}
    tue_dep: dict[str, int] = {}
    if SAT_DEPLETION:
        sat_dep, _, _, _, _ = parse_depletion_xlsx(SAT_DEPLETION, sku_translations)
    if TUE_DEPLETION:
        tue_dep, _, _, _, _ = parse_depletion_xlsx(TUE_DEPLETION, sku_translations)

    available: dict[str, int] = {}
    all_inv_skus = set(inventory.keys()) | set(sat_dep.keys()) | set(tue_dep.keys())
    for sku in all_inv_skus:
        available[sku] = inventory.get(sku, 0) - sat_dep.get(sku, 0) - tue_dep.get(sku, 0)

    # -- Override with corrected inventory if provided --
    # The weekly HAVE (corrected_inventory_path) is AUTHORITATIVE for on-hand: a SKU the
    # HAVE does NOT list has on-hand 0, NOT a stale value from the 3-week-old base INV_CSV.
    # (Bug 2026-07-14: GFLEECE et al. dropped from the HAVE showed stale 6/23 base Avail;
    # harmless only because those had 0 demand — a dropped high-demand SKU would mis-cut.)
    corrected_inv_path = settings.get("corrected_inventory_path", "")
    if corrected_inv_path and os.path.exists(corrected_inv_path):
        try:
            _ci_wb = openpyxl.load_workbook(corrected_inv_path, read_only=True, data_only=True)
            _ci_ws = _ci_wb[_ci_wb.sheetnames[0]]
            _ci_have: dict[str, int] = {}
            for _ci_row in _ci_ws.iter_rows(min_row=2, values_only=True):
                _ci_sku = (_ci_row[0] or "") if _ci_row else ""
                _ci_qty = _ci_row[6] if len(_ci_row) > 6 else None  # Column G = Corrected Qty
                if _ci_sku and _ci_qty is not None:
                    _ci_have[_normalize_sku(str(_ci_sku).strip())] = int(float(_ci_qty))
            _ci_wb.close()
            if _ci_have:
                # HAVE is the sole authority: rebuild available from it; SKUs absent → 0.
                _stale = sorted(s for s in available if s not in _ci_have and available.get(s, 0) != 0)
                available = dict(_ci_have)
                print(f"  Overrode {len(_ci_have)} SKUs from corrected inventory (AUTHORITATIVE): "
                      f"{corrected_inv_path}")
                if _stale:
                    print(f"  Zeroed {len(_stale)} SKU(s) absent from HAVE (no stale base fallback): "
                          f"{', '.join(_stale[:20])}{' …' if len(_stale) > 20 else ''}")
            else:
                print(f"  Warning: corrected inventory had 0 usable rows, keeping base: {corrected_inv_path}")
        except Exception as e:
            print(f"  Warning: Could not load corrected inventory: {e}")

    # -- Recharge --
    print("Fetching Recharge charges...")
    # Specialty/bundle accumulator (AHB-X*, BL-*) — populated by both fetchers
    specialty: dict = {"WK1": {}, "WK2": {}}
    # Bundle explosion breakdown (per-box components) for the Cut Order tab.
    bundles: dict = {"WK1": {}, "WK2": {}}

    # Shopify creds for live bundle-recipe lookups from the Recharge fetcher.
    _bstore = settings.get("shopify_store_url", "").strip()
    if _bstore and not _bstore.startswith("http"):
        _bstore = f"https://{_bstore}.myshopify.com"
    _bundle_creds = (_bstore, settings.get("shopify_access_token", "").strip())

    (
        rc_wk1,
        rc_wk2,
        rc_wk1_curations,
        rc_wk2_curations,
        rc_wk1_large,
        rc_wk2_large,
        _,
        _,
        rc_wk1_med_monthly,
        rc_wk2_med_monthly,
        rc_wk1_cmed_monthly,
        rc_wk2_cmed_monthly,
        rc_wk1_lge_monthly,
        rc_wk2_lge_monthly,
        monthly_by_week_month,
        rc_wk1_custom,
        rc_wk2_custom,
    ) = fetch_recharge_api(recharge_token, out_specialty=specialty,
                           out_bundles=bundles, shop_creds=_bundle_creds)

    # -- Shopify --
    print("Fetching Shopify orders...")
    (
        sh_wk1_addon,
        sh_wk2_addon,
        sh_wk1_curations,
        sh_wk2_curations,
        sh_wk1_large,
        sh_wk2_large,
        sh_wk1_med,
        sh_wk2_med,
        sh_wk1_lge,
        sh_wk2_lge,
    ) = fetch_shopify_orders(settings, out_specialty=specialty, out_bundles=bundles)

    # -- Normalize aliased SKUs across all demand dicts --
    if SKU_ALIASES:
        def _apply_aliases(d: dict) -> dict:
            merged: dict = {}
            for k, v in d.items():
                canonical = _normalize_sku(k)
                merged[canonical] = merged.get(canonical, 0) + v
            return merged

        rc_wk1 = _apply_aliases(rc_wk1)
        rc_wk2 = _apply_aliases(rc_wk2)
        sh_wk1_addon = _apply_aliases(sh_wk1_addon)
        sh_wk2_addon = _apply_aliases(sh_wk2_addon)

    rc_wk1_total = sum(rc_wk1.values())
    rc_wk2_total = sum(rc_wk2.values())
    rc_wk1_curs = sum(rc_wk1_curations.values())
    rc_wk2_curs = sum(rc_wk2_curations.values())
    rc_wk1_monthly = rc_wk1_med_monthly + rc_wk1_cmed_monthly + rc_wk1_lge_monthly
    rc_wk2_monthly = rc_wk2_med_monthly + rc_wk2_cmed_monthly + rc_wk2_lge_monthly
    rc_wk1_orders = rc_wk1_curs + rc_wk1_monthly + rc_wk1_custom
    rc_wk2_orders = rc_wk2_curs + rc_wk2_monthly + rc_wk2_custom
    print(f"  Recharge WK1: {rc_wk1_orders} queued orders ({rc_wk1_curs} curation + {rc_wk1_monthly} monthly + {rc_wk1_custom} custom), {rc_wk1_total} pickable SKUs")
    print(f"  Recharge WK2: {rc_wk2_orders} queued orders ({rc_wk2_curs} curation + {rc_wk2_monthly} monthly + {rc_wk2_custom} custom), {rc_wk2_total} pickable SKUs")

    sh_wk1_total = sum(sh_wk1_addon.values())
    sh_wk2_total = sum(sh_wk2_addon.values())
    # Orders count already logged by fetch_shopify_orders. This line summarizes addon SKU volume only.
    print(f"  Shopify WK1 addon SKU volume: {sh_wk1_total}")
    print(f"  Shopify WK2 addon SKU volume: {sh_wk2_total}")

    # -- First-order projection (MONG) --
    import requests as _req

    store_url = settings.get("shopify_store_url", "")
    shop_token = settings.get("shopify_access_token", "")
    fo_count = 0
    if store_url and shop_token:
        if not store_url.startswith("http"):
            store_url = f"https://{store_url}.myshopify.com"
        # Monday mode: window = last completed Sat 00:00 → Sun 23:59, extrapolate 7 days to next ship Monday
        # Non-Monday: fallback to 3-day rolling window + days-to-Friday projection
        _is_monday = datetime.now().weekday() == 0
        if _is_monday:
            _today = datetime.now().date()
            _last_sun = _today - timedelta(days=1)
            _last_sat = _today - timedelta(days=2)
            _window_start = datetime.combine(_last_sat, datetime.min.time()).isoformat()
            _window_end = datetime.combine(_last_sun, datetime.max.time()).isoformat()
            _fo_params = {
                "status": "any",
                "limit": 250,
                "created_at_min": _window_start,
                "created_at_max": _window_end,
                "fields": "id,tags,line_items",
            }
        else:
            _cutoff = (datetime.now() - timedelta(days=3)).isoformat()
            _fo_params = {
                "status": "any",
                "limit": 250,
                "created_at_min": _cutoff,
                "fields": "id,tags,line_items",
            }
        _fo_url = f"{store_url}/admin/api/2026-04/orders.json"
        _fo_orders: list[dict] = []
        _fo_page_url: str | None = _fo_url
        while _fo_page_url:
            _fo_resp = _req.get(
                _fo_page_url,
                headers={"X-Shopify-Access-Token": shop_token, "Content-Type": "application/json"},
                params=_fo_params if _fo_page_url == _fo_url else None,
                timeout=30,
            )
            _fo_data = _fo_resp.json()
            for _o in _fo_data.get("orders", []):
                if "Subscription First Order" in (_o.get("tags") or ""):
                    _fo_orders.append(_o)
            _fo_page_url = None
            _link = _fo_resp.headers.get("Link", "")
            if 'rel="next"' in _link:
                _m = _re.search(r'<([^>]+)>;\s*rel="next"', _link)
                if _m:
                    _fo_page_url = _m.group(1)
            time.sleep(0.3)

        fo_count = len(_fo_orders)
        if _is_monday:
            _daily_rate = fo_count / 2.0 if _fo_orders else 0
            _days_to_project = 7  # Mon → next Mon ship week
        else:
            _daily_rate = fo_count / 3.0 if _fo_orders else 0
            _days_to_project = max(0, (5 - datetime.now().weekday()))
        _projected = int(_daily_rate * _days_to_project)

        _mong_fo = [o for o in _fo_orders if any("MONG" in (li.get("sku") or "") for li in o.get("line_items", []))]
        _fo_skus: dict[str, float] = defaultdict(float)
        if _mong_fo:
            for _o in _mong_fo:
                for _li in _o.get("line_items", []):
                    _sku = (_li.get("sku") or "").strip()
                    if _sku.startswith(PICKABLE_PREFIXES):
                        _fo_skus[_sku] += (_li.get("quantity", 1)) / len(_mong_fo)

        _mong_pct = len(_mong_fo) / fo_count if fo_count else 0
        _mong_projected = int(_projected * _mong_pct)
        _window_label = "last Sat+Sun (2d)" if _is_monday else "rolling 3d"
        print(
            f"  First-order projection: {fo_count} in {_window_label}, "
            f"{_daily_rate:.1f}/day × {_days_to_project}d = {_projected} projected, "
            f"{_mong_projected} MONG ({_mong_pct:.0%})"
        )
        # First-order projection DISABLED — was inflating sh_wk1_addon with
        # phantom demand for subscribers who hadn't ordered yet. Real demand
        # comes from already-charged Shopify orders + queued Recharge charges.
        # Projection log line above kept for visibility only.
        # (Was: distributed _mong_projected across observed first-order SKUs.)

    # -- Merge curation counts --
    wk1_curations: dict[str, int] = defaultdict(int)
    wk2_curations: dict[str, int] = defaultdict(int)
    wk1_large: dict[str, int] = defaultdict(int)
    wk2_large: dict[str, int] = defaultdict(int)
    for d_rc, d_sh, d_out in [
        (rc_wk1_curations, sh_wk1_curations, wk1_curations),
        (rc_wk2_curations, sh_wk2_curations, wk2_curations),
        (rc_wk1_large, sh_wk1_large, wk1_large),
        (rc_wk2_large, sh_wk2_large, wk2_large),
    ]:
        for k, v in d_rc.items():
            d_out[k] += v
        for k, v in d_sh.items():
            d_out[k] += v

    # Box size counts per curation
    wk1_med: dict[str, int] = defaultdict(int, sh_wk1_med)
    wk2_med: dict[str, int] = defaultdict(int, sh_wk2_med)
    wk1_lge: dict[str, int] = defaultdict(int, sh_wk1_lge)
    wk2_lge: dict[str, int] = defaultdict(int, sh_wk2_lge)
    for cur, ct in rc_wk1_curations.items():
        lg = rc_wk1_large.get(cur, 0)
        wk1_med[cur] += ct - lg
        wk1_lge[cur] += lg
    for cur, ct in rc_wk2_curations.items():
        lg = rc_wk2_large.get(cur, 0)
        wk2_med[cur] += ct - lg
        wk2_lge[cur] += lg

    # MONTHLY boxes
    wk1_med["MONTHLY"] = wk1_med.get("MONTHLY", 0) + rc_wk1_med_monthly
    wk2_med["MONTHLY"] = wk2_med.get("MONTHLY", 0) + rc_wk2_med_monthly
    wk1_med["CMED"] = wk1_med.get("CMED", 0) + rc_wk1_cmed_monthly
    wk2_med["CMED"] = wk2_med.get("CMED", 0) + rc_wk2_cmed_monthly
    wk1_lge["MONTHLY"] = wk1_lge.get("MONTHLY", 0) + rc_wk1_lge_monthly
    wk2_lge["MONTHLY"] = wk2_lge.get("MONTHLY", 0) + rc_wk2_lge_monthly

    # Active pickable SKUs. Also ALWAYS include any SKU carrying an ingested
    # spec (Cut / Notes / wheel / slice / first-order) — Tommy's notes like
    # "Bundle Remove", "IBRES", "Kurt investigate" must render even when the SKU
    # has 0 avail and 0 demand this week, or the annotation silently vanishes
    # (bit us 2026-07-28: CH-ALMA/CH-CCC/MT-BRAS/AC-WASP notes dropped).
    spec_skus = set(settings.get("cut_order_specs", {}).keys())
    report_skus: set[str] = set(spec_skus)
    for d in (available, rc_wk1, rc_wk2, sh_wk1_addon, sh_wk2_addon):
        report_skus.update(d.keys())
    active_skus = sorted(
        sku
        for sku in report_skus
        if any(sku.startswith(p) for p in PICKABLE_PREFIXES)
        and (
            sku in spec_skus
            or available.get(sku, 0) != 0
            or rc_wk1.get(sku, 0) > 0
            or rc_wk2.get(sku, 0) > 0
            or sh_wk1_addon.get(sku, 0) > 0
            or sh_wk2_addon.get(sku, 0) > 0
        )
    )

    all_curations = sorted(
        set(
            list(wk1_curations.keys())
            + list(wk2_curations.keys())
            + list(wk1_large.keys())
            + list(wk2_large.keys())
            + list(wk1_med.keys())
            + list(wk2_med.keys())
            + list(wk1_lge.keys())
            + list(wk2_lge.keys())
        )
    )

    # Single-cohort mode: merge WK2 demand into WK1, zero WK2.
    # Both ship tags collapse into one production batch (e.g. _SHIP_04-27 + _SHIP_05-04).

    # Inject Shopify MONTHLY/CMED/LGE counts into monthly_by_week_month
    # (otherwise the MONTHLY BOX COUNTS summary shows Recharge only).
    # Ship tag month = first 7 chars after "_SHIP_" (e.g. "_SHIP_2026-04-27" → "2026-04").
    from inventory_demand_report import WK1_SHIP_TAGS as _W1TS, WK2_SHIP_TAG as _W2T
    _W1T = _W1TS[-1]  # next-week tag (last in list)
    _wk1_month = _W1T.replace("_SHIP_", "")[:7]
    _wk2_month = _W2T.replace("_SHIP_", "")[:7]

    def _add_monthly(week_key: str, month: str, box_type: str, qty: int) -> None:
        if qty <= 0:
            return
        key = (week_key, month)
        bucket = monthly_by_week_month.setdefault(key, {})
        bucket[box_type] = bucket.get(box_type, 0) + qty

    _add_monthly("WK1", _wk1_month, "MED", sh_wk1_med.get("MONTHLY", 0))
    _add_monthly("WK1", _wk1_month, "CMED", sh_wk1_med.get("CMED", 0))
    _add_monthly("WK1", _wk1_month, "LGE", sh_wk1_lge.get("MONTHLY", 0))
    _add_monthly("WK2", _wk2_month, "MED", sh_wk2_med.get("MONTHLY", 0))
    _add_monthly("WK2", _wk2_month, "CMED", sh_wk2_med.get("CMED", 0))
    _add_monthly("WK2", _wk2_month, "LGE", sh_wk2_lge.get("MONTHLY", 0))

    # WK2 dropped entirely (per user 2026-05-05 — "drop week 2 for now").
    # WK1 cohort only. If Monday run, both this-Mon and next-Mon ship tags
    # are already in WK1_SHIP_TAGS, so they're captured by sh_wk1/rc_wk1.
    monthly_by_week_month = {
        k: v for k, v in monthly_by_week_month.items() if k[0] == "WK1"
    }
    rc_wk2 = {}
    sh_wk2_addon = {}
    wk2_curations = defaultdict(int)
    wk2_large = defaultdict(int)
    wk2_med = defaultdict(int)
    for cur, ct in wk2_lge.items():
        wk1_lge[cur] += ct
    wk2_lge = defaultdict(int)

    return {
        "available": available,
        "rc_wk1": rc_wk1,
        "rc_wk2": rc_wk2,
        "sh_wk1_addon": sh_wk1_addon,
        "sh_wk2_addon": sh_wk2_addon,
        "wk1_curations": dict(wk1_curations),
        "wk2_curations": dict(wk2_curations),
        "wk1_large": dict(wk1_large),
        "wk2_large": dict(wk2_large),
        "wk1_med": dict(wk1_med),
        "wk2_med": dict(wk2_med),
        "wk1_lge": dict(wk1_lge),
        "wk2_lge": dict(wk2_lge),
        "monthly_by_week_month": monthly_by_week_month,
        "active_skus": active_skus,
        "all_curations": all_curations,
        "sku_name": sku_name,
        "bulk_weights": bulk_weights,
        "csv_rows": csv_rows,
        "rc_wk1_curs": rc_wk1_curs,
        "rc_wk2_curs": rc_wk2_curs,
        "sh_wk1_total": sum(sh_wk1_curations.values()),
        "sh_wk2_total": sum(sh_wk2_curations.values()),
        "specialty": specialty,
        "bundles": bundles,
    }


# ── Tab 2: Assignments ──────────────────────────────────────────────

# Slot definitions for MONTHLY boxes
AHB_MED_SLOTS = [
    ("Cheese 1", "CH-"),
    ("Cheese 2", "CH-"),
    ("Meat 1", "MT-"),
    ("Meat 2", "MT-"),
    ("Crackers", "AC-"),
    ("Accompaniment 1", "AC-"),
    ("Accompaniment 2", "AC-"),
    ("PR-CJAM-GEN Cheese", "CH-"),
    ("PR-CJAM-GEN Jam", "AC-"),
]
AHB_CMED_SLOTS = [
    ("Cheese 1", "CH-"),
    ("Cheese 2", "CH-"),
    ("Cheese 3", "CH-"),
    ("Cheese 4", "CH-"),
    ("Crackers", "AC-"),
    ("Accompaniment 1", "AC-"),
    ("Accompaniment 2", "AC-"),
    ("PR-CJAM-GEN Cheese", "CH-"),
    ("PR-CJAM-GEN Jam", "AC-"),
]
AHB_LGE_SLOTS = [
    ("Cheese 1", "CH-"),
    ("Cheese 2", "CH-"),
    ("Cheese 3", "CH-"),
    ("Meat 1", "MT-"),
    ("Meat 2", "MT-"),
    ("Meat 3", "MT-"),
    ("Crackers", "AC-"),
    ("Accompaniment 1", "AC-"),
    ("Accompaniment 2", "AC-"),
    ("PR-CJAM-GEN Cheese", "CH-"),
    ("PR-CJAM-GEN Jam", "AC-"),
]


def _write_assignments_on_cut_order(ws: Worksheet, data: dict, settings: dict) -> int:
    """Write assignment tables onto the Cut Order tab starting at column W (23).

    Layout:
      V (22) = spacer
      W-Z (23-26) = PR-CJAM table (Curation, Cheese SKU, W1 Count, W2 Count)
      AA (27) = spacer
      AB-AE (28-31) = CEX-EC table (Curation, Cheese SKU, W1 Lg Count, W2 Lg Count)
      MONTHLY slot tables below PR-CJAM/CEX-EC (whichever is taller), in W-Z area.

    PR-CJAM cheese SKU col = X (24), counts = Y (25), Z (26)
    CEX-EC cheese SKU col  = AC (29), counts = AD (30), AE (31)

    MONTHLY slot SKU col = X (24), counts = Y (25), Z (26) — contiguous with PR-CJAM.

    Returns the last row of the contiguous SKU range (PR-CJAM + CEX-EC + MONTHLY slots)
    used for SUMIF lookups. The SUMIF range covers col X for SKU, Y for W1, Z for W2
    across PR-CJAM rows AND MONTHLY slot rows (one contiguous block).
    CEX-EC is separate (col AC for SKU, AD for W1, AE for W2).
    """
    pr_cjam_cfg = settings.get("pr_cjam", {})
    cex_ec_cfg = settings.get("cex_ec", {})
    all_curations = data["all_curations"]
    wk1_curations = data["wk1_curations"]
    wk2_curations = data["wk2_curations"]
    wk1_large = data["wk1_large"]
    wk2_large = data["wk2_large"]
    monthly_by_week_month = data["monthly_by_week_month"]

    # Column constants
    COL_V = 22  # spacer
    COL_W = 23  # PR-CJAM: Curation
    COL_X = 24  # PR-CJAM: Cheese SKU (SUMIF lookup)
    COL_Y = 25  # PR-CJAM: W1 Count
    COL_Z = 26  # PR-CJAM: W2 Count
    COL_AA = 27  # spacer
    COL_AB = 28  # CEX-EC: Curation
    COL_AC = 29  # CEX-EC: Cheese SKU (SUMIF lookup)
    COL_AD = 30  # CEX-EC: W1 Lg Count
    COL_AE = 31  # CEX-EC: W2 Lg Count

    # Column widths for assignment area
    _set_col_widths(
        ws,
        {
            COL_V: 2,
            COL_W: 14,
            COL_X: 18,
            COL_Y: 10,
            COL_Z: 10,
            COL_AA: 2,
            COL_AB: 14,
            COL_AC: 18,
            COL_AD: 10,
            COL_AE: 10,
        },
    )

    # MONTHLY and CMED use PR-CJAM-GEN (handled by MONTHLY slot tables below),
    # so exclude them from the per-curation PR-CJAM and CEX-EC assignment rows.
    PRCJAM_GEN_CURATIONS = {"MONTHLY", "CMED"}
    prcjam_curations = [c for c in all_curations if c not in PRCJAM_GEN_CURATIONS]

    # ── PR-CJAM Section (W-Z, starting row 1) ──
    _dark_header_row(ws, 1, ["PR-CJAM ASSIGNMENTS", "", "", ""], col_start=COL_W)
    _dark_header_row(ws, 2, ["Curation", "Cheese SKU", "W1 Count", "W2 Count"], col_start=COL_W)

    prcjam_data_start = 3
    for i, cur in enumerate(prcjam_curations):
        r = prcjam_data_start + i
        cheese = ""
        cfg = pr_cjam_cfg.get(cur)
        if isinstance(cfg, dict):
            cheese = cfg.get("cheese", "")
        ws.cell(row=r, column=COL_W, value=cur).font = F_NAME
        c_cheese = ws.cell(row=r, column=COL_X, value=cheese)
        c_cheese.font = F_EDIT
        c_cheese.fill = FILL_INPUT
        ws.cell(row=r, column=COL_Y, value=wk1_curations.get(cur, 0)).font = F_NUM
        ws.cell(row=r, column=COL_Y).alignment = A_RIGHT
        ws.cell(row=r, column=COL_Z, value=wk2_curations.get(cur, 0)).font = F_NUM
        ws.cell(row=r, column=COL_Z).alignment = A_RIGHT

    prcjam_end = prcjam_data_start + len(prcjam_curations) - 1

    # ── CEX-EC Section (AB-AE, starting row 1) ──
    _dark_header_row(ws, 1, ["CEX-EC ASSIGNMENTS", "", "", ""], col_start=COL_AB)
    _dark_header_row(ws, 2, ["Curation", "Cheese SKU", "W1 CEX Count", "W2 CEX Count"], col_start=COL_AB)

    cexec_data_start = 3
    for i, cur in enumerate(prcjam_curations):
        r = cexec_data_start + i
        cheese = cex_ec_cfg.get(cur, "")
        ws.cell(row=r, column=COL_AB, value=cur).font = F_NAME
        c_cheese = ws.cell(row=r, column=COL_AC, value=cheese)
        c_cheese.font = F_EDIT
        c_cheese.fill = FILL_INPUT
        ws.cell(row=r, column=COL_AD, value=wk1_large.get(cur, 0)).font = F_NUM
        ws.cell(row=r, column=COL_AD).alignment = A_RIGHT
        ws.cell(row=r, column=COL_AE, value=wk2_large.get(cur, 0)).font = F_NUM
        ws.cell(row=r, column=COL_AE).alignment = A_RIGHT

    cexec_end = cexec_data_start + len(prcjam_curations) - 1

    # ── PR-CJAM JAM Section (AG-AJ) ──
    # PR-CJAM ships as a cheese + jam PAIR, so the jam count == the cheese count
    # per curation. The cheese block (X) already feeds +Assign; this parallel
    # block feeds the jam SKU with the SAME W1/W2 counts (Kurt 2026-07-28).
    COL_AF = 32  # spacer
    COL_AG = 33  # JAM: Curation
    COL_AH = 34  # JAM: Jam SKU (SUMIF lookup)
    COL_AI = 35  # JAM: W1 Count
    COL_AJ = 36  # JAM: W2 Count
    _set_col_widths(ws, {COL_AF: 2, COL_AG: 14, COL_AH: 18, COL_AI: 10, COL_AJ: 10})
    _dark_header_row(ws, 1, ["PR-CJAM JAM ASSIGNMENTS", "", "", ""], col_start=COL_AG)
    _dark_header_row(ws, 2, ["Curation", "Jam SKU", "W1 Count", "W2 Count"], col_start=COL_AG)
    jam_data_start = 3
    for i, cur in enumerate(prcjam_curations):
        r = jam_data_start + i
        cfg = pr_cjam_cfg.get(cur)
        jam = cfg.get("jam", "") if isinstance(cfg, dict) else ""
        ws.cell(row=r, column=COL_AG, value=cur).font = F_NAME
        c_jam = ws.cell(row=r, column=COL_AH, value=jam)
        c_jam.font = F_EDIT
        c_jam.fill = FILL_INPUT
        ws.cell(row=r, column=COL_AI, value=wk1_curations.get(cur, 0)).font = F_NUM
        ws.cell(row=r, column=COL_AI).alignment = A_RIGHT
        ws.cell(row=r, column=COL_AJ, value=wk2_curations.get(cur, 0)).font = F_NUM
        ws.cell(row=r, column=COL_AJ).alignment = A_RIGHT
    jam_end = jam_data_start + len(prcjam_curations) - 1

    # ── MONTHLY Box Slot Tables (below PR-CJAM in W-Z columns) ──
    # These go right after PR-CJAM rows so the SKU column (X) is contiguous
    # with PR-CJAM cheese SKUs for a single SUMIF range.
    month_counts: dict[str, dict[str, dict[str, int]]] = {}
    for (week, month), counts in monthly_by_week_month.items():
        if month not in month_counts:
            month_counts[month] = {
                "MED": {"wk1": 0, "wk2": 0},
                "CMED": {"wk1": 0, "wk2": 0},
                "LGE": {"wk1": 0, "wk2": 0},
            }
        wk_key = "wk1" if week == "WK1" else "wk2"
        for box_type in ("MED", "CMED", "LGE"):
            month_counts[month][box_type][wk_key] += counts.get(box_type, 0)

    # Start MONTHLY slots right after PR-CJAM data (no gap — contiguous for SUMIF)
    slot_row = prcjam_end + 1

    def _write_slot_table(
        ws_: Worksheet,
        start_row: int,
        label: str,
        slots: list[tuple[str, str]],
        w1_count: int,
        w2_count: int,
        col_start: int,
        slot_skus: dict | None = None,
    ) -> int:
        slot_skus = slot_skus or {}
        ws_.merge_cells(
            start_row=start_row,
            start_column=col_start,
            end_row=start_row,
            end_column=col_start + 3,
        )
        c = ws_.cell(row=start_row, column=col_start, value=f"{label} ({w1_count} W1 / {w2_count} W2)")
        c.font = Font(name="Calibri", size=10, bold=True, color=OK_FG)
        c.fill = PatternFill("solid", fgColor=OK_BG)
        for ci in range(col_start + 1, col_start + 4):
            ws_.cell(row=start_row, column=ci).fill = PatternFill("solid", fgColor=OK_BG)
        r = start_row
        for slot_name, _prefix in slots:
            r += 1
            ws_.cell(row=r, column=col_start, value=slot_name).font = F_NAME
            c_sku = ws_.cell(row=r, column=col_start + 1, value=slot_skus.get(slot_name) or None)
            c_sku.font = F_EDIT
            c_sku.fill = FILL_INPUT
            ws_.cell(row=r, column=col_start + 2, value=w1_count).font = F_NUM
            ws_.cell(row=r, column=col_start + 2).alignment = A_RIGHT
            ws_.cell(row=r, column=col_start + 3, value=w2_count).font = F_NUM
            ws_.cell(row=r, column=col_start + 3).alignment = A_RIGHT
        return r

    # Persisted monthly slot SKUs (snapshotted from Tommy via --ingest). Keyed
    # month → box-type label → {slot_name: sku}. Pre-fills the slot cells so
    # they survive regeneration AND the SUMIF counts the monthly components.
    mbr = settings.get("monthly_box_recipes", {})

    def _slot_skus_for(month, label_prefix):
        rows = mbr.get(month, {}).get(label_prefix, [])
        return {r[0]: r[1] for r in rows if len(r) >= 2 and r[1]}

    for month in sorted(month_counts.keys()):
        mc = month_counts[month]
        for box_type, slots, label_prefix in [
            ("MED", AHB_MED_SLOTS, "AHB-MED"),
            ("CMED", AHB_CMED_SLOTS, "AHB-CMED"),
            ("LGE", AHB_LGE_SLOTS, "AHB-LGE"),
        ]:
            if mc[box_type]["wk1"] + mc[box_type]["wk2"] > 0:
                slot_row = _write_slot_table(
                    ws,
                    slot_row,
                    f"{label_prefix} ({month})",
                    slots,
                    mc[box_type]["wk1"],
                    mc[box_type]["wk2"],
                    col_start=COL_W,
                    slot_skus=_slot_skus_for(month, label_prefix),
                )
                slot_row += 1  # next table starts on the row after last slot row
                # Note: no extra gap — keep contiguous for SUMIF

    # The last data row in the W-Z contiguous block (PR-CJAM + MONTHLY slots)
    prcjam_monthly_last_row = max(slot_row - 1, prcjam_end)

    # ── MONTHLY Box Counts summary (below the taller of the two table groups) ──
    summary_start = max(prcjam_monthly_last_row, cexec_end) + 3
    if monthly_by_week_month:
        _dark_header_row(ws, summary_start, ["MONTHLY BOX COUNTS", "", "", ""], col_start=COL_W)
        summary_start += 1
        _dark_header_row(ws, summary_start, ["Box Type", "Month", "W1 Count", "W2 Count"], col_start=COL_W)
        summary_start += 1
        for month in sorted(month_counts.keys()):
            mc = month_counts[month]
            for bt in ("MED", "CMED", "LGE"):
                if mc[bt]["wk1"] + mc[bt]["wk2"] > 0:
                    ws.cell(row=summary_start, column=COL_W, value=f"AHB-{bt}").font = F_NAME
                    ws.cell(row=summary_start, column=COL_X, value=month).font = F_NAME
                    ws.cell(row=summary_start, column=COL_Y, value=mc[bt]["wk1"]).font = F_NUM
                    ws.cell(row=summary_start, column=COL_Y).alignment = A_RIGHT
                    ws.cell(row=summary_start, column=COL_Z, value=mc[bt]["wk2"]).font = F_NUM
                    ws.cell(row=summary_start, column=COL_Z).alignment = A_RIGHT
                    summary_start += 1

    # Return info needed for SUMIF references:
    # prcjam_monthly_last_row = last row of contiguous PR-CJAM + MONTHLY slot SKU data in col X
    # cexec_end = last row of CEX-EC data in col AC
    # jam_end = last row of PR-CJAM JAM data in col AH
    return prcjam_monthly_last_row, cexec_end, jam_end


# ── Tab 3: Raw Materials ─────────────────────────────────────────────


def _build_raw_materials_tab(wb: openpyxl.Workbook, data: dict, settings: dict) -> None:
    ws = wb.create_sheet("Raw Materials")
    bulk_weights = data["bulk_weights"]
    csv_rows = data["csv_rows"]
    sku_name_fn = data["sku_name"]
    bulk_conversions = settings.get("bulk_conversions", {})

    _set_col_widths(ws, {1: 28, 2: 14, 3: 14, 4: 14, 5: 16, 6: 10})

    # ── Cheese Wheels Section ──
    row = 1
    _dark_header_row(ws, row, ["CHEESE WHEELS & BLOCKS", "", "", "", "", ""])
    row = 2
    _dark_header_row(ws, row, ["Cheese Name", "SKU", "On Hand", "Weight (lbs)", "Potential Slices", "Status"])

    # Sort by potential yield descending
    wheel_items = sorted(bulk_weights.items(), key=lambda x: x[1].get("potential_yield", 0), reverse=True)
    row = 3
    for sku, wt in wheel_items:
        name = sku_name_fn(sku) or sku
        count = wt.get("count", 0)
        weight_lbs = wt.get("weight_lbs", 0)
        potential = wt.get("potential_yield", 0)

        if potential > 200:
            status = "HIGH"
            status_fill = PatternFill("solid", fgColor=OK_BG)
            status_font = Font(name="Calibri", size=9, bold=True, color=OK_FG)
        elif potential >= 50:
            status = "MED"
            status_fill = PatternFill("solid", fgColor=TIGHT_BG)
            status_font = Font(name="Calibri", size=9, bold=True, color=TIGHT_FG)
        elif potential > 0:
            status = "LOW"
            status_fill = PatternFill("solid", fgColor=SHORTAGE_BG)
            status_font = Font(name="Calibri", size=9, bold=True, color=SHORTAGE_FG)
        else:
            status = "EMPTY"
            status_fill = PatternFill("solid", fgColor="F1F5F9")
            status_font = Font(name="Calibri", size=9, bold=True, color=MUTED)

        ws.cell(row=row, column=1, value=name).font = F_NAME
        ws.cell(row=row, column=2, value=sku).font = F_SKU
        c_count = ws.cell(row=row, column=3, value=count)
        c_count.font = F_NUM
        c_count.alignment = A_RIGHT
        c_wt = ws.cell(row=row, column=4, value=round(weight_lbs * count, 1) if weight_lbs else 0)
        c_wt.font = F_NUM
        c_wt.alignment = A_RIGHT
        c_pot = ws.cell(row=row, column=5, value=potential)
        c_pot.font = Font(name="Calibri", size=12, bold=True, color=WHEEL_POT)
        c_pot.alignment = A_RIGHT
        c_st = ws.cell(row=row, column=6, value=status)
        c_st.font = status_font
        c_st.fill = status_fill
        c_st.alignment = A_CENTER
        row += 1

    if not wheel_items:
        ws.cell(row=row, column=1, value="No cheese wheel/block data found").font = F_NUM_MUTED
        row += 1

    # ── Bulk Accompaniments Section ──
    row += 2
    _dark_header_row(ws, row, ["BULK ACCOMPANIMENTS", "", "", "", "", ""])
    row += 1
    _dark_header_row(ws, row, ["Ingredient", "SKU", "On Hand (oz)", "Packet Size (oz)", "Potential Packets", "Status"])
    row += 1

    for ingredient_name, conv in sorted(bulk_conversions.items()):
        target_sku = conv.get("sku", "")
        packet_oz = conv.get("packet_oz", 1)

        # Scan CSV rows for this ingredient
        total_on_hand = 0.0
        for csv_row in csv_rows:
            ing = csv_row.get("Ingredient", "").strip()
            if ingredient_name.lower() in ing.lower():
                try:
                    raw_total = float(csv_row.get("Total", 0) or 0)
                except (ValueError, TypeError):
                    raw_total = 0
                # Check unit — convert lbs to oz if needed
                unit = csv_row.get("Unit1", "").strip().lower()
                if "lb" in unit:
                    total_on_hand += raw_total * 16.0
                elif "kg" in unit:
                    total_on_hand += raw_total * 35.274
                else:
                    total_on_hand += raw_total  # assume oz

        potential_packets = int(total_on_hand / packet_oz) if packet_oz > 0 else 0

        if potential_packets > 200:
            status = "HIGH"
            status_fill = PatternFill("solid", fgColor=OK_BG)
            status_font = Font(name="Calibri", size=9, bold=True, color=OK_FG)
        elif potential_packets >= 50:
            status = "MED"
            status_fill = PatternFill("solid", fgColor=TIGHT_BG)
            status_font = Font(name="Calibri", size=9, bold=True, color=TIGHT_FG)
        elif potential_packets > 0:
            status = "LOW"
            status_fill = PatternFill("solid", fgColor=SHORTAGE_BG)
            status_font = Font(name="Calibri", size=9, bold=True, color=SHORTAGE_FG)
        else:
            status = "EMPTY"
            status_fill = PatternFill("solid", fgColor="F1F5F9")
            status_font = Font(name="Calibri", size=9, bold=True, color=MUTED)

        ws.cell(row=row, column=1, value=ingredient_name).font = F_NAME
        ws.cell(row=row, column=2, value=target_sku).font = F_SKU
        c_oh = ws.cell(row=row, column=3, value=round(total_on_hand, 1))
        c_oh.font = F_NUM
        c_oh.alignment = A_RIGHT
        c_ps = ws.cell(row=row, column=4, value=packet_oz)
        c_ps.font = F_NUM
        c_ps.alignment = A_RIGHT
        c_pp = ws.cell(row=row, column=5, value=potential_packets)
        c_pp.font = Font(name="Calibri", size=12, bold=True, color=WHEEL_POT)
        c_pp.alignment = A_RIGHT
        c_st = ws.cell(row=row, column=6, value=status)
        c_st.font = status_font
        c_st.fill = status_fill
        c_st.alignment = A_CENTER
        row += 1

    if not bulk_conversions:
        ws.cell(row=row, column=1, value="No bulk conversions configured in settings").font = F_NUM_MUTED

    ws.freeze_panes = "A3"
    ws.sheet_properties.tabColor = "7C3AED"


# ── Tab: Checklist ──────────────────────────────────────────────────


def _build_checklist_tab(wb: openpyxl.Workbook, data: dict, settings: dict) -> None:
    """General pre-cut sanity checks. Section 1: bundles & specialty (AHB-X*, BL-*).

    Add new sections below by following the same _section_header + table pattern.
    """
    ws = wb.create_sheet("Checklist")
    sku_name_fn = data["sku_name"]
    specialty = data.get("specialty") or {"WK1": {}, "WK2": {}}

    _set_col_widths(ws, {1: 28, 2: 38, 3: 12, 4: 12, 5: 14})

    row = 1
    _merge_title_bar(ws, row, "PRE-CUT CHECKLIST", 5)
    row += 1
    _merge_subtitle(ws, row, "General sanity checks before cutting", 5)
    row += 2

    # ── Section 1: Bundles & Specialty (AHB-X*, BL-*) ──
    _section_header(ws, row, "Bundles & Specialty (AHB-X*, BL-*)", "1E293B", "FFFFFF", 5)
    row += 1
    _dark_header_row(ws, row, ["SKU", "Name", "WK1 Qty", "WK2 Qty", "Total"])
    row += 1

    wk1 = specialty.get("WK1", {})
    wk2 = specialty.get("WK2", {})
    all_skus = sorted(set(wk1) | set(wk2))

    if not all_skus:
        ws.cell(row=row, column=1, value="No AHB-X* or BL-* SKUs in queued charges or open orders").font = F_NUM_MUTED
        row += 1
    else:
        for sku in all_skus:
            q1 = int(wk1.get(sku, 0))
            q2 = int(wk2.get(sku, 0))
            ws.cell(row=row, column=1, value=sku).font = F_SKU
            ws.cell(row=row, column=2, value=sku_name_fn(sku) or "").font = F_NAME
            c1 = ws.cell(row=row, column=3, value=q1)
            c1.font = F_NUM
            c1.alignment = A_RIGHT
            c2 = ws.cell(row=row, column=4, value=q2)
            c2.font = F_NUM
            c2.alignment = A_RIGHT
            ct = ws.cell(row=row, column=5, value=q1 + q2)
            ct.font = Font(name="Calibri", size=12, bold=True)
            ct.alignment = A_RIGHT
            row += 1

    ws.freeze_panes = "A5"
    ws.sheet_properties.tabColor = "F59E0B"


# ── Tab 1: Cut Order (main sheet) ───────────────────────────────────


def _build_cut_order_tab(
    wb: openpyxl.Workbook,
    data: dict,
    settings: dict,
    prcjam_monthly_last_row: int,
    cexec_last_row: int,
    jam_last_row: int,
) -> None:
    """Build the main Cut Order tab with urgency-grouped rows and inline assignments."""
    ws = wb.active
    ws.title = "Cut Order"
    ws.sheet_properties.tabColor = "1E293B"

    inv_settings = settings.get("inventory", {})
    pr_cjam_cfg = settings.get("pr_cjam", {})
    cex_ec_cfg = settings.get("cex_ec", {})
    # Per-SKU inputs Tommy filled last time, snapshotted via --ingest. Pre-fills
    # First Order (F), Wheel lb (M), Slice oz (N) so he doesn't re-enter the same
    # values every week. Same value → same SKU. Still input-styled → editable.
    cut_specs = settings.get("cut_order_specs", {})

    available = data["available"]
    rc_wk1 = data["rc_wk1"]
    rc_wk2 = data["rc_wk2"]
    sh_wk1_addon = data["sh_wk1_addon"]
    sh_wk2_addon = data["sh_wk2_addon"]
    active_skus = data["active_skus"]
    sku_name_fn = data["sku_name"]
    bulk_weights = data["bulk_weights"]

    # Column widths: A-U
    _set_col_widths(
        ws,
        {
            1: 14,  # A SKU
            2: 30,  # B Name
            3: 8,   # C Avail
            4: 10,  # D Recharge
            5: 10,  # E Shopify
            6: 10,  # F First Order
            7: 8,   # G +Assign
            8: 1,   # H spacer
            9: 10,  # I Final Demand
            10: 10, # J After
            11: 9,  # K Cut (slices)
            12: 9,  # L Good?
            13: 9,  # M Wheel lb
            14: 9,  # N Slice oz
            15: 9,  # O Wheels
            16: 8,  # P Safety (hidden)
            17: 34, # Q Notes (input)
        },
    )
    LAST_COL = 15
    NOTES_COL = 17
    ws.column_dimensions["P"].hidden = True

    # ── Title bar ──
    _merge_title_bar(ws, 1, f"CUT ORDER \u2014 Ship Week of {SHIP_WEEK_MONDAY}", NOTES_COL)

    rc_total = data["rc_wk1_curs"]
    sh_total = data["sh_wk1_total"]
    gen_time = datetime.now().strftime("%Y-%m-%d %H:%M")
    _merge_subtitle(ws, 2, f"Generated {gen_time}  |  RC: {rc_total} charges  |  SH: {sh_total} orders", NOTES_COL)

    # Row 3: blank
    # Row 4: column headers
    HDR_ROW = 4
    headers = [
        "SKU",
        "Name",
        "Avail",
        "Recharge",
        "Shopify",
        "First Order",
        "+Assign",
        "",  # spacer
        "Final Demand",
        "After",
        "Cut",
        "Good?",
        "Wheel lb",
        "Slice oz",
        "Wheels",
        "Safety",
    ]
    _dark_header_row(ws, HDR_ROW, headers)
    # Notes header (col Q/17) — written separately since it sits past the hidden
    # Safety column (P/16). Free-text per-SKU, persisted via cut_order_specs.
    _nh = ws.cell(HDR_ROW, NOTES_COL, "Notes")
    _nh.font = Font(name="Calibri", size=10, bold=True, color=HEADER_FG)
    _nh.fill = PatternFill("solid", fgColor=HEADER_BG)

    # SUMIF references — same sheet, two ranges:
    # 1) PR-CJAM + MONTHLY slots: col X (SKU), Y (W1 count), Z (W2 count) rows 3..prcjam_monthly_last_row
    # 2) CEX-EC: col AC (SKU), AD (W1 lg count), AE (W2 lg count) rows 3..cexec_last_row
    # +Assign = SUMIF(pr_cjam_monthly_sku, A{row}, pr_cjam_monthly_w) + SUMIF(cexec_sku, A{row}, cexec_w)
    prcjam_sku_range = f"$X$3:$X${prcjam_monthly_last_row}"
    prcjam_w1_range = f"$Y$3:$Y${prcjam_monthly_last_row}"
    prcjam_w2_range = f"$Z$3:$Z${prcjam_monthly_last_row}"
    cexec_sku_range = f"$AC$3:$AC${cexec_last_row}"
    cexec_w1_range = f"$AD$3:$AD${cexec_last_row}"
    cexec_w2_range = f"$AE$3:$AE${cexec_last_row}"
    # PR-CJAM jam pair (col AH sku, AI/AJ counts) — same counts as the cheese.
    jam_sku_range = f"$AH$3:$AH${jam_last_row}"
    jam_w1_range = f"$AI$3:$AI${jam_last_row}"
    jam_w2_range = f"$AJ$3:$AJ${jam_last_row}"

    # ── Pre-compute urgency for each SKU ──
    # We need to calculate actual numeric values for sorting, not just formulas
    # For sorting: estimate assignment demand from settings (won't match SUMIF exactly but close enough)
    def _estimate_assign_demand(sku: str, curations: dict, large: dict, pr_cfg: dict, cex_cfg: dict) -> int:
        total = 0
        for cur, ct in curations.items():
            cfg = pr_cfg.get(cur)
            if isinstance(cfg, dict):
                if cfg.get("cheese") == sku:
                    total += ct
                if cfg.get("jam") == sku:  # PR-CJAM pair — jam same count as cheese
                    total += ct
        for cur, ct in large.items():
            if cex_cfg.get(cur) == sku:
                total += ct
        return total

    wk1_curs = data["wk1_curations"]
    wk2_curs = data["wk2_curations"]
    wk1_lg = data["wk1_large"]
    wk2_lg = data["wk2_large"]

    sku_rows: list[dict] = []
    for sku in active_skus:
        avail = available.get(sku, 0)

        rc1 = rc_wk1.get(sku, 0)
        sh1 = sh_wk1_addon.get(sku, 0)
        assign1 = _estimate_assign_demand(sku, wk1_curs, wk1_lg, pr_cjam_cfg, cex_ec_cfg)
        # Demand = Recharge (queued, not yet charged) + Shopify (already charged).
        # Recharge charges convert to Shopify orders on charge date — no overlap.
        demand1 = rc1 + sh1 + assign1

        rc2 = rc_wk2.get(sku, 0)
        sh2 = sh_wk2_addon.get(sku, 0)
        assign2 = _estimate_assign_demand(sku, wk2_curs, wk2_lg, pr_cjam_cfg, cex_ec_cfg)
        demand2 = rc2 + sh2 + assign2

        after1 = avail - demand1
        after2 = after1 - demand2  # after W1, no cut yet

        cat = _sku_category(sku)

        # Urgency classification
        if after1 < 0:
            urgency = "SHORTAGE"
        elif after1 - demand2 < 0:
            urgency = "TIGHT"
        else:
            urgency = "OK"

        sku_rows.append(
            {
                "sku": sku,
                "name": sku_name_fn(sku),
                "avail": avail,
                "rc1": rc1,
                "sh1": sh1,
                "rc2": rc2,
                "sh2": sh2,
                "demand1": demand1,
                "demand2": demand2,
                "after1": after1,
                "after2": after2,
                "cat": cat,
                "urgency": urgency,
                "sort_ratio": (after1 / demand1) if demand1 > 0 else 9999,
            }
        )

    # Category sort order
    CAT_ORDER = {"CHEESE": 0, "MEAT": 1, "ACCOMPANIMENTS": 2, "OTHER": 3}

    # Single combined list — no urgency grouping. Sort by category then SKU alpha.
    all_rows = list(sku_rows)
    all_rows.sort(key=lambda r: (CAT_ORDER.get(r["cat"], 9), r["sku"]))

    # ── Write rows ──
    row = HDR_ROW  # will increment before writing

    def _write_sku_row(ws_: Worksheet, row_num: int, sr: dict) -> None:
        sku = sr["sku"]
        # A: SKU
        ws_.cell(row=row_num, column=1, value=sku).font = F_SKU
        # B: Name
        ws_.cell(row=row_num, column=2, value=sr["name"]).font = F_NAME
        # C: Avail
        c_avail = ws_.cell(row=row_num, column=3, value=sr["avail"])
        c_avail.font = F_NUM
        c_avail.alignment = A_RIGHT
        # D: Recharge (queued charges, not yet charged)
        c_rc = ws_.cell(row=row_num, column=4, value=sr["rc1"])
        c_rc.font = F_NUM
        c_rc.alignment = A_RIGHT
        # E: Shopify (already-charged orders) — no overlap with Recharge by definition
        c_sh = ws_.cell(row=row_num, column=5, value=sr["sh1"])
        c_sh.font = F_NUM
        c_sh.alignment = A_RIGHT
        # F: First Order (input — projected new-subscriber first-order demand).
        # Pre-filled from the last snapshot (--ingest); still editable.
        _spec = cut_specs.get(sku, {})
        ws_.cell(row=row_num, column=6, value=_spec.get("first_order") or None).font = F_INPUT
        ws_.cell(row=row_num, column=6).fill = FILL_INPUT
        ws_.cell(row=row_num, column=6).alignment = A_RIGHT
        # G: +Assign W1 = SUMIF(PR-CJAM+MONTHLY) + SUMIF(CEX-EC) + SUMIF(PR-CJAM JAM)
        ws_[f"G{row_num}"] = (
            f"=SUMIF({prcjam_sku_range},A{row_num},{prcjam_w1_range})"
            f"+SUMIF({cexec_sku_range},A{row_num},{cexec_w1_range})"
            f"+SUMIF({jam_sku_range},A{row_num},{jam_w1_range})"
        )
        ws_.cell(row=row_num, column=7).font = F_NUM
        ws_.cell(row=row_num, column=7).alignment = A_RIGHT
        # H: spacer
        # I: Final Demand = Recharge (D) + Shopify (E) + First Order (F) + Assign (G) + Safety (P)
        ws_[f"I{row_num}"] = f"=D{row_num}+E{row_num}+F{row_num}+G{row_num}+P{row_num}"
        ws_.cell(row=row_num, column=9).font = F_NUM_BOLD
        ws_.cell(row=row_num, column=9).alignment = A_RIGHT
        # J: After = Avail (C) - Final Demand (I)
        ws_[f"J{row_num}"] = f"=C{row_num}-I{row_num}"
        ws_.cell(row=row_num, column=10).font = F_NUM_BOLD
        ws_.cell(row=row_num, column=10).alignment = A_RIGHT
        # K: Cut (input — number of slices to cut). Pre-filled from snapshot.
        ws_.cell(row=row_num, column=11, value=_spec.get("cut") or None).font = F_INPUT
        ws_.cell(row=row_num, column=11).fill = FILL_INPUT
        ws_.cell(row=row_num, column=11).alignment = A_RIGHT
        # L: Good?
        ws_[f"L{row_num}"] = (
            f'=IF(I{row_num}=0,"",IF(J{row_num}+K{row_num}>=0,"OK","NEED "&ABS(J{row_num}+K{row_num})))'
        )
        ws_.cell(row=row_num, column=12).font = F_GOOD
        ws_.cell(row=row_num, column=12).alignment = A_CENTER
        # M: Wheel lb (input — weight of one raw wheel, lbs). Pre-filled from snapshot.
        ws_.cell(row=row_num, column=13, value=_spec.get("wheel_lb") or None).font = F_INPUT
        ws_.cell(row=row_num, column=13).fill = FILL_INPUT
        ws_.cell(row=row_num, column=13).alignment = A_RIGHT
        # N: Slice oz (input — finished slice weight, oz). Pre-filled from snapshot.
        ws_.cell(row=row_num, column=14, value=_spec.get("slice_oz") or None).font = F_INPUT
        ws_.cell(row=row_num, column=14).fill = FILL_INPUT
        ws_.cell(row=row_num, column=14).alignment = A_RIGHT
        # O: Wheels (computed) = ROUNDUP(Cut slices * Slice oz / (Wheel lb * 16)) — whole wheels
        ws_[f"O{row_num}"] = (
            f'=IF(OR(K{row_num}=0,M{row_num}=0,N{row_num}=0),"",'
            f"ROUNDUP(K{row_num}*N{row_num}/(M{row_num}*16),0))"
        )
        ws_.cell(row=row_num, column=15).font = F_NUM_BOLD
        ws_.cell(row=row_num, column=15).alignment = A_RIGHT
        # P: Safety (hidden) — flat 25 if raw demand (D+E+F+G) > 25 else 0; avoids circular ref with I
        ws_[f"P{row_num}"] = f"=IF((D{row_num}+E{row_num}+F{row_num}+G{row_num})>25,25,0)"
        ws_.cell(row=row_num, column=16).font = F_NUM
        ws_.cell(row=row_num, column=16).alignment = A_RIGHT
        # Q: Notes (input — free text). Pre-filled from snapshot; editable.
        _nc = ws_.cell(row=row_num, column=17, value=_spec.get("notes") or None)
        _nc.font = F_INPUT
        _nc.fill = FILL_INPUT
        _nc.alignment = A_LEFT

    def _write_section(ws_: Worksheet, start_row: int, label: str, bg: str, fg: str, rows_: list[dict]) -> int:
        """Write a section header + sub-grouped rows. Returns next available row."""
        if not rows_:
            return start_row
        r = start_row + 1
        _section_header(ws_, r, label, bg, fg, LAST_COL)
        r += 1

        current_cat = None
        cat_start_rows: dict[str, list[int]] = {}  # track rows per category for subtotals

        for sr in rows_:
            cat = sr["cat"]
            if cat != current_cat:
                current_cat = cat
                # Sub-category label row
                r += 1
                c_cat = ws_.cell(row=r, column=1, value=cat)
                c_cat.font = Font(name="Calibri", size=9, bold=True, color=SECTION_ACCENT)
                r += 1
                if cat not in cat_start_rows:
                    cat_start_rows[cat] = []

            _write_sku_row(ws_, r, sr)
            cat_start_rows[cat].append(r)
            r += 1

        # Subtotals per category
        for cat, data_rows in cat_start_rows.items():
            if not data_rows:
                continue
            r += 1
            ws_.cell(row=r, column=1, value=f"{cat} SUBTOTAL").font = Font(
                name="Calibri", size=9, bold=True, color=MUTED
            )
            # SUM formulas for key columns
            ws_[f"C{r}"] = f"=SUM({','.join(f'C{dr}' for dr in data_rows)})"
            ws_.cell(row=r, column=3).font = F_NUM_MUTED
            ws_.cell(row=r, column=3).alignment = A_RIGHT

            ws_[f"I{r}"] = f"=SUM({','.join(f'I{dr}' for dr in data_rows)})"
            ws_.cell(row=r, column=9).font = F_NUM_MUTED
            ws_.cell(row=r, column=9).alignment = A_RIGHT

        return r

    # Single flat section — alphabetical within category, no urgency groups.
    row = _write_section(ws, row, "SKUs", HEADER_BG, HEADER_FG, all_rows)

    last_row = row

    # ── Conditional Formatting ──
    data_start = HDR_ROW + 1

    # After (J): red < 0, green if ok
    ws.conditional_formatting.add(
        f"J{data_start}:J{last_row}",
        CellIsRule(
            operator="lessThan",
            formula=["0"],
            fill=PatternFill("solid", fgColor=SHORTAGE_BG),
            font=Font(color=SHORTAGE_FG),
        ),
    )
    ws.conditional_formatting.add(
        f"J{data_start}:J{last_row}",
        CellIsRule(
            operator="greaterThanOrEqual",
            formula=["0"],
            fill=PatternFill("solid", fgColor=OK_BG),
            font=Font(color=OK_FG),
        ),
    )

    # Good? (L)
    ws.conditional_formatting.add(
        f"L{data_start}:L{last_row}",
        FormulaRule(formula=[f'L{data_start}="OK"'], fill=PatternFill("solid", fgColor=OK_BG), font=Font(color=OK_FG)),
    )
    ws.conditional_formatting.add(
        f"L{data_start}:L{last_row}",
        FormulaRule(
            formula=[f'LEFT(L{data_start},4)="NEED"'],
            fill=PatternFill("solid", fgColor=SHORTAGE_BG),
            font=Font(color=SHORTAGE_FG),
        ),
    )

    # Input column highlights: First Order (F), Cut (K), Wheel lb (M), Slice oz (N)
    for _col in ("F", "K", "M", "N"):
        ws.conditional_formatting.add(
            f"{_col}{data_start}:{_col}{last_row}",
            CellIsRule(
                operator="greaterThan",
                formula=["0"],
                fill=PatternFill("solid", fgColor=INPUT_BG),
                font=Font(color=INPUT_FG, bold=True),
            ),
        )

    # ── Bundles & Limited-Release section (below the main table) ──
    _write_bundles_on_cut_order(ws, data, last_row + 2)

    # Freeze panes: row 4 header + columns A:B
    ws.freeze_panes = "C5"


def _write_bundles_on_cut_order(ws: Worksheet, data: dict, start_row: int) -> int:
    """List each bundle / limited-release box with its component SKUs indented
    beneath it. Components were already exploded into the demand columns above
    (added only where a container didn't already list them); this section shows
    the breakdown so the floor sees what each box contributes.
    """
    bundles = data.get("bundles") or {"WK1": {}, "WK2": {}}
    sku_name_fn = data["sku_name"]

    # Merge WK1 + WK2 per parent bundle.
    merged: dict = {}
    for wk in ("WK1", "WK2"):
        for parent, rec in (bundles.get(wk) or {}).items():
            m = merged.setdefault(parent, {"count": 0, "no_recipe": rec.get("no_recipe", False),
                                           "components": {}})
            m["count"] += rec.get("count", 0)
            m["no_recipe"] = m["no_recipe"] or rec.get("no_recipe", False)
            for cs, cd in rec.get("components", {}).items():
                c = m["components"].setdefault(cs, {"per": cd["per"], "added": 0, "already": 0,
                                                    "pickable": cd.get("pickable", True)})
                c["added"] += cd.get("added", 0)
                c["already"] += cd.get("already", 0)

    if not merged:
        return start_row

    row = start_row
    _section_header(ws, row,
                    "BUNDLES & LIMITED-RELEASE  (components already added into demand above)",
                    "1E293B", "FFFFFF", 15)
    row += 1
    for col, label in ((1, "Box / Component"), (2, "Name"), (3, "Boxes"),
                       (4, "Per box"), (5, "Added to cut"), (6, "Already in orders")):
        c = ws.cell(row=row, column=col, value=label)
        c.font = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="475569")
        c.alignment = A_RIGHT if col >= 3 else A_LEFT
    row += 1

    for parent in sorted(merged):
        m = merged[parent]
        pc = ws.cell(row=row, column=1, value=parent)
        pc.font = Font(name="Calibri", size=11, bold=True)
        ws.cell(row=row, column=2, value=sku_name_fn(parent) or "").font = Font(
            name="Calibri", size=10, bold=True)
        cc = ws.cell(row=row, column=3, value=m["count"])
        cc.font = Font(name="Calibri", size=11, bold=True)
        cc.alignment = A_RIGHT
        row += 1
        if m["no_recipe"]:
            note = ws.cell(row=row, column=2,
                           value="recipe unavailable in Simple Bundles — NOT exploded")
            note.font = Font(name="Calibri", size=9, italic=True, color="B45309")
            row += 1
            continue
        for cs in sorted(m["components"]):
            cd = m["components"][cs]
            ws.cell(row=row, column=1, value=f"   → {cs}").font = Font(
                name="Calibri", size=9, color="475569")
            nm = sku_name_fn(cs) or ""
            if not cd["pickable"]:
                nm = (nm + "  (not cut)").strip()
            ws.cell(row=row, column=2, value=nm).font = F_NAME
            for col, val in ((4, cd["per"]), (5, cd["added"]), (6, cd["already"])):
                cell = ws.cell(row=row, column=col, value=val)
                cell.font = F_NUM if col == 5 and val else F_NUM_MUTED
                cell.alignment = A_RIGHT
            row += 1
        row += 1  # spacer between bundles
    return row


# ── Main ─────────────────────────────────────────────────────────────


def main() -> str:
    settings = load_settings()
    data = _fetch_all_data(settings)

    wb = openpyxl.Workbook()

    # Raw Materials tab removed 2026-07-28 (Kurt) — not used on the floor.

    # Checklist (pre-cut sanity)
    _build_checklist_tab(wb, data, settings)

    # Tab 1: Cut Order (main) — build the sheet structure first
    ws_cut = wb.active
    ws_cut.title = "Cut Order"

    # Write assignments onto the Cut Order sheet (columns W-AJ)
    prcjam_monthly_last_row, cexec_last_row, jam_last_row = _write_assignments_on_cut_order(
        ws_cut, data, settings)

    # Build the main cut order content (columns A-U) using same-sheet SUMIF refs
    _build_cut_order_tab(wb, data, settings, prcjam_monthly_last_row, cexec_last_row, jam_last_row)

    # Reorder tabs: Cut Order first
    wb.move_sheet("Cut Order", offset=-1)

    # Save. Prefix the run day so the Monday (1st run) and Tuesday (final) builds
    # for the same ship week don't overwrite each other (Kurt 2026-07-21).
    ship_date = SHIP_WEEK_MONDAY.isoformat()
    _wd = datetime.now().weekday()
    _day = {0: "MON", 1: "TUES"}.get(_wd, ["MON", "TUES", "WED", "THU", "FRI", "SAT", "SUN"][_wd])
    out_path = os.path.join(BASE, f"{_day}_cut_order_v2_{ship_date}.xlsx")
    wb.save(out_path)

    # Recalc formulas via Excel COM (Windows) so cached values are baked into the file.
    # openpyxl writes formula strings without evaluating; without recalc, cells appear
    # blank in Excel until F9. Google Sheets re-evaluates on open, but Excel does not.
    try:
        import win32com.client
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        try:
            wb_xl = excel.Workbooks.Open(out_path)
            excel.CalculateFull()
            wb_xl.Save()
            wb_xl.Close(SaveChanges=False)
            print("  Recalc: Excel COM OK - formulas cached")
        finally:
            excel.Quit()
    except Exception as e:
        print(f"  Recalc skipped (Excel COM unavailable): {e}")

    print(f"\nExcel written to: {out_path}")
    print(f"  {len(data['active_skus'])} active SKUs")
    print(f"  Tab 1: Cut Order (alpha by category, single-cohort, assignments at cols W-AE)")
    print(f"  Blue cells = editable input")

    return out_path


# ── Google Drive Upload ──────────────────────────────────────────────

DRIVE_FOLDER_ID = "1TgvxK10tFAPJqhkYw-6u1Umnvp9wMJ3I"
DRIVE_TOKEN_PATH = os.path.join(BASE, "dist", "drive_oauth_token.json")


def upload_to_drive(file_path: str) -> str:
    """Upload XLSX to shared Google Drive folder using OAuth credentials."""
    from google.auth.transport.requests import Request
    from google.oauth2.credentials import Credentials
    from googleapiclient.discovery import build
    from googleapiclient.http import MediaIoBaseUpload

    with open(DRIVE_TOKEN_PATH, encoding="utf-8") as f:
        token_data = json.load(f)

    creds = Credentials(
        token=token_data["token"],
        refresh_token=token_data["refresh_token"],
        token_uri=token_data["token_uri"],
        client_id=token_data["client_id"],
        client_secret=token_data["client_secret"],
        scopes=token_data["scopes"],
    )
    if creds.expired:
        creds.refresh(Request())
        token_data["token"] = creds.token
        with open(DRIVE_TOKEN_PATH, "w") as f:
            json.dump(token_data, f, indent=2)

    drive = build("drive", "v3", credentials=creds)
    file_name = os.path.basename(file_path)
    with open(file_path, "rb") as fh:
        media = MediaIoBaseUpload(
            io.BytesIO(fh.read()),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            resumable=True,
        )
        f = (
            drive.files()
            .create(
                body={"name": file_name, "parents": [DRIVE_FOLDER_ID]},
                media_body=media,
                fields="id, webViewLink",
                supportsAllDrives=True,
            )
            .execute()
        )
    link = f.get("webViewLink", f"https://drive.google.com/file/d/{f['id']}")
    print(f"  Uploaded: {link}")
    return link


def _ingest_tommy_inputs(path: str) -> None:
    """Snapshot Tommy's filled cut-order inputs into settings so the next build
    pre-fills them (same value → same SKU / curation).

    Reads BY HEADER (robust to column shifts):
      - Main table: First Order (F), Wheel lb (M), Slice oz (N) per SKU
        → settings['cut_order_specs'][sku] = {first_order, wheel_lb, slice_oz}
      - PR-CJAM ASSIGNMENTS 'Cheese SKU' per curation → settings['pr_cjam'][cur]['cheese']
      - CEX-EC ASSIGNMENTS 'Cheese SKU' per curation  → settings['cex_ec'][cur]

    Explicit-only (never runs on a normal build) so it can't silently re-populate
    a deliberately-cleared rotation table.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Cut Order"] if "Cut Order" in wb.sheetnames else wb.active

    def _norm(v):
        return str(v).strip().lower() if v is not None else ""

    # -- Locate the main-table header row (has SKU + First Order) --
    hdr_row, hcol = None, {}
    for r in range(1, 8):
        labels = {_norm(ws.cell(r, c).value): c for c in range(1, 20) if ws.cell(r, c).value}
        if "sku" in labels and "first order" in labels:
            hdr_row, hcol = r, labels
            break
    specs: dict = {}
    if hdr_row:
        c_sku = hcol["sku"]
        c_f = hcol.get("first order")
        c_m = hcol.get("wheel lb")
        c_n = hcol.get("slice oz")
        c_k = hcol.get("cut")
        c_notes = hcol.get("notes")
        for r in range(hdr_row + 1, ws.max_row + 1):
            sku = ws.cell(r, c_sku).value
            if not sku or not isinstance(sku, str) or not sku.strip():
                continue
            sku = sku.strip()
            rec = {}
            for key, col in (("first_order", c_f), ("wheel_lb", c_m),
                             ("slice_oz", c_n), ("cut", c_k)):
                if col:
                    v = ws.cell(r, col).value
                    if isinstance(v, (int, float)) and v:
                        rec[key] = v
            if c_notes:
                nv = ws.cell(r, c_notes).value
                if nv is not None and str(nv).strip():
                    rec["notes"] = str(nv).strip()
            if rec:
                specs[sku] = rec

    # -- Locate assignment blocks by their title cells --
    def _read_assignment(title_substr, sku_label="cheese sku"):
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                if title_substr in _norm(ws.cell(r, c).value):
                    sub = r + 1  # sub-header row (Curation | Cheese/Jam SKU | ...)
                    # Find THIS block's Curation col, then the FIRST sku col at/after
                    # it, and STOP. Blocks sit only ~5 cols apart, so a wide scan
                    # overshoots into the next block and grabs its columns instead
                    # (2026-07-28: PR-CJAM cheese got filled from CEX-EC this way).
                    cur_col = chz_col = None
                    for cc in range(c, c + 5):
                        lbl = _norm(ws.cell(sub, cc).value)
                        if cur_col is None and lbl == "curation":
                            cur_col = cc
                        elif cur_col is not None and sku_label in lbl:
                            chz_col = cc
                            break
                    if not (cur_col and chz_col):
                        return {}
                    out = {}
                    rr = sub + 1
                    while True:
                        cur = ws.cell(rr, cur_col).value
                        if not cur or not str(cur).strip():
                            break
                        cur = str(cur).strip()
                        # stop if we've wandered into a MONTHLY slot table ("AHB-…")
                        if cur.upper().startswith("AHB-"):
                            break
                        chz = ws.cell(rr, chz_col).value
                        if chz and str(chz).strip():
                            out[cur] = str(chz).strip()
                        rr += 1
                    return out
        return {}

    prcjam_cheese = _read_assignment("pr-cjam assignments")
    cexec_cheese = _read_assignment("cex-ec assignments")
    prcjam_jam = _read_assignment("pr-cjam jam assignments", sku_label="jam sku")

    # -- Monthly-box slot fill-ins (AHB-MED/CMED/LGE (YYYY-MM)) --
    # Slot name in the title's column, SKU in the next column. Persist so the
    # regen keeps Tommy's monthly recipe AND the SUMIF counts the components.
    import re as _re
    monthly_recipes: dict = {}
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            m = _re.match(r"ahb-(med|cmed|lge)\s*\((\d{4}-\d{2})\)", _norm(ws.cell(r, c).value))
            if not m:
                continue
            bt, month = "AHB-" + m.group(1).upper(), m.group(2)
            rows = []
            rr = r + 1
            while rr <= ws.max_row:
                slot = ws.cell(rr, c).value
                if not slot or not str(slot).strip():
                    break
                if _re.match(r"ahb-(med|cmed|lge)\s*\(", _norm(slot)):
                    break
                sku = ws.cell(rr, c + 1).value
                rows.append([str(slot).strip(), str(sku).strip() if sku else "", 1])
                rr += 1
            if rows:
                monthly_recipes.setdefault(month, {})[bt] = rows
    wb.close()

    # -- Merge into settings (dist + APPDATA copies) --
    import json as _json
    targets = [
        SETTINGS_PATH,
        os.path.join(os.environ.get("APPDATA", ""), "AppyHour", "inventory_reorder_settings.json"),
    ]
    written = 0
    for tgt in targets:
        if not tgt or not os.path.exists(tgt):
            continue
        with open(tgt, encoding="utf-8") as f:
            st = _json.load(f)
        cos = st.setdefault("cut_order_specs", {})
        for sku, rec in specs.items():
            cos.setdefault(sku, {}).update(rec)
        pr = st.setdefault("pr_cjam", {})
        for cur, chz in prcjam_cheese.items():
            entry = pr.setdefault(cur, {})
            if isinstance(entry, dict):
                entry["cheese"] = chz
        for cur, jam in prcjam_jam.items():
            entry = pr.setdefault(cur, {})
            if isinstance(entry, dict):
                entry["jam"] = jam
        cx = st.setdefault("cex_ec", {})
        for cur, chz in cexec_cheese.items():
            cx[cur] = chz
        mrec = st.setdefault("monthly_box_recipes", {})
        for month, bts in monthly_recipes.items():
            mrec.setdefault(month, {}).update(bts)
        import shutil as _sh
        _sh.copy2(tgt, tgt + ".bak")
        with open(tgt, "w", encoding="utf-8") as f:
            _json.dump(st, f, indent=2)
        written += 1

    print(f"Ingested Tommy inputs from: {path}")
    print(f"  {len(specs)} SKUs with First Order/Wheel lb/Slice oz/Cut/Notes specs")
    print(f"  {len(prcjam_cheese)} PR-CJAM cheese, {sum(1 for v in prcjam_jam.values() if v)} PR-CJAM jam, "
          f"{len(cexec_cheese)} CEX-EC assignments")
    _mslots = sum(sum(1 for row in b if row[1]) for bts in monthly_recipes.values() for b in bts.values())
    print(f"  monthly slots: {_mslots} filled across "
          f"{sum(len(b) for b in monthly_recipes.values())} box-month blocks")
    print(f"  wrote {written} settings file(s) (.bak backup kept). Next build pre-fills these.")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Generate cut order XLSX v2")
    parser.add_argument("--local", action="store_true", help="Generate locally, don't upload")
    parser.add_argument("--ingest", metavar="XLSX",
                        help="Snapshot Tommy's filled F/M/N + T/Y from a cut order xlsx into settings")
    args = parser.parse_args()

    if args.ingest:
        _ingest_tommy_inputs(args.ingest)
        sys.exit(0)

    out_path = main()
    if not args.local:
        try:
            upload_to_drive(out_path)
        except Exception as e:
            print(f"  Upload failed: {e}")
            print(f"  File saved locally: {out_path}")
