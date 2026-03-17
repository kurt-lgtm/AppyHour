"""
Fulfillment Planner -- Flask + pywebview
Gamified weekly cheese fulfillment planning.
Shares settings JSON with inventory_reorder.py.
"""
from __future__ import annotations

import json
import os
import sys
import math
import re
import datetime
import csv
import io
import threading
from collections import defaultdict
from flask import Flask, render_template, jsonify, request, send_file

# ── Constants ───────────────────────────────────────────────────────────

SETTINGS_FILE = "inventory_reorder_settings.json"
CURATION_ORDER = ["MONG", "MDT", "OWC", "SPN", "ALPN", "ISUN", "HHIGH"]
EXTRA_CURATIONS = ["NMS", "BYO", "SS"]
ALL_CURATIONS = CURATION_ORDER + EXTRA_CURATIONS
WHEEL_TO_SLICE_FACTOR = 2.67

# SKUs excluded from PR-CJAM and CEX-EC assignment candidates
ASSIGNMENT_EXCLUDE = {"CH-MAFT"}

# Global extra assignment slots (single SKU across all curations)
GLOBAL_EXTRA_SLOTS = {
    "EX-EC":  {"category": "Cheese",        "prefix": "CH-"},
    "CEX-EM": {"category": "Meat",          "prefix": "MT-"},
    "EX-EM":  {"category": "Meat",          "prefix": "MT-"},
    "CEX-EA": {"category": "Accompaniment", "prefix": "AC-"},
    "EX-EA":  {"category": "Accompaniment", "prefix": "AC-"},
}

MONTHLY_BOX_TYPES = ["AHB-MED", "AHB-CMED", "AHB-LGE"]
MONTHLY_BOX_SLOTS = {
    "AHB-MED": [
        ("Cheese 1", "CH-"), ("Cheese 2", "CH-"),
        ("Meat 1", "MT-"), ("Meat 2", "MT-"),
        ("Crackers", "AC-"),
        ("Accompaniment 1", "AC-"), ("Accompaniment 2", "AC-"),
        ("PR-CJAM-GEN Cheese", "CH-"), ("PR-CJAM-GEN Jam", "AC-"),
    ],
    "AHB-CMED": [
        ("Cheese 1", "CH-"), ("Cheese 2", "CH-"),
        ("Cheese 3", "CH-"), ("Cheese 4", "CH-"),
        ("Crackers", "AC-"),
        ("Accompaniment 1", "AC-"), ("Accompaniment 2", "AC-"),
        ("PR-CJAM-GEN Cheese", "CH-"), ("PR-CJAM-GEN Jam", "AC-"),
    ],
    "AHB-LGE": [
        ("Cheese 1", "CH-"), ("Cheese 2", "CH-"), ("Cheese 3", "CH-"),
        ("Meat 1", "MT-"), ("Meat 2", "MT-"), ("Meat 3", "MT-"),
        ("Crackers", "AC-"),
        ("Accompaniment 1", "AC-"), ("Accompaniment 2", "AC-"),
        ("PR-CJAM-GEN Cheese", "CH-"), ("PR-CJAM-GEN Jam", "AC-"),
    ],
}

# ── Helpers ──────────────────────────────────────────────────────────

def _inv_qty(data):
    """Extract qty from an inventory entry (dict or raw int)."""
    return data.get("qty", 0) if isinstance(data, dict) else int(data or 0)


# ── Settings persistence (shared with main app) ────────────────────────

def _get_app_dir():
    if getattr(sys, 'frozen', False):
        # Frozen exe: settings live next to the executable
        return os.path.dirname(sys.executable)
    base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    dist = os.path.join(base, "dist", SETTINGS_FILE)
    if os.path.exists(dist):
        return os.path.join(base, "dist")
    return base


def _get_project_dir():
    """Return the project root dir (for RMFG folders, Shipments, etc.)."""
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


def load_settings():
    path = os.path.join(_get_app_dir(), SETTINGS_FILE)
    if os.path.exists(path):
        try:
            with open(path, "r") as f:
                return json.load(f)
        except Exception:
            pass
    return {}


def save_settings(data):
    path = os.path.join(_get_app_dir(), SETTINGS_FILE)
    # Safety: never overwrite with empty/tiny data if file already has content
    try:
        if os.path.exists(path) and os.path.getsize(path) > 100:
            new_json = json.dumps(data, indent=2)
            if len(new_json) < 50:
                return  # refuse to write essentially empty settings
            # Write to temp file first, then rename (atomic-ish)
            tmp_path = path + ".tmp"
            with open(tmp_path, "w") as f:
                f.write(new_json)
            # Backup before overwrite
            bak_path = path + ".bak"
            import shutil
            shutil.copy2(path, bak_path)
            os.replace(tmp_path, path)
        else:
            with open(path, "w") as f:
                json.dump(data, f, indent=2)
    except Exception:
        pass


# ── Constraint checking ────────────────────────────────────────────────

def check_constraint(curation, prcjam_cheese, cexec_cheese,
                     recipes, pr_cjam, cex_ec):
    if curation not in CURATION_ORDER:
        return "OK"
    idx = CURATION_ORDER.index(curation)
    nearby = set()
    for offset in [-2, -1, 1, 2]:
        ni = idx + offset
        if 0 <= ni < len(CURATION_ORDER):
            nb = CURATION_ORDER[ni]
            for item in recipes.get(nb, []):
                sku = item[0] if isinstance(item, (list, tuple)) else item
                if isinstance(sku, str) and sku.startswith("CH-"):
                    nearby.add(sku)
            n_pr = pr_cjam.get(nb, {})
            if isinstance(n_pr, dict) and n_pr.get("cheese"):
                nearby.add(n_pr["cheese"])
            n_ec = cex_ec.get(nb, "")
            if n_ec:
                nearby.add(n_ec)
    violations = []
    if prcjam_cheese and prcjam_cheese in nearby:
        violations.append("PR")
    if cexec_cheese and cexec_cheese in nearby:
        violations.append("EC")
    return "CONFLICT: " + "+".join(violations) if violations else "OK"


# ── Flask app ───────────────────────────────────────────────────────────

# PyInstaller frozen support: templates/static live next to the exe in _MEIPASS
if getattr(sys, 'frozen', False):
    _bundle_dir = sys._MEIPASS
    app = Flask(__name__,
                template_folder=os.path.join(_bundle_dir, 'templates'),
                static_folder=os.path.join(_bundle_dir, 'static'))
else:
    app = Flask(__name__)
STATE = {"saved": {}, "csv_demand": {}}


def _s():
    """Get current settings."""
    return STATE["saved"]


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/data")
def get_data():
    """Return all data needed for the UI."""
    s = _s()
    return jsonify({
        "inventory": s.get("inventory", {}),
        "wheel_inventory": s.get("wheel_inventory", {}),
        "open_pos": s.get("open_pos", []),
        "pr_cjam": s.get("pr_cjam", {}),
        "cex_ec": s.get("cex_ec", {}),
        "cexec_splits": s.get("cexec_splits", {}),
        "curation_recipes": s.get("curation_recipes", {}),
        "recharge_queued": s.get("recharge_queued", {}),
        "recharge_queued_resolved": s.get("recharge_queued_resolved", {}),
        "shopify_api_demand": s.get("shopify_api_demand", {}),
        "manual_demand": s.get("manual_demand", {}),
        "vendor_catalog": s.get("vendor_catalog", {}),
        "global_extras": s.get("global_extras", {}),
        "curations": ALL_CURATIONS,
        "curation_order": CURATION_ORDER,
    })


@app.route("/api/calculate", methods=["POST"])
def calculate():
    s = _s()
    inventory = s.get("inventory", {})
    wheel_inv = s.get("wheel_inventory", {})
    open_pos = s.get("open_pos", [])
    pr_cjam = s.get("pr_cjam", {})
    cex_ec = s.get("cex_ec", {})
    splits = s.get("cexec_splits", {})
    rq_resolved = s.get("recharge_queued_resolved", {})
    rq_raw = s.get("recharge_queued", {})
    shopify_demand = s.get("shopify_api_demand", {})
    manual = s.get("manual_demand", {})
    csv_demand = STATE.get("csv_demand", {})
    recipes = s.get("curation_recipes", {})

    # 1. Inventory snapshot
    inv = {}
    for sku, data in inventory.items():
        inv[sku] = data.get("qty", 0) if isinstance(data, dict) else int(data)

    for wsku, wd in wheel_inv.items():
        if isinstance(wd, dict):
            w = float(wd.get("weight_lbs", 0))
            c = int(wd.get("count", 0))
            t = wd.get("target_sku", "")
            if t and w > 0 and c > 0:
                inv[t] = inv.get(t, 0) + int(w * c * WHEEL_TO_SLICE_FACTOR)

    for po in open_pos:
        sku = po.get("sku", "")
        if sku and po.get("status", "Open") == "Open":
            try:
                inv[sku] = inv.get(sku, 0) + int(float(po.get("qty", 0)))
            except (ValueError, TypeError):
                pass

    # 2. Demand
    d_direct = defaultdict(int)
    d_prcjam = defaultdict(int)
    d_cexec = defaultdict(int)
    d_exec = defaultdict(int)

    # Recharge resolved
    for month, data in rq_resolved.items():
        for suffix, count in data.get("pr_cjam", {}).items():
            info = pr_cjam.get(suffix, {})
            ch = info.get("cheese", "") if isinstance(info, dict) else str(info)
            if ch:
                d_prcjam[ch] += int(count)

        for suffix, count in data.get("cex_ec", {}).items():
            sp = splits.get(suffix, {})
            if sp:
                total = int(count)
                rem = total
                items = list(sp.items())
                for i, (sk, ratio) in enumerate(items):
                    if i == len(items) - 1:
                        d_cexec[sk] += rem
                    else:
                        portion = int(total * float(ratio))
                        d_cexec[sk] += portion
                        rem -= portion
            else:
                ch = cex_ec.get(suffix, "")
                if ch:
                    d_cexec[ch] += int(count)

    # Recharge raw
    ge = s.get("global_extras", {})
    for month, skus in rq_raw.items():
        for sku, qty in skus.items():
            upper = sku.upper()
            if sku.startswith("CH-"):
                d_direct[sku] += int(qty)
            else:
                ge_resolved = ge.get(upper)
                if ge_resolved:
                    d_direct[ge_resolved] += int(qty)

    # Shopify
    for sku, qty in shopify_demand.items():
        if sku.startswith("CH-"):
            d_direct[sku] += int(qty)

    # CSV
    for sku, qty in csv_demand.items():
        if sku.startswith("CH-"):
            d_direct[sku] += int(qty)

    # Manual
    for sku, qty in manual.items():
        if sku.startswith("CH-"):
            d_direct[sku] += int(qty)

    # 3. Results
    all_ch = set()
    all_ch.update(k for k in inv if k.startswith("CH-"))
    all_ch.update(d_direct.keys(), d_prcjam.keys(),
                  d_cexec.keys(), d_exec.keys())

    results = []
    shortage_count = 0
    for sku in sorted(all_ch):
        avail = inv.get(sku, 0)
        dd = d_direct.get(sku, 0)
        dp = d_prcjam.get(sku, 0)
        dc = d_cexec.get(sku, 0)
        de = d_exec.get(sku, 0)
        total = dd + dp + dc + de
        net = avail - total

        if total == 0:
            status = "NO DEMAND"
        elif net < 0:
            status = "SHORTAGE"
            shortage_count += 1
        elif net < total * 0.2:
            status = "TIGHT"
        elif net > avail * 0.5 and avail > 200:
            status = "SURPLUS"
        else:
            status = "OK"

        results.append({
            "sku": sku, "available": avail,
            "direct": dd, "prcjam": dp, "cexec": dc, "exec": de,
            "total_demand": total, "net": net, "status": status,
        })

    status_order = {"SHORTAGE": 0, "TIGHT": 1, "OK": 2,
                    "SURPLUS": 3, "NO DEMAND": 4}
    results.sort(key=lambda r: (status_order.get(r["status"], 9), r["net"]))

    # Assignment demands per curation
    assign_demands = {}
    for cur in ALL_CURATIONS:
        pr_qty = sum(int(md.get("pr_cjam", {}).get(cur, 0))
                     for md in rq_resolved.values())
        ec_qty = sum(int(md.get("cex_ec", {}).get(cur, 0))
                     for md in rq_resolved.values())
        assign_demands[cur] = {"pr_qty": pr_qty, "ec_qty": ec_qty}

    # Shelf life
    today = datetime.date.today()
    shelf_items = []
    for sku, data in inventory.items():
        if not sku.startswith("CH-") or not isinstance(data, dict):
            continue
        dates = data.get("expiration_dates", [])
        if not dates:
            continue
        try:
            earliest = datetime.date.fromisoformat(dates[0])
        except (ValueError, IndexError):
            continue
        days = (earliest - today).days
        qty = data.get("qty", 0)
        if days <= 14 and qty > 0:
            shelf_items.append({
                "sku": sku, "days_left": days, "qty": qty,
                "action": "EXPIRED" if days < 0 else
                          "USE NOW" if days <= 7 else "Prioritize",
            })

    # Multi-week projections (weeks 2, 3, 4)
    weeks = []
    prev = {r["sku"]: r for r in results}
    for week_num in range(2, 6):
        week_results = []
        week_shortages = 0
        for r in results:
            sku = r["sku"]
            if not sku.startswith("CH-"):
                continue
            p = prev.get(sku, r)
            carry = max(0, p.get("proj_net", p["net"]))
            demand = r["total_demand"]
            proj_net = carry - demand
            if demand == 0:
                status = "NO DEMAND"
            elif proj_net < 0:
                status = "PLAN PO"
                week_shortages += 1
            elif proj_net < demand * 0.3:
                status = "TIGHT"
            else:
                status = "OK"
            entry = {
                "sku": sku, "carry_fwd": carry,
                "demand": demand, "net": proj_net,
                "status": status, "proj_net": proj_net,
            }
            week_results.append(entry)

        weeks.append({
            "week": week_num,
            "results": week_results,
            "shortages": week_shortages,
        })
        prev = {r["sku"]: r for r in week_results}

    return jsonify({
        "results": results,
        "shortages": shortage_count,
        "total_skus": len([r for r in results if r["total_demand"] > 0]),
        "total_units": sum(r["total_demand"] for r in results),
        "assign_demands": assign_demands,
        "shelf_life": shelf_items,
        "weeks": weeks,
    })


@app.route("/api/assignments")
def get_assignments():
    s = _s()
    pr_cjam = s.get("pr_cjam", {})
    cex_ec = s.get("cex_ec", {})
    splits = s.get("cexec_splits", {})
    recipes = s.get("curation_recipes", {})

    # Per-curation box counts from loaded demand data (STATE), fall back to settings
    prcjam_counts = STATE.get("rmfg_prcjam_sat") or _s().get("rmfg_prcjam_sat", {})
    cexec_counts = STATE.get("rmfg_cexec_sat") or _s().get("rmfg_cexec_sat", {})

    rows = []
    for cur in ALL_CURATIONS:
        info = pr_cjam.get(cur, {})
        pr_ch = info.get("cheese", "") if isinstance(info, dict) else str(info)
        ec_ch = cex_ec.get(cur, "")
        sp = splits.get(cur, {})
        split_text = (" / ".join(f"{int(float(v)*100)}% {k}"
                                 for k, v in sp.items()) if sp else "")
        constraint = check_constraint(cur, pr_ch, ec_ch, recipes, pr_cjam, cex_ec)
        pr_qty = int(prcjam_counts.get(cur, 0))
        ec_qty = int(cexec_counts.get(cur, 0))
        rows.append({
            "curation": cur, "prcjam_cheese": pr_ch,
            "cexec_cheese": ec_ch, "split": split_text,
            "constraint": constraint,
            "pr_qty": pr_qty, "ec_qty": ec_qty,
        })
    return jsonify(rows)


@app.route("/api/assign", methods=["POST"])
def set_assignment():
    data = request.json
    cur = data["curation"]
    slot = data["slot"]
    cheese = data["cheese"]
    s = _s()

    recipes = s.get("curation_recipes", {})
    pr_cjam = s.get("pr_cjam", {})
    cex_ec = s.get("cex_ec", {})

    # Validate constraint
    test_pr = cheese if slot == "prcjam" else (
        pr_cjam.get(cur, {}).get("cheese", "")
        if isinstance(pr_cjam.get(cur), dict) else "")
    test_ec = cheese if slot == "cexec" else cex_ec.get(cur, "")
    constraint = check_constraint(cur, test_pr, test_ec, recipes, pr_cjam, cex_ec)

    # Constraint is advisory — warn but allow the assignment
    if slot == "prcjam":
        if isinstance(pr_cjam.get(cur), dict):
            pr_cjam[cur]["cheese"] = cheese
        else:
            pr_cjam[cur] = {"cheese": cheese, "jam": ""}
        s["pr_cjam"] = pr_cjam
    else:
        cex_ec[cur] = cheese
        s["cex_ec"] = cex_ec

    save_settings(s)
    return jsonify({"ok": True, "constraint": constraint})


@app.route("/api/candidates/<curation>/<slot>")
def get_candidates(curation, slot):
    s = _s()
    inventory = s.get("inventory", {})
    pr_cjam = s.get("pr_cjam", {})
    cex_ec = s.get("cex_ec", {})
    recipes = s.get("curation_recipes", {})

    candidates = []
    for sku, data in inventory.items():
        if not sku.startswith("CH-"):
            continue
        if sku in ASSIGNMENT_EXCLUDE:
            continue
        qty = data.get("qty", 0) if isinstance(data, dict) else 0
        if qty <= 0:
            continue
        test_pr = sku if slot == "prcjam" else (
            pr_cjam.get(curation, {}).get("cheese", "")
            if isinstance(pr_cjam.get(curation), dict) else "")
        test_ec = sku if slot == "cexec" else cex_ec.get(curation, "")
        c = check_constraint(curation, test_pr, test_ec, recipes, pr_cjam, cex_ec)
        candidates.append({"sku": sku, "qty": qty, "constraint": c})

    candidates.sort(key=lambda x: (0 if x["constraint"] == "OK" else 1, -x["qty"]))
    return jsonify(candidates)


@app.route("/api/monthly_boxes")
def get_monthly_boxes():
    s = _s()
    recipes = s.get("monthly_box_recipes", {})
    # Order-derived counts (from Shopify + Recharge), fall back to manual
    order_counts = STATE.get("rmfg_monthly_box_counts") or s.get("rmfg_monthly_box_counts", {})
    manual_counts = s.get("monthly_box_counts", {})
    # Current month label (e.g. "2026-03")
    month = datetime.datetime.now().strftime("%Y-%m")
    month_data = recipes.get(month, {})
    result = {}
    for box_type in MONTHLY_BOX_TYPES:
        slots_template = MONTHLY_BOX_SLOTS[box_type]
        assigned = month_data.get(box_type, [])
        # Build slot list with current assignments
        slot_rows = []
        for i, (slot_name, prefix) in enumerate(slots_template):
            sku = ""
            qty = 1
            if i < len(assigned):
                entry = assigned[i]
                sku = entry[1] if len(entry) > 1 else ""
                qty = entry[2] if len(entry) > 2 else 1
            slot_rows.append({"slot": slot_name, "prefix": prefix, "sku": sku, "qty": qty})
        count = order_counts.get(box_type, manual_counts.get(box_type, 0))
        result[box_type] = {
            "slots": slot_rows,
            "count": count,
            "from_orders": box_type in order_counts,
        }
    return jsonify(result)


@app.route("/api/monthly_box_assign", methods=["POST"])
def set_monthly_box_assign():
    data = request.json
    box_type = data["box_type"]
    slot_index = data["slot_index"]
    sku = data["sku"]
    s = _s()
    month = datetime.datetime.now().strftime("%Y-%m")
    recipes = s.setdefault("monthly_box_recipes", {})
    month_data = recipes.setdefault(month, {})
    slots_template = MONTHLY_BOX_SLOTS[box_type]
    assigned = month_data.get(box_type, [])
    # Pad to correct length
    while len(assigned) < len(slots_template):
        assigned.append([slots_template[len(assigned)][0], "", 1])
    assigned[slot_index] = [slots_template[slot_index][0], sku, 1]
    month_data[box_type] = assigned
    save_settings(s)
    return jsonify({"ok": True})


@app.route("/api/monthly_box_candidates/<box_type>/<int:slot_index>")
def get_monthly_box_candidates(box_type, slot_index):
    s = _s()
    inventory = s.get("inventory", {})
    slots_template = MONTHLY_BOX_SLOTS.get(box_type, [])
    if slot_index >= len(slots_template):
        return jsonify([])
    slot_name, prefix = slots_template[slot_index]
    candidates = []
    for sku, data in inventory.items():
        if not sku.startswith(prefix):
            continue
        qty = _inv_qty(data)
        name = data.get("name", "") if isinstance(data, dict) else ""
        candidates.append({"sku": sku, "qty": qty, "name": name})
    candidates.sort(key=lambda c: c["qty"], reverse=True)
    return jsonify(candidates)


@app.route("/api/monthly_box_count", methods=["POST"])
def set_monthly_box_count():
    data = request.json
    box_type = data["box_type"]
    count = int(data.get("count", 0))
    s = _s()
    counts = s.setdefault("monthly_box_counts", {})
    counts[box_type] = count
    save_settings(s)
    return jsonify({"ok": True})


@app.route("/api/auto_assign", methods=["POST"])
def auto_assign():
    s = _s()
    inventory = s.get("inventory", {})
    pr_cjam = s.get("pr_cjam", {})
    cex_ec = s.get("cex_ec", {})
    recipes = s.get("curation_recipes", {})
    rq = s.get("recharge_queued_resolved", {})

    headroom = {}
    for sku, data in inventory.items():
        if sku.startswith("CH-") and sku not in ASSIGNMENT_EXCLUDE:
            headroom[sku] = data.get("qty", 0) if isinstance(data, dict) else 0

    consumed = defaultdict(int)
    new_pr = {}
    new_ec = {}
    changes = []
    used = set()

    for cur in ALL_CURATIONS:
        est = sum(int(md.get("pr_cjam", {}).get(cur, 0)) for md in rq.values())
        if est == 0:
            est = 50
        cands = [(sk, q - consumed.get(sk, 0))
                 for sk, q in headroom.items()
                 if q - consumed.get(sk, 0) >= est and sk not in used
                 and check_constraint(cur, sk, "", recipes, pr_cjam, cex_ec) == "OK"]
        cands.sort(key=lambda x: -x[1])
        if cands:
            best = cands[0][0]
            old = pr_cjam.get(cur, {}).get("cheese", "") if isinstance(pr_cjam.get(cur), dict) else ""
            new_pr[cur] = {"cheese": best, "jam": ""}
            consumed[best] += est
            used.add(best)
            if best != old:
                changes.append(f"PR-CJAM-{cur}: {old} -> {best}")
        else:
            new_pr[cur] = pr_cjam.get(cur, {"cheese": "", "jam": ""})

    for cur in ALL_CURATIONS:
        est = sum(int(md.get("cex_ec", {}).get(cur, 0)) for md in rq.values())
        if est == 0:
            est = 20
        pr_ch = new_pr.get(cur, {}).get("cheese", "")
        cands = [(sk, q - consumed.get(sk, 0))
                 for sk, q in headroom.items()
                 if q - consumed.get(sk, 0) >= est and sk != pr_ch
                 and check_constraint(cur, pr_ch, sk, recipes, pr_cjam, cex_ec) == "OK"]
        cands.sort(key=lambda x: -x[1])
        if cands:
            best = cands[0][0]
            old = cex_ec.get(cur, "")
            new_ec[cur] = best
            consumed[best] += est
            if best != old:
                changes.append(f"CEX-EC-{cur}: {old} -> {best}")
        else:
            new_ec[cur] = cex_ec.get(cur, "")

    if changes:
        s["pr_cjam"] = new_pr
        s["cex_ec"] = new_ec
        save_settings(s)

    return jsonify({"changes": changes, "count": len(changes)})


# ── Global Extras ─────────────────────────────────────────────────────

@app.route("/api/global_extras")
def get_global_extras():
    """Return current global extra assignments with on-hand qty."""
    s = _s()
    ge = s.get("global_extras", {})
    inventory = s.get("inventory", {})
    result = {}
    for slot, meta in GLOBAL_EXTRA_SLOTS.items():
        sku = ge.get(slot, "")
        qty = 0
        if sku:
            qty = _inv_qty(inventory.get(sku, {}))
        result[slot] = {"sku": sku, "qty": qty,
                        "category": meta["category"], "prefix": meta["prefix"]}
    return jsonify(result)


@app.route("/api/set_global_extra", methods=["POST"])
def set_global_extra():
    """Set a single global extra assignment."""
    data = request.json
    slot = data.get("slot", "")
    sku = data.get("sku", "")
    if slot not in GLOBAL_EXTRA_SLOTS:
        return jsonify({"ok": False, "error": f"Unknown slot: {slot}"}), 400
    s = _s()
    if sku:
        inventory = s.get("inventory", {})
        if sku not in inventory:
            return jsonify({"ok": False, "error": f"SKU {sku} not in inventory"}), 400
    ge = s.get("global_extras", {})
    ge[slot] = sku
    s["global_extras"] = ge
    save_settings(s)
    return jsonify({"ok": True})


@app.route("/api/global_extra_candidates/<slot>")
def get_global_extra_candidates(slot):
    """Return candidate SKUs for a global extra slot, filtered by type."""
    if slot not in GLOBAL_EXTRA_SLOTS:
        return jsonify({"error": f"Unknown slot: {slot}"}), 400
    prefix = GLOBAL_EXTRA_SLOTS[slot]["prefix"]
    s = _s()
    inventory = s.get("inventory", {})
    candidates = []
    for sku, data in inventory.items():
        if not sku.startswith(prefix):
            continue
        qty = _inv_qty(data)
        if qty <= 0:
            continue
        name = data.get("name", "") if isinstance(data, dict) else ""
        candidates.append({"sku": sku, "name": name, "qty": qty})
    candidates.sort(key=lambda x: -x["qty"])
    return jsonify(candidates)


@app.route("/api/auto_assign_extras", methods=["POST"])
def auto_assign_extras():
    """Auto-assign all 5 global extra slots using highest-qty SKU of correct type."""
    s = _s()
    inventory = s.get("inventory", {})
    ge = s.get("global_extras", {})
    changes = []

    # Group slots by prefix to avoid picking same SKU for both slots of same type
    prefix_groups = defaultdict(list)
    for slot, meta in GLOBAL_EXTRA_SLOTS.items():
        prefix_groups[meta["prefix"]].append(slot)

    used_by_prefix = defaultdict(set)
    for prefix, slots in prefix_groups.items():
        # Build candidate list sorted by qty desc
        cands = []
        for sku, data in inventory.items():
            if not sku.startswith(prefix):
                continue
            qty = _inv_qty(data)
            if qty > 0:
                cands.append((sku, qty))
        cands.sort(key=lambda x: -x[1])

        for slot in slots:
            old = ge.get(slot, "")
            best = ""
            for sku, qty in cands:
                if sku not in used_by_prefix[prefix]:
                    best = sku
                    break
            if best:
                ge[slot] = best
                used_by_prefix[prefix].add(best)
                if best != old:
                    changes.append(f"{slot}: {old or '(empty)'} -> {best}")
            else:
                ge[slot] = old

    s["global_extras"] = ge
    save_settings(s)
    return jsonify({"changes": changes, "count": len(changes)})


@app.route("/api/suggest_fixes")
def suggest_fixes():
    # Trigger a calculate first
    s = _s()
    wheel_inv = s.get("wheel_inventory", {})
    open_pos = s.get("open_pos", [])

    # We need results - recalculate inline
    calc_resp = calculate()
    calc_data = calc_resp.get_json()

    suggestions = []
    for r in calc_data["results"]:
        if r["status"] != "SHORTAGE":
            continue
        deficit = abs(r["net"])
        fixes = []
        for wsku, wd in wheel_inv.items():
            if isinstance(wd, dict) and wd.get("target_sku") == r["sku"]:
                w = float(wd.get("weight_lbs", 0))
                c = int(wd.get("count", 0))
                p = int(w * c * WHEEL_TO_SLICE_FACTOR)
                if p > 0:
                    fixes.append(f"MFG: Cut {wsku} ({c} wheels = ~{p} units)")
        for po in open_pos:
            if po.get("sku") == r["sku"] and po.get("status") == "Open":
                fixes.append(f"PO: {po.get('qty','?')} units ETA {po.get('eta','?')}")
        if not fixes:
            fixes.append("Recipe swap, partial sub, or Wednesday PO")
        suggestions.append({"sku": r["sku"], "deficit": deficit, "fixes": fixes})

    return jsonify(suggestions)


@app.route("/api/variety_check")
def variety_check():
    s = _s()
    recipes = s.get("curation_recipes", {})
    pr_cjam = s.get("pr_cjam", {})
    cex_ec = s.get("cex_ec", {})
    splits = s.get("cexec_splits", {})

    cur_cheeses = {}
    for cur in CURATION_ORDER:
        ch = set()
        for item in recipes.get(cur, []):
            sku = item[0] if isinstance(item, (list, tuple)) else item
            if isinstance(sku, str) and sku.startswith("CH-"):
                ch.add(sku)
        info = pr_cjam.get(cur, {})
        pr_ch = info.get("cheese", "") if isinstance(info, dict) else ""
        if pr_ch:
            ch.add(pr_ch)
        ec_ch = cex_ec.get(cur, "")
        if ec_ch:
            ch.add(ec_ch)
        for sk in splits.get(cur, {}):
            ch.add(sk)
        cur_cheeses[cur] = ch

    issues = []
    for i, cur in enumerate(CURATION_ORDER):
        for j in range(i + 1, min(i + 3, len(CURATION_ORDER))):
            nb = CURATION_ORDER[j]
            overlap = cur_cheeses.get(cur, set()) & cur_cheeses.get(nb, set())
            for sku in overlap:
                issues.append(f"{sku} in {cur} and {nb} ({j-i} apart)")

    pr_map = defaultdict(list)
    for cur in ALL_CURATIONS:
        info = pr_cjam.get(cur, {})
        ch = info.get("cheese", "") if isinstance(info, dict) else ""
        if ch:
            pr_map[ch].append(cur)
    for ch, curs in pr_map.items():
        if len(curs) > 1:
            issues.append(f"PR-CJAM duplicate: {ch} in {', '.join(curs)}")

    return jsonify(issues)


@app.route("/api/wed_po")
def wed_po():
    """Generate Wednesday PO lines. Uses RMFG data if loaded, else legacy."""
    inv = STATE.get("rmfg_inventory")
    sat_demand = STATE.get("rmfg_sat_demand")
    s = _s()
    vendor_catalog = s.get("vendor_catalog", {})

    # Use RMFG data if available
    bulk_weights = STATE.get("bulk_weights", {})
    if inv and sat_demand:
        all_ch = set(k for k in inv if k.startswith("CH-"))
        all_ch.update(k for k in sat_demand if k.startswith("CH-"))
        shortage_rows = []
        for sku in all_ch:
            avail = inv.get(sku, 0)
            demand = int(round(sat_demand.get(sku, 0)))
            net = avail - demand
            if net < 0:
                # Check if wheels/blocks can cover the deficit
                bw = bulk_weights.get(sku, {})
                potential = bw.get("potential_yield", 0)
                if potential > 0 and avail + potential >= demand:
                    # Wheels cover it — flag as MFG, not PO
                    continue
                # Reduce deficit by potential yield (partial coverage)
                effective_net = net + potential
                shortage_rows.append({
                    "sku": sku, "net": effective_net,
                    "status": "SHORTAGE",
                    "potential": potential,
                })
    else:
        calc_resp = calculate()
        calc_data = calc_resp.get_json()
        shortage_rows = [r for r in calc_data["results"] if r["status"] == "SHORTAGE"]

    lines = []
    for r in shortage_rows:
        deficit = abs(r["net"])
        buf = max(deficit, int(deficit * 1.2))
        vi = vendor_catalog.get(r["sku"], {})
        cq = int(vi.get("case_qty", 1)) or 1
        cases = math.ceil(buf / cq)
        lines.append({
            "sku": r["sku"], "deficit": deficit,
            "order_qty": cases * cq, "cases": cases,
            "case_qty": cq, "vendor": vi.get("vendor", "?"),
        })
    lines.sort(key=lambda x: -x["deficit"])
    return jsonify(lines)


@app.route("/api/order_list")
def order_list():
    """Consolidated order list grouped by product type, not vendor."""
    inv = STATE.get("rmfg_inventory", {})
    s = _s()
    vendor_catalog = s.get("vendor_catalog", {})
    wheel_inv = s.get("wheel_inventory", {})
    open_pos = s.get("open_pos", [])
    inventory = s.get("inventory", {})

    # Gather all results across weeks
    all_shortages = {}  # sku -> total deficit

    # Check multi-week data
    for key in ["rmfg_sat_demand", "rmfg_tue_demand", "rmfg_nsat_demand"]:
        demand = STATE.get(key, {})
        for sku, qty in demand.items():
            q = int(round(qty))
            if q > 0:
                if sku not in all_shortages:
                    all_shortages[sku] = {"demand": 0, "avail": 0}
                all_shortages[sku]["demand"] += q

    # Compute availability from inventory
    bulk_weights = STATE.get("bulk_weights", {})
    for sku in all_shortages:
        avail = inv.get(sku, 0)
        if avail == 0:
            inv_entry = inventory.get(sku, {})
            avail = int(inv_entry.get("qty", 0)) if isinstance(inv_entry, dict) else int(inv_entry or 0)
        # Add wheel/block potential from bulk weights (preferred)
        bw = bulk_weights.get(sku, {})
        if bw.get("potential_yield", 0) > 0:
            avail += bw["potential_yield"]
        else:
            # Fallback to legacy wheel_inventory
            for wsku, wd in wheel_inv.items():
                if isinstance(wd, dict) and wd.get("target_sku") == sku:
                    w = float(wd.get("weight_lbs", 0))
                    c = int(wd.get("count", 0))
                    avail += int(w * c * WHEEL_TO_SLICE_FACTOR)
        # Add open POs
        for po in open_pos:
            if po.get("sku") == sku and po.get("status", "").lower() in ("open", "ordered"):
                avail += int(po.get("qty", 0))
        all_shortages[sku]["avail"] = avail

    # Build order lines for items with net < 0
    lines = []
    for sku, d in all_shortages.items():
        net = d["avail"] - d["demand"]
        if net >= 0:
            continue
        deficit = abs(net)
        # Classify by type
        if sku.startswith("CH-"):
            category = "Cheese"
        elif sku.startswith("AC-"):
            category = "Accompaniment"
        elif sku.startswith("MT-"):
            category = "Meat"
        elif sku.startswith("PR-") or sku.startswith("CEX-"):
            category = "Pairing"
        else:
            category = "Other"

        vi = vendor_catalog.get(sku, {})
        cq = int(vi.get("case_qty", 1)) or 1
        buf = max(deficit, int(deficit * 1.2))
        cases = math.ceil(buf / cq)
        sku_name = ""
        inv_entry = inventory.get(sku, {})
        if isinstance(inv_entry, dict):
            sku_name = inv_entry.get("name", "")

        lines.append({
            "sku": sku,
            "name": sku_name,
            "category": category,
            "demand": d["demand"],
            "avail": d["avail"],
            "deficit": deficit,
            "order_qty": cases * cq,
            "cases": cases,
            "case_qty": cq,
            "vendor": vi.get("vendor", ""),
            "unit_cost": vi.get("unit_cost", ""),
        })

    # Sort by category then deficit
    cat_order = {"Cheese": 0, "Accompaniment": 1, "Meat": 2, "Pairing": 3, "Other": 4}
    lines.sort(key=lambda x: (cat_order.get(x["category"], 9), -x["deficit"]))
    return jsonify(lines)


@app.route("/api/order_list_csv")
def order_list_csv():
    """Export consolidated order list as CSV."""
    resp = order_list()
    lines = resp.get_json()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Category", "SKU", "Name", "Demand", "Available", "Deficit",
                     "Order Qty", "Cases", "Case Qty", "Vendor", "Unit Cost"])
    for r in lines:
        writer.writerow([
            r["category"], r["sku"], r["name"], r["demand"], r["avail"],
            r["deficit"], r["order_qty"], r["cases"], r["case_qty"],
            r["vendor"], r["unit_cost"],
        ])

    buf = io.BytesIO(output.getvalue().encode("utf-8-sig"))
    today = datetime.date.today().strftime("%Y%m%d")
    return send_file(buf, mimetype="text/csv", as_attachment=True,
                     download_name=f"order_list_{today}.csv")


## email_po moved to "Email Wednesday PO" section below


@app.route("/api/import_csv", methods=["POST"])
def import_csv():
    if "file" not in request.files:
        return jsonify({"error": "No file"}), 400

    f = request.files["file"]
    content = f.read().decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(content))
    headers = reader.fieldnames or []

    s = _s()
    pr_cjam = s.get("pr_cjam", {})
    cex_ec = s.get("cex_ec", {})
    splits = s.get("cexec_splits", {})

    imported = defaultdict(int)
    rows = 0

    if "All SKUs" in headers:
        for row in reader:
            rows += 1
            for sku_raw in row.get("All SKUs", "").split(","):
                sku = sku_raw.strip().upper()
                if not sku:
                    continue
                if sku.startswith("CH-"):
                    imported[sku] += 1
                elif sku.startswith("PR-CJAM-"):
                    suffix = sku.split("PR-CJAM-", 1)[1]
                    info = pr_cjam.get(suffix, {})
                    ch = info.get("cheese", "") if isinstance(info, dict) else ""
                    if ch:
                        imported[ch] += 1
                elif sku.startswith("CEX-EC-"):
                    suffix = sku.split("CEX-EC-", 1)[1]
                    sp = splits.get(suffix, {})
                    if sp:
                        for sk, ratio in sp.items():
                            imported[sk] += max(1, int(float(ratio)))
                    else:
                        ch = cex_ec.get(suffix, "")
                        if ch:
                            imported[ch] += 1
                elif sku.startswith("EX-EC-"):
                    suffix = sku.split("EX-EC-", 1)[1]
                    ch = cex_ec.get(suffix, "")
                    if ch:
                        imported[ch] += 1

    elif "line_item_sku" in headers:
        for row in reader:
            rows += 1
            sku = (row.get("line_item_sku") or "").strip().upper()
            try:
                qty = int(float(row.get("line_item_quantity", 1)))
            except (ValueError, TypeError):
                qty = 1
            if sku.startswith("CH-"):
                imported[sku] += qty
            elif sku.startswith("PR-CJAM-"):
                suffix = sku.split("PR-CJAM-", 1)[1]
                info = pr_cjam.get(suffix, {})
                ch = info.get("cheese", "") if isinstance(info, dict) else ""
                if ch:
                    imported[ch] += qty
            elif sku.startswith("CEX-EC-"):
                suffix = sku.split("CEX-EC-", 1)[1]
                ch = cex_ec.get(suffix, "")
                if ch:
                    imported[ch] += qty
    else:
        return jsonify({"error": f"Unrecognized CSV format. Headers: {headers}"}), 400

    STATE["csv_demand"] = dict(imported)
    return jsonify({
        "rows": rows,
        "skus": len([k for k in imported if k.startswith("CH-")]),
        "units": sum(imported.values()),
    })


@app.route("/api/export_csv")
def export_csv():
    """Export NET report as CSV. Uses RMFG data if loaded."""
    inv = STATE.get("rmfg_inventory")
    if inv:
        with app.test_request_context(json={}):
            calc_resp = calculate_rmfg()
            data = calc_resp.get_json() if hasattr(calc_resp, 'get_json') \
                else json.loads(calc_resp.data)
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["SKU", "Available", "Sat Demand", "NET Sat",
                         "Tue Demand", "NET Tue", "Next Sat Demand",
                         "NET Final", "Total Demand", "Status"])
        for r in data["results"]:
            writer.writerow([
                r["sku"], r["available"],
                r.get("sat_demand", 0), r.get("net_sat", r["net"]),
                r.get("tue_demand", 0), r.get("net_tue", 0),
                r.get("next_sat_demand", 0), r.get("net_final", 0),
                r["total_demand"], r["status"],
            ])
    else:
        calc_resp = calculate()
        data = calc_resp.get_json()
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["SKU", "Available", "Direct", "PRCJAM", "CEXEC",
                         "EXEC", "Total Demand", "NET", "Status"])
        for r in data["results"]:
            writer.writerow([r["sku"], r["available"], r["direct"],
                             r["prcjam"], r["cexec"], r["exec"],
                             r["total_demand"], r["net"], r["status"]])

    mem = io.BytesIO(output.getvalue().encode("utf-8"))
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    return send_file(mem, mimetype="text/csv", as_attachment=True,
                     download_name=f"fulfillment_net_{ts}.csv")


@app.route("/api/split", methods=["POST"])
def set_split():
    data = request.json
    cur = data["curation"]
    splits = data.get("splits", {})
    s = _s()
    if not s.get("cexec_splits"):
        s["cexec_splits"] = {}
    if splits:
        s["cexec_splits"][cur] = splits
    else:
        s["cexec_splits"].pop(cur, None)
    save_settings(s)
    return jsonify({"ok": True})


# ══════════════════════════════════════════════════════════════════════════
#  AUTOMATION: Full RMFG folder loading + multi-window demand + substitutions
# ══════════════════════════════════════════════════════════════════════════

# ── SKU helpers (from weekly_demand_report.py) ─────────────────────────

EQUIV = {"CH-BRIE": "CH-EBRIE"}
SKIP_PREFIXES = ("AHB-", "BL-", "PK-", "TR-", "EX-")
KNOWN_CURATIONS = {
    "MONG", "MDT", "OWC", "SPN", "ALPN", "ALPT",
    "ISUN", "HHIGH", "NMS", "BYO", "SS", "GEN", "MS",
}
_MONTHLY_PATTERNS = {"AHB-MED", "AHB-LGE", "AHB-CMED", "AHB-CUR-MS", "AHB-BVAL",
                     "AHB-MCUST-MS", "AHB-MCUST-NMS"}


def normalize_sku(sku):
    return EQUIV.get(sku, sku)


def resolve_curation_from_box_sku(sku):
    if not sku:
        return None
    sku = sku.strip().upper()
    if sku in _MONTHLY_PATTERNS:
        return "MONTHLY"
    if "MCUST-NMS" in sku:
        return "NMS"
    if "MCUST-MS" in sku or "CUR-MS" in sku or "BVAL" in sku:
        return "MS"
    for cur in KNOWN_CURATIONS:
        if cur in sku:
            return cur
    return None


def is_pickable(sku):
    upper = sku.upper()
    if any(upper.startswith(p) for p in SKIP_PREFIXES):
        return False
    if upper.startswith("PR-CJAM"):
        return False
    if upper.startswith("CEX-"):
        return False
    return bool(sku.strip())


def resolve_pr_cjam(suffix):
    s = _s()
    pr_cjam = s.get("pr_cjam", {})
    return resolve_pr_cjam_with(suffix, pr_cjam)


def resolve_cex_ec(suffix):
    s = _s()
    cex_ec = s.get("cex_ec", {})
    splits = s.get("cexec_splits", {})
    return resolve_cex_ec_with(suffix, cex_ec, splits)


def resolve_pr_cjam_with(suffix, pr_cjam_dict):
    """Pure version — takes explicit assignment dict, not settings."""
    info = pr_cjam_dict.get(suffix, {})
    ch = info.get("cheese", "") if isinstance(info, dict) else str(info)
    return {ch: 1} if ch else {}


def resolve_cex_ec_with(suffix, cex_ec_dict, splits_dict):
    """Pure version — takes explicit assignment dict, not settings."""
    sp = splits_dict.get(suffix, {})
    if sp:
        return dict(sp)
    ch = cex_ec_dict.get(suffix, "")
    return {ch: 1} if ch else {}


def resolve_demand(direct_demand, prcjam_counts, cexec_counts, pr_cjam, cex_ec, splits):
    """Resolve raw demand components using provided assignments.
    Returns (demand_dict, attribution_dict).
    - direct_demand: {sku: qty} — cheese from recipes + custom picks
    - prcjam_counts: {curation_suffix: qty}
    - cexec_counts: {curation_suffix: qty}
    - attribution: {sku: {direct: N, prcjam: {cur: N}, cexec: {cur: N}}}
    """
    demand = defaultdict(float)
    attribution = defaultdict(lambda: {"direct": 0, "prcjam": {}, "cexec": {}})

    # Direct demand
    for sku, qty in direct_demand.items():
        demand[sku] += qty
        attribution[sku]["direct"] += qty

    # PR-CJAM resolution
    for suffix, count in prcjam_counts.items():
        resolved = resolve_pr_cjam_with(suffix, pr_cjam)
        for rsku, rqty in resolved.items():
            rsku = normalize_sku(rsku)
            amt = rqty * count
            demand[rsku] += amt
            attribution[rsku]["prcjam"][suffix] = attribution[rsku]["prcjam"].get(suffix, 0) + amt

    # CEX-EC resolution
    for suffix, count in cexec_counts.items():
        if suffix == "BARE":
            continue
        resolved = resolve_cex_ec_with(suffix, cex_ec, splits)
        for rsku, rqty in resolved.items():
            rsku = normalize_sku(rsku)
            amt = rqty * count
            demand[rsku] += amt
            attribution[rsku]["cexec"][suffix] = attribution[rsku]["cexec"].get(suffix, 0) + amt

    resolved_demand = {sku: int(round(q)) for sku, q in demand.items() if q > 0}
    # Clean up attribution — convert defaultdict to plain dict
    attr_out = {}
    for sku in resolved_demand:
        a = attribution[sku]
        attr_out[sku] = {
            "direct": int(round(a["direct"])),
            "prcjam": {k: int(round(v)) for k, v in a["prcjam"].items()},
            "cexec": {k: int(round(v)) for k, v in a["cexec"].items()},
        }
    return resolved_demand, attr_out


# ── Ship Tag Helpers ──────────────────────────────────────────────────


def parse_ship_tag(tags_str: str):
    """Extract _SHIP_YYYY-MM-DD Monday date from order tags. Returns date or None."""
    m = re.search(r'_SHIP_(\d{4}-\d{2}-\d{2})', tags_str or "")
    if m:
        try:
            return datetime.date.fromisoformat(m.group(1))
        except (ValueError, TypeError):
            return None
    return None


def ship_tag_windows(ship_monday):
    """Given a _SHIP_ Monday date, return (saturday, tuesday) fulfillment dates."""
    saturday = ship_monday - datetime.timedelta(days=2)
    tuesday = ship_monday + datetime.timedelta(days=1)
    return saturday, tuesday


def current_ship_monday(ref_date=None):
    """Get the _SHIP_ Monday for the current fulfillment cycle.
    The cycle starts on the Saturday before the Monday."""
    if ref_date is None:
        ref_date = datetime.date.today()
    weekday = ref_date.weekday()  # 0=Mon .. 6=Sun
    if weekday <= 2:
        # Mon-Wed: we're in this Monday's cycle
        return ref_date - datetime.timedelta(days=weekday)
    else:
        # Thu-Sun: we're preparing for next Monday's cycle
        return ref_date + datetime.timedelta(days=(7 - weekday))


def classify_order_window(ship_monday, ref_date=None):
    """Classify whether an order with this ship tag is Saturday or Tuesday demand.
    Returns 'saturday', 'tuesday', 'next_saturday', or 'future'."""
    if ref_date is None:
        ref_date = datetime.date.today()
    current_monday = current_ship_monday(ref_date)
    if ship_monday == current_monday:
        # Current cycle — which fulfillment day depends on timing
        # After Monday 5am, Saturday is fulfilled; remaining = Tuesday
        if ref_date.weekday() >= 0 and ref_date >= current_monday:
            return "tuesday"
        return "saturday"
    elif ship_monday == current_monday + datetime.timedelta(days=7):
        return "next_saturday"
    elif ship_monday > current_monday:
        return "future"
    return "saturday"  # past tag, treat as Saturday


# ── Depletion File Parser ─────────────────────────────────────────────


def parse_depletion_xlsx(path):
    """Parse AHB_WeeklyProductionQuery XLSX depletion file.
    Returns {product_name: total_qty} and order count."""
    try:
        import openpyxl
    except ImportError as exc:
        import traceback
        traceback.print_exc()
        return {}, 0, f"openpyxl import failed: {exc}"

    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    # Find AHB product columns
    product_cols = []
    for i, h in enumerate(headers):
        if h and "AHB (S_REG):" in str(h):
            name = str(h).split(": ", 1)[1].strip() if ": " in str(h) else str(h)
            product_cols.append((i, name))

    # Sum quantities per product
    totals = defaultdict(int)
    order_count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        order_count += 1
        for idx, name in product_cols:
            val = row[idx] if idx < len(row) else None
            if val and isinstance(val, (int, float)) and val > 0:
                totals[name] += int(val)

    wb.close()
    return dict(totals), order_count, None


def map_depletion_to_skus(product_totals, sku_translations, inventory):
    """Map depletion product names to SKUs.
    Returns {sku: qty}, {product_name: sku}, [unmatched_products].
    Uses sku_translations (exact), inventory names (fuzzy), and difflib."""
    import difflib

    # Build reverse lookup: normalized inventory name -> SKU
    inv_names = {}
    for sku, data in inventory.items():
        if not isinstance(data, dict):
            continue
        name = data.get("name", "")
        if not name:
            continue
        # Strip category prefix: "Cheese Slice, X" -> "X"
        # "Crackers, X" -> "X", "Jam, X" -> "X", etc.
        parts = name.split(", ", 1)
        short = parts[1].strip() if len(parts) > 1 else name
        inv_names[short.lower()] = sku
        inv_names[name.lower()] = sku

    mapped = {}
    sku_totals = defaultdict(int)
    unmatched = []

    for product, qty in product_totals.items():
        # Skip non-depletable items (tasting guides, etc.)
        if "tasting guide" in product.lower():
            continue

        sku = None
        # 1. Exact translation
        if product in sku_translations:
            sku = sku_translations[product]
        # 2. Exact inventory name match
        elif product.lower() in inv_names:
            sku = inv_names[product.lower()]
        else:
            # 3. Fuzzy match against inventory names
            candidates = list(inv_names.keys())
            matches = difflib.get_close_matches(product.lower(), candidates,
                                                 n=1, cutoff=0.6)
            if matches:
                sku = inv_names[matches[0]]

        if sku:
            mapped[product] = sku
            sku_totals[sku] += qty
        else:
            unmatched.append({"product": product, "qty": qty})

    return dict(sku_totals), mapped, unmatched


# ── File detection ────────────────────────────────────────────────────

def detect_rmfg_files(folder):
    """Auto-detect RMFG data files in a folder."""
    files = os.listdir(folder)
    result = {
        "template_check": None,
        "product_inventory": None,
        "order_dashboard": None,
        "charges_queued": None,
        "march_charges": None,
    }
    for f in files:
        fl = f.lower()
        fp = os.path.join(folder, f)
        if not f.endswith(".csv"):
            continue
        if "template check" in fl or "template_check" in fl:
            result["template_check"] = fp
        elif "product inventory" in fl or "product_inventory" in fl:
            result["product_inventory"] = fp
        elif "order-dashboard" in fl or "order_dashboard" in fl:
            result["order_dashboard"] = fp
        elif fl.startswith("charges_queued"):
            result["charges_queued"] = fp
        elif "march charges" in fl or "march_charges" in fl:
            result["march_charges"] = fp
    return result


# ── Inventory loading ─────────────────────────────────────────────────

def load_inventory_from_files(template_path, product_inv_path,
                              po_additions=None, incoming=None,
                              corrections=None):
    """Load inventory from Template Check + Product Inventory fallback."""
    inv = {}

    # Primary: Template Check
    if template_path and os.path.exists(template_path):
        with open(template_path, encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            header = next(reader, None)
            for row in reader:
                if len(row) < 2:
                    continue
                sku = row[0].strip()
                if not sku:
                    continue
                try:
                    qty = int(float(row[1]))
                except ValueError:
                    qty = 0
                inv[sku] = qty

    # PO additions
    if po_additions:
        for sku, add_qty in po_additions.items():
            inv[sku] = inv.get(sku, 0) + add_qty

    # Incoming
    if incoming:
        for sku, add_qty in incoming.items():
            inv[sku] = inv.get(sku, 0) + add_qty

    # Corrections (override)
    if corrections:
        for sku, qty in corrections.items():
            inv[sku] = qty

    # Fallback: Product Inventory
    if product_inv_path and os.path.exists(product_inv_path):
        fallback = {}
        with open(product_inv_path, encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            for row in reader:
                if len(row) < 7:
                    continue
                sku = row[1].strip()
                if not sku or sku == "Product SKU":
                    continue
                try:
                    qty = int(float(row[6]))
                except (ValueError, IndexError):
                    qty = 0
                if sku not in fallback:
                    fallback[sku] = qty
                else:
                    fallback[sku] += qty
        for sku, qty in fallback.items():
            if sku not in inv and qty > 0:
                inv[sku] = qty

    # Apply equivalences
    for old, new in EQUIV.items():
        if old in inv:
            inv[new] = inv.get(new, 0) + inv.pop(old)

    return inv


# ── Demand parsers ────────────────────────────────────────────────────

def parse_order_dashboard(path):
    """Parse Shopify order-dashboard. Returns demand dicts + counts.
    Splits demand by _SHIP_ tag into Saturday vs Tuesday windows.
    Returns raw unresolved components alongside resolved demand."""
    all_demand = defaultdict(int)
    sat_demand = defaultdict(int)
    tue_demand = defaultdict(int)
    first_order_demand = defaultdict(int)
    recurring_demand = defaultdict(int)
    prcjam_counts = defaultdict(int)
    cexec_counts = defaultdict(int)
    direct_demand = defaultdict(int)  # cheese without PR-CJAM/CEX-EC
    first_order_prcjam = defaultdict(int)  # first-order PR-CJAM by curation
    first_order_cexec = defaultdict(int)  # first-order CEX-EC by curation
    monthly_box_order_counts = defaultdict(int)  # AHB-MED/CMED/LGE order counts
    first_order_count = 0
    recurring_count = 0
    total_order_count = 0
    ship_tags_found = defaultdict(int)

    with open(path, encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            total_order_count += 1
            tags = row.get("Order Tags", "")
            is_first = "Subscription First Order" in tags
            is_recurring = "Subscription Recurring Order" in tags
            if is_first:
                first_order_count += 1
            if is_recurring:
                recurring_count += 1

            # Track ship tags but always route to sat_demand —
            # the order dashboard CSV IS the Saturday fulfillment batch.
            # Time-dependent classify_order_window is only for live Shopify sync.
            ship_monday = parse_ship_tag(tags)
            if ship_monday:
                ship_tags_found[ship_monday.isoformat()] += 1
            target = sat_demand

            all_skus = row.get("All SKUs", "")
            skus = [s.strip() for s in all_skus.split(",") if s.strip()]

            # Count monthly box orders
            for sku in skus:
                u = sku.strip().upper()
                if u in ("AHB-MED", "AHB-CMED", "AHB-LGE"):
                    monthly_box_order_counts[u] += 1

            for sku in skus:
                upper = sku.upper()

                if upper.startswith("PR-CJAM-"):
                    suffix = upper.split("PR-CJAM-", 1)[1]
                    prcjam_counts[suffix] += 1
                    if is_first:
                        first_order_prcjam[suffix] += 1
                    # Still resolve for backward-compat sat_demand
                    resolved = resolve_pr_cjam(suffix)
                    for rsku, rqty in resolved.items():
                        rsku = normalize_sku(rsku)
                        all_demand[rsku] += rqty
                        target[rsku] += rqty
                        if is_first:
                            first_order_demand[rsku] += rqty
                        if is_recurring:
                            recurring_demand[rsku] += rqty
                    continue

                if upper.startswith("CEX-EC-"):
                    suffix = upper.split("CEX-EC-", 1)[1]
                    cexec_counts[suffix] += 1
                    if is_first:
                        first_order_cexec[suffix] += 1
                    resolved = resolve_cex_ec(suffix)
                    for rsku, rqty in resolved.items():
                        rsku = normalize_sku(rsku)
                        all_demand[rsku] += rqty
                        target[rsku] += rqty
                        if is_first:
                            first_order_demand[rsku] += rqty
                        if is_recurring:
                            recurring_demand[rsku] += rqty
                    continue

                if upper == "CEX-EC":
                    cexec_counts["BARE"] += 1
                    continue

                # Resolve global extras (bare EX-EC, CEX-EM, EX-EM, etc.)
                ge_resolved = _s().get("global_extras", {}).get(upper)
                if ge_resolved:
                    rsku = normalize_sku(ge_resolved)
                    all_demand[rsku] += 1
                    target[rsku] += 1
                    if is_first:
                        first_order_demand[rsku] += 1
                    if is_recurring:
                        recurring_demand[rsku] += 1
                    continue

                if not is_pickable(sku):
                    continue

                sku = normalize_sku(sku)
                all_demand[sku] += 1
                target[sku] += 1
                direct_demand[sku] += 1
                if is_first:
                    first_order_demand[sku] += 1
                if is_recurring:
                    recurring_demand[sku] += 1

    return {
        "all_demand": dict(all_demand),
        "sat_demand": dict(sat_demand),
        "tue_demand": dict(tue_demand),
        "first_order_demand": dict(first_order_demand),
        "recurring_demand": dict(recurring_demand),
        "first_order_count": first_order_count,
        "recurring_count": recurring_count,
        "total_orders": total_order_count,
        "prcjam_counts": dict(prcjam_counts),
        "cexec_counts": dict(cexec_counts),
        "ship_tags": dict(ship_tags_found),
        # Raw unresolved components for deferred resolution
        "direct_demand": dict(direct_demand),
        "first_order_prcjam": dict(first_order_prcjam),
        "first_order_cexec": dict(first_order_cexec),
        "monthly_box_counts": dict(monthly_box_order_counts),
    }


def parse_charges_queued(path, target_date=None):
    """Parse Recharge charges_queued CSV.
    Returns ({sku: qty}, bare_skipped, raw_components).
    raw_components = {direct_demand, prcjam_counts, cexec_counts}."""
    demand = defaultdict(float)
    direct_demand = defaultdict(float)
    prcjam_counts = defaultdict(float)
    cexec_counts = defaultdict(float)
    charges_by_id = defaultdict(list)
    bare_skipped = 0

    with open(path, encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            scheduled = row.get("scheduled_at", "")
            if target_date and target_date not in scheduled:
                continue
            cid = row.get("charge_id", "")
            sku = row.get("line_item_sku", "").strip()
            try:
                qty = int(float(row.get("line_item_quantity", "1") or "1"))
            except ValueError:
                qty = 1
            if sku:
                charges_by_id[cid].append((sku, qty))

    monthly_box_order_counts = defaultdict(int)
    for cid, items in charges_by_id.items():
        box_sku = None
        for sku, _ in items:
            if sku.upper().startswith("AHB-"):
                box_sku = sku.upper()
                if box_sku in ("AHB-MED", "AHB-CMED", "AHB-LGE"):
                    monthly_box_order_counts[box_sku] += 1
                break
        curation = resolve_curation_from_box_sku(box_sku)

        for sku, qty in items:
            upper = sku.upper()

            if upper.startswith("PR-CJAM-"):
                suffix = upper.split("PR-CJAM-", 1)[1]
                if suffix == "GEN":
                    if curation and curation not in ("MONTHLY", None):
                        suffix = curation
                    else:
                        continue
                prcjam_counts[suffix] += qty
                # Still resolve for backward-compat demand
                resolved = resolve_pr_cjam(suffix)
                for rsku, rqty in resolved.items():
                    demand[normalize_sku(rsku)] += rqty * qty
                continue

            if upper == "CEX-EC":
                bare_skipped += qty
                continue

            if upper.startswith("CEX-EC-"):
                suffix = upper.split("CEX-EC-", 1)[1]
                cexec_counts[suffix] += qty
                resolved = resolve_cex_ec(suffix)
                for rsku, rqty in resolved.items():
                    demand[normalize_sku(rsku)] += rqty * qty
                continue

            # Resolve global extras (bare EX-EC, CEX-EM, EX-EM, etc.)
            ge_resolved = _s().get("global_extras", {}).get(upper)
            if ge_resolved:
                demand[normalize_sku(ge_resolved)] += qty
                continue

            if not is_pickable(sku):
                continue

            nsku = normalize_sku(sku)
            demand[nsku] += qty
            direct_demand[nsku] += qty

    raw = {
        "direct_demand": {k: int(round(v)) for k, v in direct_demand.items()},
        "prcjam_counts": {k: int(round(v)) for k, v in prcjam_counts.items()},
        "cexec_counts": {k: int(round(v)) for k, v in cexec_counts.items()},
        "monthly_box_counts": dict(monthly_box_order_counts),
    }
    return {sku: int(round(q)) for sku, q in demand.items()}, bare_skipped, raw


def parse_march_charges(path, start_day, end_day, year=2026, month=3):
    """Parse MARCH CHARGES for a date range. Returns {sku: qty}, charge_count."""
    demand = defaultdict(float)
    charges_by_id = defaultdict(list)
    bare_skipped = 0

    with open(path, encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            scheduled = row.get("scheduled_at", "")
            date_part = scheduled.split(" ")[0]
            parts = date_part.split("/")
            if len(parts) != 3:
                continue
            try:
                m, d, y = int(parts[0]), int(parts[1]), int(parts[2])
            except ValueError:
                continue
            if not (y == year and m == month and start_day <= d <= end_day):
                continue

            cid = row.get("charge_id", "")
            sku = row.get("line_item_sku", "").strip()
            try:
                qty = int(float(row.get("line_item_quantity", "1") or "1"))
            except ValueError:
                qty = 1
            if sku:
                charges_by_id[cid].append((sku, qty))

    for cid, items in charges_by_id.items():
        box_sku = None
        for sku, _ in items:
            if sku.upper().startswith("AHB-"):
                box_sku = sku.upper()
                break
        curation = resolve_curation_from_box_sku(box_sku)

        has_pr_cjam = False
        for sku, qty in items:
            upper = sku.upper()

            if upper.startswith("PR-CJAM-"):
                suffix = upper.split("PR-CJAM-", 1)[1]
                has_pr_cjam = True
                if suffix == "GEN":
                    if curation and curation not in ("MONTHLY", None):
                        resolved = resolve_pr_cjam(curation)
                    else:
                        continue
                else:
                    resolved = resolve_pr_cjam(suffix)
                for rsku, rqty in resolved.items():
                    demand[normalize_sku(rsku)] += rqty * qty
                continue

            if upper == "CEX-EC":
                bare_skipped += qty
                continue

            if upper.startswith("CEX-EC-"):
                suffix = upper.split("CEX-EC-", 1)[1]
                resolved = resolve_cex_ec(suffix)
                for rsku, rqty in resolved.items():
                    demand[normalize_sku(rsku)] += rqty * qty
                continue

            # Resolve global extras (bare EX-EC, CEX-EM, EX-EM, etc.)
            ge_resolved = _s().get("global_extras", {}).get(upper)
            if ge_resolved:
                demand[normalize_sku(ge_resolved)] += qty
                continue

            if not is_pickable(sku):
                continue

            demand[normalize_sku(sku)] += qty

        if not has_pr_cjam and curation and curation not in ("MONTHLY", None):
            for rsku, rqty in resolve_pr_cjam(curation).items():
                demand[normalize_sku(rsku)] += rqty

    return {sku: int(round(q)) for sku, q in demand.items()}, len(charges_by_id)


# ── Load RMFG folder endpoint ─────────────────────────────────────────

@app.route("/api/load_rmfg", methods=["POST"])
def load_rmfg():
    """Load all data from an RMFG folder. Auto-detects files."""
    data = request.json or {}
    folder = data.get("folder", "")

    # Allow relative path from project root
    if not os.path.isabs(folder):
        folder = os.path.join(_get_project_dir(), folder)

    if not os.path.isdir(folder):
        return jsonify({"error": f"Folder not found: {folder}"}), 400

    files = detect_rmfg_files(folder)
    log_lines = []
    warnings = []

    # PO/incoming/corrections from request (no hardcoded defaults)
    po_additions = data.get("po_additions", {})
    incoming = data.get("incoming", {})
    corrections = data.get("corrections", {})

    # 1. Load inventory
    inv = load_inventory_from_files(
        files["template_check"], files["product_inventory"],
        po_additions, incoming, corrections,
    )
    ch_count = sum(1 for k in inv if k.startswith("CH-"))
    log_lines.append(f"Inventory: {len(inv)} SKUs ({ch_count} cheese)")

    # 2. Parse Saturday demand (order dashboard + charges_queued)
    sat_demand = defaultdict(int)
    dashboard_info = {}
    if files["order_dashboard"]:
        dashboard_info = parse_order_dashboard(files["order_dashboard"])
        for sku, qty in dashboard_info.get("sat_demand", dashboard_info.get("all_demand", {})).items():
            sat_demand[sku] += qty
        ship_tags = dashboard_info.get("ship_tags", {})
        tag_info = f", tags: {', '.join(ship_tags.keys())}" if ship_tags else ""
        log_lines.append(
            f"Dashboard: {dashboard_info['total_orders']} orders "
            f"({dashboard_info['first_order_count']} first){tag_info}"
        )
    else:
        warnings.append("No order-dashboard CSV found")

    charges_bare = 0
    rc_raw = {"direct_demand": {}, "prcjam_counts": {}, "cexec_counts": {}}
    if files["charges_queued"]:
        # Auto-detect target date from filename
        fname = os.path.basename(files["charges_queued"])
        # charges_queued-2026.03.06-... → target date 2026-03-07 (next day)
        # But actually we want all charges in the file if they're for Sat
        # Try without date filter first, then with
        queued, charges_bare, rc_raw = parse_charges_queued(files["charges_queued"])
        for sku, qty in queued.items():
            sat_demand[sku] += qty
        log_lines.append(f"Charges queued: {sum(queued.values())} items")
        if charges_bare:
            warnings.append(f"{charges_bare} bare CEX-EC skipped")
    else:
        warnings.append("No charges_queued CSV found")

    # 3. Tuesday demand (from ship tag split, or fallback to first orders × multiplier)
    tue_demand = {}
    dashboard_tue = dashboard_info.get("tue_demand", {})
    if dashboard_tue:
        tue_demand = dict(dashboard_tue)
        log_lines.append(
            f"Tuesday: {sum(dashboard_tue.values())} units from ship tag split"
        )
    elif dashboard_info.get("first_order_demand"):
        # Prefer Shopify rolling average over static multiplier
        s = _s()
        shopify_fo = s.get("shopify_first_order_demand", {})
        fo_overrides = s.get("first_order_overrides", {})
        if shopify_fo or fo_overrides:
            # Use rolling average with manual overrides
            merged = dict(shopify_fo)
            merged.update(fo_overrides)
            tue_demand = {
                sku: int(round(merged.get(sku, 0)))
                for sku in dashboard_info["first_order_demand"]
                if merged.get(sku, 0) > 0
            }
            log_lines.append(
                f"Tuesday estimate: Shopify rolling avg "
                f"({len(merged)} SKUs, {len(fo_overrides)} overrides)"
            )
        else:
            # Legacy fallback: x3 multiplier
            proj = s.get("first_order_projection", {})
            multiplier = proj.get("multiplier", 3) if proj.get("enabled", True) else 3
            tue_demand = {
                sku: qty * multiplier
                for sku, qty in dashboard_info["first_order_demand"].items()
            }
            log_lines.append(
                f"Tuesday estimate: {dashboard_info['first_order_count']} "
                f"first orders × {multiplier} (no Shopify data)"
            )

    # 4. Next Saturday demand (MARCH CHARGES)
    next_sat_demand = {}
    next_sat_charges = 0
    if files["march_charges"]:
        next_sat_demand, next_sat_charges = parse_march_charges(
            files["march_charges"], 8, 14
        )
        log_lines.append(
            f"Next Saturday: {next_sat_charges} charges (3/8-3/14)"
        )
    else:
        warnings.append("No MARCH CHARGES CSV found")

    # Store in state
    STATE["rmfg_inventory"] = inv
    STATE["rmfg_sat_demand"] = dict(sat_demand)
    STATE["rmfg_tue_demand"] = tue_demand
    STATE["rmfg_next_sat_demand"] = next_sat_demand
    STATE["rmfg_dashboard"] = dashboard_info
    STATE["rmfg_folder"] = folder
    STATE["rmfg_files"] = {k: os.path.basename(v) if v else None
                           for k, v in files.items()}

    # Store raw unresolved components for deferred resolution
    # Merge Shopify (dashboard) + Recharge (charges_queued) raw components
    sh_direct = dashboard_info.get("direct_demand", {})
    sh_prcjam = dashboard_info.get("prcjam_counts", {})
    sh_cexec = dashboard_info.get("cexec_counts", {})

    merged_direct = defaultdict(int)
    merged_prcjam = defaultdict(int)
    merged_cexec = defaultdict(int)
    for sku, q in sh_direct.items():
        merged_direct[sku] += q
    for sku, q in rc_raw.get("direct_demand", {}).items():
        merged_direct[sku] += q
    for cur, q in sh_prcjam.items():
        merged_prcjam[cur] += q
    for cur, q in rc_raw.get("prcjam_counts", {}).items():
        merged_prcjam[cur] += q
    for cur, q in sh_cexec.items():
        merged_cexec[cur] += q
    for cur, q in rc_raw.get("cexec_counts", {}).items():
        merged_cexec[cur] += q

    STATE["rmfg_direct_sat"] = dict(merged_direct)
    STATE["rmfg_prcjam_sat"] = dict(merged_prcjam)
    STATE["rmfg_cexec_sat"] = dict(merged_cexec)

    # Merge monthly box order counts from both sources
    merged_monthly_box = defaultdict(int)
    for bt, cnt in dashboard_info.get("monthly_box_counts", {}).items():
        merged_monthly_box[bt] += cnt
    for bt, cnt in rc_raw.get("monthly_box_counts", {}).items():
        merged_monthly_box[bt] += cnt
    STATE["rmfg_monthly_box_counts"] = dict(merged_monthly_box)

    # Persist to settings so quantities survive restart
    s = _s()
    s["rmfg_prcjam_sat"] = dict(merged_prcjam)
    s["rmfg_cexec_sat"] = dict(merged_cexec)
    s["rmfg_monthly_box_counts"] = dict(merged_monthly_box)
    save_settings(s)
    # Per-source raw for attribution
    STATE["rmfg_direct_rc"] = rc_raw.get("direct_demand", {})
    STATE["rmfg_prcjam_rc"] = rc_raw.get("prcjam_counts", {})
    STATE["rmfg_cexec_rc"] = rc_raw.get("cexec_counts", {})
    STATE["rmfg_direct_sh"] = dict(sh_direct)
    STATE["rmfg_prcjam_sh"] = dict(sh_prcjam)
    STATE["rmfg_cexec_sh"] = dict(sh_cexec)
    # First-order counts for projection
    STATE["rmfg_first_order_prcjam"] = dashboard_info.get("first_order_prcjam", {})
    STATE["rmfg_first_order_cexec"] = dashboard_info.get("first_order_cexec", {})

    # Auto-detect meal-type-export CSV for SKU translations
    import glob as globmod
    base_dir = _get_project_dir()
    mte_pattern = os.path.join(base_dir, "Shipments", "meal-type-export*.csv")
    mte_files = sorted(globmod.glob(mte_pattern), key=os.path.getmtime, reverse=True)
    if mte_files:
        try:
            s = _s()
            translations = s.setdefault("sku_translations", {})
            mte_count = 0
            with open(mte_files[0], encoding="utf-8-sig") as f:
                reader = csv.reader(f)
                for row in reader:
                    if len(row) >= 2:
                        sku_code = row[0].strip()
                        product_name = row[1].strip()
                        # Strip AHB (S_REG): prefix
                        if ": " in product_name:
                            product_name = product_name.split(": ", 1)[1]
                        if product_name and sku_code:
                            translations[product_name] = sku_code
                            mte_count += 1
            save_settings(s)
            log_lines.append(f"SKU translations: {mte_count} from meal-type-export")
        except Exception as e:
            warnings.append(f"meal-type-export parse error: {e}")

    # Auto-snapshot on RMFG folder load
    folder_name = os.path.basename(folder)
    _take_snapshot(f"RMFG: {folder_name}", source="rmfg_folder")

    return jsonify({
        "ok": True,
        "files": STATE["rmfg_files"],
        "log": log_lines,
        "warnings": warnings,
        "inventory_count": len(inv),
        "cheese_count": ch_count,
        "sat_skus": len(sat_demand),
        "sat_units": sum(sat_demand.values()),
    })


# ── Demand Mode (Discrete vs Churned) ─────────────────────────────────


def apply_churn_to_demand(demand, recurring_demand, churn_rates, weeks_out=0):
    """Apply churn reduction to subscription portion of demand.

    For orders already placed (weeks_out=0), minimal churn (~2% cancellations).
    For projected demand (weeks_out>0), compound churn based on settings.

    Args:
        demand: {sku: qty} total demand
        recurring_demand: {sku: qty} subscription-recurring subset
        churn_rates: settings churn_rates dict
        weeks_out: how many weeks until fulfillment (0=this week)

    Returns: {sku: adjusted_qty}
    """
    if weeks_out <= 0:
        # Current batch: apply a small cancellation buffer (2%)
        cancel_rate = 0.02
    else:
        # Use average monthly churn across tracks, convert to weekly
        monthly_rates = []
        for track, rates in churn_rates.items():
            if isinstance(rates, dict):
                # Use month_2_plus or month_3_plus as the steady-state rate
                r = rates.get("month_2_plus", rates.get("month_3_plus",
                    rates.get("month_1", 0.12)))
                monthly_rates.append(r)
        avg_monthly = sum(monthly_rates) / len(monthly_rates) if monthly_rates else 0.12
        # Compound weekly: (1 - monthly)^(weeks/4.33)
        weekly_retention = (1 - avg_monthly) ** (1 / 4.33)
        cancel_rate = 1 - weekly_retention ** weeks_out

    adjusted = {}
    for sku, qty in demand.items():
        sub_qty = recurring_demand.get(sku, 0)
        non_sub_qty = qty - sub_qty
        # Only churn the subscription portion
        churned_sub = max(0, int(round(sub_qty * (1 - cancel_rate))))
        adjusted[sku] = non_sub_qty + churned_sub

    return adjusted


def apply_unified_forecast(demand, recurring_demand, first_order_demand,
                           churn_rates, shopify_trend_data, repeat_rate,
                           reship_buffer_pct, weeks_out=0):
    """Unified forecast: churn + trend + first-order repeat + addon + reship.

    Returns {sku: adjusted_qty} for a given weeks_out horizon.
    """
    # Compute weekly retention from average monthly churn
    monthly_rates = []
    for track, rates in churn_rates.items():
        if isinstance(rates, dict):
            r = rates.get("month_2_plus", rates.get("month_3_plus",
                rates.get("month_1", 0.12)))
            monthly_rates.append(r)
    avg_monthly = sum(monthly_rates) / len(monthly_rates) if monthly_rates else 0.12
    weekly_retention = (1 - avg_monthly) ** (1 / 4.33)

    reship_mult = 1 + (reship_buffer_pct / 100.0)
    result = {}

    for sku, qty in demand.items():
        sub = recurring_demand.get(sku, 0)
        first = first_order_demand.get(sku, 0)
        addon = max(0, qty - sub - first)

        # Trend adjustment from Shopify linear regression
        trend = shopify_trend_data.get(sku, {})
        wavg = trend.get("weekly_avg", 0)
        slope = trend.get("trend_slope", 0)
        trend_adj = (1 + slope / wavg) if wavg > 0 else 1.0

        # Subscription decays with retention^weeks_out
        sub_proj = sub * (weekly_retention ** weeks_out)
        first_proj = first * repeat_rate * trend_adj
        addon_proj = addon * trend_adj

        forecast = (sub_proj + first_proj + addon_proj) * reship_mult
        result[sku] = max(0, int(round(forecast)))

    return result


# ── Full calculate using RMFG data ────────────────────────────────────

@app.route("/api/calculate_rmfg", methods=["POST"])
def calculate_rmfg():
    """Calculate NET using loaded RMFG folder data (multi-window)."""
    data = request.json or {}
    demand_mode = data.get("demand_mode", "discrete")  # "discrete" or "churned"

    inv = STATE.get("rmfg_inventory", {})
    sat_demand = STATE.get("rmfg_sat_demand", {})
    tue_demand = STATE.get("rmfg_tue_demand", {})
    next_sat_demand = STATE.get("rmfg_next_sat_demand", {})

    if not inv and not sat_demand:
        return jsonify({"error": "No RMFG data loaded. Use Load Folder first."}), 400

    # Deferred resolution: re-resolve from raw components using current settings
    attribution = {}
    raw_direct = STATE.get("rmfg_direct_sat")
    raw_prcjam = STATE.get("rmfg_prcjam_sat")
    raw_cexec = STATE.get("rmfg_cexec_sat")
    if raw_direct is not None and raw_prcjam is not None:
        s = _s()
        pr_cjam = s.get("pr_cjam", {})
        cex_ec = s.get("cex_ec", {})
        splits = s.get("cexec_splits", {})
        sat_demand, attribution = resolve_demand(
            raw_direct, raw_prcjam, raw_cexec, pr_cjam, cex_ec, splits
        )
        STATE["rmfg_sat_demand"] = sat_demand
        STATE["rmfg_attribution"] = attribution

    # Monthly box demand (AHB-MED, AHB-CMED, AHB-LGE) — individual SKU demand
    # already flows through parsers via direct_demand. Counts stored in STATE
    # for display in the Monthly Boxes panel.

    # If we have inventory (e.g. from Dropbox) but no parsed demand,
    # build demand from settings (Recharge + Shopify + manual)
    if inv and not sat_demand:
        s = _s()
        rq_resolved = s.get("recharge_queued_resolved", {})
        shopify = s.get("shopify_api_demand", {})
        manual = s.get("manual_demand", {})
        pr_cjam = s.get("pr_cjam", {})
        cex_ec = s.get("cex_ec", {})
        splits = s.get("cexec_splits", {})
        ge = s.get("global_extras", {})

        # Build weekly demand from settings sources
        settings_demand = defaultdict(int)
        # Recharge queued (use latest month)
        for month, data in rq_resolved.items():
            for suffix, count in data.get("pr_cjam", {}).items():
                info = pr_cjam.get(suffix, {})
                ch = info.get("cheese", "") if isinstance(info, dict) else str(info)
                if ch:
                    settings_demand[normalize_sku(ch)] += int(count)
            for suffix, count in data.get("cex_ec", {}).items():
                ec = cex_ec.get(suffix, "")
                if isinstance(ec, str) and ec:
                    settings_demand[normalize_sku(ec)] += int(count)
                elif isinstance(ec, dict):
                    for esku, epct in ec.items():
                        settings_demand[normalize_sku(esku)] += int(count * epct)
            for sku, qty in data.get("direct", {}).items():
                ge_resolved = ge.get(sku.upper())
                if ge_resolved:
                    settings_demand[normalize_sku(ge_resolved)] += int(qty)
                else:
                    settings_demand[normalize_sku(sku)] += int(qty)
        # Shopify API demand (weekly)
        for sku, qty in shopify.items():
            settings_demand[normalize_sku(sku)] += int(qty)
        # Manual demand
        for sku, qty in manual.items():
            settings_demand[normalize_sku(sku)] += int(qty)

        sat_demand = dict(settings_demand)
        # Estimate Tuesday as ~30% of Saturday
        tue_demand = {sku: max(1, int(q * 0.3))
                      for sku, q in sat_demand.items() if q > 0}
        next_sat_demand = dict(sat_demand)

    # Apply churn adjustment if in churned mode
    churn_info = {}
    if demand_mode == "churned":
        s = _s()
        churn_rates = s.get("churn_rates", {})
        dashboard = STATE.get("rmfg_dashboard", {})
        recurring = dashboard.get("recurring_demand", {})

        discrete_sat = dict(sat_demand)
        discrete_tue = dict(tue_demand)
        discrete_next = dict(next_sat_demand)

        sat_demand = apply_churn_to_demand(sat_demand, recurring, churn_rates, weeks_out=0)
        tue_demand = apply_churn_to_demand(tue_demand, recurring, churn_rates, weeks_out=1)
        next_sat_demand = apply_churn_to_demand(next_sat_demand, recurring, churn_rates, weeks_out=2)

        churn_info = {
            "mode": "churned",
            "sat_reduction": sum(discrete_sat.values()) - sum(sat_demand.values()),
            "tue_reduction": sum(discrete_tue.values()) - sum(tue_demand.values()),
            "next_reduction": sum(discrete_next.values()) - sum(next_sat_demand.values()),
        }
    else:
        churn_info = {"mode": "discrete"}

    # Collect all CH-* SKUs
    all_ch = set()
    all_ch.update(k for k in inv if k.startswith("CH-"))
    all_ch.update(k for k in sat_demand if k.startswith("CH-"))
    all_ch.update(k for k in tue_demand if k.startswith("CH-"))
    all_ch.update(k for k in next_sat_demand if k.startswith("CH-"))

    bulk_weights = STATE.get("bulk_weights", {})
    results = []
    shortage_count = 0
    for sku in sorted(all_ch):
        avail = inv.get(sku, 0)
        d_sat = int(round(sat_demand.get(sku, 0)))
        d_tue = int(round(tue_demand.get(sku, 0)))
        d_next = int(round(next_sat_demand.get(sku, 0)))
        total = d_sat + d_tue + d_next

        # Potential yield from wheels/blocks
        bw = bulk_weights.get(sku, {})
        potential = bw.get("potential_yield", 0)
        wheel_count = bw.get("count", 0)

        net_sat = avail - d_sat
        net_tue = net_sat - d_tue  # after Tuesday
        net_final = avail - total  # after next Saturday

        # Net including potential (if wheels were cut)
        net_with_potential = avail + potential - d_sat

        if d_sat == 0 and d_tue == 0 and d_next == 0:
            status = "NO DEMAND"
        elif net_sat < 0:
            if potential > 0 and net_with_potential >= 0:
                status = "MFG"  # short on processed, but wheels available
            else:
                status = "SHORTAGE"
            shortage_count += 1
        elif net_sat < d_sat * 0.2:
            status = "TIGHT"
        elif net_sat > avail * 0.5 and avail > 200:
            status = "SURPLUS"
        else:
            status = "OK"

        results.append({
            "sku": sku, "available": avail,
            "sat_demand": d_sat, "net_sat": net_sat,
            "tue_demand": d_tue, "net_tue": net_tue,
            "next_sat_demand": d_next, "net_final": net_final,
            "total_demand": total, "net": net_sat,
            "potential": potential, "wheel_count": wheel_count,
            "net_with_potential": net_with_potential,
            "status": status,
            # Keep compat fields for the existing UI
            "direct": d_sat, "prcjam": 0, "cexec": 0, "exec": 0,
        })

    status_order = {"SHORTAGE": 0, "TIGHT": 1, "OK": 2,
                    "SURPLUS": 3, "NO DEMAND": 4}
    results.sort(key=lambda r: (status_order.get(r["status"], 9), r["net"]))

    # Dashboard info for assignment panel
    dashboard = STATE.get("rmfg_dashboard", {})
    prcjam_counts = dashboard.get("prcjam_counts", {})
    cexec_counts = dashboard.get("cexec_counts", {})

    # Multi-week: Tue and Next Sat as week tabs
    weeks = []
    # Week 2 = Tuesday
    tue_results = []
    tue_shortages = 0
    for r in results:
        if not r["sku"].startswith("CH-"):
            continue
        carry = max(0, r["net_sat"])
        demand = r["tue_demand"]
        proj = carry - demand
        if demand == 0:
            st = "NO DEMAND"
        elif proj < 0:
            st = "PLAN PO"
            tue_shortages += 1
        elif proj < demand * 0.3:
            st = "TIGHT"
        else:
            st = "OK"
        tue_results.append({
            "sku": r["sku"], "carry_fwd": carry,
            "demand": demand, "net": proj, "status": st,
        })
    weeks.append({"week": 2, "label": "Tuesday", "results": tue_results,
                  "shortages": tue_shortages})

    # Week 3 = Next Saturday
    nsat_results = []
    nsat_shortages = 0
    for r in results:
        if not r["sku"].startswith("CH-"):
            continue
        carry = max(0, r["net_sat"] - r["tue_demand"])
        demand = r["next_sat_demand"]
        proj = carry - demand
        if demand == 0:
            st = "NO DEMAND"
        elif proj < 0:
            st = "PLAN PO"
            nsat_shortages += 1
        elif proj < demand * 0.3:
            st = "TIGHT"
        else:
            st = "OK"
        nsat_results.append({
            "sku": r["sku"], "carry_fwd": carry,
            "demand": demand, "net": proj, "status": st,
        })
    weeks.append({"week": 3, "label": "Next Sat", "results": nsat_results,
                  "shortages": nsat_shortages})

    # Weeks 4-5 (Sat +3, Sat +4) — carry forward from previous week
    prev_week = {r["sku"]: r for r in nsat_results}
    for week_num in range(4, 6):
        wk_results = []
        wk_shortages = 0
        for r in results:
            if not r["sku"].startswith("CH-"):
                continue
            p = prev_week.get(r["sku"])
            carry = max(0, p["net"]) if p else 0
            demand = r["sat_demand"]  # assume Saturday-like demand
            proj = carry - demand
            if demand == 0:
                st = "NO DEMAND"
            elif proj < 0:
                st = "PLAN PO"
                wk_shortages += 1
            elif proj < demand * 0.3:
                st = "TIGHT"
            else:
                st = "OK"
            wk_results.append({
                "sku": r["sku"], "carry_fwd": carry,
                "demand": demand, "net": proj, "status": st,
            })
        weeks.append({"week": week_num,
                      "label": f"Sat +{week_num - 1}",
                      "results": wk_results,
                      "shortages": wk_shortages})
        prev_week = {r["sku"]: r for r in wk_results}

    # Shelf life
    s = _s()
    inventory = s.get("inventory", {})
    today = datetime.date.today()
    shelf_items = []
    for sku, data in inventory.items():
        if not sku.startswith("CH-") or not isinstance(data, dict):
            continue
        dates = data.get("expiration_dates", [])
        if not dates:
            continue
        try:
            earliest = datetime.date.fromisoformat(dates[0])
        except (ValueError, IndexError):
            continue
        days = (earliest - today).days
        qty = data.get("qty", 0)
        if days <= 14 and qty > 0:
            shelf_items.append({
                "sku": sku, "days_left": days, "qty": qty,
                "action": "EXPIRED" if days < 0 else
                          "USE NOW" if days <= 7 else "Prioritize",
            })

    return jsonify({
        "results": results,
        "shortages": shortage_count,
        "total_skus": len([r for r in results if r["total_demand"] > 0]),
        "total_units": sum(r["total_demand"] for r in results),
        "prcjam_counts": prcjam_counts,
        "cexec_counts": cexec_counts,
        "weeks": weeks,
        "shelf_life": shelf_items,
        "assign_demands": {},
        "churn_info": churn_info,
        "attribution": attribution,
    })


# ── Inventory Runway ──────────────────────────────────────────────────


def _runway_saturdays(num_weeks=4):
    """Return next N Saturday dates from today."""
    today = datetime.date.today()
    days_to_sat = (5 - today.weekday()) % 7
    if days_to_sat == 0:
        days_to_sat = 7
    return [today + datetime.timedelta(days=days_to_sat + 7 * i)
            for i in range(num_weeks)]


def _fmtd(d):
    """Format date as M/DD."""
    try:
        return d.strftime("%#m/%d")  # Windows
    except (ValueError, TypeError):
        return d.strftime("%-m/%d")  # Unix


def _split_monthly_to_weekly(monthly_data, saturdays):
    """Split {month: {sku: qty}} into per-Saturday demand dicts.

    Each Saturday gets a proportional share: monthly_qty / saturdays_in_that_month.
    """
    import calendar
    # Pre-compute how many Saturdays are in each month
    sats_per_month = {}
    for sat in saturdays:
        key = sat.strftime("%Y-%m")
        if key not in sats_per_month:
            # Count Saturdays in this month
            year, month = sat.year, sat.month
            days_in_month = calendar.monthrange(year, month)[1]
            count = sum(1 for d in range(1, days_in_month + 1)
                        if datetime.date(year, month, d).weekday() == 5)
            sats_per_month[key] = count

    week_demands = [defaultdict(float) for _ in saturdays]
    for i, sat in enumerate(saturdays):
        month_key = sat.strftime("%Y-%m")
        month_skus = monthly_data.get(month_key, {})
        divisor = sats_per_month.get(month_key, 4)
        for sku, qty in month_skus.items():
            week_demands[i][sku] += qty / divisor

    return week_demands


def _extract_curation_counts(monthly_data, saturdays):
    """Extract per-curation box counts from monthly queued data, split weekly.

    Parses AHB-*CUST-* SKUs to find curation suffix and count boxes.
    Returns [{curation: count}, ...] per Saturday.
    """
    import calendar
    sats_per_month = {}
    for sat in saturdays:
        key = sat.strftime("%Y-%m")
        if key not in sats_per_month:
            year, month = sat.year, sat.month
            days_in_month = calendar.monthrange(year, month)[1]
            count = sum(1 for d in range(1, days_in_month + 1)
                        if datetime.date(year, month, d).weekday() == 5)
            sats_per_month[key] = count

    week_counts = [defaultdict(float) for _ in saturdays]
    for i, sat in enumerate(saturdays):
        month_key = sat.strftime("%Y-%m")
        month_skus = monthly_data.get(month_key, {})
        divisor = sats_per_month.get(month_key, 4)
        for sku, qty in month_skus.items():
            if not (sku.startswith("AHB-") and "CUST" in sku):
                continue
            # Extract curation: last segment after splitting
            cur = sku.split("-")[-1]
            week_counts[i][cur] += qty / divisor

    return week_counts


def _build_queued_runway_demand(recharge_queued, pr_cjam, cex_ec, splits, saturdays):
    """Build per-week demand dicts from recharge_queued monthly data.

    Returns (week_demands, week_direct, week_prcjam, week_cexec, curation_box_counts).
    week_demands: [{sku: qty}, ...] — fully resolved demand per Saturday
    week_direct/prcjam/cexec: raw components for deferred resolution
    curation_box_counts: [{curation: count}, ...] per Saturday
    """
    PICKABLE = ("CH-", "MT-", "AC-", "BL-", "EX-", "CEX-", "PR-")

    # Split monthly totals into weekly shares
    raw_weekly = _split_monthly_to_weekly(recharge_queued, saturdays)
    curation_box_counts = _extract_curation_counts(recharge_queued, saturdays)

    week_direct = []
    week_prcjam = []
    week_cexec = []

    for wi, (raw, box_counts) in enumerate(zip(raw_weekly, curation_box_counts)):
        direct = defaultdict(float)
        prcjam_counts = defaultdict(float)
        cexec_counts = defaultdict(float)

        total_boxes = sum(box_counts.values())

        for sku, qty in raw.items():
            upper = sku.upper().strip()

            # Skip box SKUs — components are already listed individually
            if upper.startswith("AHB-") or upper.startswith("TR-"):
                continue

            # PR-CJAM: distribute GEN proportionally across curations
            if upper.startswith("PR-CJAM"):
                suffix = upper.split("PR-CJAM-", 1)[-1] if "-" in upper[7:] else "GEN"
                if suffix == "GEN" and total_boxes > 0:
                    # Distribute proportionally by curation box count
                    for cur, bcount in box_counts.items():
                        prcjam_counts[cur] += qty * (bcount / total_boxes)
                elif suffix != "GEN":
                    prcjam_counts[suffix] += qty
                continue

            # CEX-EC: distribute across curations with large boxes
            if upper.startswith("CEX-EC") or upper == "CEX-EC":
                # Only large boxes get CEX-EC
                large_curs = {c: n for c, n in box_counts.items()
                              if any(raw.get(f"AHB-LCUST-{c}", 0) > 0
                                     for _ in [1])
                              # Simpler: check recharge_queued for LCUST
                              }
                # Fallback: distribute across all curations
                if upper == "CEX-EC" and total_boxes > 0:
                    # CEX-EC goes to large box curations
                    for cur, bcount in box_counts.items():
                        cexec_counts[cur] += qty * (bcount / total_boxes)
                elif "-" in upper[6:]:
                    suffix = upper.split("CEX-EC-", 1)[-1]
                    cexec_counts[suffix] += qty
                continue

            # Direct pickable SKUs
            if any(upper.startswith(p) for p in ("CH-", "MT-", "AC-", "BL-", "EX-")):
                direct[normalize_sku(sku)] += qty

        week_direct.append(dict(direct))
        week_prcjam.append(dict(prcjam_counts))
        week_cexec.append(dict(cexec_counts))

    # Resolve each week's demand using current assignments
    week_demands = []
    all_attribution = {}
    for wi in range(len(saturdays)):
        resolved, attr = resolve_demand(
            week_direct[wi], week_prcjam[wi], week_cexec[wi],
            pr_cjam, cex_ec, splits
        )
        week_demands.append(resolved)
        # Merge attribution (accumulate)
        for sku, a in attr.items():
            if sku not in all_attribution:
                all_attribution[sku] = {"direct": 0, "prcjam": {}, "cexec": {}}
            all_attribution[sku]["direct"] += a["direct"]
            for c, v in a["prcjam"].items():
                all_attribution[sku]["prcjam"][c] = \
                    all_attribution[sku]["prcjam"].get(c, 0) + v
            for c, v in a["cexec"].items():
                all_attribution[sku]["cexec"][c] = \
                    all_attribution[sku]["cexec"].get(c, 0) + v

    return week_demands, week_direct, week_prcjam, week_cexec, \
        curation_box_counts, all_attribution


@app.route("/api/runway", methods=["POST"])
def get_runway():
    """Return per-SKU weekly depletion for forecast + discrete models.

    Demand sources (in priority order):
    1. Recharge queued charges (settings) — split monthly into weekly Saturdays
    2. RMFG folder data (STATE) — legacy CSV-based path
    """
    s = _s()
    pr_cjam = s.get("pr_cjam", {})
    cex_ec = s.get("cex_ec", {})
    splits = s.get("cexec_splits", {})
    churn_rates = s.get("churn_rates", {})
    repeat_rate = s.get("repeat_rate", 0.56)
    reship_pct = s.get("reship_buffer_pct", 0.77)
    shopify_trend = s.get("shopify_trend_data", {})
    inventory_settings = s.get("inventory", {})
    recharge_queued = s.get("recharge_queued", {})

    # Compute avg weekly churn for model_params display
    monthly_rates = []
    for track, rates in churn_rates.items():
        if isinstance(rates, dict):
            r = rates.get("month_2_plus", rates.get("month_3_plus",
                rates.get("month_1", 0.12)))
            monthly_rates.append(r)
    avg_monthly = sum(monthly_rates) / len(monthly_rates) if monthly_rates else 0.12
    avg_churn_weekly = round(1 - (1 - avg_monthly) ** (1 / 4.33), 4)

    # --- Inventory source ---
    # Prefer RMFG loaded inventory, fall back to settings inventory
    inv = STATE.get("rmfg_inventory", {})
    bulk_weights = STATE.get("bulk_weights", {})
    if not inv:
        inv = {}
        for sku, data in inventory_settings.items():
            if isinstance(data, dict):
                inv[sku] = data.get("qty", 0)
            else:
                inv[sku] = int(data) if data else 0

    if not inv and not recharge_queued:
        return jsonify({"error": "No inventory or demand data. Load data or sync Recharge."}), 400

    # --- 4 Saturday windows ---
    NUM_WEEKS = 4
    saturdays = _runway_saturdays(NUM_WEEKS)
    week_labels = [_fmtd(sat) for sat in saturdays]

    # --- Demand source ---
    # Shopify API demand (weekly) — add to each Saturday window
    shopify_weekly = s.get("shopify_api_demand", {})

    if recharge_queued:
        # NEW PATH: Recharge queued charges → weekly demand
        week_demands, week_direct, week_prcjam, week_cexec, \
            curation_box_counts, attribution = _build_queued_runway_demand(
                recharge_queued, pr_cjam, cex_ec, splits, saturdays
            )

        # --- Demand model ---
        # Recurring = Recharge queued (actual scheduled charges) — NO prediction
        # Predictive = Shopify first orders + add-ons + specialty — from history
        #
        # Recharge queued is the source of truth for recurring subscription demand.
        # Shopify fulfilled history tells us the weekly rate for:
        #   - Subscription First Orders (tagged "Subscription First Order")
        #   - Add-ons and one-time purchases (no subscription tag)
        #   - Specialty boxes
        # These predictive components are added on top of Recharge actuals.

        shopify_first = s.get("shopify_first_order_demand", {})
        shopify_recurring = s.get("shopify_recurring_demand", {})
        shopify_unfulfilled = s.get("shopify_unfulfilled", {})
        first_order_overrides = s.get("first_order_overrides", {})

        # Predictive demand = first orders + add-ons (anything in Shopify
        # that's NOT a recurring subscription)
        predictive_weekly = {}
        first_order_weekly = {}
        addon_weekly = {}
        if shopify_weekly:
            for sku, total_avg in shopify_weekly.items():
                nsku = normalize_sku(sku)
                rec_avg = shopify_recurring.get(sku, 0)
                # Predictive = total Shopify avg minus recurring portion
                # (recurring already comes from Recharge queued)
                pred = max(0, total_avg - rec_avg)
                if pred > 0:
                    predictive_weekly[nsku] = pred
                # Split predictive into first-order vs addon
                fo = first_order_overrides.get(nsku,
                     first_order_overrides.get(sku,
                     shopify_first.get(nsku,
                     shopify_first.get(sku, 0))))
                fo = min(fo, pred)  # can't exceed total predictive
                if fo > 0:
                    first_order_weekly[nsku] = fo
                addon = pred - fo
                if addon > 0:
                    addon_weekly[nsku] = addon

        # Merge into week_demands:
        # week_demands already has Recharge recurring (actual, flat per week)
        # Add predictive on top of each week
        for wi in range(len(week_demands)):
            for nsku, qty in predictive_weekly.items():
                week_demands[wi][nsku] = week_demands[wi].get(nsku, 0) + qty
            # Week 0 also gets unfulfilled pending orders
            if wi == 0 and shopify_unfulfilled:
                for sku, qty in shopify_unfulfilled.items():
                    nsku = normalize_sku(sku)
                    week_demands[0][nsku] = week_demands[0].get(nsku, 0) + qty

        # For the forecast model:
        # recurring_demand = Recharge queued weekly share (no churn applied)
        # first_order_demand = Shopify predictive (gets repeat_rate projection)
        recurring_demand = {}
        first_order_demand = {}
        first_order_split = {}
        addon_split = {}
        if week_demands:
            for sku, qty in week_demands[0].items():
                pred = predictive_weekly.get(sku, 0)
                recurring_demand[sku] = max(0, int(round(qty - pred)))
                first_order_demand[sku] = int(round(pred))
                first_order_split[sku] = int(round(first_order_weekly.get(sku, 0)))
                addon_split[sku] = int(round(addon_weekly.get(sku, 0)))
    else:
        # LEGACY PATH: RMFG folder data
        raw_direct = STATE.get("rmfg_direct_sat", {})
        raw_prcjam = STATE.get("rmfg_prcjam_sat", {})
        raw_cexec = STATE.get("rmfg_cexec_sat", {})
        if raw_direct is not None and raw_prcjam is not None:
            sat_demand_resolved, attribution = resolve_demand(
                raw_direct, raw_prcjam, raw_cexec, pr_cjam, cex_ec, splits
            )
        else:
            sat_demand_resolved = STATE.get("rmfg_sat_demand", {})
            attribution = STATE.get("rmfg_attribution", {})
        tue_demand = STATE.get("rmfg_tue_demand", {})
        dashboard = STATE.get("rmfg_dashboard", {})
        recurring_demand = dashboard.get("recurring_demand", {})
        first_order_demand = dashboard.get("first_order_demand", {})
        first_order_split = dict(first_order_demand)
        addon_split = {}
        # Build 4 weekly windows from legacy data
        week_demands = [sat_demand_resolved, tue_demand,
                        sat_demand_resolved, sat_demand_resolved]

    # --- Collect all pickable SKUs ---
    RUNWAY_PREFIXES = ("CH-", "MT-", "AC-", "TR-")
    all_skus = set()
    all_skus.update(k for k in inv
                    if any(k.startswith(p) for p in RUNWAY_PREFIXES))
    for wd in week_demands:
        all_skus.update(k for k in wd
                        if any(k.startswith(p) for p in RUNWAY_PREFIXES))

    # Compute weekly retention for churn decay on recurring demand
    weekly_retention = (1 - avg_monthly) ** (1 / 4.33)

    # Archived SKUs to exclude
    archived_skus = set(s.get("archived_skus", []))

    total_forecast = 0
    sku_rows = []

    for sku in sorted(all_skus):
        # Skip archived SKUs
        if sku in archived_skus:
            continue
        avail = inv.get(sku, 0)
        bw = bulk_weights.get(sku, {})
        potential = bw.get("potential_yield", 0) if isinstance(bw, dict) else 0

        # Get SKU display name
        inv_entry = inventory_settings.get(sku, {})
        sku_name = inv_entry.get("name", "") if isinstance(inv_entry, dict) else ""

        # Determine category
        if sku.startswith("CH-"):
            category = "cheese"
        elif sku.startswith("MT-"):
            category = "meat"
        else:
            category = "accompaniment"

        # Walk through weekly windows (include wheel/bulk potential in supply)
        forecast_weeks = []
        f_carry = avail + potential

        rec_qty = recurring_demand.get(sku, 0)
        pred_qty = first_order_demand.get(sku, 0)

        # Predictive projection factors (constant across weeks)
        reship_mult = 1 + (reship_pct / 100.0)
        trend = shopify_trend.get(sku, {})
        wavg = trend.get("weekly_avg", 0)
        slope = trend.get("trend_slope", 0)
        trend_adj = (1 + slope / wavg) if wavg > 0 else 1.0
        f_predictive = pred_qty * repeat_rate * trend_adj * reship_mult

        for wi in range(NUM_WEEKS):
            # Recurring decays with weekly retention (cancellations/skips)
            f_recurring = rec_qty * (weekly_retention ** wi)

            f_demand = max(0, int(round(f_recurring + f_predictive)))

            f_carry_in = f_carry
            f_out = f_carry_in - f_demand

            # Status
            if f_demand == 0:
                status = "OK"
            elif f_out < 0:
                status = "SHORTAGE"
            elif f_out < f_demand * 0.3:
                status = "TIGHT"
            else:
                status = "OK"

            forecast_weeks.append({
                "label": week_labels[wi],
                "demand": f_demand,
                "carry_in": f_carry_in,
                "carry_out": f_out,
                "status": status,
            })

            f_carry = max(0, f_out)

        # Runway weeks: how many weeks until stock hits 0
        # Include wheel/bulk potential in total supply
        total_supply = avail + potential
        f_runway = 0.0
        remaining = total_supply
        for wi in range(NUM_WEEKS):
            fd = forecast_weeks[wi]["demand"]
            if fd <= 0:
                f_runway += 1.0
                continue
            if remaining >= fd:
                remaining -= fd
                f_runway += 1.0
            else:
                f_runway += remaining / fd
                break

        # Worst status across forecast weeks
        status_priority = {"SHORTAGE": 0, "TIGHT": 1, "OK": 2}
        worst = min((status_priority.get(w["status"], 2) for w in forecast_weeks))
        worst_status = ["SHORTAGE", "TIGHT", "OK"][worst]

        # Auto-hide: zero inventory + zero demand across all weeks
        total_demand = sum(w["demand"] for w in forecast_weeks)
        if avail == 0 and potential == 0 and total_demand == 0:
            continue

        attr = attribution.get(sku, {"direct": 0, "prcjam": {}, "cexec": {}})

        total_forecast += forecast_weeks[0]["demand"]

        demand_per_wk = int(round(rec_qty + f_predictive))
        fo_wk = first_order_split.get(sku, 0)
        addon_wk = addon_split.get(sku, 0)

        sku_rows.append({
            "sku": sku,
            "name": sku_name,
            "category": category,
            "available": avail,
            "potential": potential,
            "forecast": {
                "weeks": forecast_weeks,
                "runway_weeks": round(f_runway, 1),
            },
            "demand_per_wk": demand_per_wk,
            "recurring_per_wk": int(round(rec_qty)),
            "predicted_per_wk": int(round(f_predictive)),
            "first_order_per_wk": fo_wk,
            "addon_per_wk": addon_wk,
            "attribution": attr,
            "worst_status": worst_status,
        })

    # Sort by shortest forecast runway (worst first)
    sku_rows.sort(key=lambda r: (
        {"SHORTAGE": 0, "TIGHT": 1, "OK": 2}.get(r["worst_status"], 2),
        r["forecast"]["runway_weeks"],
    ))

    # Build assignment summary with unit counts from curation box counts
    curation_order = ["MONG", "MDT", "OWC", "SPN", "ALPN", "ISUN", "HHIGH",
                      "NMS", "BYO", "SS", "MS"]
    # Sum curation box counts across all weeks for display
    total_cur_boxes = defaultdict(float)
    if recharge_queued:
        for wk_counts in curation_box_counts:
            for cur, n in wk_counts.items():
                total_cur_boxes[cur] += n
    else:
        dashboard = STATE.get("rmfg_dashboard", {})
        prcjam_counts_raw = dashboard.get("prcjam_counts", {})
        cexec_counts_raw = dashboard.get("cexec_counts", {})

    assignments = []
    for cur in curation_order:
        pj_sku = pr_cjam.get(cur, {})
        pj_cheese = pj_sku.get("cheese", "") if isinstance(pj_sku, dict) else ""
        ce_sku = cex_ec.get(cur, "")

        if recharge_queued:
            pj_units = int(round(total_cur_boxes.get(cur, 0)))
            ce_units = int(round(total_cur_boxes.get(cur, 0)))
        else:
            pj_units = int(round(prcjam_counts_raw.get(cur, 0)))
            ce_units = int(round(cexec_counts_raw.get(cur, 0)))

        if pj_units > 0 or ce_units > 0 or pj_cheese or ce_sku:
            assignments.append({
                "curation": cur,
                "prcjam": pj_cheese,
                "prcjam_units": pj_units,
                "cexec": ce_sku,
                "cexec_units": ce_units,
            })

    return jsonify({
        "skus": sku_rows,
        "week_labels": week_labels,
        "model_params": {
            "repeat_rate": repeat_rate,
            "reship_pct": reship_pct,
            "avg_churn_weekly": avg_churn_weekly,
            "total_forecast": total_forecast,
        },
        "assignments": assignments,
    })


# ── Activity Log ──────────────────────────────────────────────────────

@app.route("/api/activity_log")
def activity_log():
    """Return unified timeline of inventory events sorted by date desc."""
    s = _s()
    days = request.args.get("days", 60, type=int)
    cutoff = (datetime.datetime.now() - datetime.timedelta(days=days)).strftime("%Y-%m-%d")
    events = []

    # Depletions (outflow)
    for d in s.get("depletion_history", []):
        date = d.get("date", "")
        if date < cutoff:
            continue
        total = d.get("total", 0)
        day_label = d.get("day", "")
        reship = d.get("reship_count", 0)
        skus = d.get("skus", {})
        sku_list = [{"sku": k, "qty": -v} for k, v in skus.items() if v]
        sku_list.sort(key=lambda x: x["qty"])
        summary = f"{day_label} fulfillment — {total} units"
        if reship:
            summary += f" ({reship} reships)"
        events.append({
            "type": "DEPLETION", "date": date, "direction": "out",
            "summary": summary, "total_units": -total,
            "skus": sku_list,
        })

    # Open POs (inflow when received, action when submitted)
    for po in s.get("open_pos", []):
        eta = po.get("eta", "")
        sku = po.get("sku", "")
        qty = po.get("qty", 0)
        status = po.get("status", "Open")
        vendor = po.get("vendor", "")
        if status == "Received":
            date = po.get("received_date", eta)
            if date < cutoff:
                continue
            events.append({
                "type": "PO_RECEIVED", "date": date, "direction": "in",
                "summary": f"PO received from {vendor} — {qty} units",
                "total_units": qty,
                "skus": [{"sku": sku, "qty": qty}],
            })
        elif eta and eta >= cutoff:
            events.append({
                "type": "PO_OPEN", "date": eta, "direction": "status",
                "summary": f"PO open — {qty} {sku} from {vendor} (ETA {eta})",
                "total_units": qty,
                "skus": [{"sku": sku, "qty": qty}],
            })

    # Production yields (inflow)
    for y in s.get("production_yield_history", []):
        date = y.get("date", "")
        if date < cutoff:
            continue
        sku = y.get("sku", "")
        actual = y.get("actual", 0)
        events.append({
            "type": "PRODUCTION", "date": date, "direction": "in",
            "summary": f"Production yield — {actual} units of {sku}",
            "total_units": actual,
            "skus": [{"sku": sku, "qty": actual}],
        })

    # Transfers (inflow/outflow depending on perspective)
    for t in s.get("transfer_history", []):
        date = t.get("date", "")
        if date < cutoff:
            continue
        sku = t.get("sku", "")
        qty = t.get("qty", 0)
        from_wh = t.get("from_warehouse", "")
        to_wh = t.get("to_warehouse", "")
        events.append({
            "type": "TRANSFER", "date": date, "direction": "in",
            "summary": f"Transfer {from_wh} → {to_wh} — {qty} {sku}",
            "total_units": qty,
            "skus": [{"sku": sku, "qty": qty}],
        })

    # Waste ledger (outflow)
    for w in s.get("waste_ledger", []):
        date = w.get("date", "")
        if date < cutoff:
            continue
        sku = w.get("sku", "")
        qty = w.get("qty", 0)
        reason = w.get("reason", "")
        events.append({
            "type": "WASTE", "date": date, "direction": "out",
            "summary": f"Waste — {qty} {sku} ({reason})",
            "total_units": -qty,
            "skus": [{"sku": sku, "qty": -qty}],
        })

    # Reconciliation events (status)
    for r in s.get("reconciliation_history", []):
        date = r.get("date", "")
        if date < cutoff:
            continue
        mismatches = r.get("mismatches", 0)
        checked = r.get("skus_checked", 0)
        events.append({
            "type": "RECONCILIATION", "date": date, "direction": "status",
            "summary": f"Reconciliation — {checked} SKUs checked, {mismatches} mismatches",
            "total_units": 0, "skus": [],
        })

    # Sort newest first
    events.sort(key=lambda e: e["date"], reverse=True)

    # Compute totals
    total_in = sum(e["total_units"] for e in events if e["direction"] == "in")
    total_out = sum(e["total_units"] for e in events if e["direction"] == "out")

    return jsonify({
        "events": events,
        "total_in": total_in,
        "total_out": total_out,
        "event_count": len(events),
    })


# ── First-Order Overrides ─────────────────────────────────────────────

@app.route("/api/first_order_override", methods=["POST"])
def set_first_order_override():
    """Set or clear a per-SKU first-order demand override."""
    data = request.get_json(force=True)
    sku = data.get("sku", "").strip()
    if not sku:
        return jsonify({"error": "sku required"}), 400
    s = _s()
    overrides = s.get("first_order_overrides", {})
    if data.get("clear"):
        overrides.pop(sku, None)
    else:
        qty = data.get("qty")
        if qty is None:
            return jsonify({"error": "qty required"}), 400
        try:
            qty_int = int(qty)
        except (TypeError, ValueError):
            return jsonify({"error": "qty must be an integer"}), 400
        if qty_int < 0 or qty_int > 10000:
            return jsonify({"error": "qty out of range (0-10000)"}), 400
        overrides[sku] = qty_int
    s["first_order_overrides"] = overrides
    save_settings(s)
    STATE["saved"] = s
    return jsonify({"ok": True, "overrides": overrides})


@app.route("/api/first_order_overrides", methods=["GET"])
def get_first_order_overrides():
    """Return current first-order overrides and rolling averages."""
    s = _s()
    overrides = s.get("first_order_overrides", {})
    rolling = s.get("shopify_first_order_demand", {})
    return jsonify({"overrides": overrides, "rolling_averages": rolling})


# ── Cut Order ─────────────────────────────────────────────────────────

@app.route("/api/cut_order", methods=["POST"])
def get_cut_order():
    """Generate cut order from current demand, inventory, and wheel data."""
    inv = STATE.get("rmfg_inventory", {})
    bulk_weights = STATE.get("bulk_weights", {})
    s = _s()
    pr_cjam = s.get("pr_cjam", {})
    cex_ec = s.get("cex_ec", {})
    splits = s.get("cexec_splits", {})

    # Re-resolve demand from raw components
    raw_direct = STATE.get("rmfg_direct_sat", {})
    raw_prcjam = STATE.get("rmfg_prcjam_sat", {})
    raw_cexec = STATE.get("rmfg_cexec_sat", {})

    if raw_direct or raw_prcjam or raw_cexec:
        sat_demand, attribution = resolve_demand(
            raw_direct, raw_prcjam, raw_cexec, pr_cjam, cex_ec, splits
        )
    else:
        sat_demand = STATE.get("rmfg_sat_demand", {})
        attribution = STATE.get("rmfg_attribution", {})

    # Also resolve Recharge vs Shopify breakdown for RC/SH columns
    sh_direct = STATE.get("rmfg_direct_sh", {})
    sh_prcjam = STATE.get("rmfg_prcjam_sh", {})
    sh_cexec = STATE.get("rmfg_cexec_sh", {})
    rc_direct = STATE.get("rmfg_direct_rc", {})
    rc_prcjam = STATE.get("rmfg_prcjam_rc", {})
    rc_cexec = STATE.get("rmfg_cexec_rc", {})

    sh_demand, _ = resolve_demand(sh_direct, sh_prcjam, sh_cexec,
                                   pr_cjam, cex_ec, splits) if sh_direct or sh_prcjam else ({}, {})
    rc_demand, _ = resolve_demand(rc_direct, rc_prcjam, rc_cexec,
                                   pr_cjam, cex_ec, splits) if rc_direct or rc_prcjam else ({}, {})

    # Collect all CH-* with demand or inventory
    all_ch = set()
    all_ch.update(k for k in inv if k.startswith("CH-"))
    all_ch.update(k for k in sat_demand if k.startswith("CH-"))

    cut_lines = []
    shortages = []
    surplus_candidates = []
    total_demand = 0
    total_wheels_to_cut = 0
    total_pcs_from_cut = 0
    shortage_count = 0

    for sku in sorted(all_ch):
        sliced = inv.get(sku, 0)
        demand = int(round(sat_demand.get(sku, 0)))
        rc_qty = int(round(rc_demand.get(sku, 0)))
        sh_qty = int(round(sh_demand.get(sku, 0)))

        bw = bulk_weights.get(sku, {})
        wheel_potential = bw.get("potential_yield", 0)
        wheel_count = bw.get("count", 0)
        wheel_weight = bw.get("weight_lbs", 0)

        gap = demand - sliced
        wheels_to_cut = 0
        pcs_from_cut = 0

        if gap > 0 and wheel_count > 0 and wheel_weight > 0:
            slices_per_wheel = int(wheel_weight * WHEEL_TO_SLICE_FACTOR)
            wheels_to_cut = min(math.ceil(gap / slices_per_wheel), wheel_count)
            pcs_from_cut = wheels_to_cut * slices_per_wheel

        net = sliced + pcs_from_cut - demand

        if demand == 0:
            status = "NO DEMAND"
        elif gap <= 0:
            status = "OK"
            if sliced > demand * 1.5 and sliced > 100:
                status = "SURPLUS"
        elif pcs_from_cut >= gap:
            status = "MFG"
        else:
            status = "SHORTAGE"
            shortage_count += 1

        sku_attr = attribution.get(sku, {})

        line = {
            "sku": sku,
            "sliced": sliced,
            "rc_demand": rc_qty,
            "sh_demand": sh_qty,
            "total_demand": demand,
            "gap": max(0, gap),
            "wheels_to_cut": wheels_to_cut,
            "wheels_available": wheel_count,
            "wheel_weight": wheel_weight,
            "pcs_from_cut": pcs_from_cut,
            "net": net,
            "status": status,
            "attribution": sku_attr,
        }
        cut_lines.append(line)
        total_demand += demand

        if status == "SHORTAGE":
            shortages.append(line)
        if status == "SURPLUS" or (status == "OK" and net > 50):
            surplus_candidates.append(line)

        if wheels_to_cut > 0:
            total_wheels_to_cut += wheels_to_cut
            total_pcs_from_cut += pcs_from_cut

    # Sort: shortages first, then by gap descending
    cut_lines.sort(key=lambda r: (
        {"SHORTAGE": 0, "MFG": 1, "OK": 2, "SURPLUS": 3, "NO DEMAND": 4}.get(r["status"], 9),
        -r["gap"]
    ))
    surplus_candidates.sort(key=lambda r: -r["net"])

    return jsonify({
        "cut_lines": cut_lines,
        "shortages": shortages,
        "surplus_candidates": surplus_candidates[:15],
        "assignments": {
            "pr_cjam": pr_cjam,
            "cex_ec": cex_ec,
            "cexec_splits": splits,
        },
        "summary": {
            "total_demand": total_demand,
            "total_wheels_to_cut": total_wheels_to_cut,
            "total_pcs_from_cut": total_pcs_from_cut,
            "shortage_count": shortage_count,
            "total_skus": len([l for l in cut_lines if l["total_demand"] > 0]),
        },
    })


@app.route("/api/demand_breakdown/<sku>")
def demand_breakdown(sku):
    """Get detailed demand attribution for a single SKU."""
    attribution = STATE.get("rmfg_attribution", {})
    attr = attribution.get(sku, {"direct": 0, "prcjam": {}, "cexec": {}})
    return jsonify(attr)


@app.route("/api/cut_order_csv")
def export_cut_order_csv():
    """Export current cut order as CSV download."""
    inv = STATE.get("rmfg_inventory", {})
    bulk_weights = STATE.get("bulk_weights", {})
    s = _s()
    pr_cjam = s.get("pr_cjam", {})
    cex_ec_d = s.get("cex_ec", {})
    splits = s.get("cexec_splits", {})

    raw_direct = STATE.get("rmfg_direct_sat", {})
    raw_prcjam = STATE.get("rmfg_prcjam_sat", {})
    raw_cexec = STATE.get("rmfg_cexec_sat", {})

    if raw_direct or raw_prcjam or raw_cexec:
        sat_demand, _ = resolve_demand(raw_direct, raw_prcjam, raw_cexec,
                                        pr_cjam, cex_ec_d, splits)
    else:
        sat_demand = STATE.get("rmfg_sat_demand", {})

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["SKU", "Sliced", "Demand", "Gap", "Wheels to Cut",
                      "Wheels Avail", "Pcs from Cut", "Net", "Status"])

    all_ch = sorted(set(k for k in inv if k.startswith("CH-")) |
                    set(k for k in sat_demand if k.startswith("CH-")))

    for sku in all_ch:
        sliced = inv.get(sku, 0)
        demand = int(round(sat_demand.get(sku, 0)))
        if demand == 0 and sliced == 0:
            continue
        bw = bulk_weights.get(sku, {})
        wc = bw.get("count", 0)
        ww = bw.get("weight_lbs", 0)
        gap = max(0, demand - sliced)
        wtc = 0
        pfc = 0
        if gap > 0 and wc > 0 and ww > 0:
            spw = int(ww * WHEEL_TO_SLICE_FACTOR)
            wtc = min(math.ceil(gap / spw), wc)
            pfc = wtc * spw
        net = sliced + pfc - demand
        st = "OK"
        if demand == 0:
            st = "NO DEMAND"
        elif gap <= 0:
            st = "OK"
        elif pfc >= gap:
            st = "MFG"
        else:
            st = "SHORTAGE"
        writer.writerow([sku, sliced, demand, gap, wtc, wc, pfc, net, st])

    output.seek(0)
    today = datetime.date.today().isoformat()
    return send_file(
        io.BytesIO(output.getvalue().encode()),
        mimetype="text/csv",
        as_attachment=True,
        download_name=f"cut_order_{today}.csv",
    )


@app.route("/api/projection_settings", methods=["GET", "POST"])
def projection_settings():
    """Get or set first-order projection settings."""
    s = _s()
    if request.method == "GET":
        proj = s.get("first_order_projection", {
            "enabled": True,
            "active_curation": "MONG",
            "multiplier": 3,
            "recipe_only": True,
        })
        proj["shopify_weeks_back"] = s.get("shopify_weeks_back", 8)
        return jsonify(proj)

    data = request.json or {}
    s["first_order_projection"] = {
        "enabled": bool(data.get("enabled", True)),
        "active_curation": data.get("active_curation", "MONG"),
        "multiplier": max(1, min(10, int(data.get("multiplier", 3)))),
        "recipe_only": bool(data.get("recipe_only", True)),
    }
    STATE["saved"] = s
    save_settings(s)
    return jsonify({"ok": True, "projection": s["first_order_projection"]})


@app.route("/api/reassignment_preview", methods=["POST"])
def reassignment_preview():
    """Preview demand impact of changing a PR-CJAM or CEX-EC assignment."""
    data = request.json or {}
    slot = data.get("slot")  # "prcjam" or "cexec"
    curation = data.get("curation", "")
    cheese = data.get("cheese", "")

    s = _s()
    pr_cjam = dict(s.get("pr_cjam", {}))
    cex_ec = dict(s.get("cex_ec", {}))
    splits = dict(s.get("cexec_splits", {}))

    raw_direct = STATE.get("rmfg_direct_sat", {})
    raw_prcjam = STATE.get("rmfg_prcjam_sat", {})
    raw_cexec = STATE.get("rmfg_cexec_sat", {})

    if not raw_direct and not raw_prcjam:
        return jsonify({"error": "No raw demand data loaded"}), 400

    # Current demand
    current_demand, _ = resolve_demand(raw_direct, raw_prcjam, raw_cexec,
                                        pr_cjam, cex_ec, splits)

    # Apply proposed change
    if slot == "prcjam":
        pr_cjam_new = dict(pr_cjam)
        entry = pr_cjam_new.get(curation, {})
        if isinstance(entry, dict):
            entry = dict(entry)
            entry["cheese"] = cheese
        else:
            entry = {"cheese": cheese, "jam": ""}
        pr_cjam_new[curation] = entry
        new_demand, _ = resolve_demand(raw_direct, raw_prcjam, raw_cexec,
                                        pr_cjam_new, cex_ec, splits)
    elif slot == "cexec":
        cex_ec_new = dict(cex_ec)
        cex_ec_new[curation] = cheese
        # Clear any split for this curation
        splits_new = dict(splits)
        splits_new.pop(curation, None)
        new_demand, _ = resolve_demand(raw_direct, raw_prcjam, raw_cexec,
                                        pr_cjam, cex_ec_new, splits_new)
    else:
        return jsonify({"error": f"Unknown slot: {slot}"}), 400

    # Compute deltas
    all_skus = set(current_demand) | set(new_demand)
    deltas = {}
    for sku in all_skus:
        old_q = current_demand.get(sku, 0)
        new_q = new_demand.get(sku, 0)
        if old_q != new_q:
            deltas[sku] = {"old": old_q, "new": new_q, "delta": new_q - old_q}

    # Check adjacency constraint
    recipes = s.get("curation_recipes", {})
    constraint = check_constraint(
        curation,
        cheese if slot == "prcjam" else pr_cjam.get(curation, {}).get("cheese", ""),
        cheese if slot == "cexec" else cex_ec.get(curation, ""),
        recipes, pr_cjam, cex_ec
    )

    return jsonify({
        "deltas": deltas,
        "constraint": constraint,
    })


# ── Substitution engine ───────────────────────────────────────────────

@app.route("/api/substitutions")
def get_substitutions():
    """Suggest cheese substitutions for shortages.
    Uses loaded RMFG data to find surplus cheeses that could replace short ones.
    Includes potential yield from wheels/blocks.
    """
    inv = STATE.get("rmfg_inventory", {})
    sat_demand = STATE.get("rmfg_sat_demand", {})
    bulk_weights = STATE.get("bulk_weights", {})

    if not inv:
        return jsonify([])

    # Build NET for all CH-*, including potential yield from wheels
    nets = {}
    for sku in set(list(inv.keys()) + list(sat_demand.keys())):
        if not sku.startswith("CH-"):
            continue
        avail = inv.get(sku, 0)
        demand = sat_demand.get(sku, 0)
        bw = bulk_weights.get(sku, {})
        potential = bw.get("potential_yield", 0)
        wheel_count = bw.get("count", 0)
        nets[sku] = {
            "available": avail, "demand": demand,
            "net": avail - demand,
            "potential": potential, "wheel_count": wheel_count,
            "net_with_potential": avail + potential - demand,
        }

    # Find shortages (still short even with processed inventory)
    # and surplus candidates
    shortages = [(sku, d) for sku, d in nets.items() if d["net"] < 0]
    surplus = [(sku, d) for sku, d in nets.items()
               if d["net"] > 50 and d["demand"] > 0]
    surplus.sort(key=lambda x: -x[1]["net"])

    suggestions = []
    for sku, sdata in sorted(shortages, key=lambda x: x[1]["net"]):
        deficit = abs(sdata["net"])
        subs = []
        for cand_sku, cdata in surplus:
            if cand_sku == sku:
                continue
            headroom = cdata["net"]
            can_cover = min(deficit, headroom)
            subs.append({
                "sku": cand_sku,
                "headroom": headroom,
                "can_cover": can_cover,
                "covers_all": can_cover >= deficit,
            })
        # Also suggest any SKU with high inventory and no demand
        for sk, d in nets.items():
            if sk == sku or d["demand"] > 0:
                continue
            if d["available"] >= deficit:
                subs.append({
                    "sku": sk,
                    "headroom": d["available"],
                    "can_cover": deficit,
                    "covers_all": True,
                    "no_demand": True,
                })
        subs.sort(key=lambda x: (-x["covers_all"], -x["can_cover"]))
        suggestions.append({
            "sku": sku,
            "deficit": deficit,
            "available": sdata["available"],
            "demand": sdata["demand"],
            "potential": sdata.get("potential", 0),
            "wheel_count": sdata.get("wheel_count", 0),
            "net_with_potential": sdata.get("net_with_potential", sdata["net"]),
            "substitutes": subs[:5],
        })

    return jsonify(suggestions)


# ── Run All endpoint ──────────────────────────────────────────────────

@app.route("/api/run_all", methods=["POST"])
def run_all():
    """One-click: load folder + calculate + suggest fixes."""
    data = request.json or {}
    folder = data.get("folder", "")

    # 1. Load folder
    load_req = type('obj', (object,), {'json': data, 'method': 'POST'})()
    with app.test_request_context(json=data):
        load_resp = load_rmfg()
        load_data = load_resp.get_json() if hasattr(load_resp, 'get_json') else json.loads(load_resp.data)
        if not load_data.get("ok"):
            return jsonify(load_data)

    # 2. Calculate
    with app.test_request_context(json={}):
        calc_resp = calculate_rmfg()
        calc_data = calc_resp.get_json() if hasattr(calc_resp, 'get_json') else json.loads(calc_resp.data)

    # 3. Substitutions
    subs = []
    with app.test_request_context():
        sub_resp = get_substitutions()
        subs = sub_resp.get_json() if hasattr(sub_resp, 'get_json') else json.loads(sub_resp.data)

    return jsonify({
        "load": load_data,
        "results": calc_data,
        "substitutions": subs,
    })


# ── List available RMFG folders ───────────────────────────────────────

@app.route("/api/rmfg_folders")
def list_rmfg_folders():
    """List RMFG_* folders in the project directory."""
    base = _get_project_dir()
    folders = []
    for name in sorted(os.listdir(base)):
        path = os.path.join(base, name)
        if os.path.isdir(path) and name.startswith("RMFG_"):
            files = detect_rmfg_files(path)
            found = sum(1 for v in files.values() if v)
            folders.append({"name": name, "files_found": found,
                            "total_files": 5})
    return jsonify(folders)


# ── Dropbox integration ───────────────────────────────────────────────

@app.route("/api/dropbox_sync", methods=["POST"])
def dropbox_sync():
    """Fetch latest inventory snapshot from Dropbox shared link."""
    import tempfile
    try:
        import requests as req
    except ImportError:
        return jsonify({"error": "requests library not installed"}), 500

    STATE["saved"] = load_settings()  # reload to pick up fresh tokens
    s = _s()
    app_key = s.get("dropbox_app_key", "")
    app_secret = s.get("dropbox_app_secret", "")
    refresh_token = s.get("dropbox_refresh_token", "")
    shared_link = s.get("dropbox_shared_link", "")

    direct_token = s.get("dropbox_access_token", "")
    has_api_creds = bool(app_key and app_secret and (refresh_token or direct_token))

    if not has_api_creds and not shared_link:
        return jsonify({"error": "Dropbox not configured. Set shared_link or app credentials in Settings."}), 400

    if has_api_creds:
        # Get access token — use refresh token if available, else use direct token
        if refresh_token:
            token_resp = req.post(
                "https://api.dropboxapi.com/oauth2/token",
                data={
                    "grant_type": "refresh_token",
                    "refresh_token": refresh_token,
                    "client_id": app_key,
                    "client_secret": app_secret,
                }, timeout=15)
            token_resp.raise_for_status()
            access_token = token_resp.json()["access_token"]
        else:
            access_token = direct_token

        headers = {"Authorization": f"Bearer {access_token}",
                   "Content-Type": "application/json"}

        # Try multiple paths — direct account path first, then shared link
        attempts = []
        if shared_link:
            # Shared link with empty path = root of shared folder
            attempts.append({"path": "", "recursive": False,
                             "shared_link": {"url": shared_link.split("?")[0],
                                             "password": None}})
        attempts.append({"path": "/!AppyHour_SHARED/Product Inventory",
                         "recursive": False})
        # Try common folder name variations
        attempts.append({"path": "/!AppyHour_SHARED/Product%20Inventory",
                         "recursive": False})

        resp = None
        last_err = ""
        for list_body in attempts:
            resp = req.post(
                "https://api.dropboxapi.com/2/files/list_folder",
                headers=headers, json=list_body, timeout=15)
            if resp.status_code == 200:
                break
            last_err = resp.text[:300]

        if resp is None or resp.status_code != 200:
            return jsonify({"error": f"Dropbox list_folder failed: {last_err}"}), 400

        entries = resp.json().get("entries", [])
        # Accept any csv/xlsx file (don't filter by name — the folder IS the inventory folder)
        inv_files = [
            e for e in entries
            if e.get("name", "").lower().endswith((".csv", ".xlsx"))
        ]
        if not inv_files:
            return jsonify({"error": "No inventory files found on Dropbox"}), 404

        inv_files.sort(key=lambda e: e.get("server_modified", ""), reverse=True)
        newest = inv_files[0]
        name = newest["name"]

        # Download using file id (works for both shared link and direct)
        file_id = newest.get("id", "")
        fpath = newest.get("path_lower", "/" + name)
        dl_arg = {"path": file_id} if file_id else {"path": fpath}
        dl_resp = req.post(
            "https://content.dropboxapi.com/2/files/download",
            headers={
                "Authorization": f"Bearer {access_token}",
                "Dropbox-API-Arg": json.dumps(dl_arg),
            }, timeout=30)
        dl_resp.raise_for_status()
    else:
        # Shared link only — try to scrape file links from the folder page
        # Fetch the folder page HTML to find file links
        page_resp = req.get(shared_link, timeout=15,
                            headers={"User-Agent": "Mozilla/5.0"})
        page_resp.raise_for_status()
        html_text = page_resp.text

        # Look for file entries in the page — Dropbox embeds JSON data
        import re
        # Find file URLs in the HTML for .csv or .xlsx files containing "product inventory"
        # Dropbox shared folder pages contain file metadata in embedded JSON
        file_links = re.findall(
            r'(https://www\.dropbox\.com/scl/fi/[^"\'\\]+\.(?:csv|xlsx)[^"\'\\]*)',
            html_text, re.IGNORECASE)

        if not file_links:
            # Try alternate pattern
            file_links = re.findall(
                r'(https://[^"\'\\]*dropbox[^"\'\\]*(?:csv|xlsx)[^"\'\\]*)',
                html_text, re.IGNORECASE)

        # Filter for inventory-related files
        inv_links = [l for l in file_links
                     if "product" in l.lower() or "inventory" in l.lower()]
        if not inv_links:
            inv_links = file_links  # fallback to any csv/xlsx

        if not inv_links:
            return jsonify({"error": "No inventory files found in shared folder. "
                           "Configure Dropbox app credentials for reliable access."}), 404

        # Use the first (most recent) link — convert to direct download
        file_url = inv_links[0]
        # Strip existing query params and add dl=1
        base_url = file_url.split("?")[0]
        rlkey_match = re.search(r'rlkey=([^&"]+)', file_url)
        dl_url = base_url + "?dl=1"
        if rlkey_match:
            dl_url += "&rlkey=" + rlkey_match.group(1)

        dl_resp = req.get(dl_url, timeout=30,
                         headers={"User-Agent": "Mozilla/5.0"})
        dl_resp.raise_for_status()

        # Infer filename
        cd = dl_resp.headers.get("Content-Disposition", "")
        if "filename=" in cd:
            name = cd.split("filename=")[-1].strip('" ').split("'")[-1]
        else:
            name = base_url.split("/")[-1] or "inventory.xlsx"

    # 4. Parse inventory
    inv = {}
    if name.lower().endswith(".xlsx"):
        try:
            import openpyxl
        except ImportError as exc:
            import traceback
            traceback.print_exc()
            return jsonify({"error": f"openpyxl import failed: {exc}"}), 500
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        tmp.write(dl_resp.content)
        tmp.close()
        try:
            wb = openpyxl.load_workbook(tmp.name, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            headers = [str(c).strip() if c else "" for c in rows[0]]
            sku_col = next((i for i, h in enumerate(headers)
                          if h.lower() in ("product sku", "sku")), None)
            rmfg_col = next((i for i, h in enumerate(headers)
                           if h.upper() == "RMFG"), None)
            total_col = next((i for i, h in enumerate(headers)
                            if h.lower() == "total"), None)
            qty_col = total_col if total_col is not None else rmfg_col

            if sku_col is None or qty_col is None:
                return jsonify({"error": f"Can't find SKU/qty columns in {name}"}), 400

            # Build dict rows for bulk weight extraction
            ingredient_col = next((i for i, h in enumerate(headers)
                                  if h.lower() == "ingredient"), None)
            q1_col = next((i for i, h in enumerate(headers)
                          if h.lower() == "quantity1"), None)
            u1_col = next((i for i, h in enumerate(headers)
                          if h.lower() == "unit1"), None)
            xlsx_rows = []
            for row in rows[1:]:
                sku = str(row[sku_col]).strip() if row[sku_col] else ""
                try:
                    qty = int(float(str(row[qty_col] or 0)))
                except (ValueError, TypeError):
                    qty = 0
                if sku:
                    inv[normalize_sku(sku)] = inv.get(normalize_sku(sku), 0) + qty
                # Collect for bulk weight extraction (include Total for wheel counts)
                total_val = row[qty_col] if qty_col is not None else 0
                xlsx_rows.append({
                    "Ingredient": str(row[ingredient_col]).strip() if ingredient_col is not None and row[ingredient_col] else "",
                    "Product SKU": sku,
                    "Quantity1": str(row[q1_col]).strip() if q1_col is not None and row[q1_col] else "",
                    "Unit1": str(row[u1_col]).strip() if u1_col is not None and row[u1_col] else "",
                    "Total": total_val,
                })
            STATE["bulk_weights"] = extract_bulk_weights(xlsx_rows)
        finally:
            os.unlink(tmp.name)
    else:
        text = dl_resp.content.decode("utf-8-sig")
        csv_rows = list(csv.DictReader(io.StringIO(text)))
        for row in csv_rows:
            sku = row.get("Product SKU", row.get("SKU", "")).strip()
            if not sku:
                continue
            try:
                qty = int(float(row.get("Total", row.get("RMFG", "0")) or 0))
            except ValueError:
                qty = 0
            sku = normalize_sku(sku)
            inv[sku] = inv.get(sku, 0) + qty
        # Extract bulk weights from Quantity1/Unit1 columns
        STATE["bulk_weights"] = extract_bulk_weights(csv_rows)

    # Also add open PO quantities
    for po in s.get("open_pos", []):
        if po.get("status", "").lower() != "received":
            sku = normalize_sku(po.get("sku", ""))
            if sku:
                inv[sku] = inv.get(sku, 0) + int(po.get("qty", 0))

    ch_count = sum(1 for k in inv if k.startswith("CH-"))
    STATE["rmfg_inventory"] = inv

    # Auto-snapshot on Dropbox sync
    snap = _take_snapshot(f"Dropbox: {name}", source="dropbox")

    # Append journal snapshot entry
    journal = s.setdefault("inventory_journal", [])
    journal.append({
        "ts": snap["timestamp"],
        "type": "snapshot",
        "snapshot_id": snap["id"],
        "label": f"Dropbox: {name}",
    })

    # Detect production events if previous snapshot exists
    all_snaps = s.get("inventory_snapshots", [])
    if len(all_snaps) >= 2:
        prev_snap = all_snaps[-2]
        # Build depletion between prev and current from pre/post pairs
        dep_between = {}
        prev_ts = prev_snap.get("timestamp", "")
        curr_ts = snap["timestamp"]
        for idx, sn in enumerate(all_snaps):
            lbl = sn.get("label", "")
            sn_ts = sn.get("timestamp", "")
            if "Pre-depletion" in lbl and prev_ts <= sn_ts <= curr_ts:
                pre_inv = sn.get("inventory", {})
                if idx + 1 < len(all_snaps):
                    post_inv = all_snaps[idx + 1].get("inventory", {})
                    for dep_sku in pre_inv:
                        d = max(0, pre_inv.get(dep_sku, 0) - post_inv.get(dep_sku, 0))
                        if d > 0:
                            dep_between[dep_sku] = dep_between.get(dep_sku, 0) + d
        prod_events = detect_production_events(prev_snap, snap, dep_between)
        for evt in prod_events:
            journal.append({
                "ts": snap["timestamp"],
                "type": "production",
                "label": f"Cut {evt['wheels_cut']} wheels of {evt['sku']}",
                "sku": evt["sku"],
                "wheels_cut": evt["wheels_cut"],
                "expected_sliced": evt["expected"],
                "actual_sliced": evt["actual"],
                "variance": evt["variance"],
            })
            if abs(evt["variance"]) > 5:
                print(f"[WARN] Production variance for {evt['sku']}: "
                      f"expected={evt['expected']}, actual={evt['actual']}, "
                      f"variance={evt['variance']}")
    # Cap journal at 200 entries
    if len(journal) > 200:
        s["inventory_journal"] = journal[-200:]
    save_settings(s)

    bw = STATE.get("bulk_weights", {})
    pot_total = sum(d.get("potential_yield", 0) for d in bw.values())
    wheel_skus = sum(1 for d in bw.values() if d.get("count", 0) > 0)

    return jsonify({
        "ok": True,
        "source": "dropbox",
        "file": name,
        "modified": newest.get("server_modified", ""),
        "inventory_count": len(inv),
        "cheese_count": ch_count,
        "wheel_skus": wheel_skus,
        "potential_yield": pot_total,
    })


@app.route("/api/dropbox_status")
def dropbox_status():
    """Check if Dropbox is configured."""
    s = _s()
    return jsonify({
        "configured": bool(
            (s.get("dropbox_app_key") and (s.get("dropbox_refresh_token") or s.get("dropbox_access_token"))) or
            s.get("dropbox_shared_link")
        ),
        "has_shared_link": bool(s.get("dropbox_shared_link")),
    })


@app.route("/api/dropbox_auth_url")
def dropbox_auth_url():
    """Generate Dropbox OAuth2 authorization URL (no redirect, copy code)."""
    s = _s()
    app_key = s.get("dropbox_app_key", "")
    if not app_key:
        return jsonify({"error": "Set dropbox_app_key in settings first"}), 400
    url = (f"https://www.dropbox.com/oauth2/authorize"
           f"?client_id={app_key}"
           f"&response_type=code"
           f"&token_access_type=offline")
    return jsonify({"url": url, "instructions": "Open this URL, authorize, then paste the code at /api/dropbox_token?code=YOUR_CODE"})


@app.route("/api/dropbox_token")
def dropbox_token():
    """Exchange authorization code for permanent refresh token."""
    try:
        import requests as req
    except ImportError:
        return "requests module not installed", 500

    code = request.args.get("code", "").strip()
    if not code:
        return ("<h2>Dropbox Token Exchange</h2>"
                "<form action='/api/dropbox_token' method='get'>"
                "<label>Paste your authorization code:</label><br><br>"
                "<input type='text' name='code' style='width:400px;padding:8px;font-size:14px' placeholder='Paste code here...'>"
                "<br><br><button type='submit' style='padding:8px 20px;font-size:14px'>Submit</button>"
                "</form>")

    s = _s()
    app_key = s.get("dropbox_app_key", "")
    app_secret = s.get("dropbox_app_secret", "")

    resp = req.post("https://api.dropboxapi.com/oauth2/token", data={
        "code": code,
        "grant_type": "authorization_code",
        "client_id": app_key,
        "client_secret": app_secret,
    }, timeout=15)

    if resp.status_code != 200:
        return f"<h2>Token exchange failed</h2><pre>{resp.text}</pre>", 400

    data = resp.json()
    refresh_token = data.get("refresh_token", "")

    if refresh_token:
        s["dropbox_refresh_token"] = refresh_token
        s.pop("dropbox_access_token", None)
        save_settings(s)
        STATE["saved"] = load_settings()  # reload cached settings
        return ("<h2>Dropbox authorized!</h2>"
                "<p>Permanent refresh token saved. You can close this tab.</p>")
    else:
        return f"<h2>No refresh token</h2><pre>{data}</pre>", 400


# ── Recharge API integration ──────────────────────────────────────────

@app.route("/api/recharge_sync", methods=["POST"])
def recharge_sync():
    """Pull queued charges from Recharge API and resolve into cheese demand."""
    try:
        import requests as req
    except ImportError:
        return jsonify({"error": "requests library not installed"}), 500

    s = _s()
    api_token = s.get("recharge_api_token", "")
    if not api_token:
        return jsonify({"error": "Recharge API token not set in Settings"}), 400

    # Fetch queued charges
    session = req.Session()
    session.headers.update({
        "X-Recharge-Access-Token": api_token,
        "Accept": "application/json",
    })

    all_charges = []
    url = "https://api.rechargeapps.com/charges"
    # Only fetch charges scheduled in the next 4 weeks
    today = datetime.date.today()
    date_min = today.isoformat()
    date_max = (today + datetime.timedelta(days=28)).isoformat()
    params = {"status": "queued", "limit": 250,
              "scheduled_at_min": date_min, "scheduled_at_max": date_max}
    import time as _time
    t0 = _time.time()
    page_count = 0
    while True:
        # Retry with backoff on 429 rate limit
        for attempt in range(3):
            resp = session.get(url, params=params, timeout=30)
            if resp.status_code == 429:
                retry_after = int(resp.headers.get("Retry-After", 2))
                _time.sleep(retry_after)
                continue
            break
        resp.raise_for_status()
        data = resp.json()
        charges = data.get("charges", [])
        if not charges:
            break
        all_charges.extend(charges)
        page_count += 1

        # Cursor-based pagination (v2021-11 API — mandatory)
        next_cursor = data.get("next_cursor")
        if next_cursor:
            params = {"cursor": next_cursor, "limit": 250}
            continue
        break
    elapsed = round(_time.time() - t0, 1)

    if not all_charges:
        return jsonify({"error": "No queued charges found"}), 404

    # Resolve into cheese demand using per-charge box context
    pr_cjam = s.get("pr_cjam", {})
    cex_ec = s.get("cex_ec", {})
    splits = s.get("cexec_splits", {})
    global_extras = s.get("global_extras", {})

    # Build 4 weekly windows from today
    today = datetime.date.today()
    # Find next 4 Saturdays
    days_to_sat = (5 - today.weekday()) % 7
    if days_to_sat == 0:
        days_to_sat = 7
    saturdays = [today + datetime.timedelta(days=days_to_sat + 7 * i)
                 for i in range(4)]

    # Bin charges by week: which Saturday window they fall into
    # Each charge's scheduled_at determines its window
    week_demands = [defaultdict(int) for _ in range(4)]
    total_by_month = defaultdict(lambda: defaultdict(float))
    total_charges_count = 0

    for charge in all_charges:
        scheduled = charge.get("scheduled_at", "")
        if not scheduled:
            continue
        total_charges_count += 1

        # Parse date
        try:
            sched_date = datetime.date.fromisoformat(scheduled[:10])
        except (ValueError, TypeError):
            continue

        # Save by month for persistence
        month_label = scheduled[:7]

        # Find which weekly window this charge falls into
        week_idx = None
        for i, sat in enumerate(saturdays):
            # Window: previous Sunday through Saturday
            window_start = sat - datetime.timedelta(days=6)
            if window_start <= sched_date <= sat:
                week_idx = i
                break
        if week_idx is None:
            # Before first window or after last — put in nearest
            if sched_date <= saturdays[0]:
                week_idx = 0
            elif sched_date <= saturdays[-1]:
                week_idx = 3
            else:
                continue  # beyond 4 weeks, skip

        # Find box SKU on this charge for curation context
        items = charge.get("line_items", [])
        box_sku = None
        for item in items:
            sku = (item.get("sku") or "").strip().upper()
            if sku.startswith("AHB-"):
                box_sku = sku
                break
        curation = resolve_curation_from_box_sku(box_sku)

        for item in items:
            sku = (item.get("sku") or "").strip()
            if not sku:
                continue
            upper = sku.upper()
            qty = int(float(item.get("quantity", 1)))

            total_by_month[month_label][sku] += qty

            # PR-CJAM resolution with curation context
            if upper.startswith("PR-CJAM-"):
                suffix = upper.split("PR-CJAM-", 1)[1]
                if suffix == "GEN":
                    # Resolve using box curation
                    if curation and curation not in ("MONTHLY", None):
                        suffix = curation
                    else:
                        continue
                info = pr_cjam.get(suffix, {})
                ch = info.get("cheese", "") if isinstance(info, dict) \
                    else str(info)
                if ch:
                    week_demands[week_idx][normalize_sku(ch)] += qty
                continue

            # CEX-EC resolution
            if upper.startswith("CEX-EC-"):
                suffix = upper.split("CEX-EC-", 1)[1]
                if suffix in splits:
                    for ssku, pct in splits[suffix].items():
                        week_demands[week_idx][normalize_sku(ssku)] += \
                            int(qty * pct)
                else:
                    ec = cex_ec.get(suffix, "")
                    if ec:
                        week_demands[week_idx][normalize_sku(ec)] += qty
                continue

            if upper == "CEX-EC":
                if curation and curation not in ("MONTHLY", None):
                    ec = cex_ec.get(curation, "")
                    if isinstance(ec, str) and ec:
                        week_demands[week_idx][normalize_sku(ec)] += qty
                continue

            # Global extras resolution (bare SKUs across all curations)
            ge_resolved = global_extras.get(upper)
            if ge_resolved:
                week_demands[week_idx][normalize_sku(ge_resolved)] += qty
                continue

            if not is_pickable(sku):
                continue

            week_demands[week_idx][normalize_sku(sku)] += qty

    # Store per-week demand
    STATE["rmfg_sat_demand"] = dict(week_demands[0])      # This Saturday
    STATE["rmfg_tue_demand"] = dict(week_demands[1])       # Tuesday / week 2
    STATE["rmfg_next_sat_demand"] = dict(week_demands[2])  # Next Saturday
    STATE["rmfg_week4_demand"] = dict(week_demands[3])     # Week 4

    # Save to settings for persistence
    s["recharge_queued"] = {m: dict(skus)
                            for m, skus in total_by_month.items()}
    save_settings(s)
    STATE["saved"] = s

    ch_demand = sum(sum(v for k, v in wd.items() if k.startswith("CH-"))
                    for wd in week_demands)
    week_labels = ["This Sat", "Tuesday", "Next Sat", "Sat +3"]
    week_summary = [
        {"label": week_labels[i],
         "date": saturdays[i].isoformat(),
         "skus": len(week_demands[i]),
         "units": sum(week_demands[i].values())}
        for i in range(4)
    ]
    return jsonify({
        "ok": True,
        "total_charges": total_charges_count,
        "months": sorted(total_by_month.keys()),
        "weeks": week_summary,
        "cheese_demand_units": ch_demand,
        "api_pages": page_count,
        "api_seconds": elapsed,
    })


@app.route("/api/shopify_sync", methods=["POST"])
def shopify_sync():
    """Pull fulfilled Shopify orders from last 4 weeks to calculate weekly demand.

    Computes per-SKU weekly average from historical fulfilled orders.
    Saves to settings as shopify_api_demand {sku: weekly_qty}.
    Should be run after or with Recharge sync.
    """
    try:
        import requests as req
    except ImportError:
        return jsonify({"error": "requests library not installed"}), 500

    s = _s()
    store = s.get("shopify_store_url", "").strip()
    token = s.get("shopify_access_token", "").strip()
    if not store or not token:
        return jsonify({"error": "Shopify store URL or access token not configured"}), 400

    if not store.startswith("http"):
        store = f"https://{store}.myshopify.com"

    api_version = "2024-01"
    session = req.Session()
    session.headers.update({
        "X-Shopify-Access-Token": token,
        "Content-Type": "application/json",
    })

    # Pull fulfilled orders for weekly demand calculation
    try:
        weeks_back = max(1, min(52, int(_s().get("shopify_weeks_back", 8))))
    except (TypeError, ValueError):
        weeks_back = 8
    cutoff = (datetime.datetime.now() - datetime.timedelta(days=weeks_back * 7)).isoformat()
    url = f"{store}/admin/api/{api_version}/orders.json"
    params = {"status": "any", "fulfillment_status": "shipped",
              "limit": 250, "created_at_min": cutoff}

    import time as _time
    t0 = _time.time()
    all_orders = []
    page_count = 0
    while url:
        # Retry with backoff on 429 rate limit
        for attempt in range(3):
            resp = session.get(url, params=params, timeout=30)
            if resp.status_code == 429:
                retry_after = float(resp.headers.get("Retry-After", 2))
                _time.sleep(retry_after)
                continue
            break
        if resp.status_code != 200:
            return jsonify({"error": f"Shopify API error {resp.status_code}: {resp.text[:200]}"}), 400
        data = resp.json()
        orders = data.get("orders", [])
        all_orders.extend(orders)
        page_count += 1

        # Link header pagination (standard for Shopify REST API)
        url = None
        params = None
        link = resp.headers.get("Link", "")
        if 'rel="next"' in link:
            import re
            m = re.search(r'<([^>]+)>;\s*rel="next"', link)
            if m:
                url = m.group(1)
                params = None

    # Also pull currently unfulfilled orders (pending demand for this week)
    unfulfilled_url = f"{store}/admin/api/{api_version}/orders.json"
    unfulfilled_cutoff = (datetime.datetime.now() - datetime.timedelta(days=14)).isoformat()
    uf_params = {"status": "open", "fulfillment_status": "unfulfilled",
                 "limit": 250, "created_at_min": unfulfilled_cutoff}
    unfulfilled_orders = []
    while unfulfilled_url:
        resp = session.get(unfulfilled_url, params=uf_params, timeout=30)
        if resp.status_code == 200:
            data = resp.json()
            uf_orders = data.get("orders", [])
            unfulfilled_orders.extend(uf_orders)
            unfulfilled_url = None
            uf_params = None
            link = resp.headers.get("Link", "")
            if 'rel="next"' in link:
                import re
                m = re.search(r'<([^>]+)>;\s*rel="next"', link)
                if m:
                    unfulfilled_url = m.group(1)
        else:
            break

    # Bucket fulfilled orders by week (Saturday boundaries)
    today = datetime.date.today()
    # Per-week SKU totals for historical average
    week_sku_totals = defaultdict(lambda: defaultdict(int))
    order_count = 0
    first_order_total = defaultdict(int)
    recurring_total = defaultdict(int)

    for order in all_orders:
        order_count += 1
        created = order.get("created_at", "")[:10]
        try:
            order_date = datetime.date.fromisoformat(created)
        except (ValueError, TypeError):
            continue

        # Determine which week this order belongs to (week number from cutoff)
        days_ago = (today - order_date).days
        week_num = days_ago // 7

        # Detect first order vs recurring
        tags = order.get("tags", "") or ""
        is_first = "Subscription First Order" in tags
        is_subscription = "Subscription" in tags

        for item in order.get("line_items", []):
            sku = (item.get("sku") or "").strip()
            if not sku:
                continue
            qty = int(float(item.get("quantity", 1)))
            nsku = normalize_sku(sku)
            if is_pickable(nsku):
                week_sku_totals[week_num][nsku] += qty
                if is_first:
                    first_order_total[nsku] += qty
                elif is_subscription:
                    recurring_total[nsku] += qty

    # Compute weekly average per SKU from fulfilled history
    num_weeks_with_data = max(1, len(week_sku_totals))
    all_skus_seen = set()
    for wk_data in week_sku_totals.values():
        all_skus_seen.update(wk_data.keys())

    shopify_weekly = {}
    for sku in all_skus_seen:
        total = sum(week_sku_totals[wk].get(sku, 0) for wk in week_sku_totals)
        shopify_weekly[sku] = round(total / num_weeks_with_data)

    # Unfulfilled orders = pending demand for current week
    unfulfilled_demand = defaultdict(int)
    for order in unfulfilled_orders:
        for item in order.get("line_items", []):
            sku = (item.get("sku") or "").strip()
            if not sku:
                continue
            qty = int(float(item.get("quantity", 1)))
            nsku = normalize_sku(sku)
            if is_pickable(nsku):
                unfulfilled_demand[nsku] += qty

    # Save to settings
    s["shopify_api_demand"] = shopify_weekly
    s["shopify_unfulfilled"] = dict(unfulfilled_demand)
    # Save first-order vs recurring split (weekly averages)
    s["shopify_first_order_demand"] = {
        sku: round(qty / num_weeks_with_data)
        for sku, qty in first_order_total.items() if qty > 0
    }
    s["shopify_recurring_demand"] = {
        sku: round(qty / num_weeks_with_data)
        for sku, qty in recurring_total.items() if qty > 0
    }
    save_settings(s)
    STATE["saved"] = s

    total_weekly = sum(shopify_weekly.values())
    total_unfulfilled = sum(unfulfilled_demand.values())
    return jsonify({
        "ok": True,
        "orders_analyzed": order_count,
        "weeks_of_history": num_weeks_with_data,
        "skus": len(shopify_weekly),
        "weekly_avg_units": total_weekly,
        "unfulfilled_orders": len(unfulfilled_orders),
        "unfulfilled_units": total_unfulfilled,
        "first_order_skus": len(s.get("shopify_first_order_demand", {})),
        "recurring_skus": len(s.get("shopify_recurring_demand", {})),
        "api_pages": page_count,
        "api_seconds": round(_time.time() - t0, 1),
    })


@app.route("/api/shopify_status")
def shopify_status():
    s = _s()
    weekly = s.get("shopify_api_demand", {})
    return jsonify({
        "configured": bool(s.get("shopify_store_url") and
                          s.get("shopify_access_token")),
        "has_data": bool(weekly),
        "sku_count": len(weekly),
        "weekly_units": sum(weekly.values()) if weekly else 0,
    })


@app.route("/api/load_settings_inventory", methods=["POST"])
def load_settings_inventory():
    """Load inventory from settings JSON into RMFG state."""
    s = _s()
    inventory = s.get("inventory", {})
    inv = {}
    for sku, data in inventory.items():
        if isinstance(data, dict):
            qty = int(float(data.get("qty", 0)))
        else:
            qty = int(float(data))
        if qty > 0:
            inv[normalize_sku(sku)] = inv.get(normalize_sku(sku), 0) + qty

    # Add wheel supply
    wheel_inv = s.get("wheel_inventory", {})
    for wsku, wd in wheel_inv.items():
        if isinstance(wd, dict):
            w = float(wd.get("weight_lbs", 0))
            c = int(wd.get("count", 0))
            t = wd.get("target_sku", "")
            if t and w > 0 and c > 0:
                inv[t] = inv.get(t, 0) + int(w * c * WHEEL_TO_SLICE_FACTOR)

    # Add open POs
    for po in s.get("open_pos", []):
        if po.get("status", "").lower() != "received":
            sku = normalize_sku(po.get("sku", ""))
            if sku:
                inv[sku] = inv.get(sku, 0) + int(float(po.get("qty", 0)))

    STATE["rmfg_inventory"] = inv
    ch_count = sum(1 for k in inv if k.startswith("CH-"))
    return jsonify({
        "ok": True,
        "source": "settings",
        "inventory_count": len(inv),
        "cheese_count": ch_count,
    })


@app.route("/api/recharge_status")
def recharge_status():
    """Check if Recharge API is configured."""
    s = _s()
    return jsonify({
        "configured": bool(s.get("recharge_api_token")),
    })


# ══════════════════════════════════════════════════════════════════════════
#  ACTION CALENDAR — multi-week task schedule (PO, MFG, Crossdock, Ship)
# ══════════════════════════════════════════════════════════════════════════

def _next_weekday(start, weekday):
    """Return the next date on or after `start` that falls on `weekday` (0=Mon)."""
    days_ahead = weekday - start.weekday()
    if days_ahead < 0:
        days_ahead += 7
    return start + datetime.timedelta(days=days_ahead)


@app.route("/api/action_calendar", methods=["POST"])
def action_calendar():
    """Generate a multi-week action calendar from current data.

    Returns 4 weeks of daily tasks:
      - PO: purchase orders to place (Wed), with SKU/qty/vendor
      - MFG: manufacturing/production orders (Wed), wheel-to-slice
      - CROSSDOCK: incoming PO arrivals to receive
      - FULFILL: Tuesday + Saturday shipment volumes
      - INVENTORY: Monday/Friday inventory check reminders
    """
    s = _s()
    inv = STATE.get("rmfg_inventory") or {}
    sat_demand = STATE.get("rmfg_sat_demand") or {}
    tue_demand = STATE.get("rmfg_tue_demand") or {}
    next_sat_demand = STATE.get("rmfg_next_sat_demand") or {}
    open_pos = s.get("open_pos", [])
    vendor_catalog = s.get("vendor_catalog", {})
    wheel_inv = s.get("wheel_inventory", {})
    inventory = s.get("inventory", {})

    today = datetime.date.today()
    weeks = []

    # Build running inventory projection
    running_inv = {}
    for sku in set(list(inv.keys()) + list(sat_demand.keys()) +
                   list(tue_demand.keys()) + list(next_sat_demand.keys())):
        if sku.startswith("CH-") or sku.startswith("MT-") or sku.startswith("AC-"):
            running_inv[sku] = inv.get(sku, 0)

    # Demand windows: [(label, demand_dict, ship_weekday)]
    # ship_weekday: 5=Saturday, 1=Tuesday
    demand_windows = [
        ("This Saturday", sat_demand, 5),
        ("Tuesday", tue_demand, 1),
        ("Next Saturday", next_sat_demand, 5),
    ]

    # For weeks 4+, estimate demand as average of known windows
    avg_demand = {}
    total_windows = 0
    for label, dd, _ in demand_windows:
        if dd:
            total_windows += 1
            for sku, qty in dd.items():
                avg_demand[sku] = avg_demand.get(sku, 0) + qty
    if total_windows > 0:
        for sku in avg_demand:
            avg_demand[sku] = int(avg_demand[sku] / total_windows)

    # Generate 4 weeks of calendar
    week_start = today - datetime.timedelta(days=today.weekday())  # Monday

    for week_idx in range(4):
        ws = week_start + datetime.timedelta(weeks=week_idx)
        we = ws + datetime.timedelta(days=6)
        days = []

        # Determine which demand window applies this week
        if week_idx == 0:
            week_demand = sat_demand
            week_tue_demand = tue_demand
        elif week_idx == 1:
            week_demand = next_sat_demand
            week_tue_demand = {sku: int(qty * 0.3) for sku, qty in
                               next_sat_demand.items()} if next_sat_demand else {}
        else:
            week_demand = avg_demand
            week_tue_demand = {sku: int(qty * 0.3) for sku, qty in
                               avg_demand.items()}

        # Compute shortages for this week
        shortages = {}
        for sku in set(list(running_inv.keys()) + list(week_demand.keys())):
            avail = running_inv.get(sku, 0)
            demand = int(round(week_demand.get(sku, 0)))
            tue_d = int(round(week_tue_demand.get(sku, 0)))
            total = demand + tue_d
            net = avail - total
            if net < 0:
                shortages[sku] = {"deficit": abs(net), "avail": avail,
                                  "demand": total}

        week_po_lines = []
        week_mfg_lines = []

        bulk_wts = STATE.get("bulk_weights", {})
        for sku, info in sorted(shortages.items(), key=lambda x: -x[1]["deficit"]):
            if not sku.startswith("CH-"):
                continue
            deficit = info["deficit"]
            buf = max(deficit, int(deficit * 1.15))

            # Check if wheels can cover via MFG (bulk_weights preferred)
            bw = bulk_wts.get(sku, {})
            mfg_possible = bw.get("potential_yield", 0)
            wheel_weight = bw.get("weight_lbs", 0)
            if mfg_possible == 0:
                # Fallback to legacy wheel_inventory
                for wsku, wd in wheel_inv.items():
                    if isinstance(wd, dict) and wd.get("target_sku") == sku:
                        w = float(wd.get("weight_lbs", 0))
                        c = int(wd.get("count", 0))
                        mfg_possible += int(w * c * WHEEL_TO_SLICE_FACTOR)
                        wheel_weight = w

            if mfg_possible >= deficit:
                per_wheel = int(wheel_weight * WHEEL_TO_SLICE_FACTOR) if wheel_weight > 0 else 1
                wheels_needed = math.ceil(deficit / max(1, per_wheel))
                week_mfg_lines.append({
                    "sku": sku, "deficit": deficit,
                    "slices_needed": deficit,
                    "wheels_needed": wheels_needed,
                    "action": "CUT",
                })
            else:
                vi = vendor_catalog.get(sku, {})
                cq = int(vi.get("case_qty", 1)) or 1
                cases = math.ceil(buf / cq)
                week_po_lines.append({
                    "sku": sku, "deficit": deficit,
                    "order_qty": cases * cq, "cases": cases,
                    "case_qty": cq,
                    "vendor": vi.get("vendor", "TBD"),
                })

        # Build daily tasks for this week
        for day_offset in range(7):
            d = ws + datetime.timedelta(days=day_offset)
            dow = d.weekday()  # 0=Mon
            tasks = []

            # Monday: inventory check
            if dow == 0:
                tasks.append({
                    "type": "INVENTORY",
                    "title": "Check inventory report",
                    "detail": "Review fulfillment center inventory CSV",
                    "priority": "medium",
                })

            # Tuesday: fulfillment ship
            if dow == 1:
                tue_total = sum(week_tue_demand.values()) if week_tue_demand else 0
                tue_skus = len([k for k, v in week_tue_demand.items()
                                if v > 0]) if week_tue_demand else 0
                tasks.append({
                    "type": "FULFILL",
                    "title": f"Tuesday ship: {tue_total} units",
                    "detail": (f"{tue_skus} SKUs, first orders + requested"
                               if tue_skus else "First orders + requested"),
                    "priority": "high",
                    "units": tue_total,
                    "skus": tue_skus,
                })

            # Wednesday: POs + MFG
            if dow == 2:
                if week_po_lines:
                    total_po_units = sum(p["order_qty"] for p in week_po_lines)
                    vendors = list(set(p["vendor"] for p in week_po_lines))
                    tasks.append({
                        "type": "PO",
                        "title": f"Place POs: {len(week_po_lines)} lines, {total_po_units} units",
                        "detail": f"Vendors: {', '.join(vendors[:5])}",
                        "priority": "high",
                        "lines": week_po_lines,
                    })
                if week_mfg_lines:
                    total_slices = sum(m["slices_needed"] for m in week_mfg_lines)
                    tasks.append({
                        "type": "MFG",
                        "title": f"Production order: {total_slices} slices",
                        "detail": f"{len(week_mfg_lines)} SKUs to cut/wrap/label",
                        "priority": "high",
                        "lines": week_mfg_lines,
                    })

            # Friday: inventory check
            if dow == 4:
                tasks.append({
                    "type": "INVENTORY",
                    "title": "Check inventory report",
                    "detail": "Review fulfillment center inventory CSV",
                    "priority": "medium",
                })

            # Saturday: main fulfillment
            if dow == 5:
                sat_total = sum(week_demand.values()) if week_demand else 0
                sat_skus = len([k for k, v in week_demand.items()
                                if v > 0]) if week_demand else 0
                shortage_count = len(shortages)
                tasks.append({
                    "type": "FULFILL",
                    "title": f"Saturday ship: {sat_total} units",
                    "detail": (f"{sat_skus} SKUs, {shortage_count} shortages"
                               if shortage_count
                               else f"{sat_skus} SKUs, all covered"),
                    "priority": "critical" if shortage_count else "high",
                    "units": sat_total,
                    "skus": sat_skus,
                    "shortages": shortage_count,
                })

            # Crossdock: check open POs with ETA this day
            for po in open_pos:
                if po.get("status") != "Open":
                    continue
                eta_str = po.get("eta", "")
                try:
                    eta = datetime.date.fromisoformat(eta_str)
                except (ValueError, TypeError):
                    continue
                if eta == d:
                    tasks.append({
                        "type": "CROSSDOCK",
                        "title": f"Receive PO: {po.get('sku', '?')}",
                        "detail": f"{po.get('qty', '?')} units from {po.get('vendor', '?')}",
                        "priority": "high",
                        "sku": po.get("sku", ""),
                        "qty": po.get("qty", 0),
                    })

            days.append({
                "date": d.isoformat(),
                "dow": d.strftime("%a"),
                "day": d.day,
                "tasks": tasks,
                "is_today": d == today,
                "is_past": d < today,
            })

        # Deplete running inventory for next week
        for sku in running_inv:
            sat_d = int(round(week_demand.get(sku, 0)))
            tue_d = int(round(week_tue_demand.get(sku, 0)))
            running_inv[sku] = max(0, running_inv[sku] - sat_d - tue_d)

        weeks.append({
            "week": week_idx + 1,
            "start": ws.isoformat(),
            "end": we.isoformat(),
            "label": f"Week {week_idx + 1}: {ws.strftime('%b %d')} - {we.strftime('%b %d')}",
            "days": days,
            "po_lines": week_po_lines,
            "mfg_lines": week_mfg_lines,
            "shortages": len(shortages),
            "total_demand": sum(week_demand.values()) if week_demand else 0,
        })

    return jsonify({"weeks": weeks, "generated": today.isoformat()})


# ── Invoice API ────────────────────────────────────────────────────────


def _get_bulk_weights() -> dict:
    """Get bulk weights from STATE (populated by Dropbox sync) or load from local CSV."""
    bw = STATE.get("bulk_weights")
    if bw:
        return bw
    # Fallback: load from local Product Inventory CSV
    base = _get_project_dir()
    for pat in ("Product Inventory*.csv", "RMFG_*/Product Inventory*.csv"):
        import glob as globmod
        files = sorted(globmod.glob(os.path.join(base, pat)),
                       key=os.path.getmtime, reverse=True)
        if files:
            try:
                with open(files[0], encoding="utf-8-sig") as f:
                    rows = list(csv.DictReader(f))
                bw = extract_bulk_weights(rows)
                STATE["bulk_weights"] = bw
                return bw
            except Exception:
                pass
    return {}


from invoice_processor import (
    parse_production_invoice, extract_invoice_id,
    match_product_to_sku, get_match_candidates, SEED_TRANSLATIONS,
    gmail_connect, search_rmfg_invoices,
    reconcile_invoice_with_pos, apply_reconciliation,
    compute_yield_ratios, annotate_invoice_yields,
    extract_bulk_weights,
)


@app.route("/api/invoice_status")
def invoice_status():
    """Config check + invoice counts."""
    s = _s()
    smtp_user = s.get("smtp_user", "")
    smtp_pass = s.get("smtp_password", "")
    invoices = s.get("production_invoices", [])
    pending = len([i for i in invoices if i.get("status") in ("pending", "partial")])
    total_charge = sum(i.get("total_production_charge", 0) for i in invoices)
    return jsonify({
        "configured": bool(smtp_user and smtp_pass),
        "total_invoices": len(invoices),
        "pending_match": pending,
        "total_production_charge": round(total_charge, 2),
        "last_sync": s.get("_last_invoice_sync", ""),
    })


_invoice_sync_state = {"running": False, "progress": "", "result": None}


def _invoice_sync_worker(force: bool):
    """Background worker for invoice sync."""
    global _invoice_sync_state
    try:
        s = _s()
        smtp_user = s.get("smtp_user", "")
        smtp_pass = s.get("smtp_password", "")

        _invoice_sync_state["progress"] = "Connecting to Gmail IMAP..."

        subject_filter = s.get("imap_search_subject", "Production Breakdown")
        processed_ids = [] if force else list(s.get("processed_invoice_ids", []))
        sku_translations = s.get("sku_translations", {})
        inventory = s.get("inventory", {})

        try:
            conn = gmail_connect(smtp_user, smtp_pass)
            _invoice_sync_state["progress"] = "Searching inbox for invoices..."
            new_emails = search_rmfg_invoices(conn, subject_filter, processed_ids)
            conn.logout()
        except Exception as e:
            _invoice_sync_state["result"] = {"error": f"IMAP connection failed: {e}"}
            _invoice_sync_state["running"] = False
            return

        total_emails = len(new_emails)
        _invoice_sync_state["progress"] = f"Found {total_emails} emails. Parsing PDFs..."

        # Only clear after successful IMAP connection
        if force:
            s["production_invoices"] = []
        invoices = s.setdefault("production_invoices", [])

        new_count = 0
        for idx, em in enumerate(new_emails):
            for att in em.get("attachments", []):
                invoice_id = extract_invoice_id(att["filename"])
                _invoice_sync_state["progress"] = f"Parsing {invoice_id} ({idx + 1}/{total_emails})..."

                if any(inv.get("id") == invoice_id for inv in invoices):
                    continue

                parsed = parse_production_invoice(att["pdf_bytes"])

                line_items = []
                unmatched = []
                for section_key in ("full_mfg", "meals", "label_only"):
                    for item in parsed.get(section_key, []):
                        sku, confidence, method = match_product_to_sku(
                            item["product_name"], sku_translations, inventory)
                        li = {
                            "section": section_key,
                            "product_name": item["product_name"],
                            "sku": sku,
                            "match_confidence": confidence,
                            "match_method": method,
                            "case_packouts": item.get("case_packouts", 0),
                            "total_yield": item.get("total_yield", 0),
                            "estimated_cost": None,
                        }
                        line_items.append(li)
                        if not sku:
                            unmatched.append(item["product_name"])

                invoice_rec = {
                    "id": invoice_id,
                    "gmail_msg_id": em["msg_id"],
                    "filename": att["filename"],
                    "received_date": em["date"],
                    "mfg_date": parsed.get("mfg_date"),
                    "status": "matched" if not unmatched else
                              "partial" if len(unmatched) < len(line_items) else "pending",
                    "total_production_charge": parsed.get("total_production_charge", 0),
                    "full_mfg_charge": parsed.get("full_mfg_totals", {}).get("charge", 0),
                    "label_only_charge": parsed.get("label_only_totals", {}).get("charge", 0),
                    "meals_charge": parsed.get("meals_totals", {}).get("charge", 0),
                    "full_mfg_cases": parsed.get("full_mfg_totals", {}).get("cases", 0),
                    "full_mfg_yield": parsed.get("full_mfg_totals", {}).get("yield", 0),
                    "label_only_yield": parsed.get("label_only_totals", {}).get("yield", 0),
                    "line_items": line_items,
                    "unmatched_products": unmatched,
                    "po_matches": [],
                    "parse_method": parsed.get("parse_method", "unknown"),
                }
                invoices.append(invoice_rec)
                new_count += 1

            processed_ids.append(em["msg_id"])

        _invoice_sync_state["progress"] = "Saving..."
        s["processed_invoice_ids"] = processed_ids
        s["_last_invoice_sync"] = datetime.date.today().isoformat()
        save_settings(s)

        _invoice_sync_state["result"] = {
            "ok": True,
            "new_invoices": new_count,
            "total_invoices": len(invoices),
            "emails_checked": total_emails,
        }
    except Exception as e:
        _invoice_sync_state["result"] = {"error": str(e)}
    finally:
        _invoice_sync_state["running"] = False


@app.route("/api/invoice_sync", methods=["POST"])
def invoice_sync():
    """Start invoice sync (runs in background thread).
    Pass {"force": true} to re-parse all invoices from Gmail."""
    global _invoice_sync_state
    s = _s()
    smtp_user = s.get("smtp_user", "")
    smtp_pass = s.get("smtp_password", "")
    if not smtp_user or not smtp_pass:
        return jsonify({"error": "SMTP/IMAP not configured. Set smtp_user and smtp_password."})

    if _invoice_sync_state["running"]:
        return jsonify({"error": "Sync already in progress", "progress": _invoice_sync_state["progress"]})

    body = request.get_json(silent=True) or {}
    force = body.get("force", False)

    _invoice_sync_state = {"running": True, "progress": "Starting...", "result": None}
    t = threading.Thread(target=_invoice_sync_worker, args=(force,), daemon=True)
    t.start()

    return jsonify({"ok": True, "started": True, "progress": "Starting..."})


@app.route("/api/invoice_sync_progress")
def invoice_sync_progress():
    """Poll sync progress."""
    return jsonify({
        "running": _invoice_sync_state["running"],
        "progress": _invoice_sync_state["progress"],
        "result": _invoice_sync_state["result"],
    })


@app.route("/api/invoices")
def list_invoices():
    """List all invoices (summary)."""
    s = _s()
    invoices = s.get("production_invoices", [])
    summaries = []
    for inv in invoices:
        products = len(inv.get("line_items", []))
        cases = sum(li.get("case_packouts", 0) for li in inv.get("line_items", []))
        yld = sum(li.get("total_yield", 0) for li in inv.get("line_items", []))
        summaries.append({
            "id": inv.get("id"),
            "mfg_date": inv.get("mfg_date"),
            "received_date": inv.get("received_date"),
            "products": products,
            "cases": cases,
            "total_yield": yld,
            "total_charge": inv.get("total_production_charge", 0),
            "status": inv.get("status", "unknown"),
            "unmatched_count": len(inv.get("unmatched_products", [])),
        })
    return jsonify({"invoices": summaries})


@app.route("/api/invoice/<invoice_id>")
def get_invoice(invoice_id):
    """Full invoice detail."""
    s = _s()
    invoices = s.get("production_invoices", [])
    for inv in invoices:
        if inv.get("id") == invoice_id:
            return jsonify(inv)
    return jsonify({"error": f"Invoice {invoice_id} not found"})


@app.route("/api/invoice_map_sku", methods=["POST"])
def invoice_map_sku():
    """Manual product→SKU mapping."""
    s = _s()
    data = request.json or {}
    product_name = data.get("product_name", "")
    sku = data.get("sku", "")
    if not product_name or not sku:
        return jsonify({"error": "product_name and sku required"})

    # Save to sku_translations
    translations = s.setdefault("sku_translations", {})
    translations[product_name] = sku

    # Update all invoices that have this unmatched product
    invoices = s.get("production_invoices", [])
    updated = 0
    for inv in invoices:
        for li in inv.get("line_items", []):
            if li.get("product_name") == product_name and not li.get("sku"):
                li["sku"] = sku
                li["match_confidence"] = 1.0
                li["match_method"] = "manual"
                updated += 1
        # Recalculate unmatched list and status
        unmatched = [li["product_name"] for li in inv.get("line_items", [])
                     if not li.get("sku")]
        inv["unmatched_products"] = unmatched
        if not unmatched:
            inv["status"] = "matched"
        elif len(unmatched) < len(inv.get("line_items", [])):
            inv["status"] = "partial"

    save_settings(s)
    return jsonify({"ok": True, "updated": updated, "product_name": product_name, "sku": sku})


@app.route("/api/invoice_match_candidates", methods=["POST"])
def invoice_match_candidates():
    """Get ranked SKU candidates for an unmatched product name."""
    s = _s()
    data = request.json or {}
    product_name = data.get("product_name", "")
    if not product_name:
        return jsonify({"error": "product_name required"})
    candidates = get_match_candidates(
        product_name, s.get("sku_translations", {}), s.get("inventory", {}))
    return jsonify({"product_name": product_name, "candidates": candidates})


@app.route("/api/invoice_auto_map", methods=["POST"])
def invoice_auto_map():
    """Auto-map all unmatched products where a strong candidate exists (score >= 0.55)."""
    s = _s()
    translations = s.setdefault("sku_translations", {})
    inventory = s.get("inventory", {})
    invoices = s.get("production_invoices", [])

    # Collect unique unmatched products
    unmatched_set = set()
    for inv in invoices:
        for p in inv.get("unmatched_products", []):
            unmatched_set.add(p)

    mapped = []
    skipped = []
    for product_name in sorted(unmatched_set):
        candidates = get_match_candidates(product_name, translations, inventory)
        recommended = [c for c in candidates if c.get("recommended")]
        if recommended:
            best = recommended[0]
            translations[product_name] = best["sku"]
            mapped.append({"product_name": product_name, "sku": best["sku"],
                           "score": best["score"], "name": best["name"]})
        else:
            top = candidates[0] if candidates else None
            skipped.append({"product_name": product_name,
                            "best_sku": top["sku"] if top else "",
                            "best_score": top["score"] if top else 0})

    # Apply mappings to all invoices
    total_updated = 0
    for inv in invoices:
        for li in inv.get("line_items", []):
            if not li.get("sku") and li["product_name"] in translations:
                li["sku"] = translations[li["product_name"]]
                li["match_confidence"] = 1.0
                li["match_method"] = "auto_recommend"
                total_updated += 1
        unmatched = [li["product_name"] for li in inv.get("line_items", [])
                     if not li.get("sku")]
        inv["unmatched_products"] = unmatched
        if not unmatched:
            inv["status"] = "matched"
        elif len(unmatched) < len(inv.get("line_items", [])):
            inv["status"] = "partial"

    save_settings(s)
    return jsonify({
        "ok": True,
        "mapped": mapped,
        "skipped": skipped,
        "total_updated": total_updated,
    })


@app.route("/api/invoice_reconcile/<invoice_id>", methods=["POST"])
def invoice_reconcile(invoice_id):
    """Run reconciliation against open POs."""
    s = _s()
    result = apply_reconciliation(invoice_id, s)
    if "error" in result:
        return jsonify(result)
    save_settings(s)
    return jsonify({"ok": True, **result})


@app.route("/api/invoice_yield_ratios")
def invoice_yield_ratios():
    """Per-SKU yield-per-case ratios computed from invoice history."""
    s = _s()
    invoices = s.get("production_invoices", [])
    bw = _get_bulk_weights()
    ratios = compute_yield_ratios(invoices, bulk_weights=bw)
    return jsonify({"ratios": ratios})


@app.route("/api/invoice_yield/<invoice_id>")
def invoice_yield_detail(invoice_id):
    """Annotated yield analysis for a single invoice."""
    s = _s()
    invoices = s.get("production_invoices", [])
    invoice = None
    for inv in invoices:
        if inv.get("id") == invoice_id:
            invoice = inv
            break
    if not invoice:
        return jsonify({"error": f"Invoice {invoice_id} not found"})

    bw = _get_bulk_weights()
    ratios = compute_yield_ratios(invoices, bulk_weights=bw)
    annotations = annotate_invoice_yields(invoice, ratios)
    return jsonify({"invoice_id": invoice_id, "annotations": annotations, "ratios": ratios})


@app.route("/api/invoice_cost_history")
def invoice_cost_history():
    """Per-SKU cost analytics."""
    s = _s()
    cost_history = s.get("production_cost_history", [])

    # Group by SKU
    by_sku = {}
    for entry in cost_history:
        sku = entry.get("sku", "")
        if sku not in by_sku:
            by_sku[sku] = {"entries": [], "total_yield": 0, "total_cost": 0}
        by_sku[sku]["entries"].append(entry)
        by_sku[sku]["total_yield"] += entry.get("yield", 0)
        by_sku[sku]["total_cost"] += entry.get("estimated_cost", 0)

    # Calculate averages
    analytics = []
    for sku, data in sorted(by_sku.items()):
        avg_cost = (data["total_cost"] / data["total_yield"]
                    if data["total_yield"] > 0 else 0)
        analytics.append({
            "sku": sku,
            "total_yield": data["total_yield"],
            "total_cost": round(data["total_cost"], 2),
            "avg_cost_per_unit": round(avg_cost, 2),
            "entries": len(data["entries"]),
            "history": data["entries"],
        })

    return jsonify({"analytics": analytics})


# ── Depletion File Endpoints ──────────────────────────────────────────


@app.route("/api/depletion_parse", methods=["POST"])
def depletion_parse():
    """Parse an uploaded depletion XLSX file. Returns product totals + SKU mapping."""
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    f = request.files["file"]
    if not f.filename.endswith(".xlsx"):
        return jsonify({"error": "Must be .xlsx file"}), 400

    # Save to temp, parse
    import tempfile
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        f.save(tmp.name)
        tmp_path = tmp.name

    try:
        product_totals, order_count, err = parse_depletion_xlsx(tmp_path)
    finally:
        os.unlink(tmp_path)

    if err:
        return jsonify({"error": err}), 500

    # Map to SKUs
    s = _s()
    translations = s.get("sku_translations", {})
    inventory = s.get("inventory", {})
    sku_totals, mapped, unmatched = map_depletion_to_skus(
        product_totals, translations, inventory)

    return jsonify({
        "ok": True,
        "filename": f.filename,
        "order_count": order_count,
        "product_count": len(product_totals),
        "total_units": sum(product_totals.values()),
        "sku_totals": sku_totals,
        "mapped": mapped,
        "unmatched": unmatched,
        "mapped_count": len(mapped),
        "unmatched_count": len(unmatched),
    })


@app.route("/api/depletion_apply", methods=["POST"])
def depletion_apply():
    """Apply depletion to current inventory. Subtracts from STATE inventory."""
    data = request.json or {}
    sku_totals = data.get("sku_totals", {})
    label = data.get("label", "Depletion")

    if not sku_totals:
        return jsonify({"error": "No SKU totals provided"}), 400

    inv = STATE.get("rmfg_inventory", {})
    if not inv:
        return jsonify({"error": "No inventory loaded"}), 400

    # Snapshot before applying
    _take_snapshot(f"Pre-depletion: {label}", source="depletion")

    applied = {}
    for sku, qty in sku_totals.items():
        before = inv.get(sku, 0)
        after = before - int(qty)
        inv[sku] = after
        applied[sku] = {"before": before, "after": after, "depleted": int(qty)}

    STATE["rmfg_inventory"] = inv

    # Log depletion in settings history
    s = _s()
    history = s.setdefault("depletion_history", [])
    history.append({
        "date": datetime.datetime.now().isoformat(),
        "file": label,
        "day": _detect_cycle_day(),
        "skus": len(applied),
        "total": sum(int(q) for q in sku_totals.values()),
        "total_orders": data.get("order_count", 0),
    })

    # Append journal entry for running inventory
    journal = s.setdefault("inventory_journal", [])
    journal.append({
        "ts": datetime.datetime.now().isoformat(),
        "type": "depletion",
        "label": label,
        "sku_deltas": {sku: -int(qty) for sku, qty in sku_totals.items()},
    })

    save_settings(s)

    # Snapshot after applying
    _take_snapshot(f"Post-depletion: {label}", source="depletion")

    return jsonify({
        "ok": True,
        "applied": applied,
        "total_depleted": sum(int(q) for q in sku_totals.values()),
        "skus_affected": len(applied),
    })


@app.route("/api/depletion_map_sku", methods=["POST"])
def depletion_map_sku():
    """Save a product name → SKU translation for future depletion files."""
    data = request.json or {}
    product = data.get("product", "")
    sku = data.get("sku", "")

    if not product or not sku:
        return jsonify({"error": "product and sku required"}), 400

    s = _s()
    translations = s.setdefault("sku_translations", {})
    translations[product] = sku
    save_settings(s)

    return jsonify({"ok": True, "product": product, "sku": sku})


@app.route("/api/import_sku_translations", methods=["POST"])
def import_sku_translations():
    """Import SKU translations from a meal-type-export CSV upload."""
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    f = request.files["file"]
    if not f.filename:
        return jsonify({"error": "Empty filename"}), 400

    try:
        content = f.read().decode("utf-8-sig")
        reader = csv.reader(io.StringIO(content))
        s = _s()
        translations = s.setdefault("sku_translations", {})
        count = 0
        for row in reader:
            if len(row) >= 2:
                sku_code = row[0].strip()
                product_name = row[1].strip()
                # Strip AHB (S_REG): prefix
                if ": " in product_name:
                    product_name = product_name.split(": ", 1)[1]
                if product_name and sku_code:
                    translations[product_name] = sku_code
                    count += 1
        save_settings(s)
        return jsonify({"ok": True, "count": count})
    except Exception as e:
        return jsonify({"error": f"Parse error: {e}"}), 400


# ── Inventory Snapshots ────────────────────────────────────────────────


def _take_snapshot(label, source="manual", cycle_day=None):
    """Save a point-in-time inventory snapshot. Returns the snapshot dict."""
    s = _s()
    inv = STATE.get("rmfg_inventory", {})
    if not inv:
        # Fallback to settings inventory
        raw = s.get("inventory", {})
        inv = {k: (v.get("qty", 0) if isinstance(v, dict) else int(v or 0))
               for k, v in raw.items()}

    # Include potential yield from wheels/blocks
    bulk_weights = STATE.get("bulk_weights", {})
    potential = {}
    for sku, bw in bulk_weights.items():
        py = bw.get("potential_yield", 0)
        wc = bw.get("count", 0)
        if py > 0:
            potential[sku] = {"yield": py, "wheels": wc,
                              "weight_lbs": bw.get("weight_lbs", 0)}

    snap = {
        "id": datetime.datetime.now().strftime("%Y%m%d_%H%M%S"),
        "timestamp": datetime.datetime.now().isoformat(),
        "label": label,
        "source": source,
        "cycle_day": cycle_day or _detect_cycle_day(),
        "inventory": dict(inv),
        "potential_yield": potential,
        "sat_demand": dict(STATE.get("rmfg_sat_demand", {})),
        "tue_demand": dict(STATE.get("rmfg_tue_demand", {})),
    }

    snapshots = s.setdefault("inventory_snapshots", [])
    snapshots.append(snap)
    save_settings(s)
    return snap


def _detect_cycle_day():
    """Detect current cycle day: friday/saturday/monday/tuesday/wednesday."""
    wd = datetime.date.today().weekday()  # 0=Mon
    return {0: "monday", 1: "tuesday", 2: "wednesday",
            4: "friday", 5: "saturday"}.get(wd, f"day_{wd}")


@app.route("/api/snapshots")
def list_snapshots():
    """List all inventory snapshots (summary only, no full inventory)."""
    s = _s()
    snapshots = s.get("inventory_snapshots", [])
    summaries = []
    for snap in snapshots:
        inv = snap.get("inventory", {})
        pot = snap.get("potential_yield", {})
        pot_total = sum(p.get("yield", 0) for p in pot.values())
        summaries.append({
            "id": snap["id"],
            "timestamp": snap["timestamp"],
            "label": snap["label"],
            "source": snap.get("source", ""),
            "cycle_day": snap.get("cycle_day", ""),
            "sku_count": len(inv),
            "total_units": sum(inv.values()),
            "cheese_units": sum(v for k, v in inv.items() if k.startswith("CH-")),
            "potential_yield": pot_total,
            "wheel_skus": len(pot),
        })
    return jsonify({"snapshots": summaries})


@app.route("/api/snapshot/<snap_id>")
def get_snapshot(snap_id):
    """Get full snapshot detail."""
    s = _s()
    for snap in s.get("inventory_snapshots", []):
        if snap["id"] == snap_id:
            return jsonify(snap)
    return jsonify({"error": "Snapshot not found"}), 404


@app.route("/api/snapshot", methods=["POST"])
def save_snapshot():
    """Manually save a snapshot of current inventory state."""
    data = request.json or {}
    label = data.get("label", "Manual snapshot")
    snap = _take_snapshot(label, source="manual")
    return jsonify({"ok": True, "snapshot": {
        "id": snap["id"], "label": snap["label"],
        "sku_count": len(snap["inventory"]),
        "total_units": sum(snap["inventory"].values()),
    }})


@app.route("/api/snapshot/<snap_id>", methods=["DELETE"])
def delete_snapshot(snap_id):
    """Delete a snapshot."""
    s = _s()
    snapshots = s.get("inventory_snapshots", [])
    s["inventory_snapshots"] = [sn for sn in snapshots if sn["id"] != snap_id]
    save_settings(s)
    return jsonify({"ok": True})


@app.route("/api/snapshot_compare")
def compare_snapshots():
    """Compare two snapshots side by side. Query: ?a=ID&b=ID"""
    s = _s()
    a_id = request.args.get("a", "")
    b_id = request.args.get("b", "")
    snapshots = {sn["id"]: sn for sn in s.get("inventory_snapshots", [])}

    snap_a = snapshots.get(a_id)
    snap_b = snapshots.get(b_id)
    if not snap_a or not snap_b:
        return jsonify({"error": "Snapshot(s) not found"}), 404

    inv_a = snap_a.get("inventory", {})
    inv_b = snap_b.get("inventory", {})
    pot_a = snap_a.get("potential_yield", {})
    pot_b = snap_b.get("potential_yield", {})
    all_skus = sorted(set(inv_a) | set(inv_b))

    rows = []
    for sku in all_skus:
        qa = inv_a.get(sku, 0)
        qb = inv_b.get(sku, 0)
        delta = qb - qa
        if qa == 0 and qb == 0:
            continue
        pa = pot_a.get(sku, {}).get("yield", 0)
        pb = pot_b.get(sku, {}).get("yield", 0)
        rows.append({
            "sku": sku,
            "qty_a": qa,
            "qty_b": qb,
            "delta": delta,
            "pct_change": round(delta / qa * 100, 1) if qa else None,
            "potential_a": pa,
            "potential_b": pb,
        })

    return jsonify({
        "a": {"id": snap_a["id"], "label": snap_a["label"],
              "timestamp": snap_a["timestamp"], "cycle_day": snap_a.get("cycle_day", "")},
        "b": {"id": snap_b["id"], "label": snap_b["label"],
              "timestamp": snap_b["timestamp"], "cycle_day": snap_b.get("cycle_day", "")},
        "rows": rows,
        "summary": {
            "total_a": sum(inv_a.values()),
            "total_b": sum(inv_b.values()),
            "net_change": sum(inv_b.values()) - sum(inv_a.values()),
            "skus_gained": sum(1 for r in rows if r["qty_a"] == 0 and r["qty_b"] > 0),
            "skus_lost": sum(1 for r in rows if r["qty_a"] > 0 and r["qty_b"] == 0),
            "skus_increased": sum(1 for r in rows if r["delta"] > 0),
            "skus_decreased": sum(1 for r in rows if r["delta"] < 0),
        },
    })


@app.route("/api/snapshot_current")
def snapshot_current():
    """Get current inventory state (not saved, just for preview/comparison)."""
    inv = STATE.get("rmfg_inventory", {})
    if not inv:
        s = _s()
        raw = s.get("inventory", {})
        inv = {k: (v.get("qty", 0) if isinstance(v, dict) else int(v or 0))
               for k, v in raw.items()}
    return jsonify({
        "inventory": inv,
        "sat_demand": STATE.get("rmfg_sat_demand", {}),
        "tue_demand": STATE.get("rmfg_tue_demand", {}),
        "sku_count": len(inv),
        "total_units": sum(inv.values()),
    })


# ── Inventory Reconciliation ───────────────────────────────────────────


@app.route("/api/reconcile_inventory", methods=["POST"])
def reconcile_inventory():
    """Compare actual inventory (Monday snapshot) vs expected (Friday - depletion + yields).

    Body: {monday_snap_id?, friday_snap_id?}
    If IDs not provided, auto-detect most recent matching snapshots.
    """
    s = _s()
    body = request.get_json(silent=True) or {}
    snapshots = s.get("inventory_snapshots", [])

    if not snapshots:
        return jsonify({"error": "No snapshots available. Sync Dropbox or save a manual snapshot first."})

    # Find Monday (actual) and Friday (baseline) snapshots
    monday_id = body.get("monday_snap_id")
    friday_id = body.get("friday_snap_id")

    monday_snap = None
    friday_snap = None

    if monday_id:
        monday_snap = next((sn for sn in snapshots if sn["id"] == monday_id), None)
    if friday_id:
        friday_snap = next((sn for sn in snapshots if sn["id"] == friday_id), None)

    # Auto-detect: find most recent Monday and most recent Friday/pre-depletion
    if not monday_snap:
        for sn in reversed(snapshots):
            day = sn.get("cycle_day", "")
            if day == "monday" or "monday" in sn.get("label", "").lower():
                monday_snap = sn
                break
        # Fallback: use the most recent snapshot as "actual"
        if not monday_snap:
            monday_snap = snapshots[-1]

    if not friday_snap:
        for sn in reversed(snapshots):
            if sn["id"] == monday_snap["id"]:
                continue
            day = sn.get("cycle_day", "")
            src = sn.get("source", "")
            label = sn.get("label", "").lower()
            # Prefer: Friday, pre-depletion, or Dropbox sync before Monday
            if (day in ("friday", "saturday") or src == "depletion"
                    or "dropbox" in label or "friday" in label):
                friday_snap = sn
                break
        # Fallback: second most recent
        if not friday_snap and len(snapshots) >= 2:
            for sn in reversed(snapshots):
                if sn["id"] != monday_snap["id"]:
                    friday_snap = sn
                    break

    if not monday_snap or not friday_snap:
        return jsonify({"error": "Need at least 2 snapshots to reconcile."})

    actual = monday_snap.get("inventory", {})
    baseline = friday_snap.get("inventory", {})

    # Gather per-SKU depletion from pre/post snapshot pairs between Fri and Mon
    fri_ts = friday_snap.get("timestamp", "")
    mon_ts = monday_snap.get("timestamp", "")
    depletion_between = {}
    depletion_entries = []

    # Extract per-SKU depletion from pre/post snapshot deltas
    for i, snap in enumerate(snapshots):
        label = snap.get("label", "")
        ts = snap.get("timestamp", "")
        if "Pre-depletion" in label and fri_ts <= ts <= mon_ts:
            pre_inv = snap.get("inventory", {})
            if i + 1 < len(snapshots):
                post_inv = snapshots[i + 1].get("inventory", {})
                for sku in pre_inv:
                    delta = max(0, pre_inv.get(sku, 0) - post_inv.get(sku, 0))
                    if delta > 0:
                        depletion_between[sku] = depletion_between.get(sku, 0) + delta

    # Also count depletion history entries in the window
    depletions = s.get("depletion_history", [])
    for dep in depletions:
        dep_ts = dep.get("date", "")
        if fri_ts <= dep_ts <= mon_ts:
            depletion_entries.append(dep)

    # Gather invoice yields between the two snapshots
    invoices = s.get("production_invoices", [])
    invoice_yields = {}  # sku -> total yield from invoices in the window
    invoice_entries = []
    for inv in invoices:
        inv_date = inv.get("mfg_date") or inv.get("received_date", "")
        # Include invoices with mfg_date in the window
        if inv_date and fri_ts[:10] <= inv_date <= mon_ts[:10]:
            invoice_entries.append(inv)
            for li in inv.get("line_items", []):
                sku = li.get("sku", "")
                yld = li.get("total_yield", 0)
                if sku and yld > 0:
                    invoice_yields[sku] = invoice_yields.get(sku, 0) + yld

    # Wheel reconciliation data from potential_yield in snapshots
    fri_pot = friday_snap.get("potential_yield", {})
    mon_pot = monday_snap.get("potential_yield", {})

    # Build reconciliation rows
    all_skus = sorted(set(list(actual.keys()) + list(baseline.keys())))
    rows = []
    total_discrepancy = 0
    flagged_count = 0
    threshold_pct = s.get("yield_reconciliation_threshold_pct", 5)
    threshold_min = s.get("yield_reconciliation_threshold_min", 2)

    for sku in all_skus:
        qty_fri = baseline.get(sku, 0)
        qty_mon = actual.get(sku, 0)
        inv_yield = invoice_yields.get(sku, 0)
        depletion_total = depletion_between.get(sku, 0)

        # Expected Monday = Friday - depletion + invoice yields
        expected = qty_fri - depletion_total + inv_yield
        diff = qty_mon - expected
        abs_diff = abs(diff)

        # Wheel reconciliation for CH-* SKUs
        wheel_fri = fri_pot.get(sku, {}).get("wheels", 0)
        wheel_mon = mon_pot.get(sku, {}).get("wheels", 0)
        expected_cutting = 0
        cutting_disc = 0
        if sku.startswith("CH-") and (sku in fri_pot or sku in mon_pot):
            wheels_cut = max(0, wheel_fri - wheel_mon)
            if wheels_cut > 0:
                weight_lbs = fri_pot.get(sku, mon_pot.get(sku, {})).get("weight_lbs", 0)
                expected_cutting = round(wheels_cut * weight_lbs * WHEEL_TO_SLICE_FACTOR)
                actual_gain = qty_mon - qty_fri + depletion_total
                cutting_disc = round(actual_gain - expected_cutting)
        pct = round((diff / expected * 100), 1) if expected > 0 else (
            100.0 if qty_mon > 0 else 0.0)
        flagged = abs_diff >= threshold_min and abs(pct) >= threshold_pct

        if flagged:
            flagged_count += 1
            total_discrepancy += abs_diff

        rows.append({
            "sku": sku,
            "friday": qty_fri,
            "monday": qty_mon,
            "invoice_yield": inv_yield,
            "sat_depletion": depletion_total,
            "expected": expected,
            "diff": diff,
            "pct": pct,
            "flagged": flagged,
            "status": "over" if diff > 0 else "under" if diff < 0 else "match",
            "wheel_fri": wheel_fri,
            "wheel_mon": wheel_mon,
            "expected_from_cutting": expected_cutting,
            "cutting_discrepancy": cutting_disc,
        })

    # Sort: flagged first (by abs diff desc), then others
    rows.sort(key=lambda r: (-int(r["flagged"]), -abs(r["diff"])))

    return jsonify({
        "ok": True,
        "monday": {
            "id": monday_snap["id"],
            "label": monday_snap.get("label", ""),
            "timestamp": monday_snap.get("timestamp", ""),
            "cycle_day": monday_snap.get("cycle_day", ""),
        },
        "friday": {
            "id": friday_snap["id"],
            "label": friday_snap.get("label", ""),
            "timestamp": friday_snap.get("timestamp", ""),
            "cycle_day": friday_snap.get("cycle_day", ""),
        },
        "rows": rows,
        "summary": {
            "total_skus": len(rows),
            "flagged": flagged_count,
            "total_discrepancy": total_discrepancy,
            "invoices_in_window": len(invoice_entries),
            "depletions_in_window": len(depletion_entries),
        },
    })


@app.route("/api/reconcile_snapshots")
def reconcile_snapshots():
    """Return snapshots grouped by cycle day for the reconciliation picker."""
    s = _s()
    snapshots = s.get("inventory_snapshots", [])
    result = []
    for sn in reversed(snapshots):
        ts = sn.get("timestamp", "")
        result.append({
            "id": sn["id"],
            "label": sn.get("label", ""),
            "timestamp": ts,
            "cycle_day": sn.get("cycle_day", ""),
            "source": sn.get("source", ""),
            "sku_count": len(sn.get("inventory", {})),
        })
    return jsonify({"snapshots": result})


# ── Production Event Detection ─────────────────────────────────────────


def detect_production_events(old_snap, new_snap, depletion_between):
    """Detect wheel→slice production events between two snapshots.

    For each CH-* SKU with potential_yield data, checks if wheels were cut
    and calculates expected vs actual sliced gain.

    Returns list of production event dicts.
    """
    old_pot = old_snap.get("potential_yield", {})
    new_pot = new_snap.get("potential_yield", {})
    old_inv = old_snap.get("inventory", {})
    new_inv = new_snap.get("inventory", {})

    all_ch = set(k for k in (set(old_pot) | set(new_pot)) if k.startswith("CH-"))
    events = []

    for sku in sorted(all_ch):
        old_wheels = old_pot.get(sku, {}).get("wheels", 0)
        new_wheels = new_pot.get(sku, {}).get("wheels", 0)
        wheels_cut = max(0, old_wheels - new_wheels)
        if wheels_cut > 0:
            weight_lbs = old_pot.get(sku, new_pot.get(sku, {})).get("weight_lbs", 0)
            expected_sliced_gain = round(wheels_cut * weight_lbs * WHEEL_TO_SLICE_FACTOR)
            old_sliced = old_inv.get(sku, 0)
            new_sliced = new_inv.get(sku, 0)
            actual_sliced_gain = new_sliced - old_sliced + depletion_between.get(sku, 0)
            variance = round(actual_sliced_gain - expected_sliced_gain)
            events.append({
                "sku": sku,
                "wheels_cut": wheels_cut,
                "weight_lbs": weight_lbs,
                "expected": expected_sliced_gain,
                "actual": round(actual_sliced_gain),
                "variance": variance,
            })

    return events


# ── Running Inventory (Journal-based) ──────────────────────────────────


def compute_running_inventory():
    """Compute current running inventory by replaying journal on last snapshot.

    Returns: {"sliced": {sku: qty}, "wheels": {sku: qty}}
    """
    s = _s()
    journal = s.get("inventory_journal", [])
    snapshots = s.get("inventory_snapshots", [])

    # Index snapshots by ID for O(1) lookup
    snap_by_id = {sn["id"]: sn for sn in snapshots}

    def _load_snap(snap_id):
        sn = snap_by_id.get(snap_id)
        if not sn:
            return {}, {}
        sl = dict(sn.get("inventory", {}))
        wh = {sku: pot.get("wheels", 0) for sku, pot in sn.get("potential_yield", {}).items()}
        return sl, wh

    # Find the most recent snapshot entry in journal
    last_snap_idx = -1
    for i, entry in enumerate(journal):
        if entry.get("type") == "snapshot":
            last_snap_idx = i

    sliced, wheels = {}, {}
    if last_snap_idx >= 0:
        sliced, wheels = _load_snap(journal[last_snap_idx].get("snapshot_id", ""))

    # Replay journal entries after the last snapshot
    for entry in journal[last_snap_idx + 1:]:
        etype = entry.get("type", "")
        if etype == "snapshot":
            sliced, wheels = _load_snap(entry.get("snapshot_id", ""))
        elif etype in ("depletion", "adjustment"):
            for sku, delta in entry.get("sku_deltas", {}).items():
                sliced[sku] = sliced.get(sku, 0) + int(delta)
            for sku, delta in entry.get("wheel_deltas", {}).items():
                wheels[sku] = wheels.get(sku, 0) + int(delta)
        elif etype == "production":
            sku = entry.get("sku", "")
            if sku:
                wheels[sku] = wheels.get(sku, 0) - entry.get("wheels_cut", 0)
                sliced[sku] = sliced.get(sku, 0) + entry.get("actual_sliced", 0)

    return {"sliced": sliced, "wheels": wheels}


@app.route("/api/running_inventory")
def running_inventory():
    """Return computed running inventory from journal replay."""
    result = compute_running_inventory()
    return jsonify({"ok": True, **result})


@app.route("/api/journal")
def get_journal():
    """Return journal entries since last Friday snapshot."""
    s = _s()
    journal = s.get("inventory_journal", [])

    # Find last Friday snapshot entry
    start_idx = 0
    for i, entry in enumerate(journal):
        if entry.get("type") == "snapshot":
            # Check if the snapshot was on a Friday
            ts = entry.get("ts", "")
            try:
                dt = datetime.datetime.fromisoformat(ts)
                if dt.weekday() == 4:  # Friday
                    start_idx = i
            except (ValueError, TypeError):
                pass

    return jsonify({"ok": True, "entries": journal[start_idx:]})


@app.route("/api/journal_entry", methods=["POST"])
def add_journal_entry():
    """Add a manual adjustment journal entry.

    Body: {label, sku_deltas: {sku: delta}, wheel_deltas?: {sku: delta}}
    """
    data = request.json or {}
    label = data.get("label", "Manual adjustment")
    sku_deltas = data.get("sku_deltas", {})
    wheel_deltas = data.get("wheel_deltas", {})

    if not sku_deltas and not wheel_deltas:
        return jsonify({"error": "No deltas provided"}), 400

    s = _s()
    journal = s.setdefault("inventory_journal", [])
    entry = {
        "ts": datetime.datetime.now().isoformat(),
        "type": "adjustment",
        "label": label,
        "sku_deltas": {k: int(v) for k, v in sku_deltas.items()},
    }
    if wheel_deltas:
        entry["wheel_deltas"] = {k: int(v) for k, v in wheel_deltas.items()}
    journal.append(entry)
    save_settings(s)

    return jsonify({"ok": True, "entry": entry})


# ── Settings Configuration UI ────────────────────────────────────────

@app.route("/api/settings_config")
def get_settings_config():
    """Return editable settings for the Settings UI."""
    s = _s()
    return jsonify({
        "vendor_catalog": s.get("vendor_catalog", {}),
        "bulk_conversions": s.get("bulk_conversions", {}),
        "reorder_points": s.get("reorder_points", {}),
        "smtp_host": s.get("smtp_host", "smtp.gmail.com"),
        "smtp_port": s.get("smtp_port", "587"),
        "smtp_user": s.get("smtp_user", ""),
        "smtp_password": s.get("smtp_password", ""),
        "depletion_email_to": s.get("depletion_email_to", ""),
        "depletion_email_from": s.get("depletion_email_from", ""),
        "auto_refresh_interval": s.get("auto_refresh_interval", 60),
        "yield_reconciliation_threshold_pct": s.get("yield_reconciliation_threshold_pct", 5),
        "yield_reconciliation_threshold_min": s.get("yield_reconciliation_threshold_min", 2),
        "expiration_warning_days": s.get("expiration_warning_days", "14"),
        "fulfillment_buffer": s.get("fulfillment_buffer", "10"),
    })


@app.route("/api/settings_config", methods=["POST"])
def update_settings_config():
    """Update specific settings from the Settings UI. Body: {key: value, ...}"""
    data = request.json or {}
    if not data:
        return jsonify({"error": "No settings provided"}), 400

    s = _s()
    allowed = {
        "vendor_catalog", "bulk_conversions", "reorder_points",
        "smtp_host", "smtp_port", "smtp_user", "smtp_password",
        "depletion_email_to", "depletion_email_from",
        "auto_refresh_interval", "yield_reconciliation_threshold_pct",
        "yield_reconciliation_threshold_min", "expiration_warning_days",
        "fulfillment_buffer",
    }
    updated = []
    for key, value in data.items():
        if key in allowed:
            s[key] = value
            updated.append(key)

    if updated:
        save_settings(s)

    return jsonify({"ok": True, "updated": updated})


@app.route("/api/vendor_catalog", methods=["POST"])
def update_vendor_catalog():
    """Add or update a vendor catalog entry. Body: {sku, vendor, unit_cost, case_qty, moq, wheel_weight_lbs}"""
    data = request.json or {}
    sku = data.get("sku", "").upper()
    if not sku:
        return jsonify({"error": "SKU required"}), 400

    s = _s()
    catalog = s.setdefault("vendor_catalog", {})
    catalog[sku] = {
        "vendor": data.get("vendor", ""),
        "unit_cost": float(data.get("unit_cost", 0)),
        "case_qty": int(data.get("case_qty", 1)),
        "moq": int(data.get("moq", 0)),
        "wheel_weight_lbs": float(data.get("wheel_weight_lbs", 0)),
    }
    save_settings(s)
    return jsonify({"ok": True, "sku": sku})


@app.route("/api/vendor_catalog/<sku>", methods=["DELETE"])
def delete_vendor_catalog(sku):
    """Remove a vendor catalog entry."""
    s = _s()
    catalog = s.get("vendor_catalog", {})
    if sku.upper() in catalog:
        del catalog[sku.upper()]
        save_settings(s)
    return jsonify({"ok": True})


# ── Undo Depletion / Audit Trail ─────────────────────────────────────

@app.route("/api/undo_depletion", methods=["POST"])
def undo_depletion():
    """Undo the last depletion by restoring the pre-depletion snapshot."""
    s = _s()
    snapshots = s.get("inventory_snapshots", [])
    dep_history = s.get("depletion_history", [])

    if not dep_history:
        return jsonify({"error": "No depletion history to undo"}), 400

    last_dep = dep_history[-1]

    # Find the pre-depletion snapshot
    pre_snap = None
    for sn in reversed(snapshots):
        label = sn.get("label", "")
        if label.startswith("Pre-depletion") and sn.get("source") == "depletion":
            pre_snap = sn
            break

    if not pre_snap:
        return jsonify({"error": "Pre-depletion snapshot not found"}), 400

    # Restore inventory from pre-depletion snapshot
    restored_inv = pre_snap.get("inventory", {})
    STATE["rmfg_inventory"] = dict(restored_inv)

    # Remove the last depletion entry
    dep_history.pop()

    # Log the undo in audit trail
    audit = s.setdefault("audit_log", [])
    audit.append({
        "timestamp": datetime.datetime.now().isoformat(),
        "action": "undo_depletion",
        "detail": f"Restored inventory from: {pre_snap['label']}",
        "depletion_file": last_dep.get("file", ""),
        "units_restored": last_dep.get("total", 0),
    })
    save_settings(s)

    # Take a new snapshot
    _take_snapshot("Post-undo: restored from " + pre_snap["label"], source="undo")

    return jsonify({
        "ok": True,
        "restored_from": pre_snap["label"],
        "units_restored": last_dep.get("total", 0),
        "skus_restored": len(restored_inv),
    })


@app.route("/api/audit_log")
def get_audit_log():
    """Get the audit trail of important actions."""
    s = _s()
    # Build audit log from multiple sources
    log_entries = list(s.get("audit_log", []))

    # Include depletion history
    for dep in s.get("depletion_history", []):
        log_entries.append({
            "timestamp": dep.get("date", ""),
            "action": "depletion_applied",
            "detail": f"{dep.get('file', 'Unknown')}: {dep.get('total', 0)} units, {dep.get('skus', 0)} SKUs",
        })

    # Include snapshots
    for sn in s.get("inventory_snapshots", []):
        log_entries.append({
            "timestamp": sn.get("timestamp", ""),
            "action": "snapshot_" + sn.get("source", "manual"),
            "detail": sn.get("label", ""),
        })

    # Sort by timestamp descending
    log_entries.sort(key=lambda x: x.get("timestamp", ""), reverse=True)
    return jsonify({"entries": log_entries[:100]})


# ── Waste / Spoilage Ledger ──────────────────────────────────────────

@app.route("/api/waste", methods=["POST"])
def record_waste():
    """Record waste/spoilage. Body: {sku, qty, reason, date?}"""
    data = request.json or {}
    sku = data.get("sku", "").upper()
    qty = int(data.get("qty", 0))
    reason = data.get("reason", "spoilage")

    if not sku or qty <= 0:
        return jsonify({"error": "sku and qty (>0) required"}), 400

    s = _s()
    ledger = s.setdefault("waste_ledger", [])
    entry = {
        "id": datetime.datetime.now().strftime("%Y%m%d_%H%M%S"),
        "timestamp": datetime.datetime.now().isoformat(),
        "sku": sku,
        "qty": qty,
        "reason": reason,
        "date": data.get("date", datetime.date.today().isoformat()),
    }
    ledger.append(entry)

    # Also subtract from current inventory if loaded
    inv = STATE.get("rmfg_inventory", {})
    if sku in inv:
        inv[sku] = max(0, inv[sku] - qty)

    # Audit log
    audit = s.setdefault("audit_log", [])
    audit.append({
        "timestamp": entry["timestamp"],
        "action": "waste_recorded",
        "detail": f"{sku}: {qty} units ({reason})",
    })
    save_settings(s)

    return jsonify({"ok": True, "entry": entry})


@app.route("/api/waste")
def get_waste_ledger():
    """Get waste/spoilage history."""
    s = _s()
    ledger = s.get("waste_ledger", [])

    # Summary by SKU
    by_sku = defaultdict(lambda: {"total": 0, "entries": 0})
    for entry in ledger:
        by_sku[entry["sku"]]["total"] += entry["qty"]
        by_sku[entry["sku"]]["entries"] += 1

    # Summary by reason
    by_reason = defaultdict(int)
    for entry in ledger:
        by_reason[entry.get("reason", "spoilage")] += entry["qty"]

    return jsonify({
        "entries": ledger,
        "by_sku": dict(by_sku),
        "by_reason": dict(by_reason),
        "total_wasted": sum(e["qty"] for e in ledger),
    })


@app.route("/api/waste/<waste_id>", methods=["DELETE"])
def delete_waste(waste_id):
    """Delete a waste entry."""
    s = _s()
    ledger = s.get("waste_ledger", [])
    s["waste_ledger"] = [e for e in ledger if e.get("id") != waste_id]
    save_settings(s)
    return jsonify({"ok": True})


# ── Reorder Points / Wed PO Templates ────────────────────────────────

@app.route("/api/reorder_points")
def get_reorder_points():
    """Get per-SKU reorder points."""
    s = _s()
    return jsonify({"reorder_points": s.get("reorder_points", {})})


@app.route("/api/reorder_points", methods=["POST"])
def update_reorder_points():
    """Set reorder point for a SKU. Body: {sku, min_stock, preferred_qty, lead_days}"""
    data = request.json or {}
    sku = data.get("sku", "").upper()
    if not sku:
        return jsonify({"error": "SKU required"}), 400

    s = _s()
    rp = s.setdefault("reorder_points", {})
    rp[sku] = {
        "min_stock": int(data.get("min_stock", 0)),
        "preferred_qty": int(data.get("preferred_qty", 0)),
        "lead_days": int(data.get("lead_days", 7)),
    }
    save_settings(s)
    return jsonify({"ok": True, "sku": sku})


@app.route("/api/wed_po_draft")
def wed_po_draft():
    """Generate a draft Wednesday PO based on reorder points + current inventory.
    Compares current inventory vs min_stock and generates order lines for anything below.
    """
    s = _s()
    inv = STATE.get("rmfg_inventory", {})
    rp = s.get("reorder_points", {})
    catalog = s.get("vendor_catalog", {})

    if not rp:
        return jsonify({"lines": [], "message": "No reorder points configured. Set them in Settings."})

    lines = []
    for sku, points in sorted(rp.items()):
        current = inv.get(sku, 0)
        min_stock = points.get("min_stock", 0)
        preferred_qty = points.get("preferred_qty", 0)
        lead_days = points.get("lead_days", 7)

        if current < min_stock:
            deficit = min_stock - current
            order_qty = max(deficit, preferred_qty) if preferred_qty > 0 else deficit

            # Check vendor catalog for case qty rounding
            vc = catalog.get(sku, {})
            case_qty = vc.get("case_qty", 1)
            if case_qty > 1:
                cases = math.ceil(order_qty / case_qty)
                order_qty = cases * case_qty
            else:
                cases = order_qty

            lines.append({
                "sku": sku,
                "current": current,
                "min_stock": min_stock,
                "deficit": deficit,
                "order_qty": order_qty,
                "cases": cases,
                "case_qty": case_qty,
                "vendor": vc.get("vendor", ""),
                "unit_cost": vc.get("unit_cost", 0),
                "line_cost": round(vc.get("unit_cost", 0) * order_qty, 2),
                "lead_days": lead_days,
            })

    total_cost = sum(l["line_cost"] for l in lines)
    return jsonify({
        "lines": lines,
        "total_lines": len(lines),
        "total_cost": round(total_cost, 2),
    })


# ── Email Wednesday PO ───────────────────────────────────────────────

@app.route("/api/email_po", methods=["POST"])
def email_po():
    """Email the Wednesday PO / order list via SMTP.
    Body: {to?, subject?, lines: [{sku, order_qty, vendor, ...}]}
    """
    import smtplib
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart

    data = request.json or {}
    lines = data.get("lines", [])
    if not lines:
        return jsonify({"error": "No order lines to email"}), 400

    s = _s()
    smtp_host = s.get("smtp_host", "smtp.gmail.com")
    smtp_port = int(s.get("smtp_port", "587"))
    smtp_user = s.get("smtp_user", "")
    smtp_pass = s.get("smtp_password", "")
    from_addr = data.get("from", s.get("depletion_email_from", smtp_user))
    to_addr = data.get("to", s.get("depletion_email_to", ""))
    subject = data.get("subject", f"Wednesday PO - {datetime.date.today().isoformat()}")

    if not smtp_user or not smtp_pass:
        return jsonify({"error": "SMTP credentials not configured. Set them in Settings."}), 400
    if not to_addr:
        return jsonify({"error": "No recipient email. Set depletion_email_to in Settings."}), 400

    # Build email body
    body = f"Production Order — {datetime.date.today().strftime('%B %d, %Y')}\n"
    body += "=" * 60 + "\n\n"

    # Group by vendor
    by_vendor = defaultdict(list)
    for l in lines:
        by_vendor[l.get("vendor", "Unknown")].append(l)

    total_cost = 0
    for vendor, vlines in sorted(by_vendor.items()):
        body += f"VENDOR: {vendor}\n"
        body += "-" * 40 + "\n"
        for l in vlines:
            cost = l.get("line_cost", 0)
            total_cost += cost
            body += f"  {l['sku']}: {l['order_qty']} units"
            if l.get("cases") and l.get("case_qty", 1) > 1:
                body += f" ({l['cases']} cases x {l['case_qty']})"
            if cost > 0:
                body += f" — ${cost:.2f}"
            body += "\n"
        body += "\n"

    body += f"TOTAL: {len(lines)} lines, ${total_cost:.2f}\n"

    try:
        msg = MIMEMultipart()
        msg["From"] = from_addr
        msg["To"] = to_addr
        msg["Subject"] = subject
        msg.attach(MIMEText(body, "plain"))

        with smtplib.SMTP(smtp_host, smtp_port, timeout=15) as server:
            server.starttls()
            server.login(smtp_user, smtp_pass)
            server.send_message(msg)

        # Audit log
        audit = s.setdefault("audit_log", [])
        audit.append({
            "timestamp": datetime.datetime.now().isoformat(),
            "action": "po_emailed",
            "detail": f"Sent to {to_addr}: {len(lines)} lines, ${total_cost:.2f}",
        })
        save_settings(s)

        return jsonify({"ok": True, "sent_to": to_addr, "lines": len(lines)})
    except Exception as e:
        return jsonify({"error": f"Email send failed: {str(e)}"}), 500


# ── Cut Order PDF + Email ─────────────────────────────────────────────

def generate_cut_order_pdf(cut_lines, summary):
    """Generate a simple PDF for the cut order. Returns BytesIO with PDF bytes."""
    from fpdf import FPDF

    pdf = FPDF()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)

    # Header
    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 10, f"Cut Order - {datetime.date.today().strftime('%B %d, %Y')}", ln=True, align="C")
    pdf.ln(4)

    # Summary line
    pdf.set_font("Helvetica", "", 10)
    total_demand = summary.get("total_demand", 0)
    total_wheels = summary.get("total_wheels_to_cut", 0)
    total_pcs = summary.get("total_pcs_from_cut", 0)
    shortage_count = summary.get("shortage_count", 0)
    pdf.cell(0, 6, f"Total Demand: {total_demand}  |  Wheels to Cut: {total_wheels}  |  Pcs from Cut: {total_pcs}  |  Shortages: {shortage_count}", ln=True, align="C")
    pdf.ln(6)

    # Table header
    col_widths = [30, 60, 20, 22, 22, 30]
    headers = ["SKU", "Product Name", "Order Qty", "Wheels", "Pcs", "Status"]
    pdf.set_font("Helvetica", "B", 9)
    pdf.set_fill_color(40, 40, 50)
    pdf.set_text_color(255, 255, 255)
    for i, h in enumerate(headers):
        pdf.cell(col_widths[i], 7, h, border=1, fill=True, align="C")
    pdf.ln()

    # Sort lines: SHORTAGE first, then MFG, OK, etc.
    status_order = {"SHORTAGE": 0, "MFG": 1, "OK": 2, "SURPLUS": 3, "NO DEMAND": 4}
    sorted_lines = sorted(cut_lines, key=lambda x: (status_order.get(x.get("status", "OK"), 9), x.get("sku", "")))

    pdf.set_text_color(0, 0, 0)
    pdf.set_font("Helvetica", "", 8)
    for line in sorted_lines:
        if line.get("total_demand", 0) == 0 and line.get("sliced", 0) == 0:
            continue
        sku = str(line.get("sku", ""))[:12]
        name = str(line.get("product_name", line.get("sku", "")))[:30]
        order_qty = str(line.get("total_demand", 0))
        wheels = str(line.get("wheels_to_cut", 0))
        pcs = str(line.get("pcs_from_cut", 0))
        status = str(line.get("status", ""))

        # Color coding for status
        if status == "SHORTAGE":
            pdf.set_fill_color(255, 220, 220)
        elif status == "MFG":
            pdf.set_fill_color(220, 240, 255)
        else:
            pdf.set_fill_color(255, 255, 255)

        fill = status in ("SHORTAGE", "MFG")
        pdf.cell(col_widths[0], 6, sku, border=1, fill=fill)
        pdf.cell(col_widths[1], 6, name, border=1, fill=fill)
        pdf.cell(col_widths[2], 6, order_qty, border=1, align="R", fill=fill)
        pdf.cell(col_widths[3], 6, wheels, border=1, align="R", fill=fill)
        pdf.cell(col_widths[4], 6, pcs, border=1, align="R", fill=fill)
        pdf.cell(col_widths[5], 6, status, border=1, align="C", fill=fill)
        pdf.ln()

    # Footer
    pdf.ln(6)
    pdf.set_font("Helvetica", "B", 9)
    pdf.cell(0, 6, f"Total: {total_demand} demand  |  {total_wheels} wheels to cut  |  {total_pcs} pcs from cut", ln=True, align="C")

    buf = io.BytesIO()
    pdf.output(buf)
    buf.seek(0)
    return buf


@app.route("/api/email_cut_order", methods=["POST"])
def email_cut_order():
    """Email cut order as PDF attachment via SMTP."""
    import smtplib
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart
    from email.mime.base import MIMEBase
    from email import encoders

    data = request.json or {}
    lines = data.get("lines", [])
    summary = data.get("summary", {})
    if not lines:
        return jsonify({"error": "No cut order lines to email"}), 400

    s = _s()
    smtp_host = s.get("smtp_host", "smtp.gmail.com")
    smtp_port = int(s.get("smtp_port", "587"))
    smtp_user = s.get("smtp_user", "")
    smtp_pass = s.get("smtp_password", "")
    from_addr = data.get("from", s.get("depletion_email_from", smtp_user))
    to_addr = data.get("to", s.get("depletion_email_to", ""))
    today = datetime.date.today().isoformat()
    subject = f"Cut Order - {today}"

    if not smtp_user or not smtp_pass:
        return jsonify({"error": "SMTP credentials not configured. Set them in Settings."}), 400
    if not to_addr:
        return jsonify({"error": "No recipient email. Set depletion_email_to in Settings."}), 400

    try:
        # Generate PDF
        pdf_buf = generate_cut_order_pdf(lines, summary)

        # Build email
        msg = MIMEMultipart()
        msg["From"] = from_addr
        msg["To"] = to_addr
        msg["Subject"] = subject

        # Text body
        body_text = f"Cut Order - {datetime.date.today().strftime('%B %d, %Y')}\n"
        body_text += f"Total demand: {summary.get('total_demand', 0)}\n"
        body_text += f"Wheels to cut: {summary.get('total_wheels_to_cut', 0)}\n"
        body_text += f"Shortages: {summary.get('shortage_count', 0)}\n\n"
        body_text += "PDF attached."
        msg.attach(MIMEText(body_text, "plain"))

        # PDF attachment
        part = MIMEBase("application", "octet-stream")
        part.set_payload(pdf_buf.read())
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", f"attachment; filename=cut_order_{today}.pdf")
        msg.attach(part)

        with smtplib.SMTP(smtp_host, smtp_port, timeout=15) as server:
            server.starttls()
            server.login(smtp_user, smtp_pass)
            server.send_message(msg)

        # Audit log
        audit = s.setdefault("audit_log", [])
        audit.append({
            "timestamp": datetime.datetime.now().isoformat(),
            "action": "cut_order_emailed",
            "detail": f"Sent to {to_addr}: {len(lines)} lines",
        })
        save_settings(s)

        return jsonify({"ok": True, "sent_to": to_addr, "lines": len(lines)})
    except Exception as e:
        return jsonify({"error": f"Email send failed: {str(e)}"}), 500


@app.route("/api/schedule_cut_order_email", methods=["POST"])
def schedule_cut_order_email():
    """Store cut order email schedule config in settings."""
    data = request.json or {}
    s = _s()
    s["cut_order_email_schedule"] = {
        "enabled": bool(data.get("enabled", True)),
        "day": data.get("day", "wednesday"),
        "time": data.get("time", "10:00"),
    }
    save_settings(s)
    return jsonify({"ok": True, "schedule": s["cut_order_email_schedule"]})


# ── SKU History for Sparklines ────────────────────────────────────────

@app.route("/api/sku_history")
def sku_history():
    """Get historical inventory levels per CH- SKU from snapshots, for sparklines.
    Returns last N data points per SKU from saved snapshots.
    """
    s = _s()
    snapshots = s.get("inventory_snapshots", [])

    # Use up to last 12 snapshots (roughly 6 weeks of Fri+Mon pairs)
    recent = snapshots[-12:] if len(snapshots) > 12 else snapshots

    # Build per-SKU time series
    history = defaultdict(list)
    timestamps = []
    for sn in recent:
        inv = sn.get("inventory", {})
        ts = sn.get("timestamp", "")[:10]
        timestamps.append(ts)
        for sku, qty in inv.items():
            if sku.startswith("CH-"):
                history[sku].append(qty)
        # Backfill zeros for SKUs not in this snapshot
        for sku in history:
            if len(history[sku]) < len(timestamps):
                history[sku].append(0)

    return jsonify({
        "timestamps": timestamps,
        "history": dict(history),
        "count": len(timestamps),
    })


# ── Supplier Lead Time Tracking ──────────────────────────────────────

@app.route("/api/po_received", methods=["POST"])
def mark_po_received():
    """Mark an open PO as received. Body: {po_index, received_date?, actual_qty?}
    Records the actual lead time for tracking.
    """
    data = request.json or {}
    po_index = data.get("po_index")
    received_date = data.get("received_date", datetime.date.today().isoformat())
    actual_qty = data.get("actual_qty")

    s = _s()
    open_pos = s.get("open_pos", [])
    if po_index is None or po_index >= len(open_pos):
        return jsonify({"error": "Invalid PO index"}), 400

    po = open_pos[po_index]
    po["status"] = "received"
    po["received_date"] = received_date

    if actual_qty is not None:
        po["actual_qty"] = int(actual_qty)

    # Calculate actual lead time if we have ETA
    eta = po.get("eta", "")
    placed = po.get("placed_date", "")
    actual_lead = None
    if placed:
        try:
            placed_dt = datetime.date.fromisoformat(placed)
            received_dt = datetime.date.fromisoformat(received_date)
            actual_lead = (received_dt - placed_dt).days
        except (ValueError, TypeError):
            pass

    # Record in lead time history
    lt_history = s.setdefault("lead_time_history", [])
    lt_history.append({
        "sku": po.get("sku", ""),
        "vendor": po.get("vendor", ""),
        "placed_date": placed,
        "eta": eta,
        "received_date": received_date,
        "expected_lead_days": None,
        "actual_lead_days": actual_lead,
        "qty": po.get("qty", 0),
        "actual_qty": actual_qty or po.get("qty", 0),
    })

    # Audit
    audit = s.setdefault("audit_log", [])
    audit.append({
        "timestamp": datetime.datetime.now().isoformat(),
        "action": "po_received",
        "detail": f"{po.get('sku', '?')}: {po.get('qty', 0)} units from {po.get('vendor', '?')}"
                  + (f", {actual_lead}d lead time" if actual_lead else ""),
    })

    save_settings(s)

    return jsonify({
        "ok": True,
        "sku": po.get("sku", ""),
        "actual_lead_days": actual_lead,
    })


@app.route("/api/lead_times")
def get_lead_times():
    """Get supplier lead time statistics."""
    s = _s()
    history = s.get("lead_time_history", [])

    if not history:
        return jsonify({"has_data": False, "entries": [], "by_vendor": {}, "by_sku": {}})

    # Stats by vendor
    by_vendor = defaultdict(list)
    for h in history:
        if h.get("actual_lead_days") is not None:
            by_vendor[h.get("vendor", "Unknown")].append(h["actual_lead_days"])

    vendor_stats = {}
    for vendor, days in by_vendor.items():
        vendor_stats[vendor] = {
            "avg_days": round(sum(days) / len(days), 1),
            "min_days": min(days),
            "max_days": max(days),
            "count": len(days),
        }

    # Stats by SKU
    by_sku = defaultdict(list)
    for h in history:
        if h.get("actual_lead_days") is not None:
            by_sku[h["sku"]].append(h["actual_lead_days"])

    sku_stats = {}
    for sku, days in by_sku.items():
        sku_stats[sku] = {
            "avg_days": round(sum(days) / len(days), 1),
            "count": len(days),
        }

    return jsonify({
        "has_data": True,
        "entries": history[-20:],
        "by_vendor": vendor_stats,
        "by_sku": sku_stats,
    })


# ── Morning Briefing ─────────────────────────────────────────────────

@app.route("/api/briefing")
def morning_briefing():
    """Generate a morning briefing summary: shortages, expiring items, coverage gaps, cycle context."""
    s = _s()
    inv = STATE.get("rmfg_inventory", {})
    sat_demand = STATE.get("rmfg_sat_demand", {})
    tue_demand = STATE.get("rmfg_tue_demand", {})
    next_sat_demand = STATE.get("rmfg_next_sat_demand", {})
    bulk_weights = STATE.get("bulk_weights", {})

    today = datetime.date.today()
    weekday = today.weekday()  # 0=Mon
    cycle_day = {0: "monday", 1: "tuesday", 2: "wednesday", 4: "friday", 5: "saturday"}.get(weekday, "midweek")

    # Day-specific action hints
    day_actions = {
        "friday": "Planning baseline day. Review NET positions and finalize assignments for Saturday fulfillment.",
        "saturday": "Main fulfillment day. Upload depletion file after shipment completes.",
        "monday": "Reconciliation day. Compare Friday baseline vs today's actual inventory.",
        "tuesday": "Second fulfillment day. Upload depletion file after shipment.",
        "wednesday": "Production order day. Generate Wed PO and submit orders to RMFG.",
        "midweek": "Mid-cycle. Review inventory and upcoming demand.",
    }

    # Shortages
    shortages = []
    tight = []
    surplus = []
    for sku in sorted(set(list(inv.keys()) + list(sat_demand.keys()))):
        if not sku.startswith("CH-"):
            continue
        avail = inv.get(sku, 0)
        d_sat = int(round(sat_demand.get(sku, 0)))
        d_tue = int(round(tue_demand.get(sku, 0)))
        d_next = int(round(next_sat_demand.get(sku, 0)))
        net = avail - d_sat
        bw = bulk_weights.get(sku, {})
        potential = bw.get("potential_yield", 0)

        if d_sat == 0 and d_tue == 0:
            continue
        if net < 0:
            shortages.append({"sku": sku, "available": avail, "demand": d_sat, "deficit": abs(net),
                              "potential": potential, "coverable": potential >= abs(net)})
        elif net < d_sat * 0.2:
            tight.append({"sku": sku, "available": avail, "demand": d_sat, "net": net})
        elif net > avail * 0.5 and avail > 200:
            surplus.append({"sku": sku, "available": avail, "demand": d_sat, "net": net})

    # Tuesday coverage check
    tue_gaps = []
    for sku in sorted(sat_demand.keys()):
        if not sku.startswith("CH-"):
            continue
        carry = max(0, inv.get(sku, 0) - int(round(sat_demand.get(sku, 0))))
        td = int(round(tue_demand.get(sku, 0)))
        if td > 0 and carry < td:
            tue_gaps.append({"sku": sku, "carry": carry, "tue_demand": td, "gap": td - carry})

    # Shelf life / expiring items
    inventory_full = s.get("inventory", {})
    expiring = []
    for sku, data in inventory_full.items():
        if not isinstance(data, dict):
            continue
        dates = data.get("expiration_dates", [])
        for d in dates:
            try:
                exp = datetime.date.fromisoformat(d) if isinstance(d, str) else d
                days_left = (exp - today).days
                if 0 <= days_left <= 7:
                    expiring.append({"sku": sku, "days_left": days_left, "date": str(exp)})
            except (ValueError, TypeError):
                pass
    expiring.sort(key=lambda x: x["days_left"])

    # Last sync info
    snapshots = s.get("inventory_snapshots", [])
    last_snapshot = snapshots[-1] if snapshots else None
    last_depletion = None
    dep_history = s.get("depletion_history", [])
    if dep_history:
        last_depletion = dep_history[-1]

    # Forecast accuracy (if we have data)
    accuracy = s.get("forecast_accuracy", [])
    recent_accuracy = accuracy[-4:] if accuracy else []

    return jsonify({
        "cycle_day": cycle_day,
        "date": today.isoformat(),
        "weekday": today.strftime("%A"),
        "action_hint": day_actions.get(cycle_day, ""),
        "shortages": shortages,
        "shortage_count": len(shortages),
        "tight": tight,
        "tight_count": len(tight),
        "surplus_count": len(surplus),
        "tue_gaps": tue_gaps,
        "expiring": expiring,
        "expiring_count": len(expiring),
        "last_snapshot": {
            "label": last_snapshot.get("label", ""),
            "timestamp": last_snapshot.get("timestamp", ""),
            "cycle_day": last_snapshot.get("cycle_day", ""),
        } if last_snapshot else None,
        "last_depletion": {
            "date": last_depletion.get("date", ""),
            "file": last_depletion.get("file", ""),
            "total": last_depletion.get("total", 0),
        } if last_depletion else None,
        "recent_accuracy": recent_accuracy,
        "total_cheese_skus": len([k for k in inv if k.startswith("CH-")]),
        "total_cheese_units": sum(v for k, v in inv.items() if k.startswith("CH-")),
    })


# ── Forecast Accuracy ────────────────────────────────────────────────

@app.route("/api/forecast_accuracy", methods=["POST"])
def record_forecast_accuracy():
    """Record actual depletion vs predicted demand for accuracy tracking.

    Called after depletion is applied. Compares pre-depletion demand predictions
    with actual depletion quantities.

    Body: {depletion_skus: {sku: actual_qty}, window: "saturday"|"tuesday", date: "YYYY-MM-DD"}
    """
    data = request.json or {}
    depletion_skus = data.get("depletion_skus", {})
    window = data.get("window", "saturday")
    record_date = data.get("date", datetime.date.today().isoformat())

    if not depletion_skus:
        return jsonify({"error": "No depletion data provided"}), 400

    # Get the demand prediction that was active at the time
    if window == "saturday":
        predicted = dict(STATE.get("rmfg_sat_demand", {}))
    else:
        predicted = dict(STATE.get("rmfg_tue_demand", {}))

    # Calculate accuracy per SKU
    sku_accuracy = []
    total_predicted = 0
    total_actual = 0
    for sku in sorted(set(list(depletion_skus.keys()) + list(predicted.keys()))):
        if not sku.startswith("CH-"):
            continue
        actual = int(depletion_skus.get(sku, 0))
        pred = int(round(predicted.get(sku, 0)))
        if pred == 0 and actual == 0:
            continue
        error = actual - pred
        pct_error = round((error / pred) * 100, 1) if pred > 0 else None
        sku_accuracy.append({
            "sku": sku, "predicted": pred, "actual": actual,
            "error": error, "pct_error": pct_error,
        })
        total_predicted += pred
        total_actual += actual

    # Overall accuracy metrics
    mape_values = [abs(s["pct_error"]) for s in sku_accuracy if s["pct_error"] is not None]
    mape = round(sum(mape_values) / len(mape_values), 1) if mape_values else None
    overall_pct = round(((total_actual - total_predicted) / total_predicted) * 100, 1) if total_predicted > 0 else None

    record = {
        "date": record_date,
        "window": window,
        "total_predicted": total_predicted,
        "total_actual": total_actual,
        "overall_pct_error": overall_pct,
        "mape": mape,
        "sku_count": len(sku_accuracy),
        "skus": sku_accuracy,
    }

    # Save to settings
    s = _s()
    history = s.setdefault("forecast_accuracy", [])
    history.append(record)
    # Keep last 52 entries (1 year of weekly records)
    if len(history) > 52:
        s["forecast_accuracy"] = history[-52:]
    save_settings(s)

    return jsonify({"ok": True, "record": record})


@app.route("/api/forecast_accuracy")
def get_forecast_accuracy():
    """Get forecast accuracy history."""
    s = _s()
    history = s.get("forecast_accuracy", [])
    return jsonify({"history": history})


@app.route("/api/forecast_accuracy/summary")
def forecast_accuracy_summary():
    """Get summary stats for forecast accuracy."""
    s = _s()
    history = s.get("forecast_accuracy", [])
    if not history:
        return jsonify({"has_data": False})

    recent = history[-8:]  # Last 8 records
    mapes = [r["mape"] for r in recent if r.get("mape") is not None]
    overall_errors = [r["overall_pct_error"] for r in recent if r.get("overall_pct_error") is not None]

    # Per-SKU trends: find consistently over/under predicted SKUs
    sku_errors = defaultdict(list)
    for r in history[-12:]:
        for sr in r.get("skus", []):
            if sr.get("pct_error") is not None:
                sku_errors[sr["sku"]].append(sr["pct_error"])

    biased_skus = []
    for sku, errors in sku_errors.items():
        if len(errors) >= 3:
            avg = sum(errors) / len(errors)
            if abs(avg) > 15:  # consistently >15% off
                biased_skus.append({
                    "sku": sku,
                    "avg_error_pct": round(avg, 1),
                    "direction": "over" if avg > 0 else "under",
                    "samples": len(errors),
                })
    biased_skus.sort(key=lambda x: abs(x["avg_error_pct"]), reverse=True)

    return jsonify({
        "has_data": True,
        "total_records": len(history),
        "recent_mape": round(sum(mapes) / len(mapes), 1) if mapes else None,
        "recent_overall_error": round(sum(overall_errors) / len(overall_errors), 1) if overall_errors else None,
        "trend": [{"date": r["date"], "window": r["window"], "mape": r.get("mape"),
                   "overall_pct_error": r.get("overall_pct_error"),
                   "total_predicted": r["total_predicted"], "total_actual": r["total_actual"]}
                  for r in recent],
        "biased_skus": biased_skus[:10],
    })


# ── Launch ──────────────────────────────────────────────────────────────

def run_webview():
    """Launch in a native window via pywebview."""
    import webview
    STATE["saved"] = load_settings()

    server = threading.Thread(
        target=lambda: app.run(port=5187, debug=False, use_reloader=False),
        daemon=True)
    server.start()

    import time
    time.sleep(0.5)

    webview.create_window(
        "Fulfillment Planner",
        "http://127.0.0.1:5187",
        width=1400, height=900, min_size=(1000, 700))
    webview.start()


def run_browser():
    """Launch in default browser."""
    STATE["saved"] = load_settings()
    import webbrowser
    webbrowser.open("http://127.0.0.1:5187")
    app.run(port=5187, debug=False)


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--browser", action="store_true",
                        help="Open in browser instead of native window")
    args = parser.parse_args()

    STATE["saved"] = load_settings()

    if args.browser:
        run_browser()
    else:
        try:
            run_webview()
        except ImportError:
            print("pywebview not installed, falling back to browser")
            run_browser()
