"""Set the weekly HAVE inventory for the cut order (the one manual input).

The cut order's "Avail" column is NOT a live feed — it comes from
`corrected_inventory_path` in settings. This helper takes the raw RMFG HAVE
export and:
  1. Writes Downloads/Corrected_Inventory_<DATE>.xlsx in the format the
     generator reads (col A = SKU, col G = Qty).
  2. Repoints `corrected_inventory_path` in BOTH settings copies
     (repo dist/ + %APPDATA%/AppyHour/).

Run this BEFORE build_cut_order_xlsx_v2.py every week. If you skip it, the
generator silently reuses whatever stale Corrected_Inventory_*.xlsx is still
pointed at (burned us 2026-06-23: a 6/16 file clobbered the 6/23 HAVE).

The HAVE export header is often mislabeled "Delta" — that's the post-Saturday
depletion value = the upcoming Monday's on-hand. It IS on-hand; use it as Avail.

Usage:
    python scripts/set_corrected_inventory.py "<HAVE export.csv|.xlsx>" <MM-DD-YY>

    col A = SKU, col B = on-hand count (other layouts: pass --sku-col / --qty-col,
    1-based). Blank/empty qty rows are skipped.
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import sys

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.dirname(HERE)  # InventoryReorder/


def _settings_paths() -> list[str]:
    paths = [os.path.join(REPO, "dist", "inventory_reorder_settings.json")]
    appdata = os.environ.get("APPDATA")
    if appdata:
        paths.append(os.path.join(appdata, "AppyHour", "inventory_reorder_settings.json"))
    return [p for p in paths if os.path.exists(p)]


def _read_have(path: str, sku_col: int, qty_col: int) -> list[tuple[str, int]]:
    """Return [(sku, qty), ...] from a CSV or XLSX HAVE export. 1-based col indexes."""
    si, qi = sku_col - 1, qty_col - 1
    rows: list[tuple[str, int]] = []
    skipped: list[str] = []      # SKUs with a blank/unparseable qty — NAMED, not counted
    ext = os.path.splitext(path)[1].lower()

    def _emit(sku_raw, qty_raw):
        sku = str(sku_raw or "").strip()
        qraw = str(qty_raw if qty_raw is not None else "").strip()
        if not sku:
            return
        if qraw == "":
            skipped.append(sku)
            return
        try:
            rows.append((sku, int(float(qraw))))
        except ValueError:
            skipped.append(f"{sku}={qraw!r}")

    if ext in (".xlsx", ".xlsm"):
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        for i, r in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:  # header
                continue
            _emit(r[si] if len(r) > si else "", r[qi] if len(r) > qi else "")
        wb.close()
    else:
        with open(path, newline="", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            next(reader, None)  # header
            for r in reader:
                if not r:
                    continue
                _emit(r[si] if len(r) > si else "", r[qi] if len(r) > qi else "")

    print(f"  Read {len(rows)} SKUs from {path}")

    # A blank qty is NOT zero — it means "not counted". The SKU is absent from the
    # corrected file, so the HAVE-authoritative rule zeroes it and any demand shows
    # as a full shortage. Name them so a phantom shortfall is never a surprise
    # (PK-TCUST shipped blank two weeks running against ~630 demand).
    if skipped:
        print(f"  !! {len(skipped)} SKU(s) have a BLANK/unparseable qty and are NOT in the "
              f"corrected file -> they will read as Avail 0, not 'uncounted':")
        print(f"     {', '.join(skipped)}")

    # Negative on-hand is an upstream count error (over-pick / mis-count). It flows
    # straight into Avail and inflates the shortfall, so never let it pass silently.
    negs = [(s, q) for s, q in rows if q < 0]
    if negs:
        print(f"  !! {len(negs)} SKU(s) have NEGATIVE on-hand — upstream count error, "
              f"inflates the cut: {', '.join(f'{s}={q}' for s, q in negs)}")

    return rows


def main() -> int:
    ap = argparse.ArgumentParser(description="Set weekly HAVE inventory for the cut order.")
    ap.add_argument("have_file", help="RMFG HAVE export (.csv or .xlsx)")
    ap.add_argument("date", help="Date tag for the output filename, e.g. 06-29 or 06-29-26")
    ap.add_argument("--sku-col", type=int, default=1, help="1-based SKU column (default 1=A)")
    ap.add_argument("--qty-col", type=int, default=2, help="1-based on-hand column (default 2=B)")
    args = ap.parse_args()

    if not os.path.exists(args.have_file):
        print(f"ERROR: HAVE file not found: {args.have_file}")
        return 1

    rows = _read_have(args.have_file, args.sku_col, args.qty_col)
    if not rows:
        print("ERROR: no rows parsed — check --sku-col/--qty-col.")
        return 1

    downloads = os.path.join(os.path.expanduser("~"), "Downloads")
    out = os.path.join(downloads, f"Corrected_Inventory_{args.date}.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["SKU", "", "", "", "", "", "Qty"])  # generator reads col A=SKU, col G=Qty
    for sku, qty in rows:
        ws.append([sku, "", "", "", "", "", qty])
    wb.save(out)
    print(f"  Wrote {out}")

    paths = _settings_paths()
    if not paths:
        print("ERROR: no settings file found to repoint.")
        return 1
    for p in paths:
        with open(p, encoding="utf-8") as f:
            d = json.load(f)
        d["corrected_inventory_path"] = out
        with open(p, "w", encoding="utf-8") as f:
            json.dump(d, f, indent=2)
        print(f"  Repointed corrected_inventory_path -> {p}")

    print("\nDone. Now run: python build_cut_order_xlsx_v2.py --local")
    return 0


if __name__ == "__main__":
    sys.exit(main())
