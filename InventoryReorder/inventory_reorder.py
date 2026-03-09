"""
Inventory Reorder System
Standalone tkinter app for forecasting demand, tracking inventory,
and generating reorder alerts using Safety Stock methodology.

Demand sources:
  1. Recharge subscriptions (API pull + churn rate)
  2. Shopify first-time orders (manual weekly forecast for bundles)
  3. Manual adjustment per SKU

Reorder Point = (Daily Usage x (Total Lead Time + Fulfillment Buffer)) + Safety Stock
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog
import json
import os
import sys
import csv
import threading
import datetime
import time
import math
import smtplib
import tempfile
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from collections import defaultdict
import uuid
from http.server import HTTPServer, BaseHTTPRequestHandler
from urllib.parse import urlparse, parse_qs

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ── version ──────────────────────────────────────────────────────────
APP_VERSION = "3.0.0"
SETTINGS_FILE = "inventory_reorder_settings.json"

# ── dark theme palette (matches GelPackCalculator) ───────────────────
_BG   = "#2b2b2b"
_BG2  = "#363636"
_BG3  = "#424242"
_FG   = "#f0f0f0"
_FG2  = "#aaaaaa"
_ACC  = "#4a7ec4"
_SEP  = "#555555"
_GREEN = "#1a6633"
_RED   = "#c0392b"
_ORANGE = "#d4650a"
_YELLOW = "#b8900a"
_BLUE   = "#3498db"

# Default bulk-to-packet conversion mappings
# {bulk_ingredient_keyword: {target_sku, packet_oz}}
DEFAULT_BULK_CONVERSIONS = {
    "Raw Hazelnuts":                    {"sku": "AC-RHAZ",  "packet_oz": 3.9},
    "Praline Pecans":                   {"sku": "AC-PRPE",  "packet_oz": 3.0},
    "Spanish Salted Marcona Almonds":   {"sku": "AC-MARC",  "packet_oz": 3.9},
    "Sweet & Smoky Almonds":            {"sku": "AC-SMAL",  "packet_oz": 3.9},
    "Dried Tart Cherries":              {"sku": "AC-DTCH",  "packet_oz": 3.9},
    "Dark Chocolate Covered Cranberries": {"sku": "AC-DCRAN", "packet_oz": 3.9},
    "Dark Chocolate Covered Almonds":   {"sku": "AC-DALM",  "packet_oz": 3.9},
    "Sun-Dried Turkish Figs":           {"sku": "AC-SDF",   "packet_oz": 3.9},
    "Piri Piri Cocktail Mix":           {"sku": "AC-PPCM",  "packet_oz": 3.9},
    "Chocolate Covered Pretzels":       {"sku": "AC-MCP",   "packet_oz": 3.9},
    "Coconut Cashews":                  {"sku": "AC-COCO",  "packet_oz": 3.9},
    "Dried Apricots":                   {"sku": "AC-APR",   "packet_oz": 3.9},
}

# SKU prefixes that represent actual shippable product (not bundles/admin)
_SHIPPABLE_PREFIXES = ("MT-", "CH-", "AC-", "PK-")

# Cross-dock timeline: Woburn receives Thursday, pickup Friday,
# available at Primary for fulfillment 2nd Saturday = ~9 days
CROSSDOCK_LEAD_DAYS = 9

# ── Cohort-based forecasting constants ───────────────────────────────

CURATION_ORDER = ["MONG", "MDT", "OWC", "SPN", "ALPT", "ISUN", "HHIGH"]

# Retention matrix: % of ORIGINAL cohort on each curation per month
# Rows = curations, columns = months 1..7
DEFAULT_RETENTION_MATRIX = {
    "MONG":  [1.00, 0.00, 0.00, 0.00, 0.00, 0.00, 0.00],
    "MDT":   [0.00, 0.57, 0.12, 0.07, 0.02, 0.00, 0.00],
    "OWC":   [0.00, 0.05, 0.30, 0.07, 0.05, 0.03, 0.00],
    "SPN":   [0.00, 0.00, 0.03, 0.15, 0.05, 0.01, 0.01],
    "ALPT":  [0.00, 0.00, 0.00, 0.03, 0.05, 0.10, 0.04],
    "ISUN":  [0.00, 0.00, 0.00, 0.00, 0.00, 0.00, 0.00],
    "HHIGH": [0.00, 0.00, 0.00, 0.00, 0.00, 0.00, 0.00],
}

DEFAULT_CHURN_RATES = {
    "MONG": {"month_1": 0.15, "month_2_plus": 0.12},
    "MS":   {"month_1": 0.15, "month_2": 0.10, "month_3_plus": 0.12},
}

REPEAT_RATE = 0.56
WHEEL_TO_SLICE_FACTOR = 2.67

# Default curation recipes: {curation: [(sku, qty), ...]}
DEFAULT_CURATION_RECIPES = {
    "MONG": [
        ("CH-MAFT", 1), ("CH-BRZ", 1), ("MT-LONZ", 1), ("MT-TUSC", 1),
        ("AC-DTCH", 1), ("AC-PRPE", 1), ("AC-TCRISP", 1),
    ],
    "MDT": [
        ("CH-SOT", 1), ("CH-TOPR", 1), ("MT-PRO", 1), ("MT-SFEN", 1),
        ("AC-SDF", 1), ("AC-LFOLIVE", 1), ("AC-RBOL", 1),
    ],
    "OWC": [
        ("CH-EBCC", 1), ("CH-FOWC", 1), ("MT-SOP", 1), ("MT-JAHH", 1),
        ("AC-SMAL", 1), ("AC-MUSTCH", 1), ("AC-PFLAT", 1),
    ],
    "SPN": [
        ("CH-WWDI", 1), ("CH-LOSC", 1), ("MT-SPAP", 1), ("MT-JAMS", 1),
        ("CH-MAU3", 1), ("AC-PPCM", 1), ("AC-RBOL", 1),
    ],
    "ALPT": [
        ("CH-IPAC", 1), ("CH-ALPHA", 1), ("MT-SLRWG", 1), ("MT-ASPK", 1),
        ("AC-PRPE", 1), ("AC-DCRAN", 1), ("AC-RBOL", 1),
    ],
    "ISUN": [
        ("CH-PVEC", 1), ("CH-WWBC", 1), ("MT-TUSC", 1), ("MT-SFEN", 1),
        ("AC-RHAZ", 1), ("AC-FLH", 1), ("AC-ACRISP", 1),
    ],
    "HHIGH": [
        ("CH-KM39", 1), ("CH-FONTAL", 1), ("MT-SPAP", 1), ("MT-JAHH", 1),
        ("AC-MARC", 1), ("AC-HON", 1), ("AC-RBOL", 1),
    ],
}

# PR-CJAM bonus: one cheese + one jam/mustard per box per curation
# Cheese must be unique across curations; jam overlap is OK
# ── Monthly curated box types (MED / CMED / LGE) ──────────────────
MONTHLY_BOX_TYPES = ["AHB-MED", "AHB-CMED", "AHB-LGE",
                     "AHB-MCUST-MS", "AHB-MCUST-NMS"]

MONTHLY_BOX_SLOTS = {
    "AHB-MED": [
        ("Cheese 1", "CH-"), ("Cheese 2", "CH-"),
        ("Meat 1", "MT-"), ("Meat 2", "MT-"),
        ("Crackers", "AC-"),
        ("Accompaniment 1", "AC-"), ("Accompaniment 2", "AC-"),
        ("PR-CJAM-GEN Cheese", "CH-"), ("PR-CJAM-GEN Jam", "AC-"),
    ],
    "AHB-CMED": [
        ("Cheese 1", "CH-"), ("Cheese 2", "CH-"),
        ("Cheese 3", "CH-"), ("Cheese 4", "CH-"),
        ("Crackers", "AC-"),
        ("Accompaniment 1", "AC-"), ("Accompaniment 2", "AC-"),
        ("PR-CJAM-GEN Cheese", "CH-"), ("PR-CJAM-GEN Jam", "AC-"),
    ],
    "AHB-LGE": [
        ("Cheese 1", "CH-"), ("Cheese 2", "CH-"),
        ("Meat 1", "MT-"), ("Meat 2", "MT-"),
        ("Crackers", "AC-"),
        ("Accompaniment 1", "AC-"), ("Accompaniment 2", "AC-"),
        ("PR-CJAM-GEN Cheese", "CH-"), ("PR-CJAM-GEN Jam", "AC-"),
        ("Extra Cheese", "CH-"), ("Extra Meat", "MT-"),
        ("CEX-ECS", "CH-"), ("EX-EC", "CH-"),
        ("CEX-EM", "MT-"), ("EX-EM", "MT-"),
    ],
}

DEFAULT_PR_CJAM = {
    "MONG": {"cheese": "CH-BLR", "jam": ""},
    "MDT": {"cheese": "CH-MAU3", "jam": ""},
    "OWC": {"cheese": "CH-RQCAV", "jam": ""},
    "SPN": {"cheese": "CH-FONTAL", "jam": ""},
    "ALPT": {"cheese": "CH-MSMG", "jam": ""},
    "ISUN": {"cheese": "CH-BAP", "jam": ""},
    "HHIGH": {"cheese": "CH-TIP", "jam": ""},
    "NMS": {"cheese": "CH-MCPC", "jam": ""},
    "BYO": {"cheese": "CH-6COM", "jam": ""},
}

# CEX-EC extra cheese assignments per curation
DEFAULT_CEX_EC = {
    "MONG": "CH-CTGOD", "MDT": "CH-RQCAV", "OWC": "CH-6COM",
    "SPN": "CH-MSMG", "HHIGH": "CH-HGCU", "BYO": "CH-HGCU",
}


# ═════════════════════════════════════════════════════════════════════
#  SETTINGS PERSISTENCE
# ═════════════════════════════════════════════════════════════════════

def _get_app_dir():
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    script_dir = os.path.dirname(os.path.abspath(__file__))
    dist_settings = os.path.join(script_dir, "dist", SETTINGS_FILE)
    if os.path.exists(dist_settings):
        return os.path.join(script_dir, "dist")
    return script_dir


def load_settings():
    path = os.path.join(_get_app_dir(), SETTINGS_FILE)
    if os.path.exists(path):
        try:
            with open(path, "r") as f:
                return json.load(f)
        except Exception:
            pass
    return {}


def save_settings(settings):
    path = os.path.join(_get_app_dir(), SETTINGS_FILE)
    try:
        with open(path, "w") as f:
            json.dump(settings, f, indent=2)
    except Exception:
        pass


# ═════════════════════════════════════════════════════════════════════
#  EXPIRATION DATE PARSING
# ═════════════════════════════════════════════════════════════════════

_EXP_DATE_FORMATS = ["%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d"]


def _parse_expiration_dates(raw_string):
    """Parse comma-separated expiration dates into sorted ISO date strings.

    Accepts MM/DD/YYYY, M/D/YY, YYYY-MM-DD formats.
    Returns a sorted list of 'YYYY-MM-DD' strings, or [].
    """
    if not raw_string or not raw_string.strip():
        return []
    results = []
    for chunk in raw_string.split(","):
        chunk = chunk.strip()
        if not chunk:
            continue
        for fmt in _EXP_DATE_FORMATS:
            try:
                dt = datetime.datetime.strptime(chunk, fmt).date()
                results.append(dt.isoformat())
                break
            except ValueError:
                continue
    results.sort()
    return results


# ═════════════════════════════════════════════════════════════════════
#  RECHARGE API CLIENT
# ═════════════════════════════════════════════════════════════════════

class RechargeClient:
    """Minimal Recharge API v1 client for pulling active subscriptions."""

    BASE_URL = "https://api.rechargeapps.com"

    def __init__(self, api_token):
        self.api_token = api_token
        self.session = None

    def _ensure_session(self):
        if self.session is None:
            try:
                import requests
                self.session = requests.Session()
                self.session.headers.update({
                    "X-Recharge-Access-Token": self.api_token,
                    "Accept": "application/json",
                    "Content-Type": "application/json",
                })
            except ImportError:
                raise RuntimeError(
                    "The 'requests' library is required for Recharge API.\n"
                    "Install with:  pip install requests"
                )

    def get_active_subscriptions(self, status="active", limit=250,
                                  progress_cb=None):
        """Yield all active subscriptions, paginating automatically."""
        self._ensure_session()
        url = f"{self.BASE_URL}/subscriptions"
        params = {"status": status, "limit": limit, "page": 1}
        all_subs = []
        while True:
            resp = self.session.get(url, params=params, timeout=30)
            resp.raise_for_status()
            data = resp.json()
            subs = data.get("subscriptions", [])
            if not subs:
                break
            all_subs.extend(subs)
            if progress_cb:
                progress_cb(len(all_subs))
            params["page"] += 1
        return all_subs

    def aggregate_sku_quantities(self, subscriptions):
        """Return {sku: total_quantity_per_interval} from subscriptions."""
        sku_qty = defaultdict(float)
        for sub in subscriptions:
            sku = sub.get("sku") or sub.get("shopify_variant_id", "UNKNOWN")
            qty = float(sub.get("quantity", 1))
            sku_qty[sku] += qty
        return dict(sku_qty)

    @staticmethod
    def build_cohorts_from_subscriptions(subscriptions):
        """Build cohort list from active subscriptions grouped by created_at month.

        Returns: [{start_month: "YYYY-MM", size: int, track: "MONG"}, ...]
        Each unique customer_id counts once, placed in their earliest created_at month.
        """
        customer_months = {}
        for sub in subscriptions:
            cid = str(sub.get("customer_id", ""))
            created = sub.get("created_at", "")
            if not cid or not created:
                continue
            try:
                dt = datetime.datetime.fromisoformat(
                    created.replace("Z", "+00:00")).replace(tzinfo=None)
                month_str = dt.strftime("%Y-%m")
            except (ValueError, AttributeError):
                continue
            if cid not in customer_months or month_str < customer_months[cid]:
                customer_months[cid] = month_str

        month_counts = defaultdict(int)
        for month in customer_months.values():
            month_counts[month] += 1

        return [{"start_month": m, "size": c, "track": "MONG"}
                for m, c in sorted(month_counts.items())]

    def get_queued_charges(self, limit=250, progress_cb=None):
        """Fetch all queued charges (committed orders not yet processed)."""
        self._ensure_session()
        url = f"{self.BASE_URL}/charges"
        params = {"status": "queued", "limit": limit}
        all_charges = []
        while True:
            resp = self.session.get(url, params=params, timeout=30)
            resp.raise_for_status()
            data = resp.json()
            charges = data.get("charges", [])
            if not charges:
                break
            all_charges.extend(charges)
            if progress_cb:
                progress_cb(len(all_charges))
            # cursor-based pagination
            next_cursor = data.get("next_cursor")
            if not next_cursor:
                break
            params = {"cursor": next_cursor, "limit": limit}
            url = f"{self.BASE_URL}/charges"
        return all_charges

    def aggregate_charges_by_month(self, charges):
        """Group queued charges by scheduled month, aggregate SKU quantities.

        Returns: {month_label: {sku: total_qty}}
        """
        by_month = defaultdict(lambda: defaultdict(float))
        for charge in charges:
            scheduled = charge.get("scheduled_at", "")
            if not scheduled:
                continue
            month_label = scheduled[:7]  # "YYYY-MM"
            for item in charge.get("line_items", []):
                sku = (item.get("sku") or "").strip()
                if not sku:
                    continue
                qty = float(item.get("quantity", 1))
                by_month[month_label][sku] += qty
        return {m: dict(skus) for m, skus in by_month.items()}

    def get_all_subscriptions(self, limit=250, progress_cb=None):
        """Fetch all subscriptions (active + cancelled) for retention analysis."""
        self._ensure_session()
        all_subs = []
        for status in ("active", "cancelled", "expired"):
            url = f"{self.BASE_URL}/subscriptions"
            params = {"status": status, "limit": limit, "page": 1}
            while True:
                resp = self.session.get(url, params=params, timeout=30)
                resp.raise_for_status()
                data = resp.json()
                subs = data.get("subscriptions", [])
                if not subs:
                    break
                all_subs.extend(subs)
                if progress_cb:
                    progress_cb(len(all_subs))
                params["page"] += 1
        return all_subs

    # Cancellation reason categories for segmentation
    CANCEL_CATEGORIES = {
        "gift": ["gift", "present", "one-time", "one time"],
        "price": ["price", "expensive", "cost", "afford", "budget"],
        "product": ["quality", "taste", "didn't like", "not what",
                     "product", "flavor"],
        "logistics": ["shipping", "delivery", "late", "damaged",
                      "missing", "lost"],
        "self_purchase": [],  # default for non-gift active subs
    }

    @staticmethod
    def categorize_cancellation(reason_text):
        """Map a cancellation reason string to a category."""
        if not reason_text:
            return "other"
        reason_lower = reason_text.lower()
        for cat, keywords in RechargeClient.CANCEL_CATEGORIES.items():
            if cat == "self_purchase":
                continue
            for kw in keywords:
                if kw in reason_lower:
                    return cat
        return "other"

    @staticmethod
    def build_retention_curves(subscriptions):
        """Build actual retention curves from subscription lifecycle data.

        Groups subscriptions by start month, calculates what % survived
        to each subsequent month.  Also builds per-cancellation-reason
        segmented curves.

        Returns:
            cohort_data: {start_month: {size, retention: [mo1_pct, ...]}}
            overall_curve: [mo1_pct, mo2_pct, ...] (averaged across cohorts)
            segmented: {reason_cat: {overall_curve: [...], cohort_count: N}}
        """
        # Group by customer + start month
        cohorts = defaultdict(lambda: {"started": 0,
                                       "active_at_month": defaultdict(int)})
        # Per-reason segmented cohorts
        seg_cohorts = defaultdict(lambda: defaultdict(
            lambda: {"started": 0, "active_at_month": defaultdict(int)}))

        for sub in subscriptions:
            created = sub.get("created_at", "")
            if not created:
                continue
            try:
                start_dt = datetime.datetime.fromisoformat(
                    created.replace("Z", "+00:00")).replace(tzinfo=None)
            except (ValueError, AttributeError):
                continue

            start_month = start_dt.strftime("%Y-%m")
            cohort = cohorts[start_month]
            cohort["started"] += 1

            # Determine cancel reason category
            cancel_reason = sub.get("cancellation_reason") or ""
            reason_cat = RechargeClient.categorize_cancellation(cancel_reason)
            status = sub.get("status", "")
            if status == "active" and not cancel_reason:
                reason_cat = "self_purchase"

            seg_cohort = seg_cohorts[reason_cat][start_month]
            seg_cohort["started"] += 1

            # Determine how many months this sub survived
            cancelled = sub.get("cancelled_at")
            if cancelled:
                try:
                    end_dt = datetime.datetime.fromisoformat(
                        cancelled.replace("Z", "+00:00")).replace(tzinfo=None)
                except (ValueError, AttributeError):
                    end_dt = datetime.datetime.now()
            elif status == "active":
                end_dt = datetime.datetime.now()
            else:
                end_dt = datetime.datetime.now()

            months_active = max(1, (
                (end_dt.year - start_dt.year) * 12 +
                end_dt.month - start_dt.month + 1))

            for m in range(min(months_active, 12)):
                cohort["active_at_month"][m] += 1
                seg_cohort["active_at_month"][m] += 1

        # Build retention percentages
        def _build_curves(cohort_dict):
            data = {}
            curves = []
            for month_label in sorted(cohort_dict.keys()):
                c = cohort_dict[month_label]
                size = c["started"]
                if size == 0:
                    continue
                retention = []
                for m in range(12):
                    active = c["active_at_month"].get(m, 0)
                    retention.append(round(active / size * 100, 1))
                data[month_label] = {"size": size, "retention": retention}
                curves.append((size, retention))

            overall = []
            if curves:
                for m in range(12):
                    total_w = 0
                    weighted_sum = 0
                    for size, ret in curves:
                        if m < len(ret):
                            weighted_sum += ret[m] * size
                            total_w += size
                    overall.append(round(weighted_sum / max(total_w, 1), 1))
            return data, overall

        cohort_data, overall = _build_curves(cohorts)

        # Build segmented curves
        segmented = {}
        for reason_cat, cat_cohorts in seg_cohorts.items():
            seg_data, seg_overall = _build_curves(cat_cohorts)
            total_subs = sum(c["started"] for c in cat_cohorts.values())
            segmented[reason_cat] = {
                "overall_curve": seg_overall,
                "cohort_count": len(seg_data),
                "total_subs": total_subs,
            }

        return cohort_data, overall, segmented


# ═════════════════════════════════════════════════════════════════════
#  SHOPIFY API CLIENT
# ═════════════════════════════════════════════════════════════════════

class ShopifyOAuth:
    """Handle Shopify OAuth flow to obtain an access token."""

    SCOPES = "read_all_orders,read_customers,read_orders,read_products"

    def __init__(self, store_url, client_id, client_secret):
        self.store_url = store_url.rstrip("/")
        if not self.store_url.startswith("http"):
            if ".myshopify.com" not in self.store_url:
                self.store_url = f"https://{self.store_url}.myshopify.com"
            else:
                self.store_url = f"https://{self.store_url}"
        self.client_id = client_id
        self.client_secret = client_secret
        self._access_token = None
        self._server = None

    def authorize(self, callback=None):
        """Start OAuth flow: open browser, run local server for callback.

        Args:
            callback: function(access_token) called on success

        Returns the access token on success, None on failure.
        """
        import webbrowser
        import hashlib
        import hmac
        from http.server import HTTPServer, BaseHTTPRequestHandler
        from urllib.parse import urlencode, urlparse, parse_qs
        try:
            import requests
        except ImportError:
            raise RuntimeError("requests library required")

        port = 21849  # local callback port
        redirect_uri = f"http://localhost:{port}/callback"
        nonce = hashlib.sha256(os.urandom(16)).hexdigest()[:16]

        # Build authorization URL
        auth_params = urlencode({
            "client_id": self.client_id,
            "scope": self.SCOPES,
            "redirect_uri": redirect_uri,
            "state": nonce,
        })
        auth_url = f"{self.store_url}/admin/oauth/authorize?{auth_params}"

        oauth_ref = self  # reference for handler closure

        class CallbackHandler(BaseHTTPRequestHandler):
            def do_GET(self):
                parsed = urlparse(self.path)
                params = parse_qs(parsed.query)

                if parsed.path != "/callback" or "code" not in params:
                    self.send_response(400)
                    self.send_header("Content-Type", "text/html")
                    self.end_headers()
                    self.wfile.write(
                        b"<h2>Authorization failed.</h2>"
                        b"<p>No authorization code received.</p>")
                    return

                code = params["code"][0]
                state = params.get("state", [""])[0]

                if state != nonce:
                    self.send_response(403)
                    self.send_header("Content-Type", "text/html")
                    self.end_headers()
                    self.wfile.write(
                        b"<h2>State mismatch - possible CSRF.</h2>")
                    return

                # Exchange code for access token
                try:
                    resp = requests.post(
                        f"{oauth_ref.store_url}/admin/oauth/access_token",
                        json={
                            "client_id": oauth_ref.client_id,
                            "client_secret": oauth_ref.client_secret,
                            "code": code,
                        },
                        timeout=15)
                    resp.raise_for_status()
                    token = resp.json().get("access_token")
                    oauth_ref._access_token = token
                    self.send_response(200)
                    self.send_header("Content-Type", "text/html")
                    self.end_headers()
                    self.wfile.write(
                        b"<h2>Authorization successful!</h2>"
                        b"<p>You can close this tab and return "
                        b"to the app.</p>")
                except Exception as e:
                    self.send_response(500)
                    self.send_header("Content-Type", "text/html")
                    self.end_headers()
                    self.wfile.write(
                        f"<h2>Token exchange failed:</h2>"
                        f"<pre>{e}</pre>".encode())

            def log_message(self, format, *args):
                pass  # suppress console output

        # Start local server in a thread, open browser
        server = HTTPServer(("localhost", port), CallbackHandler)
        server.timeout = 120  # 2 minute timeout

        webbrowser.open(auth_url)

        # Wait for the callback (blocking, but called from a thread)
        server.handle_request()
        server.server_close()

        if self._access_token and callback:
            callback(self._access_token)

        return self._access_token


class ShopifyClient:
    """Shopify Admin API client for pulling Subscription First Orders."""

    API_VERSION = "2025-01"

    def __init__(self, store_url, access_token):
        self.store_url = store_url.rstrip("/")
        # normalize: accept "mystore" or "mystore.myshopify.com" or full URL
        if not self.store_url.startswith("http"):
            if ".myshopify.com" not in self.store_url:
                self.store_url = f"https://{self.store_url}.myshopify.com"
            else:
                self.store_url = f"https://{self.store_url}"
        self.access_token = access_token
        self.session = None

    def _ensure_session(self):
        if self.session is None:
            try:
                import requests
                self.session = requests.Session()
                self.session.headers.update({
                    "X-Shopify-Access-Token": self.access_token,
                    "Accept": "application/json",
                    "Content-Type": "application/json",
                })
            except ImportError:
                raise RuntimeError(
                    "The 'requests' library is required for Shopify API.\n"
                    "Install with:  pip install requests"
                )

    def get_orders(self, tag=None, created_at_min=None, status="any",
                   limit=250, progress_cb=None):
        """Fetch orders with optional tag filter, paginating automatically."""
        self._ensure_session()
        url = (f"{self.store_url}/admin/api/{self.API_VERSION}/orders.json")
        params = {"status": status, "limit": limit}
        if tag:
            params["tag"] = tag
        if created_at_min:
            params["created_at_min"] = created_at_min

        all_orders = []
        while url:
            resp = self.session.get(url, params=params, timeout=30)
            resp.raise_for_status()
            data = resp.json()
            orders = data.get("orders", [])
            if not orders:
                break
            all_orders.extend(orders)
            if progress_cb:
                progress_cb(len(all_orders))

            # cursor-based pagination via Link header
            url = None
            params = {}  # clear params for next page (encoded in URL)
            link_header = resp.headers.get("Link", "")
            if 'rel="next"' in link_header:
                for part in link_header.split(","):
                    if 'rel="next"' in part:
                        url = part.split("<")[1].split(">")[0]
                        break

        return all_orders

    def aggregate_first_order_skus(self, orders, weeks_back=4):
        """Aggregate line-item SKUs from orders into weekly averages.

        Returns: {sku: weekly_avg_qty}
        """
        sku_totals = defaultdict(float)
        order_count = 0

        for order in orders:
            order_count += 1
            for item in order.get("line_items", []):
                sku = (item.get("sku") or "").strip()
                if not sku:
                    continue
                qty = float(item.get("quantity", 1))
                sku_totals[sku] += qty

        # compute weekly average
        if weeks_back <= 0:
            weeks_back = 1
        weekly = {}
        for sku, total in sku_totals.items():
            weekly[sku] = round(total / weeks_back, 2)

        return weekly, order_count

    def aggregate_with_trend(self, orders, weeks_back=4):
        """Aggregate orders into weekly buckets and compute trend via linear
        regression.

        Returns:
            trend_data: {sku: {weekly_avg, trend_slope, projected_next_week,
                               pct_of_orders}}
            order_count: int
            weekly_totals: [total_orders_per_week]  (oldest-first)
        """
        if not orders:
            return {}, 0, []

        # parse order dates and bucket into week indices (0 = oldest week)
        now = datetime.datetime.now()
        cutoff = now - datetime.timedelta(weeks=weeks_back)

        # per-SKU weekly buckets and per-order SKU presence
        sku_weekly = defaultdict(lambda: [0.0] * weeks_back)
        weekly_order_counts = [0] * weeks_back
        sku_order_presence = defaultdict(int)
        order_count = 0

        for order in orders:
            created = order.get("created_at", "")
            if not created:
                continue
            try:
                dt = datetime.datetime.fromisoformat(
                    created.replace("Z", "+00:00")).replace(tzinfo=None)
            except (ValueError, AttributeError):
                continue
            if dt < cutoff:
                continue

            delta = now - dt
            week_idx = min(int(delta.days / 7), weeks_back - 1)
            # invert so index 0 = oldest
            week_idx = (weeks_back - 1) - week_idx

            order_count += 1
            weekly_order_counts[week_idx] += 1

            skus_in_order = set()
            for item in order.get("line_items", []):
                sku = (item.get("sku") or "").strip()
                if not sku:
                    continue
                qty = float(item.get("quantity", 1))
                sku_weekly[sku][week_idx] += qty
                skus_in_order.add(sku)
            for sku in skus_in_order:
                sku_order_presence[sku] += 1

        # linear regression helper: y = a + b*x
        def _linreg(values):
            n = len(values)
            if n < 2:
                return 0.0
            x_mean = (n - 1) / 2.0
            y_mean = sum(values) / n
            num = sum((i - x_mean) * (v - y_mean) for i, v in enumerate(values))
            den = sum((i - x_mean) ** 2 for i in range(n))
            if den == 0:
                return 0.0
            return num / den

        trend_data = {}
        total_orders = max(order_count, 1)
        for sku, weekly_vals in sku_weekly.items():
            avg = sum(weekly_vals) / max(len(weekly_vals), 1)
            slope = _linreg(weekly_vals)
            # project next week: value at index = weeks_back (one beyond last)
            projected = max(0.0, weekly_vals[-1] + slope)
            pct = round(sku_order_presence.get(sku, 0) / total_orders * 100, 1)
            trend_data[sku] = {
                "weekly_avg": round(avg, 2),
                "trend_slope": round(slope, 3),
                "projected_next_week": round(projected, 2),
                "pct_of_orders": pct,
            }

        return trend_data, order_count, weekly_order_counts

    def build_customer_lifecycle(self, months_back=12, progress_cb=None):
        """Pull subscription orders and build customer lifecycle data.

        Uses streaming batch processing — processes each page of orders
        immediately, extracts lifecycle data, discards raw response.
        Memory: ~50MB for 100K+ orders vs ~1GB accumulating all.

        Returns:
            lifecycle: {email: {first_order_date, order_count,
                                last_order_date, months_active}}
            cohort_sizes: {month: new_subscriber_count}
            retention_by_cohort: {start_month: {month_N: active_count}}
            reship_rate: float (% of orders tagged Reship)
        """
        self._ensure_session()
        min_date = (datetime.datetime.now() -
                    datetime.timedelta(days=months_back * 30))
        min_date_str = min_date.strftime("%Y-%m-%dT00:00:00-00:00")

        # Streaming aggregation: process each page and discard raw orders
        customers = defaultdict(lambda: {
            "first_order_date": None, "orders": [], "is_reship": []})
        total_orders = 0
        reship_orders = 0

        for tag in ("Subscription First Order", "Subscription Recurring Order"):
            url = (f"{self.store_url}/admin/api/{self.API_VERSION}/orders.json")
            params = {"status": "any", "limit": 250, "tag": tag,
                      "created_at_min": min_date_str}

            while url:
                resp = self.session.get(url, params=params, timeout=30)
                resp.raise_for_status()
                data = resp.json()
                page_orders = data.get("orders", [])
                if not page_orders:
                    break

                # Process this page immediately, then discard
                for order in page_orders:
                    email = (order.get("email") or "").strip().lower()
                    if not email:
                        continue
                    created = order.get("created_at", "")
                    if not created:
                        continue
                    try:
                        dt = datetime.datetime.fromisoformat(
                            created.replace("Z", "+00:00")).replace(
                                tzinfo=None)
                    except (ValueError, AttributeError):
                        continue

                    tags = (order.get("tags") or "").lower()
                    total_orders += 1
                    is_reship = "reship" in tags
                    if is_reship:
                        reship_orders += 1

                    cust = customers[email]
                    cust["orders"].append(dt)
                    cust["is_reship"].append(is_reship)
                    if (cust["first_order_date"] is None or
                            dt < cust["first_order_date"]):
                        cust["first_order_date"] = dt

                if progress_cb:
                    progress_cb(total_orders)

                # cursor-based pagination via Link header
                url = None
                params = {}
                link_header = resp.headers.get("Link", "")
                if 'rel="next"' in link_header:
                    for part in link_header.split(","):
                        if 'rel="next"' in part:
                            url = part.split("<")[1].split(">")[0]
                            break

                del page_orders, data  # free page memory immediately

        # Build lifecycle summary
        lifecycle = {}
        cohort_sizes = defaultdict(int)
        retention_by_cohort = defaultdict(lambda: defaultdict(int))

        for email, cust in customers.items():
            first = cust["first_order_date"]
            orders = sorted(cust["orders"])
            last = orders[-1]
            months_active = max(1, (
                (last.year - first.year) * 12 +
                last.month - first.month + 1))

            start_month = first.strftime("%Y-%m")
            cohort_sizes[start_month] += 1

            lifecycle[email] = {
                "first_order_date": first.isoformat()[:10],
                "order_count": len(orders),
                "last_order_date": last.isoformat()[:10],
                "months_active": months_active,
            }

            # Track retention: for each month this customer had an order
            order_months = set()
            for o in orders:
                month_offset = ((o.year - first.year) * 12 +
                                o.month - first.month)
                order_months.add(month_offset)

            for m in order_months:
                retention_by_cohort[start_month][m] += 1

        reship_rate = round(
            reship_orders / max(total_orders, 1) * 100, 2)

        return (dict(lifecycle), dict(cohort_sizes),
                {k: dict(v) for k, v in retention_by_cohort.items()},
                reship_rate)


# ═════════════════════════════════════════════════════════════════════
#  DATA MODELS / CALCULATIONS
# ═════════════════════════════════════════════════════════════════════

def calculate_reorder_point(daily_usage, lead_time_days, safety_stock):
    """Reorder Point = (Daily Usage x Total Lead Time) + Safety Stock"""
    return (daily_usage * lead_time_days) + safety_stock


def calculate_total_lead_time(purchase_lt, production_lt, shipping_lt):
    return purchase_lt + production_lt + shipping_lt


def decompose_bundles(bundle_sku, quantity, bundle_map):
    """Break a bundle SKU into component SKUs using bundle_map.

    bundle_map: {bundle_sku: [(component_sku, qty_per_bundle), ...]}
    Returns: [(component_sku, total_qty), ...] or [(bundle_sku, quantity)]
    if not a bundle.
    """
    if bundle_sku in bundle_map:
        return [
            (comp_sku, quantity * comp_qty)
            for comp_sku, comp_qty in bundle_map[bundle_sku]
        ]
    return [(bundle_sku, quantity)]


def apply_churn_rate(quantity, churn_pct):
    """Reduce quantity by churn percentage (e.g., 5% churn -> 95% remains)."""
    return quantity * (1.0 - churn_pct / 100.0)


# ── Queued-charge curation resolution ────────────────────────────────

KNOWN_CURATIONS = set(CURATION_ORDER) | {"NMS", "BYO", "SS", "MS"}

_MONTHLY_PATTERNS = {"AHB-MED", "AHB-LGE", "AHB-CMED", "AHB-CUR-MS", "AHB-BVAL"}


def resolve_curation_from_box_sku(sku):
    """Determine curation from an AHB box SKU.

    Returns curation string (e.g. "MDT"), "MONTHLY" for monthly boxes,
    or None if unknown/empty.
    """
    if not sku:
        return None
    sku = sku.strip().upper()
    if sku in _MONTHLY_PATTERNS:
        return "MONTHLY"
    # AHB-X<CUR> shorthand (e.g. AHB-XMONG)
    if sku.startswith("AHB-X") and not sku.startswith("AHB-XL"):
        cur = sku[5:]
        return cur if cur in KNOWN_CURATIONS else None
    # AHB-MCUST-* or AHB-LCUST-* (e.g. AHB-LCUST-CORS-MDT)
    if "CUST" in sku:
        parts = sku.split("-")
        # Walk backwards to find a known curation (skip CORS variant tag)
        for seg in reversed(parts):
            if seg in KNOWN_CURATIONS:
                return seg
        return None
    return None


def resolve_queued_charges(charges):
    """Resolve generic PR-CJAM/CEX-EC SKUs using per-charge box context.

    Args:
        charges: raw charge list from Recharge API

    Returns:
        {month: {
            "pr_cjam": {curation_suffix: count},
            "cex_ec": {curation_suffix: count},
            "unresolved": int
        }}
    """
    # Group line items by charge_id
    charges_by_id = defaultdict(list)
    charge_months = {}
    for charge in charges:
        cid = charge.get("id")
        scheduled = charge.get("scheduled_at", "")
        if not cid or not scheduled:
            continue
        month_label = scheduled[:7]
        charge_months[cid] = month_label
        for item in charge.get("line_items", []):
            sku = (item.get("sku") or "").strip()
            qty = float(item.get("quantity", 1))
            if sku:
                charges_by_id[cid].append((sku, qty))

    result = defaultdict(lambda: {"pr_cjam": defaultdict(float),
                                   "cex_ec": defaultdict(float),
                                   "unresolved": 0})

    for cid, items in charges_by_id.items():
        month = charge_months.get(cid)
        if not month:
            continue

        # Find box SKU on this charge
        box_sku = None
        for sku, _qty in items:
            upper = sku.upper()
            if upper.startswith("AHB-"):
                box_sku = upper
                break

        curation = resolve_curation_from_box_sku(box_sku)

        # Track whether this charge has PR-CJAM
        has_pr_cjam = False

        for sku, qty in items:
            upper = sku.upper()

            # Already-specific PR-CJAM (e.g. PR-CJAM-MONG) — pass through
            if upper.startswith("PR-CJAM-") and upper != "PR-CJAM-GEN":
                suffix = upper.split("PR-CJAM-", 1)[1]
                result[month]["pr_cjam"][suffix] += qty
                has_pr_cjam = True

            # Generic PR-CJAM-GEN
            elif upper == "PR-CJAM-GEN":
                has_pr_cjam = True
                if curation and curation != "MONTHLY":
                    result[month]["pr_cjam"][curation] += qty
                elif curation == "MONTHLY":
                    # Monthly boxes keep PR-CJAM-GEN as-is
                    result[month]["pr_cjam"]["GEN"] += qty
                else:
                    result[month]["unresolved"] += int(qty)

            # CEX-EC (bare or with suffix)
            elif upper.startswith("CEX-EC"):
                if upper == "CEX-EC":
                    # Bare CEX-EC — resolve from box curation
                    if curation and curation != "MONTHLY":
                        result[month]["cex_ec"][curation] += qty
                    elif curation == "MONTHLY":
                        result[month]["cex_ec"]["GEN"] += qty
                    else:
                        result[month]["unresolved"] += int(qty)
                else:
                    # Already specific (e.g. CEX-EC-MONG)
                    suffix = upper.split("CEX-EC-", 1)[1] if "-" in upper[6:] else upper[6:]
                    result[month]["cex_ec"][suffix] += qty

        # Infer missing PR-CJAM for CUST boxes (every sub box gets one)
        if not has_pr_cjam and curation and curation not in ("MONTHLY", None):
            result[month]["pr_cjam"][curation] += 1

    # Convert nested defaultdicts to plain dicts
    return {m: {"pr_cjam": dict(v["pr_cjam"]),
                "cex_ec": dict(v["cex_ec"]),
                "unresolved": v["unresolved"]}
            for m, v in result.items()}


# ── Cohort-based forecast engine ─────────────────────────────────────

def forecast_cohort_demand(cohorts, retention_matrix, curation_recipes,
                           pr_cjam, cex_ec, forecast_months=3):
    """Main forecast engine: project SKU demand from cohort data.

    Args:
        cohorts: list of {start_month: "YYYY-MM", size: int, track: str}
        retention_matrix: {curation: [mo1_pct, mo2_pct, ...]}
        curation_recipes: {curation: [(sku, qty), ...]}
        pr_cjam: {curation: {cheese: sku, jam: sku}}
        cex_ec: {curation: cheese_sku}
        forecast_months: how many months ahead to forecast

    Returns:
        {month_label: {
            curation_counts: {curation: box_count},
            sku_demand: {sku: qty},
            total_boxes: int
        }}
    """
    today = datetime.date.today()
    results = {}

    for offset in range(forecast_months):
        target_date = today.replace(day=1)
        # advance by offset months
        month = target_date.month + offset
        year = target_date.year
        while month > 12:
            month -= 12
            year += 1
        target_date = target_date.replace(year=year, month=month)
        month_label = target_date.strftime("%Y-%m")

        curation_counts = defaultdict(float)
        sku_demand = defaultdict(float)
        total_boxes = 0

        for cohort in cohorts:
            start_str = cohort.get("start_month", "")
            size = cohort.get("size", 0)
            if not start_str or size <= 0:
                continue

            # calculate age in months
            try:
                start_parts = start_str.split("-")
                start_year = int(start_parts[0])
                start_month = int(start_parts[1])
            except (ValueError, IndexError):
                continue

            age = (year - start_year) * 12 + (month - start_month) + 1
            if age < 1 or age > 7:
                continue

            # use retention matrix to get curation distribution
            col_idx = age - 1  # 0-indexed
            for curation in CURATION_ORDER:
                matrix_row = retention_matrix.get(curation, [0] * 7)
                if col_idx < len(matrix_row):
                    pct = matrix_row[col_idx]
                else:
                    pct = 0
                if pct <= 0:
                    continue

                boxes = size * pct
                curation_counts[curation] += boxes
                total_boxes += boxes

                # multiply by recipe
                recipe = curation_recipes.get(curation, [])
                for sku, qty in recipe:
                    sku_demand[sku] += boxes * qty

                # PR-CJAM contribution (one cheese + one jam per box)
                pr_entry = pr_cjam.get(curation)
                if pr_entry:
                    # Support both old str format and new dict format
                    if isinstance(pr_entry, str):
                        if pr_entry:
                            sku_demand[pr_entry] += boxes
                    else:
                        cheese = pr_entry.get("cheese", "")
                        jam = pr_entry.get("jam", "")
                        if cheese:
                            sku_demand[cheese] += boxes
                        if jam:
                            sku_demand[jam] += boxes

                # CEX-EC contribution (one per box for large boxes)
                ec_sku = cex_ec.get(curation)
                if ec_sku:
                    # approximate: ~40% of boxes are large
                    sku_demand[ec_sku] += boxes * 0.4

        results[month_label] = {
            "curation_counts": dict(curation_counts),
            "sku_demand": {k: round(v, 1) for k, v in sku_demand.items()},
            "total_boxes": round(total_boxes, 1),
        }

    return results


def compute_wheel_supply(wheel_inventory, adjusted_factors=None):
    """Convert cheese wheel inventory to potential slices.

    Args:
        wheel_inventory: {sku: {weight_lbs: float, count: int, target_sku: str}}
        adjusted_factors: {sku: float} per-SKU conversion factors from yield history

    Returns:
        {target_sku: slice_count}
    """
    if adjusted_factors is None:
        adjusted_factors = {}
    supply = defaultdict(float)
    for sku, info in wheel_inventory.items():
        weight = info.get("weight_lbs", 0)
        count = info.get("count", 0)
        target = info.get("target_sku", sku)
        factor = adjusted_factors.get(target, WHEEL_TO_SLICE_FACTOR)
        slices = weight * count * factor
        if slices > 0:
            supply[target] += round(slices, 1)
    return dict(supply)


def compute_bulk_supply(inventory, bulk_conversions):
    """Convert bulk raw material inventory to potential finished packets.

    Args:
        inventory: {sku: {qty, name, unit, unit_size, ...}}
        bulk_conversions: {keyword: {sku, packet_oz}}

    Returns:
        {target_sku: packet_count}
    """
    supply = defaultdict(float)
    for inv_sku, inv_data in inventory.items():
        name = inv_data.get("name", "")
        if not name:
            continue
        # Only process bulk raw materials
        cat = inv_data.get("category", "")
        if cat != "Bulk Raw Materials":
            continue
        qty = float(inv_data.get("qty", 0))
        if qty <= 0:
            continue

        # Find matching conversion
        for keyword, conv in bulk_conversions.items():
            if keyword.lower() in name.lower():
                target_sku = conv["sku"]
                packet_oz = conv.get("packet_oz", 3.9)
                unit = inv_data.get("unit", "").lower()
                unit_size = inv_data.get("unit_size", 1)

                # Determine total oz available
                if "lb" in unit:
                    total_oz = unit_size * qty * 16
                elif "oz" in unit:
                    total_oz = unit_size * qty
                elif "kg" in unit:
                    total_oz = unit_size * qty * 35.274
                else:
                    # Assume unit_size is in lbs
                    total_oz = unit_size * qty * 16

                packets = total_oz / packet_oz if packet_oz > 0 else 0
                if packets > 0:
                    supply[target_sku] += round(packets, 1)
                break
    return dict(supply)


def compute_monthly_box_demand(monthly_box_recipes, monthly_box_counts,
                                forecast_months=3):
    """Compute SKU demand from monthly curated box recipes.

    Returns: {month_label: {sku_demand: {sku: qty}, box_counts: {type: count}}}
    """
    today = datetime.date.today()
    results = {}

    for offset in range(forecast_months):
        m = today.month + offset
        y = today.year
        while m > 12:
            m -= 12
            y += 1
        month_label = f"{y:04d}-{m:02d}"

        sku_demand = defaultdict(float)
        box_info = {}

        for box_type in MONTHLY_BOX_TYPES:
            count = monthly_box_counts.get(box_type, 0)
            if count <= 0:
                continue

            # use this month's recipe if available, else fall back to current month's
            recipe = None
            month_data = monthly_box_recipes.get(month_label, {})
            if box_type in month_data:
                recipe = month_data[box_type]
            else:
                # fall back to current month's recipe as estimate
                cur_label = f"{today.year:04d}-{today.month:02d}"
                cur_data = monthly_box_recipes.get(cur_label, {})
                if box_type in cur_data:
                    recipe = cur_data[box_type]

            if not recipe:
                continue

            box_info[box_type] = count
            for item in recipe:
                sku = item[1] if len(item) > 1 else ""
                qty = item[2] if len(item) > 2 else 1
                if sku:
                    sku_demand[sku] += count * qty

        if sku_demand or box_info:
            results[month_label] = {
                "sku_demand": {k: round(v, 1) for k, v in sku_demand.items()},
                "box_counts": box_info,
            }

    return results


def compute_reorder_alerts(forecast, inventory, open_pos, wheel_supply,
                           bulk_supply=None):
    """Compare forecast demand to supply chain, generate alerts.

    Args:
        forecast: output of forecast_cohort_demand (one month)
        inventory: {sku: {qty: float, ...}}
        open_pos: list of {sku, qty, eta, type, vendor, status}
        wheel_supply: output of compute_wheel_supply
        bulk_supply: output of compute_bulk_supply (optional)

    Returns:
        list of {sku, action, urgency, deficit, needed_by, current_supply}
    """
    sku_demand = forecast.get("sku_demand", {})
    alerts = []

    # aggregate open POs by SKU
    po_by_sku = defaultdict(float)
    for po in open_pos:
        if po.get("status", "").lower() != "received":
            po_by_sku[po["sku"]] += po.get("qty", 0)

    for sku, demand in sku_demand.items():
        if demand <= 0:
            continue

        on_hand = float(inventory.get(sku, {}).get("qty", 0))
        po_qty = po_by_sku.get(sku, 0)
        ws = wheel_supply.get(sku, 0)
        bs = (bulk_supply or {}).get(sku, 0)
        net = on_hand + po_qty + ws + bs - demand

        if net >= 0:
            continue

        deficit = abs(net)

        # determine action type
        if sku.startswith("CH-"):
            action = "MFG" if ws > 0 or sku in wheel_supply else "PO"
        elif sku.startswith("AC-"):
            action = "MFG" if bs > 0 else "PO"
        elif sku.startswith("MT-"):
            action = "PO"
        else:
            action = "PO"

        # determine urgency based on how soon we run out
        days_of_stock = (on_hand + po_qty + ws + bs) / (demand / 30) if demand > 0 else 999
        if days_of_stock <= 3:
            urgency = "CRITICAL"
        elif days_of_stock <= 10:
            urgency = "WARNING"
        else:
            urgency = "PLAN"

        alerts.append({
            "sku": sku,
            "action": action,
            "urgency": urgency,
            "deficit": round(deficit, 1),
            "needed_by": forecast.get("month_label", ""),
            "current_supply": round(on_hand + po_qty + ws + bs, 1),
            "on_hand": on_hand,
            "open_po": po_qty,
            "wheel_supply": ws,
            "bulk_supply": bs,
            "demand": demand,
        })

    # sort by urgency
    urgency_order = {"CRITICAL": 0, "WARNING": 1, "PLAN": 2}
    alerts.sort(key=lambda a: urgency_order.get(a["urgency"], 3))
    return alerts


# ═════════════════════════════════════════════════════════════════════
#  CSV COLUMN MAPPING DIALOG
# ═════════════════════════════════════════════════════════════════════

class ColumnMappingDialog(tk.Toplevel):
    """Modal dialog for mapping CSV columns to app fields."""

    REQUIRED_FIELDS = ["SKU", "Quantity On Hand"]
    OPTIONAL_FIELDS = ["Product Name", "Unit Cost",
                       "Expiration Dates"]

    def __init__(self, parent, csv_headers):
        super().__init__(parent)
        self.title("Map CSV Columns")
        self.configure(bg=_BG)
        self.transient(parent)
        self.grab_set()
        self.result = None

        self.csv_headers = csv_headers
        all_fields = self.REQUIRED_FIELDS + self.OPTIONAL_FIELDS
        self.mapping_vars = {}

        # instructions
        ttk.Label(self, text="Map your CSV columns to app fields:",
                  style="Bold.TLabel").pack(anchor="w", padx=15, pady=(15, 5))
        ttk.Label(self, text="Leave unmapped fields as '-- skip --'",
                  style="Dim.TLabel").pack(anchor="w", padx=15, pady=(0, 10))

        frame = ttk.Frame(self)
        frame.pack(fill="both", expand=True, padx=15, pady=5)

        options = ["-- skip --"] + list(csv_headers)

        for i, field in enumerate(all_fields):
            lbl_text = f"{field} *" if field in self.REQUIRED_FIELDS else field
            ttk.Label(frame, text=lbl_text).grid(
                row=i, column=0, sticky="w", padx=(0, 10), pady=3)
            var = tk.StringVar(value=self._auto_match(field, csv_headers))
            combo = ttk.Combobox(frame, textvariable=var, values=options,
                                 state="readonly", width=30)
            combo.grid(row=i, column=1, pady=3)
            self.mapping_vars[field] = var

        # buttons
        btn_frame = ttk.Frame(self)
        btn_frame.pack(fill="x", padx=15, pady=15)
        tk.Button(btn_frame, text="Cancel", command=self.destroy,
                  bg=_BG3, fg=_FG, relief="flat", padx=15, pady=5
                  ).pack(side="right", padx=(5, 0))
        tk.Button(btn_frame, text="Apply Mapping", command=self._on_apply,
                  bg=_GREEN, fg="white", relief="flat", padx=15, pady=5
                  ).pack(side="right")

        self.geometry("500x480")
        self.minsize(450, 380)

    @staticmethod
    def _auto_match(field, headers):
        """Try to auto-match field name to a CSV header."""
        field_lower = field.lower().replace(" ", "").replace("_", "")
        for h in headers:
            h_lower = h.lower().replace(" ", "").replace("_", "")
            if field_lower == h_lower or field_lower in h_lower:
                return h
        # common aliases
        aliases = {
            "sku": ["sku", "item", "itemcode", "productcode",
                    "variantsku", "productsku"],
            "quantityonhand": ["qty", "quantity", "stock", "onhand",
                               "qtyonhand", "instock", "available",
                               "total", "rmfg"],
            "productname": ["name", "title", "product", "description",
                            "productname", "producttitle", "itemname",
                            "ingredient"],
            "unitcost": ["cost", "price", "unitprice", "unitcost"],
            "expirationdates": ["expiration", "expiry", "bestby",
                                "bestbefore", "useby",
                                "expirationdates"],
        }
        if field_lower in aliases:
            for h in headers:
                h_clean = h.lower().replace(" ", "").replace("_", "")
                if h_clean in aliases[field_lower]:
                    return h
        return "-- skip --"

    def _on_apply(self):
        mapping = {}
        for field, var in self.mapping_vars.items():
            val = var.get()
            if val != "-- skip --":
                mapping[field] = val
        # validate required
        for req in self.REQUIRED_FIELDS:
            if req not in mapping:
                messagebox.showerror(
                    "Missing Required Field",
                    f"'{req}' must be mapped to a CSV column.",
                    parent=self)
                return
        self.result = mapping
        self.destroy()


# ═════════════════════════════════════════════════════════════════════
#  BUNDLE MAPPING EDITOR DIALOG
# ═════════════════════════════════════════════════════════════════════

class BundleMappingEditor(tk.Toplevel):
    """In-app editor for bundle → component SKU mappings."""

    def __init__(self, parent, bundle_map):
        super().__init__(parent)
        self.title("Bundle Mapping Editor")
        self.configure(bg=_BG)
        self.transient(parent)
        self.grab_set()
        self.result = None

        # deep copy
        self.bundle_map = {k: list(v) for k, v in bundle_map.items()}

        # ── left: bundle list ──
        left = ttk.Frame(self)
        left.pack(side="left", fill="y", padx=(15, 5), pady=15)

        ttk.Label(left, text="Bundles", style="Bold.TLabel").pack(anchor="w")
        self.bundle_listbox = tk.Listbox(
            left, bg=_BG2, fg=_FG, selectbackground=_ACC,
            width=25, height=20, relief="flat", bd=0)
        self.bundle_listbox.pack(fill="y", expand=True, pady=(5, 5))
        self.bundle_listbox.bind("<<ListboxSelect>>", self._on_bundle_select)

        btn_row = ttk.Frame(left)
        btn_row.pack(fill="x")
        tk.Button(btn_row, text="+ Add", command=self._add_bundle,
                  bg=_GREEN, fg="white", relief="flat", padx=8, pady=3
                  ).pack(side="left", padx=(0, 3))
        tk.Button(btn_row, text="- Remove", command=self._remove_bundle,
                  bg=_RED, fg="white", relief="flat", padx=8, pady=3
                  ).pack(side="left")

        # ── right: components ──
        right = ttk.Frame(self)
        right.pack(side="left", fill="both", expand=True, padx=(5, 15),
                   pady=15)

        ttk.Label(right, text="Components", style="Bold.TLabel"
                  ).pack(anchor="w")
        self.comp_tree = ttk.Treeview(
            right, columns=("sku", "qty"), show="headings", height=14)
        self.comp_tree.heading("sku", text="Component SKU")
        self.comp_tree.heading("qty", text="Qty Per Bundle")
        self.comp_tree.column("sku", width=200)
        self.comp_tree.column("qty", width=100, anchor="center")
        self.comp_tree.pack(fill="both", expand=True, pady=(5, 5))

        comp_btns = ttk.Frame(right)
        comp_btns.pack(fill="x")
        tk.Button(comp_btns, text="+ Add Component",
                  command=self._add_component,
                  bg=_GREEN, fg="white", relief="flat", padx=8, pady=3
                  ).pack(side="left", padx=(0, 3))
        tk.Button(comp_btns, text="- Remove Component",
                  command=self._remove_component,
                  bg=_RED, fg="white", relief="flat", padx=8, pady=3
                  ).pack(side="left", padx=(0, 3))
        tk.Button(comp_btns, text="Edit", command=self._edit_component,
                  bg=_BG3, fg=_FG, relief="flat", padx=8, pady=3
                  ).pack(side="left")

        # ── bottom buttons ──
        bottom = ttk.Frame(self)
        bottom.pack(fill="x", padx=15, pady=(0, 15))
        tk.Button(bottom, text="Cancel", command=self.destroy,
                  bg=_BG3, fg=_FG, relief="flat", padx=15, pady=5
                  ).pack(side="right", padx=(5, 0))
        tk.Button(bottom, text="Save", command=self._on_save,
                  bg=_GREEN, fg="white", relief="flat", padx=15, pady=5
                  ).pack(side="right")

        self._refresh_bundle_list()
        self.geometry("620x500")
        self.minsize(500, 400)

    def _refresh_bundle_list(self):
        self.bundle_listbox.delete(0, "end")
        for sku in sorted(self.bundle_map.keys()):
            self.bundle_listbox.insert("end", sku)

    def _on_bundle_select(self, event=None):
        sel = self.bundle_listbox.curselection()
        if not sel:
            return
        bundle_sku = self.bundle_listbox.get(sel[0])
        self._refresh_components(bundle_sku)

    def _refresh_components(self, bundle_sku):
        for item in self.comp_tree.get_children():
            self.comp_tree.delete(item)
        for comp_sku, qty in self.bundle_map.get(bundle_sku, []):
            self.comp_tree.insert("", "end", values=(comp_sku, qty))

    def _get_selected_bundle(self):
        sel = self.bundle_listbox.curselection()
        if not sel:
            messagebox.showinfo("Select Bundle",
                                "Select a bundle from the list first.",
                                parent=self)
            return None
        return self.bundle_listbox.get(sel[0])

    def _add_bundle(self):
        sku = simpledialog.askstring("New Bundle", "Bundle SKU:",
                                     parent=self)
        if sku and sku.strip():
            sku = sku.strip()
            if sku not in self.bundle_map:
                self.bundle_map[sku] = []
            self._refresh_bundle_list()

    def _remove_bundle(self):
        sku = self._get_selected_bundle()
        if sku and messagebox.askyesno("Confirm",
                                       f"Remove bundle '{sku}'?",
                                       parent=self):
            del self.bundle_map[sku]
            self._refresh_bundle_list()
            for item in self.comp_tree.get_children():
                self.comp_tree.delete(item)

    def _add_component(self):
        bundle = self._get_selected_bundle()
        if not bundle:
            return
        dlg = _ComponentEntryDialog(self)
        self.wait_window(dlg)
        if dlg.result:
            self.bundle_map[bundle].append(dlg.result)
            self._refresh_components(bundle)

    def _remove_component(self):
        bundle = self._get_selected_bundle()
        if not bundle:
            return
        sel = self.comp_tree.selection()
        if not sel:
            return
        idx = self.comp_tree.index(sel[0])
        del self.bundle_map[bundle][idx]
        self._refresh_components(bundle)

    def _edit_component(self):
        bundle = self._get_selected_bundle()
        if not bundle:
            return
        sel = self.comp_tree.selection()
        if not sel:
            return
        idx = self.comp_tree.index(sel[0])
        old_sku, old_qty = self.bundle_map[bundle][idx]
        dlg = _ComponentEntryDialog(self, old_sku, old_qty)
        self.wait_window(dlg)
        if dlg.result:
            self.bundle_map[bundle][idx] = dlg.result
            self._refresh_components(bundle)

    def _on_save(self):
        self.result = self.bundle_map
        self.destroy()


class _ComponentEntryDialog(tk.Toplevel):
    """Small dialog for entering a component SKU + quantity."""

    def __init__(self, parent, sku="", qty=1):
        super().__init__(parent)
        self.title("Component")
        self.configure(bg=_BG)
        self.transient(parent)
        self.grab_set()
        self.result = None

        ttk.Label(self, text="Component SKU:").pack(
            anchor="w", padx=15, pady=(15, 3))
        self.sku_var = tk.StringVar(value=sku)
        ttk.Entry(self, textvariable=self.sku_var, width=30).pack(
            padx=15, fill="x")

        ttk.Label(self, text="Qty per bundle:").pack(
            anchor="w", padx=15, pady=(10, 3))
        self.qty_var = tk.StringVar(value=str(qty))
        ttk.Entry(self, textvariable=self.qty_var, width=10).pack(
            anchor="w", padx=15)

        btn_frame = ttk.Frame(self)
        btn_frame.pack(fill="x", padx=15, pady=15)
        tk.Button(btn_frame, text="Cancel", command=self.destroy,
                  bg=_BG3, fg=_FG, relief="flat", padx=10, pady=4
                  ).pack(side="right", padx=(5, 0))
        tk.Button(btn_frame, text="OK", command=self._on_ok,
                  bg=_GREEN, fg="white", relief="flat", padx=10, pady=4
                  ).pack(side="right")

        self.geometry("320x200")

    def _on_ok(self):
        sku = self.sku_var.get().strip()
        if not sku:
            messagebox.showerror("Error", "SKU cannot be empty.", parent=self)
            return
        try:
            qty = float(self.qty_var.get())
            if qty <= 0:
                raise ValueError
        except ValueError:
            messagebox.showerror("Error", "Qty must be a positive number.",
                                 parent=self)
            return
        self.result = (sku, qty)
        self.destroy()


# ═════════════════════════════════════════════════════════════════════
#  SKU SETTINGS EDITOR DIALOG
# ═════════════════════════════════════════════════════════════════════

class SkuSettingsDialog(tk.Toplevel):
    """Edit per-SKU lead times and safety stock."""

    def __init__(self, parent, sku, sku_settings):
        super().__init__(parent)
        self.title(f"Settings — {sku}")
        self.configure(bg=_BG)
        self.transient(parent)
        self.grab_set()
        self.result = None

        s = sku_settings or {}

        fields = [
            ("Purchase Lead Time (days):", "purchase_lt",
             s.get("purchase_lt", "")),
            ("Production Lead Time (days):", "production_lt",
             s.get("production_lt", "")),
            ("Shipping Lead Time (days):", "shipping_lt",
             s.get("shipping_lt", "")),
            ("Safety Stock (units):", "safety_stock",
             s.get("safety_stock", "")),
            ("Churn Rate (%):", "churn_pct", s.get("churn_pct", "")),
        ]

        self.vars = {}
        frame = ttk.Frame(self)
        frame.pack(fill="both", expand=True, padx=15, pady=15)

        for i, (label, key, default) in enumerate(fields):
            ttk.Label(frame, text=label).grid(
                row=i, column=0, sticky="w", padx=(0, 10), pady=5)
            var = tk.StringVar(value=str(default))
            ttk.Entry(frame, textvariable=var, width=12).grid(
                row=i, column=1, pady=5, sticky="w")
            self.vars[key] = var

        ttk.Label(frame,
                  text="Leave blank to use global defaults.",
                  style="Dim.TLabel"
                  ).grid(row=len(fields), column=0, columnspan=2,
                         sticky="w", pady=(10, 0))

        btn_frame = ttk.Frame(self)
        btn_frame.pack(fill="x", padx=15, pady=(0, 15))
        tk.Button(btn_frame, text="Cancel", command=self.destroy,
                  bg=_BG3, fg=_FG, relief="flat", padx=15, pady=5
                  ).pack(side="right", padx=(5, 0))
        tk.Button(btn_frame, text="Save", command=self._on_save,
                  bg=_GREEN, fg="white", relief="flat", padx=15, pady=5
                  ).pack(side="right")

        self.geometry("380x300")

    def _on_save(self):
        result = {}
        for key, var in self.vars.items():
            val = var.get().strip()
            if val:
                try:
                    result[key] = float(val)
                except ValueError:
                    messagebox.showerror(
                        "Invalid Input",
                        f"'{val}' is not a valid number for {key}.",
                        parent=self)
                    return
        self.result = result
        self.destroy()


# ═════════════════════════════════════════════════════════════════════
#  MANUAL DEMAND EDITOR DIALOG
# ═════════════════════════════════════════════════════════════════════

class ManualDemandDialog(tk.Toplevel):
    """Edit manual weekly demand adjustments per SKU."""

    def __init__(self, parent, manual_demand):
        super().__init__(parent)
        self.title("Manual Demand Adjustments")
        self.configure(bg=_BG)
        self.transient(parent)
        self.grab_set()
        self.result = None

        self.entries = {}
        # deep copy
        self.data = dict(manual_demand)

        top = ttk.Frame(self)
        top.pack(fill="x", padx=15, pady=(15, 5))
        ttk.Label(top, text="Manual weekly demand adjustments per SKU.",
                  style="Bold.TLabel").pack(anchor="w")
        ttk.Label(top, text="These are added to Recharge + Shopify forecasts.",
                  style="Dim.TLabel").pack(anchor="w")

        # treeview
        tree_frame = ttk.Frame(self)
        tree_frame.pack(fill="both", expand=True, padx=15, pady=5)

        self.tree = ttk.Treeview(
            tree_frame, columns=("sku", "weekly_qty"), show="headings",
            height=12)
        self.tree.heading("sku", text="SKU")
        self.tree.heading("weekly_qty", text="Weekly Qty")
        self.tree.column("sku", width=250)
        self.tree.column("weekly_qty", width=100, anchor="center")
        yscroll = ttk.Scrollbar(tree_frame, orient="vertical",
                                command=self.tree.yview)
        self.tree.configure(yscrollcommand=yscroll.set)
        self.tree.pack(side="left", fill="both", expand=True)
        yscroll.pack(side="right", fill="y")

        btn_row = ttk.Frame(self)
        btn_row.pack(fill="x", padx=15, pady=5)
        tk.Button(btn_row, text="+ Add SKU", command=self._add_entry,
                  bg=_GREEN, fg="white", relief="flat", padx=8, pady=3
                  ).pack(side="left", padx=(0, 3))
        tk.Button(btn_row, text="- Remove", command=self._remove_entry,
                  bg=_RED, fg="white", relief="flat", padx=8, pady=3
                  ).pack(side="left", padx=(0, 3))
        tk.Button(btn_row, text="Edit", command=self._edit_entry,
                  bg=_BG3, fg=_FG, relief="flat", padx=8, pady=3
                  ).pack(side="left")

        bottom = ttk.Frame(self)
        bottom.pack(fill="x", padx=15, pady=(5, 15))
        tk.Button(bottom, text="Cancel", command=self.destroy,
                  bg=_BG3, fg=_FG, relief="flat", padx=15, pady=5
                  ).pack(side="right", padx=(5, 0))
        tk.Button(bottom, text="Save", command=self._on_save,
                  bg=_GREEN, fg="white", relief="flat", padx=15, pady=5
                  ).pack(side="right")

        self._refresh()
        self.geometry("450x480")

    def _refresh(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
        for sku in sorted(self.data.keys()):
            self.tree.insert("", "end", values=(sku, self.data[sku]))

    def _add_entry(self):
        sku = simpledialog.askstring("Add SKU", "SKU:", parent=self)
        if not sku or not sku.strip():
            return
        sku = sku.strip()
        qty_str = simpledialog.askstring("Weekly Qty",
                                         f"Weekly demand for {sku}:",
                                         parent=self)
        if qty_str is None:
            return
        try:
            qty = float(qty_str)
        except ValueError:
            messagebox.showerror("Error", "Invalid number.", parent=self)
            return
        self.data[sku] = qty
        self._refresh()

    def _remove_entry(self):
        sel = self.tree.selection()
        if not sel:
            return
        sku = self.tree.item(sel[0], "values")[0]
        if sku in self.data:
            del self.data[sku]
        self._refresh()

    def _edit_entry(self):
        sel = self.tree.selection()
        if not sel:
            return
        sku, old_qty = self.tree.item(sel[0], "values")
        qty_str = simpledialog.askstring(
            "Edit Weekly Qty", f"Weekly demand for {sku}:",
            initialvalue=old_qty, parent=self)
        if qty_str is None:
            return
        try:
            qty = float(qty_str)
        except ValueError:
            messagebox.showerror("Error", "Invalid number.", parent=self)
            return
        self.data[sku] = qty
        self._refresh()

    def _on_save(self):
        self.result = self.data
        self.destroy()


# ═════════════════════════════════════════════════════════════════════
#  SHOPIFY FORECAST EDITOR DIALOG
# ═════════════════════════════════════════════════════════════════════

class ShopifyForecastDialog(tk.Toplevel):
    """Edit weekly forecast for Shopify first-time order bundles."""

    def __init__(self, parent, shopify_forecast):
        super().__init__(parent)
        self.title("Shopify First-Time Order Forecast")
        self.configure(bg=_BG)
        self.transient(parent)
        self.grab_set()
        self.result = None
        self.data = dict(shopify_forecast)

        top = ttk.Frame(self)
        top.pack(fill="x", padx=15, pady=(15, 5))
        ttk.Label(top,
                  text="Weekly forecast for Shopify first-time order bundles.",
                  style="Bold.TLabel").pack(anchor="w")
        ttk.Label(top,
                  text="These are bundles — they'll be decomposed into "
                       "component SKUs.",
                  style="Dim.TLabel").pack(anchor="w")

        tree_frame = ttk.Frame(self)
        tree_frame.pack(fill="both", expand=True, padx=15, pady=5)

        self.tree = ttk.Treeview(
            tree_frame, columns=("sku", "weekly_qty"), show="headings",
            height=12)
        self.tree.heading("sku", text="Bundle / SKU")
        self.tree.heading("weekly_qty", text="Weekly Qty")
        self.tree.column("sku", width=250)
        self.tree.column("weekly_qty", width=100, anchor="center")
        yscroll = ttk.Scrollbar(tree_frame, orient="vertical",
                                command=self.tree.yview)
        self.tree.configure(yscrollcommand=yscroll.set)
        self.tree.pack(side="left", fill="both", expand=True)
        yscroll.pack(side="right", fill="y")

        btn_row = ttk.Frame(self)
        btn_row.pack(fill="x", padx=15, pady=5)
        tk.Button(btn_row, text="+ Add", command=self._add,
                  bg=_GREEN, fg="white", relief="flat", padx=8, pady=3
                  ).pack(side="left", padx=(0, 3))
        tk.Button(btn_row, text="- Remove", command=self._remove,
                  bg=_RED, fg="white", relief="flat", padx=8, pady=3
                  ).pack(side="left", padx=(0, 3))
        tk.Button(btn_row, text="Edit", command=self._edit,
                  bg=_BG3, fg=_FG, relief="flat", padx=8, pady=3
                  ).pack(side="left")

        bottom = ttk.Frame(self)
        bottom.pack(fill="x", padx=15, pady=(5, 15))
        tk.Button(bottom, text="Cancel", command=self.destroy,
                  bg=_BG3, fg=_FG, relief="flat", padx=15, pady=5
                  ).pack(side="right", padx=(5, 0))
        tk.Button(bottom, text="Save", command=self._on_save,
                  bg=_GREEN, fg="white", relief="flat", padx=15, pady=5
                  ).pack(side="right")

        self._refresh()
        self.geometry("450x480")

    def _refresh(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
        for sku in sorted(self.data.keys()):
            self.tree.insert("", "end", values=(sku, self.data[sku]))

    def _add(self):
        sku = simpledialog.askstring("Add Bundle/SKU", "Bundle or SKU:",
                                     parent=self)
        if not sku or not sku.strip():
            return
        sku = sku.strip()
        qty_str = simpledialog.askstring("Weekly Qty",
                                         f"Weekly forecast for {sku}:",
                                         parent=self)
        if qty_str is None:
            return
        try:
            qty = float(qty_str)
        except ValueError:
            messagebox.showerror("Error", "Invalid number.", parent=self)
            return
        self.data[sku] = qty
        self._refresh()

    def _remove(self):
        sel = self.tree.selection()
        if not sel:
            return
        sku = self.tree.item(sel[0], "values")[0]
        if sku in self.data:
            del self.data[sku]
        self._refresh()

    def _edit(self):
        sel = self.tree.selection()
        if not sel:
            return
        sku, old_qty = self.tree.item(sel[0], "values")
        qty_str = simpledialog.askstring(
            "Edit Weekly Qty", f"Weekly forecast for {sku}:",
            initialvalue=old_qty, parent=self)
        if qty_str is None:
            return
        try:
            qty = float(qty_str)
        except ValueError:
            messagebox.showerror("Error", "Invalid number.", parent=self)
            return
        self.data[sku] = qty
        self._refresh()

    def _on_save(self):
        self.result = self.data
        self.destroy()


# ═════════════════════════════════════════════════════════════════════
#  CURATION RECIPE EDITOR DIALOG
# ═════════════════════════════════════════════════════════════════════

class CurationRecipeDialog(tk.Toplevel):
    """Editor for curation recipes, PR-CJAM, and CEX-EC assignments."""

    def __init__(self, parent, curation_recipes, pr_cjam, cex_ec, inventory_skus=None):
        super().__init__(parent)
        self.title("Curation Recipe Editor")
        self.configure(bg=_BG)
        self.transient(parent)
        self.grab_set()
        self.result = None

        # deep copy
        self.recipes = {k: [list(c) for c in v] for k, v in curation_recipes.items()}
        self.pr_cjam = {k: dict(v) if isinstance(v, dict) else {"cheese": v, "jam": ""}
                        for k, v in pr_cjam.items()}
        self.cex_ec = dict(cex_ec)
        self.inventory_skus = set(inventory_skus or [])

        # ── left: curation list ──
        left = ttk.Frame(self)
        left.pack(side="left", fill="y", padx=(15, 5), pady=15)

        ttk.Label(left, text="Curations", style="Bold.TLabel").pack(anchor="w")
        self.cur_listbox = tk.Listbox(
            left, bg=_BG2, fg=_FG, selectbackground=_ACC,
            width=18, height=18, relief="flat", bd=0)
        self.cur_listbox.pack(fill="y", expand=True, pady=(5, 5))
        self.cur_listbox.bind("<<ListboxSelect>>", self._on_select)

        btn_row = ttk.Frame(left)
        btn_row.pack(fill="x")
        tk.Button(btn_row, text="+ Add", command=self._add_curation,
                  bg=_GREEN, fg="white", relief="flat", padx=8, pady=3
                  ).pack(side="left", padx=(0, 3))
        tk.Button(btn_row, text="- Remove", command=self._remove_curation,
                  bg=_RED, fg="white", relief="flat", padx=8, pady=3
                  ).pack(side="left")

        # ── right: recipe + assignments ──
        right = ttk.Frame(self)
        right.pack(side="left", fill="both", expand=True, padx=(5, 15), pady=15)

        # assignments row
        assign_frame = ttk.LabelFrame(right, text="Assignments", padding=5)
        assign_frame.pack(fill="x", pady=(0, 5))

        ttk.Label(assign_frame, text="PR-CJAM Cheese:").grid(
            row=0, column=0, sticky="w", padx=(0, 5))
        self.pr_cjam_cheese_var = tk.StringVar()
        ttk.Entry(assign_frame, textvariable=self.pr_cjam_cheese_var, width=15).grid(
            row=0, column=1, padx=(0, 10))

        ttk.Label(assign_frame, text="PR-CJAM Jam:").grid(
            row=0, column=2, sticky="w", padx=(0, 5))
        self.pr_cjam_jam_var = tk.StringVar()
        ttk.Entry(assign_frame, textvariable=self.pr_cjam_jam_var, width=15).grid(
            row=0, column=3, padx=(0, 15))

        ttk.Label(assign_frame, text="CEX-EC:").grid(
            row=0, column=4, sticky="w", padx=(0, 5))
        self.cex_ec_var = tk.StringVar()
        ttk.Entry(assign_frame, textvariable=self.cex_ec_var, width=15).grid(
            row=0, column=5)

        # recipe treeview
        ttk.Label(right, text="Recipe Components", style="Bold.TLabel"
                  ).pack(anchor="w")
        self.recipe_tree = ttk.Treeview(
            right, columns=("sku", "qty"), show="headings", height=10)
        self.recipe_tree.heading("sku", text="SKU")
        self.recipe_tree.heading("qty", text="Qty")
        self.recipe_tree.column("sku", width=200)
        self.recipe_tree.column("qty", width=60, anchor="center")
        self.recipe_tree.pack(fill="both", expand=True, pady=(5, 5))

        comp_btns = ttk.Frame(right)
        comp_btns.pack(fill="x")
        tk.Button(comp_btns, text="+ Add", command=self._add_item,
                  bg=_GREEN, fg="white", relief="flat", padx=8, pady=3
                  ).pack(side="left", padx=(0, 3))
        tk.Button(comp_btns, text="- Remove", command=self._remove_item,
                  bg=_RED, fg="white", relief="flat", padx=8, pady=3
                  ).pack(side="left", padx=(0, 3))
        tk.Button(comp_btns, text="Edit", command=self._edit_item,
                  bg=_BG3, fg=_FG, relief="flat", padx=8, pady=3
                  ).pack(side="left")

        # duplicate detection panel
        dup_frame = ttk.LabelFrame(right, text="Duplicate Detection", padding=5)
        dup_frame.pack(fill="x", pady=(5, 0))
        self.dup_label = ttk.Label(dup_frame, text="Select a curation to check.",
                                   style="Dim.TLabel", wraplength=400)
        self.dup_label.pack(anchor="w")

        # ── bottom buttons ──
        bottom = ttk.Frame(self)
        bottom.pack(fill="x", padx=15, pady=(0, 15))
        tk.Button(bottom, text="Cancel", command=self.destroy,
                  bg=_BG3, fg=_FG, relief="flat", padx=15, pady=5
                  ).pack(side="right", padx=(5, 0))
        tk.Button(bottom, text="Save", command=self._on_save,
                  bg=_GREEN, fg="white", relief="flat", padx=15, pady=5
                  ).pack(side="right")

        self._refresh_list()
        self.geometry("850x600")
        self.minsize(750, 450)

    def _refresh_list(self):
        self.cur_listbox.delete(0, "end")
        all_curations = list(CURATION_ORDER)
        for c in sorted(self.recipes.keys()):
            if c not in all_curations:
                all_curations.append(c)
        for c in all_curations:
            self.cur_listbox.insert("end", c)

    def _on_select(self, event=None):
        sel = self.cur_listbox.curselection()
        if not sel:
            return
        cur = self.cur_listbox.get(sel[0])
        self._save_current_assignments()
        self._current_curation = cur

        # load assignments
        pr_entry = self.pr_cjam.get(cur, {})
        if isinstance(pr_entry, str):
            self.pr_cjam_cheese_var.set(pr_entry)
            self.pr_cjam_jam_var.set("")
        else:
            self.pr_cjam_cheese_var.set(pr_entry.get("cheese", ""))
            self.pr_cjam_jam_var.set(pr_entry.get("jam", ""))
        self.cex_ec_var.set(self.cex_ec.get(cur, ""))

        # load recipe
        for item in self.recipe_tree.get_children():
            self.recipe_tree.delete(item)
        for sku, qty in self.recipes.get(cur, []):
            self.recipe_tree.insert("", "end", values=(sku, qty))

        self._check_duplicates(cur)

    def _save_current_assignments(self):
        cur = getattr(self, "_current_curation", None)
        if not cur:
            return
        pr_cheese = self.pr_cjam_cheese_var.get().strip()
        pr_jam = self.pr_cjam_jam_var.get().strip()
        ec = self.cex_ec_var.get().strip()
        if pr_cheese or pr_jam:
            self.pr_cjam[cur] = {"cheese": pr_cheese, "jam": pr_jam}
        else:
            self.pr_cjam.pop(cur, None)
        if ec:
            self.cex_ec[cur] = ec
        else:
            self.cex_ec.pop(cur, None)

    def _check_duplicates(self, curation):
        """Check if PR-CJAM cheese appears in any curation recipe.
        Only checks cheese for duplicates — jam/accompaniment overlap is OK."""
        pr_entry = self.pr_cjam.get(curation, {})
        if isinstance(pr_entry, str):
            pr_cheese = pr_entry
        else:
            pr_cheese = pr_entry.get("cheese", "")
        if not pr_cheese:
            self.dup_label.configure(text="No PR-CJAM cheese assigned.", foreground=_FG2)
            return

        conflicts = []
        for cur, recipe in self.recipes.items():
            recipe_skus = [s for s, q in recipe]
            if pr_cheese in recipe_skus:
                conflicts.append(cur)

        # check other PR-CJAM cheese assignments for uniqueness
        dup_curations = []
        for cur, entry in self.pr_cjam.items():
            if cur == curation:
                continue
            other_cheese = entry.get("cheese", "") if isinstance(entry, dict) else entry
            if other_cheese == pr_cheese:
                dup_curations.append(cur)

        msgs = []
        if conflicts:
            msgs.append(f"WARNING: {pr_cheese} also in recipe for: {', '.join(conflicts)}")
        if dup_curations:
            msgs.append(f"WARNING: {pr_cheese} also PR-CJAM for: {', '.join(dup_curations)}")

        if msgs:
            self.dup_label.configure(text="\n".join(msgs), foreground=_RED)
        else:
            self.dup_label.configure(text="No duplicates found.", foreground=_GREEN)

    def _get_selected_curation(self):
        sel = self.cur_listbox.curselection()
        if not sel:
            messagebox.showinfo("Select", "Select a curation first.", parent=self)
            return None
        return self.cur_listbox.get(sel[0])

    def _add_curation(self):
        name = simpledialog.askstring("New Curation", "Curation name:", parent=self)
        if name and name.strip():
            name = name.strip().upper()
            if name not in self.recipes:
                self.recipes[name] = []
            self._refresh_list()

    def _remove_curation(self):
        cur = self._get_selected_curation()
        if cur and messagebox.askyesno("Confirm", f"Remove '{cur}'?", parent=self):
            self.recipes.pop(cur, None)
            self.pr_cjam.pop(cur, None)
            self.cex_ec.pop(cur, None)
            self._refresh_list()
            for item in self.recipe_tree.get_children():
                self.recipe_tree.delete(item)

    def _add_item(self):
        cur = self._get_selected_curation()
        if not cur:
            return
        dlg = _ComponentEntryDialog(self)
        self.wait_window(dlg)
        if dlg.result:
            if cur not in self.recipes:
                self.recipes[cur] = []
            self.recipes[cur].append(list(dlg.result))
            self._on_select()

    def _remove_item(self):
        cur = self._get_selected_curation()
        if not cur:
            return
        sel = self.recipe_tree.selection()
        if not sel:
            return
        idx = self.recipe_tree.index(sel[0])
        del self.recipes[cur][idx]
        self._on_select()

    def _edit_item(self):
        cur = self._get_selected_curation()
        if not cur:
            return
        sel = self.recipe_tree.selection()
        if not sel:
            return
        idx = self.recipe_tree.index(sel[0])
        old_sku, old_qty = self.recipes[cur][idx]
        dlg = _ComponentEntryDialog(self, old_sku, old_qty)
        self.wait_window(dlg)
        if dlg.result:
            self.recipes[cur][idx] = list(dlg.result)
            self._on_select()

    def _on_save(self):
        self._save_current_assignments()
        self.result = {
            "recipes": {k: [tuple(c) for c in v] for k, v in self.recipes.items()},
            "pr_cjam": self.pr_cjam,
            "cex_ec": self.cex_ec,
        }
        self.destroy()


# ═════════════════════════════════════════════════════════════════════
#  MONTHLY BOX RECIPE DIALOG (MED / CMED / LGE)
# ═════════════════════════════════════════════════════════════════════

class MonthlyBoxRecipeDialog(tk.Toplevel):
    """Editor for monthly curated box recipes (AHB-MED, AHB-CMED, AHB-LGE)."""

    def __init__(self, parent, monthly_box_recipes, monthly_box_counts,
                 inventory_skus=None):
        super().__init__(parent)
        self.title("Monthly Box Recipe Editor")
        self.configure(bg=_BG)
        self.transient(parent)
        self.grab_set()
        self.result = None

        # deep copy data
        self.recipes = {}
        for month, types in monthly_box_recipes.items():
            self.recipes[month] = {}
            for btype, items in types.items():
                self.recipes[month][btype] = [list(r) for r in items]
        self.box_counts = dict(monthly_box_counts)
        self.inventory_skus = sorted(inventory_skus or [])
        self._ignored_overlaps = {}  # {(month, box_type, sku): True}

        # ── month list ──
        today = datetime.date.today()
        self._months = []
        for offset in range(-2, 3):
            m = today.month + offset
            y = today.year
            while m < 1:
                m += 12
                y -= 1
            while m > 12:
                m -= 12
                y += 1
            self._months.append(f"{y:04d}-{m:02d}")

        # ── top controls ──
        top = ttk.Frame(self)
        top.pack(fill="x", padx=15, pady=(15, 5))

        ttk.Label(top, text="Month:").pack(side="left", padx=(0, 5))
        self.month_var = tk.StringVar(value=self._months[2])  # current month
        month_cb = ttk.Combobox(top, textvariable=self.month_var,
                                values=self._months, state="readonly", width=10)
        month_cb.pack(side="left", padx=(0, 15))
        month_cb.bind("<<ComboboxSelected>>", lambda e: self._on_selection_change())

        ttk.Label(top, text="Box Type:").pack(side="left", padx=(0, 5))
        self.boxtype_var = tk.StringVar(value=MONTHLY_BOX_TYPES[0])
        bt_cb = ttk.Combobox(top, textvariable=self.boxtype_var,
                             values=MONTHLY_BOX_TYPES, state="readonly", width=12)
        bt_cb.pack(side="left", padx=(0, 15))
        bt_cb.bind("<<ComboboxSelected>>", lambda e: self._on_selection_change())

        ttk.Button(top, text="Copy from MED",
                   command=self._copy_from_med).pack(side="left", padx=(0, 5))

        # ── box count row ──
        count_frame = ttk.Frame(self)
        count_frame.pack(fill="x", padx=15, pady=(0, 5))

        ttk.Label(count_frame, text="Box Count:").pack(side="left", padx=(0, 5))
        self.count_var = tk.StringVar(value="0")
        ttk.Entry(count_frame, textvariable=self.count_var, width=8).pack(
            side="left", padx=(0, 10))
        self.count_source_var = tk.StringVar(value="")
        ttk.Label(count_frame, textvariable=self.count_source_var,
                  style="Dim.TLabel").pack(side="left")

        # ── recipe treeview ──
        tree_frame = ttk.LabelFrame(self, text="Recipe Slots", padding=5)
        tree_frame.pack(fill="both", expand=True, padx=15, pady=(0, 5))

        cols = ("slot", "prefix", "sku", "qty")
        self.recipe_tree = ttk.Treeview(
            tree_frame, columns=cols, show="headings", height=12)
        self.recipe_tree.heading("slot", text="Slot")
        self.recipe_tree.heading("prefix", text="Expected Prefix")
        self.recipe_tree.heading("sku", text="SKU")
        self.recipe_tree.heading("qty", text="Qty")
        self.recipe_tree.column("slot", width=160)
        self.recipe_tree.column("prefix", width=120, anchor="center")
        self.recipe_tree.column("sku", width=150)
        self.recipe_tree.column("qty", width=50, anchor="center")

        yscroll = ttk.Scrollbar(tree_frame, orient="vertical",
                                command=self.recipe_tree.yview)
        self.recipe_tree.configure(yscrollcommand=yscroll.set)
        self.recipe_tree.pack(side="left", fill="both", expand=True)
        yscroll.pack(side="right", fill="y")

        self.recipe_tree.tag_configure(
            "warn", background="#7a5500", foreground="white")
        self.recipe_tree.tag_configure(
            "ok", background=_BG2, foreground=_FG)

        self.recipe_tree.bind("<Double-1>", self._edit_sku)

        # edit button row
        edit_row = ttk.Frame(self)
        edit_row.pack(fill="x", padx=15, pady=(0, 5))
        tk.Button(edit_row, text="Set SKU", command=self._edit_sku,
                  bg=_BG3, fg=_FG, relief="flat", padx=8, pady=3
                  ).pack(side="left", padx=(0, 5))
        tk.Button(edit_row, text="Clear SKU", command=self._clear_sku,
                  bg=_BG3, fg=_FG, relief="flat", padx=8, pady=3
                  ).pack(side="left")

        # ── overlap detection panel ──
        overlap_frame = ttk.LabelFrame(self, text="Overlap Detection", padding=5)
        overlap_frame.pack(fill="both", padx=15, pady=(0, 5))

        self.overlap_text = tk.Text(
            overlap_frame, bg=_BG2, fg=_FG, height=5, wrap="word",
            relief="flat", bd=0, font=("Segoe UI", 9))
        self.overlap_text.pack(fill="both", expand=True)
        self.overlap_text.tag_configure("red", foreground="#ff6b6b")
        self.overlap_text.tag_configure("green", foreground="#6bff6b")
        self.overlap_text.tag_configure("blue", foreground="#6baaff")
        self.overlap_text.configure(state="disabled")

        # ── bottom buttons ──
        bottom = ttk.Frame(self)
        bottom.pack(fill="x", padx=15, pady=(0, 15))
        tk.Button(bottom, text="Cancel", command=self.destroy,
                  bg=_BG3, fg=_FG, relief="flat", padx=15, pady=5
                  ).pack(side="right", padx=(5, 0))
        tk.Button(bottom, text="Save", command=self._on_save,
                  bg=_GREEN, fg="white", relief="flat", padx=15, pady=5
                  ).pack(side="right")

        self.geometry("780x700")
        self.minsize(650, 550)

        # initial load
        self._on_selection_change()

    # ── helpers ──────────────────────────────────────────────────────

    def _current_month(self):
        return self.month_var.get()

    def _current_box_type(self):
        return self.boxtype_var.get()

    def _get_recipe(self, month, box_type):
        """Get recipe list for month/box_type, creating from template if missing."""
        if month not in self.recipes:
            self.recipes[month] = {}
        if box_type not in self.recipes[month]:
            # create from slot template
            template = MONTHLY_BOX_SLOTS.get(box_type, [])
            self.recipes[month][box_type] = [
                [slot, "", 1] for slot, prefix in template
            ]
        return self.recipes[month][box_type]

    def _on_selection_change(self):
        """Reload treeview when month or box type changes."""
        self._save_current_count()
        month = self._current_month()
        box_type = self._current_box_type()
        recipe = self._get_recipe(month, box_type)
        template = MONTHLY_BOX_SLOTS.get(box_type, [])

        # update count
        count = self.box_counts.get(box_type, 0)
        self.count_var.set(str(count))
        self.count_source_var.set("(from Recharge or manual)")

        # populate tree
        for item in self.recipe_tree.get_children():
            self.recipe_tree.delete(item)

        for i, (slot_name, prefix) in enumerate(template):
            if i < len(recipe):
                sku = recipe[i][1]
                qty = recipe[i][2]
            else:
                sku = ""
                qty = 1
            # validate prefix
            tag = "ok"
            if sku and not sku.upper().startswith(prefix) and \
               not sku.upper().startswith("EX-EA") and \
               not sku.upper().startswith("CEX-EA"):
                tag = "warn"
            self.recipe_tree.insert("", "end", values=(
                slot_name, prefix, sku, qty), tags=(tag,))

        self._run_overlap_detection()

    def _save_current_count(self):
        """Persist the box count entry before switching."""
        box_type = self._current_box_type()
        try:
            self.box_counts[box_type] = int(self.count_var.get())
        except ValueError:
            pass

    def _save_tree_to_recipe(self):
        """Write current treeview back into recipe data."""
        month = self._current_month()
        box_type = self._current_box_type()
        recipe = self._get_recipe(month, box_type)
        template = MONTHLY_BOX_SLOTS.get(box_type, [])

        children = self.recipe_tree.get_children()
        for i, item_id in enumerate(children):
            vals = self.recipe_tree.item(item_id, "values")
            sku = vals[2] if len(vals) > 2 else ""
            qty = 1
            try:
                qty = int(vals[3]) if len(vals) > 3 else 1
            except (ValueError, IndexError):
                qty = 1
            if i < len(recipe):
                recipe[i] = [template[i][0] if i < len(template) else "", sku, qty]
            else:
                slot = template[i][0] if i < len(template) else ""
                recipe.append([slot, sku, qty])

    def _edit_sku(self, event=None):
        """Edit the SKU for the selected slot."""
        sel = self.recipe_tree.selection()
        if not sel:
            sel_items = self.recipe_tree.get_children()
            if not sel_items:
                return
            # use focused item on double-click
            if event:
                item = self.recipe_tree.identify_row(event.y)
                if item:
                    sel = (item,)
                else:
                    return
            else:
                return

        item = sel[0]
        vals = self.recipe_tree.item(item, "values")
        slot_name = vals[0]
        prefix = vals[1]
        old_sku = vals[2]

        new_sku = simpledialog.askstring(
            "Set SKU",
            f"Slot: {slot_name}  (expected: {prefix})\n\nSKU:",
            initialvalue=old_sku, parent=self)
        if new_sku is None:
            return
        new_sku = new_sku.strip().upper()

        # update tree
        tag = "ok"
        if new_sku and not new_sku.startswith(prefix) and \
           not new_sku.startswith("EX-EA") and \
           not new_sku.startswith("CEX-EA"):
            tag = "warn"
            messagebox.showwarning(
                "Prefix Mismatch",
                f"'{new_sku}' does not start with expected prefix '{prefix}'.\n"
                "It has been set anyway — verify this is correct.",
                parent=self)

        self.recipe_tree.item(item, values=(slot_name, prefix, new_sku, vals[3]),
                              tags=(tag,))
        self._save_tree_to_recipe()
        self._run_overlap_detection()

    def _clear_sku(self):
        sel = self.recipe_tree.selection()
        if not sel:
            return
        item = sel[0]
        vals = self.recipe_tree.item(item, "values")
        self.recipe_tree.item(item, values=(vals[0], vals[1], "", vals[3]),
                              tags=("ok",))
        self._save_tree_to_recipe()
        self._run_overlap_detection()

    def _copy_from_med(self):
        """Copy AHB-MED recipe into current box type (useful for LGE)."""
        month = self._current_month()
        box_type = self._current_box_type()
        med_recipe = self._get_recipe(month, "AHB-MED")
        target = self._get_recipe(month, box_type)
        template = MONTHLY_BOX_SLOTS.get(box_type, [])
        med_template = MONTHLY_BOX_SLOTS.get("AHB-MED", [])

        # copy matching slots from MED
        med_by_slot = {}
        for i, (slot, prefix) in enumerate(med_template):
            if i < len(med_recipe):
                med_by_slot[slot] = med_recipe[i][1]

        for i, (slot, prefix) in enumerate(template):
            if slot in med_by_slot and i < len(target):
                target[i][1] = med_by_slot[slot]

        self._on_selection_change()

    def _run_overlap_detection(self):
        """Check current recipe for SKU overlaps with past 2 months."""
        month = self._current_month()
        box_type = self._current_box_type()
        recipe = self._get_recipe(month, box_type)

        # gather items from past 2 months
        previous_items = {}  # {sku: [(month, box_type), ...]}
        month_idx = self._months.index(month) if month in self._months else -1

        for past_offset in [1, 2]:
            past_idx = month_idx - past_offset
            if past_idx < 0 or past_idx >= len(self._months):
                continue
            past_month = self._months[past_idx]
            for bt in MONTHLY_BOX_TYPES:
                past_recipe = self.recipes.get(past_month, {}).get(bt, [])
                for item in past_recipe:
                    sku = item[1] if len(item) > 1 else ""
                    if sku:
                        previous_items.setdefault(sku, []).append(
                            (past_month, bt))

        # find overlaps in current recipe
        overlaps = []
        current_skus = set()
        for item in recipe:
            sku = item[1] if len(item) > 1 else ""
            slot = item[0] if len(item) > 0 else ""
            if sku:
                current_skus.add(sku)
                if sku in previous_items:
                    overlaps.append((sku, slot, previous_items[sku]))

        # suggest unused CH- SKUs
        all_used = set(previous_items.keys()) | current_skus
        suggestions = [s for s in self.inventory_skus
                       if s.startswith("CH-") and s not in all_used]

        # update text widget
        self.overlap_text.configure(state="normal")
        self.overlap_text.delete("1.0", "end")

        if overlaps:
            self.overlap_text.insert("end", "OVERLAPS FOUND:\n", "red")
            for sku, slot, sources in overlaps:
                for src_month, src_type in sources:
                    self.overlap_text.insert(
                        "end",
                        f"  {sku} ({slot}) was in {src_type} {src_month}\n",
                        "red")
        else:
            self.overlap_text.insert("end", "No overlaps with past 2 months.\n",
                                     "green")

        if suggestions[:10]:
            self.overlap_text.insert("end", "\nUnused CH- suggestions: ", "blue")
            self.overlap_text.insert("end", ", ".join(suggestions[:10]) + "\n",
                                     "blue")

        self.overlap_text.configure(state="disabled")

    def _on_save(self):
        self._save_tree_to_recipe()
        self._save_current_count()
        self.result = {
            "recipes": self.recipes,
            "box_counts": self.box_counts,
        }
        self.destroy()


# ═════════════════════════════════════════════════════════════════════
#  COHORT MANAGER DIALOG
# ═════════════════════════════════════════════════════════════════════

class CohortManagerDialog(tk.Toplevel):
    """Manage subscriber cohorts for forecasting."""

    def __init__(self, parent, cohorts):
        super().__init__(parent)
        self.title("Cohort Manager")
        self.configure(bg=_BG)
        self.transient(parent)
        self.grab_set()
        self.result = None

        # deep copy
        self.cohorts = [dict(c) for c in cohorts]

        top = ttk.Frame(self)
        top.pack(fill="x", padx=15, pady=(15, 5))
        ttk.Label(top, text="Subscriber cohorts by start month.",
                  style="Bold.TLabel").pack(anchor="w")
        ttk.Label(top, text="Each cohort tracks subscribers who started in that month.",
                  style="Dim.TLabel").pack(anchor="w")

        # treeview
        tree_frame = ttk.Frame(self)
        tree_frame.pack(fill="both", expand=True, padx=15, pady=5)

        self.tree = ttk.Treeview(
            tree_frame,
            columns=("start_month", "size", "track", "age"),
            show="headings", height=14)
        self.tree.heading("start_month", text="Start Month")
        self.tree.heading("size", text="Cohort Size")
        self.tree.heading("track", text="Track")
        self.tree.heading("age", text="Current Age (mo)")
        self.tree.column("start_month", width=120)
        self.tree.column("size", width=100, anchor="center")
        self.tree.column("track", width=80, anchor="center")
        self.tree.column("age", width=120, anchor="center")
        yscroll = ttk.Scrollbar(tree_frame, orient="vertical",
                                command=self.tree.yview)
        self.tree.configure(yscrollcommand=yscroll.set)
        self.tree.pack(side="left", fill="both", expand=True)
        yscroll.pack(side="right", fill="y")

        btn_row = ttk.Frame(self)
        btn_row.pack(fill="x", padx=15, pady=5)
        tk.Button(btn_row, text="+ Add Cohort", command=self._add,
                  bg=_GREEN, fg="white", relief="flat", padx=8, pady=3
                  ).pack(side="left", padx=(0, 3))
        tk.Button(btn_row, text="- Remove", command=self._remove,
                  bg=_RED, fg="white", relief="flat", padx=8, pady=3
                  ).pack(side="left", padx=(0, 3))
        tk.Button(btn_row, text="Edit", command=self._edit,
                  bg=_BG3, fg=_FG, relief="flat", padx=8, pady=3
                  ).pack(side="left", padx=(0, 10))
        tk.Button(btn_row, text="Import from Charges CSV...",
                  command=self._import_csv,
                  bg=_ACC, fg="white", relief="flat", padx=8, pady=3
                  ).pack(side="left")

        bottom = ttk.Frame(self)
        bottom.pack(fill="x", padx=15, pady=(5, 15))
        tk.Button(bottom, text="Cancel", command=self.destroy,
                  bg=_BG3, fg=_FG, relief="flat", padx=15, pady=5
                  ).pack(side="right", padx=(5, 0))
        tk.Button(bottom, text="Save", command=self._on_save,
                  bg=_GREEN, fg="white", relief="flat", padx=15, pady=5
                  ).pack(side="right")

        self._refresh()
        self.geometry("550x520")

    def _calc_age(self, start_month_str):
        try:
            parts = start_month_str.split("-")
            sy, sm = int(parts[0]), int(parts[1])
            today = datetime.date.today()
            return (today.year - sy) * 12 + (today.month - sm)
        except Exception:
            return "?"

    def _refresh(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
        for c in sorted(self.cohorts, key=lambda x: x.get("start_month", "")):
            age = self._calc_age(c.get("start_month", ""))
            self.tree.insert("", "end", values=(
                c.get("start_month", ""),
                c.get("size", 0),
                c.get("track", "MONG"),
                age,
            ))

    def _add(self):
        dlg = _CohortEntryDialog(self)
        self.wait_window(dlg)
        if dlg.result:
            self.cohorts.append(dlg.result)
            self._refresh()

    def _remove(self):
        sel = self.tree.selection()
        if not sel:
            return
        idx = self.tree.index(sel[0])
        del self.cohorts[sorted(range(len(self.cohorts)),
                                key=lambda i: self.cohorts[i].get("start_month", ""))[idx]]
        self._refresh()

    def _edit(self):
        sel = self.tree.selection()
        if not sel:
            return
        values = self.tree.item(sel[0], "values")
        idx = self.tree.index(sel[0])
        sorted_indices = sorted(range(len(self.cohorts)),
                                key=lambda i: self.cohorts[i].get("start_month", ""))
        actual_idx = sorted_indices[idx]
        dlg = _CohortEntryDialog(self, self.cohorts[actual_idx])
        self.wait_window(dlg)
        if dlg.result:
            self.cohorts[actual_idx] = dlg.result
            self._refresh()

    def _import_csv(self):
        """Import cohort sizes from a Recharge charges CSV."""
        path = filedialog.askopenfilename(
            title="Import Charges CSV",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
            parent=self)
        if not path:
            return

        try:
            with open(path, "r", newline="", encoding="utf-8-sig") as f:
                reader = csv.DictReader(f)
                rows = list(reader)
        except Exception as e:
            messagebox.showerror("Import Error", str(e), parent=self)
            return

        # group unique customers by earliest scheduled_at month
        customer_months = {}
        for row in rows:
            cid = row.get("customer_id", "").strip()
            sched = row.get("scheduled_at", "").strip()
            if not cid or not sched:
                continue
            month_str = sched[:7]  # YYYY-MM
            if cid not in customer_months or month_str < customer_months[cid]:
                customer_months[cid] = month_str

        # count per month
        month_counts = defaultdict(int)
        for cid, month in customer_months.items():
            month_counts[month] += 1

        if not month_counts:
            messagebox.showinfo("No Data",
                                "No customer/month data found in CSV.",
                                parent=self)
            return

        # merge into cohorts (replace existing months, keep others)
        existing_months = {c["start_month"]: i for i, c in enumerate(self.cohorts)}
        for month, count in sorted(month_counts.items()):
            if month in existing_months:
                self.cohorts[existing_months[month]]["size"] = count
            else:
                self.cohorts.append({
                    "start_month": month, "size": count, "track": "MONG"
                })

        self._refresh()
        messagebox.showinfo("Imported",
                            f"Imported {len(month_counts)} cohort months "
                            f"from {len(customer_months)} customers.",
                            parent=self)

    def _on_save(self):
        self.result = self.cohorts
        self.destroy()


class _CohortEntryDialog(tk.Toplevel):
    """Small dialog for entering a cohort."""

    def __init__(self, parent, data=None):
        super().__init__(parent)
        self.title("Cohort")
        self.configure(bg=_BG)
        self.transient(parent)
        self.grab_set()
        self.result = None
        data = data or {}

        ttk.Label(self, text="Start Month (YYYY-MM):").pack(
            anchor="w", padx=15, pady=(15, 3))
        self.month_var = tk.StringVar(value=data.get("start_month", ""))
        ttk.Entry(self, textvariable=self.month_var, width=15).pack(
            anchor="w", padx=15)

        ttk.Label(self, text="Cohort Size:").pack(
            anchor="w", padx=15, pady=(10, 3))
        self.size_var = tk.StringVar(value=str(data.get("size", "")))
        ttk.Entry(self, textvariable=self.size_var, width=10).pack(
            anchor="w", padx=15)

        ttk.Label(self, text="Track:").pack(
            anchor="w", padx=15, pady=(10, 3))
        self.track_var = tk.StringVar(value=data.get("track", "MONG"))
        ttk.Combobox(self, textvariable=self.track_var,
                     values=["MONG", "MS", "BYO"],
                     state="readonly", width=10).pack(anchor="w", padx=15)

        btn_frame = ttk.Frame(self)
        btn_frame.pack(fill="x", padx=15, pady=15)
        tk.Button(btn_frame, text="Cancel", command=self.destroy,
                  bg=_BG3, fg=_FG, relief="flat", padx=10, pady=4
                  ).pack(side="right", padx=(5, 0))
        tk.Button(btn_frame, text="OK", command=self._on_ok,
                  bg=_GREEN, fg="white", relief="flat", padx=10, pady=4
                  ).pack(side="right")

        self.geometry("320x280")

    def _on_ok(self):
        month = self.month_var.get().strip()
        if not month or len(month) != 7 or month[4] != "-":
            messagebox.showerror("Error", "Month must be YYYY-MM format.",
                                 parent=self)
            return
        try:
            size = int(self.size_var.get())
            if size <= 0:
                raise ValueError
        except ValueError:
            messagebox.showerror("Error", "Size must be a positive integer.",
                                 parent=self)
            return
        self.result = {
            "start_month": month,
            "size": size,
            "track": self.track_var.get(),
        }
        self.destroy()


# ═════════════════════════════════════════════════════════════════════
#  RETENTION MATRIX EDITOR DIALOG
# ═════════════════════════════════════════════════════════════════════

class RetentionMatrixDialog(tk.Toplevel):
    """Editable grid for the retention matrix and churn rates."""

    def __init__(self, parent, retention_matrix, churn_rates, repeat_rate):
        super().__init__(parent)
        self.title("Retention Matrix Editor")
        self.configure(bg=_BG)
        self.transient(parent)
        self.grab_set()
        self.result = None

        self.matrix_vars = {}
        self.churn_vars = {}

        # ── Matrix grid ──
        grid_frame = ttk.LabelFrame(self, text="Retention Matrix (% of original cohort)",
                                    padding=10)
        grid_frame.pack(fill="both", expand=True, padx=15, pady=(15, 5))

        # header row
        ttk.Label(grid_frame, text="Curation", style="Bold.TLabel").grid(
            row=0, column=0, padx=3, pady=2)
        for j in range(7):
            ttk.Label(grid_frame, text=f"Mo {j+1}", style="Bold.TLabel").grid(
                row=0, column=j+1, padx=3, pady=2)

        for i, cur in enumerate(CURATION_ORDER):
            ttk.Label(grid_frame, text=cur).grid(
                row=i+1, column=0, sticky="w", padx=3, pady=1)
            self.matrix_vars[cur] = []
            row_data = retention_matrix.get(cur, [0]*7)
            for j in range(7):
                val = row_data[j] if j < len(row_data) else 0
                var = tk.StringVar(value=f"{val:.2f}")
                entry = ttk.Entry(grid_frame, textvariable=var, width=6)
                entry.grid(row=i+1, column=j+1, padx=2, pady=1)
                self.matrix_vars[cur].append(var)

        # column sum validation
        self.sum_label = ttk.Label(grid_frame, text="", style="Dim.TLabel")
        self.sum_label.grid(row=len(CURATION_ORDER)+1, column=0,
                            columnspan=8, sticky="w", pady=(5, 0))

        # ── Churn rates ──
        churn_frame = ttk.LabelFrame(self, text="Churn Rates", padding=10)
        churn_frame.pack(fill="x", padx=15, pady=5)

        mong_churn = churn_rates.get("MONG", {})
        ms_churn = churn_rates.get("MS", {})

        churn_fields = [
            ("MONG Month 1:", "MONG_m1", mong_churn.get("month_1", 0.15)),
            ("MONG Month 2+:", "MONG_m2", mong_churn.get("month_2_plus", 0.12)),
            ("MS Month 1:", "MS_m1", ms_churn.get("month_1", 0.15)),
            ("MS Month 2:", "MS_m2", ms_churn.get("month_2", 0.10)),
            ("MS Month 3+:", "MS_m3", ms_churn.get("month_3_plus", 0.12)),
        ]

        for i, (label, key, val) in enumerate(churn_fields):
            ttk.Label(churn_frame, text=label).grid(
                row=i // 3, column=(i % 3) * 2, sticky="w", padx=(0, 3), pady=2)
            var = tk.StringVar(value=f"{val:.2f}")
            ttk.Entry(churn_frame, textvariable=var, width=6).grid(
                row=i // 3, column=(i % 3) * 2 + 1, padx=(0, 10), pady=2)
            self.churn_vars[key] = var

        # repeat rate
        ttk.Label(churn_frame, text="Repeat Rate:").grid(
            row=2, column=0, sticky="w", padx=(0, 3), pady=2)
        self.repeat_var = tk.StringVar(value=f"{repeat_rate:.2f}")
        ttk.Entry(churn_frame, textvariable=self.repeat_var, width=6).grid(
            row=2, column=1, padx=(0, 10), pady=2)

        # ── buttons ──
        bottom = ttk.Frame(self)
        bottom.pack(fill="x", padx=15, pady=(5, 15))
        tk.Button(bottom, text="Validate", command=self._validate,
                  bg=_ACC, fg="white", relief="flat", padx=10, pady=5
                  ).pack(side="left")
        tk.Button(bottom, text="Cancel", command=self.destroy,
                  bg=_BG3, fg=_FG, relief="flat", padx=15, pady=5
                  ).pack(side="right", padx=(5, 0))
        tk.Button(bottom, text="Save", command=self._on_save,
                  bg=_GREEN, fg="white", relief="flat", padx=15, pady=5
                  ).pack(side="right")

        self.geometry("650x520")
        self.minsize(550, 450)

    def _validate(self):
        """Check that column sums are <= 1.0."""
        issues = []
        for j in range(7):
            col_sum = 0
            for cur in CURATION_ORDER:
                try:
                    col_sum += float(self.matrix_vars[cur][j].get())
                except ValueError:
                    issues.append(f"Invalid value in {cur} Mo {j+1}")
                    break
            if col_sum > 1.001:
                issues.append(f"Mo {j+1} sum = {col_sum:.2f} (> 1.00)")

        if issues:
            self.sum_label.configure(
                text="Issues: " + "; ".join(issues), foreground=_RED)
        else:
            self.sum_label.configure(
                text="All columns valid.", foreground=_GREEN)

    def _on_save(self):
        matrix = {}
        for cur in CURATION_ORDER:
            row = []
            for var in self.matrix_vars[cur]:
                try:
                    row.append(float(var.get()))
                except ValueError:
                    messagebox.showerror("Error",
                                         f"Invalid value in {cur} row.",
                                         parent=self)
                    return
            matrix[cur] = row

        churn_rates = {
            "MONG": {
                "month_1": float(self.churn_vars["MONG_m1"].get()),
                "month_2_plus": float(self.churn_vars["MONG_m2"].get()),
            },
            "MS": {
                "month_1": float(self.churn_vars["MS_m1"].get()),
                "month_2": float(self.churn_vars["MS_m2"].get()),
                "month_3_plus": float(self.churn_vars["MS_m3"].get()),
            },
        }

        try:
            repeat = float(self.repeat_var.get())
        except ValueError:
            repeat = REPEAT_RATE

        self.result = {
            "retention_matrix": matrix,
            "churn_rates": churn_rates,
            "repeat_rate": repeat,
        }
        self.destroy()


# ═════════════════════════════════════════════════════════════════════
#  OPEN PO MANAGER DIALOG
# ═════════════════════════════════════════════════════════════════════

class OpenPODialog(tk.Toplevel):
    """Manage open purchase orders for supply pipeline."""

    def __init__(self, parent, open_pos):
        super().__init__(parent)
        self.title("Open Purchase Orders")
        self.configure(bg=_BG)
        self.transient(parent)
        self.grab_set()
        self.result = None

        self.data = [dict(po) for po in open_pos]

        top = ttk.Frame(self)
        top.pack(fill="x", padx=15, pady=(15, 5))
        ttk.Label(top, text="Open POs, Manufacturing Orders, and Transfers",
                  style="Bold.TLabel").pack(anchor="w")

        tree_frame = ttk.Frame(self)
        tree_frame.pack(fill="both", expand=True, padx=15, pady=5)

        cols = ("sku", "qty", "eta", "type", "vendor", "status")
        self.tree = ttk.Treeview(
            tree_frame, columns=cols, show="headings", height=14)
        self.tree.heading("sku", text="SKU")
        self.tree.heading("qty", text="Qty")
        self.tree.heading("eta", text="ETA")
        self.tree.heading("type", text="Type")
        self.tree.heading("vendor", text="Vendor")
        self.tree.heading("status", text="Status")
        self.tree.column("sku", width=130)
        self.tree.column("qty", width=70, anchor="center")
        self.tree.column("eta", width=100)
        self.tree.column("type", width=80, anchor="center")
        self.tree.column("vendor", width=120)
        self.tree.column("status", width=90, anchor="center")

        yscroll = ttk.Scrollbar(tree_frame, orient="vertical",
                                command=self.tree.yview)
        self.tree.configure(yscrollcommand=yscroll.set)
        self.tree.pack(side="left", fill="both", expand=True)
        yscroll.pack(side="right", fill="y")

        btn_row = ttk.Frame(self)
        btn_row.pack(fill="x", padx=15, pady=5)
        tk.Button(btn_row, text="+ Add", command=self._add,
                  bg=_GREEN, fg="white", relief="flat", padx=8, pady=3
                  ).pack(side="left", padx=(0, 3))
        tk.Button(btn_row, text="- Remove", command=self._remove,
                  bg=_RED, fg="white", relief="flat", padx=8, pady=3
                  ).pack(side="left", padx=(0, 3))
        tk.Button(btn_row, text="Edit", command=self._edit,
                  bg=_BG3, fg=_FG, relief="flat", padx=8, pady=3
                  ).pack(side="left", padx=(0, 3))
        tk.Button(btn_row, text="Mark Received", command=self._mark_received,
                  bg=_ACC, fg="white", relief="flat", padx=8, pady=3
                  ).pack(side="left", padx=(0, 10))
        tk.Button(btn_row, text="Import CSV...", command=self._import_csv,
                  bg=_BG3, fg=_FG, relief="flat", padx=8, pady=3
                  ).pack(side="left")

        bottom = ttk.Frame(self)
        bottom.pack(fill="x", padx=15, pady=(5, 15))
        tk.Button(bottom, text="Cancel", command=self.destroy,
                  bg=_BG3, fg=_FG, relief="flat", padx=15, pady=5
                  ).pack(side="right", padx=(5, 0))
        tk.Button(bottom, text="Save", command=self._on_save,
                  bg=_GREEN, fg="white", relief="flat", padx=15, pady=5
                  ).pack(side="right")

        self._refresh()
        self.geometry("700x520")

    def _refresh(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
        for po in self.data:
            self.tree.insert("", "end", values=(
                po.get("sku", ""), po.get("qty", 0), po.get("eta", ""),
                po.get("type", "PO"), po.get("vendor", ""),
                po.get("status", "Open"),
            ))

    def _add(self):
        dlg = _POEntryDialog(self)
        self.wait_window(dlg)
        if dlg.result:
            self.data.append(dlg.result)
            self._refresh()

    def _remove(self):
        sel = self.tree.selection()
        if not sel:
            return
        idx = self.tree.index(sel[0])
        del self.data[idx]
        self._refresh()

    def _edit(self):
        sel = self.tree.selection()
        if not sel:
            return
        idx = self.tree.index(sel[0])
        dlg = _POEntryDialog(self, self.data[idx])
        self.wait_window(dlg)
        if dlg.result:
            self.data[idx] = dlg.result
            self._refresh()

    def _mark_received(self):
        sel = self.tree.selection()
        if not sel:
            return
        idx = self.tree.index(sel[0])
        self.data[idx]["status"] = "Received"
        self._refresh()

    def _import_csv(self):
        path = filedialog.askopenfilename(
            title="Import PO CSV",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
            parent=self)
        if not path:
            return
        try:
            with open(path, "r", newline="", encoding="utf-8-sig") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    sku = (row.get("sku") or row.get("SKU") or
                           row.get("Product SKU") or "").strip()
                    if not sku:
                        continue
                    qty_str = (row.get("qty") or row.get("Qty") or
                               row.get("Quantity") or "0").strip()
                    try:
                        qty = float(qty_str.replace(",", ""))
                    except ValueError:
                        qty = 0
                    self.data.append({
                        "sku": sku,
                        "qty": qty,
                        "eta": (row.get("eta") or row.get("ETA") or
                                row.get("Expected") or ""),
                        "type": (row.get("type") or row.get("Type") or "PO"),
                        "vendor": (row.get("vendor") or row.get("Vendor") or ""),
                        "status": "Open",
                    })
            self._refresh()
        except Exception as e:
            messagebox.showerror("Import Error", str(e), parent=self)

    def _on_save(self):
        self.result = self.data
        self.destroy()


class _POEntryDialog(tk.Toplevel):
    """Small dialog for entering/editing a PO line."""

    def __init__(self, parent, data=None):
        super().__init__(parent)
        self.title("Purchase Order Entry")
        self.configure(bg=_BG)
        self.transient(parent)
        self.grab_set()
        self.result = None
        data = data or {}

        frame = ttk.Frame(self)
        frame.pack(fill="both", expand=True, padx=15, pady=15)

        fields = [
            ("SKU:", "sku", data.get("sku", "")),
            ("Quantity:", "qty", data.get("qty", "")),
            ("ETA (YYYY-MM-DD):", "eta", data.get("eta", "")),
            ("Vendor:", "vendor", data.get("vendor", "")),
        ]
        self.vars = {}
        for i, (label, key, val) in enumerate(fields):
            ttk.Label(frame, text=label).grid(
                row=i, column=0, sticky="w", padx=(0, 10), pady=3)
            var = tk.StringVar(value=str(val))
            ttk.Entry(frame, textvariable=var, width=25).grid(
                row=i, column=1, pady=3, sticky="w")
            self.vars[key] = var

        ttk.Label(frame, text="Type:").grid(
            row=len(fields), column=0, sticky="w", padx=(0, 10), pady=3)
        self.type_var = tk.StringVar(value=data.get("type", "PO"))
        ttk.Combobox(frame, textvariable=self.type_var,
                     values=["PO", "MFG", "Transfer"],
                     state="readonly", width=10).grid(
            row=len(fields), column=1, sticky="w", pady=3)

        btn_frame = ttk.Frame(self)
        btn_frame.pack(fill="x", padx=15, pady=(0, 15))
        tk.Button(btn_frame, text="Cancel", command=self.destroy,
                  bg=_BG3, fg=_FG, relief="flat", padx=10, pady=4
                  ).pack(side="right", padx=(5, 0))
        tk.Button(btn_frame, text="OK", command=self._on_ok,
                  bg=_GREEN, fg="white", relief="flat", padx=10, pady=4
                  ).pack(side="right")

        self.geometry("380x280")

    def _on_ok(self):
        sku = self.vars["sku"].get().strip()
        if not sku:
            messagebox.showerror("Error", "SKU required.", parent=self)
            return
        try:
            qty = float(self.vars["qty"].get())
        except ValueError:
            messagebox.showerror("Error", "Invalid quantity.", parent=self)
            return
        self.result = {
            "sku": sku,
            "qty": qty,
            "eta": self.vars["eta"].get().strip(),
            "type": self.type_var.get(),
            "vendor": self.vars["vendor"].get().strip(),
            "status": "Open",
        }
        self.destroy()


# ═════════════════════════════════════════════════════════════════════
#  MAIN APPLICATION
# ═════════════════════════════════════════════════════════════════════

class InventoryReorderApp:
    """Main application window with tabbed interface."""

    def __init__(self, root):
        self.root = root
        self.root.title(f"Inventory Reorder System  v{APP_VERSION}")
        self.root.geometry("1200x750")
        self.root.minsize(900, 550)
        self.root.configure(bg=_BG)

        # ── load settings ──
        self.saved = load_settings()

        # ── data stores ──
        # inventory: {sku: {qty, name, category, location, unit_cost, ...}}
        self.inventory = self.saved.get("inventory", {})
        # sku_settings: {sku: {purchase_lt, production_lt, shipping_lt,
        #                       safety_stock, churn_pct}}
        self.sku_settings = self.saved.get("sku_settings", {})
        # bundle_map: {bundle_sku: [(component_sku, qty), ...]}
        self.bundle_map = {
            k: [tuple(c) for c in v]
            for k, v in self.saved.get("bundle_map", {}).items()
        }
        # recharge_demand: {sku: weekly_qty} — last pull
        self.recharge_demand = self.saved.get("recharge_demand", {})
        # shopify_forecast: {bundle_or_sku: weekly_qty}
        self.shopify_forecast = self.saved.get("shopify_forecast", {})
        # shopify API pulled demand (separate from manual forecast)
        self.shopify_api_demand = self.saved.get("shopify_api_demand", {})
        # shopify trend data from aggregate_with_trend
        self.shopify_trend_data = self.saved.get("shopify_trend_data", {})
        # recharge queued charges: {month_label: {sku: qty}}
        self.recharge_queued = self.saved.get("recharge_queued", {})
        # resolved PR-CJAM/CEX-EC from queued charges:
        # {month: {pr_cjam: {suffix: count}, cex_ec: {suffix: count}, unresolved: int}}
        self.recharge_queued_resolved = self.saved.get(
            "recharge_queued_resolved", {})
        # manual_demand: {sku: weekly_qty}
        self.manual_demand = self.saved.get("manual_demand", {})
        # last CSV column mapping
        self.last_csv_mapping = self.saved.get("last_csv_mapping", {})
        # SKU name translations: {product_name: sku_code}
        self.sku_translations = self.saved.get("sku_translations", {})
        # archived SKUs: hidden from dashboard by default
        self.archived_skus = set(self.saved.get("archived_skus", []))
        # depletion history: [{date, file, day, skus: {sku: qty}, total,
        #                      reship_count, total_orders}]
        self.depletion_history = self.saved.get("depletion_history", [])
        # reship buffer: auto-calculated % from fulfillment history
        self.reship_buffer_pct = self.saved.get("reship_buffer_pct", 0.0)
        # retention calibration data from Recharge
        self.actual_retention = self.saved.get("actual_retention", {})
        # customer lifecycle data: {email: {first_order_date, order_count,
        #                           last_order_date, months_active}}
        self.customer_lifecycle = self.saved.get("customer_lifecycle", {})
        # customization variance: {sku: {expected, actual, variance_factor}}
        self.customization_variance = self.saved.get(
            "customization_variance", {})

        # ── cohort forecasting data ──
        self.cohorts = self.saved.get("cohorts", [])
        self.retention_matrix = self.saved.get(
            "retention_matrix", {k: list(v) for k, v in DEFAULT_RETENTION_MATRIX.items()})
        self.churn_rates = self.saved.get("churn_rates", dict(DEFAULT_CHURN_RATES))
        self.repeat_rate = self.saved.get("repeat_rate", REPEAT_RATE)
        self.curation_recipes = {
            k: [tuple(c) for c in v]
            for k, v in self.saved.get(
                "curation_recipes",
                {k: [list(c) for c in v] for k, v in DEFAULT_CURATION_RECIPES.items()}
            ).items()
        }
        _raw_pr_cjam = self.saved.get("pr_cjam", dict(DEFAULT_PR_CJAM))
        # Migrate old format: {curation: "cheese_sku"} -> {curation: {cheese, jam}}
        self.pr_cjam = {}
        for k, v in _raw_pr_cjam.items():
            if isinstance(v, str):
                self.pr_cjam[k] = {"cheese": v, "jam": ""}
            else:
                self.pr_cjam[k] = dict(v)
        self.cex_ec = self.saved.get("cex_ec", dict(DEFAULT_CEX_EC))
        self.wheel_inventory = self.saved.get("wheel_inventory", {})
        self.open_pos = self.saved.get("open_pos", [])
        self.forecast_months = self.saved.get("forecast_months", 3)
        self.last_forecast = None

        # ── monthly curated box data (MED / CMED / LGE) ──
        # {month_label: {box_type: [(slot, sku, qty), ...], ...}}
        self.monthly_box_recipes = self.saved.get("monthly_box_recipes", {})
        # {box_type: count} — from Recharge or manual override
        self.monthly_box_counts = self.saved.get("monthly_box_counts", {})
        self._monthly_box_counts_manual = self.saved.get(
            "_monthly_box_counts_manual", {})

        # ── vendor catalog: {sku: {vendor, unit_cost, case_qty, moq,
        #                           wheel_weight_lbs}} ──
        self.vendor_catalog = self.saved.get("vendor_catalog", {})

        # ── EX-EC curation assignments history ──
        # {curation: [sku, ...]} — cheeses assigned for EX-EC per curation
        self.exec_assignments = self.saved.get("exec_assignments", {})

        # ── fulfillment planner state ──
        self.fp_sat_demand = {}
        self.fp_inventory = {}
        self.fp_results = []
        self.fp_mascot_state = "idle"

        # ── notification alerts ──
        self.alerts = []  # populated on recalculate

        # ── calendar action schedule ──
        self.action_schedule = []  # [{date, type, title, details}, ...]

        # ── integrations ──
        # ClickUp
        self.clickup_api_token = self.saved.get("clickup_api_token", "")
        self.clickup_list_id = self.saved.get("clickup_list_id", "")
        # Google Calendar
        self.gcal_refresh_token = self.saved.get("gcal_refresh_token", "")
        self.gcal_client_id = self.saved.get("gcal_client_id", "")
        self.gcal_client_secret = self.saved.get("gcal_client_secret", "")
        # Dropbox
        self.dropbox_refresh_token = self.saved.get(
            "dropbox_refresh_token", "")
        self.dropbox_app_key = self.saved.get("dropbox_app_key", "")
        self.dropbox_app_secret = self.saved.get("dropbox_app_secret", "")
        self.dropbox_shared_link = self.saved.get("dropbox_shared_link", "")

        # ── reconciliation history ──
        self.reconciliation_history = self.saved.get(
            "reconciliation_history", [])

        # ── raw-to-finished conversions ──
        # {bulk_ingredient_name: {target_sku, packet_oz, unit_type}}
        self.bulk_conversions = self.saved.get(
            "bulk_conversions", dict(DEFAULT_BULK_CONVERSIONS))
        # {target_sku: [{date, wheel_sku, weight, expected, actual, variance}]}
        self.production_yield_history = self.saved.get(
            "production_yield_history", [])
        # Per-SKU adjusted conversion factors (learned from yield history)
        # {sku: adjusted_factor}  — overrides WHEEL_TO_SLICE_FACTOR
        self.adjusted_conversion_factors = self.saved.get(
            "adjusted_conversion_factors", {})

        # ── v3.0 multi-warehouse ──
        self.warehouses = self.saved.get("warehouses", {
            "Primary": {"label": "Primary Fulfillment (RMFG TX)",
                        "is_fulfillment": True},
            "Woburn": {"label": "Woburn MA", "is_fulfillment": False,
                       "capabilities": ["receive", "process", "crossdock",
                                        "store"]},
        })
        # Transfer history: [{date, sku, qty, from_warehouse, to_warehouse}]
        self.transfer_history = self.saved.get("transfer_history", [])

        # ── v3.0 processing queue ──
        # [{id, sku, source_material, target_qty, status, warehouse,
        #   created, completed, actual_yield}]
        _pq_defaults = {"id": "", "sku": "", "source_material": "",
                        "target_qty": 0, "status": "scheduled",
                        "warehouse": "Primary", "created": "",
                        "completed": None, "actual_yield": None}
        self.processing_queue = [
            {**_pq_defaults, **j}
            for j in self.saved.get("processing_queue", [])
            if isinstance(j, dict)
        ]

        # ── v3.0 yield discrepancies ──
        _yd_defaults = {"date": "", "sku": "", "type": "",
                        "expected_qty": 0, "actual_qty": 0, "variance": 0,
                        "yield_date": "", "snapshot_date": "",
                        "status": "open"}
        self.yield_discrepancies = [
            {**_yd_defaults, **d}
            for d in self.saved.get("yield_discrepancies", [])
            if isinstance(d, dict)
        ]

        # ── v3.0 yield reconciliation settings ──
        self.yield_recon_window_days = self.saved.get(
            "yield_reconciliation_window_days", 3)
        self.yield_recon_threshold_pct = self.saved.get(
            "yield_reconciliation_threshold_pct", 5)
        self.yield_recon_threshold_min = self.saved.get(
            "yield_reconciliation_threshold_min", 2)

        # ── v2.5 automation settings ──
        # Slack
        self.slack_webhook_url = self.saved.get("slack_webhook_url", "")
        self.slack_notify_critical = self.saved.get(
            "slack_notify_critical", True)
        self.slack_notify_expiring = self.saved.get(
            "slack_notify_expiring", True)
        self.slack_notify_shortfall = self.saved.get(
            "slack_notify_shortfall", True)

        # Email (SMTP) for depletion reports
        self.smtp_host = self.saved.get("smtp_host", "smtp.gmail.com")
        self.smtp_port = self.saved.get("smtp_port", "587")
        self.smtp_user = self.saved.get("smtp_user", "")
        self.smtp_password = self.saved.get("smtp_password", "")
        self.depletion_email_to = self.saved.get("depletion_email_to", "")
        self.depletion_email_from = self.saved.get("depletion_email_from", "")

        # Auto-refresh interval (minutes, 0 = disabled)
        self.auto_refresh_interval = self.saved.get(
            "auto_refresh_interval", 30)

        # File watcher
        self._file_watcher_active = False
        self._last_shipments_files = set()
        self._last_inventory_files = set()

        # Webhook server
        self.webhook_port = self.saved.get("webhook_port", 8765)
        self.webhook_secret_shopify = self.saved.get(
            "webhook_secret_shopify", "")
        self.webhook_secret_recharge = self.saved.get(
            "webhook_secret_recharge", "")
        self._webhook_server = None

        # Auto-sync flags
        self.auto_sync_clickup = self.saved.get("auto_sync_clickup", False)
        self.auto_sync_gcal = self.saved.get("auto_sync_gcal", False)
        self.auto_po_threshold = self.saved.get("auto_po_threshold", 0)

        # ── apply dark theme ──
        self._apply_dark_theme()

        # ── build UI ──
        self._build_menu()
        self._build_ui()

        # ── keyboard shortcuts ──
        self.root.bind("<F5>", lambda e: self._recalculate())
        self.root.bind("<Control-f>",
                       lambda e: self.filter_entry.focus_set()
                       if hasattr(self, 'filter_entry') else None)
        self.root.bind("<Control-s>",
                       lambda e: self._show_snapshot_comparison())
        self.root.bind("<Control-w>",
                       lambda e: self._show_workflow_guide())

        # ── initial calculation ──
        self._recalculate()

        # ── v2.5 automation startup ──
        self.root.after(1000, self._startup_automation)

    # ─────────────────────────────────────────────────────────────────
    #  WAREHOUSE HELPERS
    # ─────────────────────────────────────────────────────────────────

    def _po_qty_by_sku(self):
        """Aggregate open PO quantities by SKU (excludes received)."""
        result = defaultdict(float)
        for po in self.open_pos:
            if po.get("status", "").lower() != "received":
                result[po["sku"]] += po.get("qty", 0)
        return result

    def _bulk_source_sku_for(self, source_material):
        """Find the inventory SKU that holds the raw bulk material.

        Uses the bulk_conversions mapping to find the exact inventory
        SKU whose name matches, category is 'Bulk Raw Materials'.
        Returns (inv_sku, inv_data) or (None, None).
        """
        # First try exact match via bulk_conversions keyword
        for inv_sku, inv_data in self.inventory.items():
            name = inv_data.get("name", "")
            if (name == source_material and
                    inv_data.get("category") == "Bulk Raw Materials"):
                return inv_sku, inv_data
        # Fallback: check if source_material is a key in bulk_conversions
        # and the inventory item's name matches
        conv = self.bulk_conversions.get(source_material, {})
        if conv:
            for inv_sku, inv_data in self.inventory.items():
                if (inv_data.get("category") == "Bulk Raw Materials" and
                        inv_data.get("name", "") == source_material):
                    return inv_sku, inv_data
        return None, None

    def _qty_at(self, sku, warehouse="Primary"):
        """Get quantity of a SKU at a specific warehouse."""
        data = self.inventory.get(sku, {})
        wh_qty = data.get("warehouse_qty", {})
        if wh_qty:
            return float(wh_qty.get(warehouse, 0))
        # Legacy: single warehouse field
        if data.get("warehouse", "Primary") == warehouse:
            return float(data.get("qty", 0))
        return 0.0

    def _set_qty_at(self, sku, warehouse, qty):
        """Set quantity of a SKU at a specific warehouse.

        Initializes warehouse_qty if not present.
        Updates total qty to sum of all warehouses.
        """
        if sku not in self.inventory:
            return
        data = self.inventory[sku]
        wh_qty = data.get("warehouse_qty")
        if wh_qty is None:
            # Migrate from legacy single-warehouse model
            old_wh = data.get("warehouse", "Primary")
            old_qty = float(data.get("qty", 0))
            wh_qty = {old_wh: old_qty}
            data["warehouse_qty"] = wh_qty
        wh_qty[warehouse] = max(0.0, float(qty))
        # Clean up zero entries
        if wh_qty[warehouse] == 0:
            wh_qty.pop(warehouse, None)
        # Update total qty and primary warehouse field
        data["qty"] = sum(wh_qty.values())
        # Set warehouse field to where the majority of stock is
        if wh_qty:
            data["warehouse"] = max(wh_qty, key=wh_qty.get)
        else:
            data["warehouse"] = "Primary"

    def _primary_inventory(self):
        """Return inventory with qty reflecting Primary warehouse only.

        For SKUs with warehouse_qty, overrides qty to Primary amount.
        Excludes SKUs with zero Primary stock.
        """
        result = {}
        for sku, data in self.inventory.items():
            wh_qty = data.get("warehouse_qty")
            if wh_qty:
                primary_qty = wh_qty.get("Primary", 0)
                if primary_qty > 0:
                    result[sku] = {**data, "qty": primary_qty}
            elif data.get("warehouse", "Primary") == "Primary":
                result[sku] = data
        return result

    # ─────────────────────────────────────────────────────────────────
    #  DARK THEME
    # ─────────────────────────────────────────────────────────────────

    def _apply_dark_theme(self):
        style = ttk.Style()
        style.theme_use("clam")

        style.configure(".", background=_BG, foreground=_FG,
                        fieldbackground=_BG3, bordercolor=_SEP,
                        insertcolor=_FG,
                        font=("Segoe UI", 9))

        style.configure("TFrame", background=_BG)
        style.configure("TLabel", background=_BG, foreground=_FG)
        style.configure("Bold.TLabel", background=_BG, foreground=_FG,
                        font=("Segoe UI", 9, "bold"))
        style.configure("Dim.TLabel", background=_BG, foreground=_FG2)
        style.configure("Title.TLabel", background=_BG, foreground=_FG,
                        font=("Segoe UI", 14, "bold"))
        style.configure("Subtitle.TLabel", background=_BG, foreground=_FG,
                        font=("Segoe UI", 11, "bold"))

        style.configure("TEntry", fieldbackground=_BG3, foreground=_FG,
                        insertcolor=_FG, padding=4)

        style.configure("TCombobox", fieldbackground=_BG3, foreground=_FG,
                        padding=4)
        style.map("TCombobox",
                  fieldbackground=[("readonly", _BG3)],
                  foreground=[("readonly", _FG)])

        style.configure("TButton", background=_BG3, foreground=_FG,
                        padding=(10, 5), relief="flat")
        style.map("TButton",
                  background=[("active", "#525252"), ("pressed", "#333333")],
                  foreground=[("disabled", _FG2)])

        style.configure("Run.TButton", background="#2255a0", foreground="white",
                        font=("Segoe UI", 9, "bold"), padding=(12, 6))
        style.map("Run.TButton",
                  background=[("active", "#2d6bbd"), ("pressed", "#1a4480")])

        style.configure("Apply.TButton", background=_GREEN, foreground="white",
                        padding=(12, 6))
        style.map("Apply.TButton",
                  background=[("active", "#228040"), ("pressed", "#155528")])

        style.configure("TCheckbutton", background=_BG, foreground=_FG)
        style.map("TCheckbutton",
                  background=[("active", _BG)],
                  indicatorcolor=[("selected", _ACC), ("!selected", _BG3)])

        style.configure("TNotebook", background=_BG, bordercolor=_SEP)
        style.configure("TNotebook.Tab", background=_BG2, foreground=_FG,
                        padding=(12, 5))
        style.map("TNotebook.Tab",
                  background=[("selected", _BG)],
                  foreground=[("selected", _FG)])

        style.configure("Treeview", background=_BG2, foreground=_FG,
                        fieldbackground=_BG2, rowheight=24,
                        font=("Segoe UI", 9))
        style.configure("Treeview.Heading", background=_BG3, foreground=_FG,
                        font=("Segoe UI", 9, "bold"))
        style.map("Treeview",
                  background=[("selected", _ACC)],
                  foreground=[("selected", "white")])

        style.configure("Vertical.TScrollbar", background=_BG2,
                        troughcolor=_BG, bordercolor=_BG, arrowcolor=_FG2)
        style.configure("Horizontal.TScrollbar", background=_BG2,
                        troughcolor=_BG, bordercolor=_BG, arrowcolor=_FG2)

        style.configure("TLabelframe", background=_BG, foreground=_FG)
        style.configure("TLabelframe.Label", background=_BG, foreground=_FG,
                        font=("Segoe UI", 9, "bold"))

        style.configure("TSeparator", background=_SEP)

    # ─────────────────────────────────────────────────────────────────
    #  MENU BAR
    # ─────────────────────────────────────────────────────────────────

    def _build_menu(self):
        menubar = tk.Menu(self.root, bg=_BG2, fg=_FG, activebackground=_ACC,
                          activeforeground="white", relief="flat")

        file_menu = tk.Menu(menubar, tearoff=0, bg=_BG2, fg=_FG,
                            activebackground=_ACC, activeforeground="white")
        file_menu.add_command(label="Import Inventory CSV...",
                              command=self._import_inventory_csv)
        file_menu.add_command(label="Export Dashboard CSV...",
                              command=self._export_dashboard_csv)
        file_menu.add_separator()
        file_menu.add_command(label="Deplete && Email Report...",
                              command=self._deplete_and_email)
        file_menu.add_separator()
        file_menu.add_command(label="Save Settings",
                              command=self._save_all_settings)
        file_menu.add_separator()
        file_menu.add_command(label="Exit",
                              command=self.root.quit)
        menubar.add_cascade(label="File", menu=file_menu)

        edit_menu = tk.Menu(menubar, tearoff=0, bg=_BG2, fg=_FG,
                            activebackground=_ACC, activeforeground="white")
        edit_menu.add_command(label="Bundle Mappings...",
                              command=self._open_bundle_editor)
        edit_menu.add_command(label="Shopify Forecast...",
                              command=self._open_shopify_forecast)
        edit_menu.add_command(label="Manual Demand Adjustments...",
                              command=self._open_manual_demand)
        edit_menu.add_separator()
        edit_menu.add_command(label="Curation Recipes...",
                              command=self._open_recipe_editor)
        edit_menu.add_command(label="Monthly Box Recipes...",
                              command=self._open_monthly_box_editor)
        edit_menu.add_command(label="Cohort Manager...",
                              command=self._open_cohort_editor)
        edit_menu.add_command(label="Retention Matrix...",
                              command=self._open_retention_editor)
        edit_menu.add_command(label="Open POs...",
                              command=self._open_po_editor)
        edit_menu.add_command(label="Vendor Catalog...",
                              command=self._open_vendor_catalog)
        edit_menu.add_separator()
        edit_menu.add_command(label="Auto-PO Generator...",
                              command=self._generate_auto_po)
        edit_menu.add_command(label="Production Orders...",
                              command=self._generate_production_orders)
        menubar.add_cascade(label="Edit", menu=edit_menu)

        help_menu = tk.Menu(menubar, tearoff=0, bg=_BG2, fg=_FG,
                            activebackground=_ACC, activeforeground="white")
        help_menu.add_command(label="Workflow Guide",
                              command=self._show_workflow_guide)
        help_menu.add_separator()
        help_menu.add_command(
            label="About",
            command=lambda: messagebox.showinfo(
                "About",
                f"Inventory Reorder System v{APP_VERSION}\n\n"
                "Safety Stock reorder point calculator\n"
                "with Recharge + Shopify demand forecasting.\n\n"
                "Multi-warehouse support (Primary + Woburn)\n"
                "Bulk-aware alerts, processing queue, yield tracking."))
        menubar.add_cascade(label="Help", menu=help_menu)

        self.root.config(menu=menubar)

    # ─────────────────────────────────────────────────────────────────
    #  BUILD UI
    # ─────────────────────────────────────────────────────────────────

    def _build_ui(self):
        # ── top bar ──
        top_bar = ttk.Frame(self.root)
        top_bar.pack(fill="x", padx=15, pady=(10, 0))

        ttk.Label(top_bar, text="Inventory Reorder System",
                  style="Title.TLabel").pack(side="left")

        self.status_var = tk.StringVar(value="Ready")
        ttk.Label(top_bar, textvariable=self.status_var,
                  style="Dim.TLabel").pack(side="right", padx=(10, 0))

        # Yield discrepancy flags
        self._yield_flag_var = tk.StringVar(value="")
        self._yield_flag_btn = ttk.Button(
            top_bar, textvariable=self._yield_flag_var,
            command=self._show_yield_discrepancies)
        open_flags = sum(1 for d in self.yield_discrepancies
                         if d.get("status") == "open")
        if open_flags > 0:
            self._yield_flag_var.set(f"{open_flags} yield flags")
            self._yield_flag_btn.pack(side="right", padx=(0, 5))

        # Alert bell
        self._alert_badge_var = tk.StringVar(value="Alerts")
        ttk.Button(top_bar, textvariable=self._alert_badge_var,
                   command=self._show_alert_panel).pack(
                       side="right", padx=(0, 5))

        ttk.Separator(self.root, orient="horizontal").pack(
            fill="x", padx=15, pady=(8, 0))

        # ── notebook ──
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill="both", expand=True, padx=10, pady=10)

        self._build_dashboard_tab()
        self._build_demand_tab()
        self._build_inventory_tab()
        self._build_forecasting_tab()
        self._build_calendar_tab()
        self._build_fulfillment_tab()
        self._build_settings_tab()

    # ─────────────────────────────────────────────────────────────────
    #  TAB 1 — DASHBOARD (main reorder view)
    # ─────────────────────────────────────────────────────────────────

    def _build_dashboard_tab(self):
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text="  Dashboard  ")

        # toolbar
        toolbar = ttk.Frame(tab)
        toolbar.pack(fill="x", padx=10, pady=(10, 5))

        ttk.Button(toolbar, text="Recalculate",
                   style="Run.TButton",
                   command=self._recalculate).pack(side="left", padx=(0, 5))
        ttk.Button(toolbar, text="Export CSV",
                   command=self._export_dashboard_csv).pack(
                       side="left", padx=(0, 5))
        ttk.Button(toolbar, text="Snapshot",
                   command=self._show_snapshot_comparison).pack(
                       side="left", padx=(0, 5))
        ttk.Button(toolbar, text="Workflow Guide",
                   command=self._show_workflow_guide).pack(
                       side="left", padx=(0, 5))

        # filter
        ttk.Label(toolbar, text="Filter:").pack(side="left", padx=(20, 5))
        self.filter_var = tk.StringVar()
        self.filter_var.trace_add("write", lambda *_: self._apply_filter())
        ttk.Entry(toolbar, textvariable=self.filter_var, width=20).pack(
            side="left", padx=(0, 5))

        self.alert_only_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(toolbar, text="Alerts only",
                        variable=self.alert_only_var,
                        command=self._apply_filter).pack(
                            side="left", padx=(10, 0))

        self.show_archived_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(toolbar, text="Show Archived",
                        variable=self.show_archived_var,
                        command=self._apply_filter).pack(
                            side="left", padx=(10, 0))

        # Warehouse filter
        ttk.Label(toolbar, text="Warehouse:").pack(side="left", padx=(15, 5))
        self.dash_warehouse_var = tk.StringVar(value="All Locations")
        wh_combo = ttk.Combobox(
            toolbar, textvariable=self.dash_warehouse_var,
            values=["All Locations"] + list(self.warehouses.keys()),
            state="readonly", width=14)
        wh_combo.pack(side="left")
        wh_combo.bind("<<ComboboxSelected>>",
                       lambda e: self._apply_filter())

        # summary bar
        summary = ttk.Frame(tab)
        summary.pack(fill="x", padx=10, pady=(0, 5))

        self.summary_total_var = tk.StringVar(value="SKUs: 0")
        self.summary_alert_var = tk.StringVar(value="Alerts: 0")
        self.summary_reorder_var = tk.StringVar(value="Need Reorder: 0")

        ttk.Label(summary, textvariable=self.summary_total_var,
                  style="Bold.TLabel").pack(side="left", padx=(0, 20))
        ttk.Label(summary, textvariable=self.summary_alert_var,
                  foreground=_RED,
                  font=("Segoe UI", 9, "bold")).pack(side="left", padx=(0, 20))
        ttk.Label(summary, textvariable=self.summary_reorder_var,
                  foreground=_ORANGE,
                  font=("Segoe UI", 9, "bold")).pack(side="left", padx=(0, 20))

        self.summary_expiring_var = tk.StringVar(value="Expiring: 0")
        ttk.Label(summary, textvariable=self.summary_expiring_var,
                  foreground="#d4850a",
                  font=("Segoe UI", 9, "bold")).pack(side="left")

        # treeview
        tree_frame = ttk.Frame(tab)
        tree_frame.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        self._dash_cols = {
            "sku":           ("SKU", 150),
            "name":          ("Product Name", 180),
            "category":      ("Category", 130),
            "on_hand":       ("On Hand", 80),
            "weekly_usage":  ("Weekly Usage", 90),
            "lead_time":     ("Lead Time (d)", 95),
            "safety_stock":  ("Safety Stock", 90),
            "reorder_point": ("Reorder Point", 100),
            "surplus":       ("Surplus / Deficit", 110),
            "status":        ("Status", 90),
            "recharge_wk":   ("Recharge/wk", 90),
            "shopify_wk":    ("Shopify/wk", 85),
            "manual_wk":     ("Manual/wk", 80),
            "wheel_supply":  ("Wheel Pot.", 80),
            "bulk_supply":   ("Bulk Pot.", 80),
            "total_avail":   ("Total Avail.", 90),
            "expiration":    ("Expiration", 90),
        }
        cols = list(self._dash_cols.keys())

        yscroll = ttk.Scrollbar(tree_frame, orient="vertical")
        xscroll = ttk.Scrollbar(tree_frame, orient="horizontal")
        self.dash_tree = ttk.Treeview(
            tree_frame, columns=cols, show="headings",
            yscrollcommand=yscroll.set, xscrollcommand=xscroll.set,
            selectmode="extended")
        yscroll.configure(command=self.dash_tree.yview)
        xscroll.configure(command=self.dash_tree.xview)

        for col, (label, width) in self._dash_cols.items():
            self.dash_tree.heading(
                col, text=label,
                command=lambda c=col: self._sort_dash(c))
            anchor = "w" if col in ("sku", "name", "status") else "center"
            self.dash_tree.column(col, width=width, minwidth=50,
                                  anchor=anchor)

        self.dash_tree.pack(side="left", fill="both", expand=True)
        yscroll.pack(side="right", fill="y")
        xscroll.pack(side="bottom", fill="x")

        # row tags for alerts
        self.dash_tree.tag_configure(
            "CRITICAL", background="#8b1a1a", foreground="white")
        self.dash_tree.tag_configure(
            "WARNING", background="#7a5500", foreground="white")
        self.dash_tree.tag_configure(
            "OK", background=_BG2, foreground=_FG)
        self.dash_tree.tag_configure(
            "OVERSTOCK", background="#1a4a1a", foreground="#90ee90")

        # double-click to edit SKU settings
        self.dash_tree.bind("<Double-1>", self._on_dash_double_click)
        # right-click context menu
        self.dash_tree.bind("<Button-3>", self._on_dash_right_click)

        # sorting state
        self._dash_sort_col = "status"
        self._dash_sort_rev = False

        # store full calculated data for filtering
        self._dash_rows = []

    # ─────────────────────────────────────────────────────────────────
    #  TAB 2 — DEMAND SOURCES
    # ─────────────────────────────────────────────────────────────────

    def _build_demand_tab(self):
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text="  Demand Sources  ")

        # ── Recharge section ──
        rc_frame = ttk.LabelFrame(tab, text="Recharge Subscriptions",
                                  padding=10)
        rc_frame.pack(fill="x", padx=15, pady=(10, 5))

        row1 = ttk.Frame(rc_frame)
        row1.pack(fill="x", pady=(0, 5))

        ttk.Label(row1, text="API Token:").pack(side="left")
        self.recharge_token_var = tk.StringVar(
            value=self.saved.get("recharge_api_token", ""))
        self.recharge_token_entry = ttk.Entry(
            row1, textvariable=self.recharge_token_var, width=50, show="*")
        self.recharge_token_entry.pack(side="left", padx=(5, 10))

        self.show_token_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(row1, text="Show",
                        variable=self.show_token_var,
                        command=self._toggle_token_visibility).pack(
                            side="left", padx=(0, 10))

        ttk.Button(row1, text="Pull Subscriptions",
                   style="Run.TButton",
                   command=self._pull_recharge).pack(side="left", padx=(10, 0))

        row2 = ttk.Frame(rc_frame)
        row2.pack(fill="x")

        ttk.Label(row2, text="Default Churn Rate (%):").pack(side="left")
        self.default_churn_var = tk.StringVar(
            value=str(self.saved.get("default_churn_pct", "5")))
        ttk.Entry(row2, textvariable=self.default_churn_var, width=6).pack(
            side="left", padx=(5, 15))

        self.recharge_status_var = tk.StringVar(
            value=self._recharge_status_text())
        ttk.Label(row2, textvariable=self.recharge_status_var,
                  style="Dim.TLabel").pack(side="left")

        # ── Shopify section ──
        sp_frame = ttk.LabelFrame(tab, text="Shopify First-Time Orders",
                                  padding=10)
        sp_frame.pack(fill="x", padx=15, pady=5)

        # API credentials row 1
        sp_api_row1 = ttk.Frame(sp_frame)
        sp_api_row1.pack(fill="x", pady=(0, 5))

        ttk.Label(sp_api_row1, text="Store URL:").pack(side="left")
        self.shopify_store_var = tk.StringVar(
            value=self.saved.get("shopify_store_url", ""))
        ttk.Entry(sp_api_row1, textvariable=self.shopify_store_var,
                  width=30).pack(side="left", padx=(5, 15))

        ttk.Label(sp_api_row1, text="Access Token:").pack(side="left")
        self.shopify_token_var = tk.StringVar(
            value=self.saved.get("shopify_access_token", ""))
        self.shopify_token_entry = ttk.Entry(
            sp_api_row1, textvariable=self.shopify_token_var,
            width=35, show="*")
        self.shopify_token_entry.pack(side="left", padx=(5, 0))

        # API credentials row 2
        sp_api_row2 = ttk.Frame(sp_frame)
        sp_api_row2.pack(fill="x", pady=(0, 5))

        ttk.Label(sp_api_row2, text="API Key:").pack(side="left")
        self.shopify_api_key_var = tk.StringVar(
            value=self.saved.get("shopify_api_key", ""))
        ttk.Entry(sp_api_row2, textvariable=self.shopify_api_key_var,
                  width=30).pack(side="left", padx=(5, 15))

        ttk.Label(sp_api_row2, text="API Secret:").pack(side="left")
        self.shopify_api_secret_var = tk.StringVar(
            value=self.saved.get("shopify_api_secret", ""))
        ttk.Entry(sp_api_row2, textvariable=self.shopify_api_secret_var,
                  width=35, show="*").pack(side="left", padx=(5, 0))

        ttk.Button(sp_api_row2, text="Authorize Shopify",
                   style="Run.TButton",
                   command=self._authorize_shopify).pack(
                       side="left", padx=(15, 0))

        # Pull controls row
        sp_pull_row = ttk.Frame(sp_frame)
        sp_pull_row.pack(fill="x", pady=(0, 5))

        ttk.Label(sp_pull_row, text="Order Tag:").pack(side="left")
        self.shopify_tag_var = tk.StringVar(
            value=self.saved.get("shopify_order_tag",
                                 "Subscription First Order"))
        ttk.Entry(sp_pull_row, textvariable=self.shopify_tag_var,
                  width=25).pack(side="left", padx=(5, 10))

        ttk.Label(sp_pull_row, text="Weeks back:").pack(side="left")
        self.shopify_weeks_var = tk.StringVar(
            value=str(self.saved.get("shopify_weeks_back", "4")))
        ttk.Entry(sp_pull_row, textvariable=self.shopify_weeks_var,
                  width=4).pack(side="left", padx=(5, 10))

        ttk.Button(sp_pull_row, text="Pull Orders",
                   style="Run.TButton",
                   command=self._pull_shopify).pack(side="left", padx=(10, 5))
        ttk.Button(sp_pull_row, text="Edit Manual Overrides...",
                   command=self._open_shopify_forecast).pack(side="left")

        self.shopify_status_var = tk.StringVar(
            value=self._shopify_status_text())
        ttk.Label(sp_frame, textvariable=self.shopify_status_var,
                  style="Dim.TLabel").pack(anchor="w", pady=(5, 0))

        # ── Manual Adjustments section ──
        ma_frame = ttk.LabelFrame(tab, text="Manual Demand Adjustments",
                                  padding=10)
        ma_frame.pack(fill="x", padx=15, pady=5)

        ma_row = ttk.Frame(ma_frame)
        ma_row.pack(fill="x")

        ttk.Label(ma_row,
                  text="Optional per-SKU weekly demand overrides/additions."
                  ).pack(side="left")
        ttk.Button(ma_row, text="Edit Adjustments...",
                   command=self._open_manual_demand).pack(side="right")

        self.manual_status_var = tk.StringVar(
            value=f"{len(self.manual_demand)} SKU(s) with adjustments")
        ttk.Label(ma_frame, textvariable=self.manual_status_var,
                  style="Dim.TLabel").pack(anchor="w", pady=(5, 0))

        # ── Analytics section ──
        an_frame = ttk.LabelFrame(tab, text="Retention & Lifecycle Analytics",
                                  padding=10)
        an_frame.pack(fill="x", padx=15, pady=5)

        an_row = ttk.Frame(an_frame)
        an_row.pack(fill="x")

        ttk.Button(an_row, text="Calibrate Retention (Recharge)",
                   command=self._pull_retention_data).pack(
                       side="left", padx=(0, 5))
        ttk.Button(an_row, text="Customer Lifecycle (Shopify)",
                   command=self._pull_customer_lifecycle).pack(
                       side="left", padx=(0, 5))

        self.analytics_status_var = tk.StringVar(
            value=self._analytics_status_text())
        ttk.Label(an_frame, textvariable=self.analytics_status_var,
                  style="Dim.TLabel").pack(anchor="w", pady=(5, 0))

        # ── Demand summary ──
        ds_frame = ttk.LabelFrame(tab, text="Combined Demand Summary",
                                  padding=10)
        ds_frame.pack(fill="both", expand=True, padx=15, pady=(5, 10))

        tree_frame = ttk.Frame(ds_frame)
        tree_frame.pack(fill="both", expand=True)

        demand_cols = {
            "sku": ("SKU", 200),
            "recharge": ("Recharge/wk", 100),
            "shopify": ("Shopify/wk", 100),
            "manual": ("Manual/wk", 100),
            "total_wk": ("Total/wk", 100),
            "daily": ("Daily", 80),
        }

        yscroll = ttk.Scrollbar(tree_frame, orient="vertical")
        self.demand_tree = ttk.Treeview(
            tree_frame, columns=list(demand_cols.keys()), show="headings",
            yscrollcommand=yscroll.set, height=10)
        yscroll.configure(command=self.demand_tree.yview)

        for col, (label, width) in demand_cols.items():
            self.demand_tree.heading(col, text=label)
            anchor = "w" if col == "sku" else "center"
            self.demand_tree.column(col, width=width, minwidth=60,
                                    anchor=anchor)

        self.demand_tree.pack(side="left", fill="both", expand=True)
        yscroll.pack(side="right", fill="y")

        btn_frame = ttk.Frame(ds_frame)
        btn_frame.pack(fill="x", pady=(5, 0))
        ttk.Button(btn_frame, text="Refresh Summary",
                   command=self._refresh_demand_summary).pack(
                       side="left")

    # ─────────────────────────────────────────────────────────────────
    #  TAB 3 — INVENTORY
    # ─────────────────────────────────────────────────────────────────

    def _build_inventory_tab(self):
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text="  Inventory  ")

        toolbar = ttk.Frame(tab)
        toolbar.pack(fill="x", padx=10, pady=(10, 5))

        ttk.Button(toolbar, text="Import CSV",
                   style="Run.TButton",
                   command=self._import_inventory_csv).pack(
                       side="left", padx=(0, 5))
        ttk.Button(toolbar, text="Import Depletion",
                   command=self._import_depletion_matrix).pack(
                       side="left", padx=(0, 5))
        ttk.Button(toolbar, text="Deplete & Email",
                   style="Run.TButton",
                   command=self._deplete_and_email).pack(
                       side="left", padx=(0, 5))
        ttk.Button(toolbar, text="Undo Depletion",
                   command=self._undo_last_depletion).pack(
                       side="left", padx=(0, 5))
        ttk.Button(toolbar, text="Add SKU",
                   command=self._add_sku_manual).pack(
                       side="left", padx=(0, 5))
        ttk.Button(toolbar, text="Edit Selected",
                   command=self._edit_inventory_row).pack(
                       side="left", padx=(0, 5))
        ttk.Button(toolbar, text="Remove Selected",
                   command=self._remove_inventory_row).pack(
                       side="left", padx=(0, 5))

        ttk.Button(toolbar, text="Transfer Woburn->Primary",
                   command=self._show_transfer_dialog).pack(
                       side="left", padx=(0, 5))
        ttk.Button(toolbar, text="Transfer History",
                   command=self._show_transfer_history).pack(
                       side="left", padx=(0, 5))

        # Warehouse filter for inventory tab
        ttk.Label(toolbar, text="Warehouse:").pack(side="right", padx=(10, 5))
        self.inv_warehouse_var = tk.StringVar(value="All Locations")
        inv_wh_combo = ttk.Combobox(
            toolbar, textvariable=self.inv_warehouse_var,
            values=["All Locations"] + list(self.warehouses.keys()),
            state="readonly", width=14)
        inv_wh_combo.pack(side="right")
        inv_wh_combo.bind("<<ComboboxSelected>>",
                          lambda e: self._refresh_inventory_tree())

        self.inv_count_var = tk.StringVar(
            value=f"{len(self.inventory)} SKUs loaded")
        ttk.Label(toolbar, textvariable=self.inv_count_var,
                  style="Dim.TLabel").pack(side="right", padx=(0, 10))

        tree_frame = ttk.Frame(tab)
        tree_frame.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        inv_cols = {
            "sku":          ("SKU", 160),
            "name":         ("Product Name", 200),
            "qty":          ("Qty On Hand", 90),
            "category":     ("Category", 120),
            "location":     ("Location", 100),
            "unit_cost":    ("Unit Cost", 80),
            "exp_earliest": ("Earliest Exp.", 95),
            "exp_batches":  ("Batches", 60),
        }

        yscroll = ttk.Scrollbar(tree_frame, orient="vertical")
        xscroll = ttk.Scrollbar(tree_frame, orient="horizontal")
        self.inv_tree = ttk.Treeview(
            tree_frame, columns=list(inv_cols.keys()), show="headings",
            yscrollcommand=yscroll.set, xscrollcommand=xscroll.set,
            selectmode="extended")
        yscroll.configure(command=self.inv_tree.yview)
        xscroll.configure(command=self.inv_tree.xview)

        for col, (label, width) in inv_cols.items():
            self.inv_tree.heading(col, text=label)
            anchor = "w" if col in ("sku", "name", "category",
                                     "location") else "center"
            self.inv_tree.column(col, width=width, minwidth=60,
                                anchor=anchor)

        self.inv_tree.pack(side="left", fill="both", expand=True)
        yscroll.pack(side="right", fill="y")
        xscroll.pack(side="bottom", fill="x")

        self.inv_tree.bind("<Double-1>", self._on_inv_double_click)

        # expiration row tags
        self.inv_tree.tag_configure(
            "EXPIRED", background="#8b1a1a", foreground="white")
        self.inv_tree.tag_configure(
            "EXPIRING_SOON", background="#7a5500", foreground="white")

        self._refresh_inventory_tree()

    # ─────────────────────────────────────────────────────────────────
    #  TAB 4 — FORECASTING
    # ─────────────────────────────────────────────────────────────────

    def _build_forecasting_tab(self):
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text="  Forecasting  ")

        # toolbar
        toolbar = ttk.Frame(tab)
        toolbar.pack(fill="x", padx=10, pady=(10, 5))

        ttk.Button(toolbar, text="Run Forecast",
                   style="Run.TButton",
                   command=self._run_forecast).pack(side="left", padx=(0, 5))
        ttk.Button(toolbar, text="Edit Cohorts...",
                   command=self._open_cohort_editor).pack(
                       side="left", padx=(0, 5))
        ttk.Button(toolbar, text="Edit Recipes...",
                   command=self._open_recipe_editor).pack(
                       side="left", padx=(0, 5))
        ttk.Button(toolbar, text="Monthly Boxes...",
                   command=self._open_monthly_box_editor).pack(
                       side="left", padx=(0, 5))
        ttk.Button(toolbar, text="Edit Retention Matrix...",
                   command=self._open_retention_editor).pack(
                       side="left", padx=(0, 5))
        ttk.Button(toolbar, text="Open POs...",
                   command=self._open_po_editor).pack(
                       side="left", padx=(0, 15))

        ttk.Label(toolbar, text="Months:").pack(side="left", padx=(0, 3))
        self.horizon_var = tk.StringVar(value=str(self.forecast_months))
        horizon_spin = ttk.Spinbox(toolbar, from_=1, to=6,
                                   textvariable=self.horizon_var, width=4)
        horizon_spin.pack(side="left")

        self.forecast_status_var = tk.StringVar(
            value=f"Cohorts: {len(self.cohorts)} | "
                  f"Recipes: {len(self.curation_recipes)}")
        ttk.Label(toolbar, textvariable=self.forecast_status_var,
                  style="Dim.TLabel").pack(side="right")

        # ── top section: forecast by month (sub-notebook) ──
        top_label = ttk.Frame(tab)
        top_label.pack(fill="x", padx=10, pady=(5, 0))
        ttk.Label(top_label, text="Forecast by Month",
                  style="Subtitle.TLabel").pack(anchor="w")

        self.forecast_notebook = ttk.Notebook(tab)
        self.forecast_notebook.pack(fill="both", expand=True, padx=10, pady=(0, 5))

        # placeholder tab until forecast is run
        placeholder = ttk.Frame(self.forecast_notebook)
        self.forecast_notebook.add(placeholder, text="  No Forecast  ")
        ttk.Label(placeholder,
                  text="Click 'Run Forecast' to generate projections.\n\n"
                       "Setup: Add cohorts, configure recipes, "
                       "then run the forecast.",
                  style="Dim.TLabel", justify="center").pack(
                      expand=True, pady=40)

        # ── bottom section: reorder alerts ──
        alert_frame = ttk.LabelFrame(tab, text="Reorder Alerts", padding=5)
        alert_frame.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        # filter row
        alert_toolbar = ttk.Frame(alert_frame)
        alert_toolbar.pack(fill="x", pady=(0, 5))

        ttk.Label(alert_toolbar, text="Show:").pack(side="left", padx=(0, 5))
        self.alert_filter_var = tk.StringVar(value="All")
        for label in ("All", "PO Only", "MFG Only", "Transfer Only"):
            ttk.Radiobutton(alert_toolbar, text=label,
                            variable=self.alert_filter_var, value=label,
                            command=self._filter_alerts).pack(
                                side="left", padx=(0, 8))

        alert_tree_frame = ttk.Frame(alert_frame)
        alert_tree_frame.pack(fill="both", expand=True)

        alert_cols = ("sku", "action", "urgency", "deficit", "on_hand",
                      "open_po", "wheel_supply", "bulk_supply",
                      "demand", "current_supply")
        self.alert_tree = ttk.Treeview(
            alert_tree_frame, columns=alert_cols, show="headings", height=8)
        self.alert_tree.heading("sku", text="SKU")
        self.alert_tree.heading("action", text="Action")
        self.alert_tree.heading("urgency", text="Urgency")
        self.alert_tree.heading("deficit", text="Deficit")
        self.alert_tree.heading("on_hand", text="On Hand")
        self.alert_tree.heading("open_po", text="Open PO")
        self.alert_tree.heading("wheel_supply", text="Wheel Supply")
        self.alert_tree.heading("bulk_supply", text="Bulk Supply")
        self.alert_tree.heading("demand", text="Demand")
        self.alert_tree.heading("current_supply", text="Net Supply")
        self.alert_tree.column("sku", width=120)
        self.alert_tree.column("action", width=70, anchor="center")
        self.alert_tree.column("urgency", width=75, anchor="center")
        self.alert_tree.column("deficit", width=70, anchor="center")
        self.alert_tree.column("on_hand", width=70, anchor="center")
        self.alert_tree.column("open_po", width=70, anchor="center")
        self.alert_tree.column("wheel_supply", width=85, anchor="center")
        self.alert_tree.column("bulk_supply", width=85, anchor="center")
        self.alert_tree.column("demand", width=70, anchor="center")
        self.alert_tree.column("current_supply", width=80, anchor="center")

        yscroll = ttk.Scrollbar(alert_tree_frame, orient="vertical",
                                command=self.alert_tree.yview)
        self.alert_tree.configure(yscrollcommand=yscroll.set)
        self.alert_tree.pack(side="left", fill="both", expand=True)
        yscroll.pack(side="right", fill="y")

        # alert row tags
        self.alert_tree.tag_configure(
            "CRITICAL", background="#8b1a1a", foreground="white")
        self.alert_tree.tag_configure(
            "WARNING", background="#7a5500", foreground="white")
        self.alert_tree.tag_configure(
            "PLAN", background="#1a3a6b", foreground="#a0c4ff")

        # store alerts for filtering
        self._forecast_alerts = []

    def _run_forecast(self):
        """Execute the cohort-based forecast engine."""
        if not self.cohorts:
            messagebox.showinfo(
                "No Cohorts",
                "Add subscriber cohorts first via 'Edit Cohorts'.",
                parent=self.root)
            return

        try:
            months = int(self.horizon_var.get())
        except ValueError:
            months = 3

        self.status_var.set("Running forecast...")
        self.root.update_idletasks()

        # run forecast
        forecast = forecast_cohort_demand(
            self.cohorts, self.retention_matrix, self.curation_recipes,
            self.pr_cjam, self.cex_ec, forecast_months=months)

        # compute monthly box demand and merge additively
        box_demand = compute_monthly_box_demand(
            self.monthly_box_recipes, self.monthly_box_counts,
            forecast_months=months)

        for ml, bd in box_demand.items():
            if ml in forecast:
                for sku, qty in bd["sku_demand"].items():
                    forecast[ml]["sku_demand"][sku] = \
                        forecast[ml]["sku_demand"].get(sku, 0) + qty
            else:
                forecast[ml] = {
                    "curation_counts": {},
                    "sku_demand": dict(bd["sku_demand"]),
                    "total_boxes": 0,
                }
            # stash box_counts for display
            forecast[ml]["monthly_box_counts"] = bd.get("box_counts", {})

        # overlay queued charges: replace cohort SKU demand for covered months
        default_churn = self._get_default_churn()
        for q_month, q_skus in self.recharge_queued.items():
            if q_month not in forecast:
                continue
            # build replacement SKU demand from queued charges
            queued_demand = defaultdict(float)
            charge_count = 0
            for sku, qty in q_skus.items():
                charge_count += qty
                churn = self.sku_settings.get(sku, {}).get(
                    "churn_pct", default_churn)
                adjusted = apply_churn_rate(qty, churn)
                # Skip generic PR-CJAM-GEN and bare CEX-EC — resolved below
                upper = sku.upper()
                if upper == "PR-CJAM-GEN" or upper == "CEX-EC":
                    continue
                for comp_sku, comp_qty in decompose_bundles(
                        sku, adjusted, self.bundle_map):
                    queued_demand[comp_sku] += comp_qty

            # Add resolved PR-CJAM/CEX-EC demand as actual cheese/jam SKUs
            resolved = self.recharge_queued_resolved.get(q_month, {})
            pr_cjam_resolved = resolved.get("pr_cjam", {})
            cex_ec_resolved = resolved.get("cex_ec", {})

            for suffix, count in pr_cjam_resolved.items():
                if suffix == "GEN":
                    # Monthly box — keep as generic SKU
                    queued_demand["PR-CJAM-GEN"] += count
                    continue
                cjam = self.pr_cjam.get(suffix, {})
                cheese = cjam.get("cheese")
                jam = cjam.get("jam")
                if cheese:
                    queued_demand[cheese] += count
                if jam:
                    queued_demand[jam] += count
                # Also keep the resolved PR-CJAM-<CUR> SKU for tracking
                queued_demand[f"PR-CJAM-{suffix}"] += count

            for suffix, count in cex_ec_resolved.items():
                if suffix == "GEN":
                    queued_demand["CEX-EC"] += count
                    continue
                cheese = self.cex_ec.get(suffix)
                if cheese:
                    queued_demand[cheese] += count
                queued_demand[f"CEX-EC-{suffix}"] += count

            # replace cohort-projected demand with queued charge demand
            forecast[q_month]["sku_demand"] = dict(queued_demand)
            forecast[q_month]["queued_charge_count"] = int(charge_count)
            forecast[q_month]["queued_resolved"] = resolved

        self.last_forecast = forecast

        # clear old tabs
        for tab_id in self.forecast_notebook.tabs():
            self.forecast_notebook.forget(tab_id)

        # build month tabs
        wheel_supply = compute_wheel_supply(
            self.wheel_inventory, self.adjusted_conversion_factors)
        bulk_supply = compute_bulk_supply(
            self._primary_inventory(), self.bulk_conversions)
        all_alerts = []

        for month_label in sorted(forecast.keys()):
            month_data = forecast[month_label]
            month_tab = ttk.Frame(self.forecast_notebook)
            self.forecast_notebook.add(month_tab, text=f"  {month_label}  ")

            # split into curation breakdown and SKU demand
            paned = ttk.PanedWindow(month_tab, orient="horizontal")
            paned.pack(fill="both", expand=True, padx=5, pady=5)

            # left: curation breakdown
            left_frame = ttk.LabelFrame(paned, text="Curation Breakdown",
                                        padding=5)
            paned.add(left_frame, weight=1)

            cur_tree = ttk.Treeview(
                left_frame, columns=("curation", "boxes", "pct"),
                show="headings", height=10)
            cur_tree.heading("curation", text="Curation")
            cur_tree.heading("boxes", text="Box Count")
            cur_tree.heading("pct", text="% of Total")
            cur_tree.column("curation", width=80)
            cur_tree.column("boxes", width=80, anchor="center")
            cur_tree.column("pct", width=70, anchor="center")
            cur_tree.pack(fill="both", expand=True)

            total = month_data["total_boxes"] or 1
            for cur in CURATION_ORDER:
                boxes = month_data["curation_counts"].get(cur, 0)
                if boxes > 0:
                    pct = f"{boxes/total*100:.1f}%"
                    cur_tree.insert("", "end",
                                    values=(cur, round(boxes, 1), pct))

            # total row
            cur_tree.insert("", "end",
                            values=("TOTAL", round(month_data["total_boxes"], 1),
                                    "100%"))

            # monthly box counts (MED/CMED/LGE)
            mbc = month_data.get("monthly_box_counts", {})
            if mbc:
                cur_tree.insert("", "end", values=("───", "───", "───"))
                for bt in MONTHLY_BOX_TYPES:
                    bc = mbc.get(bt, 0)
                    if bc > 0:
                        cur_tree.insert("", "end", values=(bt, bc, ""))

            # queued charges count (if this month used queued data)
            qcc = month_data.get("queued_charge_count")
            if qcc is not None:
                cur_tree.insert("", "end", values=("───", "───", "───"))
                cur_tree.insert("", "end",
                                values=("Queued Charges", qcc, ""))

            # resolved PR-CJAM/CEX-EC breakdown
            qr = month_data.get("queued_resolved", {})
            pr_cjam_r = qr.get("pr_cjam", {})
            cex_ec_r = qr.get("cex_ec", {})
            unresolved = qr.get("unresolved", 0)
            if pr_cjam_r or cex_ec_r or unresolved:
                cur_tree.insert("", "end", values=("───", "───", "───"))
                for suffix, cnt in sorted(pr_cjam_r.items()):
                    label = f"PR-CJAM-{suffix}"
                    cur_tree.insert("", "end",
                                    values=(label, int(cnt), ""))
                for suffix, cnt in sorted(cex_ec_r.items()):
                    label = f"CEX-EC-{suffix}"
                    cur_tree.insert("", "end",
                                    values=(label, int(cnt), ""))
                if unresolved:
                    cur_tree.insert("", "end",
                                    values=("Unresolved", unresolved, "⚠"))

            # right: SKU demand
            right_frame = ttk.LabelFrame(paned, text="SKU Demand", padding=5)
            paned.add(right_frame, weight=2)

            sku_tree = ttk.Treeview(
                right_frame,
                columns=("sku", "demand", "on_hand", "open_po",
                          "wheel", "bulk", "net", "status"),
                show="headings", height=10)
            sku_tree.heading("sku", text="SKU")
            sku_tree.heading("demand", text="Forecast")
            sku_tree.heading("on_hand", text="On Hand")
            sku_tree.heading("open_po", text="Open PO")
            sku_tree.heading("wheel", text="Wheel Supply")
            sku_tree.heading("bulk", text="Bulk Supply")
            sku_tree.heading("net", text="Net Position")
            sku_tree.heading("status", text="Status")
            sku_tree.column("sku", width=100)
            sku_tree.column("demand", width=70, anchor="center")
            sku_tree.column("on_hand", width=65, anchor="center")
            sku_tree.column("open_po", width=65, anchor="center")
            sku_tree.column("wheel", width=80, anchor="center")
            sku_tree.column("bulk", width=80, anchor="center")
            sku_tree.column("net", width=80, anchor="center")
            sku_tree.column("status", width=70, anchor="center")

            sku_yscroll = ttk.Scrollbar(right_frame, orient="vertical",
                                        command=sku_tree.yview)
            sku_tree.configure(yscrollcommand=sku_yscroll.set)
            sku_tree.pack(side="left", fill="both", expand=True)
            sku_yscroll.pack(side="right", fill="y")

            sku_tree.tag_configure("short", background="#7a5500",
                                   foreground="white")
            sku_tree.tag_configure("ok", background=_BG2, foreground=_FG)

            po_by_sku = self._po_qty_by_sku()

            for sku in sorted(month_data["sku_demand"].keys()):
                demand = month_data["sku_demand"][sku]
                on_hand = float(self.inventory.get(sku, {}).get("qty", 0))
                po_qty = po_by_sku.get(sku, 0)
                ws = wheel_supply.get(sku, 0)
                bs = bulk_supply.get(sku, 0)
                net = on_hand + po_qty + ws + bs - demand
                status = "OK" if net >= 0 else "SHORT"
                tag = "ok" if net >= 0 else "short"
                sku_tree.insert("", "end", values=(
                    sku, demand, on_hand, round(po_qty, 1),
                    round(ws, 1), round(bs, 1), round(net, 1),
                    status), tags=(tag,))

            # compute alerts for this month
            month_alerts = compute_reorder_alerts(
                month_data, self.inventory, self.open_pos, wheel_supply,
                bulk_supply=bulk_supply)
            for a in month_alerts:
                a["needed_by"] = month_label
            all_alerts.extend(month_alerts)

        # populate alert tree
        self._forecast_alerts = all_alerts
        self._filter_alerts()

        self.forecast_status_var.set(
            f"Forecast: {len(forecast)} months | "
            f"Alerts: {len(all_alerts)} | "
            f"Cohorts: {len(self.cohorts)}")
        self.status_var.set(
            f"Forecast complete  |  "
            f"{datetime.datetime.now().strftime('%H:%M:%S')}")

    def _filter_alerts(self):
        """Filter alert treeview by action type."""
        for item in self.alert_tree.get_children():
            self.alert_tree.delete(item)

        filter_val = self.alert_filter_var.get()
        type_map = {"PO Only": "PO", "MFG Only": "MFG",
                    "Transfer Only": "Transfer"}
        filter_type = type_map.get(filter_val)

        for a in self._forecast_alerts:
            if filter_type and a["action"] != filter_type:
                continue
            self.alert_tree.insert("", "end", values=(
                a["sku"], a["action"], a["urgency"], a["deficit"],
                a["on_hand"], a["open_po"], a["wheel_supply"],
                a.get("bulk_supply", 0),
                a["demand"], a["current_supply"],
            ), tags=(a["urgency"],))

    def _open_cohort_editor(self):
        dlg = CohortManagerDialog(self.root, self.cohorts)
        self.root.wait_window(dlg)
        if dlg.result is not None:
            self.cohorts = dlg.result
            self.forecast_status_var.set(
                f"Cohorts: {len(self.cohorts)} | "
                f"Recipes: {len(self.curation_recipes)}")

    def _open_recipe_editor(self):
        inv_skus = list(self.inventory.keys())
        dlg = CurationRecipeDialog(
            self.root, self.curation_recipes, self.pr_cjam, self.cex_ec,
            inventory_skus=inv_skus)
        self.root.wait_window(dlg)
        if dlg.result is not None:
            self.curation_recipes = {
                k: [tuple(c) for c in v]
                for k, v in dlg.result["recipes"].items()
            }
            self.pr_cjam = dlg.result["pr_cjam"]
            self.cex_ec = dlg.result["cex_ec"]
            self.forecast_status_var.set(
                f"Cohorts: {len(self.cohorts)} | "
                f"Recipes: {len(self.curation_recipes)}")

    def _open_monthly_box_editor(self):
        inv_skus = list(self.inventory.keys())
        dlg = MonthlyBoxRecipeDialog(
            self.root, self.monthly_box_recipes, self.monthly_box_counts,
            inventory_skus=inv_skus)
        self.root.wait_window(dlg)
        if dlg.result is not None:
            self.monthly_box_recipes = dlg.result["recipes"]
            self.monthly_box_counts = dlg.result["box_counts"]

    def _open_retention_editor(self):
        dlg = RetentionMatrixDialog(
            self.root, self.retention_matrix, self.churn_rates,
            self.repeat_rate)
        self.root.wait_window(dlg)
        if dlg.result is not None:
            self.retention_matrix = dlg.result["retention_matrix"]
            self.churn_rates = dlg.result["churn_rates"]
            self.repeat_rate = dlg.result["repeat_rate"]

    def _open_po_editor(self):
        dlg = OpenPODialog(self.root, self.open_pos)
        self.root.wait_window(dlg)
        if dlg.result is not None:
            self.open_pos = dlg.result

    # ─────────────────────────────────────────────────────────────────
    #  TAB 5 — CALENDAR / ACTION SCHEDULE
    # ─────────────────────────────────────────────────────────────────

    def _build_calendar_tab(self):
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text="  Calendar  ")

        # Toolbar
        toolbar = ttk.Frame(tab)
        toolbar.pack(fill="x", padx=10, pady=(10, 5))

        ttk.Button(toolbar, text="Generate Schedule",
                   style="Run.TButton",
                   command=self._generate_action_schedule).pack(
                       side="left", padx=(0, 5))
        ttk.Button(toolbar, text="Sync to ClickUp",
                   command=self._sync_to_clickup).pack(
                       side="left", padx=(0, 5))
        ttk.Button(toolbar, text="Sync to Google Calendar",
                   command=self._sync_to_gcal).pack(
                       side="left", padx=(0, 5))
        ttk.Button(toolbar, text="Check Dropbox",
                   command=self._check_dropbox).pack(
                       side="left", padx=(0, 5))
        ttk.Button(toolbar, text="EX-EC Suggestions",
                   command=self._show_exec_suggestions).pack(
                       side="left", padx=(0, 5))
        ttk.Button(toolbar, text="Fulfillment Preview",
                   command=self._show_fulfillment_preview).pack(
                       side="left", padx=(0, 5))
        ttk.Button(toolbar, text="Reconcile Inventory",
                   command=self._run_reconciliation).pack(
                       side="left", padx=(0, 5))
        ttk.Button(toolbar, text="Processing Queue",
                   command=self._show_processing_queue).pack(
                       side="left", padx=(0, 5))

        # Month navigation
        nav_frame = ttk.Frame(tab)
        nav_frame.pack(fill="x", padx=10, pady=(0, 5))

        self._cal_month_offset = 0
        ttk.Button(nav_frame, text="< Prev",
                   command=self._cal_prev_month).pack(side="left")
        self._cal_month_label = tk.StringVar()
        ttk.Label(nav_frame, textvariable=self._cal_month_label,
                  style="Subtitle.TLabel").pack(side="left", padx=15)
        ttk.Button(nav_frame, text="Next >",
                   command=self._cal_next_month).pack(side="left")
        ttk.Button(nav_frame, text="Today",
                   command=self._cal_today).pack(side="left", padx=15)

        self._cal_status_var = tk.StringVar(value="Generate schedule first")
        ttk.Label(nav_frame, textvariable=self._cal_status_var,
                  style="Dim.TLabel").pack(side="right")

        # Calendar grid frame
        self._cal_grid_frame = ttk.Frame(tab)
        self._cal_grid_frame.pack(fill="both", expand=True, padx=10,
                                  pady=(0, 5))

        # Detail panel at bottom
        detail_frame = ttk.LabelFrame(tab, text="Day Detail", padding=5)
        detail_frame.pack(fill="x", padx=10, pady=(0, 10))

        self._cal_detail_var = tk.StringVar(
            value="Click a day cell to see scheduled actions.")
        ttk.Label(detail_frame, textvariable=self._cal_detail_var,
                  wraplength=900, justify="left").pack(
                      fill="x", padx=5, pady=5)

        self._render_calendar()

    def _cal_prev_month(self):
        self._cal_month_offset -= 1
        self._render_calendar()

    def _cal_next_month(self):
        self._cal_month_offset += 1
        self._render_calendar()

    def _cal_today(self):
        self._cal_month_offset = 0
        self._render_calendar()

    def _render_calendar(self):
        """Render the monthly calendar grid with action schedule overlay."""
        import calendar as cal_mod

        # Clear existing grid
        for w in self._cal_grid_frame.winfo_children():
            w.destroy()

        today = datetime.date.today()
        year = today.year
        month = today.month + self._cal_month_offset
        while month > 12:
            month -= 12
            year += 1
        while month < 1:
            month += 12
            year -= 1

        self._cal_month_label.set(f"{cal_mod.month_name[month]} {year}")

        # Color legend
        _type_colors = {
            "PO": "#3498db",       # blue
            "MFG": "#e67e22",      # orange
            "Transfer": "#9b59b6", # purple
            "Fulfillment": "#27ae60",  # green
            "Crossdock": "#e74c3c",    # red
            "Process": "#1abc9c",     # teal
        }

        # Day-of-week headers
        for i, day_name in enumerate(["Mon", "Tue", "Wed", "Thu",
                                      "Fri", "Sat", "Sun"]):
            lbl = tk.Label(self._cal_grid_frame, text=day_name,
                           bg=_BG3, fg=_FG, width=14,
                           font=("Segoe UI", 8, "bold"))
            lbl.grid(row=0, column=i, padx=1, pady=1, sticky="nsew")

        # Build events index by date string
        events_by_date = defaultdict(list)
        for item in self.action_schedule:
            events_by_date[item["date"]].append(item)

        # Calendar days
        month_cal = cal_mod.monthcalendar(year, month)
        for week_idx, week in enumerate(month_cal):
            for day_idx, day in enumerate(week):
                cell = tk.Frame(self._cal_grid_frame, bg=_BG2,
                                width=110, height=70)
                cell.grid(row=week_idx + 1, column=day_idx,
                          padx=1, pady=1, sticky="nsew")
                cell.grid_propagate(False)

                if day == 0:
                    continue

                date_str = f"{year}-{month:02d}-{day:02d}"
                is_today = (datetime.date(year, month, day) == today)
                day_bg = "#3a3a5a" if is_today else _BG2

                cell.configure(bg=day_bg)

                day_lbl = tk.Label(cell, text=str(day), bg=day_bg,
                                   fg=_FG if not is_today else "#ffff00",
                                   font=("Segoe UI", 9, "bold"),
                                   anchor="nw")
                day_lbl.pack(anchor="nw", padx=2)

                # Show events for this day
                day_events = events_by_date.get(date_str, [])
                for evt in day_events[:3]:  # max 3 visible
                    color = _type_colors.get(evt["type"], _FG2)
                    evt_lbl = tk.Label(
                        cell, text=evt["title"][:18],
                        bg=day_bg, fg=color,
                        font=("Segoe UI", 7), anchor="w")
                    evt_lbl.pack(anchor="w", padx=2)

                if len(day_events) > 3:
                    tk.Label(cell, text=f"+{len(day_events)-3} more",
                             bg=day_bg, fg=_FG2,
                             font=("Segoe UI", 7)).pack(anchor="w", padx=2)

                # Bind click
                def _on_click(ds=date_str, evts=day_events):
                    self._show_day_detail(ds, evts)
                cell.bind("<Button-1>", lambda e, f=_on_click: f())
                for child in cell.winfo_children():
                    child.bind("<Button-1>", lambda e, f=_on_click: f())

        # Make columns expand
        for i in range(7):
            self._cal_grid_frame.columnconfigure(i, weight=1)

        # Legend bar
        legend = ttk.Frame(self._cal_grid_frame)
        legend.grid(row=len(month_cal) + 1, column=0, columnspan=7,
                    sticky="w", pady=(5, 0))
        for action_type, color in _type_colors.items():
            tk.Label(legend, text=f"  {action_type}  ", bg=color,
                     fg="white", font=("Segoe UI", 7, "bold")).pack(
                         side="left", padx=2)

    def _show_day_detail(self, date_str, events):
        if not events:
            self._cal_detail_var.set(f"{date_str}: No actions scheduled.")
            return
        lines = [f"{date_str}: {len(events)} action(s)"]
        for evt in events:
            lines.append(
                f"  [{evt['type']}] {evt['title']}: {evt.get('details', '')}")
        self._cal_detail_var.set("\n".join(lines))

    def _generate_action_schedule(self):
        """Generate action schedule based on fulfillment dates and lead times."""
        import calendar as cal_mod

        self.action_schedule = []
        today = datetime.date.today()
        globals_ = self._get_global_defaults()

        # Fulfillment days: Tuesday (1) and Saturday (5) for next 3 months
        fulfillment_dates = []
        for offset in range(90):
            d = today + datetime.timedelta(days=offset)
            if d.weekday() in (1, 5):  # Tue=1, Sat=5
                fulfillment_dates.append(d)

        for ful_date in fulfillment_dates:
            ful_str = ful_date.isoformat()
            day_name = "TUE" if ful_date.weekday() == 1 else "SAT"

            # Fulfillment event
            self.action_schedule.append({
                "date": ful_str,
                "type": "Fulfillment",
                "title": f"Ship {day_name}",
                "details": f"Fulfillment batch for {day_name}",
            })

            # Crossdock: day before fulfillment
            crossdock = ful_date - datetime.timedelta(days=1)
            self.action_schedule.append({
                "date": crossdock.isoformat(),
                "type": "Crossdock",
                "title": f"Crossdock for {day_name}",
                "details": f"Transfer finished goods for {day_name} ship",
            })

            # MFG: Wednesday before fulfillment (cheese cutting)
            days_to_wed = (ful_date.weekday() - 2) % 7
            if days_to_wed == 0:
                days_to_wed = 7
            mfg_date = ful_date - datetime.timedelta(days=days_to_wed)
            if mfg_date >= today:
                self.action_schedule.append({
                    "date": mfg_date.isoformat(),
                    "type": "MFG",
                    "title": f"Mfg for {day_name} ship",
                    "details": "Cheese cutting & processing",
                })

            # PO order date: lead_time days before fulfillment
            lead_time = int(globals_["purchase_lt"] + globals_["shipping_lt"])
            po_date = ful_date - datetime.timedelta(days=lead_time)
            if po_date >= today:
                self.action_schedule.append({
                    "date": po_date.isoformat(),
                    "type": "PO",
                    "title": f"PO for {day_name} ship",
                    "details": (f"Order by today for {ful_str} fulfillment "
                                f"(lead time: {lead_time}d)"),
                })

        # Also add PO events for any open POs with ETAs
        for po in self.open_pos:
            eta = po.get("eta", "")
            if eta and po.get("status", "").lower() != "received":
                self.action_schedule.append({
                    "date": eta,
                    "type": "PO",
                    "title": f"PO arrive: {po['sku']}",
                    "details": (f"Qty: {po.get('qty', '?')} from "
                                f"{po.get('vendor', '?')}"),
                })

        # Add scheduled processing queue jobs
        for job in self.processing_queue:
            if job.get("status") == "scheduled":
                created = job.get("created", "")
                if created:
                    try:
                        job_date = datetime.datetime.fromisoformat(
                            created).date().isoformat()
                    except (ValueError, TypeError):
                        job_date = today.isoformat()
                else:
                    job_date = today.isoformat()
                self.action_schedule.append({
                    "date": job_date,
                    "type": "Process",
                    "title": f"Process: {job.get('sku', '?')}",
                    "details": (f"Source: {job.get('source_material', '?')}"
                                f" | Target: {job.get('target_qty', '?')}"
                                f" | {job.get('warehouse', 'Primary')}"),
                })

        self._cal_status_var.set(
            f"{len(self.action_schedule)} actions generated")
        self._render_calendar()

    # ─────────────────────────────────────────────────────────────────
    #  EX-EC EXPIRATION SUGGESTION ENGINE (Phase 4A)
    # ─────────────────────────────────────────────────────────────────

    def _show_exec_suggestions(self):
        """Suggest CH- SKUs approaching expiry for EX-EC assignment."""
        today = datetime.date.today()
        try:
            warn_days = int(self.settings_vars.get(
                "expiration_warning_days", tk.StringVar(value="14")).get())
        except (ValueError, AttributeError):
            warn_days = 14

        # Find CH- SKUs with expiration dates
        candidates = []
        for sku, inv in self.inventory.items():
            if not sku.startswith("CH-"):
                continue
            exp_dates = inv.get("expiration_dates", [])
            if not exp_dates:
                continue
            try:
                earliest_dt = datetime.date.fromisoformat(exp_dates[0])
                days_until = (earliest_dt - today).days
            except ValueError:
                continue
            if days_until > 60:  # only consider within 60 days
                continue
            qty = float(inv.get("qty", 0))
            if qty <= 0:
                continue
            candidates.append({
                "sku": sku,
                "name": inv.get("name", ""),
                "days_until_expiry": days_until,
                "qty": qty,
                "earliest_exp": exp_dates[0],
            })

        if not candidates:
            messagebox.showinfo("EX-EC Suggestions",
                                "No CH- SKUs approaching expiration.",
                                parent=self.root)
            return

        # 5-curation window duplicate prevention
        # CURATION_ORDER = ["MONG", "MDT", "OWC", "SPN", "ALPT", "ISUN",
        #                    "HHIGH"]
        # For target curation at index i, check (i-2)%7 through (i+2)%7
        def _check_conflicts(sku, target_curation):
            """Check if SKU is already used in nearby curations."""
            conflicts = []
            try:
                target_idx = CURATION_ORDER.index(target_curation)
            except ValueError:
                return conflicts
            for offset in range(-2, 3):
                check_idx = (target_idx + offset) % len(CURATION_ORDER)
                check_cur = CURATION_ORDER[check_idx]
                # Check curation recipes
                recipe_skus = [s for s, _ in self.curation_recipes.get(
                    check_cur, [])]
                if sku in recipe_skus:
                    conflicts.append(
                        f"In {check_cur} recipe")
                # Check PR-CJAM cheese assignments
                pr_entry = self.pr_cjam.get(check_cur, {})
                pr_cheese = pr_entry.get("cheese", "") if isinstance(pr_entry, dict) else pr_entry
                if pr_cheese == sku:
                    conflicts.append(
                        f"PR-CJAM in {check_cur}")
                # Check existing CEX-EC
                if self.cex_ec.get(check_cur) == sku:
                    conflicts.append(
                        f"CEX-EC in {check_cur}")
                # Check exec_assignments history
                if sku in self.exec_assignments.get(check_cur, []):
                    conflicts.append(
                        f"Previously assigned in {check_cur}")
            return conflicts

        # Sort candidates: days_until_expiry ascending, qty descending
        candidates.sort(key=lambda c: (c["days_until_expiry"], -c["qty"]))

        # Build suggestions per curation
        dlg = tk.Toplevel(self.root)
        dlg.title("EX-EC Expiration Suggestions")
        dlg.configure(bg=_BG)
        dlg.transient(self.root)
        dlg.grab_set()
        dlg.geometry("900x550")
        dlg.minsize(750, 400)

        ttk.Label(dlg, text="Suggested EX-EC Assignments (Expiring Cheese)",
                  style="Subtitle.TLabel").pack(
                      anchor="w", padx=10, pady=(10, 5))

        # Main treeview
        tree_frame = ttk.Frame(dlg)
        tree_frame.pack(fill="both", expand=True, padx=10, pady=5)

        cols = ("curation", "sku", "name", "days_exp", "qty",
                "conflicts", "status")
        tree = ttk.Treeview(tree_frame, columns=cols, show="headings",
                            height=15)
        tree.heading("curation", text="Target Curation")
        tree.heading("sku", text="SKU")
        tree.heading("name", text="Name")
        tree.heading("days_exp", text="Days to Exp")
        tree.heading("qty", text="Qty Avail")
        tree.heading("conflicts", text="Conflicts")
        tree.heading("status", text="Status")
        tree.column("curation", width=100)
        tree.column("sku", width=100)
        tree.column("name", width=150)
        tree.column("days_exp", width=80, anchor="center")
        tree.column("qty", width=70, anchor="center")
        tree.column("conflicts", width=200)
        tree.column("status", width=80, anchor="center")

        yscroll = ttk.Scrollbar(tree_frame, orient="vertical",
                                command=tree.yview)
        tree.configure(yscrollcommand=yscroll.set)
        tree.pack(side="left", fill="both", expand=True)
        yscroll.pack(side="right", fill="y")

        tree.tag_configure("ok", background=_BG2, foreground=_FG)
        tree.tag_configure("conflict", background="#7a5500",
                           foreground="white")
        tree.tag_configure("expired", background="#8b1a1a",
                           foreground="white")

        for curation in CURATION_ORDER:
            for cand in candidates:
                conflicts = _check_conflicts(cand["sku"], curation)
                status = "OK" if not conflicts else "CONFLICT"
                tag = "ok"
                if conflicts:
                    tag = "conflict"
                if cand["days_until_expiry"] < 0:
                    tag = "expired"
                    status = "EXPIRED"
                tree.insert("", "end", values=(
                    curation, cand["sku"], cand["name"],
                    cand["days_until_expiry"], cand["qty"],
                    "; ".join(conflicts) if conflicts else "None",
                    status,
                ), tags=(tag,))

        btn_frame = ttk.Frame(dlg)
        btn_frame.pack(fill="x", padx=10, pady=(5, 10))
        ttk.Button(btn_frame, text="Close",
                   command=dlg.destroy).pack(side="right")

    # ─────────────────────────────────────────────────────────────────
    #  FULFILLMENT BATCH PREVIEW (Phase 4B)
    # ─────────────────────────────────────────────────────────────────

    def _show_fulfillment_preview(self):
        """Pre-Tuesday/Saturday shortfall detection."""
        today = datetime.date.today()
        # Find next fulfillment date (Tue=1 or Sat=5)
        for offset in range(1, 8):
            d = today + datetime.timedelta(days=offset)
            if d.weekday() in (1, 5):
                next_ful = d
                break
        else:
            next_ful = today + datetime.timedelta(days=1)

        day_name = "TUE" if next_ful.weekday() == 1 else "SAT"

        # Use combined demand to estimate next fulfillment needs
        combined = self._compute_combined_demand()
        days_until_ful = (next_ful - today).days

        shortfalls = []
        for sku, demand in combined.items():
            daily = demand["daily"]
            if daily <= 0:
                continue
            on_hand = float(self.inventory.get(sku, {}).get("qty", 0))
            needed = daily * days_until_ful
            if needed > on_hand:
                shortfalls.append({
                    "sku": sku,
                    "on_hand": on_hand,
                    "needed": round(needed, 1),
                    "deficit": round(needed - on_hand, 1),
                })

        if not shortfalls:
            messagebox.showinfo(
                "Fulfillment Preview",
                f"No shortfalls expected for {day_name} {next_ful}",
                parent=self.root)
            return

        shortfalls.sort(key=lambda s: -s["deficit"])

        dlg = tk.Toplevel(self.root)
        dlg.title(f"Fulfillment Shortfall Preview — {day_name} {next_ful}")
        dlg.configure(bg=_BG)
        dlg.transient(self.root)
        dlg.grab_set()
        dlg.geometry("600x400")

        ttk.Label(dlg,
                  text=f"{len(shortfalls)} SKUs will run short by "
                       f"{day_name} {next_ful} ({days_until_ful}d away)",
                  style="Subtitle.TLabel").pack(
                      anchor="w", padx=10, pady=(10, 5))

        tree_frame = ttk.Frame(dlg)
        tree_frame.pack(fill="both", expand=True, padx=10, pady=5)

        cols = ("sku", "on_hand", "needed", "deficit")
        tree = ttk.Treeview(tree_frame, columns=cols, show="headings")
        tree.heading("sku", text="SKU")
        tree.heading("on_hand", text="On Hand")
        tree.heading("needed", text="Needed")
        tree.heading("deficit", text="Deficit")
        tree.column("sku", width=160)
        tree.column("on_hand", width=100, anchor="center")
        tree.column("needed", width=100, anchor="center")
        tree.column("deficit", width=100, anchor="center")

        yscroll = ttk.Scrollbar(tree_frame, orient="vertical",
                                command=tree.yview)
        tree.configure(yscrollcommand=yscroll.set)
        tree.pack(side="left", fill="both", expand=True)
        yscroll.pack(side="right", fill="y")

        tree.tag_configure("short", background="#7a5500", foreground="white")
        for s in shortfalls:
            tree.insert("", "end", values=(
                s["sku"], s["on_hand"], s["needed"], s["deficit"]),
                tags=("short",))

        ttk.Button(dlg, text="Close", command=dlg.destroy).pack(
            pady=(5, 10))

    # ─────────────────────────────────────────────────────────────────
    #  AUTO-RECONCILIATION (Phase 4C)
    # ─────────────────────────────────────────────────────────────────

    def _run_reconciliation(self):
        """Compare expected vs actual inventory, highlight discrepancies."""
        if not self.depletion_history:
            messagebox.showinfo(
                "Reconciliation",
                "No depletion history to reconcile against.\n"
                "Import a fulfillment depletion first, then import "
                "a new inventory CSV to compare.",
                parent=self.root)
            return

        # Expected: inventory values (current) should match
        # actual: what we'd calculate from prior on-hand minus depletions
        # For simplicity: compare current inventory against last known
        # pre-depletion values
        last_dep = self.depletion_history[-1]
        threshold_pct = 5.0

        mismatches = []
        for sku, dep_qty in last_dep.get("skus", {}).items():
            inv = self.inventory.get(sku)
            if not inv:
                continue
            actual_qty = float(inv.get("qty", 0))
            # Expected = what was there before last depletion minus depletion
            # We can't know exact pre-depletion, so flag large discrepancies
            # between current and zero (meaning we over-depleted)
            if actual_qty < 0:
                mismatches.append({
                    "sku": sku,
                    "actual": actual_qty,
                    "depleted": dep_qty,
                    "issue": "Negative inventory after depletion",
                })
            elif dep_qty > 0 and actual_qty > dep_qty * 3:
                mismatches.append({
                    "sku": sku,
                    "actual": actual_qty,
                    "depleted": dep_qty,
                    "issue": "Inventory much higher than depleted (restock?)",
                })

        # Log reconciliation
        entry = {
            "date": datetime.datetime.now().isoformat(timespec="seconds"),
            "mismatches": len(mismatches),
            "skus_checked": len(last_dep.get("skus", {})),
        }
        self.reconciliation_history.append(entry)

        if not mismatches:
            messagebox.showinfo("Reconciliation",
                                "No significant discrepancies found.",
                                parent=self.root)
            return

        dlg = tk.Toplevel(self.root)
        dlg.title("Inventory Reconciliation Results")
        dlg.configure(bg=_BG)
        dlg.transient(self.root)
        dlg.grab_set()
        dlg.geometry("600x400")

        ttk.Label(dlg,
                  text=f"{len(mismatches)} discrepancies found",
                  style="Subtitle.TLabel").pack(
                      anchor="w", padx=10, pady=(10, 5))

        tree_frame = ttk.Frame(dlg)
        tree_frame.pack(fill="both", expand=True, padx=10, pady=5)

        cols = ("sku", "actual", "depleted", "issue")
        tree = ttk.Treeview(tree_frame, columns=cols, show="headings")
        tree.heading("sku", text="SKU")
        tree.heading("actual", text="Current Qty")
        tree.heading("depleted", text="Last Depleted")
        tree.heading("issue", text="Issue")
        tree.column("sku", width=120)
        tree.column("actual", width=90, anchor="center")
        tree.column("depleted", width=90, anchor="center")
        tree.column("issue", width=250)

        yscroll = ttk.Scrollbar(tree_frame, orient="vertical",
                                command=tree.yview)
        tree.configure(yscrollcommand=yscroll.set)
        tree.pack(side="left", fill="both", expand=True)
        yscroll.pack(side="right", fill="y")

        for m in mismatches:
            tree.insert("", "end", values=(
                m["sku"], m["actual"], m["depleted"], m["issue"]))

        ttk.Button(dlg, text="Close", command=dlg.destroy).pack(
            pady=(5, 10))

    # ─────────────────────────────────────────────────────────────────
    #  CLICKUP INTEGRATION (Phase 2B)
    # ─────────────────────────────────────────────────────────────────

    def _sync_to_clickup(self):
        """Push action schedule items to ClickUp as tasks."""
        token = self.clickup_api_token
        list_id = self.clickup_list_id
        if not token or not list_id:
            messagebox.showerror(
                "ClickUp",
                "Set ClickUp API token and List ID in Settings first.",
                parent=self.root)
            return

        if not self.action_schedule:
            messagebox.showinfo("ClickUp",
                                "Generate an action schedule first.",
                                parent=self.root)
            return

        self.status_var.set("Syncing to ClickUp...")
        self.root.update_idletasks()

        def _worker():
            try:
                import requests
                session = requests.Session()
                session.headers.update({
                    "Authorization": token,
                    "Content-Type": "application/json",
                })

                created = 0
                task_ids = []  # for dependency linking
                tag_map = {"PO": "po", "MFG": "manufacturing",
                           "Transfer": "transfer",
                           "Fulfillment": "fulfillment",
                           "Crossdock": "crossdock"}

                for item in self.action_schedule:
                    # Convert date to millisecond timestamp
                    try:
                        dt = datetime.datetime.fromisoformat(item["date"])
                        ts_ms = int(dt.timestamp() * 1000)
                    except (ValueError, AttributeError):
                        continue

                    payload = {
                        "name": f"[{item['type']}] {item['title']}",
                        "description": item.get("details", ""),
                        "due_date": ts_ms,
                        "start_date": ts_ms,
                        "tags": [tag_map.get(item["type"], "other")],
                    }

                    resp = session.post(
                        f"https://api.clickup.com/api/v2/list/{list_id}/task",
                        json=payload, timeout=15)
                    if resp.status_code in (200, 201):
                        task_data = resp.json()
                        task_ids.append(task_data.get("id"))
                        created += 1

                # Link dependencies: PO → MFG → Fulfillment chains
                # Group tasks by date, link sequential types
                # (simplified: just link consecutive tasks on same date)

                self.root.after(0, lambda: self._on_clickup_done(created))
            except Exception as e:
                self.root.after(
                    0, lambda: self._on_clickup_error(str(e)))

        thread = threading.Thread(target=_worker, daemon=True)
        thread.start()

    def _on_clickup_done(self, count):
        self.status_var.set(f"ClickUp: {count} tasks created")
        messagebox.showinfo("ClickUp Sync",
                            f"Created {count} tasks in ClickUp.",
                            parent=self.root)

    def _on_clickup_error(self, error_msg):
        self.status_var.set("ClickUp sync failed")
        messagebox.showerror("ClickUp Error",
                             f"Failed to sync:\n{error_msg}",
                             parent=self.root)

    # ─────────────────────────────────────────────────────────────────
    #  GOOGLE CALENDAR INTEGRATION (Phase 2C)
    # ─────────────────────────────────────────────────────────────────

    def _sync_to_gcal(self):
        """Push action schedule items to Google Calendar."""
        if not self.gcal_client_id or not self.gcal_client_secret:
            messagebox.showerror(
                "Google Calendar",
                "Set Google Calendar Client ID and Secret in Settings first.",
                parent=self.root)
            return

        if not self.action_schedule:
            messagebox.showinfo("Google Calendar",
                                "Generate an action schedule first.",
                                parent=self.root)
            return

        self.status_var.set("Syncing to Google Calendar...")
        self.root.update_idletasks()

        def _worker():
            try:
                import requests

                # If we have a refresh token, use it; otherwise start OAuth
                if not self.gcal_refresh_token:
                    self.root.after(0, self._gcal_oauth_flow)
                    return

                # Refresh access token
                token_resp = requests.post(
                    "https://oauth2.googleapis.com/token",
                    data={
                        "client_id": self.gcal_client_id,
                        "client_secret": self.gcal_client_secret,
                        "refresh_token": self.gcal_refresh_token,
                        "grant_type": "refresh_token",
                    }, timeout=15)
                token_resp.raise_for_status()
                access_token = token_resp.json()["access_token"]

                session = requests.Session()
                session.headers.update({
                    "Authorization": f"Bearer {access_token}",
                    "Content-Type": "application/json",
                })

                created = 0
                color_map = {"PO": "9", "MFG": "6", "Transfer": "3",
                             "Fulfillment": "10", "Crossdock": "11"}

                for item in self.action_schedule:
                    event = {
                        "summary": f"[{item['type']}] {item['title']}",
                        "description": item.get("details", ""),
                        "start": {"date": item["date"]},
                        "end": {"date": item["date"]},
                        "colorId": color_map.get(item["type"], "1"),
                        "reminders": {
                            "useDefault": False,
                            "overrides": [
                                {"method": "popup", "minutes": 60},
                            ],
                        },
                    }

                    resp = session.post(
                        "https://www.googleapis.com/calendar/v3/"
                        "calendars/primary/events",
                        json=event, timeout=15)
                    if resp.status_code in (200, 201):
                        created += 1

                self.root.after(0, lambda: self._on_gcal_done(created))
            except Exception as e:
                self.root.after(
                    0, lambda: self._on_gcal_error(str(e)))

        thread = threading.Thread(target=_worker, daemon=True)
        thread.start()

    def _gcal_oauth_flow(self):
        """Run OAuth2 flow for Google Calendar."""
        import webbrowser
        from http.server import HTTPServer, BaseHTTPRequestHandler
        from urllib.parse import urlencode, urlparse, parse_qs

        port = 21849
        redirect_uri = f"http://localhost:{port}/callback"

        auth_params = urlencode({
            "client_id": self.gcal_client_id,
            "redirect_uri": redirect_uri,
            "response_type": "code",
            "scope": "https://www.googleapis.com/auth/calendar.events",
            "access_type": "offline",
            "prompt": "consent",
        })
        auth_url = f"https://accounts.google.com/o/oauth2/v2/auth?{auth_params}"

        app_ref = self

        class CallbackHandler(BaseHTTPRequestHandler):
            def do_GET(self):
                parsed = urlparse(self.path)
                params = parse_qs(parsed.query)
                if "code" not in params:
                    self.send_response(400)
                    self.send_header("Content-Type", "text/html")
                    self.end_headers()
                    self.wfile.write(b"<h2>No code received.</h2>")
                    return

                code = params["code"][0]
                try:
                    import requests
                    resp = requests.post(
                        "https://oauth2.googleapis.com/token",
                        data={
                            "client_id": app_ref.gcal_client_id,
                            "client_secret": app_ref.gcal_client_secret,
                            "code": code,
                            "grant_type": "authorization_code",
                            "redirect_uri": redirect_uri,
                        }, timeout=15)
                    resp.raise_for_status()
                    tokens = resp.json()
                    app_ref.gcal_refresh_token = tokens.get(
                        "refresh_token", "")
                    self.send_response(200)
                    self.send_header("Content-Type", "text/html")
                    self.end_headers()
                    self.wfile.write(
                        b"<h2>Google Calendar authorized!</h2>"
                        b"<p>Return to the app and sync again.</p>")
                except Exception as e:
                    self.send_response(500)
                    self.send_header("Content-Type", "text/html")
                    self.end_headers()
                    self.wfile.write(
                        f"<h2>Error: {e}</h2>".encode())

            def log_message(self, format, *args):
                pass

        def _run_server():
            server = HTTPServer(("localhost", port), CallbackHandler)
            server.timeout = 120
            server.handle_request()
            server.server_close()
            app_ref.root.after(0, lambda: app_ref.status_var.set(
                "Google Calendar authorized" if app_ref.gcal_refresh_token
                else "Google Calendar auth failed"))

        threading.Thread(target=_run_server, daemon=True).start()
        webbrowser.open(auth_url)
        self.status_var.set("Waiting for Google Calendar authorization...")

    def _on_gcal_done(self, count):
        self.status_var.set(f"Google Calendar: {count} events created")
        messagebox.showinfo("Google Calendar Sync",
                            f"Created {count} events.",
                            parent=self.root)

    def _on_gcal_error(self, error_msg):
        self.status_var.set("Google Calendar sync failed")
        messagebox.showerror("Google Calendar Error",
                             f"Failed to sync:\n{error_msg}",
                             parent=self.root)

    # ─────────────────────────────────────────────────────────────────
    #  DROPBOX INTEGRATION (Phase 3A)
    # ─────────────────────────────────────────────────────────────────

    def _check_dropbox(self):
        """Poll Dropbox for new inventory CSV files."""
        if not self.dropbox_app_key or not self.dropbox_app_secret:
            messagebox.showerror(
                "Dropbox",
                "Set Dropbox App Key and Secret in Settings first.",
                parent=self.root)
            return

        if not self.dropbox_refresh_token:
            self._dropbox_oauth_flow()
            return

        self.status_var.set("Checking Dropbox...")
        self.root.update_idletasks()

        def _worker():
            try:
                import requests

                # Refresh access token
                token_resp = requests.post(
                    "https://api.dropboxapi.com/oauth2/token",
                    data={
                        "grant_type": "refresh_token",
                        "refresh_token": self.dropbox_refresh_token,
                        "client_id": self.dropbox_app_key,
                        "client_secret": self.dropbox_app_secret,
                    }, timeout=15)
                token_resp.raise_for_status()
                access_token = token_resp.json()["access_token"]

                # List folder — use shared link API if configured
                shared_link = self.dropbox_shared_link.strip()
                list_body = {"path": "", "recursive": False} \
                    if shared_link else \
                    {"path": "/!AppyHour_SHARED/Product Inventory",
                     "recursive": False}
                if shared_link:
                    list_body["shared_link"] = {"url": shared_link}

                resp = requests.post(
                    "https://api.dropboxapi.com/2/files/list_folder",
                    headers={
                        "Authorization": f"Bearer {access_token}",
                        "Content-Type": "application/json",
                    },
                    json=list_body, timeout=15)
                resp.raise_for_status()

                entries = resp.json().get("entries", [])
                inv_files = [
                    e for e in entries
                    if e.get("name", "").lower().endswith(
                        (".csv", ".xlsx"))
                ]

                if not inv_files:
                    self.root.after(0, lambda: messagebox.showinfo(
                        "Dropbox",
                        "No inventory files found in Product Inventory.",
                        parent=self.root))
                    return

                # Sort by modified date, newest first
                inv_files.sort(
                    key=lambda e: e.get("server_modified", ""),
                    reverse=True)

                # Offer to download the newest
                newest = inv_files[0]
                name = newest["name"]
                fpath = newest.get("path_lower", "/" + newest["name"])

                self.root.after(0, lambda: self._offer_dropbox_import(
                    access_token, name, fpath,
                    shared_link=shared_link or None))

            except Exception as e:
                self.root.after(
                    0, lambda: self._on_dropbox_error(str(e)))

        thread = threading.Thread(target=_worker, daemon=True)
        thread.start()

    def _offer_dropbox_import(self, access_token, name, path,
                              shared_link=None):
        if not messagebox.askyesno(
                "Dropbox",
                f"Found: {name}\n\nDownload and import?",
                parent=self.root):
            return

        def _download():
            try:
                import requests
                import tempfile

                if shared_link:
                    # Use shared link download endpoint
                    api_arg = {"url": shared_link,
                               "path": "/" + name}
                    resp = requests.post(
                        "https://content.dropboxapi.com"
                        "/2/sharing/get_shared_link_file",
                        headers={
                            "Authorization": f"Bearer {access_token}",
                            "Dropbox-API-Arg": json.dumps(api_arg),
                        }, timeout=30)
                else:
                    resp = requests.post(
                        "https://content.dropboxapi.com/2/files/download",
                        headers={
                            "Authorization": f"Bearer {access_token}",
                            "Dropbox-API-Arg": json.dumps({"path": path}),
                        }, timeout=30)
                resp.raise_for_status()

                # Save to temp file
                suffix = ".xlsx" if name.lower().endswith(".xlsx") \
                    else ".csv"
                tmp = tempfile.NamedTemporaryFile(
                    suffix=suffix, delete=False, mode="wb")
                tmp.write(resp.content)
                tmp.close()

                self.root.after(
                    0, lambda: self._import_dropbox_csv(tmp.name, name))
            except Exception as e:
                self.root.after(
                    0, lambda: self._on_dropbox_error(str(e)))

        threading.Thread(target=_download, daemon=True).start()

    def _import_dropbox_csv(self, temp_path, original_name):
        """Import a CSV or XLSX downloaded from Dropbox."""
        try:
            if temp_path.lower().endswith(".xlsx"):
                import openpyxl
                wb = openpyxl.load_workbook(temp_path, data_only=True)
                ws = wb.active
                all_rows = list(ws.iter_rows(values_only=True))
                headers = [str(c) if c is not None else ""
                           for c in all_rows[0]]
                rows_raw = [
                    [str(c) if c is not None else "" for c in row]
                    for row in all_rows[1:]
                ]
            else:
                with open(temp_path, "r", newline="",
                          encoding="utf-8-sig") as f:
                    reader = csv.reader(f)
                    headers = next(reader)
                    rows_raw = list(reader)
        except Exception as e:
            messagebox.showerror("Import Error",
                                 f"Failed to read file:\n{e}",
                                 parent=self.root)
            return
        finally:
            try:
                os.unlink(temp_path)
            except OSError:
                pass

        if not headers:
            return

        dlg = ColumnMappingDialog(self.root, headers)
        self.root.wait_window(dlg)
        if not dlg.result:
            return

        mapping = dlg.result
        self.last_csv_mapping = mapping

        # Save pre-import inventory for yield reconciliation
        pre_import_qty = {
            sku: float(data.get("qty", 0))
            for sku, data in self.inventory.items()
            if sku.startswith("CH-")
        }

        imported = 0
        for row in rows_raw:
            if len(row) < len(headers):
                row.extend([""] * (len(headers) - len(row)))
            row_dict = dict(zip(headers, row))

            sku = row_dict.get(mapping["SKU"], "").strip()
            if not sku:
                continue

            qty_str = row_dict.get(
                mapping["Quantity On Hand"], "0").strip()
            try:
                qty = float(qty_str.replace(",", ""))
            except ValueError:
                qty = 0

            entry = {"qty": qty}
            if "Product Name" in mapping:
                entry["name"] = row_dict.get(mapping["Product Name"], "")
            # Preserve existing fields if updating
            if sku in self.inventory:
                old = self.inventory[sku]
                entry["warehouse"] = old.get("warehouse", "Primary")
                for keep in ("category", "unit", "unit_size",
                             "unit_cost", "expiration_dates"):
                    if keep in old and keep not in entry:
                        entry[keep] = old[keep]

            # Parse warehouse columns (RMFG = Primary, WIP = pending)
            rmfg_str = row_dict.get("RMFG", "").strip()
            wip_str = row_dict.get("WIP", "").strip()
            if rmfg_str or wip_str:
                rmfg_qty = 0
                wip_qty = 0
                try:
                    rmfg_qty = float(rmfg_str) if rmfg_str else 0
                except ValueError:
                    pass
                try:
                    wip_qty = float(wip_str) if wip_str else 0
                except ValueError:
                    pass
                if rmfg_qty > 0 or wip_qty > 0:
                    wh_qty = {}
                    if rmfg_qty > 0:
                        wh_qty["Primary"] = rmfg_qty
                    entry["warehouse_qty"] = wh_qty
                    entry["qty"] = rmfg_qty
                    entry["warehouse"] = "Primary"
                    if wip_qty > 0:
                        entry["wip_qty"] = wip_qty

            self.inventory[sku] = entry
            imported += 1

        self._refresh_inventory_tree()
        self._recalculate()

        # Yield reconciliation: compare pre-import qty vs snapshot
        snapshot_inv = {
            sku: float(data.get("qty", 0))
            for sku, data in self.inventory.items()
            if sku.startswith("CH-")
        }
        flags = self._reconcile_yield_vs_snapshot(
            snapshot_inv, pre_import_qty)

        flag_msg = f" ({flags} yield flags)" if flags else ""
        self.status_var.set(
            f"Dropbox import: {imported} SKUs from {original_name}"
            f"{flag_msg}")
        messagebox.showinfo("Dropbox Import",
                            f"Imported {imported} SKUs from "
                            f"{original_name}.{flag_msg}",
                            parent=self.root)

    def _dropbox_oauth_flow(self):
        """Run OAuth2 flow for Dropbox."""
        import webbrowser
        from http.server import HTTPServer, BaseHTTPRequestHandler
        from urllib.parse import urlencode, urlparse, parse_qs

        port = 21849
        redirect_uri = f"http://localhost:{port}/callback"

        auth_params = urlencode({
            "client_id": self.dropbox_app_key,
            "redirect_uri": redirect_uri,
            "response_type": "code",
            "token_access_type": "offline",
        })
        auth_url = (f"https://www.dropbox.com/oauth2/authorize?"
                    f"{auth_params}")

        app_ref = self

        class CallbackHandler(BaseHTTPRequestHandler):
            def do_GET(self):
                parsed = urlparse(self.path)
                params = parse_qs(parsed.query)
                if "code" not in params:
                    self.send_response(400)
                    self.send_header("Content-Type", "text/html")
                    self.end_headers()
                    self.wfile.write(b"<h2>No code received.</h2>")
                    return

                code = params["code"][0]
                try:
                    import requests
                    resp = requests.post(
                        "https://api.dropboxapi.com/oauth2/token",
                        data={
                            "code": code,
                            "grant_type": "authorization_code",
                            "client_id": app_ref.dropbox_app_key,
                            "client_secret": app_ref.dropbox_app_secret,
                            "redirect_uri": redirect_uri,
                        }, timeout=15)
                    resp.raise_for_status()
                    tokens = resp.json()
                    app_ref.dropbox_refresh_token = tokens.get(
                        "refresh_token", "")
                    self.send_response(200)
                    self.send_header("Content-Type", "text/html")
                    self.end_headers()
                    self.wfile.write(
                        b"<h2>Dropbox authorized!</h2>"
                        b"<p>Return to the app and check Dropbox again.</p>")
                except Exception as e:
                    self.send_response(500)
                    self.send_header("Content-Type", "text/html")
                    self.end_headers()
                    self.wfile.write(
                        f"<h2>Error: {e}</h2>".encode())

            def log_message(self, format, *args):
                pass

        def _run_server():
            server = HTTPServer(("localhost", port), CallbackHandler)
            server.timeout = 120
            server.handle_request()
            server.server_close()
            app_ref.root.after(0, lambda: app_ref.status_var.set(
                "Dropbox authorized" if app_ref.dropbox_refresh_token
                else "Dropbox auth failed"))

        threading.Thread(target=_run_server, daemon=True).start()
        webbrowser.open(auth_url)
        self.status_var.set("Waiting for Dropbox authorization...")

    def _on_dropbox_error(self, error_msg):
        self.status_var.set("Dropbox check failed")
        messagebox.showerror("Dropbox Error",
                             f"Failed:\n{error_msg}",
                             parent=self.root)

    # ═════════════════════════════════════════════════════════════════
    #  v2.5 AUTOMATION ENGINE
    # ═════════════════════════════════════════════════════════════════

    def _startup_automation(self):
        """Run all startup automation tasks."""
        # 1. Auto-detect newest inventory CSV
        self._auto_import_inventory_csv()

        # 2. Start file watcher for Shipments folder
        self._start_file_watcher()

        # 3. Start auto-refresh timer
        self._start_auto_refresh()

        # 4. Start webhook server if configured
        self._start_webhook_server()

        # 5. Start dashboard auto-refresh (every 5 min)
        self.root.after(300000, self._schedule_dashboard_refresh)

    # ─────────────────────────────────────────────────────────────────
    #  AUTO-IMPORT INVENTORY CSV ON STARTUP
    # ─────────────────────────────────────────────────────────────────

    def _auto_import_inventory_csv(self):
        """Check Dropbox first, then local folder for newest inventory CSV."""
        # Try Dropbox first if configured
        if self.dropbox_refresh_token and self.dropbox_app_key:
            self._auto_import_from_dropbox()
            return

        # Fall back to local file detection
        self._auto_import_local_csv()

    def _auto_import_local_csv(self):
        """Detect newest local Product Inventory_*.csv and offer to import."""
        script_dir = os.path.dirname(
            os.path.abspath(sys.argv[0] if sys.argv[0] else __file__))
        candidates = []
        for fn in os.listdir(script_dir):
            if (fn.lower().startswith("product inventory") and
                    fn.lower().endswith(".csv")):
                fpath = os.path.join(script_dir, fn)
                candidates.append((os.path.getmtime(fpath), fn, fpath))

        if not candidates:
            return

        candidates.sort(reverse=True)
        newest_mtime, newest_name, newest_path = candidates[0]

        last_import = self.saved.get("_last_inventory_import", "")
        if last_import == newest_name:
            return

        if messagebox.askyesno(
                "New Inventory File Detected",
                f"Found: {newest_name}\n"
                f"Modified: {datetime.datetime.fromtimestamp(newest_mtime):%Y-%m-%d %H:%M}\n\n"
                f"Import this inventory file?",
                parent=self.root):
            self._do_import_inventory_csv(newest_path)
            self.saved["_last_inventory_import"] = newest_name

    def _auto_import_from_dropbox(self):
        """Check Dropbox for newest inventory CSV and offer to import."""
        def _worker():
            try:
                import requests
                # Refresh access token
                resp = requests.post(
                    "https://api.dropboxapi.com/oauth2/token",
                    data={
                        "grant_type": "refresh_token",
                        "refresh_token": self.dropbox_refresh_token,
                        "client_id": self.dropbox_app_key,
                        "client_secret": self.dropbox_app_secret,
                    }, timeout=15)
                resp.raise_for_status()
                access_token = resp.json()["access_token"]

                # List files — use shared link API if configured
                shared_link = self.dropbox_shared_link.strip()
                list_body = {"path": "", "recursive": False} \
                    if shared_link else \
                    {"path": "/!AppyHour_SHARED/Product Inventory"}
                if shared_link:
                    list_body["shared_link"] = {"url": shared_link}

                resp = requests.post(
                    "https://api.dropboxapi.com/2/files/list_folder",
                    headers={
                        "Authorization": f"Bearer {access_token}",
                        "Content-Type": "application/json",
                    },
                    json=list_body,
                    timeout=15)
                resp.raise_for_status()
                entries = resp.json().get("entries", [])

                # Find inventory files (CSV or XLSX)
                inv_files = [
                    e for e in entries
                    if e.get("name", "").lower().endswith(
                        (".csv", ".xlsx")) and
                    "product inventory" in e.get("name", "").lower()
                ]

                if not inv_files:
                    # No Dropbox files, fall back to local
                    self.root.after(0, self._auto_import_local_csv)
                    return

                # Sort by server_modified (newest first)
                inv_files.sort(
                    key=lambda e: e.get("server_modified", ""),
                    reverse=True)
                newest = inv_files[0]
                newest_name = newest["name"]
                newest_path = newest.get("path_lower",
                                        "/" + newest["name"])
                newest_modified = newest.get("server_modified", "")

                # Check if already imported
                last_import = self.saved.get("_last_inventory_import", "")
                if last_import == newest_name:
                    return

                # Download the file
                if shared_link:
                    api_arg = {"url": shared_link,
                               "path": "/" + newest_name}
                    resp = requests.post(
                        "https://content.dropboxapi.com"
                        "/2/sharing/get_shared_link_file",
                        headers={
                            "Authorization": f"Bearer {access_token}",
                            "Dropbox-API-Arg": json.dumps(api_arg),
                        }, timeout=30)
                else:
                    resp = requests.post(
                        "https://content.dropboxapi.com/2/files/download",
                        headers={
                            "Authorization": f"Bearer {access_token}",
                            "Dropbox-API-Arg": json.dumps(
                                {"path": newest_path}),
                        }, timeout=30)
                resp.raise_for_status()

                # Save to temp file
                temp_path = os.path.join(
                    tempfile.gettempdir(), newest_name)
                with open(temp_path, "wb") as f:
                    f.write(resp.content)

                def _offer():
                    if messagebox.askyesno(
                            "New Inventory File on Dropbox",
                            f"Found on Dropbox: {newest_name}\n"
                            f"Modified: {newest_modified}\n\n"
                            f"Import this inventory file?",
                            parent=self.root):
                        self._do_import_inventory_csv(temp_path)
                        self.saved["_last_inventory_import"] = newest_name

                self.root.after(0, _offer)

            except Exception as e:
                # Dropbox failed, fall back to local
                self.root.after(0, self._auto_import_local_csv)

        threading.Thread(target=_worker, daemon=True).start()

    def _do_import_inventory_csv(self, path):
        """Import a specific inventory CSV file (reuses existing logic)."""
        self._pending_csv_path = path
        try:
            self._import_inventory_csv()
        finally:
            self._pending_csv_path = None

    # ─────────────────────────────────────────────────────────────────
    #  FILE WATCHER (Shipments folder)
    # ─────────────────────────────────────────────────────────────────

    def _start_file_watcher(self):
        """Start polling Shipments/ folder for new depletion files."""
        script_dir = os.path.dirname(
            os.path.abspath(sys.argv[0] if sys.argv[0] else __file__))
        ship_dir = os.path.join(script_dir, "Shipments")

        if not os.path.isdir(ship_dir):
            return

        # Snapshot current files
        self._ship_dir = ship_dir
        self._last_shipments_files = set(os.listdir(ship_dir))
        self._file_watcher_active = True
        self._poll_shipments()

    def _poll_shipments(self):
        """Poll Shipments folder every 30 seconds for new xlsx files."""
        if not self._file_watcher_active:
            return

        try:
            current_files = set(os.listdir(self._ship_dir))
            new_files = current_files - self._last_shipments_files
            new_xlsx = [f for f in new_files
                        if f.lower().endswith(".xlsx") and
                        "productionquery" in f.lower()]

            if new_xlsx:
                self._last_shipments_files = current_files
                newest = sorted(new_xlsx)[-1]
                full_path = os.path.join(self._ship_dir, newest)
                self.root.after(0, lambda p=full_path, n=newest:
                    self._offer_auto_depletion(p, n))
            else:
                self._last_shipments_files = current_files
        except Exception:
            pass

        # Re-schedule
        self.root.after(30000, self._poll_shipments)

    def _offer_auto_depletion(self, path, filename):
        """Offer to import a newly detected depletion file."""
        if messagebox.askyesno(
                "New Depletion File Detected",
                f"New fulfillment file detected:\n{filename}\n\n"
                f"Import and apply depletion now?",
                parent=self.root):
            self._pending_depletion_path = path
            try:
                self._import_depletion_matrix()
            finally:
                self._pending_depletion_path = None

    # ─────────────────────────────────────────────────────────────────
    #  AUTO-REFRESH TIMER
    # ─────────────────────────────────────────────────────────────────

    def _start_auto_refresh(self):
        """Start periodic auto-refresh of demand data."""
        interval = self.auto_refresh_interval
        if interval <= 0:
            return
        self._schedule_auto_refresh()

    def _schedule_auto_refresh(self):
        """Schedule next auto-refresh cycle."""
        interval = self.auto_refresh_interval
        if interval <= 0:
            return
        self.root.after(interval * 60 * 1000, self._do_auto_refresh)

    def _do_auto_refresh(self):
        """Execute one auto-refresh cycle."""
        self.status_var.set("Auto-refreshing...")

        def _refresh_worker():
            errors = []
            # Pull Recharge if token configured
            try:
                token = self.recharge_token_var.get().strip()
                if token:
                    import requests
                    client = RechargeClient(token)
                    subs = client.get_active_subscriptions()
                    demand = client.aggregate_sku_quantities(subs)
                    weekly = {k: round(v, 2) for k, v in demand.items()}
                    api_cohorts = (
                        RechargeClient.build_cohorts_from_subscriptions(subs))
                    charges = client.get_queued_charges()
                    queued_by_month = client.aggregate_charges_by_month(
                        charges)
                    queued_resolved = resolve_queued_charges(charges)
                    self.root.after(0, lambda: self._apply_recharge_refresh(
                        weekly, queued_by_month, api_cohorts,
                        queued_resolved))
            except Exception as e:
                errors.append(f"Recharge: {e}")

            # Pull Shopify if configured
            try:
                store_url = self.shopify_store_var.get().strip()
                shop_token = self.shopify_token_var.get().strip()
                if store_url and shop_token:
                    import requests
                    from collections import Counter
                    client = ShopifyClient(
                        store_url, shop_token,
                        self.shopify_api_key_var.get().strip(),
                        self.shopify_api_secret_var.get().strip())
                    tag = self.shopify_tag_var.get().strip()
                    weeks = int(self.shopify_weeks_var.get() or "8")
                    orders = client.get_orders(tag_filter=tag,
                                               weeks_back=weeks)
                    trend = client.aggregate_with_trend(orders, weeks)
                    self.root.after(0, lambda: self._apply_shopify_refresh(
                        trend))
            except Exception as e:
                errors.append(f"Shopify: {e}")

            # Check Dropbox for new inventory
            try:
                if self.dropbox_refresh_token and self.dropbox_app_key:
                    self.root.after(0, self._check_dropbox)
            except Exception as e:
                errors.append(f"Dropbox: {e}")

            def _finish():
                self._recalculate()
                if errors:
                    self.status_var.set(
                        f"Auto-refresh done with errors: "
                        f"{'; '.join(errors)}")
                else:
                    self.status_var.set(
                        f"Auto-refreshed  |  "
                        f"{datetime.datetime.now().strftime('%H:%M:%S')}")

                # Check for auto-PO triggers
                self._check_auto_po_trigger()

                # Auto-sync if enabled
                self._auto_sync_integrations()

                # Send Slack alerts if needed
                self._send_slack_alerts()

            self.root.after(0, _finish)

        threading.Thread(target=_refresh_worker, daemon=True).start()

        # Re-schedule
        self._schedule_auto_refresh()

    def _apply_recharge_refresh(self, demand, queued_by_month,
                               api_cohorts=None, queued_resolved=None):
        """Apply auto-refreshed Recharge data."""
        self.recharge_demand = demand
        if queued_by_month:
            self.recharge_queued = queued_by_month
        if queued_resolved:
            self.recharge_queued_resolved = queued_resolved
        if api_cohorts:
            existing = {c["start_month"]: i
                        for i, c in enumerate(self.cohorts)}
            for ac in api_cohorts:
                month = ac["start_month"]
                if month in existing:
                    self.cohorts[existing[month]]["size"] = ac["size"]
                else:
                    self.cohorts.append(ac)

    def _apply_shopify_refresh(self, trend):
        """Apply auto-refreshed Shopify data."""
        api_demand = {}
        trend_data = {}
        for sku, info in trend.items():
            api_demand[sku] = info.get("projected_next_week", 0)
            trend_data[sku] = info
        self.shopify_api_demand = api_demand
        self.shopify_trend_data = trend_data

    # ─────────────────────────────────────────────────────────────────
    #  SLACK NOTIFICATIONS
    # ─────────────────────────────────────────────────────────────────

    def _send_slack_alerts(self):
        """Send critical alerts to Slack via incoming webhook."""
        if not self.slack_webhook_url:
            return

        self._generate_alerts()
        if not self.alerts:
            return

        # Filter by user preferences
        filtered = []
        for alert in self.alerts:
            if (alert["severity"] == "critical" and
                    alert["category"] == "reorder" and
                    self.slack_notify_critical):
                filtered.append(alert)
            elif (alert["category"] == "expiring" and
                  self.slack_notify_expiring):
                filtered.append(alert)
            elif (alert["category"] == "fulfillment" and
                  self.slack_notify_shortfall):
                filtered.append(alert)

        if not filtered:
            return

        # Build Slack message
        blocks = []
        blocks.append({
            "type": "header",
            "text": {
                "type": "plain_text",
                "text": f"Inventory Alerts — "
                        f"{datetime.datetime.now():%Y-%m-%d %H:%M}"
            }
        })

        # Group by category
        by_cat = defaultdict(list)
        for a in filtered:
            by_cat[a["category"]].append(a)

        cat_emoji = {"reorder": ":rotating_light:",
                     "expiring": ":warning:",
                     "fulfillment": ":package:"}

        for cat, alerts_list in by_cat.items():
            emoji = cat_emoji.get(cat, ":bell:")
            lines = [f"• {a['message']}" for a in alerts_list[:15]]
            if len(alerts_list) > 15:
                lines.append(f"_...and {len(alerts_list) - 15} more_")
            blocks.append({
                "type": "section",
                "text": {
                    "type": "mrkdwn",
                    "text": f"{emoji} *{cat.upper()}*\n" + "\n".join(lines)
                }
            })

        payload = {"blocks": blocks}

        def _post():
            try:
                import requests
                requests.post(self.slack_webhook_url,
                              json=payload, timeout=10)
            except Exception:
                pass

        threading.Thread(target=_post, daemon=True).start()

    def _send_slack_message(self, text):
        """Send a simple text message to Slack."""
        if not self.slack_webhook_url:
            return

        def _post():
            try:
                import requests
                requests.post(self.slack_webhook_url,
                              json={"text": text}, timeout=10)
            except Exception:
                pass

        threading.Thread(target=_post, daemon=True).start()

    # ─────────────────────────────────────────────────────────────────
    #  ONE-CLICK DEPLETION + EMAIL
    # ─────────────────────────────────────────────────────────────────

    def _deplete_and_email(self):
        """Single action: apply depletion from xlsx, then email that same
        xlsx file to the configured recipient."""
        # Track history length to detect if depletion was actually applied
        history_len_before = len(self.depletion_history)

        # Run normal depletion (tracks the xlsx path used)
        self._depletion_source_path = None
        self._import_depletion_matrix()

        # Check if a new depletion was actually applied
        if len(self.depletion_history) <= history_len_before:
            return  # User cancelled or no depletion applied

        last = self.depletion_history[-1]
        xlsx_path = self._depletion_source_path

        if not xlsx_path or not os.path.exists(xlsx_path):
            messagebox.showwarning(
                "Deplete & Email",
                "Depletion applied but could not locate the source file "
                "for emailing.",
                parent=self.root)
            return

        # Email the original xlsx
        self._email_depletion_file(xlsx_path, last)

    def _email_depletion_file(self, xlsx_path, depletion_entry):
        """Show compose dialog then email the original depletion xlsx."""
        from_addr = self.depletion_email_from or self.smtp_user
        to_addrs = self.depletion_email_to

        if not self.smtp_user:
            messagebox.showinfo(
                "Email Not Configured",
                "Depletion applied successfully.\n\n"
                "To enable auto-email, configure SMTP settings in\n"
                "Settings > Email (Depletion Reports).",
                parent=self.root)
            return

        total_orders = depletion_entry.get("total_orders", 0)
        day = depletion_entry.get("day", "ALL")
        filename = os.path.basename(xlsx_path)

        # Build default subject: RMFG_YYYYMMDD // N Orders
        tag_prefix = self.saved.get("depletion_tag_prefix", "RMFG")
        date_tag = datetime.datetime.now().strftime("%Y%m%d")
        default_subject = (
            f"{tag_prefix}_{date_tag} // {total_orders} Orders")

        # Show compose dialog
        dlg = _DepletionEmailDialog(
            self.root, default_subject, to_addrs, from_addr,
            filename, depletion_entry)
        self.root.wait_window(dlg)

        if not dlg.result:
            return  # User cancelled

        subject = dlg.result["subject"]
        to_addr = dlg.result["to"]
        body_text = dlg.result["body"]

        msg = MIMEMultipart()
        msg["From"] = from_addr
        msg["To"] = to_addr
        msg["Subject"] = subject

        msg.attach(MIMEText(body_text, "plain"))

        # Attach the original xlsx
        with open(xlsx_path, "rb") as f:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(f.read())
            encoders.encode_base64(part)
            part.add_header(
                "Content-Disposition",
                f"attachment; filename={filename}")
            msg.attach(part)

        recipients = [a.strip() for a in to_addr.split(",") if a.strip()]

        def _send():
            try:
                server = smtplib.SMTP(self.smtp_host, int(self.smtp_port))
                server.starttls()
                server.login(self.smtp_user, self.smtp_password)
                server.sendmail(from_addr, recipients, msg.as_string())
                server.quit()
                self.root.after(0, lambda: self.status_var.set(
                    f"Depletion email sent to {len(recipients)} "
                    f"recipient(s)"))
                self._send_slack_message(
                    f":white_check_mark: Depletion report sent — "
                    f"{subject}")
            except Exception as e:
                self.root.after(0, lambda: messagebox.showerror(
                    "Email Error",
                    f"Failed to send depletion email:\n{e}",
                    parent=self.root))

        threading.Thread(target=_send, daemon=True).start()
        self.status_var.set(
            f"Sending depletion email to {len(recipients)} recipient(s)...")

    # ─────────────────────────────────────────────────────────────────
    #  SMART AUTO-PO TRIGGER
    # ─────────────────────────────────────────────────────────────────

    def _check_auto_po_trigger(self):
        """Check if forecast deficits warrant auto-PO generation."""
        threshold = self.auto_po_threshold
        if threshold <= 0 or not self.last_forecast:
            return

        wheel_supply = compute_wheel_supply(
            self.wheel_inventory, self.adjusted_conversion_factors)
        bulk_supply = compute_bulk_supply(
            self._primary_inventory(), self.bulk_conversions)
        total_critical = 0
        for month_label, month_data in self.last_forecast.items():
            alerts = compute_reorder_alerts(
                month_data, self.inventory, self.open_pos, wheel_supply,
                bulk_supply=bulk_supply)
            for alert in alerts:
                if alert["deficit"] > threshold:
                    total_critical += 1

        if total_critical > 0:
            self._send_slack_message(
                f":rotating_light: Auto-PO trigger: {total_critical} SKUs "
                f"exceed deficit threshold ({threshold} units). "
                f"Review Auto-PO Generator in the app.")

            # Auto-generate PO preview if we have vendor catalog data
            if self.vendor_catalog:
                self.root.after(500, self._generate_auto_po)

    # ─────────────────────────────────────────────────────────────────
    #  AUTO-SYNC INTEGRATIONS
    # ─────────────────────────────────────────────────────────────────

    def _auto_sync_integrations(self):
        """Auto-sync action schedule to ClickUp/GCal if enabled."""
        if not self.action_schedule:
            self._generate_action_schedule()

        if self.auto_sync_clickup and self.clickup_api_token:
            self._sync_to_clickup()

        if self.auto_sync_gcal and self.gcal_refresh_token:
            self._sync_to_gcal()

    # ─────────────────────────────────────────────────────────────────
    #  WEBHOOK SERVER (Shopify + Recharge)
    # ─────────────────────────────────────────────────────────────────

    def _start_webhook_server(self):
        """Start a lightweight HTTP server for Shopify/Recharge webhooks."""
        if not (self.webhook_secret_shopify or
                self.webhook_secret_recharge):
            return

        port = self.webhook_port
        app_ref = self

        class WebhookHandler(BaseHTTPRequestHandler):
            def _verify_hmac(self, body, secret, header_name):
                """Verify HMAC-SHA256 signature from webhook header."""
                if not secret:
                    return True  # No secret configured, skip
                import hashlib
                import hmac as hmac_mod
                import base64
                sig = self.headers.get(header_name, "")
                if not sig:
                    return False
                computed = base64.b64encode(
                    hmac_mod.new(
                        secret.encode("utf-8"), body,
                        hashlib.sha256).digest()
                ).decode("utf-8")
                return hmac_mod.compare_digest(sig, computed)

            def do_POST(self):
                content_len = int(
                    self.headers.get("Content-Length", 0))
                body = self.rfile.read(content_len)

                path = self.path.lower()

                # Verify signatures
                if "/shopify" in path:
                    if not self._verify_hmac(
                            body, app_ref.webhook_secret_shopify,
                            "X-Shopify-Hmac-SHA256"):
                        self.send_response(401)
                        self.end_headers()
                        return
                elif "/recharge" in path:
                    if not self._verify_hmac(
                            body, app_ref.webhook_secret_recharge,
                            "X-Recharge-Hmac-SHA256"):
                        self.send_response(401)
                        self.end_headers()
                        return

                try:
                    data = json.loads(body) if body else {}
                except json.JSONDecodeError:
                    self.send_response(400)
                    self.end_headers()
                    return

                if "/shopify" in path:
                    app_ref.root.after(
                        0, lambda d=data: app_ref._handle_shopify_webhook(d))
                elif "/recharge" in path:
                    app_ref.root.after(
                        0, lambda d=data: app_ref._handle_recharge_webhook(d))

                self.send_response(200)
                self.end_headers()
                self.wfile.write(b'{"status":"ok"}')

            def log_message(self, format, *args):
                pass  # Suppress logging

        def _run():
            try:
                server = HTTPServer(("0.0.0.0", port), WebhookHandler)
                app_ref._webhook_server = server
                server.serve_forever()
            except Exception:
                pass

        threading.Thread(target=_run, daemon=True).start()
        self.status_var.set(
            f"Webhook server listening on port {port}")

    def _handle_shopify_webhook(self, data):
        """Process incoming Shopify webhook (order created/fulfilled)."""
        topic = data.get("topic", "")
        # Simple handling: trigger a Shopify data refresh
        self.status_var.set(f"Shopify webhook: {topic}")
        self._send_slack_message(
            f":shopify: Shopify webhook received: {topic}")

        # Schedule a delayed refresh to batch rapid webhooks
        if hasattr(self, '_shopify_webhook_timer'):
            self.root.after_cancel(self._shopify_webhook_timer)
        self._shopify_webhook_timer = self.root.after(
            5000, self._do_auto_refresh)

    def _handle_recharge_webhook(self, data):
        """Process incoming Recharge webhook (charge created/processed)."""
        topic = data.get("topic", "")
        self.status_var.set(f"Recharge webhook: {topic}")
        self._send_slack_message(
            f":arrows_counterclockwise: Recharge webhook received: {topic}")

        # Schedule refresh
        if hasattr(self, '_recharge_webhook_timer'):
            self.root.after_cancel(self._recharge_webhook_timer)
        self._recharge_webhook_timer = self.root.after(
            5000, self._do_auto_refresh)

    # ─────────────────────────────────────────────────────────────────
    #  DASHBOARD AUTO-REFRESH
    # ─────────────────────────────────────────────────────────────────

    def _schedule_dashboard_refresh(self):
        """Refresh dashboard every 5 minutes."""
        self._recalculate()
        self.root.after(300000, self._schedule_dashboard_refresh)

    # ─────────────────────────────────────────────────────────────────
    #  VENDOR CATALOG (Phase 3B)
    # ─────────────────────────────────────────────────────────────────

    def _open_vendor_catalog(self):
        """Edit vendor catalog database."""
        dlg = _VendorCatalogDialog(self.root, self.vendor_catalog,
                                   self.inventory)
        self.root.wait_window(dlg)
        if dlg.result is not None:
            self.vendor_catalog = dlg.result

    # ─────────────────────────────────────────────────────────────────
    #  AUTO-PO DRAFTING (Phase 3C)
    # ─────────────────────────────────────────────────────────────────

    def _generate_auto_po(self):
        """Generate draft POs from reorder alerts grouped by vendor."""
        if not self.last_forecast:
            messagebox.showinfo(
                "Auto-PO",
                "Run a forecast first to identify reorder needs.",
                parent=self.root)
            return

        # Collect deficits from last forecast
        wheel_supply = compute_wheel_supply(
            self.wheel_inventory, self.adjusted_conversion_factors)
        bulk_supply = compute_bulk_supply(
            self._primary_inventory(), self.bulk_conversions)
        globals_ = self._get_global_defaults()
        po_items = defaultdict(list)  # vendor -> [(sku, qty_needed, ...)]

        for month_label, month_data in self.last_forecast.items():
            alerts = compute_reorder_alerts(
                month_data, self.inventory, self.open_pos, wheel_supply,
                bulk_supply=bulk_supply)
            for alert in alerts:
                if alert["deficit"] <= 0:
                    continue
                sku = alert["sku"]
                vendor_info = self.vendor_catalog.get(sku, {})
                vendor = vendor_info.get("vendor", "Unknown")
                case_qty = vendor_info.get("case_qty", 1)
                moq = vendor_info.get("moq", 0)
                wheel_wt = vendor_info.get("wheel_weight_lbs", 0)

                # Calculate order qty
                deficit = alert["deficit"]
                safety = self.sku_settings.get(sku, {}).get(
                    "safety_stock", globals_["safety_stock"])
                qty_needed = deficit + safety

                # Round up to case qty
                if case_qty > 1:
                    qty_needed = (
                        (int(qty_needed) + case_qty - 1) // case_qty
                    ) * case_qty

                # For cheese wheels: convert slices to wheel count
                wheel_count = 0
                if wheel_wt > 0 and sku.startswith("CH-"):
                    slices_per_wheel = wheel_wt * WHEEL_TO_SLICE_FACTOR
                    wheel_count = max(
                        1, int(qty_needed / slices_per_wheel + 0.5))

                unit_cost = vendor_info.get("unit_cost", 0)

                po_items[vendor].append({
                    "sku": sku,
                    "qty": qty_needed,
                    "wheel_count": wheel_count,
                    "unit_cost": unit_cost,
                    "total_cost": round(qty_needed * unit_cost, 2),
                    "month": month_label,
                })

        if not po_items:
            messagebox.showinfo("Auto-PO",
                                "No reorder needs found.",
                                parent=self.root)
            return

        # Show Auto-PO preview dialog
        dlg = _AutoPOPreviewDialog(self.root, po_items, self.vendor_catalog)
        self.root.wait_window(dlg)

    # ─────────────────────────────────────────────────────────────────
    #  PRODUCTION ORDER GENERATOR (Phase 3D)
    # ─────────────────────────────────────────────────────────────────

    def _generate_production_orders(self):
        """Generate Wednesday manufacturing orders for cheese cutting."""
        if not self.last_forecast:
            messagebox.showinfo(
                "Production Orders",
                "Run a forecast first.",
                parent=self.root)
            return

        # Find CH- SKUs that need manufacturing
        today = datetime.date.today()
        # Next fulfillment
        for offset in range(1, 8):
            d = today + datetime.timedelta(days=offset)
            if d.weekday() in (1, 5):
                next_ful = d
                break
        else:
            next_ful = today + datetime.timedelta(days=1)

        # Get demand for next fulfillment period
        combined = self._compute_combined_demand()
        days_until = (next_ful - today).days

        prod_orders = []
        for sku, demand in combined.items():
            if not sku.startswith("CH-"):
                continue
            daily = demand["daily"]
            if daily <= 0:
                continue
            needed = daily * days_until
            on_hand_finished = float(
                self.inventory.get(sku, {}).get("qty", 0))
            deficit = needed - on_hand_finished
            if deficit <= 0:
                continue

            # Check wheel inventory for source
            wheel_info = self.wheel_inventory.get(sku, {})
            wheel_wt = wheel_info.get("weight_lbs", 0)
            wheel_count = wheel_info.get("count", 0)

            prod_orders.append({
                "sku": sku,
                "name": self.inventory.get(sku, {}).get("name", ""),
                "target_qty": round(deficit, 1),
                "wheel_source": f"{wheel_count}x {wheel_wt}lb"
                                if wheel_wt else "N/A",
                "cuts_needed": (
                    max(1, int(deficit / max(
                        wheel_wt * WHEEL_TO_SLICE_FACTOR, 1) + 0.5))
                    if wheel_wt else 0),
            })

        # Also generate AC- processing orders from bulk raw materials
        bulk_supply = compute_bulk_supply(self.inventory, self.bulk_conversions)
        for sku, demand in combined.items():
            if not sku.startswith("AC-"):
                continue
            daily = demand["daily"]
            if daily <= 0:
                continue
            needed = daily * days_until
            on_hand_finished = float(
                self.inventory.get(sku, {}).get("qty", 0))
            deficit = needed - on_hand_finished
            if deficit <= 0:
                continue

            bulk_avail = bulk_supply.get(sku, 0)
            if bulk_avail <= 0:
                continue

            # Find source material name
            source_mat = "Unknown"
            for keyword, conv in self.bulk_conversions.items():
                if conv.get("sku") == sku:
                    source_mat = keyword
                    break

            prod_orders.append({
                "sku": sku,
                "name": self.inventory.get(sku, {}).get("name", ""),
                "target_qty": round(min(deficit, bulk_avail), 1),
                "wheel_source": f"Bulk: {source_mat}",
                "cuts_needed": 0,
                "type": "accompaniment",
                "source_material": source_mat,
            })

        if not prod_orders:
            messagebox.showinfo("Production Orders",
                                "No manufacturing needed.",
                                parent=self.root)
            return

        # Show production order dialog
        dlg = tk.Toplevel(self.root)
        dlg.title("Production Orders — Manufacturing")
        dlg.configure(bg=_BG)
        dlg.transient(self.root)
        dlg.grab_set()
        dlg.geometry("700x400")

        day_name = "TUE" if next_ful.weekday() == 1 else "SAT"
        ttk.Label(dlg,
                  text=f"Manufacturing needed for {day_name} "
                       f"{next_ful} fulfillment",
                  style="Subtitle.TLabel").pack(
                      anchor="w", padx=10, pady=(10, 5))

        tree_frame = ttk.Frame(dlg)
        tree_frame.pack(fill="both", expand=True, padx=10, pady=5)

        cols = ("sku", "name", "target_qty", "wheel_source", "cuts")
        tree = ttk.Treeview(tree_frame, columns=cols, show="headings")
        tree.heading("sku", text="SKU")
        tree.heading("name", text="Product")
        tree.heading("target_qty", text="Target Qty")
        tree.heading("wheel_source", text="Wheel Source")
        tree.heading("cuts", text="Cuts Needed")
        tree.column("sku", width=100)
        tree.column("name", width=200)
        tree.column("target_qty", width=80, anchor="center")
        tree.column("wheel_source", width=120, anchor="center")
        tree.column("cuts", width=80, anchor="center")

        yscroll = ttk.Scrollbar(tree_frame, orient="vertical",
                                command=tree.yview)
        tree.configure(yscrollcommand=yscroll.set)
        tree.pack(side="left", fill="both", expand=True)
        yscroll.pack(side="right", fill="y")

        for po in prod_orders:
            tree.insert("", "end", values=(
                po["sku"], po["name"], po["target_qty"],
                po["wheel_source"], po["cuts_needed"]))

        btn_frame = ttk.Frame(dlg)
        btn_frame.pack(fill="x", padx=10, pady=(5, 10))

        def _export_csv():
            path = filedialog.asksaveasfilename(
                title="Export Production Orders",
                defaultextension=".csv",
                filetypes=[("CSV", "*.csv")])
            if not path:
                return
            with open(path, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(["SKU", "Product", "Target Qty",
                                 "Wheel Source", "Cuts Needed"])
                for po in prod_orders:
                    writer.writerow([po["sku"], po["name"],
                                     po["target_qty"],
                                     po["wheel_source"],
                                     po["cuts_needed"]])
            messagebox.showinfo("Export",
                                f"Exported {len(prod_orders)} items.",
                                parent=dlg)

        def _add_to_queue():
            """Add all production orders to the processing queue."""
            added = 0
            for po in prod_orders:
                # Skip if already queued (same SKU + same target)
                already = any(
                    j["sku"] == po["sku"] and
                    j["status"] in ("scheduled", "in-progress") and
                    abs(j["target_qty"] - po["target_qty"]) < 0.1
                    for j in self.processing_queue)
                if already:
                    continue
                job = {
                    "id": str(uuid.uuid4())[:8],
                    "sku": po["sku"],
                    "source_material": po.get("source_material", ""),
                    "target_qty": po["target_qty"],
                    "status": "scheduled",
                    "warehouse": "Woburn" if po.get("type") == "accompaniment"
                                 else "Primary",
                    "created": datetime.datetime.now().isoformat(
                        timespec="seconds"),
                    "completed": None,
                    "actual_yield": None,
                }
                self.processing_queue.append(job)
                added += 1
            messagebox.showinfo(
                "Processing Queue",
                f"Added {added} job(s) to processing queue."
                + (" (duplicates skipped)" if added < len(prod_orders) else ""),
                parent=dlg)

        ttk.Button(btn_frame, text="Export CSV",
                   command=_export_csv).pack(side="left", padx=(0, 5))
        ttk.Button(btn_frame, text="Add to Queue",
                   command=_add_to_queue).pack(side="left", padx=(0, 5))
        ttk.Button(btn_frame, text="Record Yield...",
                   style="Run.TButton",
                   command=lambda: self._record_production_yield(
                       dlg, prod_orders)).pack(side="left", padx=(5, 5))
        ttk.Button(btn_frame, text="Yield History...",
                   command=self._show_yield_history).pack(
                       side="left", padx=(0, 5))
        ttk.Button(btn_frame, text="Close",
                   command=dlg.destroy).pack(side="right")

    # ─────────────────────────────────────────────────────────────────
    #  PRODUCTION YIELD TRACKING
    # ─────────────────────────────────────────────────────────────────

    def _record_production_yield(self, parent_dlg, prod_orders):
        """Record actual yield from a production run."""
        dlg = tk.Toplevel(self.root)
        dlg.title("Record Production Yield")
        dlg.configure(bg=_BG)
        dlg.transient(parent_dlg)
        dlg.grab_set()
        dlg.geometry("650x450")

        ttk.Label(dlg,
                  text="Enter actual slices produced per SKU",
                  style="Subtitle.TLabel").pack(
                      anchor="w", padx=10, pady=(10, 5))
        ttk.Label(dlg,
                  text="Expected is based on wheel weight x conversion "
                       "factor. Variance updates the factor over time.",
                  style="Dim.TLabel").pack(anchor="w", padx=10, pady=(0, 5))

        tree_frame = ttk.Frame(dlg)
        tree_frame.pack(fill="both", expand=True, padx=10, pady=5)

        cols = ("sku", "wheel_wt", "cuts", "expected", "actual")
        tree = ttk.Treeview(tree_frame, columns=cols, show="headings")
        tree.heading("sku", text="SKU")
        tree.heading("wheel_wt", text="Wheel (lb)")
        tree.heading("cuts", text="Wheels Cut")
        tree.heading("expected", text="Expected")
        tree.heading("actual", text="Actual")
        tree.column("sku", width=100)
        tree.column("wheel_wt", width=90, anchor="center")
        tree.column("cuts", width=90, anchor="center")
        tree.column("expected", width=90, anchor="center")
        tree.column("actual", width=90, anchor="center")

        yscroll = ttk.Scrollbar(tree_frame, orient="vertical",
                                command=tree.yview)
        tree.configure(yscrollcommand=yscroll.set)
        tree.pack(side="left", fill="both", expand=True)
        yscroll.pack(side="right", fill="y")

        # Populate with production order SKUs
        yield_entries = []
        for po in prod_orders:
            sku = po["sku"]
            wheel_info = self.wheel_inventory.get(sku, {})
            wheel_wt = wheel_info.get("weight_lbs", 0)
            cuts = po.get("cuts_needed", 0)
            factor = self.adjusted_conversion_factors.get(
                sku, WHEEL_TO_SLICE_FACTOR)
            expected = round(wheel_wt * cuts * factor, 1) if wheel_wt else 0

            iid = tree.insert("", "end", values=(
                sku, wheel_wt, cuts, expected, ""))
            yield_entries.append({
                "iid": iid, "sku": sku, "wheel_wt": wheel_wt,
                "cuts": cuts, "expected": expected})

        # Make "actual" column editable via double-click
        actual_vars = {}

        def _on_double_click(event):
            item = tree.identify_row(event.y)
            col = tree.identify_column(event.x)
            if not item or col != "#5":  # Only "actual" column
                return
            # Get cell bbox
            bbox = tree.bbox(item, col)
            if not bbox:
                return
            entry = tk.Entry(tree, bg=_BG3, fg=_FG,
                             insertbackground=_FG)
            entry.place(x=bbox[0], y=bbox[1],
                        width=bbox[2], height=bbox[3])
            current = tree.set(item, "actual")
            entry.insert(0, current)
            entry.select_range(0, "end")
            entry.focus_set()

            def _save(e=None):
                val = entry.get().strip()
                tree.set(item, "actual", val)
                actual_vars[item] = val
                entry.destroy()

            entry.bind("<Return>", _save)
            entry.bind("<FocusOut>", _save)

        tree.bind("<Double-1>", _on_double_click)

        btn_frame = ttk.Frame(dlg)
        btn_frame.pack(fill="x", padx=10, pady=(5, 10))

        def _save_yields():
            recorded = 0
            for entry in yield_entries:
                iid = entry["iid"]
                actual_str = actual_vars.get(iid, "").strip()
                if not actual_str:
                    continue
                try:
                    actual = float(actual_str)
                except ValueError:
                    continue

                expected = entry["expected"]
                sku = entry["sku"]
                variance = ((actual - expected) / expected * 100
                            if expected > 0 else 0)

                self.production_yield_history.append({
                    "date": datetime.datetime.now().isoformat(
                        timespec="seconds"),
                    "sku": sku,
                    "wheel_wt": entry["wheel_wt"],
                    "cuts": entry["cuts"],
                    "expected": expected,
                    "actual": actual,
                    "variance_pct": round(variance, 2),
                })
                recorded += 1

                # Update adjusted conversion factor (rolling avg of last 10)
                sku_history = [
                    h for h in self.production_yield_history
                    if h["sku"] == sku and h.get("actual", 0) > 0
                    and h.get("wheel_wt", 0) > 0
                    and h.get("cuts", 0) > 0
                ][-10:]
                if sku_history:
                    # Derive actual factor: actual / (weight * cuts)
                    factors = []
                    for h in sku_history:
                        total_wt = h["wheel_wt"] * h["cuts"]
                        if total_wt > 0:
                            factors.append(h["actual"] / total_wt)
                    if factors:
                        avg_factor = sum(factors) / len(factors)
                        self.adjusted_conversion_factors[sku] = round(
                            avg_factor, 3)

            if recorded:
                # Check for discrepancies against latest snapshot
                warn_msgs = []
                for entry in yield_entries:
                    iid = entry["iid"]
                    actual_str = actual_vars.get(iid, "").strip()
                    if not actual_str:
                        continue
                    sku = entry["sku"]
                    snap_qty = float(
                        self.inventory.get(sku, {}).get("qty", 0))
                    try:
                        actual_val = float(actual_str)
                    except ValueError:
                        continue
                    if snap_qty > 0 and abs(snap_qty - actual_val) > 2:
                        warn_msgs.append(
                            f"{sku}: Snapshot shows {snap_qty}, "
                            f"yield suggests {actual_val}")
                warn_text = ""
                if warn_msgs:
                    warn_text = ("\n\nNote: " +
                                 "; ".join(warn_msgs[:3]))

                messagebox.showinfo(
                    "Yield Recorded",
                    f"Recorded {recorded} yield entries.\n"
                    f"Conversion factors updated.{warn_text}",
                    parent=dlg)
                dlg.destroy()
            else:
                messagebox.showwarning(
                    "No Data",
                    "Double-click the 'Actual' column to enter "
                    "slice counts.",
                    parent=dlg)

        ttk.Button(btn_frame, text="Save Yields",
                   style="Apply.TButton",
                   command=_save_yields).pack(side="right")
        ttk.Button(btn_frame, text="Cancel",
                   command=dlg.destroy).pack(side="right", padx=(0, 5))

    def _show_yield_history(self):
        """Show production yield history with variance analysis."""
        if not self.production_yield_history:
            messagebox.showinfo("Yield History",
                                "No production yields recorded yet.",
                                parent=self.root)
            return

        dlg = tk.Toplevel(self.root)
        dlg.title("Production Yield History")
        dlg.configure(bg=_BG)
        dlg.transient(self.root)
        dlg.grab_set()
        dlg.geometry("850x550")

        # Type filter
        filter_frame = ttk.Frame(dlg)
        filter_frame.pack(fill="x", padx=10, pady=(10, 5))
        ttk.Label(filter_frame, text="Conversion Factor Summary",
                  style="Subtitle.TLabel").pack(side="left")
        yield_type_var = tk.StringVar(value="All")
        for label in ("All", "Cheese", "Accompaniment"):
            ttk.Radiobutton(filter_frame, text=label,
                            variable=yield_type_var, value=label,
                            command=lambda: _refresh_yield()).pack(
                                side="left", padx=(10, 0))

        summary_frame = ttk.Frame(dlg)
        summary_frame.pack(fill="x", padx=10, pady=(0, 5))

        sum_cols = ("sku", "type", "entries", "default_factor",
                    "actual_factor", "avg_variance")
        sum_tree = ttk.Treeview(summary_frame, columns=sum_cols,
                                show="headings", height=6)
        sum_tree.heading("sku", text="SKU")
        sum_tree.heading("type", text="Type")
        sum_tree.heading("entries", text="Records")
        sum_tree.heading("default_factor", text="Default")
        sum_tree.heading("actual_factor", text="Adjusted")
        sum_tree.heading("avg_variance", text="Avg Variance")
        sum_tree.column("sku", width=110)
        sum_tree.column("type", width=95, anchor="center")
        sum_tree.column("entries", width=65, anchor="center")
        sum_tree.column("default_factor", width=75, anchor="center")
        sum_tree.column("actual_factor", width=75, anchor="center")
        sum_tree.column("avg_variance", width=95, anchor="center")
        sum_tree.pack(fill="x")

        # Full history
        ttk.Label(dlg, text="Full History",
                  style="Subtitle.TLabel").pack(
                      anchor="w", padx=10, pady=(10, 5))

        hist_frame = ttk.Frame(dlg)
        hist_frame.pack(fill="both", expand=True, padx=10, pady=(0, 5))

        h_cols = ("date", "sku", "type", "source", "expected",
                  "actual", "variance")
        h_tree = ttk.Treeview(hist_frame, columns=h_cols,
                              show="headings")
        h_tree.heading("date", text="Date")
        h_tree.heading("sku", text="SKU")
        h_tree.heading("type", text="Type")
        h_tree.heading("source", text="Source")
        h_tree.heading("expected", text="Expected")
        h_tree.heading("actual", text="Actual")
        h_tree.heading("variance", text="Variance")
        h_tree.column("date", width=130)
        h_tree.column("sku", width=90)
        h_tree.column("type", width=85, anchor="center")
        h_tree.column("source", width=140)
        h_tree.column("expected", width=75, anchor="center")
        h_tree.column("actual", width=75, anchor="center")
        h_tree.column("variance", width=80, anchor="center")

        h_tree.tag_configure("positive", foreground="#4CAF50")
        h_tree.tag_configure("negative", foreground="#f44336")

        yscroll = ttk.Scrollbar(hist_frame, orient="vertical",
                                command=h_tree.yview)
        h_tree.configure(yscrollcommand=yscroll.set)
        h_tree.pack(side="left", fill="both", expand=True)
        yscroll.pack(side="right", fill="y")

        def _refresh_yield():
            # Clear both trees
            for item in sum_tree.get_children():
                sum_tree.delete(item)
            for item in h_tree.get_children():
                h_tree.delete(item)

            type_filter = yield_type_var.get()

            # Filter history
            filtered = []
            for h in self.production_yield_history:
                h_type = h.get("type", "cheese")
                if type_filter == "Cheese" and h_type != "cheese":
                    continue
                if type_filter == "Accompaniment" and h_type != "accompaniment":
                    continue
                filtered.append(h)

            # Summary by SKU
            by_sku = defaultdict(list)
            for h in filtered:
                by_sku[h["sku"]].append(h)

            for sku, entries in sorted(by_sku.items()):
                adj = self.adjusted_conversion_factors.get(
                    sku, WHEEL_TO_SLICE_FACTOR)
                variances = [e["variance_pct"] for e in entries
                             if "variance_pct" in e]
                avg_var = (sum(variances) / len(variances)
                           if variances else 0)
                entry_type = entries[0].get("type", "cheese")
                default_f = (f"{WHEEL_TO_SLICE_FACTOR:.2f}"
                             if entry_type == "cheese" else "1.00")
                sum_tree.insert("", "end", values=(
                    sku, entry_type.title(), len(entries),
                    default_f, f"{adj:.3f}", f"{avg_var:+.1f}%"))

            # Full history
            for h in reversed(filtered):
                var_pct = h.get("variance_pct", 0)
                tag = "positive" if var_pct >= 0 else "negative"
                h_type = h.get("type", "cheese")
                source = (h.get("source_material", "")
                          if h_type == "accompaniment" else
                          f"{h.get('wheel_wt', 0)}lb x{h.get('cuts', 0)}")
                h_tree.insert("", "end", values=(
                    h["date"][:16], h["sku"], h_type.title(),
                    source, h["expected"], h["actual"],
                    f"{var_pct:+.1f}%"), tags=(tag,))

        _refresh_yield()

        ttk.Button(dlg, text="Close", command=dlg.destroy).pack(
            pady=(5, 10))

    # ─────────────────────────────────────────────────────────────────
    #  NOTIFICATION ALERTS (Phase 4D)
    # ─────────────────────────────────────────────────────────────────

    def _generate_alerts(self):
        """Generate in-app notification alerts."""
        self.alerts = []
        today = datetime.date.today()

        try:
            warn_days = int(self.settings_vars.get(
                "expiration_warning_days", tk.StringVar(value="14")).get())
        except (ValueError, AttributeError):
            warn_days = 14

        # Expiring inventory
        for sku, inv in self.inventory.items():
            exp_dates = inv.get("expiration_dates", [])
            if not exp_dates:
                continue
            try:
                earliest = datetime.date.fromisoformat(exp_dates[0])
                days_until = (earliest - today).days
                if days_until < 0:
                    self.alerts.append({
                        "category": "expiring",
                        "severity": "critical",
                        "message": f"{sku}: EXPIRED ({exp_dates[0]})",
                    })
                elif days_until <= warn_days:
                    self.alerts.append({
                        "category": "expiring",
                        "severity": "warning",
                        "message": (f"{sku}: expires in {days_until}d "
                                    f"({exp_dates[0]})"),
                    })
            except ValueError:
                pass

        # Reorder critical (from dashboard)
        for row in self._dash_rows:
            if row["status"] in ("CRITICAL", "OUT OF STOCK"):
                self.alerts.append({
                    "category": "reorder",
                    "severity": "critical",
                    "message": f"{row['sku']}: {row['status']}",
                })

        # Fulfillment shortfall
        combined = self._compute_combined_demand()
        wheel_supply = compute_wheel_supply(
            self.wheel_inventory, self.adjusted_conversion_factors)
        bulk_supply = compute_bulk_supply(
            self._primary_inventory(), self.bulk_conversions)
        for offset in range(1, 8):
            d = today + datetime.timedelta(days=offset)
            if d.weekday() in (1, 5):
                days_until_ful = offset
                for sku, demand in combined.items():
                    daily = demand["daily"]
                    if daily <= 0:
                        continue
                    on_hand = float(self.inventory.get(
                        sku, {}).get("qty", 0))
                    ws = wheel_supply.get(sku, 0)
                    bs = bulk_supply.get(sku, 0)
                    if daily * days_until_ful > on_hand + ws + bs:
                        self.alerts.append({
                            "category": "fulfillment",
                            "severity": "warning",
                            "message": (f"{sku}: shortfall by "
                                        f"{d.isoformat()}"),
                        })
                break

        # Update bell badge
        if hasattr(self, '_alert_badge_var'):
            unread = len(self.alerts)
            self._alert_badge_var.set(
                f"Alerts ({unread})" if unread else "Alerts")

    def _show_alert_panel(self):
        """Show in-app notification panel."""
        self._generate_alerts()

        if not self.alerts:
            messagebox.showinfo("Alerts", "No alerts.",
                                parent=self.root)
            return

        dlg = tk.Toplevel(self.root)
        dlg.title(f"Alerts ({len(self.alerts)})")
        dlg.configure(bg=_BG)
        dlg.transient(self.root)
        dlg.grab_set()
        dlg.geometry("500x400")

        tree_frame = ttk.Frame(dlg)
        tree_frame.pack(fill="both", expand=True, padx=10, pady=10)

        cols = ("severity", "category", "message")
        tree = ttk.Treeview(tree_frame, columns=cols, show="headings")
        tree.heading("severity", text="Severity")
        tree.heading("category", text="Category")
        tree.heading("message", text="Message")
        tree.column("severity", width=70, anchor="center")
        tree.column("category", width=80, anchor="center")
        tree.column("message", width=300)

        yscroll = ttk.Scrollbar(tree_frame, orient="vertical",
                                command=tree.yview)
        tree.configure(yscrollcommand=yscroll.set)
        tree.pack(side="left", fill="both", expand=True)
        yscroll.pack(side="right", fill="y")

        tree.tag_configure("critical", background="#8b1a1a",
                           foreground="white")
        tree.tag_configure("warning", background="#7a5500",
                           foreground="white")

        for alert in self.alerts:
            tree.insert("", "end", values=(
                alert["severity"].upper(),
                alert["category"],
                alert["message"]),
                tags=(alert["severity"],))

        ttk.Button(dlg, text="Close", command=dlg.destroy).pack(
            pady=(0, 10))

    # ─────────────────────────────────────────────────────────────────
    #  SNAPSHOT COMPARISON (Current vs Potential)
    # ─────────────────────────────────────────────────────────────────

    def _show_snapshot_comparison(self):
        """Show Current vs Potential inventory comparison dialog."""
        dlg = tk.Toplevel(self.root)
        dlg.title("Inventory Snapshot — Current vs Potential")
        dlg.configure(bg=_BG)
        dlg.transient(self.root)
        dlg.grab_set()
        dlg.geometry("1100x600")

        # Toolbar
        top = ttk.Frame(dlg)
        top.pack(fill="x", padx=10, pady=(10, 5))

        ttk.Label(top, text="Current vs Potential Inventory",
                  style="Subtitle.TLabel").pack(side="left")

        include_mfg_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(top, text="Include planned MFG",
                        variable=include_mfg_var,
                        command=lambda: _refresh()).pack(
                            side="left", padx=(20, 0))

        ttk.Button(top, text="Export CSV",
                   command=lambda: self._export_snapshot_csv(tree)).pack(
                       side="right")

        # Treeview
        tree_frame = ttk.Frame(dlg)
        tree_frame.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        cols = ("sku", "name", "on_hand", "wip", "open_po",
                "bulk_potential", "wheel_potential", "planned_mfg",
                "total_potential", "demand_30d", "net")
        tree = ttk.Treeview(tree_frame, columns=cols, show="headings")
        tree.heading("sku", text="SKU")
        tree.heading("name", text="Name")
        tree.heading("on_hand", text="On Hand")
        tree.heading("wip", text="+ WIP")
        tree.heading("open_po", text="+ Open POs")
        tree.heading("bulk_potential", text="+ Bulk Pot.")
        tree.heading("wheel_potential", text="+ Wheel Pot.")
        tree.heading("planned_mfg", text="+ Planned MFG")
        tree.heading("total_potential", text="= Total Pot.")
        tree.heading("demand_30d", text="Demand (30d)")
        tree.heading("net", text="Net")

        tree.column("sku", width=110)
        tree.column("name", width=170)
        tree.column("on_hand", width=70, anchor="center")
        tree.column("wip", width=60, anchor="center")
        tree.column("open_po", width=75, anchor="center")
        tree.column("bulk_potential", width=80, anchor="center")
        tree.column("wheel_potential", width=80, anchor="center")
        tree.column("planned_mfg", width=85, anchor="center")
        tree.column("total_potential", width=85, anchor="center")
        tree.column("demand_30d", width=85, anchor="center")
        tree.column("net", width=70, anchor="center")

        yscroll = ttk.Scrollbar(tree_frame, orient="vertical",
                                command=tree.yview)
        tree.configure(yscrollcommand=yscroll.set)
        tree.pack(side="left", fill="both", expand=True)
        yscroll.pack(side="right", fill="y")

        # Row coloring
        tree.tag_configure("green", background="#1a4a2e", foreground="white")
        tree.tag_configure("amber", background="#7a5500", foreground="white")
        tree.tag_configure("red", background="#8b1a1a", foreground="white")

        # Legend
        legend = ttk.Frame(dlg)
        legend.pack(fill="x", padx=10, pady=(0, 10))
        for color, label in [("#1a4a2e", "Current covers demand"),
                             ("#7a5500", "Potential covers (processing needed)"),
                             ("#8b1a1a", "Neither covers demand")]:
            lbl = tk.Label(legend, text=f"  {label}  ", bg=color,
                           fg="white", font=("Segoe UI", 8))
            lbl.pack(side="left", padx=(0, 8))

        def _refresh():
            for item in tree.get_children():
                tree.delete(item)

            wheel_supply = compute_wheel_supply(
                self.wheel_inventory, self.adjusted_conversion_factors)
            bulk_supply = compute_bulk_supply(
                self._primary_inventory(), self.bulk_conversions)

            po_by_sku = self._po_qty_by_sku()

            # Planned MFG from processing queue
            mfg_by_sku = defaultdict(float)
            if include_mfg_var.get():
                for job in self.processing_queue:
                    if job.get("status") in ("scheduled", "in-progress"):
                        mfg_by_sku[job["sku"]] += job.get("target_qty", 0)

            # Get 30-day demand
            combined = self._compute_combined_demand()

            # Collect all relevant SKUs
            _PRODUCT_PREFIXES = ("CH-", "PK-", "MT-", "AC-")
            all_skus = set()
            for sku in self.inventory:
                if sku.startswith(_PRODUCT_PREFIXES):
                    all_skus.add(sku)
            for sku in combined:
                if sku.startswith(_PRODUCT_PREFIXES):
                    all_skus.add(sku)

            for sku in sorted(all_skus):
                inv = self.inventory.get(sku, {})
                on_hand = float(inv.get("qty", 0))
                wip = float(inv.get("wip_qty", 0))
                name = inv.get("name", "")
                po_qty = po_by_sku.get(sku, 0)
                bulk_pot = bulk_supply.get(sku, 0)
                wheel_pot = wheel_supply.get(sku, 0)
                planned = mfg_by_sku.get(sku, 0)
                total_pot = (on_hand + wip + po_qty + bulk_pot
                             + wheel_pot + planned)
                demand_30d = combined.get(sku, {}).get("daily", 0) * 30

                net = total_pot - demand_30d

                # Determine row color
                if on_hand >= demand_30d and demand_30d > 0:
                    tag = "green"
                elif total_pot >= demand_30d and demand_30d > 0:
                    tag = "amber"
                elif demand_30d > 0:
                    tag = "red"
                else:
                    tag = "green"

                tree.insert("", "end", values=(
                    sku, name, round(on_hand, 1),
                    round(wip, 1) if wip else "",
                    round(po_qty, 1), round(bulk_pot, 1),
                    round(wheel_pot, 1), round(planned, 1),
                    round(total_pot, 1), round(demand_30d, 1),
                    round(net, 1),
                ), tags=(tag,))

        _refresh()

    def _show_transfer_dialog(self):
        """Record inventory transfers between warehouses."""
        dlg = tk.Toplevel(self.root)
        dlg.title("Record Transfer — Woburn to Primary")
        dlg.configure(bg=_BG)
        dlg.transient(self.root)
        dlg.grab_set()
        dlg.geometry("600x450")

        ttk.Label(dlg, text="Transfer Processed Goods to Primary",
                  style="Subtitle.TLabel").pack(
                      anchor="w", padx=10, pady=(10, 5))

        # List Woburn inventory items
        tree_frame = ttk.Frame(dlg)
        tree_frame.pack(fill="both", expand=True, padx=10, pady=5)

        cols = ("sku", "name", "qty_available", "transfer_qty")
        tree = ttk.Treeview(tree_frame, columns=cols, show="headings",
                            height=10)
        tree.heading("sku", text="SKU")
        tree.heading("name", text="Name")
        tree.heading("qty_available", text="At Woburn")
        tree.heading("transfer_qty", text="Transfer Qty")
        tree.column("sku", width=110)
        tree.column("name", width=180)
        tree.column("qty_available", width=90, anchor="center")
        tree.column("transfer_qty", width=100, anchor="center")

        yscroll = ttk.Scrollbar(tree_frame, orient="vertical",
                                command=tree.yview)
        tree.configure(yscrollcommand=yscroll.set)
        tree.pack(side="left", fill="both", expand=True)
        yscroll.pack(side="right", fill="y")

        # Populate with Woburn items (check warehouse_qty or warehouse field)
        woburn_items = []
        for sku, data in self.inventory.items():
            woburn_qty = self._qty_at(sku, "Woburn")
            if woburn_qty > 0:
                tree.insert("", "end", values=(
                    sku, data.get("name", ""),
                    round(woburn_qty, 1), ""))
                woburn_items.append(sku)

        if not woburn_items:
            ttk.Label(dlg, text="No items at Woburn warehouse.",
                      foreground=_FG2).pack(pady=10)

        # Editable transfer_qty column
        transfer_vals = {}

        def _on_double_click(event):
            item = tree.identify_row(event.y)
            col = tree.identify_column(event.x)
            if not item or col != "#4":
                return
            bbox = tree.bbox(item, col)
            if not bbox:
                return
            entry = tk.Entry(tree, bg=_BG3, fg=_FG,
                             insertbackground=_FG)
            entry.place(x=bbox[0], y=bbox[1],
                        width=bbox[2], height=bbox[3])
            current = tree.set(item, "transfer_qty")
            entry.insert(0, current)
            entry.select_range(0, "end")
            entry.focus_set()

            def _save(e=None):
                val = entry.get().strip()
                tree.set(item, "transfer_qty", val)
                transfer_vals[item] = val
                entry.destroy()

            entry.bind("<Return>", _save)
            entry.bind("<FocusOut>", _save)

        tree.bind("<Double-1>", _on_double_click)

        btn_frame = ttk.Frame(dlg)
        btn_frame.pack(fill="x", padx=10, pady=(5, 10))

        def _execute_transfer():
            transferred = 0
            for item_id in tree.get_children():
                qty_str = transfer_vals.get(item_id, "").strip()
                if not qty_str:
                    continue
                try:
                    qty = float(qty_str)
                except ValueError:
                    continue
                if qty <= 0:
                    continue
                sku = tree.set(item_id, "sku")
                available = self._qty_at(sku, "Woburn")
                qty = min(qty, available)

                # Move qty from Woburn to Primary using warehouse_qty
                new_woburn = available - qty
                self._set_qty_at(sku, "Woburn", new_woburn)
                primary_now = self._qty_at(sku, "Primary")
                self._set_qty_at(sku, "Primary", primary_now + qty)

                # Log transfer
                self.transfer_history.append({
                    "date": datetime.datetime.now().isoformat(
                        timespec="seconds"),
                    "sku": sku,
                    "qty": qty,
                    "from_warehouse": "Woburn",
                    "to_warehouse": "Primary",
                })
                transferred += 1

            if transferred:
                self._refresh_inventory_tree()
                self._recalculate()
                messagebox.showinfo(
                    "Transfer Complete",
                    f"Transferred {transferred} item(s) to Primary.",
                    parent=dlg)
                dlg.destroy()
            else:
                messagebox.showwarning(
                    "No Transfer",
                    "Double-click 'Transfer Qty' to enter quantities.",
                    parent=dlg)

        ttk.Button(btn_frame, text="Execute Transfer",
                   style="Apply.TButton",
                   command=_execute_transfer).pack(side="right")
        ttk.Button(btn_frame, text="Cancel",
                   command=dlg.destroy).pack(side="right", padx=(0, 5))

    def _show_transfer_history(self):
        """Show a read-only log of past warehouse transfers."""
        dlg = tk.Toplevel(self.root)
        dlg.title("Transfer History")
        dlg.configure(bg=_BG)
        dlg.transient(self.root)
        dlg.grab_set()
        dlg.geometry("700x400")

        ttk.Label(dlg, text="Warehouse Transfer Log",
                  style="Subtitle.TLabel").pack(
                      anchor="w", padx=10, pady=(10, 5))

        tree_frame = ttk.Frame(dlg)
        tree_frame.pack(fill="both", expand=True, padx=10, pady=5)

        cols = ("date", "sku", "qty", "from_wh", "to_wh")
        tree = ttk.Treeview(tree_frame, columns=cols, show="headings")
        tree.heading("date", text="Date")
        tree.heading("sku", text="SKU")
        tree.heading("qty", text="Qty")
        tree.heading("from_wh", text="From")
        tree.heading("to_wh", text="To")
        tree.column("date", width=150)
        tree.column("sku", width=120)
        tree.column("qty", width=80, anchor="center")
        tree.column("from_wh", width=120)
        tree.column("to_wh", width=120)

        yscroll = ttk.Scrollbar(tree_frame, orient="vertical",
                                command=tree.yview)
        tree.configure(yscrollcommand=yscroll.set)
        tree.pack(side="left", fill="both", expand=True)
        yscroll.pack(side="right", fill="y")

        # Show newest first
        for t in reversed(self.transfer_history):
            tree.insert("", "end", values=(
                t.get("date", "")[:16],
                t.get("sku", ""),
                t.get("qty", 0),
                t.get("from_warehouse", ""),
                t.get("to_warehouse", ""),
            ))

        btn_frame = ttk.Frame(dlg)
        btn_frame.pack(fill="x", padx=10, pady=(5, 10))

        def _export_csv():
            path = filedialog.asksaveasfilename(
                defaultextension=".csv",
                filetypes=[("CSV", "*.csv")],
                title="Export Transfer History",
                parent=dlg)
            if not path:
                return
            with open(path, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(["Date", "SKU", "Qty", "From", "To"])
                for t in self.transfer_history:
                    writer.writerow([
                        t.get("date", ""), t.get("sku", ""),
                        t.get("qty", 0), t.get("from_warehouse", ""),
                        t.get("to_warehouse", ""),
                    ])
            messagebox.showinfo("Export",
                                f"Exported {len(self.transfer_history)} entries.",
                                parent=dlg)

        ttk.Button(btn_frame, text="Export CSV",
                   command=_export_csv).pack(side="left")
        ttk.Button(btn_frame, text="Close",
                   command=dlg.destroy).pack(side="right")

    def _export_snapshot_csv(self, tree):
        """Export snapshot treeview to CSV."""
        path = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv")],
            title="Export Snapshot CSV",
            parent=self.root)
        if not path:
            return
        with open(path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            headers = [tree.heading(c)["text"]
                       for c in tree["columns"]]
            writer.writerow(headers)
            for item in tree.get_children():
                writer.writerow(tree.item(item)["values"])
        self.status_var.set(f"Snapshot exported to {os.path.basename(path)}")

    # ─────────────────────────────────────────────────────────────────
    #  PROCESSING QUEUE
    # ─────────────────────────────────────────────────────────────────

    def _show_processing_queue(self):
        """Show processing queue dialog for managing AC-/CH- jobs."""
        dlg = tk.Toplevel(self.root)
        dlg.title("Processing Queue")
        dlg.configure(bg=_BG)
        dlg.transient(self.root)
        dlg.grab_set()
        dlg.geometry("900x550")

        ttk.Label(dlg, text="Processing Queue",
                  style="Subtitle.TLabel").pack(
                      anchor="w", padx=10, pady=(10, 5))

        # Toolbar
        q_toolbar = ttk.Frame(dlg)
        q_toolbar.pack(fill="x", padx=10, pady=(0, 5))

        def _add_job():
            """Add a new processing job from bulk materials."""
            add_dlg = tk.Toplevel(dlg)
            add_dlg.title("Add Processing Job")
            add_dlg.configure(bg=_BG)
            add_dlg.transient(dlg)
            add_dlg.grab_set()
            add_dlg.geometry("400x300")

            ttk.Label(add_dlg, text="Source Material:").pack(
                anchor="w", padx=10, pady=(10, 3))
            mat_var = tk.StringVar()
            mat_names = list(self.bulk_conversions.keys())
            ttk.Combobox(add_dlg, textvariable=mat_var,
                         values=mat_names, state="readonly",
                         width=35).pack(padx=10)

            ttk.Label(add_dlg, text="Target Qty (packets):").pack(
                anchor="w", padx=10, pady=(10, 3))
            qty_var = tk.StringVar()
            ttk.Entry(add_dlg, textvariable=qty_var, width=15).pack(
                anchor="w", padx=10)

            ttk.Label(add_dlg, text="Warehouse:").pack(
                anchor="w", padx=10, pady=(10, 3))
            wh_var = tk.StringVar(value="Woburn")
            ttk.Combobox(add_dlg, textvariable=wh_var,
                         values=list(self.warehouses.keys()),
                         state="readonly", width=15).pack(anchor="w", padx=10)

            def _save_job():
                mat = mat_var.get()
                if not mat:
                    return
                conv = self.bulk_conversions.get(mat, {})
                try:
                    target = float(qty_var.get())
                except ValueError:
                    return
                job = {
                    "id": str(uuid.uuid4())[:8],
                    "sku": conv.get("sku", ""),
                    "source_material": mat,
                    "target_qty": target,
                    "status": "scheduled",
                    "warehouse": wh_var.get(),
                    "created": datetime.datetime.now().isoformat(
                        timespec="seconds"),
                    "completed": None,
                    "actual_yield": None,
                }
                self.processing_queue.append(job)
                add_dlg.destroy()
                _refresh_queue()

            ttk.Button(add_dlg, text="Add Job", style="Apply.TButton",
                       command=_save_job).pack(pady=15)

        ttk.Button(q_toolbar, text="Add Job",
                   command=_add_job).pack(side="left", padx=(0, 5))

        def _start_selected():
            sel = tree.selection()
            if not sel:
                return
            item = sel[0]
            job_id = tree.set(item, "id")
            for job in self.processing_queue:
                if job["id"] == job_id and job["status"] == "scheduled":
                    job["status"] = "in-progress"
                    break
            _refresh_queue()

        def _complete_selected():
            sel = tree.selection()
            if not sel:
                return
            item = sel[0]
            job_id = tree.set(item, "id")
            job = None
            for j in self.processing_queue:
                if j["id"] == job_id and j["status"] == "in-progress":
                    job = j
                    break
            if not job:
                return

            # Prompt for actual yield
            actual_str = simpledialog.askstring(
                "Actual Yield",
                f"Enter actual packets produced for {job['sku']}:",
                parent=dlg)
            if not actual_str:
                return
            try:
                actual = float(actual_str)
            except ValueError:
                return

            job["status"] = "complete"
            job["completed"] = datetime.datetime.now().isoformat(
                timespec="seconds")
            job["actual_yield"] = actual

            # Update inventory: decrement raw material, increment finished
            sku = job["sku"]
            source = job["source_material"]

            # Decrement raw material (exact match via helper)
            _, raw_inv = self._bulk_source_sku_for(source)
            if raw_inv:
                conv = self.bulk_conversions.get(source, {})
                packet_oz = conv.get("packet_oz", 3.9)
                oz_used = actual * packet_oz
                unit = raw_inv.get("unit", "").lower()
                unit_size = raw_inv.get("unit_size", 1)
                if "lb" in unit:
                    lbs_used = oz_used / 16
                    raw_inv["qty"] = max(
                        0, float(raw_inv.get("qty", 0)) - lbs_used / max(unit_size, 1))
                else:
                    raw_inv["qty"] = max(
                        0, float(raw_inv.get("qty", 0)) - oz_used / max(unit_size, 1))

            # Increment finished good
            if sku in self.inventory:
                self.inventory[sku]["qty"] = (
                    float(self.inventory[sku].get("qty", 0)) + actual)
            else:
                self.inventory[sku] = {"qty": actual, "name": sku,
                                       "warehouse": "Primary"}

            # Record yield history (same as cheese tracking)
            expected = job["target_qty"]
            variance = ((actual - expected) / expected * 100
                        if expected > 0 else 0)
            self.production_yield_history.append({
                "date": datetime.datetime.now().isoformat(
                    timespec="seconds"),
                "sku": sku,
                "type": "accompaniment",
                "source_material": source,
                "expected": expected,
                "actual": actual,
                "variance_pct": round(variance, 2),
                "wheel_wt": 0,
                "cuts": 0,
            })

            # Update adjusted conversion factor for AC- items
            sku_history = [
                h for h in self.production_yield_history
                if h["sku"] == sku and h.get("actual", 0) > 0
                and h.get("source_material")
            ][-10:]
            if sku_history and len(sku_history) >= 2:
                # For AC-: track packets per lb of raw material
                conv = self.bulk_conversions.get(source, {})
                packet_oz = conv.get("packet_oz", 3.9)
                if packet_oz > 0:
                    # Factor = actual_packets / expected_packets
                    ratios = [h["actual"] / h["expected"]
                              for h in sku_history
                              if h.get("expected", 0) > 0]
                    if ratios:
                        avg_ratio = sum(ratios) / len(ratios)
                        self.adjusted_conversion_factors[sku] = round(
                            avg_ratio, 3)

            _refresh_queue()
            self._refresh_inventory_tree()
            messagebox.showinfo(
                "Job Complete",
                f"Recorded {actual} packets for {sku}.\n"
                f"Expected: {expected}, Variance: {variance:+.1f}%",
                parent=dlg)

        ttk.Button(q_toolbar, text="Start Selected",
                   command=_start_selected).pack(side="left", padx=(0, 5))
        ttk.Button(q_toolbar, text="Complete Selected",
                   style="Apply.TButton",
                   command=_complete_selected).pack(side="left", padx=(0, 5))

        # Treeview
        tree_frame = ttk.Frame(dlg)
        tree_frame.pack(fill="both", expand=True, padx=10, pady=5)

        cols = ("id", "sku", "source", "target", "status",
                "warehouse", "created", "actual")
        tree = ttk.Treeview(tree_frame, columns=cols, show="headings")
        tree.heading("id", text="ID")
        tree.heading("sku", text="SKU")
        tree.heading("source", text="Source Material")
        tree.heading("target", text="Target Qty")
        tree.heading("status", text="Status")
        tree.heading("warehouse", text="Warehouse")
        tree.heading("created", text="Created")
        tree.heading("actual", text="Actual Yield")
        tree.column("id", width=65)
        tree.column("sku", width=90)
        tree.column("source", width=160)
        tree.column("target", width=75, anchor="center")
        tree.column("status", width=85, anchor="center")
        tree.column("warehouse", width=80, anchor="center")
        tree.column("created", width=130)
        tree.column("actual", width=80, anchor="center")

        tree.tag_configure("scheduled", background=_BG2)
        tree.tag_configure("in-progress", background="#2a4a6b",
                           foreground="white")
        tree.tag_configure("complete", background="#1a4a2e",
                           foreground="white")

        yscroll = ttk.Scrollbar(tree_frame, orient="vertical",
                                command=tree.yview)
        tree.configure(yscrollcommand=yscroll.set)
        tree.pack(side="left", fill="both", expand=True)
        yscroll.pack(side="right", fill="y")

        def _refresh_queue():
            for item in tree.get_children():
                tree.delete(item)
            for job in self.processing_queue:
                tree.insert("", "end", values=(
                    job.get("id", ""),
                    job.get("sku", ""),
                    job.get("source_material", ""),
                    job.get("target_qty", 0),
                    job.get("status", ""),
                    job.get("warehouse", ""),
                    job.get("created", "")[:16] if job.get("created") else "",
                    job.get("actual_yield", "") or "",
                ), tags=(job.get("status", ""),))

        _refresh_queue()

        bottom_frame = ttk.Frame(dlg)
        bottom_frame.pack(fill="x", padx=10, pady=(0, 10))

        def _export_queue_csv():
            path = filedialog.asksaveasfilename(
                defaultextension=".csv",
                filetypes=[("CSV", "*.csv")],
                title="Export Processing Queue",
                parent=dlg)
            if not path:
                return
            with open(path, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(["ID", "SKU", "Source Material",
                                 "Target Qty", "Status", "Warehouse",
                                 "Created", "Actual Yield"])
                for job in self.processing_queue:
                    writer.writerow([
                        job.get("id", ""), job.get("sku", ""),
                        job.get("source_material", ""),
                        job.get("target_qty", 0), job.get("status", ""),
                        job.get("warehouse", ""), job.get("created", ""),
                        job.get("actual_yield", ""),
                    ])
            messagebox.showinfo(
                "Export",
                f"Exported {len(self.processing_queue)} jobs.",
                parent=dlg)

        ttk.Button(bottom_frame, text="Export CSV",
                   command=_export_queue_csv).pack(side="left")
        ttk.Button(bottom_frame, text="Close",
                   command=dlg.destroy).pack(side="right")

    # ─────────────────────────────────────────────────────────────────
    #  YIELD DISCREPANCIES & RECONCILIATION
    # ─────────────────────────────────────────────────────────────────

    def _show_yield_discrepancies(self):
        """Show yield discrepancy flags dialog."""
        dlg = tk.Toplevel(self.root)
        dlg.title("Yield Discrepancies")
        dlg.configure(bg=_BG)
        dlg.transient(self.root)
        dlg.grab_set()
        dlg.geometry("900x500")

        ttk.Label(dlg, text="Yield Discrepancy Flags",
                  style="Subtitle.TLabel").pack(
                      anchor="w", padx=10, pady=(10, 5))

        tree_frame = ttk.Frame(dlg)
        tree_frame.pack(fill="both", expand=True, padx=10, pady=5)

        cols = ("sku", "expected", "actual", "variance",
                "yield_date", "snapshot_date", "status")
        tree = ttk.Treeview(tree_frame, columns=cols, show="headings")
        tree.heading("sku", text="SKU")
        tree.heading("expected", text="Expected")
        tree.heading("actual", text="Actual")
        tree.heading("variance", text="Variance")
        tree.heading("yield_date", text="Yield Date")
        tree.heading("snapshot_date", text="Snapshot Date")
        tree.heading("status", text="Status")
        tree.column("sku", width=110)
        tree.column("expected", width=80, anchor="center")
        tree.column("actual", width=80, anchor="center")
        tree.column("variance", width=80, anchor="center")
        tree.column("yield_date", width=110, anchor="center")
        tree.column("snapshot_date", width=110, anchor="center")
        tree.column("status", width=100, anchor="center")

        tree.tag_configure("open", background="#8b1a1a", foreground="white")
        tree.tag_configure("acknowledged", background="#7a5500",
                           foreground="white")
        tree.tag_configure("resolved", background="#1a4a2e",
                           foreground="white")

        yscroll = ttk.Scrollbar(tree_frame, orient="vertical",
                                command=tree.yview)
        tree.configure(yscrollcommand=yscroll.set)
        tree.pack(side="left", fill="both", expand=True)
        yscroll.pack(side="right", fill="y")

        # Map treeview iids to list indices
        _iid_to_idx = {}

        def _refresh():
            _iid_to_idx.clear()
            for item in tree.get_children():
                tree.delete(item)
            for i, d in enumerate(self.yield_discrepancies):
                iid = tree.insert("", "end", values=(
                    d.get("sku", ""),
                    d.get("expected_qty", ""),
                    d.get("actual_qty", ""),
                    d.get("variance", ""),
                    d.get("yield_date", "")[:10] if d.get("yield_date") else "",
                    d.get("snapshot_date", "")[:10] if d.get("snapshot_date") else "",
                    d.get("status", "open"),
                ), tags=(d.get("status", "open"),))
                _iid_to_idx[iid] = i

        _refresh()

        btn_frame = ttk.Frame(dlg)
        btn_frame.pack(fill="x", padx=10, pady=(5, 10))

        def _acknowledge():
            sel = tree.selection()
            if not sel:
                return
            idx = _iid_to_idx.get(sel[0])
            if idx is not None and idx < len(self.yield_discrepancies):
                self.yield_discrepancies[idx]["status"] = "acknowledged"
            _refresh()
            self._update_yield_flag_badge()

        def _resolve():
            sel = tree.selection()
            if not sel:
                return
            idx = _iid_to_idx.get(sel[0])
            if idx is None or idx >= len(self.yield_discrepancies):
                return
            disc = self.yield_discrepancies[idx]

            # Offer to adjust inventory
            if messagebox.askyesno(
                    "Resolve Discrepancy",
                    f"Adjust {disc['sku']} inventory to match snapshot "
                    f"({disc.get('actual_qty', '?')} units)?",
                    parent=dlg):
                sku = disc["sku"]
                if sku in self.inventory:
                    self.inventory[sku]["qty"] = disc.get("actual_qty", 0)

            disc["status"] = "resolved"
            _refresh()
            self._update_yield_flag_badge()
            self._refresh_inventory_tree()
            self._recalculate()

        ttk.Button(btn_frame, text="Acknowledge",
                   command=_acknowledge).pack(side="left", padx=(0, 5))
        ttk.Button(btn_frame, text="Resolve",
                   style="Apply.TButton",
                   command=_resolve).pack(side="left", padx=(0, 5))
        ttk.Button(btn_frame, text="Close",
                   command=dlg.destroy).pack(side="right")

    def _update_yield_flag_badge(self):
        """Update the yield flag badge in the top bar."""
        open_flags = sum(1 for d in self.yield_discrepancies
                         if d.get("status") == "open")
        if open_flags > 0:
            self._yield_flag_var.set(f"{open_flags} yield flags")
            self._yield_flag_btn.pack(side="right", padx=(0, 5))
        else:
            self._yield_flag_var.set("")
            self._yield_flag_btn.pack_forget()

    def _reconcile_yield_vs_snapshot(self, snapshot_inventory,
                                     pre_import_qty=None):
        """Compare CH- SKU quantities from a Dropbox snapshot against
        recent production yield history. Creates discrepancy flags.

        Args:
            snapshot_inventory: {sku: qty} from the imported snapshot
            pre_import_qty: {sku: qty} inventory state BEFORE import
        """
        today = datetime.date.today()
        window = self.yield_recon_window_days
        threshold_pct = self.yield_recon_threshold_pct
        threshold_min = self.yield_recon_threshold_min
        new_flags = 0

        # Pre-index yield history by SKU for O(n+m) instead of O(n*m)
        yield_by_sku = defaultdict(list)
        for yh in self.production_yield_history:
            yield_by_sku[yh.get("sku", "")].append(yh)

        for sku, snap_qty in snapshot_inventory.items():
            # Find recent yield entries within the date window
            recent_yields = []
            for yh in yield_by_sku.get(sku, []):
                try:
                    yield_date = datetime.date.fromisoformat(
                        yh["date"][:10])
                    days_diff = abs((today - yield_date).days)
                    if days_diff <= window:
                        recent_yields.append(yh)
                except (ValueError, KeyError):
                    continue

            if not recent_yields:
                continue

            # Use pre-import qty (what we had before snapshot overwrote it)
            if pre_import_qty is not None:
                expected = pre_import_qty.get(sku, 0)
            else:
                expected = float(
                    self.inventory.get(sku, {}).get("qty", 0))

            actual = float(snap_qty)
            variance = actual - expected

            # Check threshold
            if abs(variance) < threshold_min:
                continue
            if expected > 0 and abs(variance / expected * 100) < threshold_pct:
                continue

            # Check if we already have this flag
            already_flagged = False
            for existing in self.yield_discrepancies:
                if (existing["sku"] == sku and
                        existing.get("snapshot_date", "")[:10] == today.isoformat()):
                    already_flagged = True
                    break
            if already_flagged:
                continue

            self.yield_discrepancies.append({
                "date": datetime.datetime.now().isoformat(
                    timespec="seconds"),
                "sku": sku,
                "type": "yield_vs_snapshot",
                "expected_qty": round(expected, 1),
                "actual_qty": round(actual, 1),
                "variance": round(variance, 1),
                "yield_date": (recent_yields[-1]["date"][:10]
                               if recent_yields else ""),
                "snapshot_date": today.isoformat(),
                "status": "open",
            })
            new_flags += 1

        if new_flags > 0:
            self._update_yield_flag_badge()
            # Optional Slack notification
            if self.slack_webhook_url and self.slack_notify_shortfall:
                self._send_slack_message(
                    f":warning: {new_flags} new yield discrepancy "
                    f"flag(s) detected after Dropbox import.")

        return new_flags

    # ─────────────────────────────────────────────────────────────────
    #  WORKFLOW GUIDE
    # ─────────────────────────────────────────────────────────────────

    def _show_workflow_guide(self):
        """Show daily/weekly workflow guide dialog."""
        dlg = tk.Toplevel(self.root)
        dlg.title("Workflow Guide")
        dlg.configure(bg=_BG)
        dlg.transient(self.root)
        dlg.grab_set()
        dlg.geometry("650x700")

        canvas = tk.Canvas(dlg, bg=_BG, highlightthickness=0)
        scrollbar = ttk.Scrollbar(dlg, orient="vertical",
                                  command=canvas.yview)
        scroll_frame = ttk.Frame(canvas)
        scroll_frame.bind("<Configure>",
                          lambda e: canvas.configure(
                              scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=scroll_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        canvas.bind_all("<MouseWheel>", _on_mousewheel)

        # Daily Workflow
        daily = ttk.LabelFrame(scroll_frame, text="Daily Workflow",
                                padding=10)
        daily.pack(fill="x", padx=15, pady=(10, 5))

        daily_steps = [
            ("1. Check Dashboard",
             "Review CRITICAL/REORDER alerts + yield discrepancy "
             "flags after auto-refresh"),
            ("2. Review yield flags",
             "If any, open Yield Discrepancies dialog; "
             "acknowledge or resolve"),
            ("3. Click 'Snapshot'",
             "See Current vs Potential; identify what processing "
             "can resolve (amber rows)"),
            ("4. Import arrivals",
             "Inventory tab: import CSV or manually update received "
             "items, set warehouse"),
            ("5. Check Calendar",
             "Today's actions: Crossdock, Process, MFG, "
             "Transfer deadlines"),
            ("6. Execute processing",
             "Processing Queue: start/complete jobs at Woburn, "
             "record actual yields"),
            ("7. Record transfers",
             "Move processed goods Woburn -> Primary"),
            ("8. Deplete & Email",
             "After fulfillment, import depletion and send report"),
        ]

        for i, (title, desc) in enumerate(daily_steps):
            ttk.Label(daily, text=title,
                      style="Bold.TLabel").grid(
                          row=i * 2, column=0, sticky="w", pady=(5, 0))
            ttk.Label(daily, text=desc, wraplength=550,
                      style="Dim.TLabel").grid(
                          row=i * 2 + 1, column=0, sticky="w",
                          padx=(15, 0), pady=(0, 3))

        # Weekly Workflow
        weekly = ttk.LabelFrame(scroll_frame,
                                 text="Weekly Workflow (Monday)",
                                 padding=10)
        weekly.pack(fill="x", padx=15, pady=(10, 15))

        weekly_steps = [
            ("1. Run Forecast",
             "Forecasting tab: generate 3-month projection"),
            ("2. Generate Production Orders",
             "For both CH- and AC- SKUs; add to processing queue"),
            ("3. Generate Schedule",
             "Populate calendar; sync to ClickUp/Google Calendar"),
            ("4. Auto-PO Review",
             "Review suggested POs, submit to vendors"),
            ("5. Snapshot Review",
             "Validate plan covers demand through the week"),
            ("6. Yield Review",
             "Check conversion factor trends, review/clear old "
             "discrepancy flags"),
        ]

        for i, (title, desc) in enumerate(weekly_steps):
            ttk.Label(weekly, text=title,
                      style="Bold.TLabel").grid(
                          row=i * 2, column=0, sticky="w", pady=(5, 0))
            ttk.Label(weekly, text=desc, wraplength=550,
                      style="Dim.TLabel").grid(
                          row=i * 2 + 1, column=0, sticky="w",
                          padx=(15, 0), pady=(0, 3))

        ttk.Button(scroll_frame, text="Close",
                   command=dlg.destroy).pack(pady=(5, 15))

    # ─────────────────────────────────────────────────────────────────
    #  TAB 6 — FULFILLMENT PLANNER
    # ─────────────────────────────────────────────────────────────────

    # Mascot ASCII art states (displayed on Canvas)
    _MASCOT = {
        "idle":     "  ___\n /o o\\\n( --- )\n  ~~~\n  zZz",
        "thinking": "  ___\n /o o\\\n( ... )\n  ~~~\n  ...",
        "happy":    "  ___\n /^ ^\\\n( =D )\n  ~~~\n \\o/",
        "worried":  "  ___\n /o o\\\n( ~~~ )\n  ~~~\n  /|\\",
        "alert":    "  ___\n /! !\\\n( !!! )\n  ~~~\n  /!\\",
        "loading":  "  ___\n /- -\\\n( nom )\n  ~~~\n [|||]",
    }

    # All extra-item SKU patterns that need resolution or skipping
    _EXTRA_SKU_INFO = {
        "CEX-EC":  {"type": "cheese", "resolve": True,  "label": "Curator's Extra Cheese"},
        "CEX-EM":  {"type": "meat",   "resolve": False, "label": "Curator's Extra Meat"},
        "CEX-EA":  {"type": "accom",  "resolve": False, "label": "Curator's Extra Accompaniment"},
        "EX-EC":   {"type": "cheese", "resolve": True,  "label": "Member Extra Cheese"},
        "EX-EM":   {"type": "meat",   "resolve": False, "label": "Member Extra Meat"},
        "EX-EA":   {"type": "accom",  "resolve": False, "label": "Member Extra Accompaniment"},
    }

    def _build_fulfillment_tab(self):
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text="  Fulfillment  ")

        # ── Top bar: mascot + toolbar ──
        top_frame = ttk.Frame(tab)
        top_frame.pack(fill="x", padx=10, pady=(10, 0))

        # Mascot panel (left side) — Canvas-based for animation
        mascot_frame = ttk.Frame(top_frame, width=130)
        mascot_frame.pack(side="left", fill="y", padx=(0, 10))
        mascot_frame.pack_propagate(False)

        self._fp_mascot_canvas = tk.Canvas(
            mascot_frame, width=120, height=80, bg=_BG2,
            highlightthickness=1, highlightbackground=_SEP)
        self._fp_mascot_canvas.pack(fill="x", padx=4, pady=4)
        self._fp_wheel_angle = 0
        self._fp_wheel_animating = False
        self._fp_draw_mascot("idle")

        self._fp_mascot_msg = tk.Label(
            mascot_frame, text="Ready to plan!",
            font=("Segoe UI", 8), bg=_BG, fg=_FG2,
            wraplength=120, justify="center")
        self._fp_mascot_msg.pack(pady=(0, 2))

        # Toolbar (right of mascot)
        toolbar = ttk.Frame(top_frame)
        toolbar.pack(side="left", fill="x", expand=True)

        btn_row1 = ttk.Frame(toolbar)
        btn_row1.pack(fill="x", pady=(0, 5))

        ttk.Button(btn_row1, text="Pull Recharge",
                   command=self._fp_pull_recharge).pack(
                       side="left", padx=(0, 5))
        ttk.Button(btn_row1, text="Pull Shopify",
                   command=self._fp_pull_shopify).pack(
                       side="left", padx=(0, 5))
        ttk.Button(btn_row1, text="Import CSV",
                   command=self._fp_import_csv).pack(
                       side="left", padx=(0, 5))
        ttk.Button(btn_row1, text="Calculate NET",
                   style="Run.TButton",
                   command=self._fp_calculate).pack(
                       side="left", padx=(0, 5))
        ttk.Button(btn_row1, text="Export CSV",
                   command=self._fp_export_csv).pack(
                       side="left", padx=(0, 5))
        ttk.Separator(btn_row1, orient="vertical").pack(
            side="left", fill="y", padx=8)
        ttk.Button(btn_row1, text="Auto-Assign",
                   command=self._fp_auto_assign).pack(
                       side="left", padx=(0, 5))
        ttk.Button(btn_row1, text="What-If",
                   command=self._fp_what_if).pack(
                       side="left", padx=(0, 5))

        btn_row2 = ttk.Frame(toolbar)
        btn_row2.pack(fill="x")

        ttk.Label(btn_row2, text="Fulfillment Date:").pack(
            side="left", padx=(0, 5))
        self._fp_date_var = tk.StringVar(
            value=datetime.datetime.now().strftime("%Y-%m-%d"))
        ttk.Entry(btn_row2, textvariable=self._fp_date_var,
                  width=12).pack(side="left", padx=(0, 10))

        ttk.Label(btn_row2, text="Filter:").pack(
            side="left", padx=(0, 5))
        self._fp_filter_var = tk.StringVar(value="CH-*")
        fp_filter_combo = ttk.Combobox(
            btn_row2, textvariable=self._fp_filter_var, width=10,
            values=["CH-*", "All", "Shortages", "Tight", "Surplus"],
            state="readonly")
        fp_filter_combo.pack(side="left", padx=(0, 15))
        fp_filter_combo.bind("<<ComboboxSelected>>",
                             lambda e: self._fp_apply_filter())

        self._fp_status_var = tk.StringVar(value="No data loaded")
        ttk.Label(btn_row2, textvariable=self._fp_status_var,
                  foreground=_FG2).pack(side="left", fill="x", expand=True)

        ttk.Separator(tab, orient="horizontal").pack(
            fill="x", padx=10, pady=8)

        # ── Main content: left=assignment grid, right=NET table ──
        content = ttk.PanedWindow(tab, orient="horizontal")
        content.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        # ── Left panel: Assignment Grid + Shelf Life ──
        left = ttk.Frame(content)
        content.add(left, weight=1)

        ttk.Label(left, text="PR-CJAM / CEX-EC Assignments",
                  font=("Segoe UI", 11, "bold")).pack(
                      anchor="w", padx=5, pady=(5, 2))
        ttk.Label(left, text="Double-click to change. "
                  "Green=OK, Red=+/-2 violation.",
                  foreground=_FG2, font=("Segoe UI", 8)).pack(
                      anchor="w", padx=5, pady=(0, 5))

        # Assignment treeview
        assign_frame = ttk.Frame(left)
        assign_frame.pack(fill="both", expand=True, padx=5, pady=5)

        assign_cols = ("curation", "prcjam_cheese", "prcjam_demand",
                       "cexec_cheese", "cexec_demand", "cexec_split",
                       "constraint")
        self._fp_assign_tree = ttk.Treeview(
            assign_frame, columns=assign_cols, show="headings",
            height=12, selectmode="browse")

        col_widths = {
            "curation": ("Curation", 75),
            "prcjam_cheese": ("PR-CJAM", 110),
            "prcjam_demand": ("Qty", 45),
            "cexec_cheese": ("CEX-EC", 110),
            "cexec_demand": ("Qty", 45),
            "cexec_split": ("Split", 60),
            "constraint": ("+/-2", 65),
        }
        for col, (heading, width) in col_widths.items():
            self._fp_assign_tree.heading(col, text=heading)
            self._fp_assign_tree.column(col, width=width, minwidth=35)

        self._fp_assign_tree.tag_configure("ok", foreground="#80cc80")
        self._fp_assign_tree.tag_configure("violation", foreground="#ff6666")
        self._fp_assign_tree.tag_configure("warning", foreground="#ffcc44")

        assign_sb = ttk.Scrollbar(assign_frame, orient="vertical",
                                  command=self._fp_assign_tree.yview)
        self._fp_assign_tree.configure(yscrollcommand=assign_sb.set)
        self._fp_assign_tree.pack(side="left", fill="both", expand=True)
        assign_sb.pack(side="right", fill="y")

        self._fp_assign_tree.bind("<Double-1>", self._fp_on_assign_dblclick)

        # Drop target highlight for assignment tree
        self._fp_assign_tree.tag_configure("drop_hover",
                                            background="#2a4a2a")

        # ── Shelf life / expiring panel ──
        shelf_frame = ttk.LabelFrame(left, text="Shelf Life (21-day cut cheese)")
        shelf_frame.pack(fill="x", padx=5, pady=(5, 0))

        self._fp_shelf_tree = ttk.Treeview(
            shelf_frame,
            columns=("sku", "days_left", "qty", "action"),
            show="headings", height=4, selectmode="browse")
        self._fp_shelf_tree.heading("sku", text="SKU")
        self._fp_shelf_tree.heading("days_left", text="Days Left")
        self._fp_shelf_tree.heading("qty", text="Qty")
        self._fp_shelf_tree.heading("action", text="Action")
        self._fp_shelf_tree.column("sku", width=100)
        self._fp_shelf_tree.column("days_left", width=65, anchor="e")
        self._fp_shelf_tree.column("qty", width=50, anchor="e")
        self._fp_shelf_tree.column("action", width=120)
        self._fp_shelf_tree.tag_configure("expired",
                                          foreground="#ff4444")
        self._fp_shelf_tree.tag_configure("expiring",
                                          foreground="#ffcc44")
        self._fp_shelf_tree.tag_configure("fresh",
                                          foreground="#80cc80")
        self._fp_shelf_tree.pack(fill="x", padx=5, pady=5)

        # ── Right panel: Saturday NET ──
        right = ttk.Frame(content)
        content.add(right, weight=2)

        right_header = ttk.Frame(right)
        right_header.pack(fill="x", padx=5, pady=(5, 2))

        ttk.Label(right_header, text="Saturday NET",
                  font=("Segoe UI", 11, "bold")).pack(
                      side="left")

        # Action buttons row
        ttk.Button(right_header, text="Suggest Fixes",
                   command=self._fp_suggest_fixes).pack(
                       side="right", padx=(5, 0))
        ttk.Button(right_header, text="Wed PO",
                   command=self._fp_generate_wed_po).pack(
                       side="right", padx=(5, 0))
        ttk.Button(right_header, text="Variety Check",
                   command=self._fp_variety_check).pack(
                       side="right", padx=(5, 0))

        # Score display
        self._fp_score_var = tk.StringVar(value="")
        tk.Label(right_header, textvariable=self._fp_score_var,
                 font=("Segoe UI", 9, "bold"), bg=_BG, fg="#f0c040").pack(
                     side="right", padx=(0, 15))

        # Summary bar
        self._fp_summary_var = tk.StringVar(value="")
        ttk.Label(right, textvariable=self._fp_summary_var,
                  foreground=_FG2, font=("Segoe UI", 9)).pack(
                      anchor="w", padx=5, pady=(0, 5))

        # Level-complete banner (hidden initially)
        self._fp_banner_frame = tk.Frame(right, bg="#1a4d1a", height=30)
        self._fp_banner_label = tk.Label(
            self._fp_banner_frame,
            text="ALL CLEAR - No shortages!",
            font=("Segoe UI", 11, "bold"), bg="#1a4d1a", fg="#80ff80")
        self._fp_banner_label.pack(pady=4)
        # Not packed yet — shown when all clear

        # Multi-week notebook
        self._fp_week_nb = ttk.Notebook(right)
        self._fp_week_nb.pack(fill="both", expand=True, padx=5, pady=5)

        # Tab: This Saturday
        this_sat_frame = ttk.Frame(self._fp_week_nb)
        self._fp_week_nb.add(this_sat_frame, text="  This Saturday  ")

        # NET treeview for this Saturday
        net_frame = ttk.Frame(this_sat_frame)
        net_frame.pack(fill="both", expand=True)

        net_cols = ("sku", "available", "direct", "prcjam", "cexec",
                    "exec", "total_demand", "net", "headroom", "status")
        self._fp_net_tree = ttk.Treeview(
            net_frame, columns=net_cols, show="headings",
            height=20, selectmode="browse")

        net_col_widths = {
            "sku": ("SKU", 95),
            "available": ("Avail", 55),
            "direct": ("Direct", 55),
            "prcjam": ("PRCJAM", 55),
            "cexec": ("CEXEC", 50),
            "exec": ("EXEC", 45),
            "total_demand": ("Total", 55),
            "net": ("NET", 60),
            "headroom": ("Headroom", 75),
            "status": ("Status", 75),
        }
        self._fp_sort_col = "net"
        self._fp_sort_rev = False
        for col, (heading, width) in net_col_widths.items():
            self._fp_net_tree.heading(
                col, text=heading,
                command=lambda c=col: self._fp_sort_by(c))
            self._fp_net_tree.column(col, width=width, minwidth=35,
                                     anchor="e" if col not in ("sku", "status", "headroom") else "w")

        # Row tags for NET coloring
        self._fp_net_tree.tag_configure("shortage",
                                        foreground="#ff4444",
                                        background="#3a1a1a")
        self._fp_net_tree.tag_configure("tight",
                                        foreground="#ffcc44",
                                        background="#3a3a1a")
        self._fp_net_tree.tag_configure("ok", foreground=_FG)
        self._fp_net_tree.tag_configure("surplus",
                                        foreground="#66cc66")
        self._fp_net_tree.tag_configure("no_demand",
                                        foreground=_FG2)

        net_sb = ttk.Scrollbar(net_frame, orient="vertical",
                               command=self._fp_net_tree.yview)
        self._fp_net_tree.configure(yscrollcommand=net_sb.set)
        self._fp_net_tree.pack(side="left", fill="both", expand=True)
        net_sb.pack(side="right", fill="y")

        # Right-click context menu on NET tree
        self._fp_net_menu = tk.Menu(self._fp_net_tree, tearoff=0,
                                    bg=_BG2, fg=_FG, activebackground=_ACC)
        self._fp_net_menu.add_command(
            label="Suggest fix for this SKU",
            command=self._fp_suggest_fix_selected)
        self._fp_net_menu.add_command(
            label="View demand breakdown",
            command=self._fp_view_demand_detail)
        self._fp_net_tree.bind("<Button-3>", self._fp_net_context_menu)

        # Drag-and-drop: drag CH-* from NET tree onto assignment grid
        self._fp_drag_data = {"sku": None, "x": 0, "y": 0}
        self._fp_drag_label = None
        self._fp_net_tree.bind("<ButtonPress-1>", self._fp_drag_start)
        self._fp_net_tree.bind("<B1-Motion>", self._fp_drag_motion)
        self._fp_net_tree.bind("<ButtonRelease-1>", self._fp_drag_drop)

        # Session score tracker
        self._fp_session_score = 0

        # Tab: Next Saturday
        next_sat_frame = ttk.Frame(self._fp_week_nb)
        self._fp_week_nb.add(next_sat_frame, text="  Next Saturday  ")

        next_net_frame = ttk.Frame(next_sat_frame)
        next_net_frame.pack(fill="both", expand=True)

        next_cols = ("sku", "carry_fwd", "next_demand", "next_net", "status")
        self._fp_next_tree = ttk.Treeview(
            next_net_frame, columns=next_cols, show="headings",
            height=20, selectmode="browse")
        self._fp_next_tree.heading("sku", text="SKU")
        self._fp_next_tree.heading("carry_fwd", text="Carry Fwd")
        self._fp_next_tree.heading("next_demand", text="Next Demand")
        self._fp_next_tree.heading("next_net", text="NET")
        self._fp_next_tree.heading("status", text="Status")
        self._fp_next_tree.column("sku", width=100)
        self._fp_next_tree.column("carry_fwd", width=80, anchor="e")
        self._fp_next_tree.column("next_demand", width=90, anchor="e")
        self._fp_next_tree.column("next_net", width=80, anchor="e")
        self._fp_next_tree.column("status", width=80)

        self._fp_next_tree.tag_configure("shortage",
                                          foreground="#ff4444",
                                          background="#3a1a1a")
        self._fp_next_tree.tag_configure("tight",
                                          foreground="#ffcc44",
                                          background="#3a3a1a")
        self._fp_next_tree.tag_configure("ok", foreground=_FG)
        self._fp_next_tree.tag_configure("plan",
                                          foreground="#6699cc")

        next_sb = ttk.Scrollbar(next_net_frame, orient="vertical",
                                command=self._fp_next_tree.yview)
        self._fp_next_tree.configure(yscrollcommand=next_sb.set)
        self._fp_next_tree.pack(side="left", fill="both", expand=True)
        next_sb.pack(side="right", fill="y")

        # CEX-EC split ratios: {curation: {sku: ratio, sku2: ratio}}
        self._fp_cexec_splits = self.saved.get("cexec_splits", {})

        # Populate assignment grid with current settings
        self._fp_refresh_assignments()
        # Init shopify orders store
        self._fp_shopify_orders = []
        self._fp_csv_demand = defaultdict(int)  # from imported CSV

        # Keyboard shortcuts
        self.root.bind("<Control-Return>", lambda e: self._fp_calculate())
        tab.bind("<Control-i>", lambda e: self._fp_import_csv())
        tab.bind("<Control-e>", lambda e: self._fp_export_csv())

    # ── Mascot / animation helpers ──

    def _fp_draw_mascot(self, state):
        """Draw mascot on Canvas. Cheese wheel with face."""
        c = self._fp_mascot_canvas
        c.delete("all")
        cx, cy = 60, 40
        r = 28

        # Cheese wheel body (circle with wedge cut)
        c.create_oval(cx - r, cy - r, cx + r, cy + r,
                      fill="#f0c040", outline="#d4a020", width=2)
        # Wedge cut (triangle)
        angle = self._fp_wheel_angle
        ax = cx + r * 0.3 * math.cos(math.radians(angle + 30))
        ay = cy + r * 0.3 * math.sin(math.radians(angle + 30))
        bx = cx + r * math.cos(math.radians(angle + 15))
        by = cy + r * math.sin(math.radians(angle + 15))
        dx = cx + r * math.cos(math.radians(angle + 45))
        dy = cy + r * math.sin(math.radians(angle + 45))
        c.create_polygon(ax, ay, bx, by, dx, dy,
                         fill=_BG2, outline="#d4a020", width=1)

        # Holes in cheese
        for ho in [(0.3, -0.2, 4), (-0.4, 0.1, 3), (0.1, 0.3, 3),
                   (-0.2, -0.3, 2)]:
            hx = cx + r * ho[0]
            hy = cy + r * ho[1]
            hr = ho[2]
            c.create_oval(hx - hr, hy - hr, hx + hr, hy + hr,
                          fill="#e0b030", outline="#d4a020")

        # Eyes
        eye_y = cy - 8
        if state == "happy":
            c.create_text(cx - 8, eye_y, text="^", font=("Consolas", 10, "bold"), fill="#333")
            c.create_text(cx + 8, eye_y, text="^", font=("Consolas", 10, "bold"), fill="#333")
        elif state == "worried":
            c.create_text(cx - 8, eye_y, text="o", font=("Consolas", 9), fill="#333")
            c.create_text(cx + 8, eye_y, text="o", font=("Consolas", 9), fill="#333")
            # Sweat drop
            c.create_text(cx + r - 2, cy - r + 8, text="'",
                          font=("Consolas", 10), fill="#66aaff")
        elif state == "alert":
            c.create_text(cx - 8, eye_y, text="!", font=("Consolas", 10, "bold"), fill="#ff4444")
            c.create_text(cx + 8, eye_y, text="!", font=("Consolas", 10, "bold"), fill="#ff4444")
        elif state == "loading":
            c.create_text(cx - 8, eye_y, text="-", font=("Consolas", 10), fill="#333")
            c.create_text(cx + 8, eye_y, text="-", font=("Consolas", 10), fill="#333")
        elif state == "thinking":
            c.create_text(cx - 8, eye_y, text="o", font=("Consolas", 9), fill="#333")
            c.create_text(cx + 8, eye_y, text="o", font=("Consolas", 9), fill="#333")
            c.create_text(cx + r + 5, cy - r, text="?",
                          font=("Consolas", 11, "bold"), fill=_ACC)
        else:  # idle
            c.create_text(cx - 8, eye_y, text="o", font=("Consolas", 9), fill="#333")
            c.create_text(cx + 8, eye_y, text="o", font=("Consolas", 9), fill="#333")

        # Mouth
        mouth_y = cy + 6
        if state == "happy":
            c.create_arc(cx - 8, mouth_y - 4, cx + 8, mouth_y + 6,
                         start=200, extent=140, style="arc",
                         outline="#333", width=2)
        elif state == "worried":
            c.create_arc(cx - 6, mouth_y, cx + 6, mouth_y + 8,
                         start=20, extent=140, style="arc",
                         outline="#333", width=2)
        elif state == "loading":
            c.create_text(cx, mouth_y + 2, text="nom",
                          font=("Consolas", 7), fill="#333")
        else:
            c.create_line(cx - 5, mouth_y + 2, cx + 5, mouth_y + 2,
                          fill="#333", width=2)

    def _fp_set_mascot(self, state, msg=""):
        self.fp_mascot_state = state
        self._fp_draw_mascot(state)
        if msg:
            self._fp_mascot_msg.configure(text=msg)
        # Start/stop wheel animation
        if state == "loading" and not self._fp_wheel_animating:
            self._fp_wheel_animating = True
            self._fp_animate_wheel()
        elif state != "loading":
            self._fp_wheel_animating = False

    def _fp_animate_wheel(self):
        """Animate cheese wheel being consumed (wedge rotates)."""
        if not self._fp_wheel_animating:
            return
        self._fp_wheel_angle = (self._fp_wheel_angle + 15) % 360
        self._fp_draw_mascot("loading")
        self.root.after(100, self._fp_animate_wheel)

    # ── Assignment grid ──

    def _fp_refresh_assignments(self):
        """Populate assignment treeview from current pr_cjam and cex_ec."""
        tree = self._fp_assign_tree
        tree.delete(*tree.get_children())

        all_curations = list(CURATION_ORDER) + ["NMS", "BYO", "SS"]
        for cur in all_curations:
            prcjam_info = self.pr_cjam.get(cur, {})
            prcjam_cheese = prcjam_info.get("cheese", "") if isinstance(
                prcjam_info, dict) else str(prcjam_info)
            prcjam_demand = ""

            cexec_cheese = self.cex_ec.get(cur, "")
            cexec_demand = ""
            splits = self._fp_cexec_splits.get(cur, {})
            if splits:
                parts = [f"{int(v*100)}% {k}" for k, v in splits.items()]
                cexec_split = " / ".join(parts)
            else:
                cexec_split = ""

            # Check +/-2 constraint
            constraint = self._fp_check_constraint(cur, prcjam_cheese,
                                                    cexec_cheese)

            tag = "ok" if constraint == "OK" else "violation"

            tree.insert("", "end", iid=cur, values=(
                cur, prcjam_cheese, prcjam_demand,
                cexec_cheese, cexec_demand, cexec_split,
                constraint,
            ), tags=(tag,))

    def _fp_check_constraint(self, curation, prcjam_cheese, cexec_cheese):
        """Check +/-2 constraint for a curation's assignments."""
        if curation not in CURATION_ORDER:
            return "OK"
        idx = CURATION_ORDER.index(curation)

        # Collect recipe cheeses from nearby curations (+/-2)
        nearby_cheeses = set()
        for offset in [-2, -1, 1, 2]:
            ni = idx + offset
            if 0 <= ni < len(CURATION_ORDER):
                neighbor = CURATION_ORDER[ni]
                recipe = self.curation_recipes.get(neighbor, [])
                for item in recipe:
                    sku = item[0] if isinstance(item, (list, tuple)) else item
                    if sku.startswith("CH-"):
                        nearby_cheeses.add(sku)
                # Also check neighbor's assignments
                n_prcjam = self.pr_cjam.get(neighbor, {})
                if isinstance(n_prcjam, dict) and n_prcjam.get("cheese"):
                    nearby_cheeses.add(n_prcjam["cheese"])
                n_cexec = self.cex_ec.get(neighbor, "")
                if n_cexec:
                    nearby_cheeses.add(n_cexec)

        violations = []
        if prcjam_cheese and prcjam_cheese in nearby_cheeses:
            violations.append("PR")
        if cexec_cheese and cexec_cheese in nearby_cheeses:
            violations.append("EC")

        return "CONFLICT: " + "+".join(violations) if violations else "OK"

    def _fp_on_assign_dblclick(self, event):
        """Double-click to reassign PR-CJAM or CEX-EC cheese."""
        tree = self._fp_assign_tree
        item = tree.focus()
        if not item:
            return

        # Determine which column was clicked
        col = tree.identify_column(event.x)
        col_idx = int(col.replace("#", "")) - 1
        cur = item  # iid is the curation name

        if col_idx in (1, 2):  # PR-CJAM cheese or demand
            self._fp_edit_assignment(cur, "prcjam")
        elif col_idx == 5:  # Split column
            self._fp_edit_split(cur)
        elif col_idx in (3, 4):  # CEX-EC cheese or demand
            self._fp_edit_assignment(cur, "cexec")

    def _fp_edit_assignment(self, curation, slot_type):
        """Open a picker to select a new cheese for this slot."""
        dlg = tk.Toplevel(self.root)
        dlg.title(f"Assign {slot_type.upper()} - {curation}")
        dlg.configure(bg=_BG)
        dlg.transient(self.root)
        dlg.grab_set()
        dlg.geometry("400x500")

        ttk.Label(dlg, text=f"Select cheese for {curation} "
                  f"{'PR-CJAM' if slot_type == 'prcjam' else 'CEX-EC'}:",
                  font=("Segoe UI", 10, "bold")).pack(
                      padx=15, pady=(15, 5), anchor="w")

        # Show eligible cheeses ranked by headroom
        cheese_frame = ttk.Frame(dlg)
        cheese_frame.pack(fill="both", expand=True, padx=15, pady=5)

        cheese_cols = ("sku", "available", "net", "constraint")
        cheese_tree = ttk.Treeview(
            cheese_frame, columns=cheese_cols, show="headings",
            height=15, selectmode="browse")
        cheese_tree.heading("sku", text="Cheese SKU")
        cheese_tree.heading("available", text="Avail")
        cheese_tree.heading("net", text="Est NET")
        cheese_tree.heading("constraint", text="+/-2")
        cheese_tree.column("sku", width=120)
        cheese_tree.column("available", width=60, anchor="e")
        cheese_tree.column("net", width=60, anchor="e")
        cheese_tree.column("constraint", width=70)

        cheese_tree.tag_configure("ok", foreground="#80cc80")
        cheese_tree.tag_configure("blocked", foreground="#ff6666")

        # Build candidate list from inventory
        candidates = []
        for sku, data in self.inventory.items():
            if not sku.startswith("CH-"):
                continue
            qty = data.get("qty", 0) if isinstance(data, dict) else 0
            if qty <= 0:
                continue

            # Check +/-2 for this candidate
            test_prcjam = sku if slot_type == "prcjam" else (
                self.pr_cjam.get(curation, {}).get("cheese", "")
                if isinstance(self.pr_cjam.get(curation), dict)
                else "")
            test_cexec = sku if slot_type == "cexec" else self.cex_ec.get(
                curation, "")
            constraint = self._fp_check_constraint(
                curation, test_prcjam, test_cexec)

            candidates.append((sku, qty, constraint))

        # Sort: OK first, then by qty descending
        candidates.sort(key=lambda x: (0 if x[2] == "OK" else 1, -x[1]))

        for sku, qty, constraint in candidates:
            tag = "ok" if constraint == "OK" else "blocked"
            cheese_tree.insert("", "end", iid=sku, values=(
                sku, qty, "", constraint), tags=(tag,))

        cheese_tree.pack(fill="both", expand=True)

        def _apply():
            sel = cheese_tree.focus()
            if not sel:
                return
            if slot_type == "prcjam":
                if isinstance(self.pr_cjam.get(curation), dict):
                    self.pr_cjam[curation]["cheese"] = sel
                else:
                    self.pr_cjam[curation] = {"cheese": sel, "jam": ""}
            else:
                self.cex_ec[curation] = sel

            self._fp_pop_icon("+1", "#80ff80")
            self._fp_update_score(1, f"{sel} assigned!")
            self.root.after(2000, lambda: self._fp_set_mascot(
                "idle", "Ready to plan!"))

            self._fp_refresh_assignments()
            if self.fp_results:
                self._fp_calculate()
            dlg.destroy()

        btn_frame = ttk.Frame(dlg)
        btn_frame.pack(fill="x", padx=15, pady=(0, 15))
        tk.Button(btn_frame, text="Cancel", command=dlg.destroy,
                  bg=_BG3, fg=_FG, relief="flat", padx=10, pady=4
                  ).pack(side="right", padx=(5, 0))
        tk.Button(btn_frame, text="Assign", command=_apply,
                  bg=_GREEN, fg="white", relief="flat", padx=10, pady=4
                  ).pack(side="right")
        cheese_tree.bind("<Double-1>", lambda e: _apply())

    # ── CSV import (offline data) ──

    def _fp_import_csv(self):
        """Import order-dashboard or charges CSV for offline planning."""
        path = filedialog.askopenfilename(
            filetypes=[("CSV", "*.csv"), ("All", "*.*")],
            title="Import demand CSV",
            parent=self.root)
        if not path:
            return

        self._fp_set_mascot("loading", "Importing CSV...")
        imported = defaultdict(int)
        row_count = 0

        try:
            with open(path, encoding="utf-8-sig") as f:
                reader = csv.DictReader(f)
                fields = reader.fieldnames or []

                if "All SKUs" in fields:
                    # Order-dashboard format
                    for row in reader:
                        row_count += 1
                        all_skus = row.get("All SKUs", "")
                        for sku in (s.strip() for s in all_skus.split(",")
                                    if s.strip()):
                            upper = sku.upper()
                            if upper.startswith("PR-CJAM-"):
                                suffix = upper.split("PR-CJAM-", 1)[1]
                                prcjam_info = self.pr_cjam.get(suffix, {})
                                cheese = (prcjam_info.get("cheese", "")
                                          if isinstance(prcjam_info, dict)
                                          else "")
                                if cheese:
                                    imported[cheese] += 1
                            elif upper.startswith("CEX-EC-"):
                                suffix = upper.split("CEX-EC-", 1)[1]
                                cheese = self.cex_ec.get(suffix, "")
                                if cheese:
                                    imported[cheese] += 1
                            elif upper == "CEX-EC" or upper == "EX-EC":
                                pass  # bare — skip
                            elif upper.startswith("CH-"):
                                imported[sku] += 1

                elif "line_item_sku" in fields:
                    # Recharge charges format
                    for row in reader:
                        sku = row.get("line_item_sku", "").strip()
                        if not sku:
                            continue
                        try:
                            qty = int(float(
                                row.get("line_item_quantity", "1") or "1"))
                        except ValueError:
                            qty = 1
                        upper = sku.upper()
                        row_count += 1
                        if upper.startswith("CH-"):
                            imported[sku] += qty
                        elif upper.startswith("PR-CJAM-"):
                            suffix = upper.split("PR-CJAM-", 1)[1]
                            if suffix == "GEN":
                                continue
                            prcjam_info = self.pr_cjam.get(suffix, {})
                            cheese = (prcjam_info.get("cheese", "")
                                      if isinstance(prcjam_info, dict)
                                      else "")
                            if cheese:
                                imported[cheese] += qty
                        elif upper.startswith("CEX-EC-"):
                            suffix = upper.split("CEX-EC-", 1)[1]
                            cheese = self.cex_ec.get(suffix, "")
                            if cheese:
                                imported[cheese] += qty
                else:
                    messagebox.showwarning(
                        "Unknown Format",
                        "CSV must have 'All SKUs' (dashboard) or "
                        "'line_item_sku' (charges) column.",
                        parent=self.root)
                    self._fp_set_mascot("idle", "Ready to plan!")
                    return

            self._fp_csv_demand = imported
            self._fp_set_mascot("happy",
                                f"{row_count} rows imported!")
            self._fp_status_var.set(
                f"CSV: {row_count} rows, {len(imported)} CH-* SKUs "
                f"from {os.path.basename(path)}")
            self.root.after(3000, lambda: self._fp_set_mascot(
                "idle", "Ready to plan!"))

        except Exception as e:
            self._fp_set_mascot("alert", "Import failed!")
            messagebox.showerror("Import Error", str(e),
                                 parent=self.root)

    # ── Filter ──

    def _fp_apply_filter(self):
        """Re-display NET tree with current filter."""
        if not self.fp_results:
            return
        filt = self._fp_filter_var.get()
        tree = self._fp_net_tree
        tree.delete(*tree.get_children())

        for r in self.fp_results:
            if r["status"] == "NO DEMAND" and r["available"] == 0:
                continue
            if filt == "CH-*" and not r["sku"].startswith("CH-"):
                continue
            if filt == "Shortages" and r["status"] != "SHORTAGE":
                continue
            if filt == "Tight" and r["status"] not in ("SHORTAGE", "TIGHT"):
                continue
            if filt == "Surplus" and r["status"] != "SURPLUS":
                continue

            headroom = self._fp_headroom_bar(r["net"], r["total_demand"])
            tree.insert("", "end", values=(
                r["sku"], r["available"],
                r["direct"], r["prcjam"],
                r["cexec"], r["exec"],
                r["total_demand"], f"{r['net']:+d}",
                headroom, r["status"],
            ), tags=(r["tag"],))

    def _fp_headroom_bar(self, net, total_demand):
        """Return a text headroom indicator."""
        if total_demand == 0:
            return ""
        if net < 0:
            return "NEED " + str(abs(net))
        ratio = net / total_demand if total_demand > 0 else 999
        if ratio > 2:
            return "+++"
        elif ratio > 1:
            return "++"
        elif ratio > 0.5:
            return "+"
        elif ratio > 0.2:
            return "~"
        else:
            return "LOW"

    # ── Context menu ──

    def _fp_net_context_menu(self, event):
        tree = self._fp_net_tree
        item = tree.identify_row(event.y)
        if item:
            tree.selection_set(item)
            tree.focus(item)
            self._fp_net_menu.post(event.x_root, event.y_root)

    def _fp_view_demand_detail(self):
        """Show demand breakdown for selected SKU."""
        sel = self._fp_net_tree.focus()
        if not sel:
            return
        vals = self._fp_net_tree.item(sel, "values")
        sku = vals[0]
        r = next((r for r in self.fp_results if r["sku"] == sku), None)
        if not r:
            return

        # Find which curations assign this cheese
        assigned_by = []
        for cur in list(CURATION_ORDER) + ["NMS", "BYO", "SS"]:
            prcjam_info = self.pr_cjam.get(cur, {})
            cheese = (prcjam_info.get("cheese", "")
                      if isinstance(prcjam_info, dict) else "")
            if cheese == sku:
                assigned_by.append(f"PR-CJAM-{cur}")
            if self.cex_ec.get(cur) == sku:
                assigned_by.append(f"CEX-EC-{cur}")

        # Find which recipes use this cheese
        in_recipes = []
        for cur, recipe in self.curation_recipes.items():
            for item in recipe:
                s = item[0] if isinstance(item, (list, tuple)) else item
                if s == sku:
                    in_recipes.append(cur)

        # Check expiration
        exp_info = ""
        inv_data = self.inventory.get(sku, {})
        if isinstance(inv_data, dict) and inv_data.get("expiration_dates"):
            dates = inv_data["expiration_dates"]
            if dates:
                exp_info = f"\nEarliest expiration: {dates[0]}"

        detail = (
            f"SKU: {sku}\n"
            f"Available: {r['available']}\n\n"
            f"Demand Breakdown:\n"
            f"  Direct (orders/addons): {r['direct']}\n"
            f"  PR-CJAM: {r['prcjam']}\n"
            f"  CEX-EC: {r['cexec']}\n"
            f"  EX-EC: {r['exec']}\n"
            f"  Total: {r['total_demand']}\n\n"
            f"NET: {r['net']:+d} ({r['status']})\n\n"
            f"Assigned by: {', '.join(assigned_by) if assigned_by else 'None'}\n"
            f"In recipes: {', '.join(in_recipes) if in_recipes else 'None'}"
            f"{exp_info}"
        )

        dlg = tk.Toplevel(self.root)
        dlg.title(f"Demand Detail - {sku}")
        dlg.configure(bg=_BG)
        dlg.transient(self.root)
        dlg.geometry("350x380")

        text = tk.Text(dlg, bg=_BG2, fg=_FG, font=("Consolas", 10),
                       wrap="word", relief="flat", padx=10, pady=10)
        text.insert("1.0", detail)
        text.configure(state="disabled")
        text.pack(fill="both", expand=True, padx=10, pady=10)

        tk.Button(dlg, text="Close", command=dlg.destroy,
                  bg=_BG3, fg=_FG, relief="flat", padx=10, pady=4
                  ).pack(pady=(0, 10))

    # ── Shortage suggestions ──

    def _fp_suggest_fix_selected(self):
        """Suggest fix for the selected shortage SKU."""
        sel = self._fp_net_tree.focus()
        if not sel:
            return
        vals = self._fp_net_tree.item(sel, "values")
        sku = vals[0]
        r = next((r for r in self.fp_results if r["sku"] == sku), None)
        if not r or r["net"] >= 0:
            messagebox.showinfo("No Shortage",
                                f"{sku} is not in shortage (NET={r['net']:+d})",
                                parent=self.root)
            return
        self._fp_show_fix_dialog(sku, r)

    def _fp_suggest_fixes(self):
        """Show fix suggestions for all shortages."""
        shortages = [r for r in self.fp_results if r["status"] == "SHORTAGE"]
        if not shortages:
            self._fp_set_mascot("happy", "No shortages!")
            messagebox.showinfo("All Clear",
                                "No shortages to fix!",
                                parent=self.root)
            return

        dlg = tk.Toplevel(self.root)
        dlg.title("Shortage Fix Suggestions")
        dlg.configure(bg=_BG)
        dlg.transient(self.root)
        dlg.geometry("600x500")

        ttk.Label(dlg, text=f"{len(shortages)} Shortages Need Fixing",
                  font=("Segoe UI", 11, "bold")).pack(
                      padx=15, pady=(15, 5), anchor="w")

        text = tk.Text(dlg, bg=_BG2, fg=_FG, font=("Consolas", 9),
                       wrap="word", relief="flat", padx=10, pady=10)
        text.pack(fill="both", expand=True, padx=10, pady=5)

        for r in shortages:
            sku = r["sku"]
            deficit = abs(r["net"])
            text.insert("end", f"\n{'=' * 50}\n", "header")
            text.insert("end",
                        f"{sku}: need {deficit} more "
                        f"(avail={r['available']}, demand={r['total_demand']})\n\n")

            suggestions = self._fp_get_suggestions(sku, deficit)
            for i, s in enumerate(suggestions, 1):
                text.insert("end", f"  {i}. {s}\n")

        text.configure(state="disabled")
        text.tag_configure("header", foreground=_ACC,
                           font=("Consolas", 9, "bold"))

        tk.Button(dlg, text="Close", command=dlg.destroy,
                  bg=_BG3, fg=_FG, relief="flat", padx=10, pady=4
                  ).pack(pady=(5, 10))

    def _fp_show_fix_dialog(self, sku, r):
        """Show fix dialog for a single shortage."""
        deficit = abs(r["net"])
        suggestions = self._fp_get_suggestions(sku, deficit)

        dlg = tk.Toplevel(self.root)
        dlg.title(f"Fix Shortage: {sku}")
        dlg.configure(bg=_BG)
        dlg.transient(self.root)
        dlg.geometry("450x350")

        ttk.Label(dlg, text=f"{sku}: need {deficit} more",
                  font=("Segoe UI", 11, "bold")).pack(
                      padx=15, pady=(15, 5), anchor="w")
        ttk.Label(dlg,
                  text=f"Available: {r['available']} | "
                       f"Demand: {r['total_demand']} | "
                       f"NET: {r['net']:+d}",
                  foreground=_FG2).pack(padx=15, anchor="w")

        text = tk.Text(dlg, bg=_BG2, fg=_FG, font=("Consolas", 10),
                       wrap="word", relief="flat", padx=10, pady=10)
        text.pack(fill="both", expand=True, padx=10, pady=10)

        text.insert("end", "Suggestions:\n\n")
        for i, s in enumerate(suggestions, 1):
            text.insert("end", f"  {i}. {s}\n\n")

        text.configure(state="disabled")

        tk.Button(dlg, text="Close", command=dlg.destroy,
                  bg=_BG3, fg=_FG, relief="flat", padx=10, pady=4
                  ).pack(pady=(0, 10))

    def _fp_get_suggestions(self, sku, deficit):
        """Generate fix suggestions for a shortage."""
        suggestions = []

        # Check wheel inventory for raw supply
        for wsku, wdata in self.wheel_inventory.items():
            if isinstance(wdata, dict) and wdata.get("target_sku") == sku:
                weight = float(wdata.get("weight_lbs", 0))
                count = int(wdata.get("count", 0))
                if weight > 0 and count > 0:
                    potential = int(weight * count * WHEEL_TO_SLICE_FACTOR)
                    wheels_needed = max(1, int(
                        deficit / (weight * WHEEL_TO_SLICE_FACTOR)) + 1)
                    suggestions.append(
                        f"MFG/Cut: {wheels_needed} wheel(s) of {wsku} "
                        f"({int(weight)} lbs each, ~{int(weight * WHEEL_TO_SLICE_FACTOR)} "
                        f"portions per wheel)")

        # Check open POs
        for po in self.open_pos:
            if po.get("sku") == sku and po.get("status") == "Open":
                eta = po.get("eta", "?")
                qty = po.get("qty", 0)
                suggestions.append(
                    f"PO incoming: {qty} units, ETA {eta} "
                    f"({'covers deficit' if int(float(qty)) >= deficit else 'partial'})")

        # Find substitution candidates
        # Which curations use this sku in PR-CJAM or CEX-EC?
        assigned_curations = []
        for cur in list(CURATION_ORDER) + ["NMS", "BYO", "SS"]:
            prcjam_info = self.pr_cjam.get(cur, {})
            cheese = (prcjam_info.get("cheese", "")
                      if isinstance(prcjam_info, dict) else "")
            if cheese == sku:
                assigned_curations.append(("PR-CJAM", cur))
            if self.cex_ec.get(cur) == sku:
                assigned_curations.append(("CEX-EC", cur))

        # For each assignment, find swap candidates
        for slot_type, cur in assigned_curations:
            if cur not in CURATION_ORDER:
                continue
            # Find cheeses with headroom that pass +/-2
            swaps = []
            for r2 in self.fp_results:
                if r2["sku"] == sku or not r2["sku"].startswith("CH-"):
                    continue
                if r2["net"] < deficit:
                    continue
                # Check constraint
                test_pr = r2["sku"] if slot_type == "PR-CJAM" else (
                    self.pr_cjam.get(cur, {}).get("cheese", "")
                    if isinstance(self.pr_cjam.get(cur), dict) else "")
                test_ec = r2["sku"] if slot_type == "CEX-EC" else (
                    self.cex_ec.get(cur, ""))
                constraint = self._fp_check_constraint(cur, test_pr, test_ec)
                if constraint == "OK":
                    swaps.append((r2["sku"], r2["net"]))

            swaps.sort(key=lambda x: -x[1])
            if swaps:
                top3 = swaps[:3]
                swap_str = ", ".join(
                    f"{s[0]}(+{s[1]})" for s in top3)
                suggestions.append(
                    f"Reassign {slot_type}-{cur}: swap to {swap_str}")

        # Partial substitution
        if deficit < 20:
            suggestions.append(
                f"Partial sub: substitute a different cheese for "
                f"{deficit} orders only")

        if not suggestions:
            suggestions.append(
                "No automatic suggestions. Consider a Wednesday PO "
                "or manual recipe swap.")

        return suggestions

    # ── Shelf life tracking ──

    def _fp_refresh_shelf_life(self):
        """Update shelf life panel with expiring cheese."""
        tree = self._fp_shelf_tree
        tree.delete(*tree.get_children())

        today = datetime.datetime.now().date()

        for sku, data in self.inventory.items():
            if not sku.startswith("CH-"):
                continue
            if not isinstance(data, dict):
                continue
            exp_dates = data.get("expiration_dates", [])
            if not exp_dates:
                continue

            for exp_str in exp_dates:
                try:
                    exp_date = datetime.datetime.strptime(
                        exp_str, "%Y-%m-%d").date()
                except (ValueError, TypeError):
                    continue

                days_left = (exp_date - today).days
                qty = data.get("qty", 0)

                if days_left < 0:
                    tag = "expired"
                    action = "EXPIRED - discard or discount"
                elif days_left <= 7:
                    tag = "expiring"
                    action = "Use ASAP - prioritize in assignments"
                elif days_left <= 14:
                    tag = "expiring"
                    action = "Plan to use this week"
                else:
                    tag = "fresh"
                    action = "OK"

                if days_left <= 14:
                    tree.insert("", "end", values=(
                        sku, days_left, qty, action), tags=(tag,))

    # ── Variety checker ──

    def _fp_variety_check(self):
        """Check that assignments maintain variety for subscribers."""
        issues = []

        # Build full cheese map per curation: recipe + PR-CJAM + CEX-EC
        cur_cheeses = {}
        for cur in CURATION_ORDER:
            cheeses = set()
            recipe = self.curation_recipes.get(cur, [])
            for item in recipe:
                sku = item[0] if isinstance(item, (list, tuple)) else item
                if sku.startswith("CH-"):
                    cheeses.add(sku)
            prcjam_info = self.pr_cjam.get(cur, {})
            pr_ch = (prcjam_info.get("cheese", "")
                     if isinstance(prcjam_info, dict) else "")
            if pr_ch:
                cheeses.add(pr_ch)
            ec_ch = self.cex_ec.get(cur, "")
            if ec_ch:
                cheeses.add(ec_ch)
            # Handle splits
            splits = self._fp_cexec_splits.get(cur, {})
            for split_sku in splits:
                cheeses.add(split_sku)
            cur_cheeses[cur] = cheeses

        # Check adjacent curations for overlap
        for i, cur in enumerate(CURATION_ORDER):
            for j in range(i + 1, min(i + 3, len(CURATION_ORDER))):
                neighbor = CURATION_ORDER[j]
                overlap = cur_cheeses[cur] & cur_cheeses[neighbor]
                if overlap:
                    dist = j - i
                    severity = "WARNING" if dist == 2 else "ISSUE"
                    for sku in overlap:
                        issues.append(
                            f"{severity}: {sku} appears in both "
                            f"{cur} and {neighbor} "
                            f"({dist} position{'s' if dist > 1 else ''} apart)")

        # Check PR-CJAM uniqueness
        pr_cheese_to_cur = defaultdict(list)
        for cur in list(CURATION_ORDER) + ["NMS", "BYO", "SS"]:
            prcjam_info = self.pr_cjam.get(cur, {})
            pr_ch = (prcjam_info.get("cheese", "")
                     if isinstance(prcjam_info, dict) else "")
            if pr_ch:
                pr_cheese_to_cur[pr_ch].append(cur)

        for cheese, curs in pr_cheese_to_cur.items():
            if len(curs) > 1:
                issues.append(
                    f"DUPLICATE PR-CJAM: {cheese} assigned to "
                    f"{', '.join(curs)} (should be unique)")

        dlg = tk.Toplevel(self.root)
        dlg.title("Variety Check")
        dlg.configure(bg=_BG)
        dlg.transient(self.root)
        dlg.geometry("500x400")

        if issues:
            self._fp_set_mascot("worried",
                                f"{len(issues)} variety issue(s)!")
            ttk.Label(dlg,
                      text=f"{len(issues)} Variety Issues Found",
                      font=("Segoe UI", 11, "bold"),
                      foreground="#ffcc44").pack(
                          padx=15, pady=(15, 5), anchor="w")
        else:
            self._fp_set_mascot("happy", "Variety looks great!")
            ttk.Label(dlg,
                      text="All Clear - Good variety!",
                      font=("Segoe UI", 11, "bold"),
                      foreground="#80cc80").pack(
                          padx=15, pady=(15, 5), anchor="w")

        text = tk.Text(dlg, bg=_BG2, fg=_FG, font=("Consolas", 10),
                       wrap="word", relief="flat", padx=10, pady=10)
        text.pack(fill="both", expand=True, padx=10, pady=5)

        if issues:
            for issue in issues:
                color = "#ff6666" if "ISSUE" in issue or "DUPLICATE" in issue else "#ffcc44"
                text.insert("end", f"  {issue}\n\n")
        else:
            text.insert("end", "  No overlap or duplicate issues.\n\n")
            text.insert("end", "  Cheeses per curation:\n")
            for cur in CURATION_ORDER:
                cheeses = sorted(cur_cheeses.get(cur, set()))
                text.insert("end", f"    {cur}: {', '.join(cheeses)}\n")

        text.configure(state="disabled")
        tk.Button(dlg, text="Close", command=dlg.destroy,
                  bg=_BG3, fg=_FG, relief="flat", padx=10, pady=4
                  ).pack(pady=(5, 10))

    # ── Wednesday PO generator ──

    def _fp_generate_wed_po(self):
        """Generate Wednesday PO to cover next-week shortfalls."""
        if not self.fp_results:
            messagebox.showinfo("No Data",
                                "Run Calculate first.",
                                parent=self.root)
            return

        # Estimate next-week demand (rough: same as this week)
        po_items = []
        for r in self.fp_results:
            if not r["sku"].startswith("CH-"):
                continue
            # Next week: carry forward this week's NET as inventory
            carry_fwd = max(0, r["net"])
            # Assume similar demand next week
            next_demand = r["total_demand"]
            next_net = carry_fwd - next_demand

            if next_net < 0:
                deficit = abs(next_net)
                # Check vendor catalog for case qty / MOQ
                vcat = self.vendor_catalog.get(r["sku"], {})
                case_qty = int(vcat.get("case_qty", 1)) if vcat else 1
                vendor = vcat.get("vendor", "") if vcat else ""
                unit_cost = float(vcat.get("unit_cost", 0)) if vcat else 0
                moq = int(vcat.get("moq", 0)) if vcat else 0

                # Round up to case qty
                order_qty = deficit
                if case_qty > 1:
                    order_qty = ((deficit + case_qty - 1) // case_qty) * case_qty
                if moq > 0 and order_qty < moq:
                    order_qty = moq

                po_items.append({
                    "sku": r["sku"],
                    "deficit": deficit,
                    "order_qty": order_qty,
                    "case_qty": case_qty,
                    "vendor": vendor,
                    "unit_cost": unit_cost,
                    "total_cost": round(order_qty * unit_cost, 2),
                    "carry_fwd": carry_fwd,
                    "next_demand": next_demand,
                })

        if not po_items:
            self._fp_set_mascot("happy", "No PO needed!")
            messagebox.showinfo("All Clear",
                                "No shortfalls projected for next week.",
                                parent=self.root)
            return

        # Show PO preview
        dlg = tk.Toplevel(self.root)
        dlg.title("Wednesday PO - Next Week Coverage")
        dlg.configure(bg=_BG)
        dlg.transient(self.root)
        dlg.geometry("700x500")

        ttk.Label(dlg,
                  text=f"Wednesday PO: {len(po_items)} items needed",
                  font=("Segoe UI", 11, "bold")).pack(
                      padx=15, pady=(15, 5), anchor="w")

        # Group by vendor
        by_vendor = defaultdict(list)
        for item in po_items:
            by_vendor[item["vendor"] or "Unknown"].append(item)

        tree_frame = ttk.Frame(dlg)
        tree_frame.pack(fill="both", expand=True, padx=10, pady=5)

        po_cols = ("sku", "deficit", "order_qty", "case_qty",
                   "vendor", "unit_cost", "total")
        po_tree = ttk.Treeview(
            tree_frame, columns=po_cols, show="headings",
            height=15, selectmode="browse")
        po_tree.heading("sku", text="SKU")
        po_tree.heading("deficit", text="Deficit")
        po_tree.heading("order_qty", text="Order Qty")
        po_tree.heading("case_qty", text="Case Qty")
        po_tree.heading("vendor", text="Vendor")
        po_tree.heading("unit_cost", text="Unit $")
        po_tree.heading("total", text="Total $")
        po_tree.column("sku", width=100)
        po_tree.column("deficit", width=65, anchor="e")
        po_tree.column("order_qty", width=75, anchor="e")
        po_tree.column("case_qty", width=65, anchor="e")
        po_tree.column("vendor", width=100)
        po_tree.column("unit_cost", width=65, anchor="e")
        po_tree.column("total", width=75, anchor="e")

        for vendor, items in sorted(by_vendor.items()):
            for item in items:
                po_tree.insert("", "end", values=(
                    item["sku"], item["deficit"], item["order_qty"],
                    item["case_qty"], item["vendor"],
                    f"${item['unit_cost']:.2f}",
                    f"${item['total_cost']:.2f}",
                ))

        po_sb = ttk.Scrollbar(tree_frame, orient="vertical",
                              command=po_tree.yview)
        po_tree.configure(yscrollcommand=po_sb.set)
        po_tree.pack(side="left", fill="both", expand=True)
        po_sb.pack(side="right", fill="y")

        # Total cost
        total_cost = sum(i["total_cost"] for i in po_items)
        total_units = sum(i["order_qty"] for i in po_items)
        ttk.Label(dlg,
                  text=f"Total: {total_units:,} units | "
                       f"${total_cost:,.2f} | "
                       f"{len(by_vendor)} vendor(s)",
                  font=("Segoe UI", 10, "bold")).pack(
                      padx=15, anchor="w")

        def _add_to_open_pos():
            today = datetime.datetime.now()
            wed = today + datetime.timedelta(
                days=(2 - today.weekday()) % 7 or 7)
            eta = (wed + datetime.timedelta(days=5)).strftime("%Y-%m-%d")

            for item in po_items:
                self.open_pos.append({
                    "sku": item["sku"],
                    "qty": item["order_qty"],
                    "eta": eta,
                    "type": "PO",
                    "vendor": item["vendor"],
                    "status": "Open",
                })

            self._fp_set_mascot("happy",
                                f"{len(po_items)} POs added!")
            messagebox.showinfo("POs Added",
                                f"{len(po_items)} items added to Open POs.\n"
                                f"ETA: {eta}",
                                parent=dlg)
            dlg.destroy()

        def _export_po_csv():
            path = filedialog.asksaveasfilename(
                defaultextension=".csv",
                filetypes=[("CSV", "*.csv")],
                initialfile=f"wednesday_po_{self._fp_date_var.get()}.csv",
                parent=dlg)
            if not path:
                return
            with open(path, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(["SKU", "Deficit", "Order Qty",
                                 "Case Qty", "Vendor", "Unit Cost",
                                 "Total Cost"])
                for item in po_items:
                    writer.writerow([
                        item["sku"], item["deficit"], item["order_qty"],
                        item["case_qty"], item["vendor"],
                        item["unit_cost"], item["total_cost"],
                    ])
            self._fp_set_mascot("happy", "PO CSV exported!")

        btn_frame = ttk.Frame(dlg)
        btn_frame.pack(fill="x", padx=15, pady=(5, 15))
        tk.Button(btn_frame, text="Close", command=dlg.destroy,
                  bg=_BG3, fg=_FG, relief="flat", padx=10, pady=4
                  ).pack(side="right", padx=(5, 0))
        tk.Button(btn_frame, text="Add to Open POs",
                  command=_add_to_open_pos,
                  bg=_GREEN, fg="white", relief="flat", padx=10, pady=4
                  ).pack(side="right", padx=(5, 0))
        tk.Button(btn_frame, text="Export CSV",
                  command=_export_po_csv,
                  bg=_ACC, fg="white", relief="flat", padx=10, pady=4
                  ).pack(side="right")

    # ── Split ratio editor ──

    def _fp_edit_split(self, curation):
        """Edit CEX-EC split ratio for a curation (e.g. MDT 64/36)."""
        dlg = tk.Toplevel(self.root)
        dlg.title(f"CEX-EC Split - {curation}")
        dlg.configure(bg=_BG)
        dlg.transient(self.root)
        dlg.grab_set()
        dlg.geometry("400x300")

        ttk.Label(dlg,
                  text=f"Split CEX-EC-{curation} across two cheeses:",
                  font=("Segoe UI", 10, "bold")).pack(
                      padx=15, pady=(15, 10), anchor="w")

        cheese_list = sorted(
            sku for sku in self.inventory
            if sku.startswith("CH-") and (
                self.inventory[sku].get("qty", 0)
                if isinstance(self.inventory[sku], dict) else 0) > 0)

        current_split = self._fp_cexec_splits.get(curation, {})
        current_primary = self.cex_ec.get(curation, "")

        frame = ttk.Frame(dlg)
        frame.pack(fill="x", padx=15, pady=5)

        ttk.Label(frame, text="Cheese A:").grid(
            row=0, column=0, sticky="w", padx=(0, 10), pady=5)
        cheese_a_var = tk.StringVar(
            value=list(current_split.keys())[0] if current_split
            else current_primary)
        ttk.Combobox(frame, textvariable=cheese_a_var,
                     values=cheese_list, width=14).grid(
            row=0, column=1, pady=5)

        ttk.Label(frame, text="Cheese B:").grid(
            row=1, column=0, sticky="w", padx=(0, 10), pady=5)
        cheese_b_var = tk.StringVar(
            value=list(current_split.keys())[1] if len(current_split) > 1
            else "")
        ttk.Combobox(frame, textvariable=cheese_b_var,
                     values=cheese_list, width=14).grid(
            row=1, column=1, pady=5)

        ttk.Label(frame, text="Ratio A (%):").grid(
            row=2, column=0, sticky="w", padx=(0, 10), pady=5)
        ratio_var = tk.StringVar(value="64")
        if current_split:
            first_val = list(current_split.values())[0]
            ratio_var.set(str(int(first_val * 100)))

        ratio_scale = ttk.Scale(frame, from_=10, to=90,
                                variable=ratio_var, orient="horizontal")
        ratio_scale.grid(row=2, column=1, sticky="ew", pady=5)

        ratio_label = ttk.Label(frame, text="64% / 36%")
        ratio_label.grid(row=3, column=1, sticky="w")

        def _update_label(*args):
            try:
                pct = int(float(ratio_var.get()))
            except (ValueError, TypeError):
                pct = 50
            pct = max(10, min(90, pct))
            ratio_label.configure(text=f"{pct}% / {100 - pct}%")

        ratio_var.trace_add("write", _update_label)
        _update_label()

        def _apply_split():
            a = cheese_a_var.get().strip()
            b = cheese_b_var.get().strip()
            if not a:
                return
            try:
                pct = int(float(ratio_var.get()))
            except (ValueError, TypeError):
                pct = 50
            pct = max(10, min(90, pct))

            if b and b != a:
                self._fp_cexec_splits[curation] = {
                    a: pct / 100.0,
                    b: (100 - pct) / 100.0,
                }
                self.cex_ec[curation] = a  # Primary
            else:
                # Single cheese, no split
                self._fp_cexec_splits.pop(curation, None)
                self.cex_ec[curation] = a

            self._fp_refresh_assignments()
            if self.fp_results:
                self._fp_calculate()
            self._fp_set_mascot("happy", f"Split set for {curation}!")
            dlg.destroy()

        def _clear_split():
            self._fp_cexec_splits.pop(curation, None)
            self._fp_refresh_assignments()
            dlg.destroy()

        btn_frame = ttk.Frame(dlg)
        btn_frame.pack(fill="x", padx=15, pady=(10, 15))
        tk.Button(btn_frame, text="Cancel", command=dlg.destroy,
                  bg=_BG3, fg=_FG, relief="flat", padx=10, pady=4
                  ).pack(side="right", padx=(5, 0))
        tk.Button(btn_frame, text="Clear Split", command=_clear_split,
                  bg=_BG3, fg=_FG, relief="flat", padx=10, pady=4
                  ).pack(side="right", padx=(5, 0))
        tk.Button(btn_frame, text="Apply", command=_apply_split,
                  bg=_GREEN, fg="white", relief="flat", padx=10, pady=4
                  ).pack(side="right")

    # ── Next-week calculation ──

    def _fp_calculate_next_week(self):
        """Calculate next Saturday NET using carry-forward + estimated demand."""
        if not self.fp_results:
            return

        tree = self._fp_next_tree
        tree.delete(*tree.get_children())

        for r in self.fp_results:
            if not r["sku"].startswith("CH-"):
                continue

            carry_fwd = max(0, r["net"])
            # Estimate next week demand = this week demand
            next_demand = r["total_demand"]
            next_net = carry_fwd - next_demand

            if next_demand == 0:
                status = "NO DEMAND"
                tag = "ok"
            elif next_net < 0:
                status = "PLAN PO"
                tag = "shortage"
            elif next_net < next_demand * 0.3:
                status = "TIGHT"
                tag = "tight"
            else:
                status = "OK"
                tag = "ok"

            # Only show items with demand or risk
            if next_demand > 0 or carry_fwd > 0:
                tree.insert("", "end", values=(
                    r["sku"], carry_fwd, next_demand,
                    f"{next_net:+d}", status), tags=(tag,))

    # ── Level-complete celebration ──

    def _fp_check_level_complete(self, shortages):
        """Show/hide the level-complete banner."""
        if shortages == 0 and self.fp_results:
            self._fp_banner_frame.pack(fill="x", padx=5, pady=(0, 5),
                                        before=self._fp_week_nb)
            # Celebration mascot
            self._fp_set_mascot("happy", "LEVEL CLEAR!")
            self._fp_pop_icon("ALL CLEAR", "#80ff80")
            # Flash the banner
            self._fp_flash_banner(0)
        else:
            self._fp_banner_frame.pack_forget()

    def _fp_flash_banner(self, count):
        """Flash the level-complete banner 3 times."""
        if count >= 6:
            self._fp_banner_frame.configure(bg="#1a4d1a")
            self._fp_banner_label.configure(bg="#1a4d1a", fg="#80ff80")
            return
        if count % 2 == 0:
            self._fp_banner_frame.configure(bg="#2a6d2a")
            self._fp_banner_label.configure(bg="#2a6d2a", fg="#ffffff")
        else:
            self._fp_banner_frame.configure(bg="#1a4d1a")
            self._fp_banner_label.configure(bg="#1a4d1a", fg="#80ff80")
        self.root.after(300, lambda: self._fp_flash_banner(count + 1))

    # ── Drag-and-drop cheese assignment ──

    def _fp_drag_start(self, event):
        """Start dragging a cheese SKU from the NET tree."""
        tree = self._fp_net_tree
        item = tree.identify_row(event.y)
        if not item:
            self._fp_drag_data["sku"] = None
            return
        vals = tree.item(item, "values")
        sku = vals[0] if vals else ""
        if not sku.startswith("CH-"):
            self._fp_drag_data["sku"] = None
            return
        self._fp_drag_data["sku"] = sku
        self._fp_drag_data["x"] = event.x_root
        self._fp_drag_data["y"] = event.y_root

    def _fp_drag_motion(self, event):
        """Show floating label while dragging."""
        if not self._fp_drag_data.get("sku"):
            return
        # Only show drag label after moving 5+ pixels
        dx = abs(event.x_root - self._fp_drag_data["x"])
        dy = abs(event.y_root - self._fp_drag_data["y"])
        if dx < 5 and dy < 5 and not self._fp_drag_label:
            return

        if not self._fp_drag_label:
            self._fp_drag_label = tk.Toplevel(self.root)
            self._fp_drag_label.overrideredirect(True)
            self._fp_drag_label.attributes("-topmost", True)
            self._fp_drag_label.attributes("-alpha", 0.85)
            lbl = tk.Label(self._fp_drag_label,
                           text=self._fp_drag_data["sku"],
                           font=("Segoe UI", 10, "bold"),
                           bg="#f0c040", fg="#333",
                           padx=8, pady=4, relief="raised", bd=1)
            lbl.pack()

        self._fp_drag_label.geometry(
            f"+{event.x_root + 12}+{event.y_root - 10}")

    def _fp_drag_drop(self, event):
        """Drop cheese onto assignment tree to assign it."""
        sku = self._fp_drag_data.get("sku")
        # Clean up drag label
        if self._fp_drag_label:
            self._fp_drag_label.destroy()
            self._fp_drag_label = None

        if not sku:
            return

        self._fp_drag_data["sku"] = None

        # Check if dropped onto the assignment tree
        assign_tree = self._fp_assign_tree
        try:
            # Get assignment tree's screen coords
            tx = assign_tree.winfo_rootx()
            ty = assign_tree.winfo_rooty()
            tw = assign_tree.winfo_width()
            th = assign_tree.winfo_height()

            if not (tx <= event.x_root <= tx + tw and
                    ty <= event.y_root <= ty + th):
                return  # Not dropped on assignment tree

            # Find which row and column
            rel_x = event.x_root - tx
            rel_y = event.y_root - ty
            row = assign_tree.identify_row(rel_y)
            col = assign_tree.identify_column(rel_x)
            if not row or not col:
                return

            col_idx = int(col.replace("#", "")) - 1
            curation = row  # iid is the curation name

            # Determine slot type from column
            if col_idx in (1, 2):
                slot = "prcjam"
            elif col_idx in (3, 4, 5):
                slot = "cexec"
            else:
                return

            # Check +/-2 constraint before assigning
            old_prcjam = ""
            old_cexec = ""
            prcjam_info = self.pr_cjam.get(curation, {})
            if isinstance(prcjam_info, dict):
                old_prcjam = prcjam_info.get("cheese", "")
            old_cexec = self.cex_ec.get(curation, "")

            test_prcjam = sku if slot == "prcjam" else old_prcjam
            test_cexec = sku if slot == "cexec" else old_cexec
            constraint = self._fp_check_constraint(
                curation, test_prcjam, test_cexec)

            if constraint != "OK":
                self._fp_set_mascot("alert",
                                    f"+/-2 violation! {sku} blocked for {curation}")
                self._fp_flash_cell(curation, "violation")
                return

            # Apply the assignment
            if slot == "prcjam":
                if isinstance(self.pr_cjam.get(curation), dict):
                    self.pr_cjam[curation]["cheese"] = sku
                else:
                    self.pr_cjam[curation] = {"cheese": sku, "jam": ""}
            else:
                self.cex_ec[curation] = sku

            # Visual celebration
            self._fp_flash_cell(curation, "success")
            self._fp_set_mascot("happy",
                                f"{sku} -> {curation} {slot.upper()}!")
            self._fp_pop_icon("+1", "#80ff80")

            # Score
            self._fp_update_score(1)

            # Refresh
            self._fp_refresh_assignments()
            if self.fp_results:
                self._fp_calculate()

        except Exception:
            pass

    def _fp_flash_cell(self, curation, flash_type):
        """Flash a row in the assignment tree for visual feedback."""
        tree = self._fp_assign_tree
        if curation not in [tree.item(iid, "values")[0]
                            for iid in tree.get_children()]:
            # Try using iid directly
            try:
                tree.item(curation)
            except tk.TclError:
                return

        if flash_type == "success":
            colors = [("#2a6d2a", "#80ff80"), (None, None)] * 2
        else:
            colors = [("#6d2a2a", "#ff8080"), (None, None)] * 2

        def do_flash(step=0):
            if step >= len(colors):
                # Restore original tag
                self._fp_refresh_assignments()
                return
            bg, fg = colors[step]
            if bg:
                tree.tag_configure("flash", background=bg, foreground=fg)
                tree.item(curation, tags=("flash",))
            else:
                tree.item(curation, tags=("ok",))
            self.root.after(150, lambda: do_flash(step + 1))

        do_flash()

    def _fp_update_score(self, delta, reason=""):
        """Update session score with optional mascot message."""
        self._fp_session_score += delta
        self._fp_score_var.set(
            f"Score: {self._fp_session_score}" if self._fp_session_score > 0
            else "")
        if reason:
            self._fp_set_mascot("happy", reason)

    def _fp_pop_icon(self, text, color="#f0c040"):
        """Show a pop-up text that floats up and fades on the mascot canvas."""
        c = self._fp_mascot_canvas
        tid = c.create_text(60, 60, text=text,
                            font=("Segoe UI", 12, "bold"), fill=color)

        def _rise(step=0):
            if step >= 10:
                c.delete(tid)
                return
            c.move(tid, 0, -3)
            self.root.after(60, lambda: _rise(step + 1))

        _rise()

    # ── Column sorting ──

    def _fp_sort_by(self, col):
        """Sort NET treeview by column."""
        if col == self._fp_sort_col:
            self._fp_sort_rev = not self._fp_sort_rev
        else:
            self._fp_sort_col = col
            self._fp_sort_rev = False

        # Numeric columns
        numeric = {"available", "direct", "prcjam", "cexec", "exec",
                   "total_demand", "net"}

        items = []
        for iid in self._fp_net_tree.get_children():
            vals = self._fp_net_tree.item(iid, "values")
            tags = self._fp_net_tree.item(iid, "tags")
            items.append((vals, tags))

        col_names = ["sku", "available", "direct", "prcjam", "cexec",
                     "exec", "total_demand", "net", "headroom", "status"]
        ci = col_names.index(col) if col in col_names else 0

        def sort_key(item):
            val = item[0][ci]
            if col in numeric:
                try:
                    return float(str(val).replace("+", "").replace(",", ""))
                except (ValueError, TypeError):
                    return 0
            return str(val).lower()

        items.sort(key=sort_key, reverse=self._fp_sort_rev)

        self._fp_net_tree.delete(*self._fp_net_tree.get_children())
        for vals, tags in items:
            self._fp_net_tree.insert("", "end", values=vals, tags=tags)

    # ── Auto-assign engine ──

    def _fp_auto_assign(self):
        """Auto-assign PR-CJAM and CEX-EC for all curations based on
        headroom and +/-2 constraints."""
        if not self.fp_results and not self.inventory:
            messagebox.showinfo("No Data",
                                "Load inventory or run Calculate first.",
                                parent=self.root)
            return

        self._fp_set_mascot("thinking", "Auto-assigning...")

        # Build headroom map from inventory
        headroom = {}
        for sku, data in self.inventory.items():
            if not sku.startswith("CH-"):
                continue
            qty = data.get("qty", 0) if isinstance(data, dict) else 0
            headroom[sku] = qty

        # If we have results, use NET for better accuracy
        if self.fp_results:
            for r in self.fp_results:
                if r["sku"].startswith("CH-"):
                    headroom[r["sku"]] = r["available"]

        # Track consumed headroom during assignment
        consumed = defaultdict(int)
        new_prcjam = {}
        new_cexec = {}
        changes = []

        # Assign PR-CJAM first (1 per box, higher volume)
        all_curations = list(CURATION_ORDER) + ["NMS", "BYO", "SS"]
        used_cheeses = set()  # PR-CJAM cheese must be unique across curations

        for cur in all_curations:
            # Estimate demand for this curation from resolved data
            est_demand = 0
            for month_data in self.recharge_queued_resolved.values():
                est_demand += int(month_data.get("pr_cjam", {}).get(cur, 0))
            if est_demand == 0:
                est_demand = 50  # default estimate

            # Get candidates sorted by remaining headroom
            candidates = []
            for sku, qty in headroom.items():
                remaining = qty - consumed.get(sku, 0)
                if remaining < est_demand:
                    continue
                if sku in used_cheeses:
                    continue
                # Check +/-2
                constraint = self._fp_check_constraint(cur, sku, "")
                if constraint != "OK":
                    continue
                candidates.append((sku, remaining))

            candidates.sort(key=lambda x: -x[1])

            if candidates:
                best = candidates[0][0]
                old = (self.pr_cjam.get(cur, {}).get("cheese", "")
                       if isinstance(self.pr_cjam.get(cur), dict) else "")
                new_prcjam[cur] = {"cheese": best, "jam": ""}
                consumed[best] += est_demand
                used_cheeses.add(best)
                if best != old:
                    changes.append(f"PR-CJAM-{cur}: {old} -> {best}")
            else:
                # Keep existing
                new_prcjam[cur] = self.pr_cjam.get(
                    cur, {"cheese": "", "jam": ""})

        # Assign CEX-EC (can reuse cheeses across curations)
        for cur in all_curations:
            est_demand = 0
            for month_data in self.recharge_queued_resolved.values():
                est_demand += int(month_data.get("cex_ec", {}).get(cur, 0))
            if est_demand == 0:
                est_demand = 20

            prcjam_cheese = new_prcjam.get(cur, {}).get("cheese", "")

            candidates = []
            for sku, qty in headroom.items():
                remaining = qty - consumed.get(sku, 0)
                if remaining < est_demand:
                    continue
                if sku == prcjam_cheese:
                    continue  # Don't use same cheese as PR-CJAM
                constraint = self._fp_check_constraint(
                    cur, prcjam_cheese, sku)
                if constraint != "OK":
                    continue
                candidates.append((sku, remaining))

            candidates.sort(key=lambda x: -x[1])

            if candidates:
                best = candidates[0][0]
                old = self.cex_ec.get(cur, "")
                new_cexec[cur] = best
                consumed[best] += est_demand
                if best != old:
                    changes.append(f"CEX-EC-{cur}: {old} -> {best}")
            else:
                new_cexec[cur] = self.cex_ec.get(cur, "")

        if not changes:
            self._fp_set_mascot("happy", "Assignments already optimal!")
            return

        # Show preview before applying
        dlg = tk.Toplevel(self.root)
        dlg.title("Auto-Assign Preview")
        dlg.configure(bg=_BG)
        dlg.transient(self.root)
        dlg.grab_set()
        dlg.geometry("500x450")

        ttk.Label(dlg, text=f"Proposed Changes ({len(changes)})",
                  font=("Segoe UI", 11, "bold")).pack(
                      padx=15, pady=(15, 5), anchor="w")

        text = tk.Text(dlg, bg=_BG2, fg=_FG, font=("Consolas", 10),
                       wrap="word", relief="flat", padx=10, pady=10)
        text.pack(fill="both", expand=True, padx=10, pady=5)
        for ch in changes:
            text.insert("end", f"  {ch}\n")
        text.configure(state="disabled")

        def _apply():
            self.pr_cjam = new_prcjam
            self.cex_ec = new_cexec
            self._fp_refresh_assignments()
            if self.fp_results:
                self._fp_calculate()
            self._fp_update_score(
                len(changes),
                f"{len(changes)} assignments updated!")
            self.root.after(3000, lambda: self._fp_set_mascot(
                "idle", "Ready to plan!"))
            dlg.destroy()

        btn_frame = ttk.Frame(dlg)
        btn_frame.pack(fill="x", padx=15, pady=(0, 15))
        tk.Button(btn_frame, text="Cancel", command=dlg.destroy,
                  bg=_BG3, fg=_FG, relief="flat", padx=10, pady=4
                  ).pack(side="right", padx=(5, 0))
        tk.Button(btn_frame, text="Apply All", command=_apply,
                  bg=_GREEN, fg="white", relief="flat", padx=10, pady=4
                  ).pack(side="right")

    # ── What-If simulation ──

    def _fp_what_if(self):
        """Open a what-if simulation dialog to test assignment changes."""
        if not self.fp_results:
            messagebox.showinfo("No Data",
                                "Run Calculate first.",
                                parent=self.root)
            return

        dlg = tk.Toplevel(self.root)
        dlg.title("What-If Simulation")
        dlg.configure(bg=_BG)
        dlg.transient(self.root)
        dlg.geometry("750x550")

        ttk.Label(dlg, text="What-If: Test assignment changes",
                  font=("Segoe UI", 11, "bold")).pack(
                      padx=15, pady=(15, 5), anchor="w")
        ttk.Label(dlg, text="Change assignments below, then click "
                  "Simulate to see the impact.",
                  foreground=_FG2, font=("Segoe UI", 9)).pack(
                      padx=15, anchor="w")

        # Editable assignment grid
        edit_frame = ttk.Frame(dlg)
        edit_frame.pack(fill="x", padx=15, pady=10)

        # Headers
        headers = ["Curation", "PR-CJAM Cheese", "CEX-EC Cheese"]
        for i, h in enumerate(headers):
            ttk.Label(edit_frame, text=h, font=("Segoe UI", 9, "bold")
                      ).grid(row=0, column=i, padx=5, pady=2, sticky="w")

        # Build cheese list for dropdowns
        cheese_list = sorted(
            sku for sku in self.inventory
            if sku.startswith("CH-") and (
                self.inventory[sku].get("qty", 0)
                if isinstance(self.inventory[sku], dict) else 0) > 0)

        wi_vars = {}
        all_curations = list(CURATION_ORDER) + ["NMS", "BYO", "SS"]
        for i, cur in enumerate(all_curations):
            ttk.Label(edit_frame, text=cur).grid(
                row=i + 1, column=0, padx=5, pady=2, sticky="w")

            prcjam_info = self.pr_cjam.get(cur, {})
            current_pr = (prcjam_info.get("cheese", "")
                          if isinstance(prcjam_info, dict) else "")
            pr_var = tk.StringVar(value=current_pr)
            ttk.Combobox(edit_frame, textvariable=pr_var,
                         values=cheese_list, width=14).grid(
                row=i + 1, column=1, padx=5, pady=2)

            current_ec = self.cex_ec.get(cur, "")
            ec_var = tk.StringVar(value=current_ec)
            ttk.Combobox(edit_frame, textvariable=ec_var,
                         values=cheese_list, width=14).grid(
                row=i + 1, column=2, padx=5, pady=2)

            wi_vars[cur] = (pr_var, ec_var)

        # Results area
        result_frame = ttk.LabelFrame(dlg, text="Simulation Results")
        result_frame.pack(fill="both", expand=True, padx=15, pady=5)

        wi_cols = ("sku", "current_net", "sim_net", "delta", "status")
        wi_tree = ttk.Treeview(
            result_frame, columns=wi_cols, show="headings",
            height=10, selectmode="browse")
        wi_tree.heading("sku", text="SKU")
        wi_tree.heading("current_net", text="Current NET")
        wi_tree.heading("sim_net", text="Sim NET")
        wi_tree.heading("delta", text="Delta")
        wi_tree.heading("status", text="Status")
        wi_tree.column("sku", width=100)
        wi_tree.column("current_net", width=90, anchor="e")
        wi_tree.column("sim_net", width=90, anchor="e")
        wi_tree.column("delta", width=80, anchor="e")
        wi_tree.column("status", width=80)

        wi_tree.tag_configure("better", foreground="#80cc80")
        wi_tree.tag_configure("worse", foreground="#ff6666")
        wi_tree.tag_configure("same", foreground=_FG2)

        wi_sb = ttk.Scrollbar(result_frame, orient="vertical",
                              command=wi_tree.yview)
        wi_tree.configure(yscrollcommand=wi_sb.set)
        wi_tree.pack(side="left", fill="both", expand=True)
        wi_sb.pack(side="right", fill="y")

        def _simulate():
            wi_tree.delete(*wi_tree.get_children())

            # Build simulated demand with new assignments
            sim_prcjam = defaultdict(int)
            sim_cexec = defaultdict(int)

            for cur, (pr_var, ec_var) in wi_vars.items():
                pr_cheese = pr_var.get().strip()
                ec_cheese = ec_var.get().strip()

                # Estimate demand for this curation
                pr_demand = 0
                ec_demand = 0
                for month_data in self.recharge_queued_resolved.values():
                    pr_demand += int(
                        month_data.get("pr_cjam", {}).get(cur, 0))
                    ec_demand += int(
                        month_data.get("cex_ec", {}).get(cur, 0))

                if pr_cheese:
                    sim_prcjam[pr_cheese] += pr_demand
                if ec_cheese:
                    sim_cexec[ec_cheese] += ec_demand

            # Compare current vs simulated
            changes = []
            all_skus = set()
            for r in self.fp_results:
                all_skus.add(r["sku"])
            all_skus.update(sim_prcjam.keys())
            all_skus.update(sim_cexec.keys())

            for sku in sorted(all_skus):
                if not sku.startswith("CH-"):
                    continue
                current = next(
                    (r for r in self.fp_results if r["sku"] == sku), None)
                if not current:
                    continue

                current_net = current["net"]
                # Recalculate with sim assignments
                avail = current["available"]
                # Remove old PR-CJAM/CEX-EC, add new
                new_total = (current["direct"] +
                             sim_prcjam.get(sku, 0) +
                             sim_cexec.get(sku, 0) +
                             current["exec"])
                sim_net = avail - new_total
                delta = sim_net - current_net

                if delta > 0:
                    tag = "better"
                elif delta < 0:
                    tag = "worse"
                else:
                    tag = "same"

                if delta != 0 or current_net < 0 or sim_net < 0:
                    status = "SHORTAGE" if sim_net < 0 else "OK"
                    changes.append((sku, current_net, sim_net, delta,
                                    status, tag))

            changes.sort(key=lambda x: x[3])  # worst delta first

            for sku, c_net, s_net, delta, status, tag in changes:
                wi_tree.insert("", "end", values=(
                    sku, f"{c_net:+d}", f"{s_net:+d}",
                    f"{delta:+d}", status), tags=(tag,))

        def _apply_sim():
            for cur, (pr_var, ec_var) in wi_vars.items():
                pr_cheese = pr_var.get().strip()
                ec_cheese = ec_var.get().strip()
                if pr_cheese:
                    self.pr_cjam[cur] = {"cheese": pr_cheese, "jam": ""}
                if ec_cheese:
                    self.cex_ec[cur] = ec_cheese

            self._fp_refresh_assignments()
            self._fp_calculate()
            self._fp_set_mascot("happy", "What-if applied!")
            dlg.destroy()

        btn_frame = ttk.Frame(dlg)
        btn_frame.pack(fill="x", padx=15, pady=(0, 15))
        tk.Button(btn_frame, text="Close", command=dlg.destroy,
                  bg=_BG3, fg=_FG, relief="flat", padx=10, pady=4
                  ).pack(side="right", padx=(5, 0))
        tk.Button(btn_frame, text="Apply Changes", command=_apply_sim,
                  bg=_GREEN, fg="white", relief="flat", padx=10, pady=4
                  ).pack(side="right", padx=(5, 0))
        tk.Button(btn_frame, text="Simulate", command=_simulate,
                  bg=_ACC, fg="white", relief="flat", padx=10, pady=4
                  ).pack(side="right")

    # ── Data pull helpers ──

    def _fp_pull_recharge(self):
        """Pull queued charges from Recharge API for fulfillment planning."""
        token = self.recharge_token_var.get().strip()
        if not token:
            messagebox.showerror("Error",
                                 "Set Recharge API token in Settings first.",
                                 parent=self.root)
            return

        self._fp_set_mascot("loading", "Pulling Recharge charges...")
        self._fp_status_var.set("Pulling Recharge queued charges...")

        def _worker():
            try:
                client = RechargeClient(token)
                charges = client.get_queued_charges(
                    progress_cb=lambda n: self.root.after(
                        0, lambda: self._fp_status_var.set(
                            f"Pulling... {n} charges")))
                queued_by_month = client.aggregate_charges_by_month(charges)
                queued_resolved = resolve_queued_charges(charges)

                self.root.after(0, lambda: self._fp_on_recharge_done(
                    queued_by_month, queued_resolved, len(charges)))
            except Exception as e:
                self.root.after(0, lambda: self._fp_on_pull_error(
                    "Recharge", str(e)))

        threading.Thread(target=_worker, daemon=True).start()

    def _fp_on_recharge_done(self, queued_by_month, queued_resolved, count):
        self.recharge_queued = queued_by_month
        self.recharge_queued_resolved = queued_resolved
        self._fp_set_mascot("happy", f"{count} charges loaded!")
        self._fp_status_var.set(
            f"Recharge: {count} charges across "
            f"{len(queued_by_month)} months")
        self.root.after(3000, lambda: self._fp_set_mascot(
            "idle", "Ready to plan!"))

    def _fp_pull_shopify(self):
        """Pull recent Shopify orders for fulfillment planning."""
        store = self.shopify_store_var.get().strip()
        token = self.shopify_token_var.get().strip()
        if not store or not token:
            messagebox.showerror("Error",
                                 "Set Shopify store URL and access token "
                                 "in Settings first.",
                                 parent=self.root)
            return

        self._fp_set_mascot("loading", "Pulling Shopify orders...")
        self._fp_status_var.set("Pulling Shopify orders...")

        def _worker():
            try:
                client = ShopifyClient(store, token)
                cutoff = (datetime.datetime.now() -
                          datetime.timedelta(days=3)).isoformat()
                orders = client.get_orders(
                    created_at_min=cutoff,
                    progress_cb=lambda n: self.root.after(
                        0, lambda: self._fp_status_var.set(
                            f"Pulling... {n} orders")))

                self.root.after(0, lambda: self._fp_on_shopify_done(
                    orders))
            except Exception as e:
                self.root.after(0, lambda: self._fp_on_pull_error(
                    "Shopify", str(e)))

        threading.Thread(target=_worker, daemon=True).start()

    def _fp_on_shopify_done(self, orders):
        # Store raw orders for the planner to process
        self._fp_shopify_orders = orders
        self._fp_set_mascot("happy", f"{len(orders)} orders loaded!")
        self._fp_status_var.set(f"Shopify: {len(orders)} orders")
        self.root.after(3000, lambda: self._fp_set_mascot(
            "idle", "Ready to plan!"))

    def _fp_on_pull_error(self, source, error_msg):
        self._fp_set_mascot("alert", f"{source} error!")
        self._fp_status_var.set(f"{source} pull failed")
        messagebox.showerror(f"{source} Error", error_msg,
                             parent=self.root)
        self.root.after(3000, lambda: self._fp_set_mascot(
            "idle", "Ready to plan!"))

    # ── Core calculation ──

    def _fp_calculate(self):
        """Calculate Saturday NET using inventory + all demand sources."""
        self._fp_set_mascot("loading", "Nom nom... crunching numbers...")

        # 1. Build inventory snapshot
        inv = {}
        for sku, data in self.inventory.items():
            if isinstance(data, dict):
                inv[sku] = data.get("qty", 0)
            else:
                inv[sku] = int(data)

        # Add wheel supply
        for wsku, wdata in self.wheel_inventory.items():
            if isinstance(wdata, dict):
                weight = float(wdata.get("weight_lbs", 0))
                count = int(wdata.get("count", 0))
                target = wdata.get("target_sku", "")
                if target and weight > 0 and count > 0:
                    wheel_supply = int(weight * count * WHEEL_TO_SLICE_FACTOR)
                    inv[target] = inv.get(target, 0) + wheel_supply

        # Add open POs
        for po in self.open_pos:
            sku = po.get("sku", "")
            if sku and po.get("status", "Open") == "Open":
                try:
                    inv[sku] = inv.get(sku, 0) + int(float(po.get("qty", 0)))
                except (ValueError, TypeError):
                    pass

        self.fp_inventory = dict(inv)

        # 2. Build demand from resolved queued charges + Shopify
        demand_direct = defaultdict(int)    # CH-* direct picks
        demand_prcjam = defaultdict(int)    # PR-CJAM resolved
        demand_cexec = defaultdict(int)     # CEX-EC resolved
        demand_exec = defaultdict(int)      # EX-EC resolved

        target_date = self._fp_date_var.get().strip()

        # From Recharge queued resolved data
        for month, data in self.recharge_queued_resolved.items():
            # PR-CJAM
            for suffix, count in data.get("pr_cjam", {}).items():
                prcjam_info = self.pr_cjam.get(suffix, {})
                cheese = (prcjam_info.get("cheese", "")
                          if isinstance(prcjam_info, dict)
                          else str(prcjam_info))
                if cheese:
                    demand_prcjam[cheese] += int(count)

            # CEX-EC (with split support)
            for suffix, count in data.get("cex_ec", {}).items():
                splits = self._fp_cexec_splits.get(suffix, {})
                if splits:
                    total = int(count)
                    remaining = total
                    split_items = list(splits.items())
                    for i, (split_sku, ratio) in enumerate(split_items):
                        if i == len(split_items) - 1:
                            demand_cexec[split_sku] += remaining
                        else:
                            portion = int(total * ratio)
                            demand_cexec[split_sku] += portion
                            remaining -= portion
                else:
                    cheese = self.cex_ec.get(suffix, "")
                    if cheese:
                        demand_cexec[cheese] += int(count)

        # From Recharge queued raw SKUs (direct CH-* items)
        for month, skus in self.recharge_queued.items():
            for sku, qty in skus.items():
                if sku.startswith("CH-"):
                    demand_direct[sku] += int(qty)
                elif sku.startswith("EX-EC"):
                    # EX-EC resolves same as CEX-EC
                    if "-" in sku[5:]:
                        suffix = sku.split("-", 2)[2] if sku.count("-") >= 2 else ""
                        cheese = self.cex_ec.get(suffix, "")
                        if cheese:
                            demand_exec[cheese] += int(qty)

        # From Shopify API demand
        for sku, qty in self.shopify_api_demand.items():
            if sku.startswith("CH-"):
                demand_direct[sku] += int(qty)

        # From Shopify raw orders (pulled via fulfillment tab)
        for order in getattr(self, '_fp_shopify_orders', []):
            for item in order.get("line_items", []):
                sku = (item.get("sku") or "").strip()
                if not sku:
                    continue
                qty = int(float(item.get("quantity", 1)))
                upper = sku.upper()
                if upper.startswith("CH-"):
                    demand_direct[sku] += qty
                elif upper.startswith("PR-CJAM-"):
                    suffix = upper.split("PR-CJAM-", 1)[1]
                    prcjam_info = self.pr_cjam.get(suffix, {})
                    cheese = (prcjam_info.get("cheese", "")
                              if isinstance(prcjam_info, dict) else "")
                    if cheese:
                        demand_prcjam[cheese] += qty
                elif upper.startswith("CEX-EC-"):
                    suffix = upper.split("CEX-EC-", 1)[1]
                    splits = self._fp_cexec_splits.get(suffix, {})
                    if splits:
                        remaining = qty
                        split_items = list(splits.items())
                        for si, (split_sku, ratio) in enumerate(split_items):
                            if si == len(split_items) - 1:
                                demand_cexec[split_sku] += remaining
                            else:
                                portion = int(qty * ratio)
                                demand_cexec[split_sku] += portion
                                remaining -= portion
                    else:
                        cheese = self.cex_ec.get(suffix, "")
                        if cheese:
                            demand_cexec[cheese] += qty
                elif upper.startswith("EX-EC-"):
                    suffix = upper.split("EX-EC-", 1)[1]
                    cheese = self.cex_ec.get(suffix, "")
                    if cheese:
                        demand_exec[cheese] += qty
                # Bare CEX-EC / EX-EC / CEX-EM / EX-EM / EX-EA — skip

        # From imported CSV demand
        for sku, qty in getattr(self, '_fp_csv_demand', {}).items():
            if sku.startswith("CH-"):
                demand_direct[sku] += qty

        # From manual demand
        for sku, qty in self.manual_demand.items():
            if sku.startswith("CH-"):
                demand_direct[sku] += int(qty)

        # 3. Build result rows for CH-* SKUs
        all_ch = set()
        all_ch.update(k for k in inv if k.startswith("CH-"))
        all_ch.update(demand_direct.keys())
        all_ch.update(demand_prcjam.keys())
        all_ch.update(demand_cexec.keys())
        all_ch.update(demand_exec.keys())

        results = []
        shortages = 0
        for sku in sorted(all_ch):
            avail = inv.get(sku, 0)
            d_direct = demand_direct.get(sku, 0)
            d_prcjam = demand_prcjam.get(sku, 0)
            d_cexec = demand_cexec.get(sku, 0)
            d_exec = demand_exec.get(sku, 0)
            total = d_direct + d_prcjam + d_cexec + d_exec
            net = avail - total

            if total == 0:
                status = "NO DEMAND"
                tag = "no_demand"
            elif net < 0:
                status = "SHORTAGE"
                tag = "shortage"
                shortages += 1
            elif net < total * 0.2:
                status = "TIGHT"
                tag = "tight"
            elif net > avail * 0.5 and avail > 200:
                status = "SURPLUS"
                tag = "surplus"
            else:
                status = "OK"
                tag = "ok"

            results.append({
                "sku": sku, "available": avail,
                "direct": d_direct, "prcjam": d_prcjam,
                "cexec": d_cexec, "exec": d_exec,
                "total_demand": total, "net": net,
                "status": status, "tag": tag,
            })

        # Sort: shortage first, then tight, then by net ascending
        status_order = {"SHORTAGE": 0, "TIGHT": 1, "OK": 2,
                        "SURPLUS": 3, "NO DEMAND": 4}
        results.sort(key=lambda r: (status_order.get(r["status"], 9),
                                     r["net"]))
        self.fp_results = results

        # 4. Populate NET treeview
        tree = self._fp_net_tree
        tree.delete(*tree.get_children())

        filt = self._fp_filter_var.get() if hasattr(
            self, '_fp_filter_var') else "CH-*"

        for r in results:
            if r["status"] == "NO DEMAND" and r["available"] == 0:
                continue
            if filt == "CH-*" and not r["sku"].startswith("CH-"):
                continue
            if filt == "Shortages" and r["status"] != "SHORTAGE":
                continue
            if filt == "Tight" and r["status"] not in ("SHORTAGE", "TIGHT"):
                continue
            if filt == "Surplus" and r["status"] != "SURPLUS":
                continue

            headroom = self._fp_headroom_bar(r["net"], r["total_demand"])
            tree.insert("", "end", values=(
                r["sku"], r["available"],
                r["direct"], r["prcjam"],
                r["cexec"], r["exec"],
                r["total_demand"], f"{r['net']:+d}",
                headroom, r["status"],
            ), tags=(r["tag"],))

        # 5. Update assignment demand counts
        self._fp_update_assign_demands(demand_prcjam, demand_cexec)

        # 6. Summary + mascot
        demand_items = [r for r in results if r["total_demand"] > 0]
        total_units = sum(r["total_demand"] for r in results)

        self._fp_summary_var.set(
            f"{len(demand_items)} SKUs with demand | "
            f"{total_units:,} total units | "
            f"{shortages} shortages")

        if shortages > 0:
            self._fp_set_mascot("worried",
                                f"{shortages} shortage{'s' if shortages > 1 else ''}! "
                                f"Check assignments.")
        elif shortages == 0 and demand_items:
            self._fp_set_mascot("happy", "All clear! No shortages.")
        else:
            self._fp_set_mascot("idle", "Ready to plan!")

        self._fp_status_var.set(
            f"Calculated: {len(demand_items)} SKUs, "
            f"{shortages} shortages | "
            f"{datetime.datetime.now().strftime('%H:%M:%S')}")

        # 7. Update shelf life panel
        self._fp_refresh_shelf_life()

        # 8. Next week projection
        self._fp_calculate_next_week()

        # 9. Level-complete check
        self._fp_check_level_complete(shortages)

    def _fp_update_assign_demands(self, demand_prcjam, demand_cexec):
        """Update PR-CJAM/CEX-EC demand counts in assignment tree."""
        tree = self._fp_assign_tree
        for iid in tree.get_children():
            vals = list(tree.item(iid, "values"))
            cur = vals[0]
            prcjam_cheese = vals[1]
            cexec_cheese = vals[3]

            # PR-CJAM demand for this cheese
            prcjam_qty = 0
            if prcjam_cheese:
                resolved = self.recharge_queued_resolved or {}
                for month_data in resolved.values():
                    prcjam_qty += int(month_data.get("pr_cjam", {}).get(cur, 0))
            vals[2] = str(prcjam_qty) if prcjam_qty else ""

            # CEX-EC demand
            cexec_qty = 0
            if cexec_cheese:
                resolved = self.recharge_queued_resolved or {}
                for month_data in resolved.values():
                    cexec_qty += int(month_data.get("cex_ec", {}).get(cur, 0))
            vals[4] = str(cexec_qty) if cexec_qty else ""

            tree.item(iid, values=vals)

    # ── Export ──

    def _fp_export_csv(self):
        """Export Saturday NET to CSV."""
        if not self.fp_results:
            messagebox.showinfo("No Data",
                                "Run Calculate first.",
                                parent=self.root)
            return

        path = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV", "*.csv")],
            initialfile=f"fulfillment_net_{self._fp_date_var.get()}.csv",
            parent=self.root)
        if not path:
            return

        with open(path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow([
                "SKU", "Available", "Direct", "PR-CJAM", "CEX-EC",
                "EX-EC", "Total Demand", "NET", "Headroom", "Status",
            ])
            for r in self.fp_results:
                headroom = self._fp_headroom_bar(r["net"],
                                                  r["total_demand"])
                writer.writerow([
                    r["sku"], r["available"],
                    r["direct"], r["prcjam"],
                    r["cexec"], r["exec"],
                    r["total_demand"], r["net"],
                    headroom, r["status"],
                ])

        self._fp_set_mascot("happy", "CSV exported!")
        self._fp_status_var.set(f"Exported to {os.path.basename(path)}")
        self.root.after(3000, lambda: self._fp_set_mascot(
            "idle", "Ready to plan!"))

    # ─────────────────────────────────────────────────────────────────
    #  TAB 7 — SETTINGS
    # ─────────────────────────────────────────────────────────────────

    def _build_settings_tab(self):
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text="  Settings  ")

        canvas = tk.Canvas(tab, bg=_BG, highlightthickness=0)
        scrollbar = ttk.Scrollbar(tab, orient="vertical",
                                  command=canvas.yview)
        scroll_frame = ttk.Frame(canvas)

        scroll_frame.bind("<Configure>",
                          lambda e: canvas.configure(
                              scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=scroll_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # enable mousewheel scrolling
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        canvas.bind_all("<MouseWheel>", _on_mousewheel)

        # ── Global defaults ──
        gd = ttk.LabelFrame(scroll_frame, text="Global Default Lead Times",
                             padding=10)
        gd.pack(fill="x", padx=15, pady=(10, 5))

        defaults = [
            ("Purchase Lead Time (days):", "default_purchase_lt",
             self.saved.get("default_purchase_lt", "7")),
            ("Production Lead Time (days):", "default_production_lt",
             self.saved.get("default_production_lt", "3")),
            ("Shipping Lead Time (days):", "default_shipping_lt",
             self.saved.get("default_shipping_lt", "5")),
            ("Default Safety Stock (units):", "default_safety_stock",
             self.saved.get("default_safety_stock", "10")),
            ("Fulfillment Buffer (days):", "fulfillment_buffer",
             self.saved.get("fulfillment_buffer", "10")),
            ("Expiration Warning (days):", "expiration_warning_days",
             self.saved.get("expiration_warning_days", "14")),
        ]
        self.settings_vars = {}

        for i, (label, key, default) in enumerate(defaults):
            ttk.Label(gd, text=label).grid(
                row=i, column=0, sticky="w", padx=(0, 10), pady=3)
            var = tk.StringVar(value=str(default))
            ttk.Entry(gd, textvariable=var, width=10).grid(
                row=i, column=1, sticky="w", pady=3)
            self.settings_vars[key] = var

        # Reship buffer (auto-calculated, editable)
        reship_row = len(defaults)
        ttk.Label(gd, text="Reship Buffer (%):").grid(
            row=reship_row, column=0, sticky="w", padx=(0, 10), pady=3)
        self.reship_var = tk.StringVar(
            value=str(self.reship_buffer_pct))
        ttk.Entry(gd, textvariable=self.reship_var, width=10).grid(
            row=reship_row, column=1, sticky="w", pady=3)

        ttk.Label(gd, text="Per-SKU settings override these globals. "
                  "Reship buffer auto-updates from fulfillment imports.",
                  style="Dim.TLabel").grid(
                      row=reship_row + 1, column=0, columnspan=2,
                      sticky="w", pady=(8, 0))

        # ── Churn ──
        ch = ttk.LabelFrame(scroll_frame,
                             text="Recharge Churn Settings", padding=10)
        ch.pack(fill="x", padx=15, pady=5)

        ttk.Label(ch, text="Default Churn Rate (%):").grid(
            row=0, column=0, sticky="w", padx=(0, 10), pady=3)
        # reuse the demand tab's variable
        ttk.Entry(ch, textvariable=self.default_churn_var, width=8).grid(
            row=0, column=1, sticky="w", pady=3)
        ttk.Label(ch,
                  text="Applied to Recharge subscription quantities. "
                       "Per-SKU overrides available via SKU settings.",
                  style="Dim.TLabel").grid(
                      row=1, column=0, columnspan=2, sticky="w", pady=(5, 0))

        # ── Bundle mappings shortcut ──
        bm = ttk.LabelFrame(scroll_frame, text="Bundle Mappings", padding=10)
        bm.pack(fill="x", padx=15, pady=5)

        bm_row = ttk.Frame(bm)
        bm_row.pack(fill="x")
        self.bundle_count_var = tk.StringVar(
            value=f"{len(self.bundle_map)} bundle(s) defined")
        ttk.Label(bm_row, textvariable=self.bundle_count_var).pack(
            side="left")
        ttk.Button(bm_row, text="Edit Bundle Mappings...",
                   command=self._open_bundle_editor).pack(side="right")

        # ── Recharge API ──
        api = ttk.LabelFrame(scroll_frame, text="Recharge API", padding=10)
        api.pack(fill="x", padx=15, pady=5)

        ttk.Label(api, text="API Token:").grid(
            row=0, column=0, sticky="w", padx=(0, 10), pady=3)
        ttk.Entry(api, textvariable=self.recharge_token_var, width=50,
                  show="*").grid(row=0, column=1, sticky="w", pady=3)
        ttk.Label(api,
                  text="Token is saved locally in settings file. "
                       "Never shared.",
                  style="Dim.TLabel").grid(
                      row=1, column=0, columnspan=2, sticky="w", pady=(3, 0))

        # ── Curation & Forecasting ──
        cf = ttk.LabelFrame(scroll_frame,
                             text="Curation & Forecasting", padding=10)
        cf.pack(fill="x", padx=15, pady=5)

        cf_row = ttk.Frame(cf)
        cf_row.pack(fill="x")
        self.cohort_count_var = tk.StringVar(
            value=f"{len(self.cohorts)} cohort(s), "
                  f"{len(self.curation_recipes)} recipe(s)")
        ttk.Label(cf_row, textvariable=self.cohort_count_var).pack(
            side="left")

        cf_btns = ttk.Frame(cf)
        cf_btns.pack(fill="x", pady=(5, 0))
        ttk.Button(cf_btns, text="Edit Recipes...",
                   command=self._open_recipe_editor).pack(
                       side="left", padx=(0, 5))
        ttk.Button(cf_btns, text="Retention Matrix...",
                   command=self._open_retention_editor).pack(
                       side="left", padx=(0, 5))
        ttk.Button(cf_btns, text="Cohort Manager...",
                   command=self._open_cohort_editor).pack(
                       side="left", padx=(0, 5))
        ttk.Button(cf_btns, text="Monthly Box Recipes...",
                   command=self._open_monthly_box_editor).pack(side="left")

        # ── Supply Pipeline ──
        sp = ttk.LabelFrame(scroll_frame,
                             text="Supply Pipeline", padding=10)
        sp.pack(fill="x", padx=15, pady=5)

        sp_row = ttk.Frame(sp)
        sp_row.pack(fill="x")
        self.po_count_var = tk.StringVar(
            value=f"{len(self.open_pos)} open PO(s)")
        ttk.Label(sp_row, textvariable=self.po_count_var).pack(side="left")
        ttk.Button(sp_row, text="Manage Open POs...",
                   command=self._open_po_editor).pack(side="right")

        sp_btns = ttk.Frame(sp)
        sp_btns.pack(fill="x", pady=(5, 0))
        ttk.Button(sp_btns, text="Vendor Catalog...",
                   command=self._open_vendor_catalog).pack(
                       side="left", padx=(0, 5))
        ttk.Button(sp_btns, text="Auto-PO Generator",
                   command=self._generate_auto_po).pack(
                       side="left", padx=(0, 5))
        ttk.Button(sp_btns, text="Production Orders",
                   command=self._generate_production_orders).pack(
                       side="left", padx=(0, 5))
        ttk.Button(sp_btns, text="Processing Queue",
                   command=self._show_processing_queue).pack(
                       side="left", padx=(0, 5))
        ttk.Button(sp_btns, text="Yield History",
                   command=self._show_yield_history).pack(
                       side="left")

        # ── Yield Reconciliation ──
        yr = ttk.LabelFrame(scroll_frame,
                             text="Yield Reconciliation", padding=10)
        yr.pack(fill="x", padx=15, pady=5)

        yr_fields = [
            ("Match Window (days):", "yield_reconciliation_window_days",
             str(self.yield_recon_window_days)),
            ("Variance Threshold (%):", "yield_reconciliation_threshold_pct",
             str(self.yield_recon_threshold_pct)),
            ("Min Variance (units):", "yield_reconciliation_threshold_min",
             str(self.yield_recon_threshold_min)),
        ]
        for i, (label, key, default) in enumerate(yr_fields):
            ttk.Label(yr, text=label).grid(
                row=i, column=0, sticky="w", padx=(0, 10), pady=3)
            var = tk.StringVar(value=default)
            ttk.Entry(yr, textvariable=var, width=10).grid(
                row=i, column=1, sticky="w", pady=3)
            self.settings_vars[key] = var

        yr_btn_row = ttk.Frame(yr)
        yr_btn_row.grid(row=len(yr_fields), column=0, columnspan=2,
                        sticky="w", pady=(5, 0))
        ttk.Button(yr_btn_row, text="View Discrepancies",
                   command=self._show_yield_discrepancies).pack(
                       side="left", padx=(0, 5))

        ttk.Label(yr, text="Flags discrepancies between yield records "
                  "and Dropbox inventory snapshots.",
                  style="Dim.TLabel").grid(
                      row=len(yr_fields) + 1, column=0, columnspan=2,
                      sticky="w", pady=(5, 0))

        # ── Warehouses ──
        wh = ttk.LabelFrame(scroll_frame,
                             text="Warehouses", padding=10)
        wh.pack(fill="x", padx=15, pady=5)

        for i, (wh_key, wh_info) in enumerate(self.warehouses.items()):
            lbl = wh_info.get("label", wh_key)
            is_ful = wh_info.get("is_fulfillment", False)
            caps = ", ".join(wh_info.get("capabilities", []))
            types = ", ".join(wh_info.get("item_types", ["All"]))
            text = f"{lbl} ({'Fulfillment' if is_ful else 'Non-fulfillment'})"
            if caps:
                text += f"  |  Capabilities: {caps}"
            if types:
                text += f"  |  Items: {types}"
            ttk.Label(wh, text=text).grid(
                row=i, column=0, sticky="w", pady=2)

        ttk.Label(wh, text="Primary warehouse inventory counts toward "
                  "fulfillment readiness. Woburn shows as potential only.",
                  style="Dim.TLabel").grid(
                      row=len(self.warehouses), column=0, sticky="w",
                      pady=(5, 0))

        # ── Integrations ──
        integ = ttk.LabelFrame(scroll_frame,
                                text="Integrations", padding=10)
        integ.pack(fill="x", padx=15, pady=5)

        # ClickUp
        cu_row = ttk.Frame(integ)
        cu_row.pack(fill="x", pady=(0, 5))
        ttk.Label(cu_row, text="ClickUp API Token:").pack(side="left")
        self.clickup_token_var = tk.StringVar(value=self.clickup_api_token)
        ttk.Entry(cu_row, textvariable=self.clickup_token_var,
                  width=40, show="*").pack(side="left", padx=(5, 10))
        ttk.Label(cu_row, text="List ID:").pack(side="left")
        self.clickup_list_var = tk.StringVar(value=self.clickup_list_id)
        ttk.Entry(cu_row, textvariable=self.clickup_list_var,
                  width=15).pack(side="left", padx=(5, 0))

        # Google Calendar
        gc_row = ttk.Frame(integ)
        gc_row.pack(fill="x", pady=(0, 5))
        ttk.Label(gc_row, text="Google Calendar Client ID:").pack(side="left")
        self.gcal_id_var = tk.StringVar(value=self.gcal_client_id)
        ttk.Entry(gc_row, textvariable=self.gcal_id_var,
                  width=30).pack(side="left", padx=(5, 10))
        ttk.Label(gc_row, text="Secret:").pack(side="left")
        self.gcal_secret_var = tk.StringVar(value=self.gcal_client_secret)
        ttk.Entry(gc_row, textvariable=self.gcal_secret_var,
                  width=25, show="*").pack(side="left", padx=(5, 0))

        # Dropbox
        db_row = ttk.Frame(integ)
        db_row.pack(fill="x", pady=(0, 5))
        ttk.Label(db_row, text="Dropbox App Key:").pack(side="left")
        self.dropbox_key_var = tk.StringVar(value=self.dropbox_app_key)
        ttk.Entry(db_row, textvariable=self.dropbox_key_var,
                  width=20).pack(side="left", padx=(5, 10))
        ttk.Label(db_row, text="App Secret:").pack(side="left")
        self.dropbox_secret_var = tk.StringVar(value=self.dropbox_app_secret)
        ttk.Entry(db_row, textvariable=self.dropbox_secret_var,
                  width=25, show="*").pack(side="left", padx=(5, 0))

        db_link_row = ttk.Frame(integ)
        db_link_row.pack(fill="x", pady=(0, 5))
        ttk.Label(db_link_row,
                  text="Shared Link (if no full access):").pack(side="left")
        self.dropbox_link_var = tk.StringVar(
            value=self.dropbox_shared_link)
        ttk.Entry(db_link_row, textvariable=self.dropbox_link_var,
                  width=50).pack(side="left", padx=(5, 0), fill="x",
                                 expand=True)

        ttk.Label(integ,
                  text="Tokens stored locally. OAuth flows use localhost:21849."
                  " Use Shared Link if folder isn't in your Dropbox.",
                  style="Dim.TLabel").pack(anchor="w", pady=(3, 0))

        # ── Slack Notifications ──
        slack = ttk.LabelFrame(scroll_frame,
                                text="Slack Notifications", padding=10)
        slack.pack(fill="x", padx=15, pady=5)

        sl_row1 = ttk.Frame(slack)
        sl_row1.pack(fill="x", pady=(0, 5))
        ttk.Label(sl_row1, text="Webhook URL:").pack(side="left")
        self.slack_url_var = tk.StringVar(value=self.slack_webhook_url)
        ttk.Entry(sl_row1, textvariable=self.slack_url_var,
                  width=50).pack(side="left", padx=(5, 0))

        sl_row2 = ttk.Frame(slack)
        sl_row2.pack(fill="x")
        self.slack_crit_var = tk.BooleanVar(
            value=self.slack_notify_critical)
        ttk.Checkbutton(sl_row2, text="Critical reorder alerts",
                        variable=self.slack_crit_var).pack(
                            side="left", padx=(0, 15))
        self.slack_exp_var = tk.BooleanVar(
            value=self.slack_notify_expiring)
        ttk.Checkbutton(sl_row2, text="Expiring inventory",
                        variable=self.slack_exp_var).pack(
                            side="left", padx=(0, 15))
        self.slack_short_var = tk.BooleanVar(
            value=self.slack_notify_shortfall)
        ttk.Checkbutton(sl_row2, text="Fulfillment shortfalls",
                        variable=self.slack_short_var).pack(side="left")

        sl_test = ttk.Frame(slack)
        sl_test.pack(fill="x", pady=(5, 0))
        ttk.Button(sl_test, text="Test Slack",
                   command=lambda: self._send_slack_message(
                       ":white_check_mark: Test from Inventory Reorder "
                       f"v{APP_VERSION}")).pack(side="left")

        # ── Email (SMTP) for Depletion Reports ──
        email_frame = ttk.LabelFrame(scroll_frame,
                                      text="Email (Depletion Reports)",
                                      padding=10)
        email_frame.pack(fill="x", padx=15, pady=5)

        em_row1 = ttk.Frame(email_frame)
        em_row1.pack(fill="x", pady=(0, 3))
        ttk.Label(em_row1, text="SMTP Host:").pack(side="left")
        self.smtp_host_var = tk.StringVar(value=self.smtp_host)
        ttk.Entry(em_row1, textvariable=self.smtp_host_var,
                  width=25).pack(side="left", padx=(5, 10))
        ttk.Label(em_row1, text="Port:").pack(side="left")
        self.smtp_port_var = tk.StringVar(value=self.smtp_port)
        ttk.Entry(em_row1, textvariable=self.smtp_port_var,
                  width=6).pack(side="left", padx=(5, 0))

        em_row2 = ttk.Frame(email_frame)
        em_row2.pack(fill="x", pady=(0, 3))
        ttk.Label(em_row2, text="SMTP User:").pack(side="left")
        self.smtp_user_var = tk.StringVar(value=self.smtp_user)
        ttk.Entry(em_row2, textvariable=self.smtp_user_var,
                  width=30).pack(side="left", padx=(5, 10))
        ttk.Label(em_row2, text="Password:").pack(side="left")
        self.smtp_pass_var = tk.StringVar(value=self.smtp_password)
        ttk.Entry(em_row2, textvariable=self.smtp_pass_var,
                  width=25, show="*").pack(side="left", padx=(5, 0))

        em_row3 = ttk.Frame(email_frame)
        em_row3.pack(fill="x", pady=(0, 3))
        ttk.Label(em_row3, text="Send To:").pack(side="left")
        self.email_to_var = tk.StringVar(value=self.depletion_email_to)
        ttk.Entry(em_row3, textvariable=self.email_to_var,
                  width=50).pack(side="left", padx=(5, 0))

        em_row4 = ttk.Frame(email_frame)
        em_row4.pack(fill="x", pady=(0, 3))
        ttk.Label(em_row4, text="From:").pack(side="left")
        self.email_from_var = tk.StringVar(
            value=self.depletion_email_from)
        ttk.Entry(em_row4, textvariable=self.email_from_var,
                  width=30).pack(side="left", padx=(5, 10))
        ttk.Label(em_row4, text="Subject Tag:").pack(side="left")
        self.depletion_tag_var = tk.StringVar(
            value=self.saved.get("depletion_tag_prefix", "RMFG"))
        ttk.Entry(em_row4, textvariable=self.depletion_tag_var,
                  width=10).pack(side="left", padx=(5, 0))

        ttk.Label(email_frame,
                  text="To: comma-separate multiple recipients. "
                       "Subject auto-formats as TAG_YYYYMMDD // N Orders. "
                       "Gmail: use App Passwords.",
                  style="Dim.TLabel").pack(anchor="w", pady=(3, 0))

        # ── Automation ──
        auto = ttk.LabelFrame(scroll_frame,
                               text="Automation", padding=10)
        auto.pack(fill="x", padx=15, pady=5)

        au_row1 = ttk.Frame(auto)
        au_row1.pack(fill="x", pady=(0, 5))
        ttk.Label(au_row1,
                  text="Auto-refresh interval (min, 0=off):").pack(
                      side="left")
        self.auto_refresh_var = tk.StringVar(
            value=str(self.auto_refresh_interval))
        ttk.Entry(au_row1, textvariable=self.auto_refresh_var,
                  width=6).pack(side="left", padx=(5, 15))

        ttk.Label(au_row1,
                  text="Auto-PO deficit threshold (0=off):").pack(
                      side="left")
        self.auto_po_var = tk.StringVar(
            value=str(self.auto_po_threshold))
        ttk.Entry(au_row1, textvariable=self.auto_po_var,
                  width=6).pack(side="left", padx=(5, 0))

        au_row2 = ttk.Frame(auto)
        au_row2.pack(fill="x", pady=(0, 5))
        self.auto_clickup_var = tk.BooleanVar(
            value=self.auto_sync_clickup)
        ttk.Checkbutton(au_row2, text="Auto-sync to ClickUp",
                        variable=self.auto_clickup_var).pack(
                            side="left", padx=(0, 15))
        self.auto_gcal_var = tk.BooleanVar(
            value=self.auto_sync_gcal)
        ttk.Checkbutton(au_row2, text="Auto-sync to Google Calendar",
                        variable=self.auto_gcal_var).pack(side="left")

        # ── Webhooks ──
        wh = ttk.LabelFrame(scroll_frame,
                              text="Webhooks (Shopify / Recharge)",
                              padding=10)
        wh.pack(fill="x", padx=15, pady=5)

        wh_row1 = ttk.Frame(wh)
        wh_row1.pack(fill="x", pady=(0, 3))
        ttk.Label(wh_row1, text="Listen Port:").pack(side="left")
        self.webhook_port_var = tk.StringVar(
            value=str(self.webhook_port))
        ttk.Entry(wh_row1, textvariable=self.webhook_port_var,
                  width=6).pack(side="left", padx=(5, 15))

        wh_row2 = ttk.Frame(wh)
        wh_row2.pack(fill="x", pady=(0, 3))
        ttk.Label(wh_row2, text="Shopify Secret:").pack(side="left")
        self.wh_shopify_var = tk.StringVar(
            value=self.webhook_secret_shopify)
        ttk.Entry(wh_row2, textvariable=self.wh_shopify_var,
                  width=30, show="*").pack(side="left", padx=(5, 10))
        ttk.Label(wh_row2, text="Recharge Secret:").pack(side="left")
        self.wh_recharge_var = tk.StringVar(
            value=self.webhook_secret_recharge)
        ttk.Entry(wh_row2, textvariable=self.wh_recharge_var,
                  width=30, show="*").pack(side="left", padx=(5, 0))

        ttk.Label(wh,
                  text="Point Shopify webhooks to http://<your-ip>:<port>"
                       "/shopify, Recharge to /recharge. "
                       "Use ngrok for public access.",
                  style="Dim.TLabel").pack(anchor="w", pady=(3, 0))

        # ── Save button ──
        save_frame = ttk.Frame(scroll_frame)
        save_frame.pack(fill="x", padx=15, pady=15)
        ttk.Button(save_frame, text="Save All Settings",
                   style="Apply.TButton",
                   command=self._save_all_settings).pack(side="left")

    # ─────────────────────────────────────────────────────────────────
    #  CORE CALCULATIONS
    # ─────────────────────────────────────────────────────────────────

    def _get_global_defaults(self):
        """Return global default values, falling back to safe defaults."""
        def _float(key, fallback):
            try:
                return float(self.settings_vars[key].get())
            except (ValueError, KeyError):
                return fallback

        return {
            "purchase_lt": _float("default_purchase_lt", 7),
            "production_lt": _float("default_production_lt", 3),
            "shipping_lt": _float("default_shipping_lt", 5),
            "safety_stock": _float("default_safety_stock", 10),
            "fulfillment_buffer": _float("fulfillment_buffer", 10),
        }

    def _get_reship_pct(self):
        """Get current reship buffer %, preferring the UI var if set."""
        try:
            return float(self.reship_var.get())
        except (ValueError, AttributeError):
            return self.reship_buffer_pct

    def _get_default_churn(self):
        try:
            return float(self.default_churn_var.get())
        except ValueError:
            return 5.0

    def _compute_combined_demand(self):
        """Combine all 3 demand sources into per-SKU weekly totals.

        Returns: {sku: {recharge_wk, shopify_wk, manual_wk, total_wk, daily}}
        """
        all_skus = set()
        default_churn = self._get_default_churn()

        # 1. Recharge demand (apply churn, decompose bundles)
        # If queued charges exist for the current month, use those
        # (converted monthly → weekly ÷ 4.33) instead of subscription counts
        recharge_by_sku = defaultdict(float)
        current_month = datetime.date.today().strftime("%Y-%m")
        queued_month_data = self.recharge_queued.get(current_month)

        if queued_month_data:
            # Use queued charges: monthly totals → weekly
            for sku, qty in queued_month_data.items():
                # Skip generic PR-CJAM-GEN and bare CEX-EC — resolved below
                upper = sku.upper()
                if upper == "PR-CJAM-GEN" or upper == "CEX-EC":
                    continue
                weekly_qty = qty / 4.33
                churn = self.sku_settings.get(sku, {}).get(
                    "churn_pct", default_churn)
                adjusted = apply_churn_rate(weekly_qty, churn)
                for comp_sku, comp_qty in decompose_bundles(
                        sku, adjusted, self.bundle_map):
                    recharge_by_sku[comp_sku] += comp_qty
                    all_skus.add(comp_sku)

            # Add resolved PR-CJAM/CEX-EC as actual cheese/jam SKU demand
            resolved = self.recharge_queued_resolved.get(current_month, {})
            for suffix, count in resolved.get("pr_cjam", {}).items():
                weekly_count = count / 4.33
                if suffix == "GEN":
                    recharge_by_sku["PR-CJAM-GEN"] += weekly_count
                    all_skus.add("PR-CJAM-GEN")
                    continue
                cjam = self.pr_cjam.get(suffix, {})
                cheese = cjam.get("cheese")
                jam = cjam.get("jam")
                if cheese:
                    recharge_by_sku[cheese] += weekly_count
                    all_skus.add(cheese)
                if jam:
                    recharge_by_sku[jam] += weekly_count
                    all_skus.add(jam)
                recharge_by_sku[f"PR-CJAM-{suffix}"] += weekly_count
                all_skus.add(f"PR-CJAM-{suffix}")
            for suffix, count in resolved.get("cex_ec", {}).items():
                weekly_count = count / 4.33
                if suffix == "GEN":
                    recharge_by_sku["CEX-EC"] += weekly_count
                    all_skus.add("CEX-EC")
                    continue
                cheese = self.cex_ec.get(suffix)
                if cheese:
                    recharge_by_sku[cheese] += weekly_count
                    all_skus.add(cheese)
                recharge_by_sku[f"CEX-EC-{suffix}"] += weekly_count
                all_skus.add(f"CEX-EC-{suffix}")
        else:
            # Fall back to subscription-based demand
            for sku, qty in self.recharge_demand.items():
                churn = self.sku_settings.get(sku, {}).get(
                    "churn_pct", default_churn)
                adjusted = apply_churn_rate(qty, churn)
                for comp_sku, comp_qty in decompose_bundles(
                        sku, adjusted, self.bundle_map):
                    recharge_by_sku[comp_sku] += comp_qty
                    all_skus.add(comp_sku)

        # 2. Shopify demand: API-pulled + manual overrides (decompose bundles)
        shopify_by_sku = defaultdict(float)
        # start with API-pulled demand
        for sku, qty in self.shopify_api_demand.items():
            for comp_sku, comp_qty in decompose_bundles(
                    sku, qty, self.bundle_map):
                shopify_by_sku[comp_sku] += comp_qty
                all_skus.add(comp_sku)
        # layer manual overrides on top (these ADD to API data)
        for sku, qty in self.shopify_forecast.items():
            for comp_sku, comp_qty in decompose_bundles(
                    sku, qty, self.bundle_map):
                shopify_by_sku[comp_sku] += comp_qty
                all_skus.add(comp_sku)

        # 3. Manual demand (already per-SKU, no decomposition)
        for sku in self.manual_demand:
            all_skus.add(sku)

        # Also include any SKU in inventory
        for sku in self.inventory:
            all_skus.add(sku)

        # Combine
        combined = {}
        for sku in all_skus:
            rc = recharge_by_sku.get(sku, 0)
            sp = shopify_by_sku.get(sku, 0)
            ma = self.manual_demand.get(sku, 0)
            total_wk = rc + sp + ma
            daily = total_wk / 7.0
            combined[sku] = {
                "recharge_wk": round(rc, 2),
                "shopify_wk": round(sp, 2),
                "manual_wk": round(ma, 2),
                "total_wk": round(total_wk, 2),
                "daily": round(daily, 4),
            }

        return combined

    def _recalculate(self):
        """Recalculate all reorder points and refresh the dashboard."""
        self.status_var.set("Calculating...")
        self.root.update_idletasks()

        combined = self._compute_combined_demand()
        globals_ = self._get_global_defaults()

        # Compute converted supply from raw materials
        wheel_supply = compute_wheel_supply(
            self.wheel_inventory, self.adjusted_conversion_factors)
        bulk_supply = compute_bulk_supply(
            self.inventory, self.bulk_conversions)

        rows = []
        alert_count = 0
        reorder_count = 0
        expiring_count = 0

        today = datetime.date.today()
        try:
            warn_days = int(self.settings_vars.get(
                "expiration_warning_days", tk.StringVar(value="14")).get())
        except (ValueError, AttributeError):
            warn_days = 14

        # Only show real product SKUs on dashboard
        _PRODUCT_PREFIXES = ("CH-", "PK-", "MT-", "AC-")

        for sku, demand in combined.items():
            # Skip non-product SKUs (bundles, admin, packaging, etc.)
            if not sku.startswith(_PRODUCT_PREFIXES):
                continue
            inv = self.inventory.get(sku, {})
            on_hand = float(inv.get("qty", 0))
            name = inv.get("name", "")
            category = inv.get("category", "")

            # per-SKU or global lead times
            ss = self.sku_settings.get(sku, {})
            purchase_lt = ss.get("purchase_lt", globals_["purchase_lt"])
            production_lt = ss.get("production_lt", globals_["production_lt"])
            shipping_lt = ss.get("shipping_lt", globals_["shipping_lt"])
            safety = ss.get("safety_stock", globals_["safety_stock"])

            total_lt = calculate_total_lead_time(
                purchase_lt, production_lt, shipping_lt)
            fulfillment_buf = globals_["fulfillment_buffer"]
            daily = demand["daily"]
            # Apply reship buffer: inflate demand by reship %
            reship_mult = 1.0 + (self._get_reship_pct() / 100.0)
            adjusted_daily = daily * reship_mult
            reorder_pt = calculate_reorder_point(
                adjusted_daily, total_lt + fulfillment_buf, safety)
            surplus = on_hand - reorder_pt

            # status
            if on_hand <= 0 and daily > 0:
                status = "OUT OF STOCK"
                tag = "CRITICAL"
                alert_count += 1
                reorder_count += 1
            elif on_hand <= reorder_pt * 0.5 and daily > 0:
                status = "CRITICAL"
                tag = "CRITICAL"
                alert_count += 1
                reorder_count += 1
            elif on_hand <= reorder_pt:
                status = "REORDER"
                tag = "WARNING"
                reorder_count += 1
            elif on_hand > reorder_pt * 3 and daily > 0:
                status = "OVERSTOCK"
                tag = "OVERSTOCK"
            else:
                status = "OK"
                tag = "OK"

            # expiration info
            exp_dates = inv.get("expiration_dates", [])
            exp_text = ""
            if exp_dates:
                try:
                    earliest_dt = datetime.date.fromisoformat(exp_dates[0])
                    days_until = (earliest_dt - today).days
                    if days_until < 0:
                        exp_text = "EXPIRED"
                        expiring_count += 1
                    elif days_until <= warn_days:
                        exp_text = f"{days_until}d ({exp_dates[0]})"
                        expiring_count += 1
                    else:
                        exp_text = exp_dates[0]
                except ValueError:
                    pass

            # Total available = on hand + converted raw supply
            ws = wheel_supply.get(sku, 0)
            bs = bulk_supply.get(sku, 0)
            raw_supply = ws + bs
            total_avail = on_hand + raw_supply

            rows.append({
                "sku": sku,
                "name": name,
                "category": category,
                "on_hand": on_hand,
                "weekly_usage": round(daily * 7, 1),
                "lead_time": total_lt,
                "safety_stock": safety,
                "reorder_point": round(reorder_pt, 1),
                "surplus": round(surplus, 1),
                "status": status,
                "recharge_wk": demand["recharge_wk"],
                "shopify_wk": demand["shopify_wk"],
                "manual_wk": demand["manual_wk"],
                "wheel_supply": round(ws, 1) if ws > 0 else "",
                "bulk_supply": round(bs, 1) if bs > 0 else "",
                "total_avail": round(total_avail, 1) if raw_supply > 0
                    else on_hand,
                "expiration": exp_text,
                "tag": tag,
                "warehouse": inv.get("warehouse", "Primary"),
            })

        self._dash_rows = rows
        self._populate_dashboard(rows)

        # update summaries
        self.summary_total_var.set(f"SKUs: {len(rows)}")
        self.summary_alert_var.set(f"Alerts: {alert_count}")
        self.summary_reorder_var.set(f"Need Reorder: {reorder_count}")
        self.summary_expiring_var.set(f"Expiring: {expiring_count}")
        self.status_var.set(
            f"Calculated {len(rows)} SKUs  |  "
            f"{datetime.datetime.now().strftime('%H:%M:%S')}")

        # also refresh demand summary
        self._refresh_demand_summary()

        # refresh notification alerts
        self._generate_alerts()

    def _populate_dashboard(self, rows):
        """Insert rows into dashboard treeview."""
        for item in self.dash_tree.get_children():
            self.dash_tree.delete(item)

        # sort
        sort_col = self._dash_sort_col
        rev = self._dash_sort_rev

        def sort_key(r):
            val = r.get(sort_col, "")
            if sort_col == "status":
                order = {"OUT OF STOCK": 0, "CRITICAL": 1, "REORDER": 2,
                         "OK": 3, "OVERSTOCK": 4}
                return order.get(val, 5)
            try:
                return float(val)
            except (ValueError, TypeError):
                return str(val).lower()

        rows_sorted = sorted(rows, key=sort_key, reverse=rev)

        for r in rows_sorted:
            values = (
                r["sku"], r["name"], r.get("category", ""),
                r["on_hand"], r["weekly_usage"],
                r["lead_time"], r["safety_stock"], r["reorder_point"],
                r["surplus"], r["status"], r["recharge_wk"],
                r["shopify_wk"], r["manual_wk"],
                r.get("wheel_supply", ""),
                r.get("bulk_supply", ""),
                r.get("total_avail", r["on_hand"]),
                r.get("expiration", ""),
            )
            self.dash_tree.insert("", "end", values=values,
                                  tags=(r["tag"],))

    def _sort_dash(self, col):
        if self._dash_sort_col == col:
            self._dash_sort_rev = not self._dash_sort_rev
        else:
            self._dash_sort_col = col
            self._dash_sort_rev = False
        self._apply_filter()

    def _apply_filter(self):
        """Filter and re-populate dashboard based on filter text,
        alert-only toggle, archive visibility, and warehouse."""
        filter_text = self.filter_var.get().strip().lower()
        alert_only = self.alert_only_var.get()
        show_archived = self.show_archived_var.get()
        wh_filter = (self.dash_warehouse_var.get()
                     if hasattr(self, 'dash_warehouse_var') else
                     "All Locations")

        filtered = []
        for r in self._dash_rows:
            if not show_archived and r["sku"] in self.archived_skus:
                continue
            if alert_only and r["status"] == "OK":
                continue
            if filter_text:
                searchable = f"{r['sku']} {r['name']}".lower()
                if filter_text not in searchable:
                    continue
            # Warehouse filter (check warehouse_qty for split inventory)
            if wh_filter and wh_filter != "All Locations":
                if self._qty_at(r["sku"], wh_filter) <= 0:
                    continue
            filtered.append(r)

        self._populate_dashboard(filtered)

    # ─────────────────────────────────────────────────────────────────
    #  DEMAND SUMMARY
    # ─────────────────────────────────────────────────────────────────

    def _refresh_demand_summary(self):
        for item in self.demand_tree.get_children():
            self.demand_tree.delete(item)

        combined = self._compute_combined_demand()
        for sku in sorted(combined.keys()):
            d = combined[sku]
            self.demand_tree.insert("", "end", values=(
                sku, d["recharge_wk"], d["shopify_wk"],
                d["manual_wk"], d["total_wk"], round(d["daily"], 2)))

    # ─────────────────────────────────────────────────────────────────
    #  CSV IMPORT / EXPORT
    # ─────────────────────────────────────────────────────────────────

    def _import_inventory_csv(self):
        path = getattr(self, '_pending_csv_path', None)
        if not path:
            path = filedialog.askopenfilename(
                title="Import Inventory CSV/XLSX",
                filetypes=[("Inventory files", "*.csv *.xlsx"),
                           ("CSV files", "*.csv"),
                           ("Excel files", "*.xlsx"),
                           ("All files", "*.*")])
        if not path:
            return

        try:
            if path.lower().endswith(".xlsx"):
                if not HAS_OPENPYXL:
                    messagebox.showerror(
                        "Import Error",
                        "openpyxl is required for XLSX import.\n"
                        "pip install openpyxl")
                    return
                wb = openpyxl.load_workbook(path, data_only=True)
                ws = wb.active
                all_rows = list(ws.iter_rows(values_only=True))
                if not all_rows:
                    messagebox.showerror("Import Error",
                                         "XLSX has no data.")
                    return
                headers = [str(c) if c is not None else ""
                           for c in all_rows[0]]
                rows_raw = [
                    [str(c) if c is not None else "" for c in row]
                    for row in all_rows[1:]
                ]
            else:
                with open(path, "r", newline="",
                          encoding="utf-8-sig") as f:
                    reader = csv.reader(f)
                    headers = next(reader)
                    rows_raw = list(reader)
        except Exception as e:
            messagebox.showerror("Import Error",
                                 f"Failed to read file:\n{e}")
            return

        if not headers:
            messagebox.showerror("Import Error",
                                 "File has no header row.")
            return

        # show column mapping dialog
        dlg = ColumnMappingDialog(self.root, headers)
        self.root.wait_window(dlg)
        if not dlg.result:
            return

        mapping = dlg.result
        self.last_csv_mapping = mapping

        # import rows — also detect cheese wheels for wheel_inventory
        imported = 0
        wheels_found = 0

        # Category parsing from Ingredient column prefix
        # Order matters: more specific prefixes first
        _INGREDIENT_CATEGORIES = [
            ("Cheese Wheel,",   "Cheese Wheels"),
            ("Cheese Slice,",   "Cheese Slices"),
            ("Crackers,",       "Crackers"),
            ("Jam,",            "Jams"),
            ("Jar,",            "Jams"),
            ("Packet,",         "Packets (Fruit & Nuts)"),
            ("Pre-Packaged Meat,", "Pre-Packaged Meats"),
            ("Bulk Meat,",      "Bulk Raw Materials"),
            ("Bulk,",           "Bulk Raw Materials"),
            ("Slice,",          "Accompaniments"),
            ("Tasting Guide,",  "Packaging"),
            ("Bag,",            "Packaging"),
            ("Boxes,",          "Packaging"),
            ("Cheese Paper,",   "Packaging"),
            ("Gel Pack,",       "Packaging"),
            ("Insulation,",     "Packaging"),
            ("Labels,",         "Packaging"),
        ]

        def _parse_category_from_ingredient(ingredient_val):
            """Derive category from Ingredient column prefix."""
            if not ingredient_val:
                return ""
            for prefix, cat in _INGREDIENT_CATEGORIES:
                if ingredient_val.startswith(prefix):
                    return cat
            return ""

        for row in rows_raw:
            if len(row) < len(headers):
                row.extend([""] * (len(headers) - len(row)))
            row_dict = dict(zip(headers, row))

            sku = row_dict.get(mapping["SKU"], "").strip()
            if not sku:
                continue

            qty_str = row_dict.get(mapping["Quantity On Hand"], "0").strip()
            try:
                qty = float(qty_str.replace(",", ""))
            except ValueError:
                qty = 0

            entry = {"qty": qty}

            # Parse category from Ingredient column
            # Try mapped "Product Name" first, then look for
            # "Ingredient" column directly
            ingredient_val = ""
            if "Product Name" in mapping:
                ingredient_val = row_dict.get(
                    mapping["Product Name"], "").strip()
            if not ingredient_val:
                ingredient_val = row_dict.get("Ingredient", "").strip()
            if not ingredient_val:
                # Fallback: first column
                ingredient_val = row[0].strip() if row else ""
            auto_cat = _parse_category_from_ingredient(ingredient_val)
            if auto_cat:
                entry["category"] = auto_cat

            if "Product Name" in mapping:
                entry["name"] = row_dict.get(mapping["Product Name"], "")
            if "Category" in mapping:
                cat = row_dict.get(mapping["Category"], "").strip()
                if cat:
                    entry["category"] = cat
            # Store unit info if available (Quantity1 + Unit1)
            unit1 = row_dict.get("Unit1", "").strip()
            qty1_str = row_dict.get("Quantity1", "").strip()
            if unit1:
                entry["unit"] = unit1
            if qty1_str:
                try:
                    entry["unit_size"] = float(
                        qty1_str.replace(",", ""))
                except ValueError:
                    pass
            if "Unit Cost" in mapping:
                cost_str = row_dict.get(mapping["Unit Cost"], "0").strip()
                cost_str = cost_str.replace("$", "").replace(",", "")
                try:
                    entry["unit_cost"] = float(cost_str)
                except ValueError:
                    entry["unit_cost"] = 0

            # expiration dates
            if "Expiration Dates" in mapping:
                raw_exp = row_dict.get(mapping["Expiration Dates"], "").strip()
                if raw_exp:
                    parsed = _parse_expiration_dates(raw_exp)
                    if parsed:
                        entry["expiration_dates"] = parsed

            # optional lead times from CSV
            for field, key in [("Purchase Lead Time", "purchase_lt"),
                               ("Production Lead Time", "production_lt"),
                               ("Shipping Lead Time", "shipping_lt"),
                               ("Safety Stock", "safety_stock")]:
                if field in mapping:
                    val_str = row_dict.get(mapping[field], "").strip()
                    if val_str:
                        try:
                            if sku not in self.sku_settings:
                                self.sku_settings[sku] = {}
                            self.sku_settings[sku][key] = float(val_str)
                        except ValueError:
                            pass

            # Parse warehouse-specific quantities (RMFG = Primary, WIP = pending)
            rmfg_str = row_dict.get("RMFG", "").strip()
            wip_str = row_dict.get("WIP", "").strip()
            has_wh_split = bool(rmfg_str or wip_str)
            if has_wh_split:
                rmfg_qty = 0
                wip_qty = 0
                try:
                    rmfg_qty = float(rmfg_str.replace(",", "")) \
                        if rmfg_str else 0
                except ValueError:
                    pass
                try:
                    wip_qty = float(wip_str.replace(",", "")) \
                        if wip_str else 0
                except ValueError:
                    pass
                if rmfg_qty > 0 or wip_qty > 0:
                    wh_qty = {}
                    if rmfg_qty > 0:
                        wh_qty["Primary"] = rmfg_qty
                    entry["warehouse_qty"] = wh_qty
                    entry["qty"] = rmfg_qty  # only Primary counts
                    entry["warehouse"] = "Primary"
                    # WIP = work in progress (pending processed items)
                    if wip_qty > 0:
                        entry["wip_qty"] = wip_qty

            self.inventory[sku] = entry
            imported += 1

            # detect cheese wheels: category-based or SKU pattern
            is_wheel = (entry.get("category") == "Cheese Wheels" or
                        ("wheel" in entry.get("name", "").lower()))
            if is_wheel and sku.startswith("CH-"):
                # Quantity1 = weight in lbs, Total = count of wheels
                weight_col = None
                count_col = None
                for h in headers:
                    h_lower = h.lower().strip()
                    if h_lower == "quantity1":
                        weight_col = h
                    elif h_lower == "total":
                        count_col = h

                weight = 0
                count = 0
                if weight_col:
                    try:
                        weight = float(
                            row_dict.get(weight_col, "0").strip()
                            .replace(",", ""))
                    except ValueError:
                        weight = 0
                if count_col:
                    try:
                        count = float(
                            row_dict.get(count_col, "0").strip()
                            .replace(",", ""))
                    except ValueError:
                        count = 0

                if weight > 0 and count > 0:
                    self.wheel_inventory[sku] = {
                        "weight_lbs": weight,
                        "count": int(count),
                        "target_sku": sku,  # same SKU after processing
                    }
                    wheels_found += 1

        self._refresh_inventory_tree()
        self._recalculate()
        self.inv_count_var.set(f"{len(self.inventory)} SKUs loaded")

        msg = f"Imported {imported} SKUs from CSV."
        if wheels_found:
            msg += f"\nDetected {wheels_found} cheese wheel entries."
        self.status_var.set(f"Imported {imported} SKUs from CSV")
        messagebox.showinfo("Import Complete", msg)

    def _import_depletion_matrix(self):
        """Import a fulfillment Excel file and subtract depleted quantities."""
        if not HAS_OPENPYXL:
            messagebox.showerror(
                "Missing Dependency",
                "openpyxl is required for Excel import.\n"
                "Install with: pip install openpyxl",
                parent=self.root)
            return

        # Step 1: Get SKU translations (from settings or CSV)
        name_to_sku = dict(self.sku_translations)  # {product_name: sku}

        if not name_to_sku:
            # Try auto-detect in Shipments folder
            script_dir = os.path.dirname(
                os.path.abspath(sys.argv[0] if sys.argv[0] else __file__))
            ship_dir = os.path.join(script_dir, "Shipments")
            auto_csv = None
            if os.path.isdir(ship_dir):
                for fn in os.listdir(ship_dir):
                    if fn.lower().endswith(".csv") and "meal-type" in fn.lower():
                        auto_csv = os.path.join(ship_dir, fn)
                        break

            if auto_csv:
                if messagebox.askyesno(
                        "Translation File Found",
                        f"Found translation file:\n{os.path.basename(auto_csv)}"
                        f"\n\nLoad this file?",
                        parent=self.root):
                    name_to_sku = self._load_translation_csv(auto_csv)
            if not name_to_sku:
                name_to_sku = self._prompt_translation_csv()
            if not name_to_sku:
                return
        else:
            # Offer to re-import if desired
            if messagebox.askyesno(
                    "SKU Translations",
                    f"{len(name_to_sku)} saved translations found.\n\n"
                    f"Use saved translations?\n"
                    f"(No = pick a new CSV file)",
                    parent=self.root):
                pass
            else:
                fresh = self._prompt_translation_csv()
                if not fresh:
                    return
                name_to_sku = fresh

        # Step 2: Select the fulfillment Excel file
        xlsx_path = getattr(self, '_pending_depletion_path', None)
        if not xlsx_path:
            xlsx_path = filedialog.askopenfilename(
                title="Select Fulfillment / Depletion Excel File",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
        if not xlsx_path:
            return

        # Store path so _deplete_and_email can find it
        self._depletion_source_path = xlsx_path

        # Read the Excel and sum quantities per product column
        try:
            wb = openpyxl.load_workbook(xlsx_path, data_only=True,
                                        read_only=True)
            ws = wb[wb.sheetnames[0]]

            rows_iter = ws.iter_rows(values_only=True)
            headers = list(next(rows_iter))

            # Locate key columns
            prod_day_idx = None
            tags_idx = None
            for idx, h in enumerate(headers):
                hl = str(h).strip().lower() if h else ""
                if hl == "productionday":
                    prod_day_idx = idx
                elif hl == "tags":
                    tags_idx = idx

            # Find product columns (those matching a translation name)
            col_map = {}  # col_index -> (product_name, sku)
            for idx, header in enumerate(headers):
                if header and str(header).strip() in name_to_sku:
                    col_map[idx] = (str(header).strip(),
                                    name_to_sku[str(header).strip()])

            # Read all data rows, tagging with production day
            all_rows = []
            production_days = set()
            total_orders = 0
            reship_count = 0
            for row in rows_iter:
                day = ""
                if prod_day_idx is not None and prod_day_idx < len(row):
                    day = str(row[prod_day_idx] or "").strip().upper()
                if day:
                    production_days.add(day)

                # Count orders and reshipped orders
                tags = ""
                if tags_idx is not None and tags_idx < len(row):
                    tags = str(row[tags_idx] or "").lower()
                has_data = any(
                    row[i] for i in col_map if i < len(row) and row[i])
                if has_data:
                    total_orders += 1
                    if "reship" in tags:
                        reship_count += 1

                all_rows.append((day, row))

            wb.close()
        except Exception as e:
            messagebox.showerror("Import Error",
                                 f"Failed to read Excel file:\n{e}",
                                 parent=self.root)
            return

        # Step 3: Filter by ProductionDay if multiple days present
        selected_day = None
        if len(production_days) > 1:
            day_dlg = _ProductionDayDialog(
                self.root, sorted(production_days))
            self.root.wait_window(day_dlg)
            if day_dlg.result is None:
                return
            selected_day = day_dlg.result  # "" means all

        # Sum quantities (with optional day filter)
        depletion_all = defaultdict(int)  # sku -> total qty (all prefixes)
        sku_names = {}
        for day, row in all_rows:
            if selected_day and day != selected_day:
                continue
            for idx, (prod_name, sku) in col_map.items():
                val = row[idx] if idx < len(row) else None
                if val is not None:
                    try:
                        depletion_all[sku] += int(float(val))
                        sku_names[sku] = prod_name
                    except (ValueError, TypeError):
                        pass

        # Separate shippable vs administrative SKUs
        depletion = {}
        admin_skus = {}
        for sku, qty in depletion_all.items():
            if qty <= 0:
                continue
            if sku.startswith(_SHIPPABLE_PREFIXES):
                depletion[sku] = qty
            else:
                admin_skus[sku] = qty

        if not depletion:
            messagebox.showinfo("No Depletion",
                                "No shippable product quantities found.",
                                parent=self.root)
            return

        # Detect unknown SKUs (in depletion but not in inventory at all)
        unknown_skus = [s for s in depletion if s not in self.inventory]

        # Count unmatched product columns
        unmatched_cols = []
        for idx, header in enumerate(headers):
            h = str(header).strip() if header else ""
            if h.startswith("AHB") and h not in name_to_sku:
                unmatched_cols.append(h)

        # Step 4: Show preview dialog
        day_label = selected_day or (
            list(production_days)[0] if len(production_days) == 1 else "ALL")
        dlg = _DepletionPreviewDialog(
            self.root, depletion, sku_names, self.inventory,
            unmatched_cols, admin_skus, unknown_skus, day_label,
            os.path.basename(xlsx_path))
        self.root.wait_window(dlg)

        if not dlg.result:
            return

        # Step 5: Apply depletion with FIFO expiration tracking
        applied = 0
        skipped_missing = []
        for sku, qty in dlg.result.items():
            if sku in self.inventory:
                inv = self.inventory[sku]
                old_qty = float(inv.get("qty", 0))
                inv["qty"] = max(0, old_qty - qty)
                applied += 1

                # FIFO: deplete from earliest expiration batches
                exp_dates = inv.get("expiration_dates", [])
                if exp_dates:
                    remaining = qty
                    new_dates = []
                    for d in exp_dates:  # already sorted earliest first
                        if remaining > 0:
                            remaining -= 1  # each date = 1 batch unit
                        else:
                            new_dates.append(d)
                    inv["expiration_dates"] = new_dates
            else:
                skipped_missing.append(sku)

        # Log to depletion history (including reship stats + warehouses)
        reship_pct = round(reship_count / max(total_orders, 1) * 100, 2)
        history_entry = {
            "date": datetime.datetime.now().isoformat(timespec="seconds"),
            "file": os.path.basename(xlsx_path),
            "day": day_label,
            "skus": dict(dlg.result),
            "warehouses": {sku: self.inventory.get(sku, {}).get(
                "warehouse", "Primary") for sku in dlg.result},
            "total": sum(dlg.result.values()),
            "total_orders": total_orders,
            "reship_count": reship_count,
            "reship_pct": reship_pct,
        }
        self.depletion_history.append(history_entry)

        # Update rolling reship buffer (average of last 10 batches)
        reship_entries = [
            h["reship_pct"] for h in self.depletion_history
            if "reship_pct" in h][-10:]
        if reship_entries:
            self.reship_buffer_pct = round(
                sum(reship_entries) / len(reship_entries), 2)

        self._refresh_inventory_tree()
        self._recalculate()

        msg = f"Depleted {applied} SKUs from inventory."
        if reship_count:
            msg += (f"\n\nReship orders: {reship_count}/{total_orders} "
                    f"({reship_pct}%)")
            msg += f"\nRolling reship buffer: {self.reship_buffer_pct}%"
        if skipped_missing:
            msg += (f"\n\n{len(skipped_missing)} SKUs not in inventory "
                    f"(skipped):\n" +
                    "\n".join(skipped_missing[:10]))
            if len(skipped_missing) > 10:
                msg += f"\n... and {len(skipped_missing) - 10} more"
        self.status_var.set(f"Depletion applied: {applied} SKUs updated")
        messagebox.showinfo("Depletion Complete", msg, parent=self.root)

    def _load_translation_csv(self, path):
        """Load a SKU name translation CSV and save to settings."""
        name_to_sku = {}
        try:
            with open(path, "r", newline="", encoding="utf-8-sig") as f:
                reader = csv.reader(f)
                for row in reader:
                    if len(row) >= 2:
                        sku_code = row[0].strip()
                        product_name = row[1].strip()
                        if sku_code and product_name:
                            name_to_sku[product_name] = sku_code
        except Exception as e:
            messagebox.showerror("Import Error",
                                 f"Failed to read translation CSV:\n{e}",
                                 parent=self.root)
            return {}

        if name_to_sku:
            self.sku_translations = name_to_sku
        return name_to_sku

    def _prompt_translation_csv(self):
        """Prompt the user to select a translation CSV."""
        trans_path = filedialog.askopenfilename(
            title="Select SKU Name Translation CSV",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")])
        if not trans_path:
            return {}
        return self._load_translation_csv(trans_path)

    def _undo_last_depletion(self):
        """Reverse the most recent depletion."""
        if not self.depletion_history:
            messagebox.showinfo("Nothing to Undo",
                                "No depletion history found.",
                                parent=self.root)
            return

        last = self.depletion_history[-1]
        msg = (f"Undo depletion from {last['date'][:10]}?\n"
               f"File: {last['file']}\n"
               f"Day: {last['day']}\n"
               f"Total units: {last['total']}\n"
               f"SKUs affected: {len(last['skus'])}")

        if not messagebox.askyesno("Undo Depletion", msg, parent=self.root):
            return

        # Add quantities back (depletions come from Primary warehouse)
        restored = 0
        warehouses = last.get("warehouses", {})
        for sku, qty in last["skus"].items():
            if sku in self.inventory:
                self.inventory[sku]["qty"] = (
                    float(self.inventory[sku].get("qty", 0)) + qty)
                # Restore warehouse if it was recorded
                if sku in warehouses:
                    self.inventory[sku]["warehouse"] = warehouses[sku]
                restored += 1

        self.depletion_history.pop()
        self._refresh_inventory_tree()
        self._recalculate()

        self.status_var.set(f"Undo complete: restored {restored} SKUs")
        messagebox.showinfo("Undo Complete",
                            f"Restored {restored} SKUs to prior quantities.",
                            parent=self.root)

    def _export_dashboard_csv(self):
        if not self._dash_rows:
            messagebox.showinfo("Nothing to Export",
                                "Run a calculation first.")
            return

        path = filedialog.asksaveasfilename(
            title="Export Dashboard",
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv")])
        if not path:
            return

        try:
            with open(path, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                cols = list(self._dash_cols.keys())
                writer.writerow([self._dash_cols[c][0] for c in cols])
                for r in self._dash_rows:
                    writer.writerow([r.get(c, "") for c in cols])
            messagebox.showinfo("Export Complete",
                                f"Exported {len(self._dash_rows)} rows.")
        except Exception as e:
            messagebox.showerror("Export Error", str(e))

    # ─────────────────────────────────────────────────────────────────
    #  INVENTORY MANAGEMENT
    # ─────────────────────────────────────────────────────────────────

    def _refresh_inventory_tree(self):
        for item in self.inv_tree.get_children():
            self.inv_tree.delete(item)

        today = datetime.date.today()
        try:
            warn_days = int(self.settings_vars.get(
                "expiration_warning_days", tk.StringVar(value="14")).get())
        except (ValueError, AttributeError):
            warn_days = 14

        # Warehouse filter
        wh_filter = (self.inv_warehouse_var.get()
                     if hasattr(self, 'inv_warehouse_var') else
                     "All Locations")

        for sku in sorted(self.inventory.keys()):
            # Apply warehouse filter (supports warehouse_qty splits)
            if wh_filter and wh_filter != "All Locations":
                if self._qty_at(sku, wh_filter) <= 0:
                    continue
            inv = self.inventory[sku]
            exp_dates = inv.get("expiration_dates", [])
            earliest_str = ""
            batch_count = ""
            tag = ()
            if exp_dates:
                batch_count = str(len(exp_dates))
                earliest = exp_dates[0]  # already sorted ISO
                earliest_str = earliest
                try:
                    earliest_dt = datetime.date.fromisoformat(earliest)
                    days_until = (earliest_dt - today).days
                    if days_until < 0:
                        tag = ("EXPIRED",)
                    elif days_until <= warn_days:
                        tag = ("EXPIRING_SOON",)
                except ValueError:
                    pass

            # Show warehouse breakdown if split
            wh_qty = inv.get("warehouse_qty", {})
            if wh_qty and len(wh_qty) > 1:
                wh_label = " / ".join(
                    f"{w}:{round(q,1)}" for w, q in wh_qty.items() if q > 0)
            else:
                wh_label = inv.get("warehouse",
                                   inv.get("location", "Primary"))

            self.inv_tree.insert("", "end", values=(
                sku,
                inv.get("name", ""),
                inv.get("qty", 0),
                inv.get("category", ""),
                wh_label,
                inv.get("unit_cost", ""),
                earliest_str,
                batch_count,
            ), tags=tag)

    def _add_sku_manual(self):
        sku = simpledialog.askstring("Add SKU", "Enter SKU:", parent=self.root)
        if not sku or not sku.strip():
            return
        sku = sku.strip()
        if sku in self.inventory:
            messagebox.showinfo("Exists", f"SKU '{sku}' already exists.",
                                parent=self.root)
            return

        qty_str = simpledialog.askstring(
            "Quantity", f"Quantity on hand for {sku}:", parent=self.root)
        if qty_str is None:
            return
        try:
            qty = float(qty_str)
        except ValueError:
            qty = 0

        name = simpledialog.askstring(
            "Product Name", f"Product name for {sku} (optional):",
            parent=self.root) or ""

        self.inventory[sku] = {"qty": qty, "name": name}
        self._refresh_inventory_tree()
        self.inv_count_var.set(f"{len(self.inventory)} SKUs loaded")

    def _edit_inventory_row(self):
        sel = self.inv_tree.selection()
        if not sel:
            messagebox.showinfo("Select", "Select a row first.",
                                parent=self.root)
            return

        values = self.inv_tree.item(sel[0], "values")
        sku = values[0]

        dlg = _InventoryEditDialog(self.root, sku, self.inventory.get(sku, {}))
        self.root.wait_window(dlg)
        if dlg.result:
            self.inventory[sku] = dlg.result
            self._refresh_inventory_tree()

    def _remove_inventory_row(self):
        sel = self.inv_tree.selection()
        if not sel:
            return
        skus = [self.inv_tree.item(s, "values")[0] for s in sel]
        if not messagebox.askyesno(
                "Confirm", f"Remove {len(skus)} SKU(s) from inventory?",
                parent=self.root):
            return
        for sku in skus:
            self.inventory.pop(sku, None)
        self._refresh_inventory_tree()
        self.inv_count_var.set(f"{len(self.inventory)} SKUs loaded")

    def _on_inv_double_click(self, event):
        self._edit_inventory_row()

    def _on_dash_right_click(self, event):
        """Show context menu on dashboard right-click."""
        item = self.dash_tree.identify_row(event.y)
        if not item:
            return
        self.dash_tree.selection_set(item)
        sku = self.dash_tree.item(item, "values")[0]

        menu = tk.Menu(self.root, tearoff=0, bg=_BG2, fg=_FG,
                       activebackground=_ACC, activeforeground="white")
        if sku in self.archived_skus:
            menu.add_command(label=f"Unarchive {sku}",
                             command=lambda: self._unarchive_sku(sku))
        else:
            menu.add_command(label=f"Archive {sku}",
                             command=lambda: self._archive_sku(sku))
        menu.add_separator()
        menu.add_command(label="Edit SKU Settings",
                         command=lambda: self._open_sku_settings(sku))
        menu.tk_popup(event.x_root, event.y_root)

    def _archive_sku(self, sku):
        self.archived_skus.add(sku)
        self._apply_filter()
        self._save_settings()

    def _unarchive_sku(self, sku):
        self.archived_skus.discard(sku)
        self._apply_filter()
        self._save_settings()

    def _open_sku_settings(self, sku):
        existing = self.sku_settings.get(sku, {})
        dlg = SkuSettingsDialog(self.root, sku, existing)
        self.root.wait_window(dlg)
        if dlg.result is not None:
            if dlg.result:
                self.sku_settings[sku] = dlg.result
            else:
                self.sku_settings.pop(sku, None)
            self._recalculate()

    def _on_dash_double_click(self, event):
        sel = self.dash_tree.selection()
        if not sel:
            return
        sku = self.dash_tree.item(sel[0], "values")[0]
        self._open_sku_settings(sku)

    # ─────────────────────────────────────────────────────────────────
    #  RECHARGE API
    # ─────────────────────────────────────────────────────────────────

    def _toggle_token_visibility(self):
        show = "" if self.show_token_var.get() else "*"
        self.recharge_token_entry.configure(show=show)

    def _recharge_status_text(self):
        n = len(self.recharge_demand)
        if n:
            return f"Last pull: {n} SKUs loaded"
        return "No data pulled yet"

    def _pull_recharge(self):
        token = self.recharge_token_var.get().strip()
        if not token:
            messagebox.showerror("Error",
                                 "Enter a Recharge API token first.",
                                 parent=self.root)
            return

        self.status_var.set("Pulling Recharge subscriptions...")
        self.root.update_idletasks()

        def _worker():
            try:
                client = RechargeClient(token)

                # Stage 1: pull active subscriptions
                def progress(count):
                    self.root.after(0, lambda: self.status_var.set(
                        f"Pulling Recharge... {count} subscriptions"))

                subs = client.get_active_subscriptions(
                    progress_cb=progress)
                sku_qty = client.aggregate_sku_quantities(subs)

                weekly = {}
                for sku, qty in sku_qty.items():
                    weekly[sku] = round(qty, 2)

                # Stage 2: build cohorts from subscription created_at dates
                self.root.after(0, lambda: self.status_var.set(
                    "Building cohorts from subscriptions..."))
                api_cohorts = RechargeClient.build_cohorts_from_subscriptions(
                    subs)

                # Stage 3: pull queued charges
                self.root.after(0, lambda: self.status_var.set(
                    "Pulling Recharge queued charges..."))

                def charge_progress(count):
                    self.root.after(0, lambda: self.status_var.set(
                        f"Pulling Recharge... {count} queued charges"))

                charges = client.get_queued_charges(
                    progress_cb=charge_progress)
                queued_by_month = client.aggregate_charges_by_month(charges)
                queued_resolved = resolve_queued_charges(charges)

                self.root.after(0, lambda: self._on_recharge_done(
                    weekly, queued_by_month, api_cohorts, queued_resolved))
            except Exception as e:
                self.root.after(
                    0, lambda: self._on_recharge_error(str(e)))

        thread = threading.Thread(target=_worker, daemon=True)
        thread.start()

    def _on_recharge_done(self, weekly, queued_by_month=None,
                          api_cohorts=None, queued_resolved=None):
        self.recharge_demand = weekly
        self.recharge_queued = queued_by_month or {}
        self.recharge_queued_resolved = queued_resolved or {}
        # extract monthly box counts (only overwrite if not manually set)
        for bt in MONTHLY_BOX_TYPES:
            if bt in weekly and bt not in self._monthly_box_counts_manual:
                self.monthly_box_counts[bt] = int(weekly[bt])

        # Merge API-derived cohorts (update sizes for existing months,
        # add new months, preserve manually-added track overrides)
        cohort_msg = ""
        if api_cohorts:
            existing = {c["start_month"]: i
                        for i, c in enumerate(self.cohorts)}
            added = 0
            updated = 0
            for ac in api_cohorts:
                month = ac["start_month"]
                if month in existing:
                    self.cohorts[existing[month]]["size"] = ac["size"]
                    updated += 1
                else:
                    self.cohorts.append(ac)
                    added += 1
            cohort_msg = f", cohorts: {updated} updated + {added} new"

        # Resolution stats
        resolve_msg = ""
        if self.recharge_queued_resolved:
            total_resolved = sum(
                sum(v["pr_cjam"].values()) + sum(v["cex_ec"].values())
                for v in self.recharge_queued_resolved.values())
            total_unresolved = sum(
                v["unresolved"]
                for v in self.recharge_queued_resolved.values())
            resolve_msg = (f", {int(total_resolved)} PR-CJAM/CEX-EC resolved"
                          f", {total_unresolved} unresolved")

        queued_months = len(self.recharge_queued)
        self.recharge_status_var.set(
            f"Last pull: {len(weekly)} SKUs + {queued_months} queued month(s)"
            f"{cohort_msg}{resolve_msg}  |  "
            f"{datetime.datetime.now().strftime('%H:%M:%S')}")
        self.status_var.set(
            f"Recharge: {len(weekly)} SKUs, {queued_months} queued month(s)"
            f"{cohort_msg}{resolve_msg}")
        self._recalculate()

    def _on_recharge_error(self, error_msg):
        self.status_var.set("Recharge pull failed")
        messagebox.showerror("Recharge Error",
                             f"Failed to pull subscriptions:\n{error_msg}",
                             parent=self.root)

    # ─────────────────────────────────────────────────────────────────
    #  SHOPIFY API
    # ─────────────────────────────────────────────────────────────────

    def _shopify_status_text(self):
        api_n = len(self.shopify_api_demand)
        manual_n = len(self.shopify_forecast)
        parts = []
        if api_n:
            parts.append(f"API: {api_n} SKUs pulled")
        if manual_n:
            parts.append(f"Manual: {manual_n} override(s)")
        if not parts:
            parts.append("No data — pull orders or add manual overrides")
        return " | ".join(parts)

    def _authorize_shopify(self):
        """Run Shopify OAuth flow to get an access token."""
        store = self.shopify_store_var.get().strip()
        client_id = self.shopify_api_key_var.get().strip()
        client_secret = self.shopify_api_secret_var.get().strip()

        if not store or not client_id or not client_secret:
            messagebox.showerror(
                "Error",
                "Enter Store URL, API Key (Client ID), and "
                "API Secret (Client Secret) first.",
                parent=self.root)
            return

        self.status_var.set(
            "Opening browser for Shopify authorization...")
        self.root.update_idletasks()

        def _worker():
            try:
                oauth = ShopifyOAuth(store, client_id, client_secret)
                token = oauth.authorize()
                if token:
                    self.root.after(
                        0, lambda: self._on_shopify_auth_done(token))
                else:
                    self.root.after(
                        0, lambda: self._on_shopify_auth_error(
                            "No token received — authorization may have "
                            "timed out or been denied."))
            except Exception as e:
                self.root.after(
                    0, lambda: self._on_shopify_auth_error(str(e)))

        thread = threading.Thread(target=_worker, daemon=True)
        thread.start()

    def _on_shopify_auth_done(self, token):
        self.shopify_token_var.set(token)
        self.status_var.set("Shopify authorized successfully")
        self.shopify_status_var.set(
            f"Authorized  |  "
            f"{datetime.datetime.now().strftime('%H:%M:%S')}")
        messagebox.showinfo("Shopify Authorized",
                            "Access token obtained and saved.\n"
                            "You can now Pull Orders.",
                            parent=self.root)

    def _on_shopify_auth_error(self, error_msg):
        self.status_var.set("Shopify authorization failed")
        messagebox.showerror("Shopify Auth Error",
                             f"Authorization failed:\n{error_msg}",
                             parent=self.root)

    def _pull_shopify(self):
        store = self.shopify_store_var.get().strip()
        token = self.shopify_token_var.get().strip()
        if not store or not token:
            messagebox.showerror(
                "Error",
                "Enter Store URL and click 'Authorize Shopify' first "
                "to obtain an access token.",
                parent=self.root)
            return

        tag = self.shopify_tag_var.get().strip()
        try:
            weeks_back = int(self.shopify_weeks_var.get())
            if weeks_back < 1:
                weeks_back = 4
        except ValueError:
            weeks_back = 4

        self.status_var.set("Pulling Shopify orders...")
        self.root.update_idletasks()

        def _worker():
            try:
                client = ShopifyClient(store, token)

                # calculate date range
                min_date = (datetime.datetime.now() -
                            datetime.timedelta(weeks=weeks_back))
                min_date_str = min_date.strftime("%Y-%m-%dT00:00:00-00:00")

                def progress(count):
                    self.root.after(0, lambda: self.status_var.set(
                        f"Pulling Shopify... {count} orders"))

                orders = client.get_orders(
                    tag=tag if tag else None,
                    created_at_min=min_date_str,
                    progress_cb=progress)

                weekly, order_count = client.aggregate_first_order_skus(
                    orders, weeks_back=weeks_back)

                # trend analysis
                trend_data, _, weekly_totals = client.aggregate_with_trend(
                    orders, weeks_back=weeks_back)

                self.root.after(0, lambda: self._on_shopify_done(
                    weekly, order_count, weeks_back, trend_data,
                    weekly_totals))
            except Exception as e:
                self.root.after(
                    0, lambda: self._on_shopify_error(str(e)))

        thread = threading.Thread(target=_worker, daemon=True)
        thread.start()

    def _on_shopify_done(self, weekly, order_count, weeks_back,
                         trend_data=None, weekly_totals=None):
        self.shopify_trend_data = trend_data or {}

        # Use projected values from trend instead of flat averages
        if self.shopify_trend_data:
            projected = {}
            for sku, td in self.shopify_trend_data.items():
                projected[sku] = round(td["projected_next_week"], 2)
            self.shopify_api_demand = projected
        else:
            self.shopify_api_demand = weekly

        # compute overall trend direction from weekly totals
        trend_text = ""
        if weekly_totals and len(weekly_totals) >= 2:
            first_half = sum(weekly_totals[:len(weekly_totals)//2])
            second_half = sum(weekly_totals[len(weekly_totals)//2:])
            if first_half > 0:
                pct_change = (second_half - first_half) / first_half * 100
                arrow = "\u2191" if pct_change >= 0 else "\u2193"
                trend_text = f"  Trend: {arrow}{abs(pct_change):.1f}%"

        self.shopify_status_var.set(
            f"API: {len(self.shopify_api_demand)} SKUs from {order_count} "
            f"orders ({weeks_back}wk){trend_text} | "
            f"Manual: {len(self.shopify_forecast)} override(s) | "
            f"{datetime.datetime.now().strftime('%H:%M:%S')}")
        self.status_var.set(
            f"Shopify: {order_count} orders \u2192 "
            f"{len(self.shopify_api_demand)} SKUs{trend_text}")
        self._recalculate()

    def _on_shopify_error(self, error_msg):
        self.status_var.set("Shopify pull failed")
        messagebox.showerror("Shopify Error",
                             f"Failed to pull orders:\n{error_msg}",
                             parent=self.root)

    # ─────────────────────────────────────────────────────────────────
    #  RETENTION & LIFECYCLE ANALYTICS
    # ─────────────────────────────────────────────────────────────────

    def _analytics_status_text(self):
        parts = []
        if self.actual_retention:
            parts.append(f"Retention: {len(self.actual_retention)} cohorts")
        if self.customer_lifecycle:
            parts.append(f"Lifecycle: {len(self.customer_lifecycle)} customers")
        if self.reship_buffer_pct > 0:
            parts.append(f"Reship buffer: {self.reship_buffer_pct}%")
        return " | ".join(parts) if parts else "No analytics data yet"

    def _pull_retention_data(self):
        """Pull all subscriptions from Recharge and build retention curves."""
        token = self.recharge_token_var.get().strip()
        if not token:
            messagebox.showerror("Error",
                                 "Enter a Recharge API token first.",
                                 parent=self.root)
            return

        self.status_var.set("Pulling all subscriptions for retention analysis...")
        self.root.update_idletasks()

        def _worker():
            try:
                client = RechargeClient(token)

                def progress(count):
                    self.root.after(0, lambda: self.status_var.set(
                        f"Retention analysis... {count} subscriptions"))

                subs = client.get_all_subscriptions(progress_cb=progress)
                cohort_data, overall, segmented = \
                    client.build_retention_curves(subs)

                self.root.after(0, lambda: self._on_retention_done(
                    cohort_data, overall, len(subs), segmented))
            except Exception as e:
                self.root.after(
                    0, lambda: self._on_retention_error(str(e)))

        thread = threading.Thread(target=_worker, daemon=True)
        thread.start()

    def _on_retention_done(self, cohort_data, overall, total_subs,
                           segmented=None):
        self.actual_retention = cohort_data
        self.analytics_status_var.set(self._analytics_status_text())
        self.status_var.set(
            f"Retention: {len(cohort_data)} cohorts from {total_subs} subs")

        # Show comparison dialog with segmentation
        dlg = _RetentionComparisonDialog(
            self.root, cohort_data, overall, self.retention_matrix,
            segmented=segmented)
        self.root.wait_window(dlg)

        if dlg.result:
            # User chose to update the retention matrix
            self.retention_matrix = dlg.result
            self.status_var.set("Retention matrix updated from actual data")

    def _on_retention_error(self, error_msg):
        self.status_var.set("Retention pull failed")
        messagebox.showerror("Retention Error",
                             f"Failed to pull subscriptions:\n{error_msg}",
                             parent=self.root)

    def _pull_customer_lifecycle(self):
        """Pull customer lifecycle data from Shopify."""
        store = self.shopify_store_var.get().strip()
        token = self.shopify_token_var.get().strip()
        if not store or not token:
            messagebox.showerror(
                "Error",
                "Enter Shopify store URL and access token first.",
                parent=self.root)
            return

        self.status_var.set("Pulling customer lifecycle data...")
        self.root.update_idletasks()

        def _worker():
            try:
                client = ShopifyClient(store, token)

                def progress(count):
                    self.root.after(0, lambda: self.status_var.set(
                        f"Lifecycle analysis... {count} orders"))

                lifecycle, cohort_sizes, retention, reship_rate = (
                    client.build_customer_lifecycle(
                        months_back=12, progress_cb=progress))

                self.root.after(0, lambda: self._on_lifecycle_done(
                    lifecycle, cohort_sizes, retention, reship_rate))
            except Exception as e:
                self.root.after(
                    0, lambda: self._on_lifecycle_error(str(e)))

        thread = threading.Thread(target=_worker, daemon=True)
        thread.start()

    def _on_lifecycle_done(self, lifecycle, cohort_sizes, retention,
                           reship_rate):
        self.customer_lifecycle = lifecycle

        # Update reship buffer from Shopify data if we don't have
        # fulfillment-based data yet
        if not any("reship_pct" in h for h in self.depletion_history):
            self.reship_buffer_pct = reship_rate
            if hasattr(self, 'reship_var'):
                self.reship_var.set(str(reship_rate))

        self.analytics_status_var.set(self._analytics_status_text())

        # Show lifecycle summary
        total_customers = len(lifecycle)
        active = sum(1 for c in lifecycle.values()
                     if c["months_active"] > 1)
        avg_months = (sum(c["months_active"] for c in lifecycle.values()) /
                      max(total_customers, 1))

        # Build retention summary from Shopify data
        retention_lines = []
        for month in sorted(cohort_sizes.keys())[-6:]:
            size = cohort_sizes[month]
            ret = retention.get(month, {})
            months_data = [str(ret.get(m, 0)) for m in range(
                min(6, len(ret)))]
            retention_lines.append(
                f"  {month}: {size} new → " + ", ".join(months_data))

        msg = (f"Customers tracked: {total_customers}\n"
               f"Multi-month subscribers: {active}\n"
               f"Avg months active: {avg_months:.1f}\n"
               f"Shopify reship rate: {reship_rate}%\n\n"
               f"Recent cohort retention (active per month):\n" +
               "\n".join(retention_lines))

        self.status_var.set(
            f"Lifecycle: {total_customers} customers, "
            f"reship {reship_rate}%")
        messagebox.showinfo("Customer Lifecycle", msg, parent=self.root)

    def _on_lifecycle_error(self, error_msg):
        self.status_var.set("Lifecycle pull failed")
        messagebox.showerror("Lifecycle Error",
                             f"Failed to pull orders:\n{error_msg}",
                             parent=self.root)

    # ─────────────────────────────────────────────────────────────────
    #  EDITOR DIALOGS
    # ─────────────────────────────────────────────────────────────────

    def _open_bundle_editor(self):
        dlg = BundleMappingEditor(self.root, self.bundle_map)
        self.root.wait_window(dlg)
        if dlg.result is not None:
            self.bundle_map = dlg.result
            self.bundle_count_var.set(
                f"{len(self.bundle_map)} bundle(s) defined")
            self._recalculate()

    def _open_shopify_forecast(self):
        dlg = ShopifyForecastDialog(self.root, self.shopify_forecast)
        self.root.wait_window(dlg)
        if dlg.result is not None:
            self.shopify_forecast = dlg.result
            self.shopify_status_var.set(
                f"{len(self.shopify_forecast)} bundle(s) configured")
            self._recalculate()

    def _open_manual_demand(self):
        dlg = ManualDemandDialog(self.root, self.manual_demand)
        self.root.wait_window(dlg)
        if dlg.result is not None:
            self.manual_demand = dlg.result
            self.manual_status_var.set(
                f"{len(self.manual_demand)} SKU(s) with adjustments")
            self._recalculate()

    # ─────────────────────────────────────────────────────────────────
    #  SAVE / LOAD
    # ─────────────────────────────────────────────────────────────────

    def _save_all_settings(self):
        settings = {
            # API
            "recharge_api_token": self.recharge_token_var.get().strip(),
            "default_churn_pct": self.default_churn_var.get().strip(),

            # global defaults
            "default_purchase_lt": self.settings_vars.get(
                "default_purchase_lt", tk.StringVar()).get(),
            "default_production_lt": self.settings_vars.get(
                "default_production_lt", tk.StringVar()).get(),
            "default_shipping_lt": self.settings_vars.get(
                "default_shipping_lt", tk.StringVar()).get(),
            "default_safety_stock": self.settings_vars.get(
                "default_safety_stock", tk.StringVar()).get(),
            "fulfillment_buffer": self.settings_vars.get(
                "fulfillment_buffer", tk.StringVar()).get(),
            "expiration_warning_days": self.settings_vars.get(
                "expiration_warning_days", tk.StringVar(value="14")).get(),

            # data
            "inventory": self.inventory,
            "sku_settings": self.sku_settings,
            "bundle_map": {
                k: [list(c) for c in v]
                for k, v in self.bundle_map.items()
            },
            "recharge_demand": self.recharge_demand,
            "recharge_queued": self.recharge_queued,
            "recharge_queued_resolved": self.recharge_queued_resolved,
            "shopify_forecast": self.shopify_forecast,
            "shopify_api_demand": self.shopify_api_demand,
            "shopify_trend_data": self.shopify_trend_data,
            "shopify_store_url": self.shopify_store_var.get().strip(),
            "shopify_access_token": self.shopify_token_var.get().strip(),
            "shopify_api_key": self.shopify_api_key_var.get().strip(),
            "shopify_api_secret": self.shopify_api_secret_var.get().strip(),
            "shopify_order_tag": self.shopify_tag_var.get().strip(),
            "shopify_weeks_back": self.shopify_weeks_var.get().strip(),
            "manual_demand": self.manual_demand,
            "last_csv_mapping": self.last_csv_mapping,
            "sku_translations": self.sku_translations,
            "archived_skus": list(self.archived_skus),
            "depletion_history": self.depletion_history,
            "reship_buffer_pct": self._get_reship_pct(),
            "actual_retention": self.actual_retention,
            "customer_lifecycle": self.customer_lifecycle,
            "customization_variance": self.customization_variance,

            # cohort forecasting
            "cohorts": self.cohorts,
            "retention_matrix": self.retention_matrix,
            "churn_rates": self.churn_rates,
            "repeat_rate": self.repeat_rate,
            "curation_recipes": {
                k: [list(c) for c in v]
                for k, v in self.curation_recipes.items()
            },
            "pr_cjam": self.pr_cjam,
            "cex_ec": self.cex_ec,
            "wheel_inventory": self.wheel_inventory,
            "open_pos": self.open_pos,
            "forecast_months": int(self.horizon_var.get()) if hasattr(self, 'horizon_var') else self.forecast_months,

            # monthly curated boxes
            "monthly_box_recipes": self.monthly_box_recipes,
            "monthly_box_counts": self.monthly_box_counts,
            "_monthly_box_counts_manual": self._monthly_box_counts_manual,

            # vendor catalog
            "vendor_catalog": self.vendor_catalog,

            # raw-to-finished conversion & yield tracking
            "bulk_conversions": self.bulk_conversions,
            "production_yield_history": self.production_yield_history,
            "adjusted_conversion_factors": self.adjusted_conversion_factors,

            # EX-EC assignments history
            "exec_assignments": self.exec_assignments,
            "cexec_splits": self._fp_cexec_splits,

            # integrations
            "clickup_api_token": (self.clickup_token_var.get().strip()
                                  if hasattr(self, 'clickup_token_var')
                                  else self.clickup_api_token),
            "clickup_list_id": (self.clickup_list_var.get().strip()
                                if hasattr(self, 'clickup_list_var')
                                else self.clickup_list_id),
            "gcal_refresh_token": self.gcal_refresh_token,
            "gcal_client_id": (self.gcal_id_var.get().strip()
                               if hasattr(self, 'gcal_id_var')
                               else self.gcal_client_id),
            "gcal_client_secret": (self.gcal_secret_var.get().strip()
                                   if hasattr(self, 'gcal_secret_var')
                                   else self.gcal_client_secret),
            "dropbox_refresh_token": self.dropbox_refresh_token,
            "dropbox_app_key": (self.dropbox_key_var.get().strip()
                                if hasattr(self, 'dropbox_key_var')
                                else self.dropbox_app_key),
            "dropbox_app_secret": (self.dropbox_secret_var.get().strip()
                                   if hasattr(self, 'dropbox_secret_var')
                                   else self.dropbox_app_secret),
            "dropbox_shared_link": (self.dropbox_link_var.get().strip()
                                    if hasattr(self, 'dropbox_link_var')
                                    else self.dropbox_shared_link),

            # reconciliation
            "reconciliation_history": self.reconciliation_history,

            # v2.5 automation
            "slack_webhook_url": (self.slack_url_var.get().strip()
                                  if hasattr(self, 'slack_url_var')
                                  else self.slack_webhook_url),
            "slack_notify_critical": (self.slack_crit_var.get()
                                      if hasattr(self, 'slack_crit_var')
                                      else self.slack_notify_critical),
            "slack_notify_expiring": (self.slack_exp_var.get()
                                      if hasattr(self, 'slack_exp_var')
                                      else self.slack_notify_expiring),
            "slack_notify_shortfall": (self.slack_short_var.get()
                                       if hasattr(self, 'slack_short_var')
                                       else self.slack_notify_shortfall),
            "smtp_host": (self.smtp_host_var.get().strip()
                          if hasattr(self, 'smtp_host_var')
                          else self.smtp_host),
            "smtp_port": (self.smtp_port_var.get().strip()
                          if hasattr(self, 'smtp_port_var')
                          else self.smtp_port),
            "smtp_user": (self.smtp_user_var.get().strip()
                          if hasattr(self, 'smtp_user_var')
                          else self.smtp_user),
            "smtp_password": (self.smtp_pass_var.get().strip()
                              if hasattr(self, 'smtp_pass_var')
                              else self.smtp_password),
            "depletion_email_to": (self.email_to_var.get().strip()
                                   if hasattr(self, 'email_to_var')
                                   else self.depletion_email_to),
            "depletion_email_from": (self.email_from_var.get().strip()
                                     if hasattr(self, 'email_from_var')
                                     else self.depletion_email_from),
            "depletion_tag_prefix": (self.depletion_tag_var.get().strip()
                                     if hasattr(self, 'depletion_tag_var')
                                     else self.saved.get(
                                         "depletion_tag_prefix", "RMFG")),
            "auto_refresh_interval": int(
                self.auto_refresh_var.get() or "0")
                if hasattr(self, 'auto_refresh_var') else
                self.auto_refresh_interval,
            "auto_po_threshold": int(
                self.auto_po_var.get() or "0")
                if hasattr(self, 'auto_po_var') else
                self.auto_po_threshold,
            "auto_sync_clickup": (self.auto_clickup_var.get()
                                   if hasattr(self, 'auto_clickup_var')
                                   else self.auto_sync_clickup),
            "auto_sync_gcal": (self.auto_gcal_var.get()
                                if hasattr(self, 'auto_gcal_var')
                                else self.auto_sync_gcal),
            "webhook_port": int(
                self.webhook_port_var.get() or "8765")
                if hasattr(self, 'webhook_port_var') else
                self.webhook_port,
            "webhook_secret_shopify": (
                self.wh_shopify_var.get().strip()
                if hasattr(self, 'wh_shopify_var')
                else self.webhook_secret_shopify),
            "webhook_secret_recharge": (
                self.wh_recharge_var.get().strip()
                if hasattr(self, 'wh_recharge_var')
                else self.webhook_secret_recharge),
            "_last_inventory_import": self.saved.get(
                "_last_inventory_import", ""),

            # v3.0 multi-warehouse & processing
            "warehouses": self.warehouses,
            "transfer_history": self.transfer_history,
            "processing_queue": self.processing_queue,
            "yield_discrepancies": self.yield_discrepancies,
            "yield_reconciliation_window_days": int(
                self.settings_vars.get(
                    "yield_reconciliation_window_days",
                    tk.StringVar(value="3")).get() or "3"),
            "yield_reconciliation_threshold_pct": int(
                self.settings_vars.get(
                    "yield_reconciliation_threshold_pct",
                    tk.StringVar(value="5")).get() or "5"),
            "yield_reconciliation_threshold_min": int(
                self.settings_vars.get(
                    "yield_reconciliation_threshold_min",
                    tk.StringVar(value="2")).get() or "2"),
        }
        # Sync integration vars back to instance attributes
        if hasattr(self, 'clickup_token_var'):
            self.clickup_api_token = self.clickup_token_var.get().strip()
        if hasattr(self, 'clickup_list_var'):
            self.clickup_list_id = self.clickup_list_var.get().strip()
        if hasattr(self, 'gcal_id_var'):
            self.gcal_client_id = self.gcal_id_var.get().strip()
        if hasattr(self, 'gcal_secret_var'):
            self.gcal_client_secret = self.gcal_secret_var.get().strip()
        if hasattr(self, 'dropbox_key_var'):
            self.dropbox_app_key = self.dropbox_key_var.get().strip()
        if hasattr(self, 'dropbox_secret_var'):
            self.dropbox_app_secret = self.dropbox_secret_var.get().strip()
        if hasattr(self, 'dropbox_link_var'):
            self.dropbox_shared_link = self.dropbox_link_var.get().strip()
        # Sync v2.5 automation vars
        if hasattr(self, 'slack_url_var'):
            self.slack_webhook_url = self.slack_url_var.get().strip()
        if hasattr(self, 'slack_crit_var'):
            self.slack_notify_critical = self.slack_crit_var.get()
        if hasattr(self, 'slack_exp_var'):
            self.slack_notify_expiring = self.slack_exp_var.get()
        if hasattr(self, 'slack_short_var'):
            self.slack_notify_shortfall = self.slack_short_var.get()
        if hasattr(self, 'smtp_host_var'):
            self.smtp_host = self.smtp_host_var.get().strip()
        if hasattr(self, 'smtp_port_var'):
            self.smtp_port = self.smtp_port_var.get().strip()
        if hasattr(self, 'smtp_user_var'):
            self.smtp_user = self.smtp_user_var.get().strip()
        if hasattr(self, 'smtp_pass_var'):
            self.smtp_password = self.smtp_pass_var.get().strip()
        if hasattr(self, 'email_to_var'):
            self.depletion_email_to = self.email_to_var.get().strip()
        if hasattr(self, 'email_from_var'):
            self.depletion_email_from = self.email_from_var.get().strip()
        if hasattr(self, 'auto_refresh_var'):
            try:
                self.auto_refresh_interval = int(
                    self.auto_refresh_var.get() or "0")
            except ValueError:
                pass
        if hasattr(self, 'auto_po_var'):
            try:
                self.auto_po_threshold = int(
                    self.auto_po_var.get() or "0")
            except ValueError:
                pass
        if hasattr(self, 'auto_clickup_var'):
            self.auto_sync_clickup = self.auto_clickup_var.get()
        if hasattr(self, 'auto_gcal_var'):
            self.auto_sync_gcal = self.auto_gcal_var.get()
        # Sync yield reconciliation settings
        for key, attr in [
            ("yield_reconciliation_window_days", "yield_recon_window_days"),
            ("yield_reconciliation_threshold_pct", "yield_recon_threshold_pct"),
            ("yield_reconciliation_threshold_min", "yield_recon_threshold_min"),
        ]:
            if key in self.settings_vars:
                try:
                    setattr(self, attr, int(
                        self.settings_vars[key].get() or "0"))
                except ValueError:
                    pass

        save_settings(settings)
        self.saved = settings
        self.status_var.set(
            f"Settings saved  |  "
            f"{datetime.datetime.now().strftime('%H:%M:%S')}")


# ═════════════════════════════════════════════════════════════════════
#  INVENTORY EDIT DIALOG
# ═════════════════════════════════════════════════════════════════════


class _RetentionComparisonDialog(tk.Toplevel):
    """Compare actual retention curves vs. the current modeled matrix."""

    def __init__(self, parent, cohort_data, overall_curve, current_matrix,
                 segmented=None):
        super().__init__(parent)
        self.result = None
        self.title("Retention Calibration — Actual vs. Modeled")
        self.configure(bg=_BG)
        self.transient(parent)
        self.grab_set()

        self.geometry("850x650")
        self.minsize(700, 400)

        # Header
        ttk.Label(self,
                  text=f"Actual retention from {len(cohort_data)} cohorts",
                  style="Subtitle.TLabel").pack(
                      anchor="w", padx=10, pady=(10, 5))

        # Overall curve comparison
        curve_frame = ttk.LabelFrame(self, text="Overall Retention Curve "
                                     "(actual vs. model)", padding=8)
        curve_frame.pack(fill="x", padx=10, pady=(0, 5))

        # Build comparison table
        # Current matrix: sum retention across curations per month
        model_curve = []
        if current_matrix:
            for m in range(7):
                total = sum(
                    vals[m] if m < len(vals) else 0
                    for vals in current_matrix.values())
                model_curve.append(round(total, 1))

        header_row = ttk.Frame(curve_frame)
        header_row.pack(fill="x")
        labels = ["", "Mo 1", "Mo 2", "Mo 3", "Mo 4", "Mo 5",
                  "Mo 6", "Mo 7"]
        for i, lbl in enumerate(labels):
            w = 50 if i == 0 else 65
            anchor = "w" if i == 0 else "center"
            ttk.Label(header_row, text=lbl, width=w // 7,
                      anchor=anchor,
                      font=("Segoe UI", 9, "bold")).pack(
                          side="left", padx=2)

        # Actual row
        actual_row = ttk.Frame(curve_frame)
        actual_row.pack(fill="x")
        ttk.Label(actual_row, text="Actual", width=7,
                  foreground=_ACC).pack(side="left", padx=2)
        for m in range(7):
            val = overall_curve[m] if m < len(overall_curve) else "—"
            ttk.Label(actual_row, text=f"{val}%", width=9,
                      anchor="center").pack(side="left", padx=2)

        # Model row
        model_row = ttk.Frame(curve_frame)
        model_row.pack(fill="x")
        ttk.Label(model_row, text="Model", width=7,
                  foreground="#aaa").pack(side="left", padx=2)
        for m in range(7):
            val = model_curve[m] if m < len(model_curve) else "—"
            ttk.Label(model_row, text=f"{val}%", width=9,
                      anchor="center").pack(side="left", padx=2)

        # Delta row
        delta_row = ttk.Frame(curve_frame)
        delta_row.pack(fill="x")
        ttk.Label(delta_row, text="Delta", width=7).pack(
            side="left", padx=2)
        for m in range(7):
            act = overall_curve[m] if m < len(overall_curve) else 0
            mod = model_curve[m] if m < len(model_curve) else 0
            delta = round(act - mod, 1)
            color = "#66bb6a" if delta >= 0 else "#ef5350"
            sign = "+" if delta > 0 else ""
            lbl = tk.Label(delta_row, text=f"{sign}{delta}%",
                           bg=_BG, fg=color, width=9, anchor="center",
                           font=("Segoe UI", 9))
            lbl.pack(side="left", padx=2)

        # Per-cohort detail
        tree_frame = ttk.LabelFrame(self, text="Per-Cohort Retention (%)",
                                    padding=5)
        tree_frame.pack(fill="both", expand=True, padx=10, pady=5)

        cols = ("cohort", "size", "mo1", "mo2", "mo3", "mo4", "mo5",
                "mo6", "mo7")
        yscroll = ttk.Scrollbar(tree_frame, orient="vertical")
        self.tree = ttk.Treeview(
            tree_frame, columns=cols, show="headings",
            yscrollcommand=yscroll.set, height=10)
        yscroll.configure(command=self.tree.yview)

        col_defs = [
            ("cohort", "Cohort", 80, "w"),
            ("size", "Size", 60, "center"),
            ("mo1", "Mo 1", 60, "center"),
            ("mo2", "Mo 2", 60, "center"),
            ("mo3", "Mo 3", 60, "center"),
            ("mo4", "Mo 4", 60, "center"),
            ("mo5", "Mo 5", 60, "center"),
            ("mo6", "Mo 6", 60, "center"),
            ("mo7", "Mo 7", 60, "center"),
        ]
        for col_id, label, width, anchor in col_defs:
            self.tree.heading(col_id, text=label)
            self.tree.column(col_id, width=width, anchor=anchor)

        for month_label in sorted(cohort_data.keys()):
            cd = cohort_data[month_label]
            ret = cd["retention"]
            vals = [month_label, cd["size"]]
            vals.extend(f"{ret[m]}%" if m < len(ret) else "—"
                        for m in range(7))
            self.tree.insert("", "end", values=tuple(vals))

        self.tree.pack(side="left", fill="both", expand=True)
        yscroll.pack(side="right", fill="y")

        # Segmented retention by cancellation reason
        if segmented:
            seg_frame = ttk.LabelFrame(
                self, text="Retention by Cancellation Reason", padding=5)
            seg_frame.pack(fill="x", padx=10, pady=(5, 0))

            seg_labels = {
                "gift": "Gift",
                "price": "Price",
                "product": "Product Quality",
                "logistics": "Logistics",
                "self_purchase": "Active (self-purchase)",
                "other": "Other",
            }
            # Header
            seg_hdr = ttk.Frame(seg_frame)
            seg_hdr.pack(fill="x")
            cols = ["Category", "Subs", "Mo1", "Mo2", "Mo3",
                    "Mo4", "Mo5", "Mo6", "Mo7"]
            for i, col in enumerate(cols):
                w = 15 if i == 0 else 7
                ttk.Label(seg_hdr, text=col, width=w,
                          anchor="center" if i > 0 else "w",
                          font=("Segoe UI", 8, "bold")).pack(
                              side="left", padx=1)

            for reason_cat in ["self_purchase", "gift", "price",
                               "product", "logistics", "other"]:
                seg_data = segmented.get(reason_cat)
                if not seg_data or seg_data["total_subs"] == 0:
                    continue
                row = ttk.Frame(seg_frame)
                row.pack(fill="x")
                label = seg_labels.get(reason_cat, reason_cat)
                ttk.Label(row, text=label, width=15, anchor="w",
                          font=("Segoe UI", 8)).pack(side="left", padx=1)
                ttk.Label(row, text=str(seg_data["total_subs"]),
                          width=7, anchor="center",
                          font=("Segoe UI", 8)).pack(side="left", padx=1)
                curve = seg_data["overall_curve"]
                for m in range(7):
                    val = f"{curve[m]}%" if m < len(curve) else "—"
                    ttk.Label(row, text=val, width=7, anchor="center",
                              font=("Segoe UI", 8)).pack(
                                  side="left", padx=1)

        # Buttons
        btn_frame = ttk.Frame(self)
        btn_frame.pack(fill="x", padx=10, pady=(5, 10))

        ttk.Label(btn_frame,
                  text="Apply updates the retention matrix with actual data.",
                  style="Dim.TLabel").pack(side="left")
        ttk.Button(btn_frame, text="Apply to Retention Matrix",
                   style="Run.TButton",
                   command=lambda: self._apply(overall_curve,
                                               current_matrix)
                   ).pack(side="right", padx=(5, 0))
        ttk.Button(btn_frame, text="Close",
                   command=self.destroy).pack(side="right")

    def _apply(self, overall_curve, current_matrix):
        """Update the retention matrix using actual data.

        Scales each curation's retention proportionally based on the
        overall actual-vs-model ratio per month.
        """
        if not current_matrix or not overall_curve:
            self.destroy()
            return

        # Calculate per-month scale factors
        new_matrix = {}
        for curation, vals in current_matrix.items():
            new_vals = list(vals)
            # Sum model values per month across curations
            for m in range(min(len(vals), len(overall_curve))):
                model_total = sum(
                    cv[m] if m < len(cv) else 0
                    for cv in current_matrix.values())
                if model_total > 0:
                    actual_total = overall_curve[m]
                    scale = actual_total / model_total
                    new_vals[m] = round(vals[m] * scale, 1)
            new_matrix[curation] = new_vals

        self.result = new_matrix
        self.destroy()


class _ProductionDayDialog(tk.Toplevel):
    """Let user choose which production day to deplete."""

    def __init__(self, parent, days):
        super().__init__(parent)
        self.result = None
        self.title("Select Production Day")
        self.configure(bg=_BG)
        self.transient(parent)
        self.grab_set()
        self.geometry("300x180")
        self.resizable(False, False)

        ttk.Label(self, text="Multiple production days found.\n"
                  "Which day do you want to deplete?",
                  justify="center").pack(padx=10, pady=(15, 10))

        self._var = tk.StringVar(value=days[0])
        for day in days:
            ttk.Radiobutton(self, text=day, variable=self._var,
                            value=day).pack(anchor="w", padx=40)
        ttk.Radiobutton(self, text="All Days", variable=self._var,
                        value="").pack(anchor="w", padx=40)

        btn = ttk.Frame(self)
        btn.pack(fill="x", padx=10, pady=(10, 10))
        ttk.Button(btn, text="OK", style="Run.TButton",
                   command=self._ok).pack(side="right", padx=(5, 0))
        ttk.Button(btn, text="Cancel",
                   command=self.destroy).pack(side="right")

    def _ok(self):
        self.result = self._var.get()
        self.destroy()


class _DepletionPreviewDialog(tk.Toplevel):
    """Preview depletion quantities before applying to inventory."""

    def __init__(self, parent, depletion, sku_names, inventory,
                 unmatched, admin_skus, unknown_skus, day_label, file_name):
        super().__init__(parent)
        self.result = None
        self.depletion = dict(depletion)
        self.title("Depletion Preview")
        self.configure(bg=_BG)
        self.transient(parent)
        self.grab_set()

        self.geometry("820x650")
        self.minsize(700, 500)

        # ── header info ──
        header = ttk.Frame(self)
        header.pack(fill="x", padx=10, pady=(10, 0))
        total_units = sum(depletion.values())
        ttk.Label(header,
                  text=f"File: {file_name}  |  Day: {day_label}",
                  style="Subtitle.TLabel").pack(anchor="w")

        info = ttk.Frame(self)
        info.pack(fill="x", padx=10, pady=(2, 5))
        ttk.Label(info,
                  text=f"{len(depletion)} SKUs  |  "
                       f"{total_units} total units to deplete"
                  ).pack(side="left")

        # Warnings row
        warn_parts = []
        if unmatched:
            warn_parts.append(f"{len(unmatched)} unmatched columns")
        if unknown_skus:
            warn_parts.append(
                f"{len(unknown_skus)} SKUs not in inventory")
        if admin_skus:
            warn_parts.append(
                f"{len(admin_skus)} admin SKUs excluded")
        if warn_parts:
            ttk.Label(info, text="  |  ".join(warn_parts),
                      foreground="#ff9966").pack(side="right")

        # ── production summary by category ──
        cat_totals = defaultdict(int)
        cat_labels = {"MT-": "Meats", "CH-": "Cheese",
                      "AC-": "Accessories", "PK-": "Packaging"}
        for sku, qty in depletion.items():
            prefix = sku.split("-")[0] + "-" if "-" in sku else "Other"
            cat_totals[cat_labels.get(prefix, "Other")] += qty

        summary_frame = ttk.Frame(self)
        summary_frame.pack(fill="x", padx=10, pady=(0, 5))
        summary_parts = [f"{cat}: {qty}" for cat, qty in
                         sorted(cat_totals.items())]
        ttk.Label(summary_frame,
                  text="Pick List Summary:  " + "   |   ".join(summary_parts),
                  foreground=_ACC).pack(anchor="w")

        # ── treeview ──
        tree_frame = ttk.Frame(self)
        tree_frame.pack(fill="both", expand=True, padx=10, pady=(0, 5))

        cols = ("sku", "name", "category", "deplete", "on_hand", "after")
        yscroll = ttk.Scrollbar(tree_frame, orient="vertical")
        self.tree = ttk.Treeview(
            tree_frame, columns=cols, show="headings",
            yscrollcommand=yscroll.set)
        yscroll.configure(command=self.tree.yview)

        col_defs = [
            ("sku",      "SKU",          110, "w"),
            ("name",     "Product Name", 230, "w"),
            ("category", "Category",      85, "center"),
            ("deplete",  "Deplete Qty",   80, "center"),
            ("on_hand",  "On Hand",       75, "center"),
            ("after",    "After",         75, "center"),
        ]
        for col_id, label, width, anchor in col_defs:
            self.tree.heading(col_id, text=label)
            self.tree.column(col_id, width=width, anchor=anchor)

        self.tree.tag_configure("WARNING", background="#7a5500",
                                foreground="white")
        self.tree.tag_configure("NEGATIVE", background="#8b1a1a",
                                foreground="white")
        self.tree.tag_configure("UNKNOWN", background="#4a2a6b",
                                foreground="white")

        for sku in sorted(depletion.keys()):
            qty = depletion[sku]
            name = sku_names.get(sku, "")
            if ": " in name:
                name = name.split(": ", 1)[1]

            prefix = sku.split("-")[0] + "-" if "-" in sku else ""
            category = cat_labels.get(prefix, "Other")

            on_hand = float(inventory.get(sku, {}).get("qty", 0))
            after = on_hand - qty

            tag = ()
            if sku in unknown_skus:
                tag = ("UNKNOWN",)
            elif after < 0:
                tag = ("NEGATIVE",)
            elif after == 0:
                tag = ("WARNING",)

            self.tree.insert("", "end", values=(
                sku, name, category, qty, int(on_hand), int(after)
            ), tags=tag)

        self.tree.pack(side="left", fill="both", expand=True)
        yscroll.pack(side="right", fill="y")

        # ── legend ──
        legend = ttk.Frame(self)
        legend.pack(fill="x", padx=10, pady=(0, 3))
        for color, label in [("#8b1a1a", "Goes negative"),
                             ("#7a5500", "Hits zero"),
                             ("#4a2a6b", "Not in inventory")]:
            lbl = tk.Label(legend, text=f"  {label}  ", bg=color,
                           fg="white", font=("Segoe UI", 8))
            lbl.pack(side="left", padx=(0, 8))

        # ── buttons ──
        btn_frame = ttk.Frame(self)
        btn_frame.pack(fill="x", padx=10, pady=(3, 10))

        ttk.Button(btn_frame, text="Apply Depletion",
                   style="Run.TButton",
                   command=self._apply).pack(side="right", padx=(5, 0))
        ttk.Button(btn_frame, text="Cancel",
                   command=self.destroy).pack(side="right")

    def _apply(self):
        self.result = self.depletion
        self.destroy()


# ═════════════════════════════════════════════════════════════════════

class _InventoryEditDialog(tk.Toplevel):
    """Edit a single inventory entry."""

    def __init__(self, parent, sku, data):
        super().__init__(parent)
        self.title(f"Edit Inventory — {sku}")
        self.configure(bg=_BG)
        self.transient(parent)
        self.grab_set()
        self.result = None
        self._existing_exp = data.get("expiration_dates", [])

        fields = [
            ("Product Name:", "name", data.get("name", "")),
            ("Category:", "category", data.get("category", "")),
            ("Unit Cost:", "unit_cost", data.get("unit_cost", "")),
        ]

        self.vars = {}
        self._wh_qty = data.get("warehouse_qty", {})
        frame = ttk.Frame(self)
        frame.pack(fill="both", expand=True, padx=15, pady=15)

        ttk.Label(frame, text=f"SKU: {sku}", style="Bold.TLabel").grid(
            row=0, column=0, columnspan=2, sticky="w", pady=(0, 10))

        for i, (label, key, default) in enumerate(fields, start=1):
            ttk.Label(frame, text=label).grid(
                row=i, column=0, sticky="w", padx=(0, 10), pady=3)
            var = tk.StringVar(value=str(default))
            ttk.Entry(frame, textvariable=var, width=30).grid(
                row=i, column=1, pady=3, sticky="w")
            self.vars[key] = var

        # Per-warehouse quantities
        next_row = len(fields) + 1
        if self._wh_qty:
            # Show per-warehouse qty fields
            ttk.Label(frame, text="Qty at Primary:").grid(
                row=next_row, column=0, sticky="w", padx=(0, 10), pady=3)
            self.primary_qty_var = tk.StringVar(
                value=str(self._wh_qty.get("Primary", 0)))
            ttk.Entry(frame, textvariable=self.primary_qty_var,
                      width=15).grid(
                row=next_row, column=1, pady=3, sticky="w")
            next_row += 1

            ttk.Label(frame, text="Qty at Woburn:").grid(
                row=next_row, column=0, sticky="w", padx=(0, 10), pady=3)
            self.woburn_qty_var = tk.StringVar(
                value=str(self._wh_qty.get("Woburn", 0)))
            ttk.Entry(frame, textvariable=self.woburn_qty_var,
                      width=15).grid(
                row=next_row, column=1, pady=3, sticky="w")
            next_row += 1
        else:
            # Legacy single-location: total qty + warehouse dropdown
            ttk.Label(frame, text="Quantity On Hand:").grid(
                row=next_row, column=0, sticky="w", padx=(0, 10), pady=3)
            qty_var = tk.StringVar(value=str(data.get("qty", 0)))
            ttk.Entry(frame, textvariable=qty_var, width=15).grid(
                row=next_row, column=1, pady=3, sticky="w")
            self.vars["qty"] = qty_var
            next_row += 1

            ttk.Label(frame, text="Warehouse:").grid(
                row=next_row, column=0, sticky="w", padx=(0, 10), pady=3)
            self.warehouse_var = tk.StringVar(
                value=data.get("warehouse",
                               data.get("location", "Primary")))
            ttk.Combobox(frame, textvariable=self.warehouse_var,
                         values=["Primary", "Woburn"], state="readonly",
                         width=27).grid(
                row=next_row, column=1, pady=3, sticky="w")
            next_row += 1

        # expiration dates field
        ttk.Label(frame, text="Expiration Dates:").grid(
            row=next_row, column=0, sticky="w", padx=(0, 10), pady=3)
        exp_default = ", ".join(self._existing_exp)
        self.exp_var = tk.StringVar(value=exp_default)
        ttk.Entry(frame, textvariable=self.exp_var, width=30).grid(
            row=next_row, column=1, pady=3, sticky="w")
        ttk.Label(frame, text="Comma-separated: MM/DD/YYYY or YYYY-MM-DD",
                  style="Dim.TLabel").grid(
            row=next_row + 1, column=0, columnspan=2,
            sticky="w", pady=(0, 3))

        btn_frame = ttk.Frame(self)
        btn_frame.pack(fill="x", padx=15, pady=(0, 15))
        tk.Button(btn_frame, text="Cancel", command=self.destroy,
                  bg=_BG3, fg=_FG, relief="flat", padx=15, pady=5
                  ).pack(side="right", padx=(5, 0))
        tk.Button(btn_frame, text="Save", command=self._on_save,
                  bg=_GREEN, fg="white", relief="flat", padx=15, pady=5
                  ).pack(side="right")

        self.geometry("400x380")

    def _on_save(self):
        result = {}
        for key, var in self.vars.items():
            val = var.get().strip()
            if key == "qty":
                try:
                    result[key] = float(val.replace(",", ""))
                except ValueError:
                    result[key] = 0
            elif key == "unit_cost":
                val_clean = val.replace("$", "").replace(",", "")
                try:
                    result[key] = float(val_clean) if val_clean else 0
                except ValueError:
                    result[key] = 0
            else:
                result[key] = val
        # Per-warehouse quantities
        if self._wh_qty:
            try:
                p = float(self.primary_qty_var.get().replace(",", ""))
            except (ValueError, AttributeError):
                p = 0
            try:
                w = float(self.woburn_qty_var.get().replace(",", ""))
            except (ValueError, AttributeError):
                w = 0
            wh_qty = {}
            if p > 0:
                wh_qty["Primary"] = p
            if w > 0:
                wh_qty["Woburn"] = w
            result["warehouse_qty"] = wh_qty
            result["qty"] = p + w
            result["warehouse"] = ("Primary" if p >= w else "Woburn") \
                if wh_qty else "Primary"
        elif hasattr(self, 'warehouse_var'):
            result["warehouse"] = self.warehouse_var.get()
        # expiration dates
        parsed = _parse_expiration_dates(self.exp_var.get())
        if parsed:
            result["expiration_dates"] = parsed
        elif self._existing_exp and not self.exp_var.get().strip():
            pass  # user cleared dates, don't preserve old ones
        elif self._existing_exp:
            result["expiration_dates"] = self._existing_exp
        self.result = result
        self.destroy()


# ═════════════════════════════════════════════════════════════════════
#  VENDOR CATALOG DIALOG
# ═════════════════════════════════════════════════════════════════════


class _VendorCatalogDialog(tk.Toplevel):
    """Editable vendor catalog database."""

    def __init__(self, parent, vendor_catalog, inventory):
        super().__init__(parent)
        self.result = None
        self.catalog = dict(vendor_catalog)
        self.title("Vendor Catalog")
        self.configure(bg=_BG)
        self.transient(parent)
        self.grab_set()
        self.geometry("900x500")
        self.minsize(700, 350)

        toolbar = ttk.Frame(self)
        toolbar.pack(fill="x", padx=10, pady=(10, 5))

        ttk.Button(toolbar, text="Add SKU",
                   command=self._add_entry).pack(side="left", padx=(0, 5))
        ttk.Button(toolbar, text="Edit Selected",
                   command=self._edit_entry).pack(side="left", padx=(0, 5))
        ttk.Button(toolbar, text="Remove Selected",
                   command=self._remove_entry).pack(side="left", padx=(0, 5))
        ttk.Button(toolbar, text="Auto-Populate from Inventory",
                   command=lambda: self._auto_populate(inventory)).pack(
                       side="left", padx=(0, 5))

        tree_frame = ttk.Frame(self)
        tree_frame.pack(fill="both", expand=True, padx=10, pady=5)

        cols = ("sku", "vendor", "unit_cost", "case_qty", "moq",
                "wheel_weight")
        self.tree = ttk.Treeview(tree_frame, columns=cols,
                                 show="headings")
        self.tree.heading("sku", text="SKU")
        self.tree.heading("vendor", text="Vendor")
        self.tree.heading("unit_cost", text="Unit Cost")
        self.tree.heading("case_qty", text="Case Qty")
        self.tree.heading("moq", text="MOQ")
        self.tree.heading("wheel_weight", text="Wheel Wt (lbs)")
        self.tree.column("sku", width=120)
        self.tree.column("vendor", width=150)
        self.tree.column("unit_cost", width=80, anchor="center")
        self.tree.column("case_qty", width=70, anchor="center")
        self.tree.column("moq", width=60, anchor="center")
        self.tree.column("wheel_weight", width=100, anchor="center")

        yscroll = ttk.Scrollbar(tree_frame, orient="vertical",
                                command=self.tree.yview)
        self.tree.configure(yscrollcommand=yscroll.set)
        self.tree.pack(side="left", fill="both", expand=True)
        yscroll.pack(side="right", fill="y")

        self._refresh()

        btn_frame = ttk.Frame(self)
        btn_frame.pack(fill="x", padx=10, pady=(5, 10))
        ttk.Button(btn_frame, text="Save", style="Run.TButton",
                   command=self._save).pack(side="right", padx=(5, 0))
        ttk.Button(btn_frame, text="Cancel",
                   command=self.destroy).pack(side="right")

    def _refresh(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
        for sku in sorted(self.catalog.keys()):
            v = self.catalog[sku]
            self.tree.insert("", "end", values=(
                sku, v.get("vendor", ""), v.get("unit_cost", ""),
                v.get("case_qty", ""), v.get("moq", ""),
                v.get("wheel_weight_lbs", "")))

    def _add_entry(self):
        sku = simpledialog.askstring("Add SKU", "Enter SKU:",
                                     parent=self)
        if not sku:
            return
        sku = sku.strip()
        if sku in self.catalog:
            messagebox.showinfo("Exists", f"{sku} already in catalog.",
                                parent=self)
            return
        self.catalog[sku] = {
            "vendor": "", "unit_cost": 0, "case_qty": 1,
            "moq": 0, "wheel_weight_lbs": 0}
        self._refresh()

    def _edit_entry(self):
        sel = self.tree.selection()
        if not sel:
            return
        sku = self.tree.item(sel[0], "values")[0]
        entry = self.catalog.get(sku, {})

        dlg = tk.Toplevel(self)
        dlg.title(f"Edit {sku}")
        dlg.configure(bg=_BG)
        dlg.transient(self)
        dlg.grab_set()
        dlg.geometry("350x250")

        fields = [
            ("Vendor:", "vendor", entry.get("vendor", "")),
            ("Unit Cost:", "unit_cost", str(entry.get("unit_cost", ""))),
            ("Case Qty:", "case_qty", str(entry.get("case_qty", ""))),
            ("MOQ:", "moq", str(entry.get("moq", ""))),
            ("Wheel Weight (lbs):", "wheel_weight_lbs",
             str(entry.get("wheel_weight_lbs", ""))),
        ]
        vars_ = {}
        for i, (label, key, val) in enumerate(fields):
            ttk.Label(dlg, text=label).grid(
                row=i, column=0, sticky="w", padx=10, pady=3)
            var = tk.StringVar(value=val)
            ttk.Entry(dlg, textvariable=var, width=25).grid(
                row=i, column=1, padx=5, pady=3)
            vars_[key] = var

        def _ok():
            self.catalog[sku] = {
                "vendor": vars_["vendor"].get().strip(),
                "unit_cost": float(vars_["unit_cost"].get() or 0),
                "case_qty": int(float(vars_["case_qty"].get() or 1)),
                "moq": int(float(vars_["moq"].get() or 0)),
                "wheel_weight_lbs": float(
                    vars_["wheel_weight_lbs"].get() or 0),
            }
            dlg.destroy()
            self._refresh()

        ttk.Button(dlg, text="OK", style="Run.TButton",
                   command=_ok).grid(row=len(fields), column=1,
                                     sticky="e", padx=10, pady=10)

    def _remove_entry(self):
        sel = self.tree.selection()
        if not sel:
            return
        sku = self.tree.item(sel[0], "values")[0]
        self.catalog.pop(sku, None)
        self._refresh()

    def _auto_populate(self, inventory):
        """Populate vendor catalog from inventory data."""
        added = 0
        for sku, inv in inventory.items():
            if sku not in self.catalog:
                self.catalog[sku] = {
                    "vendor": "",
                    "unit_cost": inv.get("unit_cost", 0),
                    "case_qty": 1,
                    "moq": 0,
                    "wheel_weight_lbs": 0,
                }
                added += 1
        self._refresh()
        messagebox.showinfo("Auto-Populate",
                            f"Added {added} SKUs from inventory.",
                            parent=self)

    def _save(self):
        self.result = self.catalog
        self.destroy()


# ═════════════════════════════════════════════════════════════════════
#  AUTO-PO PREVIEW DIALOG
# ═════════════════════════════════════════════════════════════════════


class _AutoPOPreviewDialog(tk.Toplevel):
    """Preview auto-generated POs grouped by vendor."""

    def __init__(self, parent, po_items, vendor_catalog):
        super().__init__(parent)
        self.result = None
        self.title("Auto-PO Preview")
        self.configure(bg=_BG)
        self.transient(parent)
        self.grab_set()
        self.geometry("900x550")
        self.minsize(700, 400)

        ttk.Label(self, text="Draft Purchase Orders by Vendor",
                  style="Subtitle.TLabel").pack(
                      anchor="w", padx=10, pady=(10, 5))

        # Vendor summary
        summary_frame = ttk.Frame(self)
        summary_frame.pack(fill="x", padx=10, pady=(0, 5))

        total_vendors = len(po_items)
        total_items = sum(len(v) for v in po_items.values())
        total_cost = sum(
            sum(i["total_cost"] for i in items)
            for items in po_items.values())
        ttk.Label(summary_frame,
                  text=f"{total_vendors} vendors  |  "
                       f"{total_items} line items  |  "
                       f"Est. total: ${total_cost:,.2f}",
                  style="Bold.TLabel").pack(side="left")

        # Sub-notebook per vendor
        vnb = ttk.Notebook(self)
        vnb.pack(fill="both", expand=True, padx=10, pady=5)

        for vendor in sorted(po_items.keys()):
            items = po_items[vendor]
            vtab = ttk.Frame(vnb)
            vendor_cost = sum(i["total_cost"] for i in items)
            vnb.add(vtab, text=f"  {vendor} (${vendor_cost:,.0f})  ")

            tree_frame = ttk.Frame(vtab)
            tree_frame.pack(fill="both", expand=True, padx=5, pady=5)

            cols = ("sku", "qty", "wheels", "unit_cost", "total",
                    "month")
            tree = ttk.Treeview(tree_frame, columns=cols,
                                show="headings")
            tree.heading("sku", text="SKU")
            tree.heading("qty", text="Qty")
            tree.heading("wheels", text="Wheels")
            tree.heading("unit_cost", text="Unit Cost")
            tree.heading("total", text="Total Cost")
            tree.heading("month", text="Needed By")
            tree.column("sku", width=120)
            tree.column("qty", width=70, anchor="center")
            tree.column("wheels", width=70, anchor="center")
            tree.column("unit_cost", width=80, anchor="center")
            tree.column("total", width=90, anchor="center")
            tree.column("month", width=80, anchor="center")

            yscroll = ttk.Scrollbar(tree_frame, orient="vertical",
                                    command=tree.yview)
            tree.configure(yscrollcommand=yscroll.set)
            tree.pack(side="left", fill="both", expand=True)
            yscroll.pack(side="right", fill="y")

            for item in items:
                tree.insert("", "end", values=(
                    item["sku"], item["qty"],
                    item["wheel_count"] if item["wheel_count"] else "",
                    f"${item['unit_cost']:.2f}" if item["unit_cost"] else "",
                    f"${item['total_cost']:.2f}" if item["total_cost"] else "",
                    item["month"]))

        btn_frame = ttk.Frame(self)
        btn_frame.pack(fill="x", padx=10, pady=(5, 10))

        def _export_csv():
            path = filedialog.asksaveasfilename(
                title="Export PO",
                defaultextension=".csv",
                filetypes=[("CSV", "*.csv")])
            if not path:
                return
            with open(path, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(["Vendor", "SKU", "Qty", "Wheels",
                                 "Unit Cost", "Total Cost", "Needed By"])
                for vendor, items in po_items.items():
                    for item in items:
                        writer.writerow([
                            vendor, item["sku"], item["qty"],
                            item["wheel_count"], item["unit_cost"],
                            item["total_cost"], item["month"]])
            messagebox.showinfo("Export",
                                "PO exported successfully.",
                                parent=self)

        ttk.Button(btn_frame, text="Export CSV",
                   command=_export_csv).pack(side="left", padx=(0, 5))
        ttk.Button(btn_frame, text="Close",
                   command=self.destroy).pack(side="right")


# ═════════════════════════════════════════════════════════════════════
#  DEPLETION EMAIL COMPOSE DIALOG
# ═════════════════════════════════════════════════════════════════════


class _DepletionEmailDialog(tk.Toplevel):
    """Compose dialog for depletion report email."""

    def __init__(self, parent, default_subject, default_to, from_addr,
                 filename, depletion_entry):
        super().__init__(parent)
        self.result = None
        self.title("Send Depletion Report")
        self.configure(bg=_BG)
        self.transient(parent)
        self.grab_set()
        self.geometry("600x480")
        self.minsize(500, 400)

        total_orders = depletion_entry.get("total_orders", 0)
        total = depletion_entry.get("total", 0)
        day = depletion_entry.get("day", "ALL")
        reship = depletion_entry.get("reship_count", 0)
        reship_pct = depletion_entry.get("reship_pct", 0)

        # From
        row = ttk.Frame(self)
        row.pack(fill="x", padx=15, pady=(15, 3))
        ttk.Label(row, text="From:", width=10, anchor="e").pack(
            side="left", padx=(0, 5))
        ttk.Label(row, text=from_addr, style="Dim.TLabel").pack(
            side="left")

        # To
        row = ttk.Frame(self)
        row.pack(fill="x", padx=15, pady=(0, 3))
        ttk.Label(row, text="To:", width=10, anchor="e").pack(
            side="left", padx=(0, 5))
        self.to_var = tk.StringVar(value=default_to)
        ttk.Entry(row, textvariable=self.to_var, width=55).pack(
            side="left", fill="x", expand=True)

        ttk.Label(self,
                  text="    Separate multiple addresses with commas",
                  style="Dim.TLabel").pack(anchor="w", padx=15)

        # Subject
        row = ttk.Frame(self)
        row.pack(fill="x", padx=15, pady=(5, 3))
        ttk.Label(row, text="Subject:", width=10, anchor="e").pack(
            side="left", padx=(0, 5))
        self.subject_var = tk.StringVar(value=default_subject)
        ttk.Entry(row, textvariable=self.subject_var, width=55).pack(
            side="left", fill="x", expand=True)

        # Attachment info
        row = ttk.Frame(self)
        row.pack(fill="x", padx=15, pady=(5, 3))
        ttk.Label(row, text="Attachment:", width=10, anchor="e").pack(
            side="left", padx=(0, 5))
        ttk.Label(row, text=filename, style="Dim.TLabel").pack(
            side="left")

        # Depletion summary (read-only info)
        info = ttk.LabelFrame(self, text="Depletion Summary", padding=5)
        info.pack(fill="x", padx=15, pady=(5, 3))
        summary = (
            f"Production Day: {day}  |  "
            f"Total Orders: {total_orders}  |  "
            f"Units Depleted: {total}  |  "
            f"Reship: {reship} ({reship_pct}%)")
        ttk.Label(info, text=summary).pack(anchor="w")

        # Notes (editable body)
        ttk.Label(self, text="Notes:",
                  style="Bold.TLabel").pack(
                      anchor="w", padx=15, pady=(8, 3))
        self.body_text = tk.Text(
            self, height=10, bg=_BG3, fg=_FG, insertbackground=_FG,
            font=("Segoe UI", 10), wrap="word", relief="flat",
            padx=8, pady=8)
        self.body_text.pack(fill="both", expand=True, padx=15)

        # Buttons
        btn_frame = ttk.Frame(self)
        btn_frame.pack(fill="x", padx=15, pady=12)
        ttk.Button(btn_frame, text="Send",
                   style="Run.TButton",
                   command=self._send).pack(side="right", padx=(5, 0))
        ttk.Button(btn_frame, text="Cancel",
                   command=self.destroy).pack(side="right")

        # Focus the notes field
        self.body_text.focus_set()

    def _send(self):
        to = self.to_var.get().strip()
        subject = self.subject_var.get().strip()
        body = self.body_text.get("1.0", "end").strip()

        if not to:
            messagebox.showwarning("Missing Recipient",
                                   "Please enter at least one email address.",
                                   parent=self)
            return
        if not subject:
            messagebox.showwarning("Missing Subject",
                                   "Please enter a subject line.",
                                   parent=self)
            return

        self.result = {
            "to": to,
            "subject": subject,
            "body": body,
        }
        self.destroy()


# ═════════════════════════════════════════════════════════════════════
#  ENTRY POINT
# ═════════════════════════════════════════════════════════════════════

def main():
    root = tk.Tk()
    # set icon if available
    try:
        root.iconbitmap(default="")
    except Exception:
        pass
    app = InventoryReorderApp(root)

    # save on close
    def on_close():
        app._save_all_settings()
        root.destroy()

    root.protocol("WM_DELETE_WINDOW", on_close)
    root.mainloop()


if __name__ == "__main__":
    main()
