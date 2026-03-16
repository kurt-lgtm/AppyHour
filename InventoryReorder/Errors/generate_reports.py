"""Generate error order spreadsheets from the error CSV."""
import csv
import re
import json
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC = r"C:\Users\Work\Claude Projects\AppyHour\InventoryReorder\Errors\appyhour-error-orders-2026-03-11.csv"
OUT_DIR = r"C:\Users\Work\Claude Projects\AppyHour\InventoryReorder\Errors"
SETTINGS = r"C:\Users\Work\Claude Projects\AppyHour\InventoryReorder\dist\inventory_reorder_settings.json"

# SKUs that should never appear as CEX-EC resolution (unless paid)
ASSIGNMENT_EXCLUDE = {"CH-MAFT"}


def style_sheet(ws, headers, data_rows, col_widths):
    hdr_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    border = Border(
        bottom=Side(style="thin", color="333333"),
        right=Side(style="thin", color="333333"),
    )
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    for r, row_data in enumerate(data_rows, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for c, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def get_ship_tag(tags):
    m = re.search(r"_SHIP_(\d{4}-\d{2}-\d{2})", tags or "")
    return m.group(1) if m else ""


def get_total(row):
    try:
        return float(row.get("Total", "0").replace(",", ""))
    except ValueError:
        return 0.0


def has_purchased_extras(items):
    return bool(re.search(r"\d+x (CEX-E[A-Z]|EX-E[A-Z])", items))


def main():
    rows_all = []
    with open(SRC, encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            rows_all.append(row)

    # Classify
    extras_orders = []
    error_orders = []

    for row in rows_all:
        triggered = row.get("Rules Triggered", "")
        items = row.get("All Items", "")
        is_overfill = "overfill" in triggered.lower()
        has_extras = has_purchased_extras(items)

        if is_overfill and has_extras:
            extras_orders.append(row)

        has_rule2 = "Rule 2" in triggered
        has_rule3 = "Rule 3" in triggered
        true_overfill = is_overfill and not has_extras

        if has_rule2 or has_rule3 or true_overfill:
            error_orders.append(row)

    # ============================================================
    # Workbook 1: True Errors
    # ============================================================
    wb1 = Workbook()

    # --- Rule 3: Repeats ---
    ws3 = wb1.active
    ws3.title = "Repeats (Rule 3)"
    r3_rows = [r for r in error_orders if "Rule 3" in r.get("Rules Triggered", "")]
    r3_rows.sort(key=lambda r: -int(r.get("Rule 3 Overlap Count", "0") or "0"))

    headers3 = [
        "Order", "Date", "Customer", "Email", "Total", "Overlap Count",
        "Severity", "Previous Order", "Repeated SKUs", "Other Rules", "All Items", "Ship Tag",
    ]
    data3 = []
    for r in r3_rows:
        n = int(r.get("Rule 3 Overlap Count", "0") or "0")
        if n >= 5:
            sev = "CRITICAL"
        elif n >= 3:
            sev = "HIGH"
        elif n >= 2:
            sev = "MEDIUM"
        else:
            sev = "LOW"

        triggered = r.get("Rules Triggered", "")
        other = " | ".join(p.strip() for p in triggered.split("|") if "Rule 3" not in p).strip()

        data3.append([
            r.get("Order", ""), r.get("Date", ""), r.get("Customer", ""),
            r.get("Email", ""), get_total(r), n, sev,
            r.get("Rule 3 Previous Order", ""),
            r.get("Rule 3 Repeated SKUs", ""), other,
            r.get("All Items", ""), get_ship_tag(r.get("Tags", "")),
        ])

    style_sheet(ws3, headers3, data3, [10, 11, 18, 28, 9, 10, 10, 12, 35, 25, 80, 12])

    sev_fills = {
        "CRITICAL": PatternFill(start_color="cc0000", end_color="cc0000", fill_type="solid"),
        "HIGH": PatternFill(start_color="cc6600", end_color="cc6600", fill_type="solid"),
        "MEDIUM": PatternFill(start_color="ccaa00", end_color="ccaa00", fill_type="solid"),
        "LOW": PatternFill(start_color="336633", end_color="336633", fill_type="solid"),
    }
    for r_idx, d in enumerate(data3, 2):
        sev = d[6]
        if sev in sev_fills:
            ws3.cell(row=r_idx, column=7).fill = sev_fills[sev]
            ws3.cell(row=r_idx, column=7).font = Font(color="FFFFFF", bold=True)

    # --- Rule 2: Duplicates ---
    ws2 = wb1.create_sheet("Duplicates (Rule 2)")
    r2_rows = [r for r in error_orders if "Rule 2" in r.get("Rules Triggered", "")]

    headers2 = [
        "Order", "Date", "Customer", "Email", "Total",
        "Duplicate Details", "Other Rules", "All Items", "Ship Tag",
    ]
    data2 = []
    for r in r2_rows:
        triggered = r.get("Rules Triggered", "")
        other = " | ".join(p.strip() for p in triggered.split("|") if "Rule 2" not in p).strip()
        details = r.get("Error Details", "")
        r2_details = " | ".join(p.strip() for p in details.split("|") if "Rule 2" in p)

        data2.append([
            r.get("Order", ""), r.get("Date", ""), r.get("Customer", ""),
            r.get("Email", ""), get_total(r), r2_details, other,
            r.get("All Items", ""), get_ship_tag(r.get("Tags", "")),
        ])

    style_sheet(ws2, headers2, data2, [10, 11, 18, 28, 9, 50, 25, 80, 12])

    # --- True Overfills ---
    ws5 = wb1.create_sheet("True Overfills")
    r5_rows = [
        r for r in error_orders
        if "overfill" in r.get("Rules Triggered", "").lower()
        and not has_purchased_extras(r.get("All Items", ""))
    ]

    headers5 = [
        "Order", "Date", "Customer", "Email", "Total",
        "Overfill Details", "Other Rules", "All Items", "Ship Tag",
    ]
    data5 = []
    for r in r5_rows:
        triggered = r.get("Rules Triggered", "")
        other = " | ".join(
            p.strip() for p in triggered.split("|") if "overfill" not in p.lower()
        ).strip()
        details = r.get("Error Details", "")
        overfill_details = " | ".join(
            p.strip() for p in details.split("|")
            if "overfill" in p.lower() or "Rule 5" in p or "Rule 6" in p
        )

        data5.append([
            r.get("Order", ""), r.get("Date", ""), r.get("Customer", ""),
            r.get("Email", ""), get_total(r), overfill_details, other,
            r.get("All Items", ""), get_ship_tag(r.get("Tags", "")),
        ])

    style_sheet(ws5, headers5, data5, [10, 11, 18, 28, 9, 50, 25, 80, 12])

    # --- Structural Flags ---
    wsf = wb1.create_sheet("Structural Flags")

    headers_f = [
        "Order", "Date", "Customer", "Email", "Total",
        "Flags", "Flag Count", "All Items", "Ship Tag",
    ]
    data_f = []

    for row in rows_all:
        items_str = row.get("All Items", "")
        items = re.findall(r"(\d+)x ([A-Z0-9-]+)", items_str)
        skus = [sku for _, sku in items]
        flags = []

        sku_counts = Counter()
        for q, s in items:
            sku_counts[s] += int(q)

        # Flag: Excluded SKU present without paid extra
        has_paid = any(s.startswith("EX-") for s in skus)
        for ex in ASSIGNMENT_EXCLUDE:
            if sku_counts.get(ex, 0) > 0 and not has_paid:
                flags.append(f"EXCLUDED {ex} x{sku_counts[ex]} (unpaid)")

        # Flag: Any food SKU 3+ times
        for s, c in sku_counts.items():
            if c >= 3 and s.startswith(("CH-", "MT-", "AC-")):
                flags.append(f"3+ COPIES: {s} x{c}")

        # Flag: Missing box SKU
        has_box = any(s.startswith("AHB-") for s in skus)
        if not has_box and len(skus) > 3:
            flags.append("NO BOX SKU")

        # Flag: Missing category
        if has_box:
            has_ch = any(s.startswith("CH-") for s in skus)
            has_mt = any(s.startswith("MT-") for s in skus)
            has_ac = any(s.startswith("AC-") for s in skus)
            is_cmed = "CMED" in items_str
            if not has_ch:
                flags.append("NO CHEESE")
            if not has_mt and not is_cmed:
                flags.append("NO MEAT")
            if not has_ac:
                flags.append("NO ACCOMPANIMENT")

        # Flag: Bare CEX-EC (no suffix)
        has_bare_cexec = "CEX-EC" in skus
        has_cexec_suffix = any(s.startswith("CEX-EC-") for s in skus)
        if has_bare_cexec and not has_cexec_suffix:
            flags.append("BARE CEX-EC (missing cheese)")

        if not flags:
            continue

        data_f.append([
            row.get("Order", ""), row.get("Date", ""), row.get("Customer", ""),
            row.get("Email", ""), get_total(row), " | ".join(flags), len(flags),
            items_str, get_ship_tag(row.get("Tags", "")),
        ])

    # Sort by flag count descending
    data_f.sort(key=lambda r: -r[6])

    style_sheet(wsf, headers_f, data_f, [10, 11, 18, 28, 9, 55, 8, 80, 12])

    # Color rows by severity
    red_fill = PatternFill(start_color="330000", end_color="330000", fill_type="solid")
    orange_fill = PatternFill(start_color="332200", end_color="332200", fill_type="solid")
    for r_idx, d in enumerate(data_f, 2):
        flag_count = d[6]
        if flag_count >= 3:
            for c in range(1, len(headers_f) + 1):
                wsf.cell(row=r_idx, column=c).fill = red_fill
        elif flag_count >= 2:
            for c in range(1, len(headers_f) + 1):
                wsf.cell(row=r_idx, column=c).fill = orange_fill

    # --- Summary sheet ---
    wss = wb1.create_sheet("Summary")
    wb1.move_sheet("Summary", offset=-4)

    repeat_skus = Counter()
    for r in r3_rows:
        for s in r.get("Rule 3 Repeated SKUs", "").split(","):
            s = s.strip()
            if s:
                repeat_skus[s] += 1

    summary = [
        ["Error Order Analysis", "", "", ""],
        ["Report Date", "2026-03-11", "", ""],
        ["Ship Date", "2026-03-16 (Saturday)", "", ""],
        [],
        ["Rule", "Orders", "Revenue", "Notes"],
        ["Repeats (Rule 3)", len(r3_rows),
         sum(d[4] for d in data3), "Items customer already received"],
        ["  Critical (5+)", sum(1 for d in data3 if d[6] == "CRITICAL"), "", ""],
        ["  High (3-4)", sum(1 for d in data3 if d[6] == "HIGH"), "", ""],
        ["  Medium (2)", sum(1 for d in data3 if d[6] == "MEDIUM"), "", ""],
        ["  Low (1)", sum(1 for d in data3 if d[6] == "LOW"), "", ""],
        ["Duplicates (Rule 2)", len(r2_rows),
         sum(d[4] for d in data2), "Same item appears twice"],
        ["True Overfills", len(r5_rows),
         sum(d[4] for d in data5), "Too many items, no purchased extras"],
        [],
        ["Total True Errors", len(error_orders),
         sum(get_total(r) for r in error_orders), ""],
        [],
        ["Top Repeated SKUs", "Count", "", ""],
    ]
    for sku, cnt in repeat_skus.most_common(10):
        summary.append([sku, cnt, "", ""])

    for r_idx, row_data in enumerate(summary, 1):
        for c, val in enumerate(row_data, 1):
            if val == "":
                continue
            cell = wss.cell(row=r_idx, column=c, value=val)
            if r_idx == 1:
                cell.font = Font(bold=True, size=14)
            elif r_idx == 5 or r_idx == 16:
                cell.font = Font(bold=True, size=11)
            if c == 3 and isinstance(val, (int, float)) and val > 0:
                cell.number_format = "$#,##0.00"

    wss.column_dimensions["A"].width = 25
    wss.column_dimensions["B"].width = 12
    wss.column_dimensions["C"].width = 14
    wss.column_dimensions["D"].width = 40

    path1 = OUT_DIR + r"\error-orders-true-errors-2026-03-11.xlsx"
    wb1.save(path1)
    print(f"Saved: {path1}")
    print(f"  Repeats: {len(r3_rows)} | Duplicates: {len(r2_rows)} | True Overfills: {len(r5_rows)}")

    # ============================================================
    # Workbook 2: Overfills with purchased extras
    # ============================================================
    wb2 = Workbook()
    wse = wb2.active
    wse.title = "Overfill with Extras"

    headers_e = [
        "Order", "Date", "Customer", "Email", "Total", "Box Type",
        "Extras Purchased", "Food Items", "Error Details", "All Items", "Ship Tag",
    ]
    data_e = []
    for r in extras_orders:
        items = r.get("All Items", "")
        total = get_total(r)
        ship = get_ship_tag(r.get("Tags", ""))

        box_match = re.search(r"AHB-[A-Z]+-[A-Z]+", items)
        box_type = box_match.group(0) if box_match else ""

        extras_list = re.findall(r"(\d+)x (CEX-E[A-Z]+|EX-E[A-Z]+)", items)
        extras_str = ", ".join(f"{q}x {sku}" for q, sku in extras_list)

        food_count = len(re.findall(r"\d+x (CH-|MT-|AC-|CEX-E|EX-E)", items))

        data_e.append([
            r.get("Order", ""), r.get("Date", ""), r.get("Customer", ""),
            r.get("Email", ""), total, box_type, extras_str, food_count,
            r.get("Error Details", ""), items, ship,
        ])

    style_sheet(wse, headers_e, data_e, [10, 11, 18, 28, 9, 22, 30, 10, 50, 80, 12])

    # Summary tab
    wss2 = wb2.create_sheet("Summary")
    wb2.move_sheet("Summary", offset=-1)

    extras_by_type = Counter()
    for r in extras_orders:
        items = r.get("All Items", "")
        for _, sku in re.findall(r"(\d+)x (CEX-E[A-Z]+|EX-E[A-Z]+)", items):
            extras_by_type[sku] += 1

    s2 = [
        ["Overfill Orders with Purchased Extras", "", ""],
        ["These are NOT errors - customers paid for add-ons", "", ""],
        [],
        ["Total Orders", len(extras_orders),
         sum(get_total(r) for r in extras_orders)],
        [],
        ["Extra Type", "Orders", ""],
    ]
    for sku, cnt in extras_by_type.most_common():
        s2.append([sku, cnt, ""])

    for r_idx, row_data in enumerate(s2, 1):
        for c, val in enumerate(row_data, 1):
            if val == "":
                continue
            cell = wss2.cell(row=r_idx, column=c, value=val)
            if r_idx == 1:
                cell.font = Font(bold=True, size=14)
            elif r_idx == 6:
                cell.font = Font(bold=True, size=11)
            if c == 3 and isinstance(val, (int, float)) and val > 0:
                cell.number_format = "$#,##0.00"

    wss2.column_dimensions["A"].width = 40
    wss2.column_dimensions["B"].width = 12
    wss2.column_dimensions["C"].width = 14

    path2 = OUT_DIR + r"\overfill-with-extras-2026-03-11.xlsx"
    wb2.save(path2)
    print(f"Saved: {path2}")
    print(f"  Orders with purchased extras: {len(extras_orders)}")


if __name__ == "__main__":
    main()
