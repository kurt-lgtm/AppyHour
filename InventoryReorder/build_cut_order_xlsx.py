#!/usr/bin/env python
"""
Cut Order Excel — single sheet, clean layout.

Columns: SKU | Name | Avail | Demand Wk1 | After Wk1 | Cut Wk1 (input) | Good? |
         Demand Wk2 | After Wk2 | Cut Wk2 (input) | Good?

PR-CJAM and CEX-EC assignment tables on the right side of the same sheet.
Demand includes SUMIF for PR-CJAM/CEX-EC — change cheese SKU, demand updates.
No borders. No zero rows.
"""

import argparse
import io
import json
import os
import sys
import time
from collections import defaultdict

BASE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, BASE)
from inventory_demand_report import (
    load_settings, load_inventory_csv, parse_depletion_xlsx,
    fetch_recharge_api, fetch_shopify_orders, PICKABLE_PREFIXES,
    WK1_START, WK1_END, WK2_START, WK2_END,
    INV_CSV, SAT_DEPLETION, TUE_DEPLETION,
)


def main():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.formatting.rule import CellIsRule

    settings = load_settings()
    sku_translations = settings.get("sku_translations", {})
    recharge_token = settings.get("recharge_api_token", "")
    pr_cjam_cfg = settings.get("pr_cjam", {})
    cex_ec_cfg = settings.get("cex_ec", {})
    inv_settings = settings.get("inventory", {})

    def sku_name(sku):
        data = inv_settings.get(sku, {})
        return data.get("name", "") if isinstance(data, dict) else ""

    # -- Load data --
    print("Loading inventory...")
    # Try standard inventory CSV first; fall back to template check format
    try:
        inventory = load_inventory_csv(INV_CSV)
    except (KeyError, ValueError):
        inventory = {}
    # Template check fallback: "a" column = SKU, "Available M/DD" column = qty
    if not inventory:
        import csv as _csv
        with open(INV_CSV, encoding="utf-8-sig") as _f:
            reader = _csv.reader(_f)
            hdr = next(reader)
            # Find the "Available" column (partial match)
            avail_col = next((i for i, h in enumerate(hdr) if "available" in (h or "").lower()), 4)
            for row in reader:
                sku = (row[0] if row else "").strip()
                if sku and sku.startswith(PICKABLE_PREFIXES):
                    try:
                        inventory[sku] = int(float(row[avail_col] or 0))
                    except (ValueError, IndexError):
                        pass
        print(f"  Loaded {len(inventory)} SKUs from template check (col {avail_col})")

    print("Parsing depletions...")
    sat_dep = {}
    tue_dep = {}
    if SAT_DEPLETION:
        sat_dep, _, _, _, _ = parse_depletion_xlsx(SAT_DEPLETION, sku_translations)
    if TUE_DEPLETION:
        tue_dep, _, _, _, _ = parse_depletion_xlsx(TUE_DEPLETION, sku_translations)

    available = {}
    all_skus = set(inventory.keys()) | set(sat_dep.keys()) | set(tue_dep.keys())
    for sku in all_skus:
        available[sku] = inventory.get(sku, 0) - sat_dep.get(sku, 0) - tue_dep.get(sku, 0)

    print("Fetching Recharge charges...")
    (rc_wk1, rc_wk2, rc_wk1_curations, rc_wk2_curations,
     rc_wk1_large, rc_wk2_large, _, _,
     rc_wk1_med_monthly, rc_wk2_med_monthly,
     rc_wk1_cmed_monthly, rc_wk2_cmed_monthly,
     rc_wk1_lge_monthly, rc_wk2_lge_monthly) = fetch_recharge_api(recharge_token)

    # Diagnostic: WK1 vs WK2 Recharge charge counts
    rc_wk1_total = sum(rc_wk1.values())
    rc_wk2_total = sum(rc_wk2.values())
    rc_wk1_curs = sum(rc_wk1_curations.values())
    rc_wk2_curs = sum(rc_wk2_curations.values())
    print(f"  Recharge WK1: {rc_wk1_total} pickable SKUs, {rc_wk1_curs} curation charges")
    print(f"  Recharge WK2: {rc_wk2_total} pickable SKUs, {rc_wk2_curs} curation charges")

    print("Fetching Shopify orders...")
    (sh_wk1_addon, sh_wk2_addon,
     sh_wk1_curations, sh_wk2_curations,
     sh_wk1_large, sh_wk2_large,
     sh_wk1_med, sh_wk2_med,
     sh_wk1_lge, sh_wk2_lge) = fetch_shopify_orders(settings)

    # -- First-order projection (MONG) --
    # Count "Subscription First Order" tagged orders from last 3 days, project forward
    import requests as _req
    from datetime import datetime, timedelta

    store_url = settings.get("shopify_store_url", "")
    shop_token = settings.get("shopify_access_token", "")
    if store_url and shop_token:
        if not store_url.startswith("http"):
            store_url = f"https://{store_url}.myshopify.com"
        _cutoff = (datetime.now() - timedelta(days=3)).isoformat()
        _fo_url = f"{store_url}/admin/api/2024-01/orders.json"
        _fo_params = {"status": "any", "limit": 250, "created_at_min": _cutoff,
                      "fields": "id,tags,line_items"}
        _fo_orders = []
        _fo_page_url = _fo_url
        while _fo_page_url:
            _fo_resp = _req.get(_fo_page_url, headers={
                "X-Shopify-Access-Token": shop_token,
                "Content-Type": "application/json",
            }, params=_fo_params if _fo_page_url == _fo_url else None, timeout=30)
            _fo_data = _fo_resp.json()
            for _o in _fo_data.get("orders", []):
                if "Subscription First Order" in (_o.get("tags") or ""):
                    _fo_orders.append(_o)
            _fo_page_url = None
            _link = _fo_resp.headers.get("Link", "")
            if 'rel="next"' in _link:
                import re as _re
                _m = _re.search(r'<([^>]+)>;\s*rel="next"', _link)
                if _m:
                    _fo_page_url = _m.group(1)
            time.sleep(0.3)

        _daily_rate = len(_fo_orders) / 3.0 if _fo_orders else 0
        # Project to Friday (ship cutoff): days remaining from now
        _days_to_friday = max(0, (5 - datetime.now().weekday()))  # 0=Mon, 4=Fri
        _projected = int(_daily_rate * _days_to_friday)

        # Build per-SKU profile from MONG first orders
        _mong_fo = [o for o in _fo_orders if any(
            "MONG" in (li.get("sku") or "") for li in o.get("line_items", []))]
        _fo_skus = defaultdict(float)
        if _mong_fo:
            for _o in _mong_fo:
                for _li in _o.get("line_items", []):
                    _sku = (_li.get("sku") or "").strip()
                    if _sku.startswith(PICKABLE_PREFIXES):
                        _fo_skus[_sku] += (_li.get("quantity", 1)) / len(_mong_fo)

        # Add projected demand to WK1 Shopify addon counts
        _mong_pct = len(_mong_fo) / len(_fo_orders) if _fo_orders else 0
        _mong_projected = int(_projected * _mong_pct)
        print(f"  First-order projection: {len(_fo_orders)} in 3d, "
              f"{_daily_rate:.0f}/day, {_projected} projected, "
              f"{_mong_projected} MONG ({_mong_pct:.0%})")
        for _sku, _rate in _fo_skus.items():
            _add = int(_rate * _mong_projected)
            if _add > 0:
                sh_wk1_addon[_sku] = sh_wk1_addon.get(_sku, 0) + _add

        # Add projected MONG first orders to curation counts too —
        # each first order gets a PR-CJAM + (some fraction) CEX-EC
        if _mong_projected > 0:
            # Compute large box ratio BEFORE adding projection
            _sh_mong_lg = sh_wk1_large.get("MONG", 0)
            _sh_mong_total = sh_wk1_curations.get("MONG", 1)
            _lg_ratio = _sh_mong_lg / _sh_mong_total if _sh_mong_total > 0 else 0.0
            _proj_lg = int(_mong_projected * _lg_ratio)
            sh_wk1_curations["MONG"] = sh_wk1_curations.get("MONG", 0) + _mong_projected
            sh_wk1_large["MONG"] = sh_wk1_large.get("MONG", 0) + _proj_lg

    # Merge curation counts: Recharge + Shopify
    wk1_curations = defaultdict(int)
    wk2_curations = defaultdict(int)
    wk1_large = defaultdict(int)
    wk2_large = defaultdict(int)
    for d_rc, d_sh, d_out in [
        (rc_wk1_curations, sh_wk1_curations, wk1_curations),
        (rc_wk2_curations, sh_wk2_curations, wk2_curations),
        (rc_wk1_large, sh_wk1_large, wk1_large),
        (rc_wk2_large, sh_wk2_large, wk2_large),
    ]:
        for k, v in d_rc.items():
            d_out[k] += v
        for k, v in d_sh.items():
            d_out[k] += v
    wk1_curations = dict(wk1_curations)
    wk2_curations = dict(wk2_curations)
    wk1_large = dict(wk1_large)
    wk2_large = dict(wk2_large)

    # Box size counts per curation (Recharge + Shopify)
    wk1_med = defaultdict(int, sh_wk1_med)
    wk2_med = defaultdict(int, sh_wk2_med)
    wk1_lge = defaultdict(int, sh_wk1_lge)
    wk2_lge = defaultdict(int, sh_wk2_lge)
    for cur, ct in rc_wk1_curations.items():
        lg = rc_wk1_large.get(cur, 0)
        wk1_med[cur] += ct - lg
        wk1_lge[cur] += lg
    for cur, ct in rc_wk2_curations.items():
        lg = rc_wk2_large.get(cur, 0)
        wk2_med[cur] += ct - lg
        wk2_lge[cur] += lg

    # MONTHLY boxes — Recharge counts (Shopify MONTHLY/CMED already in sh_wk*_med)
    wk1_med["MONTHLY"] = wk1_med.get("MONTHLY", 0) + rc_wk1_med_monthly
    wk2_med["MONTHLY"] = wk2_med.get("MONTHLY", 0) + rc_wk2_med_monthly
    wk1_med["CMED"] = wk1_med.get("CMED", 0) + rc_wk1_cmed_monthly
    wk2_med["CMED"] = wk2_med.get("CMED", 0) + rc_wk2_cmed_monthly
    wk1_lge["MONTHLY"] = wk1_lge.get("MONTHLY", 0) + rc_wk1_lge_monthly
    wk2_lge["MONTHLY"] = wk2_lge.get("MONTHLY", 0) + rc_wk2_lge_monthly

    wk1_med = dict(wk1_med)
    wk2_med = dict(wk2_med)
    wk1_lge = dict(wk1_lge)
    wk2_lge = dict(wk2_lge)

    # Filter to active pickable SKUs only
    report_skus = set()
    for d in (available, rc_wk1, rc_wk2, sh_wk1_addon, sh_wk2_addon):
        report_skus.update(d.keys())

    active_skus = sorted(
        sku for sku in report_skus
        if any(sku.startswith(p) for p in PICKABLE_PREFIXES)
        and (available.get(sku, 0) != 0
             or rc_wk1.get(sku, 0) > 0 or rc_wk2.get(sku, 0) > 0
             or sh_wk1_addon.get(sku, 0) > 0 or sh_wk2_addon.get(sku, 0) > 0)
    )

    all_curations = sorted(
        set(list(wk1_curations.keys()) + list(wk2_curations.keys())
            + list(wk1_large.keys()) + list(wk2_large.keys())
            + list(wk1_med.keys()) + list(wk2_med.keys())
            + list(wk1_lge.keys()) + list(wk2_lge.keys()))
    )

    # -- Build Excel --
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cut Order"

    from openpyxl.utils import get_column_letter

    # Styles
    hdr_font = Font(name="Space Mono", bold=True, size=11, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="2B2B2B")
    sku_font = Font(name="Space Mono", size=10)
    name_font = Font(name="DM Sans", size=10)
    num_font = Font(name="Rajdhani", size=12)
    bold_num = Font(name="Rajdhani", size=12, bold=True)
    input_font = Font(name="Rajdhani", size=12, bold=True, color="0000CC")
    input_fill = PatternFill("solid", fgColor="E8E8FF")
    short_fill = PatternFill("solid", fgColor="FFCCCC")
    ok_fill = PatternFill("solid", fgColor="CCFFCC")
    tight_fill = PatternFill("solid", fgColor="FFFFCC")
    section_font = Font(name="Space Mono", bold=True, size=10, color="00AA00")
    edit_font = Font(name="DM Sans", size=10, bold=True, color="0000CC")
    center = Alignment(horizontal="center")

    # ====== PR-CJAM table (columns N-Q) ======
    prcjam_col_start = 14  # N
    prcjam_headers = ["Curation", "PR-CJAM Cheese", "Wk1 Ct", "Wk2 Ct"]
    for ci, h in enumerate(prcjam_headers):
        c = ws.cell(row=1, column=prcjam_col_start + ci, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = center

    for i, cur in enumerate(all_curations):
        row = i + 2
        cheese = ""
        cfg = pr_cjam_cfg.get(cur)
        if isinstance(cfg, dict):
            cheese = cfg.get("cheese", "")
        ws.cell(row=row, column=prcjam_col_start, value=cur).font = name_font
        ws.cell(row=row, column=prcjam_col_start + 1, value=cheese).font = edit_font
        ws.cell(row=row, column=prcjam_col_start + 2, value=wk1_curations.get(cur, 0)).font = num_font
        ws.cell(row=row, column=prcjam_col_start + 3, value=wk2_curations.get(cur, 0)).font = num_font

    # ====== CEX-EC table (columns S-V) ======
    cexec_col_start = 19  # S
    cexec_headers = ["Curation", "CEX-EC Cheese", "Wk1 Ct", "Wk2 Ct"]
    for ci, h in enumerate(cexec_headers):
        c = ws.cell(row=1, column=cexec_col_start + ci, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = center

    for i, cur in enumerate(all_curations):
        row = i + 2
        cheese = cex_ec_cfg.get(cur, "")
        ws.cell(row=row, column=cexec_col_start, value=cur).font = name_font
        ws.cell(row=row, column=cexec_col_start + 1, value=cheese).font = edit_font
        ws.cell(row=row, column=cexec_col_start + 2, value=wk1_large.get(cur, 0)).font = num_font
        ws.cell(row=row, column=cexec_col_start + 3, value=wk2_large.get(cur, 0)).font = num_font

    # ====== MONTHLY box slot assignment tables (below curation tables) ======
    # MONTHLY boxes (plain AHB-MED/AHB-LGE) have their pickable items excluded
    # from rc_wk1/rc_wk2 — ALL demand flows through these editable slot tables.
    # User assigns specific SKUs to each slot; SUMIF picks them up in demand.

    # Slot definitions — must match MONTHLY_BOX_SLOTS in fulfillment_web/app.py
    AHB_MED_SLOTS = [
        ("Cheese 1", "CH-"), ("Cheese 2", "CH-"),
        ("Meat 1", "MT-"), ("Meat 2", "MT-"),
        ("Crackers", "AC-"),
        ("Accompaniment 1", "AC-"), ("Accompaniment 2", "AC-"),
        ("PR-CJAM-GEN Cheese", "CH-"), ("PR-CJAM-GEN Jam", "AC-"),
    ]
    AHB_CMED_SLOTS = [
        ("Cheese 1", "CH-"), ("Cheese 2", "CH-"),
        ("Cheese 3", "CH-"), ("Cheese 4", "CH-"),
        ("Crackers", "AC-"),
        ("Accompaniment 1", "AC-"), ("Accompaniment 2", "AC-"),
        ("PR-CJAM-GEN Cheese", "CH-"), ("PR-CJAM-GEN Jam", "AC-"),
    ]
    AHB_LGE_SLOTS = [
        ("Cheese 1", "CH-"), ("Cheese 2", "CH-"), ("Cheese 3", "CH-"),
        ("Meat 1", "MT-"), ("Meat 2", "MT-"), ("Meat 3", "MT-"),
        ("Crackers", "AC-"),
        ("Accompaniment 1", "AC-"), ("Accompaniment 2", "AC-"),
        ("PR-CJAM-GEN Cheese", "CH-"), ("PR-CJAM-GEN Jam", "AC-"),
    ]

    monthly_med_w1 = wk1_med.get("MONTHLY", 0)
    monthly_med_w2 = wk2_med.get("MONTHLY", 0)
    monthly_lge_w1 = wk1_lge.get("MONTHLY", 0)
    monthly_lge_w2 = wk2_lge.get("MONTHLY", 0)

    box_start_row = len(all_curations) + 4  # 2 rows gap

    # Helper to write a slot table section
    def _write_slot_table(ws, start_row, label, slots, w1_count, w2_count):
        ws.cell(row=start_row, column=prcjam_col_start,
                value=f"{label} ({w1_count} Wk1 / {w2_count} Wk2)").font = section_font
        r = start_row
        for slot_name, _prefix in slots:
            r += 1
            ws.cell(row=r, column=prcjam_col_start, value=slot_name).font = name_font
            ws.cell(row=r, column=prcjam_col_start + 1).font = edit_font
            ws.cell(row=r, column=prcjam_col_start + 1).fill = input_fill
            ws.cell(row=r, column=prcjam_col_start + 2, value=w1_count).font = num_font
            ws.cell(row=r, column=prcjam_col_start + 3, value=w2_count).font = num_font
        return r

    # -- AHB-MED slot table --
    slot_row = _write_slot_table(ws, box_start_row, "AHB-MED",
                                  AHB_MED_SLOTS, monthly_med_w1, monthly_med_w2)

    # -- AHB-CMED slot table --
    # CMED boxes resolve to MONTHLY via _MONTHLY_PATTERNS, count alongside MED
    # but have a different recipe (4 cheese, no meat). Need separate counts.
    monthly_cmed_w1 = wk1_med.get("CMED", 0)
    monthly_cmed_w2 = wk2_med.get("CMED", 0)
    slot_row = _write_slot_table(ws, slot_row + 2, "AHB-CMED",
                                  AHB_CMED_SLOTS, monthly_cmed_w1, monthly_cmed_w2)

    # -- AHB-LGE slot table --
    slot_row = _write_slot_table(ws, slot_row + 2, "AHB-LGE",
                                  AHB_LGE_SLOTS, monthly_lge_w1, monthly_lge_w2)

    last_slot_row = slot_row

    # SUMIF references — extend range to include all MONTHLY slot assignment rows
    sumif_last = last_slot_row
    # PR-CJAM cheese col = O, wk1 count = P, wk2 count = Q
    prcjam_cheese = f"$O$2:$O${sumif_last}"
    prcjam_w1 = f"$P$2:$P${sumif_last}"
    prcjam_w2 = f"$Q$2:$Q${sumif_last}"
    # CEX-EC cheese col = T, wk1 count = U, wk2 count = V
    cexec_cheese = f"$T$2:$T${sumif_last}"
    cexec_w1 = f"$U$2:$U${sumif_last}"
    cexec_w2 = f"$V$2:$V${sumif_last}"

    # ====== Main cut order table (columns A-L) ======
    # A:SKU  B:Name  C:Avail  D:Demand Wk1  E:After Wk1  F:Cut Wk1  G:Good? Wk1
    # H:Demand Wk2  I:After Wk2  J:Cut Wk2  K:Good? Wk2
    main_headers = [
        ("SKU", 14),
        ("Name", 36),
        ("Avail", 9),
        ("Demand Wk1", 11),
        ("After Wk1", 10),
        ("Cut Wk1", 9),
        ("Good?", 9),
        ("", 2),  # spacer
        ("Demand Wk2", 11),
        ("After Wk2", 10),
        ("Cut Wk2", 9),
        ("Good?", 9),
    ]

    for ci, (h, w) in enumerate(main_headers):
        col = ci + 1
        c = ws.cell(row=1, column=col, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = center
        ws.column_dimensions[chr(64 + col) if col <= 26 else ""].width = w

    # Set column widths properly
    for ci, (_, w) in enumerate(main_headers):
        ws.column_dimensions[get_column_letter(ci + 1)].width = w

    # Assignment table column widths
    ws.column_dimensions[get_column_letter(prcjam_col_start)].width = 10
    ws.column_dimensions[get_column_letter(prcjam_col_start + 1)].width = 16
    ws.column_dimensions[get_column_letter(prcjam_col_start + 2)].width = 8
    ws.column_dimensions[get_column_letter(prcjam_col_start + 3)].width = 8
    ws.column_dimensions[get_column_letter(cexec_col_start)].width = 10
    ws.column_dimensions[get_column_letter(cexec_col_start + 1)].width = 16
    ws.column_dimensions[get_column_letter(cexec_col_start + 2)].width = 8
    ws.column_dimensions[get_column_letter(cexec_col_start + 3)].width = 8

    ws.freeze_panes = "A2"

    # Category separators + data rows
    current_prefix = None
    cat_labels = {"CH-": "CHEESE", "MT-": "MEAT", "AC-": "ACCOMPANIMENTS"}
    row = 1

    for sku in active_skus:
        prefix = sku[:3]
        if prefix != current_prefix:
            current_prefix = prefix
            row += 1
            label = cat_labels.get(prefix, prefix)
            ws.cell(row=row, column=1, value=label).font = section_font
            ws.cell(row=row, column=2).font = section_font

        row += 1
        avail = available.get(sku, 0)
        rc1 = rc_wk1.get(sku, 0)
        sh1 = sh_wk1_addon.get(sku, 0)
        rc2 = rc_wk2.get(sku, 0)
        sh2 = sh_wk2_addon.get(sku, 0)

        # A: SKU
        ws.cell(row=row, column=1, value=sku).font = sku_font
        # B: Name
        ws.cell(row=row, column=2, value=sku_name(sku)).font = name_font
        # C: Available
        ws.cell(row=row, column=3, value=avail).font = num_font

        # D: Demand Wk1 = RC_direct + SH_addon + SUMIF(PRCJAM) + SUMIF(CEXEC)
        ws[f"D{row}"] = (
            f'={rc1}+{sh1}'
            f'+SUMIF({prcjam_cheese},A{row},{prcjam_w1})'
            f'+SUMIF({cexec_cheese},A{row},{cexec_w1})'
        )
        ws.cell(row=row, column=4).font = bold_num

        # E: After Wk1 = Avail - Demand
        ws[f"E{row}"] = f'=C{row}-D{row}'
        ws.cell(row=row, column=5).font = bold_num

        # F: Cut Wk1 (user input)
        ws.cell(row=row, column=6).font = input_font
        ws.cell(row=row, column=6).fill = input_fill

        # G: Good? Wk1 = After Wk1 + Cut Wk1
        ws[f"G{row}"] = f'=IF(D{row}=0,"",IF(E{row}+F{row}>=0,"OK","NEED "&ABS(E{row}+F{row})))'
        ws.cell(row=row, column=7).font = Font(name="Space Mono", size=9, bold=True)
        ws.cell(row=row, column=7).alignment = center

        # H: spacer (empty)

        # I: Demand Wk2 = RC_direct + SH_addon + SUMIF(PRCJAM) + SUMIF(CEXEC)
        ws[f"I{row}"] = (
            f'={rc2}+{sh2}'
            f'+SUMIF({prcjam_cheese},A{row},{prcjam_w2})'
            f'+SUMIF({cexec_cheese},A{row},{cexec_w2})'
        )
        ws.cell(row=row, column=9).font = bold_num

        # J: After Wk2 = (After Wk1 + Cut Wk1) - Demand Wk2
        ws[f"J{row}"] = f'=(E{row}+F{row})-I{row}'
        ws.cell(row=row, column=10).font = bold_num

        # K: Cut Wk2 (user input)
        ws.cell(row=row, column=11).font = input_font
        ws.cell(row=row, column=11).fill = input_fill

        # L: Good? Wk2
        ws[f"L{row}"] = f'=IF(I{row}=0,"",IF(J{row}+K{row}>=0,"OK","NEED "&ABS(J{row}+K{row})))'
        ws.cell(row=row, column=12).font = Font(name="Space Mono", size=9, bold=True)
        ws.cell(row=row, column=12).alignment = center

    last_row = row

    # Conditional formatting on After columns and Good columns
    ws.conditional_formatting.add(
        f"E2:E{last_row}",
        CellIsRule(operator="lessThan", formula=["0"], fill=short_fill))
    ws.conditional_formatting.add(
        f"E2:E{last_row}",
        CellIsRule(operator="greaterThanOrEqual", formula=["0"], fill=ok_fill))
    ws.conditional_formatting.add(
        f"J2:J{last_row}",
        CellIsRule(operator="lessThan", formula=["0"], fill=short_fill))
    ws.conditional_formatting.add(
        f"J2:J{last_row}",
        CellIsRule(operator="greaterThanOrEqual", formula=["0"], fill=ok_fill))

    # Good? columns: green for OK, red for NEED
    from openpyxl.formatting.rule import FormulaRule
    ws.conditional_formatting.add(
        f"G2:G{last_row}",
        FormulaRule(formula=[f'G2="OK"'], fill=ok_fill))
    ws.conditional_formatting.add(
        f"G2:G{last_row}",
        FormulaRule(formula=[f'LEFT(G2,4)="NEED"'], fill=short_fill))
    ws.conditional_formatting.add(
        f"L2:L{last_row}",
        FormulaRule(formula=[f'L2="OK"'], fill=ok_fill))
    ws.conditional_formatting.add(
        f"L2:L{last_row}",
        FormulaRule(formula=[f'LEFT(L2,4)="NEED"'], fill=short_fill))

    # -- Save --
    ship_date = WK1_END.isoformat()
    out_path = os.path.join(BASE, f"cut_order_{ship_date}.xlsx")
    wb.save(out_path)
    print(f"\nExcel written to: {out_path}")
    print(f"  {len(active_skus)} active SKUs (zeroes removed)")
    print(f"  PR-CJAM + CEX-EC tables on same sheet (columns N-V)")
    print(f"  Blue columns = your input (Cut Wk1, Cut Wk2)")
    print(f"  Change PR-CJAM/CEX-EC cheese SKU -> demand auto-updates")

    return out_path


# -- Upload to Google Drive --

DRIVE_FOLDER_ID = "1TgvxK10tFAPJqhkYw-6u1Umnvp9wMJ3I"
DRIVE_TOKEN_PATH = os.path.join(BASE, "dist", "drive_oauth_token.json")


def upload_to_drive(file_path):
    """Upload XLSX to shared Google Drive folder using OAuth credentials."""
    from google.auth.transport.requests import Request
    from google.oauth2.credentials import Credentials
    from googleapiclient.discovery import build
    from googleapiclient.http import MediaIoBaseUpload

    with open(DRIVE_TOKEN_PATH, encoding="utf-8") as f:
        token_data = json.load(f)

    creds = Credentials(
        token=token_data["token"],
        refresh_token=token_data["refresh_token"],
        token_uri=token_data["token_uri"],
        client_id=token_data["client_id"],
        client_secret=token_data["client_secret"],
        scopes=token_data["scopes"],
    )
    if creds.expired:
        creds.refresh(Request())
        token_data["token"] = creds.token
        with open(DRIVE_TOKEN_PATH, "w") as f:
            json.dump(token_data, f, indent=2)

    drive = build("drive", "v3", credentials=creds)

    file_name = os.path.basename(file_path)
    with open(file_path, "rb") as fh:
        media = MediaIoBaseUpload(
            io.BytesIO(fh.read()),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            resumable=True,
        )
        f = drive.files().create(
            body={"name": file_name, "parents": [DRIVE_FOLDER_ID]},
            media_body=media,
            fields="id, webViewLink",
            supportsAllDrives=True,
        ).execute()

    link = f.get("webViewLink", f"https://drive.google.com/file/d/{f['id']}")
    print(f"  Uploaded: {link}")
    return link


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Generate cut order XLSX")
    parser.add_argument("--local", action="store_true", help="Generate locally, don't upload")
    args = parser.parse_args()

    out_path = main()
    if not args.local:
        try:
            upload_to_drive(out_path)
        except Exception as e:
            print(f"  Upload failed: {e}")
            print(f"  File saved locally: {out_path}")
