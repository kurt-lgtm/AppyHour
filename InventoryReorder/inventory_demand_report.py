#!/usr/bin/env python

# /// script
# requires-python = ">=3.10"
# dependencies = ["openpyxl", "requests"]
# ///

"""
Inventory & Demand Report -- Weeks of 3/21 and 3/28
===================================================
1. Load 3/14 inventory snapshot (Total column)
2. Subtract depletions: 03-16 Sat + 03-17 Tue shipments
3. Week 1 demand (_SHIP_2026-03-23): Recharge queued charges + first orders
4. Week 2 demand (_SHIP_2026-03-30): Recharge queued charges + first orders
5. Output: SKU | Available | Wk1 | After Wk1 | Wk2
6. PR-CJAM and CEX-EC totals at bottom

Recharge API: MUST use X-Recharge-Version: 2021-11 for cursor pagination.
Without it, v1 silently caps at 250 results with no next_cursor.
"""

import csv
import json
import os
import sys
import time
from collections import defaultdict
from datetime import date, datetime, timedelta

# Canonical Shopify line-item netting (shopify-line-items rule) — single source
# of truth for "NEVER count removed/refunded items" across ALL builds.
_MCP_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "AppyHourMCP")
if _MCP_DIR not in sys.path:
    sys.path.insert(0, _MCP_DIR)
from utils import active_line_items  # noqa: E402

# -- Paths --
BASE = os.path.dirname(os.path.abspath(__file__))
SETTINGS_PATH = os.path.join(BASE, "dist", "inventory_reorder_settings.json")
INV_CSV = r"C:\Users\Work\Claude Projects\AppyHour\InventoryReorder\Product Inventory_2026-06-23_HAVE.csv"
SHIPMENTS = os.path.join(BASE, "Shipments")
SAT_DEPLETION = ""  # HAVE inventory CSV is current (Kurt 6/22) — post-depletion
TUE_DEPLETION = ""

# Ship week boundaries — DYNAMIC (no more weekly hand-editing; see Cut Order Rules.md
# "make everything dynamic"). Single-cohort: the builder drops WK2 and emits only WK1,
# so the Recharge scheduled_at window must cover the whole cut in WK1:
#   from today → the UPCOMING SATURDAY (the ship-week billing boundary — Kurt's rule:
#   "the recharge pull should be up to Saturday, it's always Saturday").
# This is correct for both Monday (this week's Tue-ship remainder is still queued +
# next week) and Tuesday runs (this week already processed out of queued). Shopify
# Mon/Tue tag handling is separate, via _wk1_ship_tags(today).
# Override by hardcoding these four if a run ever needs a custom window.
def _coming_saturday(d: date) -> date:
    return d + timedelta(days=(5 - d.weekday()) % 7)  # 5 = Saturday; today if already Sat

_TODAY = date.today()
# Recharge bills a week BEFORE ship (Kurt 2026-07-07: "we charge the orders a week before
# we actually ship them"). So the CHARGE window is the ship week's PRECEDING Sunday→Saturday:
# it ENDS on the upcoming Saturday ("always Saturday" — the billing boundary) and starts the
# Sunday 6 days earlier. The ship week it serves starts the Monday AFTER that Saturday
# (WK1_END + 2), used ONLY as the display label — NOT to anchor the window (anchoring on the
# ship week pulls a week too far forward = next week's cohort; that inflated AHB-MED 242→375).
WK1_END = _coming_saturday(_TODAY)                  # charge-week Saturday — the billing boundary
WK1_START = WK1_END - timedelta(days=6)             # Sunday of the charge week (Sun→Sat)
_SHIP_WEEK_MONDAY = WK1_END + timedelta(days=2)     # Monday of the SERVED ship week (label only)
WK2_START = WK1_END + timedelta(days=1)     # WK2 degenerate (builder drops it)
WK2_END = WK1_END                           # also caps the API scheduled_at_max pull


def _assert_window_invariants():
    """Hard gate on the charge-week window (Cut Order Rules.md invariant, 2026-07-07).

    The 7/13 cut was built with the window anchored on the SHIP week instead of the
    CHARGE week — one week forward — and over-pulled AHB-MED 242->375. The window is
    now computed correctly by construction above, but these values are documented as
    hand-overridable; this assert makes a bad override die loudly instead of shipping
    a wrong cut. Bypass for a deliberate custom window: AH_CUT_WINDOW_OVERRIDE=1.
    """
    if os.environ.get("AH_CUT_WINDOW_OVERRIDE") == "1":
        print("⚠️  AH_CUT_WINDOW_OVERRIDE=1 — charge-week window invariants NOT enforced")
        return
    problems = []
    if WK1_END.weekday() != 5:
        problems.append(f"WK1_END {WK1_END} is not a Saturday")
    if WK1_END != _coming_saturday(_TODAY):
        problems.append(f"WK1_END {WK1_END} != coming Saturday {_coming_saturday(_TODAY)} "
                        "(window must end at the NEXT billing boundary from today, "
                        "never anchored on the ship week)")
    if WK1_START != WK1_END - timedelta(days=6):
        problems.append(f"WK1_START {WK1_START} != WK1_END-6 (window must be Sun→Sat)")
    if _SHIP_WEEK_MONDAY != WK1_END + timedelta(days=2) or _SHIP_WEEK_MONDAY.weekday() != 0:
        problems.append(f"ship-week label {_SHIP_WEEK_MONDAY} != WK1_END+2 Monday")
    if problems:
        raise SystemExit(
            "CUT-ORDER WINDOW INVARIANT VIOLATION — refusing to build a wrong cut:\n  "
            + "\n  ".join(problems)
            + "\n(see Cut Order Rules.md 'Charge-week window'; deliberate custom window: "
            "set AH_CUT_WINDOW_OVERRIDE=1)")


_assert_window_invariants()

PICKABLE_PREFIXES = ("CH-", "MT-", "AC-", "MR-", "PK-")

# ── Bundle / limited-release recipes ────────────────────────────────
# A "bundle" (BL-*) or limited-release box (AHB-X*) ships several component
# SKUs. The recipe (component -> qty per box) lives in Shopify's Simple Bundles
# variant metafield `simple_bundles.bundled_variants`. Recharge-native bundles
# (no Simple Bundles metafield) are supplied via BUNDLE_RECIPE_OVERRIDES.
#
# DOUBLE-COUNT RULE (Kurt 2026-07-28): a component is added to demand ONLY when
# it is NOT already a line item in the SAME container (order/charge). Shopify
# orders — and some Recharge charges (e.g. BL-BLR4 charge 1816643580) — already
# carry the components as their own lines, which the pickable loop counts; the
# explosion adds only the components that are genuinely missing. Never re-add
# what a container already lists.
BUNDLE_RECIPE_OVERRIDES = {
    # Recharge-native bundle, absent from Simple Bundles. Confirmed live from
    # Recharge charge 1816643580 + Kurt 2026-07-28: 4 Baked Lemon Ricotta +
    # 2 Blackberry Balsamic jam.
    "BL-BLR4": {"CH-BLR": 4, "AC-BLBALS": 2},
}
_BUNDLE_RECIPE_CACHE: dict = {}


def _is_bundle_sku(sku: str) -> bool:
    u = (sku or "").upper()
    return u.startswith("BL-") or u.startswith("AHB-X")


def _get_bundle_recipe(store: str, token: str, sku: str, _seen=None) -> dict:
    """Return {component_sku: qty_per_box} for a bundle/limited-release SKU.

    Source: Shopify Simple Bundles variant metafield, then overrides. Nested
    bundles (a component that is itself BL-/AHB-X) are expanded recursively.
    Cached per run. Returns {} when no recipe is found (caller must not fabricate).
    """
    import requests

    key = (sku or "").upper()
    if key in _BUNDLE_RECIPE_CACHE:
        return _BUNDLE_RECIPE_CACHE[key]
    _seen = _seen or set()
    if key in _seen:  # cycle guard for nested bundles
        return {}
    _seen.add(key)

    raw = dict(BUNDLE_RECIPE_OVERRIDES.get(key, {}))
    if not raw and store and token:
        try:
            gql = f"{store}/admin/api/2026-04/graphql.json"
            q = ('query($q:String!){productVariants(first:1,query:$q){edges{node{'
                 'metafield(namespace:"simple_bundles",key:"bundled_variants"){value}}}}}')
            r = requests.post(
                gql,
                headers={"X-Shopify-Access-Token": token, "Content-Type": "application/json"},
                json={"query": q, "variables": {"q": f"sku:{sku}"}},
                timeout=30,
            )
            edges = r.json().get("data", {}).get("productVariants", {}).get("edges", [])
            mf = edges[0]["node"].get("metafield") if edges else None
            if mf:
                for c in json.loads(mf["value"]):
                    cs = c.get("sku")
                    if cs:
                        raw[cs] = raw.get(cs, 0) + int(c.get("quantity_in_bundle") or 1)
        except Exception as e:
            print(f"  WARN: bundle recipe fetch failed for {sku}: {e}")

    recipe: dict = {}
    for cs, qty in raw.items():
        if _is_bundle_sku(cs):  # nested bundle → expand
            for ss, sq in _get_bundle_recipe(store, token, cs, _seen).items():
                recipe[ss] = recipe.get(ss, 0) + sq * qty
        else:
            recipe[cs] = recipe.get(cs, 0) + qty
    _BUNDLE_RECIPE_CACHE[key] = recipe
    return recipe


def _explode_bundles(present_qty: dict, store: str, token: str,
                     target_demand: dict, out_bundles, week_key: str) -> None:
    """Add missing bundle components to demand for one container.

    present_qty: {sku: qty} of line items actually in this order/charge.
    For each bundle parent present, add each recipe component × parent_qty to
    target_demand ONLY if that component is absent from present_qty (else it's
    already counted). Records per-bundle breakdown into out_bundles for display.
    """
    present = set(present_qty)
    for sku in list(present):
        if not _is_bundle_sku(sku):
            continue
        n = present_qty.get(sku, 1)
        recipe = _get_bundle_recipe(store, token, sku)
        rec = None
        if out_bundles is not None:
            rec = out_bundles.setdefault(week_key, {}).setdefault(
                sku, {"count": 0, "components": {}, "no_recipe": not recipe})
            rec["count"] += n
        for cs, per in recipe.items():
            pickable = any(cs.startswith(p) for p in PICKABLE_PREFIXES)
            already = cs in present
            if not already and pickable:
                target_demand[cs] = target_demand.get(cs, 0) + n * per
            if rec is not None:
                comp = rec["components"].setdefault(
                    cs, {"per": per, "added": 0, "already": 0, "pickable": pickable})
                if already:
                    comp["already"] += n * per
                elif pickable:
                    comp["added"] += n * per

# Curation resolution
KNOWN_CURATIONS = {
    "MONG",
    "MDT",
    "OWC",
    "SPN",
    "ALPN",
    "ALPT",
    "ISUN",
    "HHIGH",
    "NMS",
    "BYO",
    "SS",
    "GEN",
    "MS",
}
_MONTHLY_PATTERNS = {"AHB-MED", "AHB-LGE", "AHB-CMED"}


def resolve_curation(sku):
    if not sku:
        return None
    sku = sku.strip().upper()
    if sku in _MONTHLY_PATTERNS:
        return "MONTHLY"
    if "-MCUST-NMS" in sku:
        return "NMS"
    if "-MCUST-MS" in sku or "-CUR-MS" in sku:
        return "MS"
    # Hyphen-anchored match — prevents substring false-positives (e.g. MS in NMS, SS in HHIGH).
    # Iterate by length desc so longer curations win when nested.
    for cur in sorted(KNOWN_CURATIONS, key=len, reverse=True):
        if sku.endswith("-" + cur) or ("-" + cur + "-") in sku:
            return cur
    return None


def is_large_box(sku):
    """True if the box is a LARGE size. Used for MED/CMED/LGE box-total bucketing ONLY.

    NOTE: box size does NOT imply CEX-EC. CEX-EC rides on medium custom boxes too
    (e.g. AHB-MCUST-MDT) and is absent from many large boxes. Use cexec_curation()
    for CEX-EC counting, never this. (Kurt 2026-06-26)
    """
    s = (sku or "").strip().upper()
    return s.startswith("AHB-L") or s.startswith("AHB-LCUST")


def cexec_curation(line_items, box_curation):
    """Return the curation a CEX-EC extra-cheese pick belongs to, or None if the
    order carries no CEX-EC line item.

    CEX-EC resolves ONLY when an actual CEX-EC* line item is present — never from
    box size (Kurt feedback 2026-06-26). Resolution:
      - suffixed CEX-EC-<CUR> (e.g. created post-charge in Shopify) → that curation
      - bare CEX-EC (the queued-recharge case) → the box's curation
    Guards against CEX-EM / CEX-EA (extra meat / accompaniment), which are not cheese.
    """
    suffix_cur = None
    has_cex = False
    for it in line_items:
        sku = (it.get("sku") or "").strip().upper()
        if sku != "CEX-EC" and not sku.startswith("CEX-EC-"):
            continue
        has_cex = True
        if sku.startswith("CEX-EC-"):
            for tok in sku[len("CEX-EC-"):].split("-"):
                if tok in KNOWN_CURATIONS:
                    suffix_cur = tok
    if not has_cex:
        return None
    return suffix_cur or box_curation


def load_settings():
    with open(SETTINGS_PATH, "r") as f:
        return json.load(f)


# -- Step 1: Load inventory --


def load_inventory_csv(path):
    inv = {}
    with open(path, encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            sku = (row.get("Product SKU") or "").strip()
            if not sku:
                continue
            # Prefer RMFG column (fulfillment warehouse) over Total (which includes CA)
            raw = row.get("RMFG") or row.get("Total", "0") or "0"
            raw = raw.strip() if isinstance(raw, str) else str(raw)
            try:
                inv[sku] = int(float(raw)) if raw else 0
            except (ValueError, TypeError):
                inv[sku] = 0
    return inv


# -- Step 2: Parse depletion XLSX --


def parse_depletion_xlsx(path, sku_translations):
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    product_cols = []
    for i, h in enumerate(headers):
        if h and "AHB (S_REG):" in str(h):
            name = str(h).split(": ", 1)[1].strip() if ": " in str(h) else str(h)
            product_cols.append((i, name))

    tags_idx = None
    for i, h in enumerate(headers):
        if h and str(h).strip().lower() == "tags":
            tags_idx = i
            break

    totals = defaultdict(int)
    first_order_totals = defaultdict(int)
    order_count = 0
    first_order_count = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        order_count += 1
        is_first = False
        if tags_idx is not None:
            tags_val = row[tags_idx] if tags_idx < len(row) else None
            if tags_val and "Subscription First Order" in str(tags_val):
                is_first = True
                first_order_count += 1

        for idx, name in product_cols:
            val = row[idx] if idx < len(row) else None
            if val and isinstance(val, (int, float)) and val > 0:
                totals[name] += int(val)
                if is_first:
                    first_order_totals[name] += int(val)

    wb.close()

    sku_totals = defaultdict(int)
    first_sku_totals = defaultdict(int)
    unmatched = []

    for product, qty in totals.items():
        if "tasting guide" in product.lower():
            continue
        sku = sku_translations.get(product)
        if sku:
            sku_totals[sku] += qty
        else:
            unmatched.append((product, qty))

    for product, qty in first_order_totals.items():
        if "tasting guide" in product.lower():
            continue
        sku = sku_translations.get(product)
        if sku:
            first_sku_totals[sku] += qty

    return dict(sku_totals), order_count, first_order_count, dict(first_sku_totals), unmatched


# -- Step 3: Fetch Recharge via API (v2021-11) --


def _load_recharge_csv(path):
    """Load queued-charges CSV export into API-shape charge dicts.

    Expected CSV header: charge_id,customer_id,scheduled_at,email,line_item_title,
                          line_item_quantity,line_item_sku
    """
    import csv as _csv
    from datetime import datetime as _dt
    charges_by_id = {}
    with open(path, encoding="utf-8-sig", newline="") as f:
        r = _csv.reader(f)
        next(r)
        for row in r:
            if not row or not (row[0] or "").strip().isdigit():
                continue
            cid = row[0].strip()
            sched_raw = (row[2] or "").strip() if len(row) > 2 else ""
            sku = (row[6] or "").strip() if len(row) > 6 else ""
            try:
                qty = int(float(row[5])) if len(row) > 5 and row[5] else 1
            except ValueError:
                qty = 1
            # parse "5/20/2026 0:00" -> ISO
            sched_iso = ""
            try:
                sched_iso = _dt.strptime(sched_raw.split(" ")[0], "%m/%d/%Y").date().isoformat()
            except Exception:
                try:
                    sched_iso = _dt.fromisoformat(sched_raw[:10]).date().isoformat()
                except Exception:
                    sched_iso = sched_raw[:10]
            ch = charges_by_id.setdefault(cid, {"id": cid, "scheduled_at": sched_iso, "line_items": []})
            ch["line_items"].append({"sku": sku, "quantity": qty})
    return list(charges_by_id.values())


_CHARGE_COUNT_LEDGER = r"C:\Users\Work\Claude Projects\_outputs\cache\cut_order_charge_counts.json"


def _warn_charge_count_drift(total, threshold_pct=40):
    """Loud sanity banner when this week's queued-charge pull jumps vs the prior run.

    The wrong-window 7/13 pull was 2525 charges vs the correct 1227 — a >100% jump
    that nothing flagged. Real demand doesn't double week-over-week; a jump this size
    means the window (or tag scope) is wrong. Warn-only: growth is possible, a human
    decides. Fire-and-forget ledger — never fails the pull.
    """
    try:
        ledger = []
        if os.path.exists(_CHARGE_COUNT_LEDGER):
            with open(_CHARGE_COUNT_LEDGER, encoding="utf-8") as f:
                ledger = json.load(f)
        window = f"{WK1_START.isoformat()}..{WK1_END.isoformat()}"
        prior = next((e for e in reversed(ledger) if e["window"] != window), None)
        if prior and prior["total"] > 0:
            delta_pct = (total - prior["total"]) / prior["total"] * 100
            if abs(delta_pct) > threshold_pct:
                print(f"\n{'!' * 78}\n"
                      f"⚠️  CHARGE-COUNT DRIFT: {total} queued charges this window vs "
                      f"{prior['total']} last window ({delta_pct:+.0f}%).\n"
                      f"   Real demand rarely moves >{threshold_pct}% week-over-week — check the "
                      f"scheduled_at window\n"
                      f"   ({window}) before trusting this cut (wrong-window burn: 7/13, 2525 vs 1227).\n"
                      f"{'!' * 78}\n")
        ledger = [e for e in ledger if e["window"] != window][-7:]
        ledger.append({"window": window, "total": total,
                       "pulled_at": datetime.now().isoformat(timespec="seconds")})
        os.makedirs(os.path.dirname(_CHARGE_COUNT_LEDGER), exist_ok=True)
        with open(_CHARGE_COUNT_LEDGER, "w", encoding="utf-8") as f:
            json.dump(ledger, f, indent=1)
    except Exception as e:
        print(f"  (charge-count ledger skipped: {type(e).__name__}: {e})")


def fetch_recharge_api(api_token, out_specialty=None, out_bundles=None, shop_creds=None):
    """Fetch queued charges. Returns pickable SKU demand + curation counts per week.

    If `out_specialty` is provided as a dict like {"WK1": {}, "WK2": {}}, it is
    populated with totals for any SKU starting with "AHB-X" or "BL-" found in
    queued charges within the WK1/WK2 windows. Used by the cut order checklist.

    If env var RECHARGE_CSV_PATH is set, load charges from CSV export instead of API
    (admin export is authoritative; API has been observed to truncate).
    """
    import requests

    # Shopify creds for live bundle-recipe lookup (Simple Bundles metafield).
    _bundle_store, _bundle_token = shop_creds if shop_creds else (None, None)

    csv_path = os.environ.get("RECHARGE_CSV_PATH", "").strip()
    if csv_path and os.path.exists(csv_path):
        all_charges = _load_recharge_csv(csv_path)
        print(f"  Loaded {len(all_charges)} charges from CSV: {csv_path}")
    else:
        session = requests.Session()
        session.headers.update(
            {
                "X-Recharge-Access-Token": api_token,
                "Accept": "application/json",
                "X-Recharge-Version": "2021-11",
            }
        )

        all_charges = []
        params = {
            "status": "queued",
            "limit": 250,
            "sort_by": "id-asc",
            "scheduled_at_min": WK1_START.isoformat(),
            "scheduled_at_max": WK2_END.isoformat(),
        }
        page = 0

        while True:
            page += 1
            for attempt in range(3):
                try:
                    resp = session.get("https://api.rechargeapps.com/charges", params=params, timeout=60)
                    resp.raise_for_status()
                    break
                except Exception:
                    if attempt < 2:
                        time.sleep(2)
                    else:
                        raise
            data = resp.json()
            charges = data.get("charges", [])
            if not charges:
                break
            all_charges.extend(charges)
            sys.stdout.write(f"\r  Fetched {len(all_charges)} charges (page {page})...")
            sys.stdout.flush()

            next_cursor = data.get("next_cursor")
            if not next_cursor:
                break
            # Cursor requests: ONLY cursor + limit
            params = {"cursor": next_cursor, "limit": 250}
            time.sleep(0.5)

        print(f"\r  Fetched {len(all_charges)} total queued charges.     ")

    _warn_charge_count_drift(len(all_charges))

    # Per-week: pickable SKU demand + curation counts
    wk1_skus = defaultdict(int)
    wk2_skus = defaultdict(int)
    wk1_curations = defaultdict(int)  # {curation: charge_count}
    wk2_curations = defaultdict(int)
    wk1_cexec = defaultdict(int)  # {curation: cexec_line_item_count}
    wk2_cexec = defaultdict(int)
    wk1_med_total = 0  # total AHB-MED/MCUST boxes (excludes CMED)
    wk2_med_total = 0
    wk1_cmed_total = 0  # total AHB-CMED boxes
    wk2_cmed_total = 0
    wk1_lge_total = 0  # total AHB-LGE/LCUST boxes
    wk2_lge_total = 0
    # Monthly box counts split by (week, charge_month) — recipes differ per month
    # Key: (week, charge_month) e.g. ("WK1", "2026-04"), ("WK2", "2026-05")
    monthly_by_week_month = defaultdict(lambda: {"MED": 0, "CMED": 0, "LGE": 0})
    charges_per_date = defaultdict(int)
    wk1_custom_total = 0  # AHB- charges where curation unresolved (custom boxes)
    wk2_custom_total = 0

    for charge in all_charges:
        sched = (charge.get("scheduled_at") or "")[:10]
        if not sched:
            continue
        try:
            d = date.fromisoformat(sched)
        except ValueError:
            continue

        charges_per_date[sched] += 1
        is_wk1 = WK1_START <= d <= WK1_END
        is_wk2 = WK2_START <= d <= WK2_END
        if not is_wk1 and not is_wk2:
            continue

        # Find box SKU and curation for this charge
        box_sku = None
        for item in charge.get("line_items", []):
            sku = (item.get("sku") or "").strip()
            if sku.upper().startswith("AHB-"):
                box_sku = sku
                break

        if box_sku:
            upper_box = box_sku.upper()
            cur = resolve_curation(box_sku)
            is_lg = is_large_box(box_sku)

            # Count MONTHLY-only MED/LGE boxes (plain AHB-MED, AHB-LGE, AHB-CMED, etc.)
            # Custom curations (AHB-MCUST-MDT, AHB-LCUST-OWC) are already in per-curation tables
            if cur == "MONTHLY":
                is_cmed = "CMED" in upper_box
                charge_month = sched[:7]  # "2026-04" or "2026-05"
                week_label = "WK1" if is_wk1 else "WK2"
                if is_lg:
                    if is_wk1:
                        wk1_lge_total += 1
                    else:
                        wk2_lge_total += 1
                    monthly_by_week_month[(week_label, charge_month)]["LGE"] += 1
                elif is_cmed:
                    if is_wk1:
                        wk1_cmed_total += 1
                    else:
                        wk2_cmed_total += 1
                    monthly_by_week_month[(week_label, charge_month)]["CMED"] += 1
                else:
                    if is_wk1:
                        wk1_med_total += 1
                    else:
                        wk2_med_total += 1
                    monthly_by_week_month[(week_label, charge_month)]["MED"] += 1

            # Curation counts (non-MONTHLY only)
            if cur and cur != "MONTHLY":
                # CEX-EC counts off the actual CEX-EC line item, NOT box size.
                cex_cur = cexec_curation(charge.get("line_items", []), cur)
                if is_wk1:
                    wk1_curations[cur] += 1
                    if cex_cur:
                        wk1_cexec[cex_cur] += 1
                else:
                    wk2_curations[cur] += 1
                    if cex_cur:
                        wk2_cexec[cex_cur] += 1

            # Custom boxes (AHB-MCUST-TRAY / AHB-LCUST-TRAY) — no curation/MONTHLY match
            if cur is None:
                if is_wk1:
                    wk1_custom_total += 1
                else:
                    wk2_custom_total += 1

        # Sum pickable SKU quantities (all charges, including MONTHLY boxes)
        for item in charge.get("line_items", []):
            sku = (item.get("sku") or "").strip()
            if not sku:
                continue
            qty = int(float(item.get("quantity", 1)))

            # Specialty/bundle accumulator (AHB-X*, BL-*) — checklist source
            if out_specialty is not None:
                upper = sku.upper()
                if upper.startswith("AHB-X") or upper.startswith("BL-"):
                    bucket = out_specialty["WK1"] if is_wk1 else out_specialty["WK2"]
                    bucket[sku] = bucket.get(sku, 0) + qty

            if not any(sku.startswith(p) for p in PICKABLE_PREFIXES):
                continue
            if is_wk1:
                wk1_skus[sku] += qty
            else:
                wk2_skus[sku] += qty

        # Explode bundle/limited-release boxes present in this charge → add ONLY
        # components missing from the charge's own lines. Some charges already
        # list them (BL-BLR4 charge 1816643580 carries CH-BLR×4 + AC-BLBALS×2);
        # those are not re-added. Same double-count rule as the Shopify side.
        if _bundle_store:
            present_qty: dict = {}
            for item in charge.get("line_items", []):
                s = (item.get("sku") or "").strip()
                if s:
                    present_qty[s] = present_qty.get(s, 0) + int(float(item.get("quantity", 1)))
            _explode_bundles(
                present_qty, _bundle_store, _bundle_token,
                wk1_skus if is_wk1 else wk2_skus,
                out_bundles, "WK1" if is_wk1 else "WK2",
            )

    return (
        dict(wk1_skus),
        dict(wk2_skus),
        dict(wk1_curations),
        dict(wk2_curations),
        dict(wk1_cexec),
        dict(wk2_cexec),
        len(all_charges),
        dict(charges_per_date),
        wk1_med_total,
        wk2_med_total,
        wk1_cmed_total,
        wk2_cmed_total,
        wk1_lge_total,
        wk2_lge_total,
        dict(monthly_by_week_month),
        wk1_custom_total,
        wk2_custom_total,
    )


# -- Step 4: Fetch Shopify orders for upcoming ship weeks --

# Ship week tag dates (Monday of each week) — DYNAMIC (no weekly hand-editing).
# WK1 = current ship Monday; WK2 = the week beyond WK1_SHIP_TAGS (dropped by single-cohort,
# kept non-overlapping so it never double-classifies a WK1 order).
_THIS_MONDAY = _TODAY - timedelta(days=_TODAY.weekday())
WK1_SHIP_TAG = f"_SHIP_{_THIS_MONDAY.isoformat()}"
WK2_SHIP_TAG = f"_SHIP_{(_THIS_MONDAY + timedelta(days=14)).isoformat()}"

# Day-of-week ship-tag selection:
#   Monday run → WK1 includes BOTH this Mon's ship tag and next Mon's
#                (this week's cohort still being prepped + next week's queueing).
#   Tuesday+   → WK1 includes ONLY next Mon's ship tag (this week already cut).
def _wk1_ship_tags(today=None):
    from datetime import date as _date, timedelta as _td
    today = today or _date.today()
    # Find next Monday (or today if Mon)
    days_to_mon = (0 - today.weekday()) % 7
    next_mon = today + _td(days=days_to_mon if days_to_mon != 0 else 7)
    tags = [f"_SHIP_{next_mon.isoformat()}"]
    if today.weekday() == 0:  # Monday
        this_mon = today
        tags.insert(0, f"_SHIP_{this_mon.isoformat()}")
    return tags

WK1_SHIP_TAGS = _wk1_ship_tags()


def fetch_shopify_orders(settings, out_specialty=None, out_bundles=None):
    """Fetch open Shopify orders for WK1/WK2 ship tags.

    Returns per-week:
      addon_skus   — pickable SKU demand from non-subscription orders (no AHB-* box)
      curations    — {curation: count} from subscription orders (has AHB-* box)
      large_boxes  — {curation: count} where box is AHB-L*
      med_counts   — {curation: count} where box is AHB-M* (MED/MCUST)
      lge_counts   — {curation: count} where box is AHB-L* (LGE/LCUST)
    """
    import requests

    store = settings.get("shopify_store_url", "").strip()
    token = settings.get("shopify_access_token", "").strip()
    if not store or not token:
        print("  WARNING: Shopify credentials not configured, skipping")
        empty = {}
        return (empty, empty, empty, empty, empty, empty, empty, empty, empty, empty)

    if not store.startswith("http"):
        store = f"https://{store}.myshopify.com"

    session = requests.Session()
    session.headers.update(
        {
            "X-Shopify-Access-Token": token,
            "Content-Type": "application/json",
        }
    )

    # Fetch all open/unfulfilled orders (no date cutoff — ship tag is the filter)
    import re

    url = f"{store}/admin/api/2026-04/orders.json"
    params = {"status": "open", "fulfillment_status": "unfulfilled", "limit": 250}

    all_orders = []
    while url:
        print(f"  Shopify: fetching page {len(all_orders) // 250 + 1}...", flush=True)
        resp = session.get(url, params=params, timeout=60)
        if resp.status_code != 200:
            print(f"  ERROR: Shopify API {resp.status_code}: {resp.text[:200]}")
            empty = {}
            return (empty, empty, empty, empty, empty, empty, empty, empty, empty, empty)
        data = resp.json()
        all_orders.extend(data.get("orders", []))
        url = None
        params = None
        link = resp.headers.get("Link", "")
        if 'rel="next"' in link:
            m = re.search(r'<([^>]+)>;\s*rel="next"', link)
            if m:
                url = m.group(1)
        time.sleep(0.3)

    print(f"  Shopify: {len(all_orders)} open orders fetched")

    # Classify orders by ship week
    wk1_addon = defaultdict(int)
    wk2_addon = defaultdict(int)
    wk1_curations = defaultdict(int)
    wk2_curations = defaultdict(int)
    wk1_cexec = defaultdict(int)
    wk2_cexec = defaultdict(int)
    wk1_med = defaultdict(int)
    wk2_med = defaultdict(int)
    wk1_lge = defaultdict(int)
    wk2_lge = defaultdict(int)
    wk1_count = 0
    wk2_count = 0

    for order in all_orders:
        tags = order.get("tags", "") or ""

        is_wk1 = any(t in tags for t in WK1_SHIP_TAGS)
        is_wk2 = WK2_SHIP_TAG in tags
        if not is_wk1 and not is_wk2:
            continue

        if is_wk1:
            wk1_count += 1
        else:
            wk2_count += 1

        # Extract line items — NEVER count removed items.
        # MANDATORY rule (shopify-line-items skill): net removed/refunded lines
        # via the canonical active_line_items() helper — do NOT read raw
        # order["line_items"], Shopify keeps refunded/edited-out lines at their
        # ORIGINAL quantity so raw counting phantom-counts food taken off the
        # order (burned us on MT-FS-BRAS #165907, swapped to MT-IBRES 7/24).
        # Belt-and-suspenders: also drop fulfillable_quantity==0 for the rare
        # order-edit removal that emits no refund_line_item. Orders here are
        # status=unfulfilled, so fulfillable_quantity is the true remaining qty.
        line_items = [
            it for it in active_line_items(order)
            if it.get("fulfillable_quantity") != 0
        ]
        item_skus = {}
        # Collect all AHB-* SKUs then pick primary box (not addon like AHB-CUR-MS / AHB-BVAL)
        ahb_skus = []
        for item in line_items:
            sku = (item.get("sku") or "").strip()
            if not sku:
                continue
            qty = int(float(item.get("quantity", 1)))
            item_skus[sku] = item_skus.get(sku, 0) + qty
            upper = sku.upper()
            if upper.startswith("AHB-"):
                ahb_skus.append(upper)
            # Specialty/bundle accumulator (AHB-X*, BL-*) — checklist source
            if out_specialty is not None and (upper.startswith("AHB-X") or upper.startswith("BL-")):
                bucket = out_specialty["WK1"] if is_wk1 else out_specialty["WK2"]
                bucket[sku] = bucket.get(sku, 0) + qty
        # Primary box priority: plain monthly > MCUST/LCUST curation > anything else
        # Skip AHB-CUR-* and AHB-BVAL which are addon/bundle SKUs
        _ADDON_AHB = ("AHB-CUR-", "AHB-BVAL")
        box_sku = None
        for s in ahb_skus:
            if s in ("AHB-MED", "AHB-LGE", "AHB-CMED"):
                box_sku = s
                break
        if not box_sku:
            for s in ahb_skus:
                if (s.startswith("AHB-MCUST-") or s.startswith("AHB-LCUST-")) and not s.startswith(_ADDON_AHB):
                    box_sku = s
                    break
        if not box_sku:
            for s in ahb_skus:
                if not s.startswith(_ADDON_AHB):
                    box_sku = s
                    break

        if box_sku:
            # Subscription order — count ALL pickable SKUs as Shopify demand.
            # Recharge queued charges and Shopify unfulfilled orders are ADDITIVE:
            # RC queued = not yet processed, SH unfulfilled = already processed.
            target = wk1_addon if is_wk1 else wk2_addon
            for item in line_items:
                sku = (item.get("sku") or "").strip()
                if not sku:
                    continue
                upper = sku.upper()
                if any(upper.startswith(p) for p in PICKABLE_PREFIXES):
                    qty = int(float(item.get("quantity", 1)))
                    target[sku] += qty
            cur = resolve_curation(box_sku)
            is_lg = is_large_box(box_sku)

            if cur == "MONTHLY":
                # Plain AHB-MED/AHB-LGE/AHB-CMED — count as box totals
                is_cmed = "CMED" in box_sku
                if is_lg:
                    if is_wk1:
                        wk1_lge["MONTHLY"] = wk1_lge.get("MONTHLY", 0) + 1
                    else:
                        wk2_lge["MONTHLY"] = wk2_lge.get("MONTHLY", 0) + 1
                elif is_cmed:
                    if is_wk1:
                        wk1_med["CMED"] = wk1_med.get("CMED", 0) + 1
                    else:
                        wk2_med["CMED"] = wk2_med.get("CMED", 0) + 1
                else:
                    if is_wk1:
                        wk1_med["MONTHLY"] = wk1_med.get("MONTHLY", 0) + 1
                    else:
                        wk2_med["MONTHLY"] = wk2_med.get("MONTHLY", 0) + 1
            elif cur:
                # Custom curation — count in per-curation tables.
                # Box size (_lge/_med) and CEX-EC (_cexec) are INDEPENDENT:
                # _lge/_med = physical box size; _cexec = CEX-EC line-item presence.
                cex_cur = cexec_curation(line_items, cur)
                if is_wk1:
                    wk1_curations[cur] += 1
                    if is_lg:
                        wk1_lge[cur] += 1
                    else:
                        wk1_med[cur] += 1
                    if cex_cur:
                        wk1_cexec[cex_cur] += 1
                else:
                    wk2_curations[cur] += 1
                    if is_lg:
                        wk2_lge[cur] += 1
                    else:
                        wk2_med[cur] += 1
                    if cex_cur:
                        wk2_cexec[cex_cur] += 1
        else:
            # Non-subscription order — count pickable SKUs as addon demand
            target = wk1_addon if is_wk1 else wk2_addon
            for sku, qty in item_skus.items():
                upper = sku.upper()
                if any(upper.startswith(p) for p in PICKABLE_PREFIXES):
                    target[sku] += qty

        # Explode bundle/limited-release boxes → add ONLY components missing from
        # this order's own lines (Shopify orders already list them, so nothing
        # double-counts). Records the per-bundle breakdown for the Cut Order tab.
        _explode_bundles(
            item_skus, store, token,
            wk1_addon if is_wk1 else wk2_addon,
            out_bundles, "WK1" if is_wk1 else "WK2",
        )

    print(
        f"  Shopify WK1: {wk1_count} orders ({sum(wk1_curations.values())} subs, "
        f"{wk1_count - sum(wk1_curations.values())} addon-only)"
    )
    print(
        f"  Shopify WK2: {wk2_count} orders ({sum(wk2_curations.values())} subs, "
        f"{wk2_count - sum(wk2_curations.values())} addon-only)"
    )

    return (
        dict(wk1_addon),
        dict(wk2_addon),
        dict(wk1_curations),
        dict(wk2_curations),
        dict(wk1_cexec),
        dict(wk2_cexec),
        dict(wk1_med),
        dict(wk2_med),
        dict(wk1_lge),
        dict(wk2_lge),
    )


# -- Main --


def main():
    settings = load_settings()
    sku_translations = settings.get("sku_translations", {})
    recharge_token = settings.get("recharge_api_token", "")
    pr_cjam = settings.get("pr_cjam", {})
    cex_ec = settings.get("cex_ec", {})

    # 1. Load inventory
    print("Loading inventory from 3/14 snapshot...")
    inventory = load_inventory_csv(INV_CSV)
    print(f"  {len(inventory)} SKUs loaded.")

    # 2. Parse depletions
    print("\nParsing Saturday 3/16 depletion...")
    sat_dep, sat_orders, sat_first, sat_first_skus, sat_unmatched = parse_depletion_xlsx(
        SAT_DEPLETION, sku_translations
    )
    print(f"  {sat_orders} orders, {sum(sat_dep.values())} items, {sat_first} first orders.")

    print("Parsing Tuesday 3/17 depletion...")
    tue_dep, tue_orders, tue_first, tue_first_skus, tue_unmatched = parse_depletion_xlsx(
        TUE_DEPLETION, sku_translations
    )
    print(f"  {tue_orders} orders, {sum(tue_dep.values())} items, {tue_first} first orders.")

    all_unmatched = set()
    for name, qty in sat_unmatched + tue_unmatched:
        all_unmatched.add(name)
    if all_unmatched:
        print(f"\n  WARNING: {len(all_unmatched)} unmatched product names:")
        for name in sorted(all_unmatched):
            print(f"    - {name}")

    # 3. Available inventory
    available = {}
    all_skus = set(inventory.keys()) | set(sat_dep.keys()) | set(tue_dep.keys())
    for sku in all_skus:
        available[sku] = inventory.get(sku, 0) - sat_dep.get(sku, 0) - tue_dep.get(sku, 0)

    # 4. Recharge demand
    if not recharge_token:
        print("\nNo Recharge API token!")
        return

    print("\nFetching Recharge demand from API (v2021-11)...")
    (
        rc_wk1,
        rc_wk2,
        wk1_curations,
        wk2_curations,
        wk1_cexec,
        wk2_cexec,
        total_charges,
        charges_per_date,
        wk1_med_monthly,
        wk2_med_monthly,
        _wk1_cmed,
        _wk2_cmed,
        wk1_lge_monthly,
        wk2_lge_monthly,
    ) = fetch_recharge_api(recharge_token)

    print(f"\n  Charges per date:")
    wk1_ct = wk2_ct = 0
    for d in sorted(charges_per_date.keys()):
        dd = date.fromisoformat(d)
        if WK1_START <= dd <= WK1_END:
            wk1_ct += charges_per_date[d]
            marker = " <-- WK1"
        elif WK2_START <= dd <= WK2_END:
            wk2_ct += charges_per_date[d]
            marker = " <-- WK2"
        else:
            marker = ""
        print(f"    {d}: {charges_per_date[d]:>5} charges{marker}")
    print(f"  WK1 charges: {wk1_ct:,}  |  WK2 charges: {wk2_ct:,}")

    # 5. First order projections
    wk1_first = dict(sat_first_skus)  # project Sat 3/16 first order profile
    wk2_first = dict(tue_first_skus)  # project Tue 3/17 first order profile

    # 6. Combine demands
    wk1_total = defaultdict(int)
    for sku, qty in rc_wk1.items():
        wk1_total[sku] += qty
    for sku, qty in wk1_first.items():
        wk1_total[sku] += qty

    wk2_total = defaultdict(int)
    for sku, qty in rc_wk2.items():
        wk2_total[sku] += qty
    for sku, qty in wk2_first.items():
        wk2_total[sku] += qty

    wk1_total = dict(wk1_total)
    wk2_total = dict(wk2_total)

    # -- Report --
    report_skus = set()
    for d in (available, wk1_total, wk2_total):
        report_skus.update(d.keys())
    report_skus = sorted(sku for sku in report_skus if any(sku.startswith(p) for p in PICKABLE_PREFIXES))

    inv_settings = settings.get("inventory", {})

    def sku_name(sku):
        data = inv_settings.get(sku, {})
        return data.get("name", "") if isinstance(data, dict) else ""

    print("\n" + "=" * 100)
    print(f"  {'SKU':<14} {'Name':<35} {'Avail':>7} {'Wk1':>7} {'AftWk1':>7} {'Wk2':>7}")
    print(f"  {'':14} {'':35} {'(now)':>7} {'(3/23)':>7} {'':>7} {'(3/30)':>7}")
    print("  " + "-" * 96)

    categories = {"CH-": [], "MT-": [], "AC-": []}
    for sku in report_skus:
        for prefix in PICKABLE_PREFIXES:
            if sku.startswith(prefix):
                categories[prefix].append(sku)
                break

    cat_labels = {"CH-": "CHEESE", "MT-": "MEAT", "AC-": "ACCOMPANIMENTS"}

    for prefix in PICKABLE_PREFIXES:
        skus = categories[prefix]
        if not skus:
            continue
        active = [s for s in skus if available.get(s, 0) != 0 or wk1_total.get(s, 0) > 0 or wk2_total.get(s, 0) > 0]
        if not active:
            continue

        print(f"\n  -- {cat_labels[prefix]} " + "-" * 85)
        for sku in active:
            avail = available.get(sku, 0)
            w1 = wk1_total.get(sku, 0)
            after_w1 = avail - w1
            w2 = wk2_total.get(sku, 0)

            flag = ""
            if after_w1 - w2 < 0:
                flag = " *** SHORT"
            elif avail < 0:
                flag = " (deficit)"

            name = sku_name(sku)[:35]
            print(f"  {sku:<14} {name:<35} {avail:>7,} {w1:>7,} {after_w1:>7,} {w2:>7,}{flag}")

    # -- PR-CJAM Totals --
    print(f"\n  {'=' * 96}")
    print(f"  PR-CJAM TOTALS (1 per box, all curations)")
    print(f"  {'-' * 96}")
    print(f"  {'Curation':<12} {'Cheese':<14} {'Name':<35} {'Wk1':>7} {'Wk2':>7}")
    print(f"  {'-' * 96}")

    cjam_wk1_totals = defaultdict(int)  # {cheese_sku: total}
    cjam_wk2_totals = defaultdict(int)

    for cur in sorted(set(list(wk1_curations.keys()) + list(wk2_curations.keys()))):
        cjam = pr_cjam.get(cur, {})
        cheese = cjam.get("cheese", "")
        if not cheese:
            continue
        w1 = wk1_curations.get(cur, 0)
        w2 = wk2_curations.get(cur, 0)
        cjam_wk1_totals[cheese] += w1
        cjam_wk2_totals[cheese] += w2
        name = sku_name(cheese)[:35]
        print(f"  {cur:<12} {cheese:<14} {name:<35} {w1:>7,} {w2:>7,}")

    print(f"  {'-' * 96}")
    print(f"  {'TOTALS BY CHEESE':}")
    for cheese in sorted(set(list(cjam_wk1_totals.keys()) + list(cjam_wk2_totals.keys()))):
        w1 = cjam_wk1_totals[cheese]
        w2 = cjam_wk2_totals[cheese]
        name = sku_name(cheese)[:35]
        print(f"  {'':12} {cheese:<14} {name:<35} {w1:>7,} {w2:>7,}")

    # -- CEX-EC Totals --
    print(f"\n  {'=' * 96}")
    print(f"  CEX-EC TOTALS (1 per order carrying a CEX-EC line item)")
    print(f"  {'-' * 96}")
    print(f"  {'Curation':<12} {'Cheese':<14} {'Name':<35} {'Wk1':>7} {'Wk2':>7}")
    print(f"  {'-' * 96}")

    cexec_wk1_totals = defaultdict(int)
    cexec_wk2_totals = defaultdict(int)

    for cur in sorted(set(list(wk1_cexec.keys()) + list(wk2_cexec.keys()))):
        cheese = cex_ec.get(cur, "")
        if not cheese:
            continue
        w1 = wk1_cexec.get(cur, 0)
        w2 = wk2_cexec.get(cur, 0)
        cexec_wk1_totals[cheese] += w1
        cexec_wk2_totals[cheese] += w2
        name = sku_name(cheese)[:35]
        print(f"  {cur:<12} {cheese:<14} {name:<35} {w1:>7,} {w2:>7,}")

    print(f"  {'-' * 96}")
    print(f"  {'TOTALS BY CHEESE':}")
    for cheese in sorted(set(list(cexec_wk1_totals.keys()) + list(cexec_wk2_totals.keys()))):
        w1 = cexec_wk1_totals[cheese]
        w2 = cexec_wk2_totals[cheese]
        name = sku_name(cheese)[:35]
        print(f"  {'':12} {cheese:<14} {name:<35} {w1:>7,} {w2:>7,}")

    # -- Summary --
    print("\n" + "=" * 100)
    print(f"Inventory: 3/14 RMFG snapshot")
    print(f"Depletions: Sat 3/16 ({sat_orders:,} orders) + Tue 3/17 ({tue_orders:,} orders)")
    print(f"Wk1: RC={sum(rc_wk1.values()):,} + FO={sum(wk1_first.values()):,} = {sum(wk1_total.values()):,} items")
    print(f"Wk2: RC={sum(rc_wk2.values()):,} + FO={sum(wk2_first.values()):,} = {sum(wk2_total.values()):,} items")

    shortages = []
    for sku in report_skus:
        avail = available.get(sku, 0)
        net = avail - wk1_total.get(sku, 0) - wk2_total.get(sku, 0)
        if net < 0:
            shortages.append((sku, net))
    if shortages:
        print(f"\n*** {len(shortages)} SKUs SHORT after 2 weeks:")
        for sku, deficit in sorted(shortages, key=lambda x: x[1]):
            print(f"    {sku:<14} {sku_name(sku)[:30]:<30} {deficit:>+8,}")

    # -- CSV --
    csv_path = os.path.join(BASE, "inventory_demand_report.csv")
    with open(csv_path, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["SKU", "Name", "Available", "Wk1_Total", "After_Wk1", "Wk2_Total"])
        for sku in report_skus:
            avail = available.get(sku, 0)
            w1 = wk1_total.get(sku, 0)
            after = avail - w1
            w2 = wk2_total.get(sku, 0)
            writer.writerow([sku, sku_name(sku), avail, w1, after, w2])

        # PR-CJAM section
        writer.writerow([])
        writer.writerow(["PR-CJAM", "Curation", "Cheese", "Wk1", "Wk2"])
        for cur in sorted(set(list(wk1_curations.keys()) + list(wk2_curations.keys()))):
            cheese = pr_cjam.get(cur, {}).get("cheese", "")
            if cheese:
                writer.writerow(["", cur, cheese, wk1_curations.get(cur, 0), wk2_curations.get(cur, 0)])
        writer.writerow(["PR-CJAM TOTALS"])
        for cheese in sorted(set(list(cjam_wk1_totals.keys()) + list(cjam_wk2_totals.keys()))):
            writer.writerow(["", "", cheese, cjam_wk1_totals[cheese], cjam_wk2_totals[cheese]])

        # CEX-EC section
        writer.writerow([])
        writer.writerow(["CEX-EC", "Curation", "Cheese", "Wk1", "Wk2"])
        for cur in sorted(set(list(wk1_cexec.keys()) + list(wk2_cexec.keys()))):
            cheese = cex_ec.get(cur, "")
            if cheese:
                writer.writerow(["", cur, cheese, wk1_cexec.get(cur, 0), wk2_cexec.get(cur, 0)])
        writer.writerow(["CEX-EC TOTALS"])
        for cheese in sorted(set(list(cexec_wk1_totals.keys()) + list(cexec_wk2_totals.keys()))):
            writer.writerow(["", "", cheese, cexec_wk1_totals[cheese], cexec_wk2_totals[cheese]])

    print(f"\nCSV written to: {csv_path}")


if __name__ == "__main__":
    main()
